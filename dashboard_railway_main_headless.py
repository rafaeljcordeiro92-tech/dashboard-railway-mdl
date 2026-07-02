# VERSAO: DASH2_0_V10_7_RATEIO_MES_CREDIARISTA_FIX
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pandas as pd
import time
import re
import os
import webbrowser
import random
import json
import unicodedata
import urllib.request
import urllib.error
import ssl
import tempfile
import shutil
import sys
from zoneinfo import ZoneInfo

# V6.8: evita erro UnicodeEncodeError no Windows/cp1252 quando o log imprime emojis.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

LOGIN = "administrativo01.moveisdolar"
SENHA = "mdladm01"
URL   = "https://smart.sgisistemas.com.br"
APP_TZ = ZoneInfo(os.getenv("APP_TZ", "America/Sao_Paulo"))

DASHBOARD_BUILD_VERSION = "V10.7"
DASHBOARD_BUILD_TAG = "DASH2_0_V10_7_RATEIO_MES_CREDIARISTA_FIX"

def now_brasilia():
    return datetime.now(APP_TZ)

# ===== DATAS
hoje        = now_brasilia()
data_inicio = (hoje - timedelta(days=90)).strftime("%d/%m/%Y")
data_fim    = (hoje - timedelta(days=15)).strftime("%d/%m/%Y")

# ===== CHROME
IS_RAILWAY = bool(os.getenv("RAILWAY_ENVIRONMENT") or os.getenv("RUN_ON_RAILWAY") == "1")
pasta = os.path.dirname(os.path.abspath(__file__))
download_dir = pasta if not IS_RAILWAY else tempfile.gettempdir()

from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import WebDriverException, SessionNotCreatedException


def _resolve_bin(env_names, candidates):
    """Resolve binário por variável de ambiente, caminho fixo ou PATH."""
    for env in env_names:
        val = os.getenv(env)
        if val and os.path.exists(val):
            return val
    for cand in candidates:
        if cand and os.path.exists(cand):
            return cand
    for name in ["chromium", "chromium-browser", "google-chrome", "google-chrome-stable", "chromedriver"]:
        found = shutil.which(name)
        if found:
            if "driver" in name and "CHROMEDRIVER" not in env_names:
                continue
            if "driver" not in name and "CHROME" not in "".join(env_names):
                continue
            return found
    return None


def _make_chrome_options(profile_dir):
    opts = Options()

    if IS_RAILWAY:
        opts.page_load_strategy = "eager"
        # Headless deixa o Chrome mais estável no Railway, mesmo com Xvfb disponível.
        if os.getenv("DISABLE_CHROME_HEADLESS", "0") != "1":
            opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        # NÃO usar porta fixa 9222: em Railway pode dar conflito e o Chrome fecha.
        opts.add_argument("--remote-debugging-port=0")
        opts.add_argument("--remote-debugging-address=127.0.0.1")
        opts.add_argument("--disable-software-rasterizer")
        opts.add_argument("--disable-extensions")
        opts.add_argument("--disable-background-networking")
        opts.add_argument("--disable-background-timer-throttling")
        opts.add_argument("--disable-backgrounding-occluded-windows")
        opts.add_argument("--disable-renderer-backgrounding")
        opts.add_argument("--hide-scrollbars")
        opts.add_argument("--mute-audio")
        opts.add_argument("--no-first-run")
        opts.add_argument("--no-default-browser-check")
        opts.add_argument("--disable-features=VizDisplayCompositor,TranslateUI")
        opts.add_argument("--disable-setuid-sandbox")
        opts.add_argument("--disable-infobars")
        opts.add_argument("--ignore-certificate-errors")
        opts.add_argument("--allow-insecure-localhost")
        opts.add_argument("--disable-crash-reporter")
        opts.add_argument("--disable-sync")
        opts.add_argument("--metrics-recording-only")
        opts.add_argument("--disable-default-apps")
        # Perfil único por execução: evita "Chrome instance exited" por profile lock.
        opts.add_argument(f"--user-data-dir={profile_dir}")
        opts.add_argument(f"--data-path={os.path.join(profile_dir, 'data')}")
        opts.add_argument(f"--disk-cache-dir={os.path.join(profile_dir, 'cache')}")
    else:
        opts.add_argument("--start-maximized")

    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "profile.default_content_settings.popups": 0,
    }
    opts.add_experimental_option("prefs", prefs)

    chrome_bin = _resolve_bin(
        ["GOOGLE_CHROME_BIN", "CHROME_BIN"],
        ["/usr/bin/chromium", "/usr/bin/chromium-browser", "/usr/bin/google-chrome", "/usr/bin/google-chrome-stable"]
    )
    if chrome_bin and os.path.exists(chrome_bin):
        opts.binary_location = chrome_bin

    return opts


def iniciar_driver_chrome():
    chrome_driver_bin = _resolve_bin(
        ["CHROMEDRIVER", "CHROMEDRIVER_PATH"],
        ["/usr/bin/chromedriver", "/usr/local/bin/chromedriver"]
    )

    ultimo_erro = None
    for tentativa in range(1, 4):
        profile_dir = os.path.join(
            tempfile.gettempdir(),
            f"chrome-profile-mdl-{os.getpid()}-{int(time.time()*1000)}-{tentativa}"
        )
        os.makedirs(profile_dir, exist_ok=True)
        opts = _make_chrome_options(profile_dir)

        try:
            print(f"🧭 Iniciando Chrome tentativa {tentativa}/3")
            print(f"   IS_RAILWAY={IS_RAILWAY} DISPLAY={os.getenv('DISPLAY')}")
            print(f"   Chrome binary={getattr(opts, 'binary_location', '')}")
            print(f"   ChromeDriver={chrome_driver_bin or 'Selenium Manager'}")
            print(f"   Profile={profile_dir}")

            if chrome_driver_bin and os.path.exists(chrome_driver_bin):
                drv = webdriver.Chrome(service=Service(chrome_driver_bin), options=opts)
            else:
                drv = webdriver.Chrome(options=opts)
            return drv

        except (SessionNotCreatedException, WebDriverException) as e:
            ultimo_erro = e
            print(f"⚠️ Falha ao iniciar Chrome na tentativa {tentativa}/3: {e}")
            time.sleep(2)

    print("❌ Não foi possível iniciar Chrome/Chromedriver após 3 tentativas.")
    print(f"   GOOGLE_CHROME_BIN={os.getenv('GOOGLE_CHROME_BIN')}")
    print(f"   CHROME_BIN={os.getenv('CHROME_BIN')}")
    print(f"   CHROMEDRIVER={os.getenv('CHROMEDRIVER')}")
    print(f"   CHROMEDRIVER_PATH={os.getenv('CHROMEDRIVER_PATH')}")
    raise ultimo_erro


driver = iniciar_driver_chrome()
driver.set_page_load_timeout(120)
wait = WebDriverWait(driver, 40)

# V10.6: no Chrome headless/Railway, às vezes o clique em Gerar acontece,
# mas o Chromium não grava o arquivo em /tmp sem liberar o download via CDP.
# Mantém também os prefs normais, mas força o downloadPath aqui.
def habilitar_download_headless():
    try:
        os.makedirs(download_dir, exist_ok=True)
    except Exception:
        pass
    for _cmd in ("Page.setDownloadBehavior", "Browser.setDownloadBehavior"):
        try:
            driver.execute_cdp_cmd(_cmd, {"behavior": "allow", "downloadPath": download_dir})
            print(f"✅ Download headless liberado via CDP ({_cmd}): {download_dir}")
            return True
        except Exception as _e_cdp:
            _last_cdp_err = _e_cdp
    try:
        print(f"⚠️ Não consegui liberar download via CDP; seguindo com prefs. Erro: {_last_cdp_err}")
    except Exception:
        print("⚠️ Não consegui liberar download via CDP; seguindo com prefs.")
    return False

habilitar_download_headless()


# ===== SELENIUM HELPERS
def esperar_sumir_overlays(timeout=10):
    fim = time.time() + timeout
    overlays = [
        ".blockUI", ".modal-backdrop", ".loading", ".spinner",
        ".ui-widget-overlay", ".select2-drop-mask"
    ]
    while time.time() < fim:
        try:
            ativos = driver.execute_script("""
                const sels = arguments[0];
                return sels.some(sel =>
                    Array.from(document.querySelectorAll(sel)).some(el => {
                        const s = window.getComputedStyle(el);
                        return s.display !== 'none' && s.visibility !== 'hidden' && el.offsetParent !== null;
                    })
                );
            """, overlays)
            if not ativos:
                return True
        except Exception:
            return True
        time.sleep(0.2)
    return False


def localizar_xpath_clickavel(xpath, timeout=20):
    elem = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center', inline:'center'});", elem
    )
    esperar_sumir_overlays()
    try:
        elem = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
    except Exception:
        elem = driver.find_element(By.XPATH, xpath)
    return elem


def clicar_seguro_xpath(xpath, timeout=20):
    ultimo_erro = None
    for _ in range(3):
        try:
            elem = localizar_xpath_clickavel(xpath, timeout=timeout)
            try:
                elem.click()
            except Exception:
                driver.execute_script("arguments[0].click();", elem)
            return elem
        except Exception as e:
            ultimo_erro = e
            time.sleep(0.5)
    raise ultimo_erro


def abrir_multiselect_por_label(label_texto, timeout=20):
    xpath_botao = (
        f"//label[contains(normalize-space(.),'{label_texto}')]/following::button[1]"
    )
    botao = clicar_seguro_xpath(xpath_botao, timeout=timeout)
    time.sleep(1)
    return botao


def marcar_multiselect_por_label(label_texto, valores, timeout=20, limpar_antes=True):
    abrir_multiselect_por_label(label_texto, timeout=timeout)
    xpath_ul = f"//label[contains(normalize-space(.),'{label_texto}')]/following::ul[1]"
    container = wait.until(EC.presence_of_element_located((By.XPATH, xpath_ul)))
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center', inline:'nearest'});", container
    )
    time.sleep(0.5)
    checkboxes = container.find_elements(By.XPATH, ".//input[@type='checkbox']")
    valores = {str(v) for v in valores}

    if limpar_antes:
        for c in checkboxes:
            if c.is_selected():
                driver.execute_script("arguments[0].click();", c)
                time.sleep(0.05)

    for c in checkboxes:
        if c.get_attribute("value") in valores and not c.is_selected():
            driver.execute_script("arguments[0].click();", c)
            time.sleep(0.05)

    driver.execute_script("document.body.click();")
    time.sleep(0.5)
    return checkboxes


def normalizar_texto_match(texto):
    s = str(texto or "").strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def set_input_value_safe(by, locator, value):
    el = wait.until(EC.presence_of_element_located((by, locator)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    try:
        el.click()
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.DELETE)
        el.send_keys(value)
    except Exception:
        driver.execute_script("arguments[0].value = arguments[1];", el, value)
        driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", el)
    return el


def clicar_primeiro_filtro_metas():
    xpaths = [
        "//button[contains(.,'Filtrar')]",
        "//a[contains(.,'Filtrar')]",
        "//button[@type='submit' and contains(.,'Filtrar')]",
    ]
    ultimo = None
    for xp in xpaths:
        try:
            return clicar_seguro_xpath(xp, timeout=10)
        except Exception as e:
            ultimo = e
    if ultimo:
        raise ultimo



def extrair_linhas_tabela_metas():
    """
    Lê a tabela de metas de forma robusta usando os atributos data-title/data-value.
    O SGI muda os IDs das metas todo mês (ex.: 208, 209, 210, 211), então não usamos ID fixo.
    """
    wait.until(EC.presence_of_all_elements_located((By.XPATH, "//table//tbody/tr")))
    linhas = []

    # Primeira tentativa: via JavaScript, pegando data-title/data-value da própria tabela.
    try:
        rows = driver.execute_script("""
            const out = [];
            document.querySelectorAll('table tbody tr').forEach(tr => {
                const obj = {};
                tr.querySelectorAll('td').forEach((td, idx) => {
                    const title = (td.getAttribute('data-title') || '').trim();
                    const val = (td.getAttribute('data-value') || td.innerText || '').trim();
                    if (title) obj[title] = val;
                    obj['col_' + idx] = val;
                    const a = td.querySelector('a[href]');
                    if (a && !obj['_href_id']) obj['_href_id'] = a.getAttribute('href') || '';
                    if (a && (a.getAttribute('href')||'').includes('/metas/consulta/')) obj['_href_resultado'] = a.getAttribute('href') || '';
                });
                if (Object.keys(obj).length) out.push(obj);
            });
            return out;
        """) or []

        for obj in rows:
            id_meta = str(obj.get('ID') or obj.get('Id') or obj.get('id') or obj.get('col_0') or '').strip()
            id_meta = re.sub(r"\D", "", id_meta)
            if not id_meta:
                continue

            tipo = str(obj.get('Tipo') or obj.get('tipo') or obj.get('col_1') or '').strip()
            escopo = str(obj.get('Escopo') or obj.get('escopo') or obj.get('col_2') or '').strip()
            descricao = str(obj.get('Descrição') or obj.get('Descricao') or obj.get('descricao') or obj.get('col_3') or '').strip()
            data_inicial = str(obj.get('Data Inicial') or obj.get('data inicial') or obj.get('col_4') or '').strip()
            data_final = str(obj.get('Data Final') or obj.get('data final') or obj.get('col_5') or '').strip()

            href_resultado = str(obj.get('_href_resultado') or '').strip()
            if href_resultado.startswith('/'):
                href_resultado = URL + href_resultado
            if not href_resultado:
                href_resultado = f"{URL}/metas/consulta/{id_meta}"

            linhas.append({
                "id": id_meta,
                "tipo": tipo,
                "escopo": escopo,
                "descricao": descricao,
                "data_inicial": data_inicial,
                "data_final": data_final,
                "href_resultado": href_resultado,
            })

        if linhas:
            return linhas
    except Exception as e:
        print(f"⚠️ Leitura JS da tabela de metas falhou, tentando Selenium comum: {e}")

    # Fallback: leitura Selenium tradicional.
    for tr in driver.find_elements(By.XPATH, "//table//tbody/tr"):
        tds = tr.find_elements(By.TAG_NAME, "td")
        if len(tds) < 4:
            continue
        vals = []
        for td in tds:
            vals.append((td.get_attribute('data-value') or td.text or '').strip())
        id_meta = re.sub(r"\D", "", vals[0] if vals else "")
        if not id_meta:
            continue
        item = {
            "id": id_meta,
            "tipo": vals[1] if len(vals) > 1 else "",
            "escopo": vals[2] if len(vals) > 2 else "",
            "descricao": vals[3] if len(vals) > 3 else "",
            "data_inicial": vals[4] if len(vals) > 4 else "",
            "data_final": vals[5] if len(vals) > 5 else "",
            "href_resultado": f"{URL}/metas/consulta/{id_meta}",
        }
        linhas.append(item)
    return linhas


def linha_meta_match(item, spec):
    tipo = normalizar_texto_match(item.get("tipo"))
    escopo = normalizar_texto_match(item.get("escopo"))
    desc = normalizar_texto_match(item.get("descricao"))

    tipo_spec = normalizar_texto_match(spec.get("tipo", ""))
    escopo_spec = normalizar_texto_match(spec.get("escopo", ""))

    if tipo_spec and tipo_spec != tipo:
        return False
    if escopo_spec and escopo_spec != escopo:
        return False

    partes = [normalizar_texto_match(p) for p in spec.get("descricao_contem", [])]
    return all(parte in desc for parte in partes if parte)


def escolher_melhor_linha_meta(linhas, spec):
    """
    Escolhe a meta do mês pela linha exibida na tela.
    Não usa IDs fixos, porque o SGI cria novos IDs todo mês.
    """
    candidatos = [x for x in linhas if linha_meta_match(x, spec)]
    if not candidatos:
        return None

    def id_int(x):
        return int(re.sub(r"\D", "", str(x.get("id") or "0")) or "0")

    candidatos.sort(key=id_int, reverse=True)
    escolhido = dict(candidatos[0])
    id_meta = str(escolhido.get("id", "")).strip()
    if id_meta and not escolhido.get("href_resultado"):
        escolhido["href_resultado"] = f"{URL}/metas/consulta/{id_meta}"
    return escolhido

def limpar_df_tabela(df):
    df = df.copy()
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    return df


def encontrar_df_resultado_metas():
    # V6.0: primeiro tenta ler o HTML inteiro da página. Isso evita erro
    # stale element quando o Sólidus troca/recarrega a tabela no meio da leitura.
    try:
        from io import StringIO
        html_src = driver.page_source or ""
        dfs = pd.read_html(StringIO(html_src))
        candidatos = []
        for _df in dfs:
            try:
                cols_txt = " ".join(str(c) for c in list(_df.columns)).upper()
                body_txt = " ".join(str(x) for x in _df.head(3).values.flatten()).upper()
                all_txt = cols_txt + " " + body_txt
                if "ATINGIDO" in all_txt and "PER" in all_txt and "META" in all_txt and ("REALIZADO" in all_txt or "PROJETADO" in all_txt):
                    candidatos.append(_df)
            except Exception:
                pass
        if candidatos:
            candidatos.sort(key=lambda d: (len(d), len(d.columns)), reverse=True)
            return limpar_df_tabela(candidatos[0])
    except Exception as _e_read_html:
        pass

    # Fallback: lê direto da tabela HTML visível via Selenium.
    possiveis = [
        "//table[@id='tabela_metas']",
        "//table[contains(@class,'tabela-metas')]",
        "(//table[contains(@class,'table') and .//tbody/tr])[1]",
    ]
    tabela = None
    ultimo_erro = None

    for xp in possiveis:
        try:
            tabela = wait.until(EC.presence_of_element_located((By.XPATH, xp)))
            if tabela and tabela.find_elements(By.XPATH, ".//tbody/tr"):
                break
        except Exception as e:
            ultimo_erro = e
            tabela = None

    if tabela is None:
        raise Exception(f"Tabela de resultados das metas não encontrada na tela: {ultimo_erro}")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tabela)
    time.sleep(0.6)

    headers = []
    for th in tabela.find_elements(By.XPATH, ".//thead//th"):
        txt = th.text.strip()
        if txt:
            headers.append(re.sub(r"\s+", " ", txt))

    # Alguns layouts têm a 1ª coluna sem texto no header. Ajusta pelo data-title do tbody.
    linhas = []
    trs = tabela.find_elements(By.XPATH, ".//tbody/tr[td]")
    for tr in trs:
        tds = tr.find_elements(By.XPATH, "./td")
        vals = []
        for td in tds:
            txt = td.text.strip()
            if not txt:
                txt = td.get_attribute("data-value") or ""
            vals.append(re.sub(r"\s+", " ", txt).strip())
        if any(v for v in vals):
            linhas.append(vals)

        # Le tfoot - o SGI coloca a linha de Total aqui
    trs_tfoot = tabela.find_elements(By.XPATH, ".//tfoot/tr[td]")
    for tr in trs_tfoot:
        tds = tr.find_elements(By.XPATH, "./td")
        vals = []
        for td in tds:
            txt = td.text.strip()
            if not txt:
                txt = td.get_attribute("data-value") or ""
            vals.append(re.sub(r"\\s+", " ", txt).strip())
        if any(v for v in vals):
            linhas.append(vals)

    if not linhas:
        raise Exception("Tabela de metas encontrada, mas sem linhas no tbody")

    max_cols = max(len(x) for x in linhas)
    while len(headers) < max_cols:
        idx = len(headers)
        if idx == 0:
            # tenta descobrir pelo data-title da primeira célula
            try:
                data_title = trs[0].find_elements(By.XPATH, "./td")[0].get_attribute("data-title") or ""
                headers.append(data_title.strip() or "Nome")
            except Exception:
                headers.append("Nome")
        else:
            headers.append(f"Coluna_{idx+1}")

    headers = headers[:max_cols]
    linhas = [row + [""] * (max_cols - len(row)) for row in linhas]
    df = pd.DataFrame(linhas, columns=headers)

    # Normaliza nomes esperados das colunas
    ren = {}
    for i, c in enumerate(df.columns):
        uc = normalizar_texto_match(c)
        nome = c
        if uc in ("FILIAL", "VENDEDOR", "SUBGRUPO", "FILIAL/VENDEDOR", "VENDEDOR/SUBGRUPO"):
            nome = c.title()
        elif "META(R$) TOTAL" in uc:
            nome = "Meta(R$) Total"
        elif "REALIZADO(R$) TOTAL" in uc:
            nome = "Realizado(R$) Total"
        elif "ATINGIDO TOTAL" in uc:
            nome = "Atingido Total"
        elif "META(R$) PERIODO" in uc:
            nome = "Meta(R$) Período"
        elif "REALIZADO(R$) PERIODO" in uc:
            nome = "Realizado(R$) Período"
        elif "ATINGIDO PERIODO" in uc:
            nome = "Atingido Período"
        elif "PROJETADO(R$)" in uc:
            nome = "Projetado(R$)"
        ren[c] = nome
    df = df.rename(columns=ren)

    # Garante nomes únicos: ex. Vendedor, Vendedor_2
    cols = []
    usados = {}
    for c in df.columns:
        if c not in usados:
            usados[c] = 1
            cols.append(c)
        else:
            usados[c] += 1
            cols.append(f"{c}_{usados[c]}")
    df.columns = cols
    return limpar_df_tabela(df)


def valor_float_brasil(v):
    try:
        s = str(v).strip()
        if s in ("", "nan", "None"):
            return None
        # Se já é um número puro (ex: 0.1995 vindo da seção numérica do xlsx),
        # não faz replace de "." que quebraria o valor
        s_clean = s.replace("%", "").strip()
        # Detecta formato brasileiro (tem vírgula como decimal): "13.031,84"
        if "," in s_clean and s_clean.index(",") > s_clean.index(".") if "." in s_clean else False:
            # Formato BR: remove pontos de milhar, troca vírgula por ponto
            s_clean = s_clean.replace(".", "").replace(",", ".")
        elif "," in s_clean and "." not in s_clean:
            # Só vírgula: "13031,84" -> "13031.84"
            s_clean = s_clean.replace(",", ".")
        # else: formato puro com ponto: "13031.84" ou "0.1995" — usa direto
        return float(s_clean)
    except Exception:
        return None


def nome_arquivo_contas_valido(fname):
    s = str(fname or "").lower().strip()
    if s.startswith("~$"):
        return False
    if not (s.endswith(".xls") or s.endswith(".xlsx")):
        return False

    # nunca aceitar relatórios de margem ou arquivos auxiliares
    bloqueados = [
        "margem",
        "margens",
        "metas_vendas_mes_atual",
        "dashboard",
        "historico",
        "fechamentos",
        "credenciais",
        "config_meta",
        "cobrancas_log",
    ]
    if any(b in s for b in bloqueados):
        return False

    # aceitar apenas nomes típicos do Contas a Receber
    permitidos = [
        "relatorio_contas",
        "contas_pagar_receber",
        "contas_receber",
    ]
    return any(p in s for p in permitidos)


def _set_input_by_id_safely(input_id, valor):
    el = wait.until(EC.presence_of_element_located((By.ID, input_id)))
    try:
        driver.execute_script("arguments[0].removeAttribute('readonly'); arguments[0].removeAttribute('disabled');", el)
    except Exception:
        pass
    driver.execute_script("arguments[0].value = '';", el)
    if valor is not None and str(valor) != "":
        el.send_keys(str(valor))
    try:
        driver.execute_script("arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", el)
        driver.execute_script("arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));", el)
    except Exception:
        pass
    return el


def _selecionar_situacao_quitados():
    """
    Tenta selecionar Situação = Quitados no relatório de Contas a Receber.
    O SGI pode mudar id/name, então tentamos várias formas.
    """
    tentativas = [
        (By.ID, "situacao"),
        (By.ID, "situacao_titulo"),
        (By.ID, "situacao_conta"),
        (By.NAME, "filtros[situacao]"),
        (By.NAME, "filtros[situacao_titulo]"),
    ]
    for by, sel in tentativas:
        try:
            el = driver.find_element(by, sel)
            s = Select(el)
            # tenta por value primeiro
            for val in ["quitado", "quitados", "Q", "2", "3"]:
                try:
                    s.select_by_value(val)
                    print(f"✅ Situação Quitados OK via {sel}={val}")
                    return True
                except Exception:
                    pass
            # tenta por texto
            for opt in s.options:
                txt = (opt.text or "").strip().lower()
                if "quitad" in txt:
                    s.select_by_visible_text(opt.text)
                    print(f"✅ Situação Quitados OK via texto: {opt.text}")
                    return True
        except Exception:
            pass

    # fallback por label
    try:
        xp = "//label[contains(translate(normalize-space(.),'ÇÃÕÁÉÍÓÚÂÊÔÀÈÌÒÙ','CAOAEIOUAEOA EIOU'),'Situacao')]/following::select[1]"
        el = driver.find_element(By.XPATH, xp)
        s = Select(el)
        for opt in s.options:
            if "quitad" in (opt.text or "").lower():
                s.select_by_visible_text(opt.text)
                print(f"✅ Situação Quitados OK via label: {opt.text}")
                return True
    except Exception:
        pass

    print("⚠️ Não consegui confirmar Situação = Quitados automaticamente")
    return False


def _parse_quitados_xls_180d(caminho_quitados, colmap):
    """
    Lê o relatório de quitados e devolve lista de títulos quitados.
    Mantém vendedor do ERP como referência, mas a atribuição do mérito será feita no dashboard
    cruzando com cobrancas_log.json pelo último cobrador antes do pagamento.
    """
    dfq = pd.read_excel(caminho_quitados, header=None, engine="openpyxl")
    quitados = []
    vend_atual = None
    agora = now_brasilia()

    def _limpa(v):
        s = str(v or "").strip()
        return "" if s in ("nan", "None") else s

    def _limpar_nome_erp_local(txt):
        try:
            return limpar_nome_erp(txt)
        except NameError:
            s = re.sub(r"^Vendedor:\s*", "", str(txt).strip(), flags=re.IGNORECASE)
            s = re.sub(r"^\d+\s*-\s*", "", s)
            s = re.sub(r"^C\.\d+\s*-\s*", "", s, flags=re.IGNORECASE)
            return s.strip()

    def _limpar_nome_display_local(txt):
        try:
            return limpar_nome_display(txt)
        except NameError:
            s = re.sub(r"\s*\(GER\s*F\d+\)\s*$", "", str(txt).strip(), flags=re.IGNORECASE)
            s = re.sub(r"\s*\(F\d+\)\s*$", "", s, flags=re.IGNORECASE)
            return s.strip()

    for i in range(len(dfq)):
        row = dfq.iloc[i]
        if observacao_deve_ignorar(row, colmap):
            continue
        c0 = _limpa(row[colmap["filial"]]) if len(row) > colmap["filial"] else ""
        c1 = _limpa(row[colmap["cliente"]]) if len(row) > colmap["cliente"] else ""

        if "Vendedor:" in c0:
            vend_atual = _limpar_nome_erp_local(c0)
            continue

        if not vend_atual:
            continue
        if c1 in ("", "Cliente", "Filial"):
            continue
        if c0.upper().startswith("TOTAL") or c1.upper().startswith("TOTAL"):
            continue

        venc = _parse_data(row[colmap["vencimento"]]) if len(row) > colmap["vencimento"] else None
        pagto = _parse_data(row[colmap["pagamento"]]) if len(row) > colmap["pagamento"] else None
        if not venc or not pagto:
            continue

        pago = tratar_valor(row[colmap["pago_total"]]) if len(row) > colmap["pago_total"] else 0.0
        if pago <= 0:
            continue

        try:
            dias_atraso_pg = max(0, (pagto - venc).days)
        except Exception:
            dias_atraso_pg = 0

        if dias_atraso_pg >= 60:
            faixa = "grave"
        elif 30 <= dias_atraso_pg < 60:
            faixa = "alerta"
        elif 15 <= dias_atraso_pg < 30:
            faixa = "atencao"
        else:
            # quitou antes de entrar na regra de cobrança
            continue

        filial_txt = c0
        mfil = re.search(r"FILIAL\s*(\d+)", filial_txt.upper())
        filial_key = f"F{int(mfil.group(1))}" if mfil else ""

        quitados.append({
            "cliente": c1[:80],
            "vendedor_erp": _limpar_nome_display_local(vend_atual),
            "filial": filial_key,
            "filial_label": filial_txt,
            "lancamento": _limpa(row[colmap["num_lancamento"]]) if len(row) > colmap["num_lancamento"] else "",
            "titulo": _limpa(row[colmap["num_titulo"]]) if len(row) > colmap["num_titulo"] else "",
            "parcela": _limpa(row[colmap["num_parcela"]]) if len(row) > colmap["num_parcela"] else "",
            "vencimento": _limpa(row[colmap["vencimento"]]) if len(row) > colmap["vencimento"] else "",
            "pagamento": _limpa(row[colmap["pagamento"]]) if len(row) > colmap["pagamento"] else "",
            "pago": round(float(pago), 2),
            "dias_atraso_pagamento": dias_atraso_pg,
            "faixa": faixa,
            "forma_pagamento": _limpa(row[colmap["forma_pagamento"]]) if len(row) > colmap["forma_pagamento"] else "",
            "conta_caixa": _limpa(row[colmap["conta_caixa"]]) if len(row) > colmap["conta_caixa"] else "",
            "origem": "relatorio_quitados_180d",
        })

    print(f"✅ Quitados 180d lidos: {len(quitados)} título(s)")
    return quitados


def coletar_quitados_180d_contas_receber():
    """
    Relatório complementar para conciliar cobranças feitas com pagamentos que saíram da faixa 15-90.
    Filtros:
    - Vencimento inicial em branco
    - Vencimento final = hoje
    - Situação = Quitados
    - + Data último pagamento inicial = hoje - 180 dias
    - Data último pagamento final = hoje
    - Formas de pagamento iguais ao relatório atual
    """
    data_hoje = now_brasilia().strftime("%d/%m/%Y")
    data_180 = (now_brasilia() - timedelta(days=180)).strftime("%d/%m/%Y")

    print("\n🔎 Iniciando relatório complementar de QUITADOS 180 dias...")
    driver.get(URL + "/relatorio_contas_receber")
    wait.until(EC.presence_of_element_located((By.ID, "data_vencimento_inicial")))
    Select(driver.find_element(By.ID, "data_vencimento")).select_by_value("intervalo")
    time.sleep(1.5)

    # Vencimento inicial em branco / final hoje
    _set_input_by_id_safely("data_vencimento_inicial", "")
    _set_input_by_id_safely("data_vencimento_final", data_hoje)

    # Situação = Quitados
    _selecionar_situacao_quitados()

    # Filiais
    try:
        driver.find_element(By.XPATH, "//label[contains(text(),'Filiais')]/following::button[1]").click()
        time.sleep(1.5)
        filiais_sel = ["1", "2", "3", "4", "5", "6", "7", "8", "10"]
        container = driver.find_element(By.XPATH, "//label[contains(text(),'Filiais')]/following::ul[1]")
        for c in container.find_elements(By.XPATH, ".//input[@type='checkbox']"):
            if c.get_attribute("value") in filiais_sel:
                driver.execute_script("arguments[0].click();", c)
        print("✅ Filiais OK quitados")
        time.sleep(1.5)
    except Exception as e:
        print(f"⚠️ Erro filiais quitados: {e}")

    # Abre +
    try:
        clicar_seguro_xpath("//span[contains(@class,'glyphicon-plus')]", timeout=20)
        print("✅ Mais filtros quitados")
        time.sleep(1.5)
    except Exception as e:
        print(f"⚠️ Não abriu + quitados: {e}")

    # Data último pagamento
    try:
        _set_input_by_id_safely("data_ultimo_pagamento_inicial", data_180)
        _set_input_by_id_safely("data_ultimo_pagamento_final", data_hoje)
        print(f"✅ Data último pagamento: {data_180} até {data_hoje}")
    except Exception as e:
        print(f"⚠️ Erro datas último pagamento: {e}")

    # Formas de pagamento
    try:
        marcar_multiselect_por_label("Forma de Pagamento", ["3", "47", "17"], timeout=20, limpar_antes=True)
        print("✅ Formas OK quitados")
        time.sleep(1.5)
    except Exception as e:
        print(f"⚠️ Erro formas quitados: {e}")

    try:
        Select(wait.until(EC.presence_of_element_located((By.ID, "_formato")))).select_by_value("xls")
        print("✅ Formato XLS quitados")
        time.sleep(1.5)
    except Exception as e:
        print(f"⚠️ Erro formato XLS quitados: {e}")

    arquivos_antes = set(f for f in os.listdir(download_dir) if nome_arquivo_contas_valido(f))
    driver.find_element(By.ID, "gerar").click()
    print("📥 Gerando XLS quitados... aguardando download...")

    caminho_q = None
    for tentativa in range(75):
        time.sleep(2)
        baixando = any(f.endswith(".crdownload") or f.endswith(".tmp") for f in os.listdir(download_dir))
        if tentativa % 5 == 0:
            print(f"⏳ Quitados tentativa {tentativa}/75 | baixando={baixando}")
        if baixando:
            continue
        novos = set(f for f in os.listdir(download_dir) if nome_arquivo_contas_valido(f)) - arquivos_antes
        if novos:
            caminho_q = max([os.path.join(download_dir, f) for f in novos], key=os.path.getctime)
            print(f"✅ Download quitados OK: {caminho_q}")
            break

    if not caminho_q:
        print("⚠️ Nenhum XLS de quitados baixado. Seguindo sem conciliação extra.")
        return {"xlsx_path": None, "json_path": None, "dados": {"quitados": [], "erro": "download_nao_encontrado"}}

    quitados = _parse_quitados_xls_180d(caminho_q, COL)
    out_json = os.path.join(pasta, "quitados_180d_contas_receber.json")
    out_xlsx = os.path.join(pasta, "quitados_180d_contas_receber.xlsx")
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump({
            "gerado_em": now_brasilia().isoformat(),
            "periodo_pagamento": {"inicio": data_180, "fim": data_hoje},
            "criterio": "quitados pagos em ate 180 dias apos cobranca, conciliados no dashboard pelo ultimo cobrador antes do pagamento",
            "quitados": quitados,
        }, f, ensure_ascii=False, indent=2)
    try:
        pd.DataFrame(quitados).to_excel(out_xlsx, index=False)
    except Exception:
        out_xlsx = None

    print(f"💾 Quitados JSON: {out_json}")
    if out_xlsx:
        print(f"💾 Quitados XLSX: {out_xlsx}")

    return {"xlsx_path": out_xlsx, "json_path": out_json, "dados": {"quitados": quitados}}


def _is_total_row(reg):
    """Detecta se uma linha é a linha de Total (tfoot ou primeira coluna contém "Total")."""
    # Verifica colunas conhecidas de identificação
    for k in ["Filial", "Vendedor", "Subgrupo", "Nome", "col_0"]:
        v = str(reg.get(k, "") or "").strip()
        if re.search(r"\btotal\b", v, re.IGNORECASE):
            return True
    # Verifica a primeira coluna de dados (não meta)
    for k, v in reg.items():
        if k.startswith("_"):
            continue
        sv = str(v or "").strip()
        if re.search(r"\btotal\b", sv, re.IGNORECASE):
            return True
        break  # só verifica a primeira coluna de dados
    return False


def df_meta_para_registros(df, spec, origem):
    saida = []
    for _, row in df.iterrows():
        reg = {}
        for c in df.columns:
            v = row[c]
            if isinstance(v, pd.Series):
                v = " | ".join("" if pd.isna(x) else str(x).strip() for x in v.tolist())
            reg[str(c)] = "" if pd.isna(v) else str(v).strip()
        reg["_meta_chave"] = spec["chave"]
        reg["_meta_label"] = spec["label"]
        reg["_meta_tipo"] = origem.get("tipo")
        reg["_meta_escopo"] = origem.get("escopo")
        reg["_meta_descricao"] = origem.get("descricao")
        reg["_meta_id"] = origem.get("id")
        reg["_meta_href"] = origem.get("href_resultado")
        # Marca linha de total (tfoot)
        reg["_is_total"] = _is_total_row(reg)
        for k in list(reg.keys()):
            if "META(R$)" in k.upper() or "REALIZADO(R$)" in k.upper() or "PROJETADO(R$)" in k.upper():
                reg[k + "_float"] = valor_float_brasil(reg[k])
            if "ATINGIDO" in k.upper():
                reg[k + "_float"] = valor_float_brasil(reg[k])
        saida.append(reg)
    return saida




def _extrair_totais_do_xlsx(xlsx_path):
    """
    Extrai os totais das abas de metas a partir das células finais do XLSX.
    Regra pedida pelo usuário:
      - venda_filial_meta           -> linha 11
      - servico_filial_ouro_fob     -> linha 10
      - venda_filial_subgrupo_20k   -> linha 10

    Mapeamento usado:
      B = Meta Total
      D = Atingido Total (%)
      E = Meta Período
      F = Realizado (valor exibido no card)
      H = Projetado

    Se a célula fixa não existir, faz fallback para localizar a última linha "Total"
    ou somar as linhas de filial.
    """
    from openpyxl import load_workbook

    def _br_float(v):
        try:
            s = str(v).strip()
            if s in ("", "None", "nan"):
                return 0.0
            s = s.replace("R$", "").replace("%", "").strip()
            if "," in s:
                s = s.replace(".", "").replace(",", ".")
            return float(s)
        except Exception:
            return 0.0

    def _sheet_by_prefix(wb, nome_base):
        for s in wb.sheetnames:
            if s == nome_base or s.startswith(nome_base[:20]):
                return wb[s]
        return None

    def _fallback_from_sheet(ws):
        # 1) tenta achar a última linha "Total"
        max_row = ws.max_row
        for r in range(max_row, 1, -1):
            a = str(ws.cell(r, 1).value or "").strip().upper()
            if a == "TOTAL":
                return {
                    "meta_total":  round(_br_float(ws.cell(r, 2).value), 2),
                    "real_total":  round(_br_float(ws.cell(r, 6).value), 2),
                    "ating_total": round(_br_float(ws.cell(r, 4).value), 2),
                    "meta_per":    round(_br_float(ws.cell(r, 5).value), 2),
                    "real_per":    round(_br_float(ws.cell(r, 6).value), 2),
                    "proj":        round(_br_float(ws.cell(r, 8).value), 2),
                }

        # 2) fallback final: soma linhas FILIAL
        meta_total = real_total = meta_per = proj = 0.0
        for r in range(2, max_row + 1):
            a = str(ws.cell(r, 1).value or "").strip().upper()
            if a.startswith("FILIAL"):
                meta_total += _br_float(ws.cell(r, 2).value)
                real_total += _br_float(ws.cell(r, 6).value)
                meta_per   += _br_float(ws.cell(r, 5).value)
                proj       += _br_float(ws.cell(r, 8).value)
        ating = round((real_total / meta_total * 100.0), 2) if meta_total > 0 else 0.0
        return {
            "meta_total": round(meta_total, 2),
            "real_total": round(real_total, 2),
            "ating_total": ating,
            "meta_per": round(meta_per, 2),
            "real_per": round(real_total, 2),
            "proj": round(proj, 2),
        }

    _MAP = {
        "venda_filial_meta":       {"sheet": "venda_filial_meta", "row": 11},
        "servico_filial_ouro_fob": {"sheet": "servico_filial_ouro_fob", "row": 10},
        "venda_filial_subgrupo_20k": {"sheet": "venda_filial_subgrupo_20k", "row": 10},
    }
    out = {}

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"⚠️ Não consegui abrir xlsx de metas: {e}")
        return out

    for chave, cfg in _MAP.items():
        ws = _sheet_by_prefix(wb, cfg["sheet"])
        if ws is None:
            print(f"⚠️ Sheet não encontrada no xlsx: {cfg['sheet']}")
            out[chave] = {"meta_total": 0.0, "real_total": 0.0, "ating_total": 0.0, "meta_per": 0.0, "real_per": 0.0, "proj": 0.0}
            continue

        r = int(cfg["row"])
        a = str(ws.cell(r, 1).value or "").strip().upper()
        # Usa a linha fixa se houver conteúdo e parecer ser a linha final/total
        if a in ("TOTAL",) or ws.cell(r, 2).value not in (None, ""):
            meta_total = round(_br_float(ws.cell(r, 2).value), 2)
            real_total = round(_br_float(ws.cell(r, 6).value), 2)
            ating_total = round(_br_float(ws.cell(r, 4).value), 2)
            meta_per = round(_br_float(ws.cell(r, 5).value), 2)
            real_per = round(_br_float(ws.cell(r, 6).value), 2)
            proj = round(_br_float(ws.cell(r, 8).value), 2)
            item = {
                "meta_total": meta_total,
                "real_total": real_total,
                "ating_total": ating_total,
                "meta_per": meta_per,
                "real_per": real_per,
                "proj": proj,
            }
        else:
            item = _fallback_from_sheet(ws)

        out[chave] = item
        print(f"✅ Total xlsx [{chave}]: meta={item['meta_total']:.2f}  real={item['real_total']:.2f}  ating={item['ating_total']:.2f}%  proj={item['proj']:.2f}")

    return out



def coletar_metas_vendas_mes_atual():
    primeiro_dia = hoje.replace(day=1).strftime("%d/%m/%Y")
    prox_mes = (hoje.replace(day=28) + timedelta(days=4)).replace(day=1)
    ultimo_dia = (prox_mes - timedelta(days=1)).strftime("%d/%m/%Y")

    print("\n📊 Iniciando coleta de metas/vendas do mês atual...")
    driver.get(URL + "/metas")
    wait.until(EC.presence_of_element_located((By.ID, "data_inicial_maior_igual")))
    set_input_value_safe(By.ID, "data_inicial_maior_igual", primeiro_dia)
    set_input_value_safe(By.ID, "data_final_menor_igual", ultimo_dia)
    clicar_primeiro_filtro_metas()
    time.sleep(2)

    linhas = extrair_linhas_tabela_metas()
    print(f"📋 Metas listadas na tela: {len(linhas)}")

    # IDs das metas mudam todo mês. Buscar sempre pelo Tipo + Escopo + Descrição da tela filtrada.
    specs = [
        {"chave": "venda_filial_meta", "label": "Venda / Filial / Meta Filial", "tipo": "Venda", "escopo": "Filial", "descricao_contem": ["META FILIAL"]},
        {"chave": "venda_filial_vendedor_meta", "label": "Venda / Filial-Vendedor / Meta Vendedor", "tipo": "Venda", "escopo": "Filial/Vendedor", "descricao_contem": ["META VENDEDOR"]},
        {"chave": "venda_filial_subgrupo_20k", "label": "Venda / Filial-Subgrupo / Caminhão 20K", "tipo": "Venda", "escopo": "Filial/Subgrupo", "descricao_contem": ["CAMINHAO", "20K"]},
        {"chave": "servico_filial_ouro_fob", "label": "Serviço / Filial / Ouro-FOB", "tipo": "Serviço", "escopo": "Filial", "descricao_contem": ["OURO", "FOB"]},
        {"chave": "servico_filial_vendedor_ouro_fob", "label": "Serviço / Filial-Vendedor / Ouro-FOB", "tipo": "Serviço", "escopo": "Filial/Vendedor", "descricao_contem": ["OURO", "FOB"]},
        {"chave": "venda_vendedor_subgrupo_20k", "label": "Venda / Vendedor-Subgrupo / Caminhão 20K", "tipo": "Venda", "escopo": "Vendedor/Subgrupo", "descricao_contem": ["CAMINHAO", "20K"]},
    ]

    resultados = {
        "coletado_em": now_brasilia().strftime("%Y-%m-%d %H:%M:%S"),
        "periodo_meta": {"data_inicial": primeiro_dia, "data_final": ultimo_dia},
        "lista_index": linhas,
        "metas": {}
    }

    xlsx_path = os.path.join(pasta, "metas_vendas_mes_atual.xlsx")
    json_path = os.path.join(pasta, "metas_vendas_mes_atual.json")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        pd.DataFrame(linhas).to_excel(writer, sheet_name="index_metas", index=False)

        for spec in specs:
            origem = escolher_melhor_linha_meta(linhas, spec)
            if not origem:
                print(f"⚠️ Meta não encontrada: {spec['label']}")
                resultados["metas"][spec["chave"]] = {"ok": False, "erro": "meta_nao_encontrada", "spec": spec}
                continue

            try:
                print(f"➡️ Abrindo resultados: {spec['label']} -> {origem['href_resultado']}")
                driver.get(origem["href_resultado"])
                time.sleep(2)
                df_result = encontrar_df_resultado_metas()
                df_result = limpar_df_tabela(df_result)
                print(f"   ↳ Colunas lidas: {list(df_result.columns)}")
                print(f"   ↳ Primeiras linhas: {len(df_result)}")
                resultados["metas"][spec["chave"]] = {
                    "ok": True,
                    "spec": spec,
                    "origem": origem,
                    "colunas": list(df_result.columns),
                    "linhas": df_meta_para_registros(df_result, spec, origem),
                }
                sheet = spec["chave"][:31]
                df_result.to_excel(writer, sheet_name=sheet, index=False)
                print(f"✅ {spec['label']}: {len(df_result)} linha(s)")
            except Exception as e:
                print(f"⚠️ Erro em {spec['label']}: {e}")
                resultados["metas"][spec["chave"]] = {
                    "ok": False,
                    "spec": spec,
                    "origem": origem,
                    "erro": str(e),
                }

    # ── Pós-coleta: relê o xlsx para extrair totais numéricos limpos ──────────────
    # O xlsx já está fechado aqui. Relemos e injetamos a linha Total em cada spec
    # que a precisa, garantindo que _sales_emp receba os valores corretos.
    _totais_xlsx = _extrair_totais_do_xlsx(xlsx_path)
    for _chave_t, _tot in _totais_xlsx.items():
        _meta_obj_t = resultados["metas"].get(_chave_t)
        if not isinstance(_meta_obj_t, dict) or not _meta_obj_t.get("ok"):
            continue
        _linhas_t = _meta_obj_t.get("linhas") or []
        # Remove qualquer linha Total anterior (evita duplicata)
        _linhas_t = [_r for _r in _linhas_t if not _r.get("_is_total")]
        # Constrói linha Total canônica com todos os campos _float prontos
        _total_reg = {
            "Filial": "Total",
            "Meta(R$) Total":           str(_tot["meta_total"]),
            "Realizado(R$) Total":       str(_tot["real_total"]),
            "Atingido Total":            f"{_tot['ating_total']:.2f}%",
            "Meta(R$) Período":          str(_tot["meta_per"]),
            "Realizado(R$) Período":     str(_tot["real_per"]),
            "Projetado(R$)":             str(_tot["proj"]),
            "Meta(R$) Total_float":      _tot["meta_total"],
            "Realizado(R$) Total_float": _tot["real_total"],
            "Atingido Total_float":      _tot["ating_total"],
            "Meta(R$) Período_float":    _tot["meta_per"],
            "Realizado(R$) Período_float": _tot["real_per"],
            "Projetado(R$)_float":       _tot["proj"],
            "_is_total": True,
            "_meta_chave": _chave_t,
        }
        _linhas_t.append(_total_reg)
        resultados["metas"][_chave_t]["linhas"] = _linhas_t
    # ── Fim do pós-coleta ──────────────────────────────────────────────────────────

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)

    print(f"💾 Metas JSON: {json_path}")
    print(f"💾 Metas XLSX: {xlsx_path}")
    return {"json_path": json_path, "xlsx_path": xlsx_path, "dados": resultados}




# =========================================
# 🎯 META DIÁRIA PELO CONTROLE DE META DO SÓLIDUS
# Regra V5.3:
#   - Não usa mais dias úteis do dashboard para bater meta diária.
#   - Abre o mesmo Controle de Meta do SGI/Sólidus.
#   - Em /metas/consulta/{id}, filtra Consulta de hoje até hoje.
#   - Tipo Venda, escopo Filial e Filial/Vendedor.
#   - Quem tiver Atingido Período >= 100% entra no mural Meta diária BATIDA.
# =========================================

def _clicar_consultar_resultado_metas():
    xpaths = [
        "//button[contains(normalize-space(.),'Consultar')]",
        "//a[contains(normalize-space(.),'Consultar')]",
        "//input[@type='submit' and contains(@value,'Consultar')]",
        "//button[@type='submit']",
    ]
    ultimo = None
    for xp in xpaths:
        try:
            btn = wait.until(EC.element_to_be_clickable((By.XPATH, xp)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
            driver.execute_script("arguments[0].click();", btn)
            time.sleep(1.8)
            return True
        except Exception as e:
            ultimo = e
    if ultimo:
        print(f"⚠️ Não consegui clicar em Consultar no resultado da meta: {ultimo}")
    return False


def _set_periodo_consulta_meta_dia(data_br):
    """Seta Data Inicial/Final da área Consulta no resultado da meta."""
    # IDs enviados pelo usuário no print do Sólidus.
    try:
        set_input_value_safe(By.ID, "data_inicial_consulta", data_br)
        set_input_value_safe(By.ID, "data_final_consulta", data_br)
        _clicar_consultar_resultado_metas()
        return True
    except Exception as e:
        print(f"⚠️ IDs data_inicial_consulta/data_final_consulta não funcionaram: {e}")

    # Fallback por name/atributo, caso o Sólidus mude algo.
    try:
        ini = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='data_inicial_consulta' or @atributo='data_inicial_consulta']")))
        fim = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@name='data_final_consulta' or @atributo='data_final_consulta']")))
        for el in (ini, fim):
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            try:
                el.click(); el.send_keys(Keys.CONTROL, 'a'); el.send_keys(Keys.DELETE); el.send_keys(data_br)
            except Exception:
                driver.execute_script("arguments[0].value = arguments[1];", el, data_br)
            driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", el)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", el)
        _clicar_consultar_resultado_metas()
        return True
    except Exception as e:
        print(f"⚠️ Fallback datas da meta diária falhou: {e}")
        return False


def coletar_metas_vendas_dia_atual(metas_index=None):
    data_dia = hoje.strftime("%d/%m/%Y")
    print(f"\n🎯 Iniciando coleta de META DIÁRIA pelo Sólidus ({data_dia})...")

    # Reaproveita o índice já lido pela coleta mensal quando possível.
    linhas = []
    try:
        linhas = list((metas_index or {}).get("lista_index") or [])
    except Exception:
        linhas = []

    if not linhas:
        try:
            primeiro_dia = hoje.replace(day=1).strftime("%d/%m/%Y")
            prox_mes = (hoje.replace(day=28) + timedelta(days=4)).replace(day=1)
            ultimo_dia = (prox_mes - timedelta(days=1)).strftime("%d/%m/%Y")
            driver.get(URL + "/metas")
            wait.until(EC.presence_of_element_located((By.ID, "data_inicial_maior_igual")))
            set_input_value_safe(By.ID, "data_inicial_maior_igual", primeiro_dia)
            set_input_value_safe(By.ID, "data_final_menor_igual", ultimo_dia)
            clicar_primeiro_filtro_metas()
            time.sleep(2)
            linhas = extrair_linhas_tabela_metas()
        except Exception as e:
            print(f"⚠️ Não consegui carregar índice de metas para meta diária: {e}")
            linhas = []

    specs = [
        # VENDA DIÁRIA: usada no mural de meta diária e no card Venda Diária.
        {"chave": "venda_filial_meta", "label": "Meta diária / Venda / Filial", "tipo": "Venda", "escopo": "Filial", "descricao_contem": ["META FILIAL"]},
        {"chave": "venda_filial_vendedor_meta", "label": "Meta diária / Venda / Filial-Vendedor", "tipo": "Venda", "escopo": "Filial/Vendedor", "descricao_contem": ["META VENDEDOR"]},

        # V6.5: SERVIÇO DIÁRIO também precisa ser coletado hoje-hoje.
        # Antes o card Venda Diária somava venda do dia com diferença de snapshot de serviço,
        # o que podia mostrar valores acumulados antigos como se fossem do dia.
        {"chave": "servico_filial_ouro_fob", "label": "Meta diária / Serviço / Filial", "tipo": "Serviço", "escopo": "Filial", "descricao_contem": ["OURO", "FOB"]},
        {"chave": "servico_filial_vendedor_ouro_fob", "label": "Meta diária / Serviço / Filial-Vendedor", "tipo": "Serviço", "escopo": "Filial/Vendedor", "descricao_contem": ["OURO", "FOB"]},
    ]

    resultados = {
        "coletado_em": now_brasilia().strftime("%Y-%m-%d %H:%M:%S"),
        "data_consulta": data_dia,
        "fonte": "SGI/Solidus /metas/consulta filtrado por data atual",
        "versao_coleta": "V6.5",
        "regra": "Somente colunas do PERIODO após Consulta hoje-hoje; nunca usa Total para mural de meta diária.",
        "lista_index": linhas,
        "metas": {},
    }

    xlsx_path = os.path.join(pasta, "metas_vendas_dia_atual.xlsx")
    json_path = os.path.join(pasta, "metas_vendas_dia_atual.json")

    # V5.5: guarda uma cópia da coleta anterior para fallback caso o Sólidus recarregue a tabela
    # no meio da leitura e gere stale element reference em apenas uma das metas.
    _metas_dia_anterior = {}
    try:
        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as _fmda:
                _metas_dia_anterior = json.load(_fmda) or {}
    except Exception:
        _metas_dia_anterior = {}

    try:
        writer_ctx = pd.ExcelWriter(xlsx_path, engine="openpyxl")
    except Exception:
        writer_ctx = None

    if writer_ctx:
        writer = writer_ctx.__enter__()
        try:
            pd.DataFrame(linhas).to_excel(writer, sheet_name="index_metas", index=False)
        except Exception:
            pass
    else:
        writer = None

    try:
        for spec in specs:
            origem = escolher_melhor_linha_meta(linhas, spec)
            if not origem:
                print(f"⚠️ Meta diária não encontrada: {spec['label']}")
                resultados["metas"][spec["chave"]] = {"ok": False, "erro": "meta_nao_encontrada", "spec": spec}
                continue
            _ultimo_erro_meta_dia = None
            _sucesso_meta_dia = False
            for _tentativa_meta_dia in range(1, 4):
                try:
                    print(f"➡️ Meta diária: abrindo {spec['label']} -> {origem['href_resultado']} (tentativa {_tentativa_meta_dia}/3)")
                    driver.get(origem["href_resultado"])
                    time.sleep(1.8)
                    _set_periodo_consulta_meta_dia(data_dia)
                    time.sleep(2.2)
                    df_result = encontrar_df_resultado_metas()
                    df_result = limpar_df_tabela(df_result)
                    linhas_resultado = df_meta_para_registros(df_result, spec, origem)
                    resultados["metas"][spec["chave"]] = {
                        "ok": True,
                        "spec": spec,
                        "origem": origem,
                        "data_consulta": data_dia,
                        "coleta_diaria_v59": True,
                        "colunas": list(df_result.columns),
                        "linhas": linhas_resultado,
                    }
                    if writer is not None:
                        df_result.to_excel(writer, sheet_name=spec["chave"][:31], index=False)

                    batidos = 0
                    for r in linhas_resultado:
                        if r.get("_is_total"):
                            continue
                        ating = 0.0
                        for k, v in r.items():
                            if "ATINGIDO" in str(k).upper() and "PER" in str(k).upper() and str(k).endswith("_float"):
                                ating = max(ating, float(v or 0))
                        if ating >= 100:
                            batidos += 1
                    print(f"✅ {spec['label']}: {len(df_result)} linha(s), {batidos} acima de 100% no período do dia")
                    _sucesso_meta_dia = True
                    break
                except Exception as e:
                    _ultimo_erro_meta_dia = e
                    print(f"⚠️ Erro meta diária {spec['label']} tentativa {_tentativa_meta_dia}/3: {e}")
                    time.sleep(1.2 * _tentativa_meta_dia)
            if not _sucesso_meta_dia:
                # Fallback: usa a última coleta válida do mesmo dia para não apagar vendedores/filiais do mural.
                _prev_meta = (((_metas_dia_anterior or {}).get("metas") or {}).get(spec["chave"]) or {})
                # V5.8: só usa fallback se a coleta anterior também foi V5.7.
                # Isso evita reaproveitar JSON antigo onde o mural pegava Total/mês em vez do Período do dia.
                if (
                    _prev_meta.get("ok")
                    and _prev_meta.get("coleta_diaria_v59")
                    and str((_metas_dia_anterior or {}).get("data_consulta") or "") == str(data_dia)
                    and str((_metas_dia_anterior or {}).get("versao_coleta") or "") in ("V5.9", "V6.0", "V6.2", "V6.3", "V6.4", "V6.5")
                ):
                    resultados["metas"][spec["chave"]] = _prev_meta
                    try:
                        print(f"♻️ Meta diária {spec['label']}: usando última coleta válida do dia por fallback ({len(_prev_meta.get('linhas') or [])} linha(s)).")
                    except Exception:
                        pass
                else:
                    resultados["metas"][spec["chave"]] = {"ok": False, "spec": spec, "origem": origem, "erro": str(_ultimo_erro_meta_dia)}
    finally:
        if writer_ctx:
            writer_ctx.__exit__(None, None, None)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)
    print(f"💾 Meta diária Sólidus JSON: {json_path}")
    print(f"💾 Meta diária Sólidus XLSX: {xlsx_path}")
    return {"json_path": json_path, "xlsx_path": xlsx_path, "dados": resultados}

# =========================================
# 📈 RELATÓRIO DE MARGEM BRUTA / RENTABILIDADE
# Coleta 2 relatórios do SGI:
#   1) Lucratividade por FILIAL
#   2) Lucratividade por VENDEDOR
# Alimenta o dashboard com Margem Bruta (%) do mês atual.
# =========================================

def _select_by_text_contains(select_el, texto_alvo):
    texto_alvo_n = normalizar_texto_match(texto_alvo)
    sel = Select(select_el)
    for opt in sel.options:
        if texto_alvo_n in normalizar_texto_match(opt.text):
            sel.select_by_visible_text(opt.text)
            try:
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", select_el)
            except Exception:
                pass
            return True
    return False


def _label_contains_xpath(txt):
    return f"//label[contains(normalize-space(.),'{txt}')] | //strong[contains(normalize-space(.),'{txt}')] | //div[contains(normalize-space(.),'{txt}') and not(*)]"


def _find_relatorio_block_by_label(label_texto):
    ultimo = None
    xp = _label_contains_xpath(label_texto)
    candidatos = [
        f"({xp})[1]/ancestor::div[contains(@class,'row') or contains(@class,'col') or contains(@class,'form-group')][1]",
        f"({xp})[1]/parent::*",
        f"({xp})[1]/ancestor::div[1]",
    ]
    for cp in candidatos:
        try:
            bloco = wait.until(EC.presence_of_element_located((By.XPATH, cp)))
            if bloco:
                return bloco
        except Exception as e:
            ultimo = e
    raise ultimo or Exception(f"Bloco do label {label_texto} não encontrado")


def _set_relatorio_periodo_emissao(data_inicial, data_final):
    bloco = _find_relatorio_block_by_label('Período de Emissão')
    try:
        sels = [s for s in bloco.find_elements(By.TAG_NAME, 'select') if s.is_displayed()]
        if sels:
            try:
                Select(sels[0]).select_by_value('intervalo')
            except Exception:
                _select_by_text_contains(sels[0], 'Intervalo')
            driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", sels[0])
    except Exception:
        pass

    inputs = [i for i in bloco.find_elements(By.XPATH, ".//input[not(@type='hidden')]") if i.is_displayed()]
    if len(inputs) < 2:
        # fallback geral: pega os 2 primeiros inputs visíveis de data/texto do formulário
        form = wait.until(EC.presence_of_element_located((By.XPATH, "//form[contains(@action,'relatorio_margens_brutas')]")))
        inputs = [i for i in form.find_elements(By.XPATH, ".//input[not(@type='hidden')]") if i.is_displayed()]
    if len(inputs) < 2:
        raise Exception('Não encontrei os 2 campos de data do relatório de margens')

    for el, valor in [(inputs[0], data_inicial), (inputs[1], data_final)]:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        try:
            el.click()
            el.send_keys(Keys.CONTROL, 'a')
            el.send_keys(Keys.DELETE)
            el.send_keys(valor)
        except Exception:
            driver.execute_script("arguments[0].value = arguments[1];", el, valor)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", el)
        driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", el)


def _select_relatorio_por_label(label_texto, texto_opcao=None, value=None):
    bloco = _find_relatorio_block_by_label(label_texto)
    selects = [s for s in bloco.find_elements(By.TAG_NAME, 'select') if s.is_displayed()]
    if not selects:
        # fallback: primeiro select visível após o label no DOM
        xp = f"({_label_contains_xpath(label_texto)})[1]/following::select[1]"
        selects = [wait.until(EC.presence_of_element_located((By.XPATH, xp)))]
    el = selects[0]
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    sel = Select(el)
    if value is not None:
        try:
            sel.select_by_value(str(value))
        except Exception:
            pass
    if texto_opcao is not None:
        if not _select_by_text_contains(el, texto_opcao):
            try:
                sel.select_by_visible_text(texto_opcao)
            except Exception:
                # último fallback: tentativa por option text normalizado
                achou = False
                alvo = normalizar_texto_match(texto_opcao)
                for opt in sel.options:
                    if alvo == normalizar_texto_match(opt.text):
                        sel.select_by_visible_text(opt.text)
                        achou = True
                        break
                if not achou:
                    raise Exception(f'Opção {texto_opcao} não encontrada em {label_texto}')
    driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", el)
    return el



def _tratar_valor_margem(v):
    try:
        s = str(v).strip()
        if s in ("", "nan", "None"):
            return 0.0
        s = s.replace("R$", "").replace("%", "").strip()
        s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


def _limpar_nome_margem(nome):
    s = str(nome or "").strip()
    s = re.sub(r"^Vendedor:\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"^\d+\s*-\s*", "", s)
    s = re.sub(r"^C\.\d+\s*-\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*\(GER\s*F\d+\)\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*\(F\d+\)\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _norm_key_margem(nome):
    s = _limpar_nome_margem(nome).upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _filial_key_from_text(texto):
    s = str(texto or "").upper()
    m = re.search(r"FILIAL\s*0*(\d{1,3})", s)
    if not m:
        return ""
    n = int(m.group(1))
    return "FDEP" if n in (90, 99) else f"F{n}"


def _select_all_multiselect_by_label(label_texto):
    abrir_multiselect_por_label(label_texto, timeout=20)
    xpath_ul = f"//label[contains(normalize-space(.),'{label_texto}')]/following::ul[1]"
    container = wait.until(EC.presence_of_element_located((By.XPATH, xpath_ul)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", container)
    time.sleep(0.5)
    checkboxes = container.find_elements(By.XPATH, ".//input[@type='checkbox']")
    for c in checkboxes:
        try:
            if not c.is_selected():
                driver.execute_script("arguments[0].click();", c)
                time.sleep(0.05)
        except Exception:
            pass
    driver.execute_script("document.body.click();")
    time.sleep(0.5)
    return len(checkboxes)


def _nome_arquivo_margem_valido(fname):
    s = str(fname or '').lower()
    if s.startswith('~$'):
        return False
    if not (s.endswith('.xls') or s.endswith('.xlsx')):
        return False
    return ('margem' in s) or ('margens' in s)


def _ler_xls_ou_xlsx(caminho):
    erros = []
    for engine in ('openpyxl', None):
        try:
            if engine:
                return pd.read_excel(caminho, header=None, engine=engine)
            return pd.read_excel(caminho, header=None)
        except Exception as e:
            erros.append(f'read_excel[{engine}]: {e}')
    try:
        tabelas = pd.read_html(caminho, decimal=',', thousands='.')
        if tabelas:
            return tabelas[0]
    except Exception as e:
        erros.append(f'read_html: {e}')
    raise Exception('Não consegui ler XLS/XLSX de margem: ' + ' | '.join(erros))


def _parse_relatorio_margem_xls(caminho, tipo):
    df_m = _ler_xls_ou_xlsx(caminho).fillna('')
    filiais = {}
    vendedores = {}
    empresa = {}

    for _, row in df_m.iterrows():
        vals = [str(x).strip() for x in row.tolist()]
        vals += [''] * max(0, 8 - len(vals))

        if tipo == 'filial':
            filial_txt = vals[0]
            norm0 = normalizar_texto_match(filial_txt)
            if not filial_txt or norm0 == 'FILIAL':
                continue
            if norm0 == 'TOTAL':
                empresa = {
                    'label': 'TOTAL',
                    'margem_bruta_pct': round(_tratar_valor_margem(vals[5]), 2),
                    'valor_total_liquido': _tratar_valor_margem(vals[1]),
                    'valor_total': _tratar_valor_margem(vals[2]),
                    'custo_total': _tratar_valor_margem(vals[3]),
                    'margem_bruta_valor': _tratar_valor_margem(vals[4]),
                    'markup_realizado': _tratar_valor_margem(vals[6]) if len(vals) > 6 else 0.0,
                    'pmr_vendas_prazo': _tratar_valor_margem(vals[7]) if len(vals) > 7 else 0.0,
                }
                continue
            if 'FILIAL' not in norm0:
                continue
            fk = _filial_key_from_text(filial_txt)
            if not fk:
                continue
            margem_pct = _tratar_valor_margem(vals[5])
            filiais[fk] = {
                'filial': fk,
                'label': filial_txt,
                'margem_bruta_pct': round(margem_pct, 2),
                'valor_total_liquido': _tratar_valor_margem(vals[1]),
                'valor_total': _tratar_valor_margem(vals[2]),
                'custo_total': _tratar_valor_margem(vals[3]),
                'margem_bruta_valor': _tratar_valor_margem(vals[4]),
                'markup_realizado': _tratar_valor_margem(vals[6]) if len(vals) > 6 else 0.0,
            }

        else:
            filial_txt = vals[0]
            vendedor_txt = vals[1]
            if not vendedor_txt or normalizar_texto_match(vendedor_txt) in ('VENDEDOR', 'TOTAL'):
                continue
            fk = _filial_key_from_text(filial_txt)
            if not fk:
                continue
            nome = _limpar_nome_margem(vendedor_txt)
            if not nome:
                continue
            margem_pct = _tratar_valor_margem(vals[6])
            key = f'{_norm_key_margem(nome)}_{fk}'
            vendedores[key] = {
                'nome': nome,
                'filial': fk,
                'key': key,
                'label_filial': filial_txt,
                'margem_bruta_pct': round(margem_pct, 2),
                'valor_total_liquido': _tratar_valor_margem(vals[2]),
                'valor_total': _tratar_valor_margem(vals[3]),
                'custo_total': _tratar_valor_margem(vals[4]),
                'margem_bruta_valor': _tratar_valor_margem(vals[5]),
                'markup_realizado': _tratar_valor_margem(vals[7]) if len(vals) > 7 else 0.0,
            }

    return {'filiais': filiais, 'vendedores': vendedores, 'empresa': empresa}


def _gerar_relatorio_margem(tipo):
    primeiro_dia = hoje.replace(day=1).strftime("%d/%m/%Y")
    prox_mes = (hoje.replace(day=28) + timedelta(days=4)).replace(day=1)
    ultimo_dia = (prox_mes - timedelta(days=1)).strftime("%d/%m/%Y")

    driver.get(URL + "/relatorio_margens_brutas")
    wait.until(EC.presence_of_element_located((By.ID, "gerar")))
    time.sleep(1.5)
    esperar_sumir_overlays(10)

    _set_relatorio_periodo_emissao(primeiro_dia, ultimo_dia)

    try:
        _select_relatorio_por_label("Tipo de Custo", texto_opcao="Custo Médio Gerencial")
    except Exception as e:
        print(f"⚠️ Não consegui selecionar Tipo de Custo automaticamente: {e}")

    try:
        _select_relatorio_por_label("Lucratividade por", texto_opcao=("Filial" if tipo == "filial" else "Vendedor"))
    except Exception as e:
        print(f"⚠️ Não consegui selecionar Lucratividade por {tipo}: {e}")

    try:
        total_ops = _select_all_multiselect_by_label("Operações de Devolução")
        print(f"✅ Operações de Devolução: {total_ops} opções marcadas")
    except Exception as e:
        print(f"⚠️ Não consegui marcar Operações de Devolução: {e}")

    try:
        Select(wait.until(EC.presence_of_element_located((By.ID, "_formato")))).select_by_value("xls")
        print("✅ Formato XLS para Margem Bruta")
    except Exception as e:
        print(f"⚠️ Não consegui selecionar formato XLS de Margem Bruta: {e}")

    arquivos_antes = set(f for f in os.listdir(download_dir) if _nome_arquivo_margem_valido(f))

    gerar_btn = wait.until(EC.presence_of_element_located((By.ID, 'gerar')))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", gerar_btn)
    esperar_sumir_overlays(10)
    try:
        wait.until(EC.element_to_be_clickable((By.ID, 'gerar'))).click()
    except Exception:
        driver.execute_script("arguments[0].click();", gerar_btn)
    print(f"📈 Gerando Margem Bruta por {tipo.upper()}...")

    caminho = None
    for _ in range(60):
        time.sleep(2)
        baixando = any(f.endswith('.crdownload') or f.endswith('.tmp') for f in os.listdir(download_dir))
        if baixando:
            continue
        novos = set(f for f in os.listdir(download_dir) if _nome_arquivo_margem_valido(f)) - arquivos_antes
        if novos:
            caminho = max([os.path.join(download_dir, f) for f in novos], key=os.path.getctime)
            break

    if not caminho:
        todos = [os.path.join(download_dir, f) for f in os.listdir(download_dir) if _nome_arquivo_margem_valido(f)]
        if not todos:
            raise Exception('Nenhum arquivo XLS/XLSX de margem foi baixado')
        caminho = max(todos, key=os.path.getctime)
        print(f"⚠️ Usando último arquivo de margem encontrado: {caminho}")
    else:
        print(f"✅ Download margem OK: {caminho}")

    return _parse_relatorio_margem_xls(caminho, tipo)


def coletar_margens_brutas_mes_atual():
    print("\n📈 Iniciando coleta de Margem Bruta/Rentabilidade do mês atual...")
    json_path = os.path.join(pasta, "margens_brutas_mes_atual.json")
    xlsx_path = os.path.join(pasta, "margens_brutas_mes_atual.xlsx")
    resultados = {
        "coletado_em": now_brasilia().strftime("%Y-%m-%d %H:%M:%S"),
        "periodo": {
            "data_inicial": hoje.replace(day=1).strftime("%d/%m/%Y"),
            "data_final": ((hoje.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)).strftime("%d/%m/%Y"),
        },
        "filiais": {},
        "vendedores": {},
        "empresa": {},
        "ok": False,
    }

    try:
        res_fil = _gerar_relatorio_margem("filial")
        resultados["filiais"] = res_fil.get("filiais", {})
        resultados["empresa"] = res_fil.get("empresa", {})
        print(f"✅ Margem por FILIAL: {len(resultados['filiais'])} registro(s)")
    except Exception as e:
        resultados["erro_filial"] = str(e)
        print(f"⚠️ Erro ao coletar margem por FILIAL: {e}")

    try:
        res_vend = _gerar_relatorio_margem("vendedor")
        resultados["vendedores"] = res_vend.get("vendedores", {})
        print(f"✅ Margem por VENDEDOR: {len(resultados['vendedores'])} registro(s)")
    except Exception as e:
        resultados["erro_vendedor"] = str(e)
        print(f"⚠️ Erro ao coletar margem por VENDEDOR: {e}")

    resultados["ok"] = bool(resultados.get("filiais") or resultados.get("vendedores"))

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)

    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            pd.DataFrame(list(resultados.get("filiais", {}).values())).to_excel(writer, sheet_name="filiais", index=False)
            pd.DataFrame([resultados.get("empresa", {})]).to_excel(writer, sheet_name="empresa", index=False)
            pd.DataFrame(list(resultados.get("vendedores", {}).values())).to_excel(writer, sheet_name="vendedores", index=False)
    except Exception as e:
        print(f"⚠️ Não consegui salvar XLSX de margens: {e}")
        xlsx_path = None

    print(f"💾 Margens JSON: {json_path}")
    if xlsx_path:
        print(f"💾 Margens XLSX: {xlsx_path}")
    return {"json_path": json_path, "xlsx_path": xlsx_path, "dados": resultados}



# ===== LOGIN
driver.get(URL)
time.sleep(8)
print("TITLE:", driver.title)
print("URL ATUAL:", driver.current_url)

campo_usuario = None
seletores_usuario = [
    (By.NAME, "usuario"),
    (By.ID, "usuario"),
    (By.CSS_SELECTOR, "input[name='usuario']"),
    (By.CSS_SELECTOR, "input[id='usuario']"),
    (By.XPATH, "//input[contains(@name,'usuario') or contains(@id,'usuario')]"),
    (By.XPATH, "//input[@type='text']"),
]

for by, sel in seletores_usuario:
    try:
        campo_usuario = WebDriverWait(driver, 8).until(
            EC.presence_of_element_located((by, sel))
        )
        if campo_usuario:
            print(f"✅ Campo usuário encontrado em: {by} = {sel}")
            break
    except Exception:
        pass

if not campo_usuario:
    print("❌ Campo usuário não encontrado")
    print("URL:", driver.current_url)
    print("TITLE:", driver.title)
    try:
        print(driver.page_source[:5000])
    except Exception:
        pass
    try:
        driver.save_screenshot(os.path.join(download_dir, "debug_login.png"))
    except Exception:
        pass
    raise Exception("Campo usuário não encontrado na tela de login")

try:
    campo_usuario.click()
    campo_usuario.send_keys(Keys.CONTROL, "a")
    campo_usuario.send_keys(Keys.DELETE)
except Exception:
    pass
campo_usuario.send_keys(LOGIN)

senha_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='password']")))
senha_field.send_keys(SENHA)
senha_field.send_keys(Keys.ENTER)

try:
    wait.until(EC.element_to_be_clickable((By.ID, "botao_prosseguir_informa_local_trabalho"))).click()
    print("✅ Filial OK")
except Exception:
    print("⚠️ Tela de filial não apareceu")
time.sleep(5)

# ===== RELATÓRIO
driver.get(URL + "/relatorio_contas_receber")
wait.until(EC.presence_of_element_located((By.ID, "data_vencimento_inicial")))
try:
    _sel_data_venc = driver.find_element(By.ID, "data_vencimento")
    Select(_sel_data_venc).select_by_value("intervalo")
    driver.execute_script("arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", _sel_data_venc)
except Exception as _e_sel_data:
    print(f"⚠️ Erro selecionando intervalo de vencimento: {_e_sel_data}")
time.sleep(2)

# V10.6: seta datas com eventos change/blur. Em algumas execuções headless o SGI
# não registrava o valor quando era apenas value='' + send_keys, e o download não começava.
_set_input_by_id_safely("data_vencimento_inicial", data_inicio)
_set_input_by_id_safely("data_vencimento_final", data_fim)
print(f"🧾 Filtro contas a receber: vencimento {data_inicio} até {data_fim}")

# Filiais
try:
    driver.find_element(By.XPATH, "//label[contains(text(),'Filiais')]/following::button[1]").click()
    time.sleep(2)
    filiais_sel = ["1", "2", "3", "4", "5", "6", "7", "8", "10"]
    container = driver.find_element(By.XPATH, "//label[contains(text(),'Filiais')]/following::ul[1]")

    for c in container.find_elements(By.XPATH, ".//input[@type='checkbox']"):
        if c.get_attribute("value") in filiais_sel:
            driver.execute_script("arguments[0].click();", c)

    print("✅ Filiais OK")
    time.sleep(3)
except Exception as e:
    print(f"⚠️ Erro filiais: {e}")

# 🔥 PULA TOTALIZAR POR VENDEDOR
# O XLS já virá totalizado corretamente, então vai direto pro botão "+"
try:
    clicar_seguro_xpath("//span[contains(@class,'glyphicon-plus')]", timeout=20)
    print("✅ Mais filtros")
    time.sleep(2)
except Exception as e:
    print(f"⚠️ Não abriu +: {e}")

# ===== FORMAS DE PAGAMENTO
try:
    marcar_multiselect_por_label("Forma de Pagamento", ["3", "47", "17"], timeout=20, limpar_antes=True)
    print("✅ Formas OK")
    time.sleep(2)
except Exception as e:
    print(f"⚠️ Erro formas: {e}")

try:
    Select(wait.until(EC.presence_of_element_located((By.ID, "_formato")))).select_by_value("xls")
    print("✅ Formato XLS")
    time.sleep(2)
except Exception as e:
    print(f"⚠️ Erro formato XLS: {e}")

# Conta Caixa já vem no relatório automaticamente — sem ação Selenium necessária

# ===== GERAR E AGUARDAR DOWNLOAD
def _listar_xls_contas_download():
    try:
        return [f for f in os.listdir(download_dir) if nome_arquivo_contas_valido(f)]
    except Exception:
        return []

def _tem_download_em_andamento():
    try:
        return any(f.endswith(".crdownload") or f.endswith(".tmp") for f in os.listdir(download_dir))
    except Exception:
        return False

def _clicar_gerar_relatorio_principal():
    btn = wait.until(EC.presence_of_element_located((By.ID, "gerar")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    esperar_sumir_overlays(15)
    try:
        wait.until(EC.element_to_be_clickable((By.ID, "gerar"))).click()
    except Exception:
        driver.execute_script("arguments[0].click();", btn)
    try:
        driver.execute_script("arguments[0].dispatchEvent(new MouseEvent('click', {bubbles:true,cancelable:true,view:window}));", btn)
    except Exception:
        pass

def _aguardar_xls_contas(arquivos_antes, inicio_click_ts, timeout_seg=240):
    fim_wait = time.time() + timeout_seg
    ultimo_print = 0
    while time.time() < fim_wait:
        time.sleep(2)
        if _tem_download_em_andamento():
            if time.time() - ultimo_print > 10:
                print("⏳ Ainda baixando XLS de Contas a Receber...")
                ultimo_print = time.time()
            continue
        atuais = set(_listar_xls_contas_download())
        novos = atuais - set(arquivos_antes)
        candidatos = []
        for f in novos:
            p = os.path.join(download_dir, f)
            try:
                if os.path.getsize(p) > 0:
                    candidatos.append(p)
            except Exception:
                pass
        # fallback: se o navegador reusou nome/sobrescreveu algum arquivo, pega XLS criado após o clique
        for f in atuais:
            p = os.path.join(download_dir, f)
            try:
                if os.path.getctime(p) >= (inicio_click_ts - 3) and os.path.getsize(p) > 0 and p not in candidatos:
                    candidatos.append(p)
            except Exception:
                pass
        if candidatos:
            return max(candidatos, key=os.path.getctime)
        if time.time() - ultimo_print > 20:
            print(f"⏳ Aguardando XLS... arquivos atuais válidos={len(atuais)}")
            ultimo_print = time.time()
    return None

arquivos_antes = set(_listar_xls_contas_download())
caminho = None
for _tent_dl in range(1, 4):
    _inicio_click = time.time()
    print(f"📥 Gerando XLS de Contas a Receber... tentativa {_tent_dl}/3")
    try:
        _clicar_gerar_relatorio_principal()
    except Exception as _e_click_dl:
        print(f"⚠️ Falha ao clicar Gerar na tentativa {_tent_dl}: {_e_click_dl}")
    caminho = _aguardar_xls_contas(arquivos_antes, _inicio_click, timeout_seg=180 if _tent_dl < 3 else 300)
    if caminho:
        print(f"✅ Download OK: {caminho}")
        break
    try:
        print(f"⚠️ Nenhum XLS encontrado na tentativa {_tent_dl}. URL atual: {driver.current_url} | TITLE: {driver.title}")
        print("📂 Arquivos atuais:", os.listdir(download_dir))
        driver.save_screenshot(os.path.join(download_dir, f"debug_sem_xls_contas_tentativa_{_tent_dl}.png"))
    except Exception as _e_diag_dl:
        print(f"⚠️ Falha ao registrar diagnóstico do download: {_e_diag_dl}")
    time.sleep(3)

if not caminho:
    print(f"📂 download_dir: {download_dir}")
    try:
        print("📂 Arquivos atuais:", os.listdir(download_dir))
    except Exception as e:
        print(f"⚠️ Não consegui listar download_dir: {e}")
    try:
        _html_debug_path = os.path.join(download_dir, "debug_sem_xls_contas_page.html")
        with open(_html_debug_path, "w", encoding="utf-8", errors="ignore") as _fh_dbg:
            _fh_dbg.write(driver.page_source or "")
        print(f"🧪 Debug HTML salvo: {_html_debug_path}")
    except Exception:
        pass

    driver.quit()
    raise Exception("Nenhum XLS válido de Contas a Receber foi baixado nesta execução")

# ===== LEITURA DO XLS
if not nome_arquivo_contas_valido(os.path.basename(caminho)):
    raise Exception(f"Arquivo selecionado não é o relatório de contas a receber: {caminho}")
df_raw = pd.read_excel(caminho, header=None, engine="openpyxl")
print(f"📋 {df_raw.shape[0]} linhas lidas")

metas_vendas_info = {"json_path": None, "xlsx_path": None, "dados": {}}
try:
    metas_vendas_info = coletar_metas_vendas_mes_atual()
except Exception as e:
    print(f"⚠️ Erro ao coletar metas/vendas do SGI: {e}")

metas_vendas_dia_info = {"json_path": None, "xlsx_path": None, "dados": {}}
try:
    metas_vendas_dia_info = coletar_metas_vendas_dia_atual(metas_vendas_info.get("dados", {}))
except Exception as e:
    print(f"⚠️ Erro ao coletar meta diária SGI/Sólidus: {e}")

margens_brutas_info = {"json_path": None, "xlsx_path": None, "dados": {}}
try:
    margens_brutas_info = coletar_margens_brutas_mes_atual()
except Exception as e:
    print(f"⚠️ Erro ao coletar Margem Bruta/Rentabilidade do SGI: {e}")

# ===== MAPA DE COLUNAS DO XLS NOVO
COL = {
    "filial": 0,
    "cliente": 1,
    "contato": 2,
    "conta_caixa": 3,
    "restricao_credito": 4,
    "historico": 5,
    "num_lancamento": 6,
    "forma_pagamento": 7,
    "prev_real": 8,
    "num_titulo": 9,
    "num_parcela": 10,
    "emissao": 11,
    "vencimento": 12,
    "pagamento": 13,
    "nominal": 14,
    "pendente": 15,
    "pago_total": 16,
    "juros_total": 17,
    # Relatório novo: S=Observações (18) e T=Avalistas (19).
    # Relatório antigo: S=Avalistas (18). Detecta pela quantidade de colunas.
    "observacoes": 18 if df_raw.shape[1] > 19 else None,
    "avalistas": 19 if df_raw.shape[1] > 19 else 18,
}
print(f"🧭 Colunas detectadas: Observações={COL.get('observacoes')} Avalistas={COL.get('avalistas')}")


def _row_val_safe(row, idx):
    try:
        if idx is None:
            return ""
        if idx < 0 or idx >= len(row):
            return ""
        s = str(row[idx]).strip()
        return "" if s in ("nan", "None") else s
    except Exception:
        return ""


def _norm_obs_block(txt):
    s = unicodedata.normalize("NFKD", str(txt or "").upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def observacao_deve_ignorar(row, colmap=None):
    """Ignora títulos ASPIS/SISPUMUT/SISPUMUTI pela coluna Observações."""
    cm = colmap or COL
    obs_idx = cm.get("observacoes")
    if obs_idx is None:
        return False
    obs = _norm_obs_block(_row_val_safe(row, obs_idx))
    if not obs:
        return False
    return any(term in obs for term in ("ASPIS", "SISPUMUT", "SISPUMUTI", "SISPUM", "SISPU"))


# ===== HELPERS
def tratar_valor(v):
    try:
        v = str(v).strip()
        if v in ("","nan","None"): return 0.0
        return float(v.replace(".","").replace(",","."))
    except:
        return 0.0


def extrair_filial_nome(nome):
    """
    Gerente: (GERFx) → filial='Fx', is_gerente=True
    Ativo:   (Fx)    → filial='Fx', is_gerente=False
    Inativo: sem tag → None, False
    """
    s = str(nome).upper()
    m = re.search(r"\(GER\s*F(\d+)\)", s)
    if m: return f"F{int(m.group(1))}", True
    m = re.search(r"\(F(\d+)\)", s)
    if m: return f"F{int(m.group(1))}", False
    return None, False


def limpar_nome_erp(texto):
    """Remove prefixo numérico do ERP"""
    s = re.sub(r"^Vendedor:\s*", "", str(texto).strip(), flags=re.IGNORECASE)
    s = re.sub(r"^\d+\s*-\s*", "", s)
    s = re.sub(r"^C\.\d+\s*-\s*", "", s, flags=re.IGNORECASE)
    return s.strip()


def limpar_nome_display(nome):
    """Remove sufixo (Fx) ou (GERFx)"""
    s = re.sub(r"\s*\(GER\s*F\d+\)\s*$", "", str(nome).strip(), flags=re.IGNORECASE)
    s = re.sub(r"\s*\(F\d+\)\s*$",       "", s,                  flags=re.IGNORECASE)
    return s.strip()


def normalizar_login(texto):
    """Primeiro token 100% letras"""
    for token in str(texto).strip().split():
        if token.isalpha():
            nfkd = unicodedata.normalize("NFKD", token)
            return "".join(c for c in nfkd if not unicodedata.combining(c)).lower()

    nfkd = unicodedata.normalize("NFKD", str(texto).strip())
    sem  = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"[^a-zA-Z]","",sem)[:10].lower() or "vendedor"


# =========================================
# 🔥 NOVOS HELPERS (WHATSAPP + CONTATO)
# =========================================

def extrair_telefones(texto):
    """
    Extrai múltiplos números de telefone de um campo
    """
    s = str(texto or "").strip()

    if not s or s.lower() in ("nan", "none"):
        return []

    # separa por ; , / |
    partes = re.split(r"[;,/|]+", s)

    telefones = []

    for parte in partes:
        numero = "".join(re.findall(r"\d", parte))

        if len(numero) >= 10:
            if not numero.startswith("55"):
                numero = "55" + numero

            telefones.append(numero)

    # remove duplicados mantendo ordem
    vistos = set()
    unicos = []

    for n in telefones:
        if n not in vistos:
            vistos.add(n)
            unicos.append(n)

    return unicos


def primeiro_nome_cliente(nome):
    nome = str(nome or "").strip()
    if not nome:
        return "Cliente"
    return nome.split()[0].title()


def mensagem_cobranca_padrao(cliente, vencimento, valor_pendente):
    primeiro_nome = primeiro_nome_cliente(cliente)

    valor_fmt = f"R$ {float(valor_pendente):,.2f}"
    valor_fmt = valor_fmt.replace(",", "X").replace(".", ",").replace("X", ".")

    return (
        f"Olá, {primeiro_nome} tudo bem? Aqui é da Lojas MDL - Móveis do Lar. "
        f"Passando para lembrar que tem uma parcelinha vencida na data de {vencimento}, "
        f"no valor de {valor_fmt}. Caso o pagamento já tenha sido realizado, por gentileza, "
        f"desconsidere esta mensagem. Se precisar do boleto, chave PIX ou tiver qualquer dúvida, "
        f"fico à disposição para ajudar. Agradecemos a atenção e desejamos um excelente dia."
    )

# ===== PARSE LINHA A LINHA
registros = []
vendedor_atual = None

for i in range(len(df_raw)):
    row = df_raw.iloc[i]
    if observacao_deve_ignorar(row):
        continue
    cel0 = str(row[COL["filial"]]).strip()

    if "Vendedor:" in cel0:
        vendedor_atual = limpar_nome_erp(cel0)
        continue

    if vendedor_atual and cel0.upper().startswith("FILIAL"):
        m = re.search(r"FILIAL\s+(\d+)", cel0.upper())
        if not m:
            continue

        pendente = tratar_valor(row[COL["pendente"]])
        pago = tratar_valor(row[COL["pago_total"]])

        registros.append({
            "vendedor": vendedor_atual,
            "filial_num": int(m.group(1)),
            "pendente": pendente,
            "pago": pago,
        })





df = pd.DataFrame(registros)

# =========================================
# 👥 CAMADA ADMINISTRATIVA DE STATUS DE COLABORADORES (V5.6)
# Fonte: credenciais_dashboard.json -> colaborador_status.
# Objetivo: quando alguém sair/entrar/trocar função, o Master controla se o usuário
# participa de cobrança, clientes sem movimento, aniversariantes e murais, sem depender
# apenas da tag (F2)/(GERF2) no nome do SGI.
# Em MODO_TESTE_LOCAL=1 o script usa o JSON local e NÃO tenta publicar no FTP.
# =========================================
def _colab_norm_key_py(txt):
    s = unicodedata.normalize("NFKD", str(txt or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _colab_status_key_py(nome, filial, is_gerente=False):
    tipo = "GERENTE" if bool(is_gerente) else "VENDEDOR"
    return f"{_colab_norm_key_py(limpar_nome_display(limpar_nome_erp(nome)))}|{str(filial or '').strip().upper()}|{tipo}"

def _colab_default_status_py(login='', nome='', filial='', is_gerente=False):
    return {
        "login": str(login or '').strip().lower(),
        "nome": limpar_nome_display(limpar_nome_erp(nome)),
        "filial": str(filial or '').strip().upper(),
        "tipo": "Gerente" if bool(is_gerente) else "Vendedor",
        "status": "ativo",
        "participa_cobrancas": True,
        "participa_sem_movimento": True,
        "participa_aniversariantes": True,
        "participa_murais": True,
        "data_entrada": "",
        "data_saida": "",
        "substituto": "",
        "obs": "",
    }

def _load_colab_status_state_py():
    path = os.path.join(pasta, "credenciais_dashboard.json")
    data = {}
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
    except Exception:
        data = {}
    raw = data.get("colaborador_status") if isinstance(data, dict) else {}
    return raw if isinstance(raw, dict) else {}

COLAB_STATUS_MAP = _load_colab_status_state_py()

# V5.8: localiza status por nome+filial para bloquear desligados mesmo quando
# o usuário vem de metas/vendas e não da carteira de cobrança.
def _colab_status_by_nome_filial_py(nome, filial):
    try:
        nn = _colab_norm_key_py(limpar_nome_display(limpar_nome_erp(nome)))
        ff = str(filial or '').strip().upper()
        for _k, _st in (COLAB_STATUS_MAP or {}).items():
            if not isinstance(_st, dict):
                continue
            if _colab_norm_key_py(_st.get('nome') or '') == nn and str(_st.get('filial') or '').strip().upper() == ff:
                base = dict(_st)
                status = str(base.get('status') or 'ativo').lower().strip()
                base['status'] = 'inativo' if status in ('inativo','desligado','bloqueado','0','false') else 'ativo'
                return base
    except Exception:
        pass
    return None

def _colab_esta_liberado_py(nome, filial, is_gerente=False, flag='participa_cobrancas'):
    try:
        st = _colab_get_status_py(nome, filial, is_gerente)
        alt = _colab_status_by_nome_filial_py(nome, filial) or {}
        if alt:
            st.update(alt)
        return st.get('status') == 'ativo' and bool(st.get(flag, True))
    except Exception:
        return True

def _colab_get_status_py(nome, filial, is_gerente=False):
    key = _colab_status_key_py(nome, filial, is_gerente)
    st = COLAB_STATUS_MAP.get(key) or {}
    if not isinstance(st, dict):
        st = {}
    base = _colab_default_status_py('', nome, filial, is_gerente)
    base.update(st)
    # compatibilidade com valores antigos
    status = str(base.get('status') or 'ativo').lower().strip()
    base['status'] = 'inativo' if status in ('inativo','desligado','bloqueado','0','false') else 'ativo'
    for flag in ('participa_cobrancas','participa_sem_movimento','participa_aniversariantes','participa_murais'):
        base[flag] = bool(base.get(flag, True))
    return base

def _colab_participa_cobranca_py(nome, filial, is_gerente=False):
    st = _colab_get_status_py(nome, filial, is_gerente)
    return st.get('status') == 'ativo' and bool(st.get('participa_cobrancas', True))

# Totaliza BRUTO (deve bater com "Total Geral" do arquivo)
total_bruto_p  = df["pendente"].sum()
total_bruto_pg = df["pago"].sum()
print(f"\n📊 TOTAL BRUTO (deve bater com Total Geral do XLS):")
print(f"   Pendente : R$ {total_bruto_p:>12,.2f}")
print(f"   Pago     : R$ {total_bruto_pg:>12,.2f}")

# ===== IDENTIFICAR TIPO DE CADA VENDEDOR
df[["filial_vendedor","is_gerente"]] = df["vendedor"].apply(
    lambda n: pd.Series(extrair_filial_nome(n))
)

# Filial do registro (onde está no ERP)
df["filial_erp"] = df["filial_num"].apply(
    lambda n: "FDEP" if n in (90,99) else f"F{n}"
)

# V5.6: aplica status administrativo ANTES do rateio.
# Se o Master marcar um vendedor/gerente como inativo ou sem participação em cobrança,
# a tag (F2)/(GERF2) é ignorada nesta execução e a carteira dele entra como inativa
# para ser redistribuída aos ativos da mesma filial.
_colab_bloqueados_rateio = []
try:
    for _idx_colab, _row_colab in df.iterrows():
        _fil_tag_colab = _row_colab.get("filial_vendedor")
        if pd.notna(_fil_tag_colab) and str(_fil_tag_colab).strip():
            _nome_colab = str(_row_colab.get("vendedor") or "")
            _is_ger_colab = bool(_row_colab.get("is_gerente"))
            _st_colab = _colab_get_status_py(_nome_colab, _fil_tag_colab, _is_ger_colab)
            if _st_colab.get('status') != 'ativo' or not bool(_st_colab.get('participa_cobrancas', True)):
                df.at[_idx_colab, "filial_erp"] = str(_fil_tag_colab).strip().upper()
                df.at[_idx_colab, "filial_vendedor"] = None
                df.at[_idx_colab, "is_gerente"] = False
                _colab_bloqueados_rateio.append(f"{limpar_nome_display(_nome_colab)} ({_fil_tag_colab})")
    if _colab_bloqueados_rateio:
        print("👥 Status administrativo aplicado no rateio: " + "; ".join(_colab_bloqueados_rateio[:20]) + ("..." if len(_colab_bloqueados_rateio)>20 else ""))
except Exception as _e_colab_status:
    print(f"⚠️ Não consegui aplicar status administrativo no rateio: {_e_colab_status}")

# Filial definitiva:
# - ativo/gerente → usa a filial do nome (ex: SANDY (F1) → F1)
# - inativo       → usa a filial do ERP (onde estão os créditos)
df["filial"] = df.apply(
    lambda r: r["filial_vendedor"] if r["filial_vendedor"] else r["filial_erp"],
    axis=1
)

# Separa ANTES de agrupar para não perder inativos com múltiplas filiais
df_ativos_raw   = df[df["filial_vendedor"].notna()].copy()
df_inativos_raw = df[df["filial_vendedor"].isna()].copy()

# Lookup de filial real para inativos (usando df já processado)
_filial_inativo_lookup = {}
_nomes_fdep = set()
for _, _rr in df_inativos_raw.iterrows():
    _nd_i = limpar_nome_display(limpar_nome_erp(str(_rr["vendedor"])))
    _fi   = str(_rr.get("filial_erp","")).strip()
    if _fi and _fi not in ("nan","None",""):
        if _fi == "FDEP":
            _nomes_fdep.add(_nd_i)
        _filial_inativo_lookup[_nd_i] = _fi


# ===== PARSE CLIENTES E RECEBIMENTOS POR FAIXA
# Lógica:
#   recebido_faixa  = pago nos títulos com pagto > data_corte (delta do período)
#                     inclui ativos + inativos + FDEP (para rateio na meta)
#   clientes_cobrar = títulos PENDENTES por faixa (para cobrança)
from datetime import datetime as _dt
_hoje = _dt.now()

def _parse_data(v):
    try:
        s = str(v).strip()
        if s in ("nan", "None", ""):
            return None
        return _dt.strptime(s, "%d/%m/%Y")
    except:
        return None

# ===== RELATÓRIO COMPLEMENTAR: QUITADOS 180 DIAS
# IMPORTANTE: roda aqui porque neste ponto os helpers limpar_nome_erp, limpar_nome_display,
# tratar_valor e _parse_data já foram definidos.
quitados_180_info = {"json_path": None, "xlsx_path": None, "dados": {"quitados": []}}
try:
    quitados_180_info = coletar_quitados_180d_contas_receber()
except Exception as e:
    print(f"⚠️ Erro ao coletar quitados 180d: {e}")
    quitados_180_info = {"json_path": None, "xlsx_path": None, "dados": {"quitados": [], "erro": str(e)}}


clientes_cobrar   = []
recebido_faixa    = {}   # {key: {grave, alerta, atencao, is_ativo, is_fdep, filial}}
clientes_key_hoje = set()
_vend_cli_atual   = None

# Data de corte para recebimentos do período (V10.5):
# A meta de recebimentos por faixa é MENSAL. Portanto, durante todo o mês
# a cobrança deve somar pagamentos desde o dia 01 da competência.
# Nunca usa snapshot futuro (ex.: 01/07 em execução de 30/06) nem "último snapshot
# disponível" como corte, pois isso zera os recebimentos de junho antes da virada.
import json as _json_tmp

_mes_atual_str = os.getenv("COBRANCA_COMPETENCIA", now_brasilia().strftime("%Y-%m")).strip()[:7]
if not re.match(r"^\d{4}-\d{2}$", _mes_atual_str):
    _mes_atual_str = now_brasilia().strftime("%Y-%m")

_month_start_str = f"{_mes_atual_str}-01"
_month_start_parse = _dt.strptime(_month_start_str, "%Y-%m-%d")
_data_corte_parse = _month_start_parse
_ref_path = os.path.join(pasta, "cache_historico", f"snapshot_referencia_{_mes_atual_str}.json")
os.makedirs(os.path.dirname(_ref_path), exist_ok=True)

_ref_precisa_regravar = True
try:
    if os.path.exists(_ref_path):
        with open(_ref_path, encoding="utf-8") as _f_tmp:
            _ref_tmp = _json_tmp.load(_f_tmp) or {}
        _ref_data_raw = str(_ref_tmp.get("data") or "")[:10]
        if _ref_data_raw != _month_start_str:
            print(f"⚠️ Referência mensal inválida para {_mes_atual_str}: {_ref_data_raw or 'vazia'} → corrigindo para {_month_start_str}")
        else:
            _ref_precisa_regravar = False
except Exception as _e_ref_mes:
    print(f"⚠️ Não consegui ler referência mensal; vou recriar: {_e_ref_mes}")

if _ref_precisa_regravar:
    try:
        with open(_ref_path, "w", encoding="utf-8") as _f_tmp:
            _json_tmp.dump({
                "data": _month_start_str,
                "gerado_em": now_brasilia().isoformat(),
                "origem": "inicio_mes_forcado_v10_7",
                "observacao": "Corte mensal de recebimentos por faixa; evita reset antes da virada do mês."
            }, _f_tmp, ensure_ascii=False, indent=2)
    except Exception as _e_ref_write:
        print(f"⚠️ Não consegui gravar referência mensal V10.5: {_e_ref_write}")

print(f"📅 Referência do mês (base): {_data_corte_parse.strftime('%d/%m/%Y')} → acumulando recebimentos desde início do mês")

for _i in range(len(df_raw)):
    _row = df_raw.iloc[_i]
    if observacao_deve_ignorar(_row):
        continue

    _c0 = str(_row[COL["filial"]]).strip()
    _c1 = str(_row[COL["cliente"]]).strip()
    _contato = str(_row[COL["contato"]]).strip() if len(_row) > COL["contato"] else ""
    _avalista = _row_val_safe(_row, COL.get("avalistas"))

    if "Vendedor:" in _c0:
        _vend_cli_atual = limpar_nome_erp(_c0)
        continue

    if _c1 in ("nan", "", "Cliente", "Filial"):
        continue
    if _c0.upper().startswith("FILIAL") and _c1 in ("nan", "", "Cliente", "Filial"):
        continue
    if not _vend_cli_atual:
        continue

    _venc     = _parse_data(_row[COL["vencimento"]])
    _pagto    = _parse_data(_row[COL["pagamento"]])
    _pendente = tratar_valor(_row[COL["pendente"]])
    _pago_val = tratar_valor(_row[COL["pago_total"]])

    if not _venc:
        continue

    # 🔥 IGNORA CONTA CAIXA 100
    if str(_row[COL["conta_caixa"]]).strip() == "Caixa Filial 100":
        continue

    _dias_venc = (_hoje - _venc).days

    if _dias_venc >= 60:
        _faixa = "grave"
    elif 30 <= _dias_venc < 60:
        _faixa = "alerta"
    elif 15 <= _dias_venc < 30:
        _faixa = "atencao"
    else:
        _faixa = None

    _fv, _ig = extrair_filial_nome(_vend_cli_atual)
    _nd_parse = limpar_nome_display(_vend_cli_atual)

    # =========================================
    # DEFINIÇÃO DE FILIAL
    # =========================================
    if _fv is not None:
        _fv_key   = _fv.strip().upper()
        _is_ativo = True
        _is_fdep  = False
    else:
        _fv_lookup = _filial_inativo_lookup.get(_nd_parse, "")

        # tenta extrair do nome
        if not _fv_lookup and "F" in _vend_cli_atual:
            _fv_lookup, _ = extrair_filial_nome(_vend_cli_atual)

        # tenta via cliente
        if not _fv_lookup:
            _fv_lookup = _filial_inativo_lookup.get(_c1[:30], "")

        if _fv_lookup == "FDEP" or _nd_parse in _nomes_fdep:
            _fv_key   = "FDEP"
            _is_ativo = False
            _is_fdep  = True
        elif _fv_lookup:
            _fv_key   = str(_fv_lookup).strip().upper()
            _is_ativo = False
            _is_fdep  = False
        else:
            _filial_txt = str(_row[COL["filial"]]).upper()
            m_fil = re.search(r"FILIAL\s*(\d+)", _filial_txt)

            if m_fil:
                _fv_key = f"F{int(m_fil.group(1))}"
            else:
                _fv_key = "OUTROS"

            _is_ativo = False
            _is_fdep  = False

    _fv_key = str(_fv_key).strip().upper()
    _key_vend = f"{_nd_parse}_{_fv_key}"

    # =========================================
    # RECEBIMENTOS
    # =========================================
    if _faixa and _pagto and _pago_val > 0 and _pagto >= _data_corte_parse:
        if _key_vend not in recebido_faixa:
            recebido_faixa[_key_vend] = {
                "grave": 0.0,
                "alerta": 0.0,
                "atencao": 0.0,
                "filial": _fv_key,
                "is_ativo": _is_ativo,
                "is_fdep": _is_fdep,
                "vendedor_nome": _nd_parse,
            }

        recebido_faixa[_key_vend][_faixa] += _pago_val

    # =========================================
    # CLIENTES A COBRAR
    # =========================================
    if _faixa and _pendente > 0:
        _titulo  = str(_row[COL["num_titulo"]]).strip()
        _parcela = str(_row[COL["num_parcela"]]).strip()

        _titulo_key = f"{_c1[:30]}_{_titulo}_{_parcela}"
        clientes_key_hoje.add(_titulo_key)

        _telefones = extrair_telefones(_contato)
        _mensagem = mensagem_cobranca_padrao(
            cliente=_c1[:50],
            vencimento=str(_row[COL["vencimento"]]).strip(),
            valor_pendente=_pendente,
        )

        clientes_cobrar.append({
            "vendedor": _vend_cli_atual,
            "filial": _fv_key,
            "is_ativo": _is_ativo,
            "is_fdep": _is_fdep,
            "cliente": _c1[:50],
            "contato": _contato,
            "telefones": _telefones,
            "avalista": _avalista[:50],
            "restricao": str(_row[COL["restricao_credito"]])[:40],
            "lancamento": str(_row[COL["num_lancamento"]]).strip(),
            "titulo": _titulo,
            "parcela": _parcela,
            "vencimento": str(_row[COL["vencimento"]]).strip(),
            "pagamento": str(_row[COL["pagamento"]]).strip(),
            "dias": _dias_venc,
            "pendente": _pendente,
            "pago": _pago_val,
            "faixa": _faixa,
            "titulo_key": _titulo_key,
            "cliente_key": normalizar_texto_match(_c1[:50]),
            "cobranca_key": "|".join([str(_c1[:50]).strip().upper(), str(_titulo).strip(), str(_parcela).strip(), str(_row[COL["vencimento"]]).strip()]),
            "mensagem_whatsapp": _mensagem,
        })


# Contagem
_ng = sum(1 for c in clientes_cobrar if c['faixa']=='grave')
_na = sum(1 for c in clientes_cobrar if c['faixa']=='alerta')
_nt = sum(1 for c in clientes_cobrar if c['faixa']=='atencao')
_nr = sum(1 for v in recebido_faixa.values() if v.get('is_ativo'))
print(f"👥 Clientes a cobrar: {len(clientes_cobrar)} ({_ng}g/{_na}a/{_nt}t)")
print(f"💰 Vendedores com recebimento no período: {_nr}")

print(f"\U0001f465 Clientes a cobrar: {len(clientes_cobrar)} títulos ({_ng} grave / {_na} alerta / {_nt} atenção)")

# =========================================
# COBRANÇA10 — helpers e config
# =========================================
import hashlib as _hashlib

COBRANCA10_LOGIN = "cobranca10"
COBRANCA10_NOME = "Cobrança10"
COBRANCA10_FILIAL = "FTER"
COBRANCA10_RATEIO_DEFAULT = 20.0

def _load_cobranca_global_rateio_pct():
    try:
        _cfgp = os.path.join(pasta, "cache_historico", "config_meta.json")
        if os.path.exists(_cfgp):
            with open(_cfgp, "r", encoding="utf-8") as _fcg:
                _rawcg = json.load(_fcg)
            _globalcg = _rawcg.get("global", _rawcg) if isinstance(_rawcg, dict) else {}
            _pctcg = float(_globalcg.get("cobranca_global_rateio_pct", COBRANCA10_RATEIO_DEFAULT) or COBRANCA10_RATEIO_DEFAULT)
            return max(0.0, min(100.0, _pctcg))
    except Exception:
        pass
    return COBRANCA10_RATEIO_DEFAULT

COBRANCA10_RATEIO = _load_cobranca_global_rateio_pct() / 100.0

def _load_crediaristas_config_do_meta():
    default = [
        {"login": "crediaristaf02_01", "nome": "Crediarista F2 01", "filial": "F2", "pct": 100},
        {"login": "crediaristaf03_01", "nome": "Crediarista F3 01", "filial": "F3", "pct": 100},
        {"login": "crediaristaf04_01", "nome": "Crediarista F4 01", "filial": "F4", "pct": 100},
        {"login": "crediaristaf05_01", "nome": "Crediarista F5 01", "filial": "F5", "pct": 100},
        {"login": "crediaristaf06_01", "nome": "Crediarista F6 01", "filial": "F6", "pct": 100},
        {"login": "crediaristaf08_01", "nome": "Crediarista F8 01", "filial": "F8", "pct": 100},
        {"login": "crediaristaf09_01", "nome": "Crediarista F9 01", "filial": "F9", "pct": 100},
    ]
    try:
        _cfgp = os.path.join(pasta, "cache_historico", "config_meta.json")
        if os.path.exists(_cfgp):
            with open(_cfgp, "r", encoding="utf-8") as _fc:
                _raw = json.load(_fc)
            _glob = _raw.get("global", _raw) if isinstance(_raw, dict) else {}
            rows = _glob.get("crediaristas_config") or []
            clean = []
            for r in rows:
                if not isinstance(r, dict): continue
                filial = str(r.get("filial") or "").upper().strip()
                if filial and not filial.startswith("F"): filial = "F" + filial.zfill(2).lstrip("0")
                if filial.startswith("F0"): filial = "F" + filial[2:]
                login = str(r.get("login") or "").lower().strip()
                nome = str(r.get("nome") or "").strip() or f"Crediarista {filial}"
                try: pct = float(str(r.get("pct", 100)).replace(",", "."))
                except Exception: pct = 100.0
                pct = max(0.0, min(100.0, pct))
                if filial and login and pct > 0: clean.append({"login": login, "nome": nome, "filial": filial, "pct": pct})
            if clean: return clean
    except Exception as e:
        print(f"⚠️ Não consegui carregar crediaristas_config: {e}")
    return default

CREDIARISTAS_CONFIG = _load_crediaristas_config_do_meta()
CREDIARISTAS_FILIAIS = {r["filial"]: r["login"] for r in CREDIARISTAS_CONFIG}
CREDIARISTAS_NOMES = {r["login"]: r.get("nome") or f"Crediarista {r['filial']}" for r in CREDIARISTAS_CONFIG}

def nome_crediarista_filial(filial, login=None):
    if login and str(login).lower() in CREDIARISTAS_NOMES:
        return CREDIARISTAS_NOMES[str(login).lower()]
    return f"CREDIARISTA{str(filial).upper()}"

def cobranca_row_key_py(r):
    return "|".join([
        str(r.get("cliente", r.get("nome", "")) or "").strip().upper(),
        str(r.get("titulo", "") or "").strip(),
        str(r.get("parcela", "") or "").strip(),
        str(r.get("vencimento", "") or "").strip(),
    ])

def cliente_grupo_key_py(r):
    """Chave de cliente para manter todos os títulos do mesmo cliente/CPF no mesmo destino operacional."""
    base = str(r.get("cpf") or r.get("documento") or r.get("cliente") or r.get("nome") or "").strip().upper()
    base = unicodedata.normalize("NFKD", base)
    base = "".join(c for c in base if not unicodedata.combining(c))
    base = re.sub(r"[^A-Z0-9]+", " ", base)
    base = re.sub(r"\s+", " ", base).strip()
    return base[:80] or cobranca_row_key_py(r)

# =========================================
# COBRANÇA10 — 20% global dos títulos de cobrança
# Cada título vai para 1 destino apenas, sem duplicidade.
# =========================================
_clientes_cobrar_sorted = sorted(clientes_cobrar, key=lambda x: (cobranca_row_key_py(x), -float(x.get("pendente", 0) or 0)))
_clientes_cobrar_unicos = {}
for _c_tmp in _clientes_cobrar_sorted:
    _k_tmp = cobranca_row_key_py(_c_tmp)
    if _k_tmp not in _clientes_cobrar_unicos:
        _clientes_cobrar_unicos[_k_tmp] = _c_tmp
_clientes_cobrar_base = list(_clientes_cobrar_unicos.values())
_clientes_cobrar_base.sort(key=lambda x: x["pendente"], reverse=True)
_qtd_terceiro = int(round(len(_clientes_cobrar_base) * COBRANCA10_RATEIO))
_clientes_cobrar_hash = sorted(_clientes_cobrar_base, key=lambda x: _hashlib.md5(cobranca_row_key_py(x).encode("utf-8")).hexdigest())
_clientes_terceiro_lista = _clientes_cobrar_hash[:_qtd_terceiro]
_clientes_terceiro_keys = {cobranca_row_key_py(x) for x in _clientes_terceiro_lista}
clientes_cobrar = [x for x in clientes_cobrar if cobranca_row_key_py(x) not in _clientes_terceiro_keys]
print(f"🤝 Cobrança10 separado com {len(_clientes_terceiro_lista)} títulos ({COBRANCA10_RATEIO*100:.0f}% do total único)")

# Ativos: agrupa por (vendedor, filial_vendedor) — pertence a uma filial pelo nome
ativos = df_ativos_raw.groupby(["vendedor","filial_vendedor","is_gerente"]).agg(
    pendente=("pendente","sum"), pago=("pago","sum")
).reset_index()
# Coluna "filial" nos ativos = filial_vendedor (necessária para o passo 2)
ativos["filial"] = ativos["filial_vendedor"]

# Garante uma linha "gerente/filial" sintética quando a filial tem vendedor ativo,
# mas não tem gerente ativo no relatório. Assim o rateio 60/40 continua valendo:
# 60% fica na filial/gerente e 40% vai para os vendedores.
_gerentes_sint = []
for _fil_syn in sorted(set(str(x) for x in ativos["filial_vendedor"].dropna().tolist())):
    _sub_syn = ativos[ativos["filial_vendedor"] == _fil_syn]
    if _sub_syn.empty:
        continue
    _tem_vend_syn = bool((_sub_syn["is_gerente"] == False).any())
    _tem_ger_syn = bool((_sub_syn["is_gerente"] == True).any())
    if _tem_vend_syn and not _tem_ger_syn:
        _gerentes_sint.append({
            "vendedor": f"GERENTE {_fil_syn}",
            "filial_vendedor": _fil_syn,
            "is_gerente": True,
            "pendente": 0.0,
            "pago": 0.0,
            "filial": _fil_syn,
            "_synthetic_manager": True,
        })

if _gerentes_sint:
    ativos = pd.concat([ativos, pd.DataFrame(_gerentes_sint)], ignore_index=True)
    print(f"🧩 Gerentes sintéticos criados para manter rateio 60/40: {', '.join(sorted(x['filial_vendedor'] for x in _gerentes_sint))}")

# Inativos: agrupa por (vendedor, filial_erp) — preserva a filial de cada crédito
inativos = df_inativos_raw.groupby(["vendedor","filial_erp"]).agg(
    pendente=("pendente","sum"), pago=("pago","sum")
).reset_index()
inativos.rename(columns={"filial_erp":"filial"}, inplace=True)
inativos["is_gerente"] = False

print(f"\n👥 Ativos (vendedor+gerente): {ativos['vendedor'].nunique()}")
print(f"👥 Inativos:                  {inativos['vendedor'].nunique()}")

# =========================================
# RATEIO: regra 60% gerente / 40% vendedores
# =========================================
PESO_GER  = 0.60
PESO_VEND = 0.40

ORDEM_FILIAIS = ["F1","F2","F3","F4","F5","F6","F8","F9"]

def ratear(ativos_df, filial, total_p, total_pg):
    """Distribui total_p/total_pg entre ativos da filial. Gerentes: 60%, vendedores: 40%."""
    mask = ativos_df["filial_vendedor"] == filial
    ger  = ativos_df[mask &  ativos_df["is_gerente"]]
    vend = ativos_df[mask & ~ativos_df["is_gerente"]]
    ng, nv = len(ger), len(vend)
    if ng == 0 and nv == 0:
        return ativos_df
    updates = {}
    if ng > 0 and nv > 0:
        for idx in ger.index:  updates[idx] = (total_p*PESO_GER/ng,   total_pg*PESO_GER/ng)
        for idx in vend.index: updates[idx] = (total_p*PESO_VEND/nv,  total_pg*PESO_VEND/nv)
    elif ng > 0:
        for idx in ger.index:  updates[idx] = (total_p/ng,  total_pg/ng)
    else:
        for idx in vend.index: updates[idx] = (total_p/nv,  total_pg/nv)
    for idx,(dp,dpg) in updates.items():
        ativos_df.loc[idx,"pendente"] += dp
        ativos_df.loc[idx,"pago"]     += dpg
    return ativos_df

# =========================================
# PASSO 1 — INATIVOS → ATIVOS DA MESMA FILIAL
# Filiais sem ativo ficam como bloco consolidado
# =========================================
filiais_sem_ativo = {}   # ex: {"F5": {"pendente":45169, "pago":77222}}

# Passo 1 trata apenas filiais normais — FDEP dos inativos é tratado no Passo 2
inat_normais = inativos[inativos["filial"] != "FDEP"].copy()
inat_fdep    = inativos[inativos["filial"] == "FDEP"].copy()

inat_agg = (
    inat_normais.groupby("filial", dropna=True)[["pendente","pago"]]
    .sum().reset_index()
)
for _, row_in in inat_agg.iterrows():
    filial   = row_in["filial"]
    total_p  = row_in["pendente"]
    total_pg = row_in["pago"]
    if total_p == 0 and total_pg == 0: continue

    tem_ativo = (ativos["filial_vendedor"] == filial).any()
    if tem_ativo:
        # V6.0: quando colaborador sai, a carteira PENDENTE pode ser redistribuída
        # para os ativos da filial, mas o RECEBIDO histórico do ex-colaborador não deve
        # inflar comissão/ranking de quem ficou no lugar. Mantém o pago como consolidado
        # da filial para fechar totais, sem jogar em vendedor/gerente ativo.
        ativos = ratear(ativos, filial, total_p, 0.0)
        filiais_sem_ativo.setdefault(f"{filial}_INATIVOS_HIST", {"pendente": 0.0, "pago": 0.0})
        filiais_sem_ativo[f"{filial}_INATIVOS_HIST"]["pago"] += float(total_pg or 0)
        print(f"  ↳ Inativos→{filial}: pendente rateado={total_p:,.2f}; pago histórico NÃO rateado={total_pg:,.2f}")
    else:
        filiais_sem_ativo[filial] = {"pendente": total_p, "pago": total_pg}
        print(f"  ↳ Inativos→{filial}: SEM ATIVO, consolidado pend={total_p:,.2f} pago={total_pg:,.2f}")

# =========================================
# PASSO 2 — FDEP (F90/F99) → FILIAIS NORMAIS
# Peso por filial = pendente+pago dos ativos daquela filial
# Dentro de cada filial aplica regra 60/40 gerente/vendedor
# Filiais sem ativo recebem a fatia proporcional no bloco consolidado
# =========================================
mask_dep  = ativos["filial"] == "FDEP"
df_dep    = ativos[mask_dep].copy()
df_normal = ativos[~mask_dep].copy()

# Soma também o FDEP vindo de inativos (vendedores inativos com créditos na F90/F99)
fdep_inat_p  = inat_fdep["pendente"].sum() if not inat_fdep.empty else 0.0
fdep_inat_pg = inat_fdep["pago"].sum()     if not inat_fdep.empty else 0.0

# pesos e tw expostos fora do bloco condicional para uso na serialização
pesos = {}
tw    = 0.0
fdep_total_rateado_p  = 0.0
fdep_total_rateado_pg = 0.0

if not df_dep.empty or (fdep_inat_p + fdep_inat_pg) > 0:
    total_dep_p  = df_dep["pendente"].sum() + fdep_inat_p
    # V6.0: pago de FDEP vindo de inativos fica histórico/consolidado; não entra no rateio dos ativos.
    total_dep_pg = df_dep["pago"].sum()
    if fdep_inat_pg:
        filiais_sem_ativo.setdefault("FDEP_INATIVOS_HIST", {"pendente": 0.0, "pago": 0.0})
        filiais_sem_ativo["FDEP_INATIVOS_HIST"]["pago"] += float(fdep_inat_pg or 0)
    fdep_total_rateado_p  = total_dep_p
    fdep_total_rateado_pg = total_dep_pg
    print(f"\n💰 FDEP a ratear: pend={total_dep_p:,.2f}  pago={total_dep_pg:,.2f} (ativos={df_dep['pendente'].sum():,.2f} + inativos pend={fdep_inat_p:,.2f}; pago inativo histórico={fdep_inat_pg:,.2f})")

    # Calcula peso de cada filial
    for f in ORDEM_FILIAIS:
        sub = df_normal[df_normal["filial_vendedor"] == f]
        if not sub.empty:
            pesos[f] = (sub["pendente"] + sub["pago"]).sum()
        elif f in filiais_sem_ativo:
            pesos[f] = filiais_sem_ativo[f]["pendente"] + filiais_sem_ativo[f]["pago"]

    tw = sum(pesos.values())
    if tw > 0:
        for filial, peso in pesos.items():
            prop   = peso / tw
            vp     = total_dep_p  * prop
            vpg    = total_dep_pg * prop
            if filial in filiais_sem_ativo:
                filiais_sem_ativo[filial]["pendente"] += vp
                filiais_sem_ativo[filial]["pago"]     += vpg
                print(f"  ↳ FDEP→{filial} (sem ativo): pend={vp:,.2f} pago={vpg:,.2f}")
            else:
                df_normal = ratear(df_normal, filial, vp, vpg)
                print(f"  ↳ FDEP→{filial}: pend={vp:,.2f} pago={vpg:,.2f}")
    print("✅ Rateio FDEP concluído")
else:
    print("ℹ️  Nenhum registro FDEP")

# ===== RESULTADO FINAL (ativos + blocos consolidados)
df_vend = df_normal.copy()
df_vend["nome_exibicao"] = df_vend["vendedor"].apply(limpar_nome_display)
df_vend["total"]         = df_vend["pendente"] + df_vend["pago"]

total_por_filial = df_vend.groupby("filial_vendedor")["pendente"].sum().rename("total_filial")
df_vend = df_vend.merge(total_por_filial, on="filial_vendedor", how="left")
df_vend["perc_filial"] = (df_vend["pendente"] / df_vend["total_filial"].replace(0,1) * 100).round(2)

df_vend["filial_ordem"] = pd.Categorical(df_vend["filial_vendedor"], categories=ORDEM_FILIAIS, ordered=True)
df_vend = df_vend.sort_values(
    ["filial_ordem","is_gerente","pendente"], ascending=[True,False,False]
).reset_index(drop=True)

# ===== VERIFICAÇÃO DE TOTAIS (deve bater com o arquivo bruto)
total_vend_p  = df_vend["pendente"].sum()
total_vend_pg = df_vend["pago"].sum()
total_cons_p  = sum(v["pendente"] for v in filiais_sem_ativo.values())
total_cons_pg = sum(v["pago"]     for v in filiais_sem_ativo.values())
total_final_p  = total_vend_p  + total_cons_p
total_final_pg = total_vend_pg + total_cons_pg

print(f"\n{'='*55}")
print(f"📊 TOTAIS POR FILIAL (após rateio):")
for f in ORDEM_FILIAIS:
    sub = df_vend[df_vend["filial_vendedor"] == f]
    if not sub.empty:
        tag = ""
    elif f in filiais_sem_ativo:
        sub_p  = filiais_sem_ativo[f]["pendente"]
        sub_pg = filiais_sem_ativo[f]["pago"]
        print(f"  {f} [sem ativo]: pend={sub_p:>12,.2f}  pago={sub_pg:>12,.2f}")
        continue
    else:
        continue
    print(f"  {f}: pend={sub['pendente'].sum():>12,.2f}  pago={sub['pago'].sum():>12,.2f}  ({len(sub)} vendedores)")

print(f"\n  TOTAL ativos   : pend={total_vend_p:>12,.2f}  pago={total_vend_pg:>12,.2f}")
print(f"  TOTAL consolid.: pend={total_cons_p:>12,.2f}  pago={total_cons_pg:>12,.2f}")
print(f"  TOTAL GERAL    : pend={total_final_p:>12,.2f}  pago={total_final_pg:>12,.2f}")
print(f"  BRUTO original : pend={total_bruto_p:>12,.2f}  pago={total_bruto_pg:>12,.2f}")
diff_p  = total_final_p  - total_bruto_p
diff_pg = total_final_pg - total_bruto_pg
print(f"  DIFERENÇA      : pend={diff_p:>+12,.2f}  pago={diff_pg:>+12,.2f}  {'✅ OK' if abs(diff_p)<0.02 and abs(diff_pg)<0.02 else '⚠️ VERIFICAR'}")
print(f"{'='*55}")


# =========================================
# V23 — USUÁRIOS COMERCIAIS AUTOMÁTICOS
# Gera usuários também a partir das metas de vendas/serviços.
# Assim, vendedor/gerente novo que já aparece no SGI em metas/vendas
# ganha login no dashboard mesmo antes de aparecer na carteira de cobrança.
# A cobrança/rateio continua usando a regra vigente do relatório de cobrança.
# =========================================
def _v23_norm_name_key(nome):
    s = limpar_nome_display(limpar_nome_erp(str(nome or ""))).strip()
    s = unicodedata.normalize("NFKD", s.upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _v23_filial_from_any_text(*vals):
    joined = " ".join(str(v or "") for v in vals)
    filial, is_ger = extrair_filial_nome(joined)
    if filial:
        return filial.upper(), bool(is_ger)
    m = re.search(r"FILIAL\s*0*(\d{1,3})", joined.upper())
    if m:
        n = int(m.group(1))
        return (f"F{n}" if n not in (90, 99) else "FDEP"), False
    m = re.search(r"\bF\s*0*(\d{1,2})\b", joined.upper())
    if m:
        return f"F{int(m.group(1))}", False
    return "", False

def _v23_nome_from_meta_row(row):
    vals = []
    for k, v in (row or {}).items():
        if str(k).startswith("_"):
            continue
        sv = str(v or "").strip()
        if not sv or sv.lower() in ("nan", "none", "total"):
            continue
        if re.search(r"^\s*(R\$)?\s*-?\d+([.,]\d+)?\s*%?\s*$", sv):
            continue
        vals.append(sv)

    for sv in vals:
        if re.search(r"\((GER\s*)?F\d+\)", sv, flags=re.I):
            return limpar_nome_display(limpar_nome_erp(sv)).strip(), sv

    for sv in vals:
        if "FILIAL" not in sv.upper() and len(sv) >= 3:
            return limpar_nome_display(limpar_nome_erp(sv)).strip(), sv

    return "", ""

def _v23_extrair_usuarios_das_metas_vendas(metas_info):
    out = {}
    dados = (metas_info or {}).get("dados", {}) if isinstance(metas_info, dict) else {}
    metas = dados.get("metas", {}) if isinstance(dados, dict) else {}

    for chave in ("venda_filial_vendedor_meta", "servico_filial_vendedor_ouro_fob", "venda_vendedor_subgrupo_20k"):
        obj = metas.get(chave, {}) if isinstance(metas, dict) else {}
        linhas = obj.get("linhas", []) if isinstance(obj, dict) else []
        if not isinstance(linhas, list):
            continue
        for row in linhas:
            if not isinstance(row, dict) or row.get("_is_total"):
                continue
            nome, raw_nome = _v23_nome_from_meta_row(row)
            if not nome or nome.upper() == "TOTAL":
                continue

            # V5.8: NÃO recria usuário comercial automático quando o nome do SGI
            # não tem tag operacional no próprio nome, ex.: (F6) ou (GERF6).
            # Isso evita colaborador desligado voltar para murais/rateios apenas
            # porque ainda existe linha antiga em metas/vendas.
            raw_nome_txt = str(raw_nome or "")
            if not re.search(r"\((?:GER\s*)?F\s*\d+\)", raw_nome_txt, flags=re.I):
                continue

            filial, is_ger = _v23_filial_from_any_text(raw_nome, *[row.get(k) for k in row.keys() if not str(k).startswith("_")])
            if not filial or filial == "FDEP":
                continue
            if re.search(r"\(GER\s*F\d+\)", str(raw_nome), flags=re.I) or re.search(r"\bGERENTE\b", str(raw_nome), flags=re.I):
                is_ger = True

            if not _colab_esta_liberado_py(nome, filial, is_ger, 'participa_cobrancas'):
                continue

            key = f"{_v23_norm_name_key(nome)}_{filial}_{1 if is_ger else 0}"
            out[key] = {"nome": nome, "filial": filial, "is_gerente": bool(is_ger), "origem": chave}
    return list(out.values())

_usuarios_vendas_auto = _v23_extrair_usuarios_das_metas_vendas(metas_vendas_info)
_existentes_v23 = set()
for _, _r_exist in df_vend.iterrows():
    _existentes_v23.add(f"{_v23_norm_name_key(_r_exist.get('nome_exibicao', _r_exist.get('vendedor','')))}_{str(_r_exist.get('filial_vendedor','')).upper()}_{1 if bool(_r_exist.get('is_gerente')) else 0}")

_novos_v23 = []
for _u_auto in _usuarios_vendas_auto:
    _key_auto = f"{_v23_norm_name_key(_u_auto['nome'])}_{_u_auto['filial']}_{1 if _u_auto.get('is_gerente') else 0}"
    if _key_auto in _existentes_v23:
        continue
    _novos_v23.append({
        "vendedor": f"{_u_auto['nome']} ({_u_auto['filial']})",
        "filial_vendedor": _u_auto["filial"],
        "is_gerente": bool(_u_auto.get("is_gerente")),
        "pendente": 0.0,
        "pago": 0.0,
        "filial": _u_auto["filial"],
        "_synthetic_sales_user": True,
        "nome_exibicao": _u_auto["nome"],
        "total": 0.0,
        "total_filial": 0.0,
        "perc_filial": 0.0,
        "filial_ordem": _u_auto["filial"],
    })

if _novos_v23:
    df_vend = pd.concat([df_vend, pd.DataFrame(_novos_v23)], ignore_index=True)
    try:
        df_vend["filial_ordem"] = pd.Categorical(df_vend["filial_vendedor"], categories=ORDEM_FILIAIS, ordered=True)
        df_vend = df_vend.sort_values(["filial_ordem","is_gerente","pendente"], ascending=[True,False,False]).reset_index(drop=True)
    except Exception:
        pass
    print("✅ V23_USUARIOS_VENDAS_AUTO: criados usuários vindos das metas de vendas/serviços:", ", ".join([f"{u['nome_exibicao']} ({u['filial_vendedor']})" for u in _novos_v23]))
else:
    print("ℹ️ V23_USUARIOS_VENDAS_AUTO: nenhum usuário novo vindo das metas de vendas/serviços.")


# ===== V27D: RATEIO DE NOVOS VENDEDORES ATIVOS VINDOS DAS METAS =====
# Quando um vendedor novo aparece nas metas de vendas/serviços, ele precisa entrar no bloco
# dos 40% de vendedores da filial. Mantém o total da filial e redistribui apenas entre
# vendedores não-gerentes da filial, sem mexer nos gerentes/senhas/bloqueios.
try:
    _filiais_com_novo_vendas = sorted(set(str(x.get("filial_vendedor", "")).upper() for x in _novos_v23 if not bool(x.get("is_gerente"))))
except Exception:
    _filiais_com_novo_vendas = []
if _filiais_com_novo_vendas:
    for _fil_v27d in _filiais_com_novo_vendas:
        _mask_vends = (df_vend["filial_vendedor"].astype(str).str.upper() == _fil_v27d) & (~df_vend["is_gerente"].astype(bool))
        _idx_vends = list(df_vend[_mask_vends].index)
        if len(_idx_vends) <= 1:
            continue
        _tot_p_vends = float(df_vend.loc[_idx_vends, "pendente"].sum())
        _tot_pg_vends = float(df_vend.loc[_idx_vends, "pago"].sum())
        if abs(_tot_p_vends) < 0.01 and abs(_tot_pg_vends) < 0.01:
            continue
        _share_p = round(_tot_p_vends / len(_idx_vends), 2)
        _share_pg = round(_tot_pg_vends / len(_idx_vends), 2)
        for _j, _idx in enumerate(_idx_vends):
            df_vend.loc[_idx, "pendente"] = _share_p
            df_vend.loc[_idx, "pago"] = _share_pg
        # ajuste de centavos no último vendedor para manter total exato
        _last = _idx_vends[-1]
        df_vend.loc[_last, "pendente"] += round(_tot_p_vends - float(df_vend.loc[_idx_vends, "pendente"].sum()), 2)
        df_vend.loc[_last, "pago"] += round(_tot_pg_vends - float(df_vend.loc[_idx_vends, "pago"].sum()), 2)
        print(f"✅ V27D rateio novos vendedores: {_fil_v27d} dividido entre {len(_idx_vends)} vendedores do bloco de 40%")
    df_vend["total"] = df_vend["pendente"].astype(float) + df_vend["pago"].astype(float)
    _total_por_filial_v27d = df_vend.groupby("filial_vendedor")["pendente"].sum().rename("total_filial")
    df_vend = df_vend.drop(columns=[c for c in ["total_filial"] if c in df_vend.columns], errors="ignore").merge(_total_por_filial_v27d, on="filial_vendedor", how="left")
    df_vend["perc_filial"] = (df_vend["pendente"] / df_vend["total_filial"].replace(0, 1) * 100).round(2)
    df_vend["filial_ordem"] = pd.Categorical(df_vend["filial_vendedor"], categories=ORDEM_FILIAIS, ordered=True)
    df_vend = df_vend.sort_values(["filial_ordem", "is_gerente", "pendente"], ascending=[True, False, False]).reset_index(drop=True)


# ===== V10.7: REATIVA RATEIO DE FILIAL QUE GANHOU USUÁRIO PELAS METAS =====
# Caso clássico: F5 não tinha colaborador ativo no relatório de cobrança, então a carteira
# ficou em filiais_sem_ativo. Depois a rotina V23 cria gerente/vendedor a partir das metas
# de vendas do SGI; se não transferir o bloco consolidado para esses usuários, a tela da
# filial fica com pendente/pago zerados, mas ainda aparece recebimento por faixa.
try:
    _filiais_drenadas_v107 = []
    for _fil_v107 in list(filiais_sem_ativo.keys()):
        if _fil_v107 not in ORDEM_FILIAIS:
            continue
        _bloco_v107 = filiais_sem_ativo.get(_fil_v107) or {}
        _pend_bloco_v107 = float(_bloco_v107.get('pendente', 0) or 0)
        _pago_bloco_v107 = float(_bloco_v107.get('pago', 0) or 0)
        if abs(_pend_bloco_v107) < 0.01 and abs(_pago_bloco_v107) < 0.01:
            continue
        _tem_ativos_v107 = (df_vend['filial_vendedor'].astype(str).str.upper() == _fil_v107).any()
        if not _tem_ativos_v107:
            continue
        df_vend = ratear(df_vend, _fil_v107, _pend_bloco_v107, _pago_bloco_v107)
        # Remove o consolidado da filial para ela não ser considerada "sem ativo" depois.
        del filiais_sem_ativo[_fil_v107]
        _filiais_drenadas_v107.append(f"{_fil_v107}: pend={_pend_bloco_v107:,.2f} pago={_pago_bloco_v107:,.2f}")
    if _filiais_drenadas_v107:
        df_vend['total'] = df_vend['pendente'].astype(float) + df_vend['pago'].astype(float)
        _total_por_filial_v107 = df_vend.groupby('filial_vendedor')['pendente'].sum().rename('total_filial')
        df_vend = df_vend.drop(columns=[c for c in ['total_filial'] if c in df_vend.columns], errors='ignore').merge(_total_por_filial_v107, on='filial_vendedor', how='left')
        df_vend['perc_filial'] = (df_vend['pendente'] / df_vend['total_filial'].replace(0, 1) * 100).round(2)
        df_vend['filial_ordem'] = pd.Categorical(df_vend['filial_vendedor'], categories=ORDEM_FILIAIS, ordered=True)
        df_vend = df_vend.sort_values(['filial_ordem', 'is_gerente', 'pendente'], ascending=[True, False, False]).reset_index(drop=True)
        print('✅ V10.7 rateio consolidado drenado para usuários ativos/metas: ' + '; '.join(_filiais_drenadas_v107))
except Exception as _e_v107_dreno:
    print(f'⚠️ V10.7 falhou ao drenar filial sem ativo para usuários de metas: {_e_v107_dreno}')

# ===== SALVAR CSV
csv_path = os.path.join(pasta, "dashboard_vendedores.csv")
df_vend.to_csv(csv_path, index=False)
print(f"\n💾 CSV salvo: {csv_path}")

# =========================================
# 🔐 CREDENCIAIS
# =========================================
SENHA_MASTER = "mdladm01"
LOGIN_MASTER = "master"
SENHA_DIRETOR = "mdldir01"
LOGIN_DIRETOR = "diretorcomercial"
LOGIN_PAINEL = "painel"
SENHA_PAINEL = "painelmdl10"
EMAIL_RECUPERACAO = "sac@moveisdolar.com.br"
cred_path    = os.path.join(pasta, "credenciais_vendedores.txt")
cred_state_path = os.path.join(pasta, "credenciais_dashboard.json")

def _load_json_safe(path, default):
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, type(default)) else default
    except Exception:
        pass
    return default

# Tenta baixar o estado remoto das credenciais para manter as senhas alteradas no dashboard
try:
    import ftplib
    from io import BytesIO
    _ftp = ftplib.FTP()
    _ftp.connect("moveisdolar.com.br", 21, timeout=20)
    _ftp.login("moveisdolar3", "Deg27ll02mdl2301#")
    _ftp.encoding = "utf-8"
    _ftp.cwd("/public_html/colaborador")
    with open(cred_state_path, "wb") as _fcred:
        _ftp.retrbinary("RETR credenciais_dashboard.json", _fcred.write)
    _ftp.quit()
    print("🔄 Credenciais remotas sincronizadas do FTP.")
except Exception:
    pass

# Carrega senhas existentes do TXT antigo para não trocar entre execuções
creds_salvas = {}
if os.path.exists(cred_path):
    with open(cred_path, "r", encoding="utf-8") as f:
        for linha in f:
            linha = linha.strip()
            if "|" in linha and not linha.startswith(("=","-")):
                p = [x.strip() for x in linha.split("|")]
                if len(p) >= 4 and p[0] and p[2]:
                    creds_salvas[f"{p[0]}_{p[3]}"] = p[2]

cred_state = _load_json_safe(cred_state_path, {"users": {}, "director": {}, "password_reset_requests": []})

def _norm_login_lookup_py(txt):
    s = unicodedata.normalize("NFKD", str(txt or "").strip().upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _buscar_login_existente_por_nome_filial(nome, filial, is_gerente=None):
    """Mantém login customizado alterado pelo Master entre uma execução e outra."""
    try:
        alvo_nome = _norm_login_lookup_py(limpar_nome_display(nome))
        alvo_filial = str(filial or "").strip().upper()
        for _login_ant, _u_ant in (cred_state.get("users", {}) or {}).items():
            if not isinstance(_u_ant, dict):
                continue
            if str(_u_ant.get("filial") or "").strip().upper() != alvo_filial:
                continue
            if _norm_login_lookup_py(_u_ant.get("nome") or "") != alvo_nome:
                continue
            if is_gerente is not None and bool(_u_ant.get("is_gerente")) != bool(is_gerente):
                # não mistura gerente com vendedor de mesmo nome/filial
                continue
            return str(_u_ant.get("login") or _login_ant).strip().lower() or str(_login_ant).strip().lower()
    except Exception:
        pass
    return ""

credenciais = {}
linhas_txt  = []
auth_users  = {}

for _, row in df_vend.iterrows():
    nome_exib = str(row["nome_exibicao"]).strip()
    filial    = str(row["filial_vendedor"]).strip()
    is_ger    = bool(row["is_gerente"])
    login     = normalizar_login(nome_exib)
    chave     = f"{login}_{filial}"
    senha_inicial = creds_salvas.get(chave) or (login.upper() + str(random.randint(100,999)))
    login_fin = login
    if login in credenciais and credenciais[login]["filial"] != filial:
        login_fin = f"{login}_{filial.lower()}"

    _login_custom = _buscar_login_existente_por_nome_filial(nome_exib, filial, is_gerente=is_ger)
    if _login_custom:
        login_fin = _login_custom

    estado_ant = cred_state.get("users", {}).get(login_fin, {})
    senha_atual = estado_ant.get("password") or senha_inicial
    precisa_trocar = bool(estado_ant.get("must_change_password", senha_atual == senha_inicial))

    credenciais[login_fin] = {
        "senha":      senha_atual,
        "senha_inicial": senha_inicial,
        "nome":       nome_exib,
        "filial":     filial,
        "is_gerente": is_ger,
        "pendente":   round(float(row["pendente"]),   2),
        "pago":       round(float(row["pago"]),       2),
        "total":      round(float(row["total"]),      2),
        "perc_filial":round(float(row["perc_filial"]),2),
    }
    auth_users[login_fin] = {
        "login": login_fin,
        "password": senha_atual,
        "initial_password": senha_inicial,
        "must_change_password": precisa_trocar,
        "nome": nome_exib,
        "filial": filial,
        "is_gerente": is_ger,
        "email_recuperacao": EMAIL_RECUPERACAO,
    }
    tipo = "GERENTE" if is_ger else "VENDEDOR"
    linhas_txt.append(f"{login_fin} | {nome_exib} | {senha_atual} | {filial} | {tipo}")

# Usuário especial de cobrança terceirizada
_chave_cob10 = COBRANCA10_LOGIN
_senha_cob10_inicial = creds_salvas.get(f"{COBRANCA10_LOGIN}_{COBRANCA10_FILIAL}") or ("COB10" + str(random.randint(100,999)))
_estado_cob10 = cred_state.get("users", {}).get(_chave_cob10, {})
_senha_cob10 = _estado_cob10.get("password") or _senha_cob10_inicial
_precisa_cob10 = bool(_estado_cob10.get("must_change_password", _senha_cob10 == _senha_cob10_inicial))
_terceiro_pendente = sum(float(x.get("pendente", 0) or 0) for x in _clientes_terceiro_lista)
credenciais[_chave_cob10] = {
    "senha": _senha_cob10,
    "senha_inicial": _senha_cob10_inicial,
    "nome": COBRANCA10_NOME,
    "filial": COBRANCA10_FILIAL,
    "is_gerente": False,
    "is_terceiro": True,
    "pendente": round(_terceiro_pendente, 2),
    "pago": 0.0,
    "total": round(_terceiro_pendente, 2),
    "perc_filial": 100.0,
    "only_cobranca": True,
}
auth_users[_chave_cob10] = {
    "login": _chave_cob10,
    "password": _senha_cob10,
    "initial_password": _senha_cob10_inicial,
    "must_change_password": _precisa_cob10,
    "nome": COBRANCA10_NOME,
    "filial": COBRANCA10_FILIAL,
    "is_gerente": False,
    "is_terceiro": True,
    "only_cobranca": True,
    "email_recuperacao": EMAIL_RECUPERACAO,
}
linhas_txt.append(f"{_chave_cob10} | {COBRANCA10_NOME} | {_senha_cob10} | {COBRANCA10_FILIAL} | COBRANÇA TERCEIRO")


# Inicialização antecipada para evitar NameError antes da montagem final
clientes_crediarista_js = {}
recebimentos_crediarista_js = {}

# Usuários CREDIARISTAS por filial
for _fil_cred, _login_cred in CREDIARISTAS_FILIAIS.items():
    _nome_cred = nome_crediarista_filial(_fil_cred, _login_cred)
    _senha_ini = creds_salvas.get(f"{_login_cred}_{_fil_cred}") or creds_salvas.get(_login_cred) or (_login_cred.upper() + str(random.randint(100,999)))
    _estado_cred = cred_state.get("users", {}).get(_login_cred, {})
    _senha_cred = _estado_cred.get("password") or _senha_ini
    _troca_cred = bool(_estado_cred.get("must_change_password", _senha_cred == _senha_ini))
    _pend_cred = sum(float(x.get("pendente", 0) or 0) for _fx in ['grave','alerta','atencao'] for x in clientes_crediarista_js.get(_login_cred, {}).get(_fx, []))
    _pago_cred = sum(float(x.get("pago", 0) or 0) for _fx in ['grave','alerta','atencao'] for x in recebimentos_crediarista_js.get(_login_cred, {}).get(_fx, []))
    credenciais[_login_cred] = {
        "senha": _senha_cred,
        "senha_inicial": _senha_ini,
        "nome": _nome_cred,
        "filial": _fil_cred,
        "is_gerente": False,
        "is_crediarista": True,
        "only_cobranca": True,
        "pendente": round(_pend_cred, 2),
        "pago": round(_pago_cred, 2),
        "total": round(_pend_cred + _pago_cred, 2),
        "perc_filial": 100.0,
    }
    auth_users[_login_cred] = {
        "login": _login_cred,
        "password": _senha_cred,
        "initial_password": _senha_ini,
        "must_change_password": _troca_cred,
        "nome": _nome_cred,
        "filial": _fil_cred,
        "is_gerente": False,
        "is_crediarista": True,
        "only_cobranca": True,
        "email_recuperacao": EMAIL_RECUPERACAO,
    }
    linhas_txt.append(f"{_login_cred} | {_nome_cred} | {_senha_cred} | {_fil_cred} | CREDIARISTA")


# Usuário somente visualização da tela inicial
_estado_painel = cred_state.get("users", {}).get(LOGIN_PAINEL, {})
_senha_painel = _estado_painel.get("password") or SENHA_PAINEL
_troca_painel = bool(_estado_painel.get("must_change_password", False))
credenciais[LOGIN_PAINEL] = {
    "senha": _senha_painel,
    "senha_inicial": SENHA_PAINEL,
    "nome": "Painel",
    "filial": "",
    "is_gerente": False,
    "is_viewer": True,
    "pendente": 0.0,
    "pago": 0.0,
    "total": 0.0,
    "perc_filial": 0.0,
}
auth_users[LOGIN_PAINEL] = {
    "login": LOGIN_PAINEL,
    "password": _senha_painel,
    "initial_password": SENHA_PAINEL,
    "must_change_password": _troca_painel,
    "nome": "Painel",
    "filial": "",
    "is_gerente": False,
    "is_viewer": True,
    "email_recuperacao": EMAIL_RECUPERACAO,
}
linhas_txt.append(f"{LOGIN_PAINEL} | Painel | {_senha_painel} | --- | VISUALIZAÇÃO")

# V5.6: preserva e publica a camada administrativa de status/participação.
_status_map_final = {}
_status_antigo = cred_state.get("colaborador_status", {}) if isinstance(cred_state.get("colaborador_status", {}), dict) else {}
for _login_st, _u_st in list(auth_users.items()):
    if not isinstance(_u_st, dict):
        continue
    _nome_st = _u_st.get("nome") or _login_st
    _fil_st = _u_st.get("filial") or ""
    _is_ger_st = bool(_u_st.get("is_gerente"))
    _key_st = _colab_status_key_py(_nome_st, _fil_st, _is_ger_st)
    _prev_st = dict(_status_antigo.get(_key_st) or {})
    _base_st = _colab_default_status_py(_login_st, _nome_st, _fil_st, _is_ger_st)
    _base_st.update(_prev_st)
    _base_st["login"] = _login_st
    _base_st["nome"] = _nome_st
    _base_st["filial"] = str(_fil_st or '').strip().upper()
    if _u_st.get("is_terceiro"):
        _base_st["tipo"] = "Cobrança terceiro"
    elif _u_st.get("is_crediarista"):
        _base_st["tipo"] = "Crediarista"
    elif _u_st.get("is_viewer"):
        _base_st["tipo"] = "Painel"
    else:
        _base_st["tipo"] = "Gerente" if _is_ger_st else "Vendedor"
    _base_st["status"] = "inativo" if str(_base_st.get("status") or "ativo").lower().strip() in ("inativo","desligado","bloqueado","0","false") else "ativo"
    for _flag_st in ("participa_cobrancas","participa_sem_movimento","participa_aniversariantes","participa_murais"):
        _base_st[_flag_st] = bool(_base_st.get(_flag_st, True))
        _u_st[_flag_st] = _base_st[_flag_st]
        if _login_st in credenciais:
            credenciais[_login_st][_flag_st] = _base_st[_flag_st]
    _u_st["status_operacional"] = _base_st["status"]
    _u_st["access_disabled"] = (_base_st["status"] != "ativo")
    _u_st["data_entrada"] = _base_st.get("data_entrada", "")
    _u_st["data_saida"] = _base_st.get("data_saida", "")
    _u_st["substituto"] = _base_st.get("substituto", "")
    _u_st["obs"] = _base_st.get("obs", "")
    if _login_st in credenciais:
        credenciais[_login_st]["status_operacional"] = _base_st["status"]
        credenciais[_login_st]["access_disabled"] = (_base_st["status"] != "ativo")
        credenciais[_login_st]["data_saida"] = _base_st.get("data_saida", "")
        credenciais[_login_st]["substituto"] = _base_st.get("substituto", "")
    _status_map_final[_key_st] = _base_st

# Mantém registros antigos inativos mesmo que não tenham vindo no SGI nesta execução, para histórico/configuração.
for _key_old_st, _old_st in _status_antigo.items():
    if _key_old_st not in _status_map_final and isinstance(_old_st, dict) and str(_old_st.get("status") or "").lower() == "inativo":
        _status_map_final[_key_old_st] = _old_st

cred_state["users"] = auth_users
cred_state["colaborador_status"] = _status_map_final
cred_state["director"] = {
    "login": LOGIN_DIRETOR,
    "password": cred_state.get("director", {}).get("password") or SENHA_DIRETOR,
    "initial_password": SENHA_DIRETOR,
    "must_change_password": bool(cred_state.get("director", {}).get("must_change_password", True)),
    "nome": "Diretor Comercial",
    "email_recuperacao": EMAIL_RECUPERACAO,
}
cred_state.setdefault("password_reset_requests", [])
# V23: bloqueio geral de acesso controlado pelo Master.
# Mantém o estado remoto entre execuções; Master/Diretor continuam podendo entrar para desbloquear.
cred_state["access_blocked"] = bool(cred_state.get("access_blocked", False))
cred_state.setdefault("access_blocked_reason", "Sistema em atualização. Aguarde liberação pelo Master.")
cred_state.setdefault("access_blocked_at", "")
cred_state.setdefault("access_unblocked_at", "")

with open(cred_state_path, "w", encoding="utf-8") as f:
    json.dump(cred_state, f, ensure_ascii=False, indent=2)

with open(cred_path, "w", encoding="utf-8") as f:
    f.write("=" * 70 + "\n")
    f.write("   CREDENCIAIS DASHBOARD – MÓVEIS DOLAR\n")
    f.write("=" * 70 + "\n\n")
    f.write(f"  MASTER             login: {LOGIN_MASTER}   senha: {SENHA_MASTER}\n")
    f.write(f"  DIRETOR COMERCIAL  login: {LOGIN_DIRETOR}   senha atual: {cred_state['director']['password']}\n")
    f.write(f"  PAINEL            login: {LOGIN_PAINEL}   senha atual: {_senha_painel}\n\n")
    f.write("-" * 70 + "\n")
    f.write(f"  {'LOGIN':<20} {'NOME':<35} {'SENHA':<14} {'FILIAL':<6} TIPO\n")
    f.write("-" * 70 + "\n")
    for linha in sorted(linhas_txt):
        p = [x.strip() for x in linha.split("|")]
        if len(p) >= 5:
            f.write(f"  {p[0]:<20} {p[1]:<35} {p[2]:<14} {p[3]:<6} {p[4]}\n")
    f.write("\n" + "=" * 70 + "\n")

print(f"🔐 Credenciais: {cred_path} ({len(credenciais)} vendedores)")
js_auth_state = json.dumps(cred_state, ensure_ascii=False)


# =========================================
# 💾 MÓDULO DE CACHE / HISTÓRICO / METAS (v2)
# Nova lógica: 3 gráficos por faixa + 1 geral
# Grave=20% do pendente grave, Alerta=15%, Atenção=10%
# Geral: ponderado 60/30/10 dos 3 gráficos
# =========================================
import json
from datetime import datetime, timedelta, date
from pathlib import Path

CACHE_DIR  = os.path.join(pasta, "cache_historico")
Path(CACHE_DIR).mkdir(exist_ok=True)

hoje_str = now_brasilia().strftime("%Y-%m-%d")
mes_str  = now_brasilia().strftime("%Y-%m")

# Carrega configurações de meta (definidas pelo master no dashboard)
# Estrutura: config_meta.json tem "global" (padrão) e "individual" (por vendedor/filial)

_config_meta_path = os.path.join(CACHE_DIR, "config_meta.json")
_config_meta_existed_before = os.path.exists(_config_meta_path)
_config_meta_loaded_from_remote = False
REMOTE_PUBLIC_BASE = "https://moveisdolar.com.br/colaborador"
REMOTE_CONFIG_URL = REMOTE_PUBLIC_BASE + "/config_meta.json"
_hist_dash_path = os.path.join(CACHE_DIR, 'historico_dashboard.json')
_fechamento_path = os.path.join(CACHE_DIR, 'fechamentos_mensais.json')
REMOTE_HIST_URL = REMOTE_PUBLIC_BASE + '/historico_dashboard.json'
REMOTE_FECHAMENTO_URL = REMOTE_PUBLIC_BASE + '/fechamentos_mensais.json'
try:
    with urllib.request.urlopen(REMOTE_HIST_URL, timeout=10) as _resp_hist:
        _remote_hist = _resp_hist.read().decode('utf-8', errors='ignore').strip()
    if _remote_hist and _remote_hist.startswith('{'):
        with open(_hist_dash_path, 'w', encoding='utf-8') as _f_hist_local:
            _f_hist_local.write(_remote_hist)
        print('🌐 Histórico dashboard sincronizado do servidor')
except Exception:
    pass
try:
    with urllib.request.urlopen(REMOTE_FECHAMENTO_URL, timeout=10) as _resp_fech:
        _remote_fech = _resp_fech.read().decode('utf-8', errors='ignore').strip()
    if _remote_fech and _remote_fech.startswith('{'):
        with open(_fechamento_path, 'w', encoding='utf-8') as _f_fech_local:
            _f_fech_local.write(_remote_fech)
        print('🌐 Fechamentos mensais sincronizados do servidor')
except Exception:
    pass
# V10.5: config_meta.json do FTP é fonte de verdade; aceita HTTPS com SSL quebrado e fallback FTP.
# Aceita tanto o JSON cru {global,individual} quanto resposta da API {ok:true,data:{...}}.
def _normalizar_config_meta_payload(_data):
    try:
        if not isinstance(_data, dict):
            return None

        # API PHP pode devolver {ok:true,data:{global:{...},individual:{...}}}
        if isinstance(_data.get("data"), dict):
            _data = _data.get("data")

        if isinstance(_data.get("global"), dict) or isinstance(_data.get("individual"), dict):
            return {
                "global": _data.get("global") if isinstance(_data.get("global"), dict) else {},
                "individual": _data.get("individual") if isinstance(_data.get("individual"), dict) else {},
            }

        # Formato antigo: o próprio objeto era o global.
        _chaves_global = {
            "grave_pct", "alerta_pct", "atencao_pct", "peso_grave", "peso_alerta", "peso_atencao",
            "vendas_min_pct", "servicos_min_pct", "gerente_vendas_min_pct", "gerente_servicos_min_pct",
            "vendedor_policy", "gerente_policy", "camp_meta_diaria_vend", "camp_meta_diaria_ger",
            "camp_dindin_vend", "camp_dindin_ger", "camp_admin", "comissao_pagamento_texto",
        }
        if any(k in _data for k in _chaves_global):
            return {"global": _data, "individual": {}}
    except Exception:
        pass
    return None


def _config_meta_tem_conteudo(_payload):
    try:
        if not isinstance(_payload, dict):
            return False
        _glob = _payload.get("global") if isinstance(_payload.get("global"), dict) else {}
        _ind = _payload.get("individual") if isinstance(_payload.get("individual"), dict) else {}
        _chaves_criticas = (
            "comissao_pagamento_texto", "gerente_vendas_min_pct", "gerente_servicos_min_pct",
            "camp_meta_diaria_vend", "camp_meta_diaria_ger", "vendedor_policy", "gerente_policy",
            "crediaristas_config",
        )
        return bool(_ind) or any(k in _glob for k in _chaves_criticas)
    except Exception:
        return False


def _ler_config_meta_arquivo(_path):
    try:
        if not _path or not os.path.exists(_path):
            return None
        with open(_path, "r", encoding="utf-8") as _f_cfg_read:
            _raw = json.load(_f_cfg_read)
        _payload = _normalizar_config_meta_payload(_raw)
        if _config_meta_tem_conteudo(_payload):
            return _payload
    except Exception as _e_cfg_file:
        print(f"⚠️ Config meta local inválida em {_path}: {_e_cfg_file}")
    return None


def _baixar_config_meta_remota():
    _urls = [
        REMOTE_CONFIG_URL,
        REMOTE_PUBLIC_BASE + "/config_meta_api.php",
    ]

    # V10.5: alguns ambientes Railway/Locaweb falham na cadeia SSL do HTTPS.
    # Para NÃO cair em default, tenta HTTPS com contexto normal, depois sem verificação,
    # e por fim baixa o JSON diretamente via FTP.
    _contexts = [None]
    try:
        _contexts.append(ssl._create_unverified_context())
    except Exception:
        pass

    for _url in _urls:
        for _ctx in _contexts:
            try:
                _sep = "&" if "?" in _url else "?"
                _req_url = _url + _sep + "_=" + str(int(time.time()))
                if _ctx is None:
                    _resp_obj = urllib.request.urlopen(_req_url, timeout=15)
                else:
                    _resp_obj = urllib.request.urlopen(_req_url, timeout=15, context=_ctx)
                with _resp_obj as _resp_cfg:
                    _remote_cfg = _resp_cfg.read().decode("utf-8", errors="ignore").strip()
                if not _remote_cfg or not _remote_cfg.startswith("{"):
                    continue
                _payload = _normalizar_config_meta_payload(json.loads(_remote_cfg))
                if not _config_meta_tem_conteudo(_payload):
                    continue
                with open(_config_meta_path, "w", encoding="utf-8") as _f_cfg_local:
                    json.dump(_payload, _f_cfg_local, ensure_ascii=False, indent=2)
                _modo_ssl = "SSL normal" if _ctx is None else "SSL sem verificação"
                print(f"🌐 Config meta sincronizada do servidor ({_modo_ssl}): {_url}")
                return _payload
            except Exception as _e_cfg_remote:
                _modo_ssl = "SSL normal" if _ctx is None else "SSL sem verificação"
                print(f"⚠️ Não consegui baixar config meta de {_url} ({_modo_ssl}): {_e_cfg_remote}")

    # Fallback FTP: este é o mais fiel ao arquivo que fica em /public_html/colaborador.
    try:
        import ftplib
        from io import BytesIO
        _bio = BytesIO()
        _ftp_cfg = ftplib.FTP()
        _ftp_cfg.connect("moveisdolar.com.br", 21, timeout=20)
        _ftp_cfg.login("moveisdolar3", "Deg27ll02mdl2301#")
        _ftp_cfg.encoding = "utf-8"
        _ftp_cfg.cwd("/public_html/colaborador")
        _ftp_cfg.retrbinary("RETR config_meta.json", _bio.write)
        _ftp_cfg.quit()
        _raw_ftp = _bio.getvalue().decode("utf-8", errors="ignore").strip()
        _payload = _normalizar_config_meta_payload(json.loads(_raw_ftp))
        if _config_meta_tem_conteudo(_payload):
            with open(_config_meta_path, "w", encoding="utf-8") as _f_cfg_local:
                json.dump(_payload, _f_cfg_local, ensure_ascii=False, indent=2)
            print("🌐 Config meta sincronizada diretamente do FTP: config_meta.json")
            return _payload
    except Exception as _e_cfg_ftp:
        print(f"⚠️ Não consegui baixar config_meta.json via FTP: {_e_cfg_ftp}")

    return None

_config_meta_payload_remoto = _baixar_config_meta_remota()
if _config_meta_payload_remoto:
    _config_meta_loaded_from_remote = True
_config_meta_default_global = {
    "grave_pct":   20.0,
    "alerta_pct":  15.0,
    "atencao_pct": 10.0,
    "peso_grave":  60.0,
    "peso_alerta": 30.0,
    "peso_atencao":10.0,
    "vendas_min_pct": 80.0,
    "servicos_min_pct": 80.0,
    "gerente_vendas_min_pct": 80.0,
    "gerente_servicos_min_pct": 80.0,
    "vendedor_rentab_min_mercantil_pct": 80.0,
    "gerente_rentab_min_mercantil_pct": 80.0,
    "vendedor_policy": [],
    "gerente_policy": [],
    "vendedor_policy_headers": [],
    "gerente_policy_headers": [],
    "camp_cobranca_terceiro": [
        {"faixa": "atencao", "pct": "1.00"},
        {"faixa": "alerta", "pct": "2.00"},
        {"faixa": "grave", "pct": "3.00"}
    ],
    "camp_cob_crediarista": [
        {"faixa": "atencao", "pct": "1.00"},
        {"faixa": "alerta", "pct": "2.00"},
        {"faixa": "grave", "pct": "3.00"}
    ],
    "cob_cred_rateio_filial_pct": 50.0,
    "cob_cred_rateio_cred_pct": 50.0,
    "cobranca_global_rateio_pct": 20.0,
    "comissao_pagamento_texto": "A comissão reinicia a cada mês e o pagamento é previsto para o dia 25 do mês seguinte.",
    "telegram_contacts": [],
    "aniversario_msg_template": "Olá, {primeiro_nome}! Feliz aniversário! 🎂🎉\n\nAqui é da Lojas MDL - Móveis do Lar. Desejamos muita saúde, paz e felicidades neste dia especial. 😍😍\n\nPreparamos condições especiais para você comemorar com a gente.\n🕺🎉🤩",
    "reativacao_msg_template": "Olá, {primeiro_nome}! Tudo bem? 😊\n\nAqui é da Lojas MDL - Móveis do Lar. Estamos com saudades de você! Faz um tempinho que você não aparece na loja.  🥹\n\nVenha conhecer nossas novidades e aproveitar condições especiais que preparamos para nossos clientes. 👈👈😍😍",
    "crediaristas_config": [
        {"login": "crediaristaf02_01", "nome": "Crediarista F2 01", "filial": "F2", "pct": 100},
        {"login": "crediaristaf03_01", "nome": "Crediarista F3 01", "filial": "F3", "pct": 100},
        {"login": "crediaristaf04_01", "nome": "Crediarista F4 01", "filial": "F4", "pct": 100},
        {"login": "crediaristaf05_01", "nome": "Crediarista F5 01", "filial": "F5", "pct": 100},
        {"login": "crediaristaf06_01", "nome": "Crediarista F6 01", "filial": "F6", "pct": 100},
        {"login": "crediaristaf08_01", "nome": "Crediarista F8 01", "filial": "F8", "pct": 100},
        {"login": "crediaristaf09_01", "nome": "Crediarista F9 01", "filial": "F9", "pct": 100}
    ],
}
_config_meta_payload = (
    _config_meta_payload_remoto
    or _ler_config_meta_arquivo(_config_meta_path)
    or _ler_config_meta_arquivo(os.path.join(pasta, "config_meta.json"))
)

if _config_meta_payload:
    _cfg_global = _config_meta_payload.get("global") if isinstance(_config_meta_payload.get("global"), dict) else {}
    _cfg_individual = _config_meta_payload.get("individual") if isinstance(_config_meta_payload.get("individual"), dict) else {}
    CONFIG_META = {**_config_meta_default_global, **_cfg_global}
    CONFIG_META_IND = _cfg_individual
    try:
        # Mantém um cache normalizado para as próximas rotinas e para o monitor Telegram.
        with open(_config_meta_path, "w", encoding="utf-8") as _f:
            json.dump({"global": CONFIG_META, "individual": CONFIG_META_IND}, _f, ensure_ascii=False, indent=2)
    except Exception as _e_cfg_cache:
        print(f"⚠️ Não consegui salvar cache local do config_meta.json: {_e_cfg_cache}")
else:
    CONFIG_META = _config_meta_default_global.copy()
    CONFIG_META_IND = {}
    with open(_config_meta_path, "w", encoding="utf-8") as _f:
        json.dump({"global": CONFIG_META, "individual": {}}, _f, ensure_ascii=False, indent=2)
    print("⚠️ Config meta não encontrada no FTP/local; usando defaults de segurança V10.5.")

def get_config_meta(key):
    """Retorna config de meta para um vendedor/filial, com fallback para global."""
    aliases = [str(key or ''), str(key or '').upper()]
    if isinstance(key, str) and '::' in key:
        _kind, _pure = key.split('::', 1)
        aliases += [_pure, _pure.upper(), ('vend:' if _kind == 'VEND' else 'fil:') + _pure]
    for ak in aliases:
        if ak in CONFIG_META_IND:
            return {**CONFIG_META, **CONFIG_META_IND[ak]}
    return CONFIG_META

print(f"⚙️  Config meta global: Grave={CONFIG_META['grave_pct']}% Alerta={CONFIG_META['alerta_pct']}% Atenção={CONFIG_META['atencao_pct']}% | Pesos: {CONFIG_META['peso_grave']}/{CONFIG_META['peso_alerta']}/{CONFIG_META['peso_atencao']} | Gerente mínimo={CONFIG_META.get('gerente_vendas_min_pct')}%/{CONFIG_META.get('gerente_servicos_min_pct')}% | Comissão texto={CONFIG_META.get('comissao_pagamento_texto')}")
print(f"⚙️  Configs individuais: {len(CONFIG_META_IND)} sobreposições")

# Histórico mensal de comissão de cobrança (pagamento configurável no config_meta.json)
COMISSAO_HIST_PATH = os.path.join(pasta, "historico_comissao_cobranca.json")

def _row_key_cob_json(r):
    return "|".join([
        str(r.get("cliente", r.get("nome", "")) or "").strip().upper(),
        str(r.get("titulo", "") or "").strip(),
        str(r.get("parcela", "") or "").strip(),
        str(r.get("vencimento", "") or "").strip(),
    ])

def _mes_data_br(s):
    try:
        d = _dt.strptime(str(s or '').strip(), "%d/%m/%Y")
        return d.strftime("%Y-%m")
    except Exception:
        return ''

def _proximo_pagamento_mes(mes_ref):
    try:
        y,m = [int(x) for x in str(mes_ref).split('-')]
        if m == 12:
            y2,m2 = y+1,1
        else:
            y2,m2 = y,m+1
        return f"10/{m2:02d}/{y2}"
    except Exception:
        return ''

def _ftp_json(nome, default):
    try:
        import ftplib
        from io import BytesIO
        bio = BytesIO()
        ftp = ftplib.FTP()
        ftp.connect("moveisdolar.com.br", 21, timeout=20)
        ftp.login("moveisdolar3", "Deg27ll02mdl2301#")
        ftp.encoding = 'utf-8'
        ftp.cwd('/public_html/colaborador')
        ftp.retrbinary(f'RETR {nome}', bio.write)
        ftp.quit()
        bio.seek(0)
        return json.loads(bio.read().decode('utf-8'))
    except Exception:
        return default

# Atualiza métricas finais dos crediaristas após montar carteiras/recebimentos reais
for _fil_cred, _login_cred in CREDIARISTAS_FILIAIS.items():
    _pend_cred = sum(float(x.get('pendente', 0) or 0) for _fx in ['grave','alerta','atencao'] for x in clientes_crediarista_js.get(_login_cred, {}).get(_fx, []))
    _pago_cred = sum(float(x.get('pago', 0) or 0) for _fx in ['grave','alerta','atencao'] for x in recebimentos_crediarista_js.get(_login_cred, {}).get(_fx, []))
    if _login_cred in credenciais:
        credenciais[_login_cred]['pendente'] = round(_pend_cred, 2)
        credenciais[_login_cred]['pago'] = round(_pago_cred, 2)
        credenciais[_login_cred]['total'] = round(_pend_cred + _pago_cred, 2)

def _historico_comissao_cobranca10():
    hist = _load_json_safe(COMISSAO_HIST_PATH, {"months": {}})
    logs = _ftp_json('cobrancas_log.json', [])
    mes_atual = _dt.now().strftime('%Y-%m')
    cfg_rows = (CONFIG_META.get('camp_cobranca_terceiro') or _config_meta_default_global.get('camp_cobranca_terceiro') or [])
    pct_fx = {str(x.get('faixa','')).lower(): float(str(x.get('pct',0)).replace(',','.')) for x in cfg_rows if x}
    logs_mes = [x for x in (logs or []) if str(x.get('usuario','')).lower() in (COBRANCA10_LOGIN, COBRANCA10_NOME.lower()) and str(x.get('server_time',''))[:7] == mes_atual]
    keys = {_row_key_cob_json(x) for x in logs_mes}
    recebido = {'atencao':0.0,'alerta':0.0,'grave':0.0}
    for fx in ['atencao','alerta','grave']:
        for r in recebimentos_terceiro_js.get(fx, []):
            if _row_key_cob_json(r) in keys and _mes_data_br(r.get('pagamento','')) == mes_atual:
                recebido[fx] += float(r.get('pago',0) or 0)
    com = {fx: round(recebido[fx] * (pct_fx.get(fx,0)/100.0), 2) for fx in recebido}
    hist.setdefault('months', {})[mes_atual] = {
        'mes': mes_atual,
        'usuario': COBRANCA10_LOGIN,
        'nome': COBRANCA10_NOME,
        'pagamento_previsto_em': _proximo_pagamento_mes(mes_atual),
        'recebido': recebido,
        'comissao': com,
        'total_comissao': round(sum(com.values()), 2),
        'updated_at': now_brasilia().isoformat(),
    }
    with open(COMISSAO_HIST_PATH, 'w', encoding='utf-8') as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)
    return hist


def cache_path(d):   return os.path.join(CACHE_DIR, f"snapshot_{d}.json")
def meta_path(m):    return os.path.join(CACHE_DIR, f"meta_{m}.json")
def ref_path(m):     return os.path.join(CACHE_DIR, f"snapshot_referencia_{m}.json")

def _load_json_file(_path, _default):
    try:
        if os.path.exists(_path):
            with open(_path, encoding='utf-8') as _f:
                _data = json.load(_f)
            if isinstance(_data, type(_default)):
                return _data
    except Exception:
        pass
    return _default

def _list_snapshot_dates():
    _out=[]
    for _name in os.listdir(CACHE_DIR):
        if _name.startswith('snapshot_') and _name.endswith('.json') and 'referencia' not in _name:
            _date=_name.replace('snapshot_','').replace('.json','')
            if len(_date)==10:
                _out.append(_date)
    return sorted(set(_out))

def _find_last_snapshot_before_date(_date_str):
    _cands=[d for d in _list_snapshot_dates() if d < _date_str]
    if not _cands:
        return None
    _last=_cands[-1]
    _path=cache_path(_last)
    try:
        with open(_path, encoding='utf-8') as _f:
            _snap=json.load(_f)
        _snap['_path']=_path
        return _snap
    except Exception:
        return None

def _find_last_snapshot_of_month(_month_str):
    _cands=[d for d in _list_snapshot_dates() if str(d).startswith(_month_str)]
    if not _cands:
        return None
    _last=_cands[-1]
    try:
        with open(cache_path(_last), encoding='utf-8') as _f:
            _snap=json.load(_f)
        _snap['_path']=cache_path(_last)
        return _snap
    except Exception:
        return None

def ensure_month_reference_locked(_month_str, _today_str):
    """V10.5: a referência mensal de cobrança deve ser sempre o dia 01 da competência.
    Isso evita que um snapshot futuro ou do fim do mês zere os recebimentos por faixa.
    """
    _rp = ref_path(_month_str)
    _base_data = f'{_month_str}-01'
    _precisa = True
    try:
        if os.path.exists(_rp):
            with open(_rp, encoding='utf-8') as _f:
                _cur = json.load(_f) or {}
            if str(_cur.get('data') or '')[:10] == _base_data:
                return _rp
            print(f'⚠️ Referência mensal antiga/inválida em {_rp}: {_cur.get("data")} → corrigindo para {_base_data}')
    except Exception as _e_ref_lock:
        print(f'⚠️ Erro lendo referência mensal em {_rp}: {_e_ref_lock}')

    _ref_payload = {
        'data': _base_data,
        'gerado_em': now_brasilia().isoformat(),
        'origem': 'inicio_mes_forcado_v10_7',
        'snapshot_origem_path': '',
        'snapshot_origem_data': _base_data,
        'observacao': 'Corte mensal fixo para recebimentos por faixa; não usar snapshot futuro nem fim do mês.',
    }
    with open(_rp, 'w', encoding='utf-8') as _f:
        json.dump(_ref_payload, _f, ensure_ascii=False, indent=2)
    print(f'🧷 Referência mensal travada/corrigida: {_rp} → base {_ref_payload["data"]}')
    return _rp

def fechar_mes_anterior_travado(_mes_atual, _hist_dash, _config_global, _config_ind):
    _fech = _load_json_file(_fechamento_path, {'months':{}})
    _fech.setdefault('months', {})
    _first_day = datetime.strptime(_mes_atual + '-01', '%Y-%m-%d')
    _prev_month = (_first_day - timedelta(days=1)).strftime('%Y-%m')
    if _prev_month in _fech['months']:
        return _fech
    _prev_meta_path = meta_path(_prev_month)
    if not os.path.exists(_prev_meta_path):
        return _fech
    try:
        with open(_prev_meta_path, encoding='utf-8') as _f:
            _prev_meta = json.load(_f)
    except Exception:
        return _fech
    _last_prev_snap = _find_last_snapshot_of_month(_prev_month)
    _prev_dates = sorted([d for d in (_hist_dash.get('dates', {}) or {}).keys() if str(d).startswith(_prev_month)])
    _last_hist_date = _prev_dates[-1] if _prev_dates else None
    _resumo_final = (_hist_dash.get('dates', {}).get(_last_hist_date, {}) if _last_hist_date else {})
    _entry = {
        'mes': _prev_month,
        'fechado_em': now_brasilia().isoformat(),
        'meta_file': os.path.basename(_prev_meta_path),
        'meta_final': _prev_meta,
        'empresa_final': _resumo_final.get('empresa', {}),
        'filiais_finais': _resumo_final.get('filiais', {}),
        'vendedores_finais': _resumo_final.get('vendedores', {}),
        'ultimo_dia_historico': _last_hist_date,
        'snapshot_final_data': (_last_prev_snap or {}).get('data', _last_hist_date),
        'snapshot_final_path': (_last_prev_snap or {}).get('_path', ''),
        'config_global_fechamento': _config_global,
        'config_individual_fechamento': _config_ind,
    }
    _fech['months'][_prev_month] = _entry
    with open(_fechamento_path, 'w', encoding='utf-8') as _f:
        json.dump(_fech, _f, ensure_ascii=False, indent=2)
    print(f'📦 Fechamento mensal travado salvo: {_prev_month}')
    return _fech


ensure_month_reference_locked(mes_str, hoje_str)

# ── Base mensal opcional via relatório de início do mês ───────────────
MESES_PT = {
    "jan": "01", "janeiro": "01",
    "fev": "02", "fevereiro": "02",
    "mar": "03", "marco": "03", "março": "03",
    "abr": "04", "abril": "04",
    "mai": "05", "maio": "05",
    "jun": "06", "junho": "06",
    "jul": "07", "julho": "07",
    "ago": "08", "agosto": "08",
    "set": "09", "setembro": "09",
    "out": "10", "outubro": "10",
    "nov": "11", "novembro": "11",
    "dez": "12", "dezembro": "12",
}

def _month_from_filename(_name):
    _n = os.path.basename(str(_name)).lower()
    _n = _n.replace("ç", "c")
    for _mes_nome, _mes_num in sorted(MESES_PT.items(), key=lambda x: len(x[0]), reverse=True):
        if _mes_nome in _n:
            _m = re.search(r"(20\d{2}|\d{2})", _n)
            if _m:
                _yy = _m.group(1)
                _yy = f"20{_yy}" if len(_yy) == 2 else _yy
                return f"{_yy}-{_mes_num}"
    return None

def _find_base_report_for_month(_month_str):
    _cands = []
    for _root in [pasta, CACHE_DIR]:
        try:
            for _name in os.listdir(_root):
                _low = _name.lower()
                if ("meta inicio" in _low or "meta_inicio" in _low) and _low.endswith((".xls", ".xlsx")):
                    if _month_from_filename(_name) == _month_str:
                        _cands.append(os.path.join(_root, _name))
        except Exception:
            pass
    if not _cands:
        return None
    _cands = sorted(set(_cands), key=lambda p: os.path.getmtime(p), reverse=True)
    return _cands[0]

def _parse_base_report_faixas(_path, _month_str):
    _base_date = datetime.strptime(f"{_month_str}-01", "%Y-%m-%d")
    _dfb = pd.read_excel(_path, header=None)
    _vend_atual = None
    _out_v = {}
    _out_f = {}

    for _i in range(len(_dfb)):
        _row = _dfb.iloc[_i]
        if observacao_deve_ignorar(_row):
            continue
        _c0 = str(_row[COL["filial"]]).strip()
        _c1 = str(_row[COL["cliente"]]).strip()

        if "Vendedor:" in _c0:
            _vend_atual = limpar_nome_erp(_c0)
            continue

        if _c1 in ("nan", "", "Cliente", "Filial"):
            continue
        if _c0.upper().startswith("FILIAL") and _c1 in ("nan", "", "Cliente", "Filial"):
            continue
        if not _vend_atual:
            continue

        _venc = _parse_data(_row[COL["vencimento"]])
        _pend = tratar_valor(_row[COL["pendente"]])
        if not _venc or _pend <= 0:
            continue
        if str(_row[COL["conta_caixa"]]).strip() == "Caixa Filial 100":
            continue

        _dias = (_base_date - _venc).days
        if _dias >= 60:
            _fx = "grave"
        elif 30 <= _dias < 60:
            _fx = "alerta"
        elif 15 <= _dias < 30:
            _fx = "atencao"
        else:
            continue

        _fv, _ig = extrair_filial_nome(_vend_atual)
        _nome = limpar_nome_display(_vend_atual)

        if _fv is not None:
            _filial = str(_fv).strip().upper()
        else:
            _lookup = _filial_inativo_lookup.get(_nome, "")
            if _lookup == "FDEP" or _nome in _nomes_fdep:
                continue
            elif _lookup:
                _filial = str(_lookup).strip().upper()
            else:
                _fil_txt = str(_row[COL["filial"]]).upper()
                _mfil = re.search(r"FILIAL\s*(\d+)", _fil_txt)
                _filial = f"F{int(_mfil.group(1))}" if _mfil else "OUTROS"

        _key = f"{_nome}_{_filial}"
        _out_v.setdefault(_key, {"grave": 0.0, "alerta": 0.0, "atencao": 0.0})
        _out_f.setdefault(_filial, {"grave": 0.0, "alerta": 0.0, "atencao": 0.0})
        _out_v[_key][_fx] += float(_pend)
        _out_f[_filial][_fx] += float(_pend)

    for _d in (_out_v, _out_f):
        for _k, _vals in _d.items():
            for _fx in ["grave", "alerta", "atencao"]:
                _vals[_fx] = round(float(_vals.get(_fx, 0) or 0), 2)
    return {"data": f"{_month_str}-01", "vendedores": _out_v, "filiais": _out_f, "arquivo": os.path.basename(_path)}

BASE_MENSAL_INFO = {"data": f"{mes_str}-01", "vendedores": {}, "filiais": {}, "arquivo": ""}
_base_report_path = _find_base_report_for_month(mes_str)
if _base_report_path:
    try:
        BASE_MENSAL_INFO = _parse_base_report_faixas(_base_report_path, mes_str)
        _base_meta_path = os.path.join(CACHE_DIR, f"base_meta_{mes_str}.json")
        with open(_base_meta_path, "w", encoding="utf-8") as _fbm:
            json.dump(BASE_MENSAL_INFO, _fbm, ensure_ascii=False, indent=2)
        with open(ref_path(mes_str), "w", encoding="utf-8") as _fr:
            json.dump({
                "data": BASE_MENSAL_INFO["data"],
                "gerado_em": now_brasilia().isoformat(),
                "origem": "arquivo_base_relatorio",
                "arquivo_base": os.path.basename(_base_report_path),
                "snapshot_origem_data": BASE_MENSAL_INFO["data"],
            }, _fr, ensure_ascii=False, indent=2)
        print(f"🧷 Base mensal carregada do relatório: {os.path.basename(_base_report_path)}")
    except Exception as _e_base:
        print(f"⚠️ Erro lendo base mensal {_base_report_path}: {_e_base}")
else:
    _base_meta_path = os.path.join(CACHE_DIR, f"base_meta_{mes_str}.json")
    if os.path.exists(_base_meta_path):
        try:
            with open(_base_meta_path, "r", encoding="utf-8") as _fbm:
                BASE_MENSAL_INFO = json.load(_fbm)
            print(f"🧷 Base mensal reutilizada do cache: {os.path.basename(_base_meta_path)}")
        except Exception:
            pass

# ── Snapshot de hoje ──────────────────────────────────────────────────
snapshot_hoje = {
    "data": hoje_str, "gerado_em": now_brasilia().isoformat(),
    "total_p": total_final_p, "total_pg": total_final_pg,
    "filiais": {}, "vendedores": {}
}

for f in ORDEM_FILIAIS:
    sub = df_vend[df_vend["filial_vendedor"] == f]
    if not sub.empty:
        snapshot_hoje["filiais"][f] = {
            "pendente": round(sub["pendente"].sum(), 2),
            "pago":     round(sub["pago"].sum(), 2), "sem_ativo": False
        }
    elif f in filiais_sem_ativo:
        snapshot_hoje["filiais"][f] = {
            "pendente": round(filiais_sem_ativo[f]["pendente"], 2),
            "pago":     round(filiais_sem_ativo[f]["pago"], 2), "sem_ativo": True
        }

for _, row in df_vend.iterrows():
    key = f"{row['nome_exibicao']}_{row['filial_vendedor']}"
    snapshot_hoje["vendedores"][key] = {
        "nome": str(row["nome_exibicao"]), "filial": str(row["filial_vendedor"]),
        "pendente": round(float(row["pendente"]), 2),
        "pago":     round(float(row["pago"]), 2),
        "total":    round(float(row["total"]), 2),
        "is_gerente": bool(row["is_gerente"]),
        # Totais de pago por faixa (para calcular delta no próximo dia)
        "pago_grave":   round(recebido_faixa.get(
            f"{row['nome_exibicao']}_{row['filial_vendedor']}", {}).get('grave', 0.0), 2),
        "pago_alerta":  round(recebido_faixa.get(
            f"{row['nome_exibicao']}_{row['filial_vendedor']}", {}).get('alerta', 0.0), 2),
        "pago_atencao": round(recebido_faixa.get(
            f"{row['nome_exibicao']}_{row['filial_vendedor']}", {}).get('atencao', 0.0), 2),
    }

# Salva chaves dos títulos de hoje para identificar novos amanhã
snapshot_hoje["clientes_keys"] = list(clientes_key_hoje)

with open(cache_path(hoje_str), "w", encoding="utf-8") as f:
    json.dump(snapshot_hoje, f, ensure_ascii=False, indent=2)
print(f"💾 Snapshot salvo: {cache_path(hoje_str)}")

# ── Comparativos ──────────────────────────────────────────────────────
def load_snapshot(d):
    p = cache_path(d)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f: return json.load(f)
    return None

ontem_str  = (now_brasilia() - timedelta(days=1)).strftime("%Y-%m-%d")
snap_ontem = load_snapshot(ontem_str)
semana_str = (now_brasilia() - timedelta(days=7)).strftime("%Y-%m-%d")
snap_semana= load_snapshot(semana_str)

if not snap_ontem:
    for d in range(2, 30):
        snap_ontem = load_snapshot((now_brasilia()-timedelta(days=d)).strftime("%Y-%m-%d"))
        if snap_ontem: break

# Identifica clientes NOVOS agora que snap_ontem está disponível
_clientes_ontem = set(snap_ontem.get("clientes_keys", [])) if snap_ontem else set()
_novos_keys = clientes_key_hoje - _clientes_ontem
for _c in clientes_cobrar:
    _c['novo'] = _c['titulo_key'] in _novos_keys
_nn = sum(1 for _c in clientes_cobrar if _c.get('novo'))
print(f"   🆕 Clientes novos hoje: {_nn}")

# ── Variações ─────────────────────────────────────────────────────────
var_vendedores = {}
if snap_ontem:
    for key, vd in snapshot_hoje["vendedores"].items():
        ant = snap_ontem.get("vendedores", {}).get(key)
        if ant:
            delta = round(vd["pago"] - ant["pago"], 2)
            perc  = round(delta / ant["pago"] * 100, 1) if ant["pago"] > 0 else None
            var_vendedores[key] = {
                "delta_pago": delta, "perc_pago": perc,
                "periodo": snap_ontem["data"]
            }

var_filiais = {}
if snap_ontem:
    for f, fd in snapshot_hoje["filiais"].items():
        ant = snap_ontem.get("filiais", {}).get(f)
        if ant:
            delta = round(fd["pago"] - ant["pago"], 2)
            perc  = round(delta / ant["pago"] * 100, 1) if ant["pago"] > 0 else None
            var_filiais[f] = {
                "delta_pago": delta, "perc_pago": perc,
                "periodo": snap_ontem["data"]
            }

# ── META MENSAL (nova lógica por faixa) ──────────────────────────────
# Cada faixa tem meta própria baseada no PENDENTE daquela faixa:
#   Grave  → meta = 20% do pendente grave
#   Alerta → meta = 15% do pendente alerta
#   Atenção→ meta = 10% do pendente atenção
# Gráfico geral = média ponderada 60/30/10 dos 3 gráficos
#
# O pendente por faixa é estimado do total pendente com os pesos 50/30/20
# (grave≈50%, alerta≈30%, atenção≈20% do total — estimativa para o alvo)
# O recebimento por faixa é o delta acumulado × peso da faixa

meta_file = meta_path(mes_str)
if os.path.exists(meta_file):
    with open(meta_file, encoding="utf-8") as f:
        meta_mes = json.load(f)
    meta_mes["ultima_atualizacao"] = hoje_str
    meta_mes["snapshots_count"] = meta_mes.get("snapshots_count", 0) + 1
else:
    meta_mes = {
        "mes": mes_str, "criado_em": hoje_str,
        "ultima_atualizacao": hoje_str, "snapshots_count": 1,
        "vendedores": {}, "filiais": {}
    }

def calc_pendente_faixas(pendente_total):
    """Estima pendente por faixa a partir do total pendente."""
    return {
        "grave":   round(pendente_total * 0.50, 2),
        "alerta":  round(pendente_total * 0.30, 2),
        "atencao": round(pendente_total * 0.20, 2),
    }

def get_base_faixas_vendedor(key, pendente_total):
    _base = (BASE_MENSAL_INFO.get("vendedores", {}) or {}).get(key)
    if isinstance(_base, dict) and sum(float(_base.get(_fx, 0) or 0) for _fx in ["grave","alerta","atencao"]) > 0:
        return {
            "grave": round(float(_base.get("grave", 0) or 0), 2),
            "alerta": round(float(_base.get("alerta", 0) or 0), 2),
            "atencao": round(float(_base.get("atencao", 0) or 0), 2),
        }
    return calc_pendente_faixas(pendente_total)

def get_base_faixas_filial(filial, pendente_total):
    _base = (BASE_MENSAL_INFO.get("filiais", {}) or {}).get(filial)
    if isinstance(_base, dict) and sum(float(_base.get(_fx, 0) or 0) for _fx in ["grave","alerta","atencao"]) > 0:
        return {
            "grave": round(float(_base.get("grave", 0) or 0), 2),
            "alerta": round(float(_base.get("alerta", 0) or 0), 2),
            "atencao": round(float(_base.get("atencao", 0) or 0), 2),
        }
    return calc_pendente_faixas(pendente_total)

def calc_metas_alvo(faixas, cfg=None):
    """Calcula alvo de meta de cada faixa usando config (global ou individual)."""
    c = cfg or CONFIG_META
    return {
        "grave_alvo":   round(faixas["grave"]   * c["grave_pct"]   / 100, 2),
        "alerta_alvo":  round(faixas["alerta"]  * c["alerta_pct"]  / 100, 2),
        "atencao_alvo": round(faixas["atencao"] * c["atencao_pct"] / 100, 2),
    }

def calc_perc_geral(grave_perc, alerta_perc, atencao_perc, cfg=None):
    """Gráfico geral ponderado pelos pesos da config."""
    c = cfg or CONFIG_META
    pg = c["peso_grave"]   / 100
    pa = c["peso_alerta"]  / 100
    pt = c["peso_atencao"] / 100
    return round(grave_perc * pg + alerta_perc * pa + atencao_perc * pt, 1)

# =========================================================================
# RECEBIMENTOS E META — FONTE ÚNICA
# Regra:
#   1. Lê todos os títulos pagos no período (pagto > data_corte, conta caixa ≠ 100)
#   2. Classifica por faixa de vencimento (grave/alerta/atenção)
#   3. Distribui inativos e FDEP para ativos da filial (60% gerente / 40% vendedores)
#   4. recebimentos_det_js = relatório visual (com detalhes dos títulos)
#   5. recebido_faixa = somatório por faixa por vendedor (para os gráficos de meta)
#   OS DOIS SÃO DERIVADOS DA MESMA FONTE → valores sempre batem
# =========================================================================

# Passo A: Lê títulos brutos do XLS (todos os vendedores)
# _rec_raw[chave] = {grave:[], alerta:[], atencao:[], is_ativo, is_fdep, filial}
_rec_raw = {}
_vd2 = None

for _i2 in range(len(df_raw)):
    _row2 = df_raw.iloc[_i2]
    if observacao_deve_ignorar(_row2):
        continue
    _c02 = str(_row2[COL["filial"]]).strip()
    _c12 = str(_row2[COL["cliente"]]).strip()

    if "Vendedor:" in _c02:
        _vd2 = limpar_nome_erp(_c02)
        continue

    if _c12 in ("nan", "", "Cliente", "Filial"):
        continue
    if not _vd2:
        continue

    _venc2 = _parse_data(_row2[COL["vencimento"]])
    _pagto2 = _parse_data(_row2[COL["pagamento"]])
    _pago2 = tratar_valor(_row2[COL["pago_total"]])

    if not _venc2 or not _pagto2 or _pago2 <= 0:
        continue
    if _pagto2 < _data_corte_parse:
        continue

    if str(_row2[COL["conta_caixa"]]).strip() == "Caixa Filial 100":
        continue

    _dias2 = (_hoje - _venc2).days

    if _dias2 >= 60:
        _f2 = "grave"
    elif 30 <= _dias2 < 60:
        _f2 = "alerta"
    elif 15 <= _dias2 < 30:
        _f2 = "atencao"
    else:
        continue

    _fv2, _ = extrair_filial_nome(_vd2)
    _nome2 = limpar_nome_display(_vd2)

    if _fv2 is not None:
        _fv2k = _fv2
        _is_at2 = True
        _is_fp2 = False
    else:
        _lk2 = _filial_inativo_lookup.get(_nome2, "")
        if _lk2 == "FDEP" or _nome2 in _nomes_fdep:
            _fv2k = "FDEP"
            _is_at2 = False
            _is_fp2 = True
        elif _lk2:
            _fv2k = _lk2
            _is_at2 = False
            _is_fp2 = False
        else:
            _fv2k = "OUTROS"
            _is_at2 = False
            _is_fp2 = False

    _kv2 = f"{_nome2}_{_fv2k}"

    if _kv2 not in _rec_raw:
        _rec_raw[_kv2] = {
            "grave": [],
            "alerta": [],
            "atencao": [],
            "is_ativo": _is_at2,
            "is_fdep": _is_fp2,
            "filial": _fv2k,
        }

    _rec_raw[_kv2][_f2].append({
        "cliente": _c12[:40],
        "dias": _dias2,
        "pago": _pago2,
        "vencimento": str(_row2[COL["vencimento"]]).strip(),
        "pagamento": str(_row2[COL["pagamento"]]).strip(),
        "parcela": str(_row2[COL["num_parcela"]]).strip(),
        "titulo": str(_row2[COL["num_titulo"]]).strip(),
        "vendedor": (_nome2 + ("" if _is_at2 else (" [FDEP]" if _is_fp2 else " [Inativo]")))[:30],
    })

# Passo B: Separa ativos, inativos por filial, e FDEP (lista única)
_inat_por_fil = {}   # {filial: {faixa: [títulos]}}
_fdep_lista   = {'grave':[], 'alerta':[], 'atencao':[]}

for _kv2, _rb in _rec_raw.items():
    if _rb['is_ativo']:
        continue
    elif _rb['is_fdep']:
        for _fx in ['grave','alerta','atencao']:
            _fdep_lista[_fx].extend(_rb[_fx])
    else:
        _ff2 = _rb['filial']
        if _ff2 not in _inat_por_fil:
            _inat_por_fil[_ff2] = {'grave':[], 'alerta':[], 'atencao':[]}
        for _fx in ['grave','alerta','atencao']:
            _inat_por_fil[_ff2][_fx].extend(_rb[_fx])

# Distribui FDEP por filial proporcional ao peso (SEM duplicar títulos)
_fdep_por_fil = {}
if tw > 0:
    _fils_ord = sorted([(f, pesos.get(f,0)) for f in ORDEM_FILIAIS if pesos.get(f,0)>0],
                       key=lambda x: x[1], reverse=True)
    for _fx in ['grave','alerta','atencao']:
        _todos = _fdep_lista[_fx][:]
        _n = len(_todos); _s = 0
        for _ff_f, _w_f in _fils_ord:
            _q = round(_n * _w_f / tw)
            _e = min(_s + _q, _n)
            if _ff_f not in _fdep_por_fil:
                _fdep_por_fil[_ff_f] = {'grave':[], 'alerta':[], 'atencao':[]}
            _fdep_por_fil[_ff_f][_fx] = _todos[_s:_e]
            _s = _e

# Passo C: Constrói recebimentos_det_js (relatório visual)
# Chaves finais: "NOME_Fx" para ativos, "Filial Fx_Fx" para filiais sem ativo
_chaves_validas   = {f"{r['nome_exibicao']}_{r['filial_vendedor']}" for _, r in df_vend.iterrows()}
_chaves_sem_ativo = {f"Filial {f}_{f}" for f in filiais_sem_ativo}

recebimentos_det_js = {}

# 1) Títulos próprios dos ativos
for _kv2, _rb in _rec_raw.items():
    if _rb['is_ativo']:
        recebimentos_det_js[_kv2] = {
            'grave':  list(_rb['grave']),
            'alerta': list(_rb['alerta']),
            'atencao':list(_rb['atencao']),
        }

# 2) Distribui inativos e FDEP: 60% gerente / 40% vendedores por filial
for _fil_r in ORDEM_FILIAIS:
    _vends_r2   = [r['nome_exibicao'] for _, r in df_vend[
        (df_vend['filial_vendedor']==_fil_r)&(~df_vend['is_gerente'])].iterrows()]
    _gerents_r2 = [r['nome_exibicao'] for _, r in df_vend[
        (df_vend['filial_vendedor']==_fil_r)&( df_vend['is_gerente'])].iterrows()]
    _sem_at_r   = _fil_r in filiais_sem_ativo

    for _src2 in [_inat_por_fil.get(_fil_r,{}), _fdep_por_fil.get(_fil_r,{})]:
        for _fx in ['grave','alerta','atencao']:
            _titulos2 = _src2.get(_fx, [])
            if not _titulos2: continue

            if _sem_at_r:
                # Filial sem ativo: 100% vai para a filial
                _kf2 = f"Filial {_fil_r}_{_fil_r}"
                if _kf2 not in recebimentos_det_js:
                    recebimentos_det_js[_kf2] = {'grave':[],'alerta':[],'atencao':[]}
                recebimentos_det_js[_kf2][_fx].extend(_titulos2)
            else:
                _ng2  = round(len(_titulos2) * 0.60)
                _gp2  = _titulos2[:_ng2]
                _vp2  = _titulos2[_ng2:]
                # Gerente recebe 60%
                for _g2 in _gerents_r2:
                    _kg2 = f"{_g2}_{_fil_r}"
                    if _kg2 not in recebimentos_det_js:
                        recebimentos_det_js[_kg2] = {'grave':[],'alerta':[],'atencao':[]}
                    recebimentos_det_js[_kg2][_fx].extend(_gp2)
                # Vendedores recebem 40% em round-robin
                if _vends_r2:
                    for _ii2, _t2 in enumerate(_vp2):
                        _kv2d = f"{_vends_r2[_ii2 % len(_vends_r2)]}_{_fil_r}"
                        if _kv2d not in recebimentos_det_js:
                            recebimentos_det_js[_kv2d] = {'grave':[],'alerta':[],'atencao':[]}
                        recebimentos_det_js[_kv2d][_fx].append(_t2)
                elif _gerents_r2:
                    # Sem vendedores ativos: tudo vai para o gerente
                    for _t2 in _vp2:
                        _kg2 = f"{_gerents_r2[0]}_{_fil_r}"
                        if _kg2 not in recebimentos_det_js:
                            recebimentos_det_js[_kg2] = {'grave':[],'alerta':[],'atencao':[]}
                        recebimentos_det_js[_kg2][_fx].append(_t2)

# Filtra: mantém só chaves de vendedores ativos + filiais sem ativo
recebimentos_det_js = {
    k: {'grave': sorted(v['grave'], key=lambda x: x['pago'], reverse=True),
        'alerta':sorted(v['alerta'],key=lambda x: x['pago'], reverse=True),
        'atencao':sorted(v['atencao'],key=lambda x: x['pago'],reverse=True)}
    for k, v in recebimentos_det_js.items()
    if k in _chaves_validas or k in _chaves_sem_ativo
}

recebimentos_terceiro_js = {'grave': [], 'alerta': [], 'atencao': []}
for _krt, _vrt in recebimentos_det_js.items():
    for _fxrt in ['grave', 'alerta', 'atencao']:
        for _rrt in (_vrt.get(_fxrt) or []):
            if cobranca_row_key_py(_rrt) in _clientes_terceiro_keys:
                recebimentos_terceiro_js[_fxrt].append(_rrt)
for _fxrt in ['grave', 'alerta', 'atencao']:
    recebimentos_terceiro_js[_fxrt].sort(key=lambda x: float(x.get('pago', 0) or 0), reverse=True)

# Passo D: Deriva recebido_faixa — MESMA FONTE que o relatório
# recebido_faixa[chave] = soma dos valores pagos por faixa
# GARANTE que meta e relatório usam exatamente os mesmos números
recebido_faixa = {}

for _kd, _vd_det in recebimentos_det_js.items():
    _nome_d   = _kd.rsplit('_', 1)[0] if '_' in _kd else _kd
    _filial_d = _kd.rsplit('_', 1)[1] if '_' in _kd else 'OUTROS'

    # 🔥 identifica se é filial sem ativo
    _is_filial_sem_ativo = _kd in _chaves_sem_ativo

    recebido_faixa[_kd] = {
        'grave':   sum(t['pago'] for t in _vd_det['grave']),
        'alerta':  sum(t['pago'] for t in _vd_det['alerta']),
        'atencao': sum(t['pago'] for t in _vd_det['atencao']),

        # 🔥 CORREÇÃO PRINCIPAL
        'is_ativo': (_kd not in _chaves_sem_ativo),

        'is_fdep':  False,
        'filial':   _filial_d,
    }

    # 🔥 CORREÇÃO FINAL (ESSA RESOLVE A F5)
    if _kd.startswith("Filial "):
        recebido_faixa[_kd]['is_ativo'] = False


_pre_inat_rec = {f: {'grave':0.0,'alerta':0.0,'atencao':0.0} for f in ORDEM_FILIAIS}
_pre_fdep_rec = {f: {'grave':0.0,'alerta':0.0,'atencao':0.0} for f in ORDEM_FILIAIS}
_pre_ativos_filial = {f: 0 for f in ORDEM_FILIAIS}  # conta vendedores ativos p/ divisão

for _kk, _rv in recebido_faixa.items():
    _is_at = _rv.get('is_ativo', False)
    _is_fp = _rv.get('is_fdep', False)
    _ff_k  = _rv.get('filial', '')

    # 🔥 ATIVOS → contam normalmente
    if _is_at and _ff_k in _pre_ativos_filial:
        _pre_ativos_filial[_ff_k] += 1

    # 🔥 INATIVOS + FILIAL SEM ATIVO (CORRIGIDO)
    elif not _is_at and not _is_fp:

        # 🔥 DETECTA SE É BLOCO DE FILIAL (ex: "Filial F5_F5")
        if _kk.startswith("Filial "):
            if _ff_k in _pre_inat_rec:
                for _fx in ['grave','alerta','atencao']:
                    _pre_inat_rec[_ff_k][_fx] += _rv.get(_fx, 0.0)

        else:
            # 🔥 INATIVO NORMAL
            if _ff_k in _pre_inat_rec:
                for _fx in ['grave','alerta','atencao']:
                    _pre_inat_rec[_ff_k][_fx] += _rv.get(_fx, 0.0)

    # 🔥 FDEP → rateio proporcional
    elif _is_fp:
        for _ff2 in ORDEM_FILIAIS:
            _w2 = pesos.get(_ff2, 0)
            if tw > 0 and _w2 > 0:
                for _fx in ['grave','alerta','atencao']:
                    _pre_fdep_rec[_ff2][_fx] += _rv.get(_fx, 0.0) * _w2 / tw

# Atualiza meta de vendedores
for key, vd in snapshot_hoje["vendedores"].items():
    ant_pago = 0.0
    if snap_ontem:
        ant_vd = snap_ontem.get("vendedores", {}).get(key)
        if ant_vd: ant_pago = ant_vd["pago"]
    delta = max(vd["pago"] - ant_pago, 0)

    _cfg_key = get_config_meta(key)  # config individual ou global
    if key not in meta_mes["vendedores"]:
        faixas = get_base_faixas_vendedor(key, vd["pendente"])
        alvos  = calc_metas_alvo(faixas, _cfg_key)
        meta_mes["vendedores"][key] = {
            "nome": vd["nome"], "filial": vd["filial"], "is_gerente": vd["is_gerente"],
            # Recebimentos acumulados por faixa
            "grave_rec":   0.0, "alerta_rec":  0.0, "atencao_rec": 0.0,
            # Alvos fixos por faixa
            "grave_alvo":   alvos["grave_alvo"],
            "alerta_alvo":  alvos["alerta_alvo"],
            "atencao_alvo": alvos["atencao_alvo"],
            # Pendentes da faixa (base do alvo)
            "grave_pend":   faixas["grave"],
            "alerta_pend":  faixas["alerta"],
            "atencao_pend": faixas["atencao"],
            "snapshots": []
        }

    vm = meta_mes["vendedores"][key]
    # Acumula delta pelos pesos das faixas
    # Recebimento real do período por faixa (usa pré-cálculo feito antes do loop)
    _filial_vend = vd.get("filial", "")
    _is_ger      = vd.get("is_gerente", False)

    # Próprio: apenas se é vendedor ativo registrado em recebido_faixa
    _rv_proprio = recebido_faixa.get(key, {})
    _grave_delta   = _rv_proprio.get('grave',   0.0) if _rv_proprio.get('is_ativo') else 0.0
    _alerta_delta  = _rv_proprio.get('alerta',  0.0) if _rv_proprio.get('is_ativo') else 0.0
    _atencao_delta = _rv_proprio.get('atencao', 0.0) if _rv_proprio.get('is_ativo') else 0.0

    # Parcela de inativos + FDEP da filial
    if _filial_vend:
        _n_at = max(_pre_ativos_filial.get(_filial_vend, 1), 1)
        for _fx in ['grave','alerta','atencao']:
            _tot = (_pre_inat_rec.get(_filial_vend, {}).get(_fx, 0.0) +
                    _pre_fdep_rec.get(_filial_vend, {}).get(_fx, 0.0))
            _share = _tot * 0.60 if _is_ger else _tot * 0.40 / _n_at
            if _fx == 'grave':    _grave_delta   += _share
            elif _fx == 'alerta': _alerta_delta  += _share
            else:                 _atencao_delta += _share

    # 🔥 CORREÇÃO: não acumular novamente a cada reprocessamento do mesmo período.
    # O gráfico do vendedor deve refletir os valores reais atuais do período, assim como já ocorre nas filiais.
    vm["grave_rec"]   = round(_grave_delta,   2)
    vm["alerta_rec"]  = round(_alerta_delta,  2)
    vm["atencao_rec"] = round(_atencao_delta, 2)
    vm["snapshots"].append({
        "data": hoje_str, "delta": delta,
        "grave_delta": _grave_delta, "alerta_delta": _alerta_delta, "atencao_delta": _atencao_delta
    })

    # Migração: se registro antigo não tem campos de faixa, recalcula
    if "grave_alvo" not in vm:
        faixas_mig = get_base_faixas_vendedor(key, vd.get("pendente", 0))
        alvos_mig  = calc_metas_alvo(faixas_mig)
        vm.update({
            "grave_alvo":   alvos_mig["grave_alvo"],
            "alerta_alvo":  alvos_mig["alerta_alvo"],
            "atencao_alvo": alvos_mig["atencao_alvo"],
            "grave_pend":   faixas_mig["grave"],
            "alerta_pend":  faixas_mig["alerta"],
            "atencao_pend": faixas_mig["atencao"],
        })

    # Base mensal do relatório, se existir, prevalece como alvo do mês
    _faixas_base_vm = get_base_faixas_vendedor(key, vd.get("pendente", 0))
    _alvos_base_vm = calc_metas_alvo(_faixas_base_vm, _cfg_key)
    vm["grave_pend"] = _faixas_base_vm["grave"]
    vm["alerta_pend"] = _faixas_base_vm["alerta"]
    vm["atencao_pend"] = _faixas_base_vm["atencao"]
    vm["grave_alvo"] = _alvos_base_vm["grave_alvo"]
    vm["alerta_alvo"] = _alvos_base_vm["alerta_alvo"]
    vm["atencao_alvo"] = _alvos_base_vm["atencao_alvo"]

    # Percentuais por faixa
    ga = vm["grave_alvo"];   gr = vm["grave_rec"]
    aa = vm["alerta_alvo"];  ar = vm["alerta_rec"]
    ta = vm["atencao_alvo"]; tr = vm["atencao_rec"]
    vm["grave_perc"]   = round(min(gr/ga*100, 100), 1) if ga > 0 else 0.0
    vm["alerta_perc"]  = round(min(ar/aa*100, 100), 1) if aa > 0 else 0.0
    vm["atencao_perc"] = round(min(tr/ta*100, 100), 1) if ta > 0 else 0.0
    vm["perc_meta"]    = calc_perc_geral(vm["grave_perc"], vm["alerta_perc"], vm["atencao_perc"], _cfg_key)

# Atualiza meta de filiais
for f, fd in snapshot_hoje["filiais"].items():

    ant_fd_pago = 0.0
    if snap_ontem:
        ant_fd = snap_ontem.get("filiais", {}).get(f, {})
        ant_fd_pago = ant_fd.get("pago", 0)

    delta_f = max(fd["pago"] - ant_fd_pago, 0)

    # =========================================
    # CRIA META SE NÃO EXISTIR
    # =========================================
    if f not in meta_mes["filiais"]:
        faixas_f = get_base_faixas_filial(f, fd["pendente"])
        alvos_f  = calc_metas_alvo(faixas_f)

        meta_mes["filiais"][f] = {
            "grave_rec": 0.0,
            "alerta_rec": 0.0,
            "atencao_rec": 0.0,

            "grave_alvo":   alvos_f["grave_alvo"],
            "alerta_alvo":  alvos_f["alerta_alvo"],
            "atencao_alvo": alvos_f["atencao_alvo"],

            "grave_pend":   faixas_f["grave"],
            "alerta_pend":  faixas_f["alerta"],
            "atencao_pend": faixas_f["atencao"],
        }

    fm = meta_mes["filiais"][f]

    # =========================================
    # 🔥 CORREÇÃO DEFINITIVA
    # =========================================
    _f_grave_delta  = 0.0
    _f_alerta_delta = 0.0
    _f_atenc_delta  = 0.0

    for _rv in recebido_faixa.values():
        if _rv.get("filial") == f:
            _f_grave_delta  += _rv.get("grave", 0.0)
            _f_alerta_delta += _rv.get("alerta", 0.0)
            _f_atenc_delta  += _rv.get("atencao", 0.0)

    # 🔥 VALOR REAL DO PERÍODO (NÃO ACUMULA)
    fm["grave_rec"]   = round(_f_grave_delta,  2)
    fm["alerta_rec"]  = round(_f_alerta_delta, 2)
    fm["atencao_rec"] = round(_f_atenc_delta,  2)

    # =========================================
    # 🔄 SEGURANÇA (migração)
    # =========================================
    if "grave_alvo" not in fm:
        faixas_mig_f = get_base_faixas_filial(f, fd.get("pendente", 0))
        alvos_mig_f  = calc_metas_alvo(faixas_mig_f)

        fm.update({
            "grave_alvo":   alvos_mig_f["grave_alvo"],
            "alerta_alvo":  alvos_mig_f["alerta_alvo"],
            "atencao_alvo": alvos_mig_f["atencao_alvo"],
            "grave_pend":   faixas_mig_f["grave"],
            "alerta_pend":  faixas_mig_f["alerta"],
            "atencao_pend": faixas_mig_f["atencao"],
        })

    # Base mensal do relatório, se existir, prevalece como alvo do mês
    _faixas_base_fm = get_base_faixas_filial(f, fd.get("pendente", 0))
    _alvos_base_fm = calc_metas_alvo(_faixas_base_fm)
    fm["grave_pend"] = _faixas_base_fm["grave"]
    fm["alerta_pend"] = _faixas_base_fm["alerta"]
    fm["atencao_pend"] = _faixas_base_fm["atencao"]
    fm["grave_alvo"] = _alvos_base_fm["grave_alvo"]
    fm["alerta_alvo"] = _alvos_base_fm["alerta_alvo"]
    fm["atencao_alvo"] = _alvos_base_fm["atencao_alvo"]

    # =========================================
    # 📊 % META
    # =========================================
    ga = fm["grave_alvo"];   gr = fm["grave_rec"]
    aa = fm["alerta_alvo"];  ar = fm["alerta_rec"]
    ta = fm["atencao_alvo"]; tr = fm["atencao_rec"]

    fm["grave_perc"]   = round(min(gr/ga*100,100),1) if ga>0 else 0.0
    fm["alerta_perc"]  = round(min(ar/aa*100,100),1) if aa>0 else 0.0
    fm["atencao_perc"] = round(min(tr/ta*100,100),1) if ta>0 else 0.0

    fm["perc_meta"] = calc_perc_geral(
        fm["grave_perc"],
        fm["alerta_perc"],
        fm["atencao_perc"]
    )

# =========================================
# 💾 SALVA
# =========================================
with open(meta_file, "w", encoding="utf-8") as f:
    json.dump(meta_mes, f, ensure_ascii=False, indent=2)

print(f"🎯 Meta do mês atualizada: {meta_file}")
print("\n===== DEBUG FILIAL F5 =====")

for k, v in recebido_faixa.items():
    if "F5" in str(v.get("filial")):
        print(k, v)
# ── Média trimestral ──────────────────────────────────────────────────
meses_trim = []
for i in range(3):
    m = (datetime.now() - timedelta(days=30*i)).strftime("%Y-%m")
    if os.path.exists(meta_path(m)):
        with open(meta_path(m), encoding="utf-8") as f:
            meses_trim.append(json.load(f))

media_trimestral_vend = {}
if meses_trim:
    all_keys = set()
    for mm in meses_trim: all_keys.update(mm.get("vendedores",{}).keys())
    for key in all_keys:
        percs = [mm["vendedores"][key]["perc_meta"]
                 for mm in meses_trim if key in mm.get("vendedores",{})]
        if percs:
            ref = next(mm["vendedores"][key] for mm in meses_trim if key in mm.get("vendedores",{}))
            media_trimestral_vend[key] = {
                "nome": ref["nome"], "filial": ref["filial"],
                "media_perc": round(sum(percs)/len(percs), 1),
                "meses_count": len(percs)
            }

media_trimestral_fil = {}
for f in ORDEM_FILIAIS:
    percs_f = [mm["filiais"].get(f,{}).get("perc_meta",0)
               for mm in meses_trim if f in mm.get("filiais",{})]
    if percs_f:
        media_trimestral_fil[f] = round(sum(percs_f)/len(percs_f), 1)


# (filtro aplicado anteriormente)


# Pré-calcula recebimentos de inativos e FDEP por filial (para rateio na meta)
# Feito FORA do loop para não recalcular a cada vendedor
# ── Destaque da semana ─────────────────────────────────────────────────
destaque_semana = None
if snap_semana:
    melhor_f, melhor_d = None, -999999
    for f, fd in snapshot_hoje["filiais"].items():
        d = fd["pago"] - snap_semana.get("filiais",{}).get(f,{}).get("pago",0)
        if d > melhor_d: melhor_d, melhor_f = d, f
    if melhor_f:
        ref_pago = snap_semana["filiais"].get(melhor_f,{}).get("pago",1)
        destaque_semana = {
            "filial": melhor_f,
            "delta_pago": round(melhor_d, 2),
            "perc_ganho": round(melhor_d/ref_pago*100,1) if ref_pago>0 else 0,
            "periodo": snap_semana["data"]
        }
        print(f"🏆 Destaque da semana: {melhor_f} (+R$ {melhor_d:,.2f})")

print(f"📊 Histórico: {len(list(Path(CACHE_DIR).glob('snapshot_*.json')))} snapshots disponíveis")



# =========================================
# SERIALIZAR DADOS PARA O HTML (v2)
# =========================================
js_creds = json.dumps(credenciais, ensure_ascii=False)
js_metas_vendas = json.dumps(metas_vendas_info.get('dados', {}), ensure_ascii=False)
js_metas_vendas_dia = json.dumps(metas_vendas_dia_info.get('dados', {}), ensure_ascii=False)


def _v54_num_pct_meta_diaria(v):
    """Converte '104,20%' / 104.2 / '1.234,56%' para float."""
    try:
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v or "").strip().replace("%", "").replace("R$", "").replace(" ", "")
        if not s:
            return 0.0
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return 0.0

def _v54_cell(row, *keys):
    for k in keys:
        if isinstance(row, dict) and row.get(k) not in (None, ""):
            return row.get(k)
    return ""


def _v101_money_float(v):
    """Converte valor BR de dinheiro para float. Usado para validar meta diária sem confundir Projetado/Total."""
    try:
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v or "").strip()
        if not s or s.lower() in ("nan", "none", "null"):
            return 0.0
        s = s.replace("R$", "").replace("%", "").replace(" ", "")
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


def _v101_meta_diaria_validada(row):
    """V10.1: meta diária só é considerada batida se Realizado Período / Meta Período >= 100.

    Proteções:
    - Nunca usa Projetado nem Realizado Total para disparar mural/Telegram.
    - Se Realizado Período vier maior que Realizado Total, considera linha desalinhada e ignora.
    - Se percentual do SGI divergir do cálculo, prevalece o cálculo Realizado Período / Meta Período.
    """
    meta_txt = _v54_cell(row, "Meta (R$) Período", "Meta (R$) Periodo", "Meta(R$) Período", "Meta(R$) Periodo")
    real_txt = _v54_cell(row, "Realizado (R$) Período", "Realizado (R$) Periodo", "Realizado(R$) Período", "Realizado(R$) Periodo")
    total_txt = _v54_cell(row, "Realizado (R$) Total", "Realizado(R$) Total")
    meta_n = _v101_money_float(meta_txt)
    real_n = _v101_money_float(real_txt)
    total_n = _v101_money_float(total_txt)

    if meta_n <= 0 or real_n <= 0:
        return False, 0.0, meta_txt or "R$ 0,00", real_txt or "R$ 0,00"

    # Em um relatório hoje-hoje, o realizado do período nunca pode ser maior que o total acumulado do mês.
    # Quando isso acontece, normalmente a tabela foi lida desalinhada e pegou Projetado como Realizado Período.
    if total_n > 0 and real_n > (total_n + 0.01):
        return False, 0.0, meta_txt or "R$ 0,00", real_txt or "R$ 0,00"

    ating_calc = round((real_n / meta_n) * 100.0, 2)

    # V10.2: blindagem contra leitura desalinhada do Sólidus.
    # Quando a tabela muda/arrasta coluna, às vezes Realizado Período vira Projetado/Total
    # e aparecem absurdos como 1297% no mural. Meta diária só aceita valores plausíveis.
    ating_txt = _v54_cell(row, "Atingido Período", "Atingido Periodo", "Atingido Período_float", "Atingido Periodo_float")
    ating_sgi = _v54_num_pct_meta_diaria(ating_txt)
    if ating_sgi > 0 and abs(ating_sgi - ating_calc) > max(2.0, ating_calc * 0.05):
        return False, 0.0, meta_txt or "R$ 0,00", real_txt or "R$ 0,00"
    # Nenhuma loja/vendedor deve disparar alerta diário acima de 500%; isso indica coluna errada.
    if ating_calc > 500:
        return False, 0.0, meta_txt or "R$ 0,00", real_txt or "R$ 0,00"

    return ating_calc >= 100.0, ating_calc, meta_txt or "R$ 0,00", real_txt or "R$ 0,00"

def _v54_meta_diaria_precalc(dados):
    """Pré-calcula no Python quem bateu meta diária para evitar erro JS no HTML."""
    out = []
    try:
        metas = (dados or {}).get("metas", {}) or {}
        # V8.4: nunca exibe meta diária batida se o JSON não for da data atual.
        # Evita mural mostrar metas batidas de outro dia quando a coleta falha ou fica stale.
        if str((dados or {}).get("data_consulta") or "").strip() != hoje.strftime("%d/%m/%Y"):
            return []
        if str((dados or {}).get("versao_coleta") or "") not in ("V5.9", "V6.0", "V6.2", "V6.3", "V6.4", "V6.5"):
            # Proteção: não renderiza metadados antigos da meta diária, pois algumas versões antigas
            # podiam exibir valores totais/mensais no mural.
            return []
        def add_row(row, kind):
            if not isinstance(row, dict) or row.get("_is_total"):
                return
            ok_meta_dia, ating, meta, realizado = _v101_meta_diaria_validada(row)
            if not ok_meta_dia:
                return
            if kind == "filial":
                raw = str(_v54_cell(row, "Filial", "Nome") or "Filial").strip()
                filial = _filial_key_from_text(raw) or ""
                nome = raw
                if filial and filial in ORDEM_FILIAIS:
                    nome = f"Filial {filial[1:]}"
                info = f"{raw} · {ating:.2f}% no dia · {realizado} / {meta}"
            else:
                raw_fil = str(_v54_cell(row, "Vendedor", "Filial") or "").strip()
                filial = _filial_key_from_text(raw_fil) or ""
                nome = str(_v54_cell(row, "Vendedor_2", "Nome_2", "Nome", "Vendedor") or "Vendedor").strip()
                # V5.8: se for colaborador desligado/inativo, não entra no mural de meta diária.
                if not _colab_esta_liberado_py(nome, filial, False, 'participa_murais'):
                    return
                info = f"{filial or raw_fil} · {ating:.2f}% no dia · {realizado} / {meta}"
            out.append({
                "nome": nome,
                "info": info.replace(".", ",", 1) if False else info,
                "filial": filial,
                "tipo": kind,
                "kind": kind,
                "ating": round(float(ating), 2),
                "realizado_periodo": str(realizado),
                "meta_periodo": str(meta),
            })
        _mf = (metas.get("venda_filial_meta", {}) or {})
        if _mf.get("coleta_diaria_v59"):
            for row in _mf.get("linhas", []) or []:
                add_row(row, "filial")
        _mv = (metas.get("venda_filial_vendedor_meta", {}) or {})
        if _mv.get("coleta_diaria_v59"):
            for row in _mv.get("linhas", []) or []:
                add_row(row, "vendedor")
        out.sort(key=lambda x: (-float(x.get("ating") or 0), str(x.get("nome") or "")))
    except Exception as e:
        print(f"⚠️ Falha ao pré-calcular meta diária batida: {e}")
    return out

_meta_diaria_batida_precalc = _v54_meta_diaria_precalc(metas_vendas_dia_info.get('dados', {}) or {})
print(f"🎯 Meta diária BATIDA pré-calculada para o HTML: {len(_meta_diaria_batida_precalc)} item(ns)")
js_meta_diaria_batida_precalc = json.dumps(_meta_diaria_batida_precalc, ensure_ascii=False)
js_margens_brutas = json.dumps(margens_brutas_info.get('dados', {}), ensure_ascii=False)
# placeholders; valores reais serão definidos após montar _sales_emp/_sales_fil/_sales_vend
js_sales_empresa = json.dumps({}, ensure_ascii=False)
js_rent_empresa = json.dumps((margens_brutas_info.get('dados', {}) or {}).get('empresa', {}), ensure_ascii=False)

# relatório de serviços do mês atual (gerado pelo worker de vendas)
_servicos_rel_path = os.path.join(pasta, "relatorio_servicos_mes_atual.json")
_servicos_rel_data = {}
try:
    if os.path.exists(_servicos_rel_path):
        with open(_servicos_rel_path, "r", encoding="utf-8") as _fsrv:
            _servicos_rel_data = json.load(_fsrv) or {}
except Exception as _e:
    print(f"⚠️ Falha ao carregar relatorio_servicos_mes_atual.json: {_e}")
js_servicos_relatorio = json.dumps(_servicos_rel_data, ensure_ascii=False)

# Inativos e FDEP por filial
bruto_inat_p  = {f: 0.0 for f in ORDEM_FILIAIS + ["FDEP"]}
bruto_inat_pg = {f: 0.0 for f in ORDEM_FILIAIS + ["FDEP"]}
for _, r in inativos.iterrows():
    f = r["filial"]
    if f in bruto_inat_p:
        bruto_inat_p[f] += r["pendente"]; bruto_inat_pg[f] += r["pago"]

fdep_para_filial_p  = {}
fdep_para_filial_pg = {}
if tw > 0:
    for f, w in pesos.items():
        fdep_para_filial_p[f]  = round(fdep_total_rateado_p  * w / tw, 2)
        fdep_para_filial_pg[f] = round(fdep_total_rateado_pg * w / tw, 2)

def get_meta_vend(key):
    return meta_mes.get("vendedores", {}).get(key, {})

def get_meta_fil(f):
    return meta_mes.get("filiais", {}).get(f, {})

MARGENS_DADOS = margens_brutas_info.get("dados", {}) or {}
MARGENS_FILIAIS = MARGENS_DADOS.get("filiais", {}) or {}
MARGENS_VENDEDORES = MARGENS_DADOS.get("vendedores", {}) or {}

def margem_filial_pct(f):
    try:
        return round(float((MARGENS_FILIAIS.get(str(f), {}) or {}).get("margem_bruta_pct", 0) or 0), 2)
    except Exception:
        return 0.0

# Mesmo padrão usado na coleta de margens: nome normalizado + filial.
def margem_vendedor_pct(nome, filial):
    try:
        k = f"{_norm_key_margem(nome)}_{filial}"
        return round(float((MARGENS_VENDEDORES.get(k, {}) or {}).get("margem_bruta_pct", 0) or 0), 2)
    except Exception:
        return 0.0

# Vendedores por filial (somente NÃO gerentes para o painel individual)
todos_js = {}
for filial in ORDEM_FILIAIS:
    sub = df_vend[df_vend["filial_vendedor"] == filial]
    if sub.empty: continue
    # Somente vendedores (não gerentes) no painel individual
    sub_vend = sub[~sub["is_gerente"]]
    if sub_vend.empty:
        todos_js[filial] = []
        continue

    total_inat_p = bruto_inat_p.get(filial, 0.0)
    total_inat_pg= bruto_inat_pg.get(filial, 0.0)
    fdep_p       = fdep_para_filial_p.get(filial, 0.0)
    fdep_pg      = fdep_para_filial_pg.get(filial, 0.0)
    total_filial_pend = sub["pendente"].sum()

    todos_js[filial] = []
    for _, r in sub_vend.sort_values("pendente", ascending=False).iterrows():
        pf = float(r["pendente"])
        pg = float(r["pago"])
        prop = pf / total_filial_pend if total_filial_pend > 0 else 0
        inat_p = round(total_inat_p * prop, 2)
        fdp    = round(fdep_p * prop, 2)
        prop_p = round(pf - inat_p - fdp, 2)
        key    = f"{r['nome_exibicao']}_{filial}"
        mv     = get_meta_vend(key)
        tv     = media_trimestral_vend.get(key, {})
        vv     = var_vendedores.get(key, {})

        todos_js[filial].append({
            "nome": str(r["nome_exibicao"]), "filial": filial,
            "is_gerente": False,
            "pendente": round(pf,2), "pago": round(pg,2),
            "total": round(float(r["total"]),2),
            "perc_filial": round(float(r["perc_filial"]),2),
            "proprio_p": max(prop_p,0), "inat_p": inat_p, "fdep_p": fdp,
            "perc_inat": round(inat_p/pf*100,1) if pf>0 else 0,
            "perc_fdep": round(fdp/pf*100,1)    if pf>0 else 0,
            # variação
            "var_pago_perc":  vv.get("perc_pago"),
            "var_pago_delta": vv.get("delta_pago", 0),
            "var_periodo":    vv.get("periodo",""),
            # metas por faixa
            "grave_perc":    mv.get("grave_perc", 0),
            "alerta_perc":   mv.get("alerta_perc", 0),
            "atencao_perc":  mv.get("atencao_perc", 0),
            "perc_meta":     mv.get("perc_meta", 0),
            "grave_rec":     mv.get("grave_rec", 0),
            "alerta_rec":    mv.get("alerta_rec", 0),
            "atencao_rec":   mv.get("atencao_rec", 0),
            "grave_alvo":    mv.get("grave_alvo", 0),
            "alerta_alvo":   mv.get("alerta_alvo", 0),
            "atencao_alvo":  mv.get("atencao_alvo", 0),
            "grave_pend":    mv.get("grave_pend", 0),
            "alerta_pend":   mv.get("alerta_pend", 0),
            "atencao_pend":  mv.get("atencao_pend", 0),
            # trimestral
            "trim_media_perc": tv.get("media_perc", 0),
            "trim_meses":      tv.get("meses_count", 0),
            # rentabilidade / margem bruta SGI
            "rentabilidade_pct": margem_vendedor_pct(r["nome_exibicao"], filial),
        })

# Filiais (inclui totais de gerentes + vendedores)
filiais_js = {}

for f in ORDEM_FILIAIS:
    sub = df_vend[df_vend["filial_vendedor"] == f]

    inat_p = round(bruto_inat_p.get(f,0.0), 2)
    fdep_p = round(fdep_para_filial_p.get(f,0.0), 2)

    vf = var_filiais.get(f, {})
    mf = get_meta_fil(f) or {}
    tf = media_trimestral_fil.get(f, 0)

    # =========================================
    # 🔹 FILIAL COM VENDEDOR
    # =========================================
    if not sub.empty:
        pt = round(sub["pendente"].sum(), 2)
        pg = round(sub["pago"].sum(), 2)

        filiais_js[f] = {
            "pendente": pt,
            "pago": pg,
            "total": round(sub["total"].sum(), 2),
            "sem_ativo": False,

            "inat_p": inat_p,
            "fdep_p": fdep_p,

            "perc_inat": round(inat_p/pt*100,1) if pt>0 else 0,
            "perc_fdep": round(fdep_p/pt*100,1) if pt>0 else 0,

            "var_pago_perc":  vf.get("perc_pago"),
            "var_pago_delta": vf.get("delta_pago",0),
            "var_periodo":    vf.get("periodo",""),

            # 🔥 META
            "grave_perc":    mf.get("grave_perc",0),
            "alerta_perc":   mf.get("alerta_perc",0),
            "atencao_perc":  mf.get("atencao_perc",0),
            "perc_meta":     mf.get("perc_meta",0),

            "grave_rec":     mf.get("grave_rec",0),
            "alerta_rec":    mf.get("alerta_rec",0),
            "atencao_rec":   mf.get("atencao_rec",0),

            "grave_alvo":    mf.get("grave_alvo",0),
            "alerta_alvo":   mf.get("alerta_alvo",0),
            "atencao_alvo":  mf.get("atencao_alvo",0),
            "grave_pend":    mf.get("grave_pend",0),
            "alerta_pend":   mf.get("alerta_pend",0),
            "atencao_pend":  mf.get("atencao_pend",0),

            "trim_media_perc": tf,
            "rentabilidade_pct": margem_filial_pct(f),
        }

    # =========================================
    # 🔹 FILIAL SEM VENDEDOR (🔥 CORRETO AGORA)
    # =========================================
    elif f in filiais_sem_ativo:
        pt = round(filiais_sem_ativo[f]["pendente"], 2)
        pg = round(filiais_sem_ativo[f]["pago"], 2)

        filiais_js[f] = {
            "pendente": pt,
            "pago": pg,
            "total": round(pt + pg, 2),
            "sem_ativo": True,

            "inat_p": inat_p,
            "fdep_p": fdep_p,

            "perc_inat": 100.0,
            "perc_fdep": round(fdep_p/pt*100,1) if pt>0 else 0,

            "var_pago_perc": None,
            "var_pago_delta": 0,
            "var_periodo": "",

            # 🔥 AQUI ESTAVA O ERRO → NÃO PODE ZERAR
            "grave_perc":    mf.get("grave_perc",0),
            "alerta_perc":   mf.get("alerta_perc",0),
            "atencao_perc":  mf.get("atencao_perc",0),
            "perc_meta":     mf.get("perc_meta",0),

            "grave_rec":     mf.get("grave_rec",0),
            "alerta_rec":    mf.get("alerta_rec",0),
            "atencao_rec":   mf.get("atencao_rec",0),

            "grave_alvo":    mf.get("grave_alvo",0),
            "alerta_alvo":   mf.get("alerta_alvo",0),
            "atencao_alvo":  mf.get("atencao_alvo",0),
            "grave_pend":    mf.get("grave_pend",0),
            "alerta_pend":   mf.get("alerta_pend",0),
            "atencao_pend":  mf.get("atencao_pend",0),

            "trim_media_perc": tf,
            "rentabilidade_pct": margem_filial_pct(f),
        }

    # =========================================
    # 🔹 SEGURANÇA (caso raro)
    # =========================================
    else:
        filiais_js[f] = {
            "pendente": 0,
            "pago": 0,
            "total": 0,
            "sem_ativo": True,

            "inat_p": 0,
            "fdep_p": 0,

            "perc_inat": 0,
            "perc_fdep": 0,

            "var_pago_perc": None,
            "var_pago_delta": 0,
            "var_periodo": "",

            "grave_perc": 0,
            "alerta_perc": 0,
            "atencao_perc": 0,
            "perc_meta": 0,

            "grave_rec": 0,
            "alerta_rec": 0,
            "atencao_rec": 0,

            "grave_alvo": 0,
            "alerta_alvo": 0,
            "atencao_alvo": 0,
            "grave_pend": 0,
            "alerta_pend": 0,
            "atencao_pend": 0,

            "trim_media_perc": 0,
            "rentabilidade_pct": 0,
        }

filiais_js_ordered = {f: filiais_js[f] for f in ORDEM_FILIAIS if f in filiais_js}


# ── V10.7: histórico mensal explícito de recebimentos por faixa ─────────────
# Serve para manter fechado o mês anterior no FTP quando virar o mês, enquanto o mês atual
# passa a contar somente pagamentos da nova competência.
RECEBIMENTOS_MENSAL_PATH = os.path.join(CACHE_DIR, 'historico_recebimentos_mensais.json')
def _sum_meta_mes_obj(_meta_obj):
    _out = {'grave_rec':0.0,'alerta_rec':0.0,'atencao_rec':0.0,'grave_alvo':0.0,'alerta_alvo':0.0,'atencao_alvo':0.0,'grave_pend':0.0,'alerta_pend':0.0,'atencao_pend':0.0}
    for _vals in ((_meta_obj or {}).get('filiais') or {}).values():
        if not isinstance(_vals, dict):
            continue
        for _k in list(_out.keys()):
            _out[_k] += float(_vals.get(_k, 0) or 0)
    return {k: round(v, 2) for k, v in _out.items()}

def _salvar_historico_recebimentos_mensais_v107():
    try:
        _hist_rec = _load_json_file(RECEBIMENTOS_MENSAL_PATH, {'months': {}})
        _hist_rec.setdefault('months', {})
        # Fecha mês anterior se houver meta_<mes>.json e ainda não estiver salvo.
        try:
            _first = datetime.strptime(mes_str + '-01', '%Y-%m-%d')
            _prev_mes = (_first - timedelta(days=1)).strftime('%Y-%m')
        except Exception:
            _prev_mes = ''
        if _prev_mes and _prev_mes not in _hist_rec['months']:
            _prev_meta_file = meta_path(_prev_mes)
            if os.path.exists(_prev_meta_file):
                with open(_prev_meta_file, encoding='utf-8') as _fpm:
                    _prev_meta = json.load(_fpm) or {}
                _hist_rec['months'][_prev_mes] = {
                    'mes': _prev_mes,
                    'fechado_em': now_brasilia().isoformat(),
                    'origem': os.path.basename(_prev_meta_file),
                    'empresa': _sum_meta_mes_obj(_prev_meta),
                    'filiais': _prev_meta.get('filiais', {}),
                    'vendedores': _prev_meta.get('vendedores', {}),
                }
                print(f'📦 V10.7 histórico de recebimentos fechado: {_prev_mes}')
        # Salva/atualiza mês atual com o estado corrente.
        _hist_rec['months'][mes_str] = {
            'mes': mes_str,
            'atualizado_em': now_brasilia().isoformat(),
            'empresa': {
                'grave_rec': round(sum(float(v.get('grave_rec',0) or 0) for v in filiais_js_ordered.values()), 2),
                'alerta_rec': round(sum(float(v.get('alerta_rec',0) or 0) for v in filiais_js_ordered.values()), 2),
                'atencao_rec': round(sum(float(v.get('atencao_rec',0) or 0) for v in filiais_js_ordered.values()), 2),
                'grave_alvo': round(sum(float(v.get('grave_alvo',0) or 0) for v in filiais_js_ordered.values()), 2),
                'alerta_alvo': round(sum(float(v.get('alerta_alvo',0) or 0) for v in filiais_js_ordered.values()), 2),
                'atencao_alvo': round(sum(float(v.get('atencao_alvo',0) or 0) for v in filiais_js_ordered.values()), 2),
            },
            'filiais': filiais_js_ordered,
            'vendedores': {f'{r.get("nome")}_{r.get("filial")}': r for arr in todos_js.values() for r in arr},
        }
        with open(RECEBIMENTOS_MENSAL_PATH, 'w', encoding='utf-8') as _fhrm:
            json.dump(_hist_rec, _fhrm, ensure_ascii=False, indent=2)
        print(f'🧾 V10.7 histórico mensal de recebimentos salvo: {RECEBIMENTOS_MENSAL_PATH}')
        return _hist_rec
    except Exception as _e_hrm:
        print(f'⚠️ V10.7 erro salvando histórico mensal de recebimentos: {_e_hrm}')
        return {'months': {}}

HIST_RECEBIMENTOS_MENSAIS = _salvar_historico_recebimentos_mensais_v107()

# ── Histórico diário persistente (Master / Diretor Comercial) ───────────────
def _load_hist_dash():
    if os.path.exists(_hist_dash_path):
        try:
            with open(_hist_dash_path, encoding='utf-8') as _fh:
                _d = json.load(_fh)
            if isinstance(_d, dict):
                return _d
        except Exception:
            pass
    return {"dates": {}}

hist_dash = _load_hist_dash()
hist_dash.setdefault("dates", {})

_hist_empresa = {
    "pendente": round(sum(v.get("pendente", 0) for v in filiais_js_ordered.values()), 2),
    "recebido": round(sum(v.get("pago", 0) for v in filiais_js_ordered.values()), 2),
    "grave": round(sum(v.get("grave_pend", 0) for v in filiais_js_ordered.values()), 2),
    "alerta": round(sum(v.get("alerta_pend", 0) for v in filiais_js_ordered.values()), 2),
    "atencao": round(sum(v.get("atencao_pend", 0) for v in filiais_js_ordered.values()), 2),
    "config_global": CONFIG_META,
}

_hist_vends = {}
for _f_hist, _arr_hist in todos_js.items():
    for _v_hist in _arr_hist:
        _k_hist = f"{_v_hist.get('nome','')}_{_v_hist.get('filial','')}"
        _hist_vends[_k_hist] = {
            "nome": _v_hist.get("nome",""),
            "filial": _v_hist.get("filial",""),
            "pendente": round(float(_v_hist.get("pendente",0) or 0), 2),
            "recebido": round(float(_v_hist.get("pago",0) or 0), 2),
            "total": round(float(_v_hist.get("total",0) or 0), 2),
            "grave_pend": round(float(_v_hist.get("grave_pend",0) or 0), 2),
            "alerta_pend": round(float(_v_hist.get("alerta_pend",0) or 0), 2),
            "atencao_pend": round(float(_v_hist.get("atencao_pend",0) or 0), 2),
            "grave_rec": round(float(_v_hist.get("grave_rec",0) or 0), 2),
            "alerta_rec": round(float(_v_hist.get("alerta_rec",0) or 0), 2),
            "atencao_rec": round(float(_v_hist.get("atencao_rec",0) or 0), 2),
            "grave_alvo": round(float(_v_hist.get("grave_alvo",0) or 0), 2),
            "alerta_alvo": round(float(_v_hist.get("alerta_alvo",0) or 0), 2),
            "atencao_alvo": round(float(_v_hist.get("atencao_alvo",0) or 0), 2),
            "perc_meta": round(float(_v_hist.get("perc_meta",0) or 0), 1),
            "config_usada": get_config_meta(_k_hist),
        }

_hist_fils = {}
for _f_hist, _fd_hist in filiais_js_ordered.items():
    _hist_fils[_f_hist] = {
        "nome": f"Filial {_f_hist}",
        "filial": _f_hist,
        "pendente": round(float(_fd_hist.get("pendente",0) or 0), 2),
        "recebido": round(float(_fd_hist.get("pago",0) or 0), 2),
        "total": round(float(_fd_hist.get("total",0) or 0), 2),
        "grave_pend": round(float(_fd_hist.get("grave_pend",0) or 0), 2),
        "alerta_pend": round(float(_fd_hist.get("alerta_pend",0) or 0), 2),
        "atencao_pend": round(float(_fd_hist.get("atencao_pend",0) or 0), 2),
        "grave_rec": round(float(_fd_hist.get("grave_rec",0) or 0), 2),
        "alerta_rec": round(float(_fd_hist.get("alerta_rec",0) or 0), 2),
        "atencao_rec": round(float(_fd_hist.get("atencao_rec",0) or 0), 2),
        "grave_alvo": round(float(_fd_hist.get("grave_alvo",0) or 0), 2),
        "alerta_alvo": round(float(_fd_hist.get("alerta_alvo",0) or 0), 2),
        "atencao_alvo": round(float(_fd_hist.get("atencao_alvo",0) or 0), 2),
        "perc_meta": round(float(_fd_hist.get("perc_meta",0) or 0), 1),
        "config_usada": get_config_meta(f"FILIAL::{_f_hist}"),
    }


def _safe_pct_num(v):
    try:
        return round(float(v), 2)
    except Exception:
        return 0.0

def _norm_txt(v):
    s = str(v or "").strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s)

def _filial_key_from_text(txt):
    s = _norm_txt(txt)
    m = re.search(r"FILIAL\s+(\d{1,2})(?:/\d{1,2})?", s)
    if m:
        return f"F{int(m.group(1))}"
    return ""

def _ensure_sales_bucket(dct, key, nome="", filial=""):
    if key not in dct:
        dct[key] = {
            "nome": nome or key,
            "filial": filial or "",
            "venda_realizado_total": 0.0,
            "venda_atingido_total": 0.0,
            "venda_projetado": 0.0,
            "servico_realizado_total": 0.0,
            "servico_atingido_total": 0.0,
            "servico_projetado": 0.0,
            "caminhao_realizado_total": 0.0,
            "caminhao_atingido_total": 0.0,
            "caminhao_projetado": 0.0,
        }
    return dct[key]

# ── Carrega totais empresa direto do xlsx (fonte mais confiável) ─────────────
_xlsx_metas_path = os.path.join(pasta, "metas_vendas_mes_atual.xlsx")
_totais_direto = _extrair_totais_do_xlsx(_xlsx_metas_path) if os.path.exists(_xlsx_metas_path) else {}

_t_venda    = _totais_direto.get("venda_filial_meta", {})
_t_servico  = _totais_direto.get("servico_filial_ouro_fob", {})
_t_caminhao = _totais_direto.get("venda_filial_subgrupo_20k", {})

_sales_emp = {
    "venda_realizado_total":  _t_venda.get("real_total", 0.0),
    "venda_atingido_total":   _t_venda.get("ating_total", 0.0),
    "venda_projetado":        _t_venda.get("proj", 0.0),
    "venda_meta_total":       _t_venda.get("meta_total", 0.0),
    "venda_meta_periodo":     _t_venda.get("meta_per", 0.0),
    "servico_realizado_total":  _t_servico.get("real_total", 0.0),
    "servico_atingido_total":   _t_servico.get("ating_total", 0.0),
    "servico_projetado":        _t_servico.get("proj", 0.0),
    "servico_meta_total":       _t_servico.get("meta_total", 0.0),
    "servico_meta_periodo":     _t_servico.get("meta_per", 0.0),
    "caminhao_realizado_total": _t_caminhao.get("real_total", 0.0),
    "caminhao_atingido_total":  _t_caminhao.get("ating_total", 0.0),
    "caminhao_projetado":       _t_caminhao.get("proj", 0.0),
    "caminhao_meta_total":      _t_caminhao.get("meta_total", 0.0),
    "caminhao_meta_periodo":    _t_caminhao.get("meta_per", 0.0),
}
print(f"📊 _sales_emp carregado do xlsx:")
print(f"   Mercantil  : R$ {_sales_emp['venda_realizado_total']:>12,.2f}  proj R$ {_sales_emp['venda_projetado']:>12,.2f}  ating {_sales_emp['venda_atingido_total']:.2f}%")
print(f"   Serviços   : R$ {_sales_emp['servico_realizado_total']:>12,.2f}  proj R$ {_sales_emp['servico_projetado']:>12,.2f}  ating {_sales_emp['servico_atingido_total']:.2f}%")
print(f"   Caminhão   : R$ {_sales_emp['caminhao_realizado_total']:>12,.2f}  proj R$ {_sales_emp['caminhao_projetado']:>12,.2f}  ating {_sales_emp['caminhao_atingido_total']:.2f}%")
_sales_fil = {}
_sales_vend = {}

for _meta_key, _meta_obj in (metas_vendas_info.get("dados", {}).get("metas", {}) or {}).items():
    if not isinstance(_meta_obj, dict) or not _meta_obj.get("ok"):
        continue
    _rows = _meta_obj.get("linhas", []) or []
    # Separa linhas normais das linhas de total (tfoot)
    _rows_normal = [_r for _r in _rows if not _r.get("_is_total")]
    _rows_total  = [_r for _r in _rows if _r.get("_is_total")]
    # Usa a linha de total se disponível (é a mais precisa — vem direto do tfoot do SGI)
    _total_row   = _rows_total[-1] if _rows_total else None

    def _val_real(r):
        return float(r.get("Realizado (R$) Total_float") or r.get("Realizado(R$) Total_float") or 0)
    def _val_meta(r):
        return float(r.get("Meta (R$) Total_float") or r.get("Meta(R$) Total_float") or 0)
    def _val_meta_per(r):
        return float(r.get("Meta (R$) Período_float") or r.get("Meta(R$) Período_float") or 0)
    def _val_ating(r):
        return float(r.get("Atingido Total_float") or 0)
    def _val_proj(r):
        return float(r.get("Projetado (R$)_float") or r.get("Projetado(R$)_float") or 0)

    for _r in _rows_normal:
        _real  = _val_real(_r)
        _ating = _val_ating(_r)
        _proj  = _val_proj(_r)
        if _meta_key == "venda_filial_meta":
            _fil = _filial_key_from_text(_r.get("Filial", ""))
            if _fil:
                _b = _ensure_sales_bucket(_sales_fil, _fil, nome=f"Filial {_fil}", filial=_fil)
                _b["venda_realizado_total"] = _real
                _b["venda_atingido_total"]  = _ating
                _b["venda_projetado"]       = _proj
                _b["venda_meta_total"]      = _val_meta(_r)
                _b["venda_meta_periodo"]    = _val_meta_per(_r)
        elif _meta_key == "servico_filial_ouro_fob":
            _fil = _filial_key_from_text(_r.get("Filial", ""))
            if _fil:
                _b = _ensure_sales_bucket(_sales_fil, _fil, nome=f"Filial {_fil}", filial=_fil)
                _b["servico_realizado_total"] = _real
                _b["servico_atingido_total"]  = _ating
                _b["servico_projetado"]       = _proj
                _b["servico_meta_total"]      = _val_meta(_r)
                _b["servico_meta_periodo"]    = _val_meta_per(_r)
        elif _meta_key == "venda_filial_subgrupo_20k":
            _fil = _filial_key_from_text(_r.get("Filial", ""))
            if _fil:
                _b = _ensure_sales_bucket(_sales_fil, _fil, nome=f"Filial {_fil}", filial=_fil)
                _b["caminhao_realizado_total"] = _real
                _b["caminhao_atingido_total"]  = _ating
                _b["caminhao_projetado"]       = _proj
                _b["caminhao_meta_total"]      = _val_meta(_r)
                _b["caminhao_meta_periodo"]    = _val_meta_per(_r)
        elif _meta_key == "venda_filial_vendedor_meta":
            _nome = str(_r.get("Vendedor_2") or _r.get("Vendedor") or "").strip()
            _fil  = _filial_key_from_text(_r.get("Vendedor", "") or _r.get("Filial", ""))
            if _nome:
                _k = f"{_nome}_{_fil}" if _fil else _nome
                _b = _ensure_sales_bucket(_sales_vend, _k, nome=_nome, filial=_fil)
                _b["venda_realizado_total"] = _real
                _b["venda_atingido_total"]  = _ating
                _b["venda_projetado"]       = _proj
        elif _meta_key == "servico_filial_vendedor_ouro_fob":
            _nome = str(_r.get("Vendedor_2") or _r.get("Vendedor") or "").strip()
            _fil  = _filial_key_from_text(_r.get("Vendedor", "") or _r.get("Filial", ""))
            if _nome:
                _k = f"{_nome}_{_fil}" if _fil else _nome
                _b = _ensure_sales_bucket(_sales_vend, _k, nome=_nome, filial=_fil)
                _b["servico_realizado_total"] = _real
                _b["servico_atingido_total"]  = _ating
                _b["servico_projetado"]       = _proj

    # _sales_emp já foi carregado do xlsx diretamente (ver inicialização acima).
    # Este loop só precisa popular _sales_fil e _sales_vend (por filial/vendedor).
    pass
for _bucket in list(_sales_fil.values()) + list(_sales_vend.values()):
    for _k in list(_bucket.keys()):
        if isinstance(_bucket[_k], float):
            _bucket[_k] = round(_bucket[_k], 2)

# agora que _sales_emp foi montado, serializa corretamente para o HTML
js_sales_empresa = json.dumps(_sales_emp, ensure_ascii=False)
js_rent_empresa = json.dumps((margens_brutas_info.get('dados', {}) or {}).get('empresa', {}), ensure_ascii=False)

# relatório de serviços do mês atual (gerado pelo worker de vendas)
_servicos_rel_path = os.path.join(pasta, "relatorio_servicos_mes_atual.json")
_servicos_rel_data = {}
try:
    if os.path.exists(_servicos_rel_path):
        with open(_servicos_rel_path, "r", encoding="utf-8") as _fsrv:
            _servicos_rel_data = json.load(_fsrv) or {}
except Exception as _e:
    print(f"⚠️ Falha ao carregar relatorio_servicos_mes_atual.json: {_e}")
js_servicos_relatorio = json.dumps(_servicos_rel_data, ensure_ascii=False)

hist_dash.setdefault("sales_dates", {})
hist_dash.setdefault("sales_months", {})

# =========================================
# 🎯 HISTÓRICO REAL DA META DIÁRIA
# Regra corrigida:
#   - Meta diária = Meta Total / dias úteis configurados
#   - Ponto diário só é gerado se a VENDA DO DIA bater a meta diária.
#   - Não soma pelo acumulado dividido; salva dia a dia para valer no próximo mês.
# =========================================

def _last_sales_date_before(_today):
    try:
        _dates = sorted([d for d in (hist_dash.get("sales_dates", {}) or {}).keys() if str(d) < str(_today)])
        return _dates[-1] if _dates else None
    except Exception:
        return None


def _sales_lookup_prev(_prev_date, scope, key):
    if not _prev_date:
        return 0.0
    try:
        return float((hist_dash.get("sales_dates", {}).get(_prev_date, {}).get(scope, {}).get(key, {}) or {}).get("venda_realizado_total", 0) or 0)
    except Exception:
        return 0.0


def _dias_uteis_campanha_padrao(ent_type, ent_key):
    try:
        _cfg = get_config_meta(ent_key)
        if ent_type == "filial":
            _rows = (_cfg.get("camp_meta_diaria_ger") or [])
        else:
            _rows = (_cfg.get("camp_meta_diaria_vend") or [])
        if _rows:
            return max(1, int(float(str(_rows[0].get("dias_uteis", 25)).replace(",", "."))))
    except Exception:
        pass
    return 26


def _meta_total_from_rows(meta_key, lookup_filial=None, lookup_nome=None):
    try:
        _rows = metas_vendas_info.get("dados", {}).get("metas", {}).get(meta_key, {}).get("linhas", []) or []
        for _r in _rows:
            if lookup_filial:
                if _filial_key_from_text(_r.get("Filial", "") or _r.get("Vendedor", "")) != lookup_filial:
                    continue
            if lookup_nome:
                _nome_row = str(_r.get("Vendedor_2") or _r.get("Vendedor") or "").strip()
                if _norm_key_margem(_nome_row) != _norm_key_margem(lookup_nome):
                    continue
            return float(_r.get("Meta (R$) Total_float") or _r.get("Meta(R$) Total_float") or 0)
    except Exception:
        pass
    return 0.0


def _atualizar_meta_diaria_historico():
    _prev_date = _last_sales_date_before(hoje_str)
    _hist_md = hist_dash.setdefault("daily_meta", {})
    _hist_md.setdefault("dates", {})
    _hist_md.setdefault("months", {})
    _today_entry = {"filiais": {}, "vendedores": {}, "prev_date": _prev_date, "updated_at": now_brasilia().isoformat()}
    _month_entry = _hist_md["months"].setdefault(mes_str, {"filiais": {}, "vendedores": {}})

    # Filiais
    for _f, _b in (_sales_fil or {}).items():
        _meta_total = _meta_total_from_rows("venda_filial_meta", lookup_filial=_f)
        _dias = _dias_uteis_campanha_padrao("filial", f"FILIAL::{_f}")
        _meta_dia = (_meta_total / _dias) if _meta_total > 0 else 0.0
        _real_hoje_acum = float(_b.get("venda_realizado_total", 0) or 0)
        _real_ant = _sales_lookup_prev(_prev_date, "filiais", _f)
        _real_dia = max(0.0, _real_hoje_acum - _real_ant)
        _bateu = bool(_meta_dia > 0 and _real_dia >= _meta_dia)
        _today_entry["filiais"][_f] = {
            "filial": _f,
            "meta_total": round(_meta_total, 2),
            "dias_uteis": _dias,
            "meta_diaria": round(_meta_dia, 2),
            "realizado_dia": round(_real_dia, 2),
            "realizado_total": round(_real_hoje_acum, 2),
            "ponto": 1 if _bateu else 0,
        }
        _me = _month_entry["filiais"].setdefault(_f, {"pontos": 0, "dias": {}})
        # Recalcula o ponto deste dia sem duplicar se o script rodar mais de uma vez.
        _old = int((_me.get("dias", {}).get(hoje_str, {}) or {}).get("ponto", 0) or 0)
        _new = 1 if _bateu else 0
        _me["pontos"] = max(0, int(_me.get("pontos", 0) or 0) - _old + _new)
        _me.setdefault("dias", {})[hoje_str] = _today_entry["filiais"][_f]

    # Vendedores
    for _k, _b in (_sales_vend or {}).items():
        _nome = _b.get("nome", "")
        _fil = _b.get("filial", "")
        _meta_total = _meta_total_from_rows("venda_filial_vendedor_meta", lookup_filial=_fil, lookup_nome=_nome)
        _dias = _dias_uteis_campanha_padrao("vendedor", f"VEND::{_limpar_nome_margem(_nome)}_{_fil}")
        _meta_dia = (_meta_total / _dias) if _meta_total > 0 else 0.0
        _real_hoje_acum = float(_b.get("venda_realizado_total", 0) or 0)
        _real_ant = _sales_lookup_prev(_prev_date, "vendedores", _k)
        _real_dia = max(0.0, _real_hoje_acum - _real_ant)
        _bateu = bool(_meta_dia > 0 and _real_dia >= _meta_dia)
        _today_entry["vendedores"][_k] = {
            "nome": _nome,
            "filial": _fil,
            "key": _k,
            "meta_total": round(_meta_total, 2),
            "dias_uteis": _dias,
            "meta_diaria": round(_meta_dia, 2),
            "realizado_dia": round(_real_dia, 2),
            "realizado_total": round(_real_hoje_acum, 2),
            "ponto": 1 if _bateu else 0,
        }
        _me = _month_entry["vendedores"].setdefault(_k, {"pontos": 0, "dias": {}})
        _old = int((_me.get("dias", {}).get(hoje_str, {}) or {}).get("ponto", 0) or 0)
        _new = 1 if _bateu else 0
        _me["pontos"] = max(0, int(_me.get("pontos", 0) or 0) - _old + _new)
        _me.setdefault("dias", {})[hoje_str] = _today_entry["vendedores"][_k]

    _hist_md["dates"][hoje_str] = _today_entry
    print(f"🎯 Meta diária histórica atualizada: {hoje_str} (base anterior: {_prev_date or 'sem histórico'})")

_atualizar_meta_diaria_historico()

hist_dash["sales_dates"][hoje_str] = {
    "empresa": _sales_emp,
    "filiais": _sales_fil,
    "vendedores": _sales_vend,
    "updated_at": now_brasilia().isoformat(),
}
hist_dash["sales_months"][mes_str] = {
    "empresa": _sales_emp,
    "filiais": _sales_fil,
    "vendedores": _sales_vend,
    "updated_at": now_brasilia().isoformat(),
}

hist_dash["dates"][hoje_str] = {
    "empresa": _hist_empresa,
    "vendedores": _hist_vends,
    "filiais": _hist_fils,
    "updated_at": now_brasilia().isoformat(),
}
_fech_mensal = fechar_mes_anterior_travado(mes_str, hist_dash, CONFIG_META, CONFIG_META_IND)
hist_dash['months_closed'] = _fech_mensal.get('months', {})
with open(_hist_dash_path, "w", encoding="utf-8") as _f_hist_out:
    json.dump(hist_dash, _f_hist_out, ensure_ascii=False, indent=2)
print(f"🗂️ Histórico diário salvo: {_hist_dash_path}")

# Serializa relatório de clientes a cobrar
# Regras:
#   - Ativos: vão direto para o índice do próprio vendedor
#   - Inativos e FDEP: 60% para o gerente/filial, 40% dividido entre vendedores ativos
#   - Cada CLIENTE vai para apenas 1 destinatário (sem duplicatas)
#   - Determinismo: mesmo cliente sempre vai para mesmo destinatário (hash do nome)
import random as _random_mod

clientes_js = {}       # {filial: {faixa: [...]}} — para visualização da filial/gerente
clientes_por_vend_js = {}  # {nome_vend: {faixa: [...]}} — para painel individual
clientes_terceiro_js = {"grave": [], "alerta": [], "atencao": []}

for _ct in _clientes_terceiro_lista:
    clientes_terceiro_js[_ct["faixa"]].append({
        "nome": _ct.get("cliente", _ct.get("nome", "")),
        "cliente": _ct.get("cliente", ""),
        "restricao": _ct.get("restricao", ""),
        "vencimento": _ct.get("vencimento", ""),
        "pagamento": _ct.get("pagamento", ""),
        "dias": _ct.get("dias", 0),
        "pendente": _ct.get("pendente", 0),
        "pago": _ct.get("pago", 0.0),
        "parcela": _ct.get("parcela", ""),
        "titulo": _ct.get("titulo", ""),
        "avalista": _ct.get("avalista", ""),
        "contato": _ct.get("contato", ""),
        "telefones": _ct.get("telefones", []),
        "mensagem_whatsapp": _ct.get("mensagem_whatsapp", ""),
        "vendedor": COBRANCA10_NOME,
        "novo": _ct.get("novo", False),
        "cliente_key": _ct.get("cliente_key", cliente_grupo_key_py(_ct)),
        "cobranca_key": _ct.get("cobranca_key", cobranca_row_key_py(_ct)),
    })

for f in ORDEM_FILIAIS:
    clientes_js[f] = {"grave": [], "alerta": [], "atencao": []}

# Separa: ativos vs inativos/FDEP por filial
_cli_ativos    = [c for c in clientes_cobrar if c.get('is_ativo')]
_cli_inat_fdep = [c for c in clientes_cobrar if not c.get('is_ativo')]

# ── Clientes ATIVOS: vão para o próprio vendedor ──────────────────────
for c in _cli_ativos:
    _nd = limpar_nome_display(c['vendedor'])

    if _nd not in clientes_por_vend_js:
        clientes_por_vend_js[_nd] = {"grave": [], "alerta": [], "atencao": []}

    clientes_por_vend_js[_nd][c['faixa']].append({
        "nome": c["cliente"],
        "cliente": c["cliente"],
        "restricao": c["restricao"],
        "vencimento": c["vencimento"],
        "pagamento": c.get("pagamento", ""),
        "dias": c["dias"],
        "pendente": c["pendente"],
        "pago": c.get("pago", 0.0),
        "parcela": c["parcela"],
        "titulo": c.get("titulo", ""),
        "avalista": c.get("avalista", ""),
        "contato": c.get("contato", ""),
        "telefones": c.get("telefones", []),
        "mensagem_whatsapp": c.get("mensagem_whatsapp", ""),
        "vendedor": _nd[:25],
        "novo": c.get("novo", False),
    })

    _ff = c["filial"]
    if _ff in clientes_js:
        clientes_js[_ff][c['faixa']].append({
            "nome": c["cliente"],
            "cliente": c["cliente"],
            "restricao": c["restricao"],
            "vencimento": c["vencimento"],
            "pagamento": c.get("pagamento", ""),
            "dias": c["dias"],
            "pendente": c["pendente"],
            "pago": c.get("pago", 0.0),
            "parcela": c["parcela"],
            "titulo": c.get("titulo", ""),
            "avalista": c.get("avalista", ""),
            "contato": c.get("contato", ""),
            "telefones": c.get("telefones", []),
            "mensagem_whatsapp": c.get("mensagem_whatsapp", ""),
            "vendedor": _nd[:25],
            "novo": c.get("novo", False),
            "cliente_key": c.get("cliente_key", cliente_grupo_key_py(c)),
            "cobranca_key": c.get("cobranca_key", cobranca_row_key_py(c)),
        })

# ── Inativos e FDEP: distribui 60% gerente/40% vendedores por filial ──
# Agrupa por filial de destino
_inat_por_filial = {}   # {filial: [clientes]}
for c in _cli_inat_fdep:
    # Determina filial de destino
    if c.get('is_fdep'):
        # FDEP: distribui por peso proporcional (mesmo rateio do pendente)
        _dest_filial = None
        if tw > 0:
            _hash_n = int(_hashlib.md5(c['cliente'][:30].encode()).hexdigest(), 16)
            _filiais_ord2 = sorted(pesos.items(), key=lambda x: x[1], reverse=True)
            _acum2 = 0
            _lim2  = (_hash_n % 10000) / 10000.0
            for _ff2, _ww2 in _filiais_ord2:
                _acum2 += _ww2 / tw
                if _lim2 <= _acum2:
                    _dest_filial = _ff2; break
            if _dest_filial is None and _filiais_ord2:
                _dest_filial = _filiais_ord2[0][0]
    else:
        # Inativo: vai para a filial onde estava (filial no nome do cliente cobrar)
        _dest_filial = c['filial'] if c['filial'] in ORDEM_FILIAIS else None

    if _dest_filial:
        if _dest_filial not in _inat_por_filial:
            _inat_por_filial[_dest_filial] = []
        _inat_por_filial[_dest_filial].append(c)

# Para cada filial, distribui 60% gerente / 40% vendedores.
# Importante: NÃO remove títulos do mesmo cliente. Agrupa por cliente/CPF apenas para
# definir um único destino operacional; todos os títulos desse cliente seguem juntos.
for _filial_dest, _clientes_inat in _inat_por_filial.items():
    _clientes_grupos = {}
    for c in _clientes_inat:
        k = cliente_grupo_key_py(c)
        _clientes_grupos.setdefault(k, []).append(c)

    _grupos = list(_clientes_grupos.values())
    _grupos.sort(key=lambda g: sum(float(x.get('pendente', 0) or 0) for x in g), reverse=True)

    _total = len(_grupos)
    _n_ger = round(_total * 0.60)   # 60% dos grupos para gerente
    _n_vend = _total - _n_ger       # 40% dos grupos para vendedores

    _grupos_gerente = _grupos[:_n_ger]
    _grupos_vends   = _grupos[_n_ger:]

    _para_gerente = [c for g in _grupos_gerente for c in g]
    _para_vends   = [c for g in _grupos_vends for c in g]
    _lista        = [c for g in _grupos for c in g]

    # Adiciona ao painel da filial (gerente vê tudo da filial)
    for c in _lista:
        _tag = " [FDEP]" if c.get("is_fdep") else " [Inativo]"

        clientes_js[_filial_dest][c["faixa"]].append({
            "nome": c.get("cliente", c.get("nome", "")),
            "cliente": c.get("cliente", ""),
            "restricao": c.get("restricao", ""),
            "vencimento": c.get("vencimento", ""),
            "pagamento": c.get("pagamento", ""),
            "dias": c.get("dias", 0),
            "pendente": c.get("pendente", 0),
            "pago": c.get("pago", 0.0),
            "parcela": c.get("parcela", ""),
            "titulo": c.get("titulo", ""),
            "avalista": c.get("avalista", ""),
            "contato": c.get("contato", ""),
            "telefones": c.get("telefones", []),
            "mensagem_whatsapp": c.get("mensagem_whatsapp", ""),
            "vendedor": limpar_nome_display(c["vendedor"])[:20] + _tag,
            "novo": c.get("novo", False),
            "cliente_key": c.get("cliente_key", cliente_grupo_key_py(c)),
            "cobranca_key": c.get("cobranca_key", cobranca_row_key_py(c)),
        })

    # Distribui os 40% entre vendedores ativos da filial
    _vends_ativos = [
        r["nome_exibicao"]
        for _, r in df_vend[
            (df_vend["filial_vendedor"] == _filial_dest) & (~df_vend["is_gerente"])
        ].iterrows()
    ]

    if _vends_ativos and _para_vends:
        _n_vends_ativos = len(_vends_ativos)

        for _idx_c, c in enumerate(_para_vends):
            # Distribui round-robin pelos vendedores ativos
            _vend_dest = _vends_ativos[_idx_c % _n_vends_ativos]
            _tag = " [FDEP]" if c.get("is_fdep") else " [Inativo]"

            if _vend_dest not in clientes_por_vend_js:
                clientes_por_vend_js[_vend_dest] = {
                    "grave": [],
                    "alerta": [],
                    "atencao": []
                }

            clientes_por_vend_js[_vend_dest][c["faixa"]].append({
                "nome": c.get("cliente", c.get("nome", "")),
                "cliente": c.get("cliente", ""),
                "restricao": c.get("restricao", ""),
                "vencimento": c.get("vencimento", ""),
                "pagamento": c.get("pagamento", ""),
                "dias": c.get("dias", 0),
                "pendente": c.get("pendente", 0),
                "pago": c.get("pago", 0.0),
                "parcela": c.get("parcela", ""),
                "titulo": c.get("titulo", ""),
                "avalista": c.get("avalista", ""),
                "contato": c.get("contato", ""),
                "telefones": c.get("telefones", []),
                "mensagem_whatsapp": c.get("mensagem_whatsapp", ""),
                "vendedor": limpar_nome_display(c["vendedor"])[:20] + _tag,
                "novo": c.get("novo", False),
            })

print(f"📋 Clientes por filial (gerente):")
for f in ORDEM_FILIAIS:
    t = sum(len(clientes_js[f][x]) for x in ["grave", "alerta", "atencao"])
    if t:
        print(f"  {f}: {t} títulos")

print(f"📋 Clientes por vendedor: {len(clientes_por_vend_js)} vendedores com títulos")
# Ordena por valor pendente descendente
for f in clientes_js:
    for faixa in ["grave","alerta","atencao"]:
        clientes_js[f][faixa].sort(key=lambda x: x["pendente"], reverse=True)

# ── CREDIARISTAS: configuráveis por usuário/filial/percentual ─────────────────
clientes_crediarista_js = {}
recebimentos_crediarista_js = {}
for _cred_cfg in CREDIARISTAS_CONFIG:
    _fil_cred = str(_cred_cfg.get('filial', '')).upper()
    _login_cred = str(_cred_cfg.get('login', '')).lower()
    _pct_cred = max(0.0, min(100.0, float(_cred_cfg.get('pct', 100) or 100)))
    _src_cli = clientes_js.get(_fil_cred, {}) or {}
    clientes_crediarista_js[_login_cred] = {'grave': [], 'alerta': [], 'atencao': []}
    for _fx in ['grave','alerta','atencao']:
        _base_fx = list(_src_cli.get(_fx, []) or [])
        _base_fx = sorted(_base_fx, key=lambda r: _hashlib.md5((cobranca_row_key_py(r)+'|'+_login_cred).encode('utf-8')).hexdigest())
        _n = int(round(len(_base_fx) * (_pct_cred/100.0)))
        clientes_crediarista_js[_login_cred][_fx] = _base_fx[:_n]
    recebimentos_crediarista_js[_login_cred] = {'grave': [], 'alerta': [], 'atencao': []}


def aplicar_anti_duplicidade_operacional_carteiras():
    """
    V7.9 HOTFIX: remove duplicidade sem zerar a lista dos vendedores.

    Regra nova:
    - Cobrança10 continua exclusiva: se o título caiu no lote do terceiro, sai das outras listas operacionais.
    - Vendedor x vendedor: mantém somente um responsável vendedor para o mesmo título.
    - Crediariasta x crediarista: mantém somente um responsável crediarista para o mesmo título.
    - Vendedor x crediarista: NÃO remove do vendedor. As listas são tratadas como trilhas operacionais separadas,
      porque a regra antiga estava fazendo os vendedores perderem Para cobrar / Novos / Aguardando.
    - Filial/gerente continua sendo visão consolidada da carteira, não deve roubar lista individual do vendedor.
    """
    removidos = []

    def _key(r):
        try:
            return cobranca_row_key_py(r)
        except Exception:
            return ''

    # 1) Títulos do terceiro têm prioridade/exclusividade.
    terceiro_keys = set()
    try:
        for fx in ['grave','alerta','atencao']:
            for r in (clientes_terceiro_js or {}).get(fx, []) or []:
                k = _key(r)
                if k and k.count('|') >= 2:
                    terceiro_keys.add(k)
    except Exception:
        pass

    def _remove_keys_de_map(map_obj, tipo, keys_to_remove):
        total = 0
        try:
            for dono, data in list((map_obj or {}).items()):
                for fx in ['grave','alerta','atencao']:
                    nova = []
                    for r in (data or {}).get(fx, []) or []:
                        k = _key(r)
                        if k in keys_to_remove:
                            total += 1
                            removidos.append({'key': k, 'removido_de': f'{tipo}:{dono}', 'mantido_em': 'terceiro:Cobrança10', 'motivo': 'cobranca10_exclusiva'})
                        else:
                            nova.append(r)
                    data[fx] = nova
        except Exception:
            pass
        return total

    _remove_keys_de_map(clientes_crediarista_js, 'crediarista', terceiro_keys)
    _remove_keys_de_map(clientes_por_vend_js, 'vendedor', terceiro_keys)

    # 2) Deduplica dentro do próprio grupo de vendedores ou crediaristas, sem cruzar os dois grupos.
    def _dedupe_interno(map_obj, tipo):
        seen = {}
        try:
            donos = sorted(list((map_obj or {}).keys()), key=lambda x: str(x))
            for dono in donos:
                data = (map_obj or {}).get(dono) or {}
                for fx in ['grave','alerta','atencao']:
                    nova = []
                    for r in data.get(fx, []) or []:
                        k = _key(r)
                        if not k or k.count('|') < 2:
                            nova.append(r)
                            continue
                        if k not in seen:
                            seen[k] = {'tipo': tipo, 'dono': dono, 'faixa': fx}
                            nova.append(r)
                        else:
                            removidos.append({'key': k, 'removido_de': f'{tipo}:{dono}', 'mantido_em': f"{seen[k]['tipo']}:{seen[k]['dono']}", 'motivo': f'duplicado_interno_{tipo}'})
                    data[fx] = nova
        except Exception:
            pass

    _dedupe_interno(clientes_crediarista_js, 'crediarista')
    _dedupe_interno(clientes_por_vend_js, 'vendedor')

    if removidos:
        print(f"🧯 Anti-duplicidade operacional V7.9 aplicado: {len(removidos)} título(s) removido(s). Vendedor x crediarista preservado.")
        try:
            with open(os.path.join(pasta, 'relatorio_duplicidades_resolvidas.json'), 'w', encoding='utf-8') as f:
                json.dump({'gerado_em': now_brasilia().isoformat(), 'versao': 'V7.9', 'regra': 'cobranca10 exclusiva; dedupe interno por grupo; vendedor x crediarista preservado', 'total_removidos': len(removidos), 'removidos': removidos[:1500]}, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    else:
        print('✅ Anti-duplicidade operacional V7.9: nada para remover.')


aplicar_anti_duplicidade_operacional_carteiras()

_historico_comissao_cobranca10()



# =========================================================
# 🧡 DASHBOARD 2.0 - clientes sem movimento / dedupe carteira
# =========================================================
def _digits_only(v):
    return re.sub(r"\D+", "", str(v or ""))


def _parse_cliente_codigo_nome(raw):
    s = re.sub(r"\s+", " ", str(raw or "").strip())
    codigo = ""
    nome = s
    m = re.match(r"^(?:C\.)?\s*([0-9]+|Sem Código)\s*-\s*(.+)$", s, flags=re.I)
    if m:
        codigo = m.group(1).strip()
        nome = m.group(2).strip()
    nome = re.sub(r"\s+", " ", nome).strip()
    primeiro = nome.split()[0].title() if nome else "Cliente"
    return codigo, nome, primeiro




CLIENTES_SEM_MOV_FILIAL_VALUE_TO_KEY = {
    "1": "F1", "2": "F2", "3": "F3", "4": "F4", "5": "F5", "6": "F6",
    "7": "F8", "8": "F9", "10": "F1", "90": "F1", "99": "F1",
}
CLIENTES_SEM_MOV_FILIAL_TAG_BY_VALUE = {
    "1": "FILIAL_01", "2": "FILIAL_02", "3": "FILIAL_03", "4": "FILIAL_04", "5": "FILIAL_05", "6": "FILIAL_06",
    "7": "FILIAL_08", "8": "FILIAL_09", "10": "FILIAL_90_99",
}
CLIENTES_SEM_MOV_CIDADES_PERMITIDAS = {
    "CASTRO", "CARAMBEI", "CARAMBEÍ", "PIRAI DO SUL", "PIRAÍ DO SUL", "SOCAVAO", "SOCAVÃO",
    "ABAPAN", "ABAPA", "ABAPÃ", "TIBAGI", "VENTANIA", "TELEMACO BORBA", "TELÊMACO BORBA",
    "PONTA GROSSA", "TRONCO", "ITAIACOCA", "ITAICACOCA", "CARAMBEI PR", "CASTRO PR",
}

def _norm_cidade_mdl(v):
    s = str(v or "").strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _cidade_cliente_sem_movimento_permitida(cidade, uf=""):
    nc = _norm_cidade_mdl(cidade)
    nuf = _norm_cidade_mdl(uf)
    permitidas = {_norm_cidade_mdl(x) for x in CLIENTES_SEM_MOV_CIDADES_PERMITIDAS}
    if not nc:
        return False
    if nuf and nuf not in {"PR", "PARANA"}:
        return False
    if nc in permitidas:
        return True
    # Mantém alguns distritos/bairros próximos que podem vir compostos no SGI.
    for base in permitidas:
        if base and (nc.startswith(base + " ") or nc.endswith(" " + base) or base in nc):
            return True
    return False
def _extrair_telefones_mdl(contatos):
    """
    Extrai somente números prováveis de WhatsApp/celular.
    Mantém o número interno com DDI 55 para abrir wa.me, mas evita:
      - telefone fixo com 10 dígitos (ex.: 42 3224-2622)
      - número vindo misturado de e-mail/código
      - duplicidades
    Regras aceitas:
      - 42 9XXXX-XXXX / XX 9XXXX-XXXX -> 55 + DDD + 9 dígitos
      - 9XXXX-XXXX sem DDD -> assume DDD 42
    """
    txt = str(contatos or "")
    nums = []
    seen = set()

    # Captura trechos telefônicos comuns, mas valida depois.
    candidatos = []
    for m in re.finditer(r"(?:\+?55\s*)?(?:\(?\d{2}\)?\s*)?\d{4,5}[-\s]?\d{4}", txt):
        candidatos.append(_digits_only(m.group(0)))

    for raw in candidatos:
        if raw.startswith("55") and len(raw) > 11:
            raw = raw[2:]

        # Quando vier só 9 dígitos de celular sem DDD, assume DDD 42.
        if len(raw) == 9 and raw.startswith("9"):
            local = "42" + raw
        elif len(raw) >= 11:
            local = raw[-11:]
        else:
            # 10 dígitos normalmente é fixo; não usar para WhatsApp.
            continue

        # Precisa ser DDD + celular 9 dígitos.
        if len(local) != 11:
            continue
        if local[2] != "9":
            continue
        if len(set(local[-8:])) <= 1:
            continue

        final = "55" + local
        if final not in seen:
            seen.add(final)
            nums.append(final)
    return nums


def _filial_key_from_sem_movimento_filename(path):
    name = os.path.basename(str(path)).upper()
    if "FILIAL_90_99" in name or "FILIAL 90" in name or "90_99" in name:
        return "F1"
    m = re.search(r"FILIAL[_\s-]*(\d+)", name)
    if m:
        n = int(m.group(1))
        if n in (90, 99):
            return "F1"
        if n in (7,):
            # No SGI o value 7 representa FILIAL 08. Se algum arquivo antigo ficou FILIAL_07,
            # tratamos como F8 para não criar F7 no dashboard.
            return "F8"
        if n == 8:
            # Arquivo antigo FILIAL_08 é F8; arquivo novo value 8 será renomeado como FILIAL_09.
            return "F8"
        if n == 9:
            return "F9"
        return f"F{n}"
    return "F1"


def parse_clientes_sem_movimento_xls(path, filial_key=None):
    filial_key = filial_key or _filial_key_from_sem_movimento_filename(path)
    try:
        df = pd.read_excel(path, engine="openpyxl")
    except Exception:
        try:
            df = pd.read_excel(path)
        except Exception as e:
            print(f"⚠️ Não consegui ler clientes sem movimento {path}: {e}")
            return []
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    out = []
    for _, row in df.iterrows():
        cliente_raw = row.get("Cliente", "")
        codigo, nome, primeiro = _parse_cliente_codigo_nome(cliente_raw)
        if not nome or nome.lower() in {"nan", "cliente"}:
            continue
        cidade = str(row.get("Cidade", "") or "")
        uf = str(row.get("UF", "") or "")
        if not _cidade_cliente_sem_movimento_permitida(cidade, uf):
            continue
        contatos = str(row.get("Contatos", "") or "")
        telefones = _extrair_telefones_mdl(contatos)
        if not telefones:
            continue
        dias = row.get("Dias Sem Movimento", 0)
        try:
            dias = int(float(str(dias).replace(",", ".")))
        except Exception:
            dias = 0
        # Regra MDL: reativação de 45 dias até no máximo 5 anos sem comprar.
        # Muito antigo tende a ter telefone desatualizado e polui a carteira.
        if dias < 45 or dias > 1825:
            continue
        out.append({
            "filial": "F1" if str(filial_key).upper() in {"F90", "F99", "90", "99"} else str(filial_key).upper(),
            "codigo": codigo,
            "cliente": nome,
            "primeiro_nome": primeiro,
            "dias_sem_movimento": dias,
            "ultimo_movimento": str(row.get("Data do Último Movimento", "") or ""),
            "tipo_movimento": str(row.get("Tipo de Movimento", "") or ""),
            "cidade": cidade,
            "bairro": str(row.get("Bairro", "") or ""),
            "uf": uf,
            "contatos_raw": contatos,
            "telefones": telefones,
            "origem": os.path.basename(str(path)),
        })
    return out





# =========================================================
# 🎂 Aniversariantes do dia - XLS + Selenium
# =========================================================
ANIVERSARIANTES_FILIAL_VALUE_TO_KEY = CLIENTES_SEM_MOV_FILIAL_VALUE_TO_KEY.copy()
ANIVERSARIANTES_FILIAL_TAG_BY_VALUE = CLIENTES_SEM_MOV_FILIAL_TAG_BY_VALUE.copy()

def _filial_key_from_aniversario_text(v):
    s = str(v or "").upper()
    if "90/99" in s or "FILIAL 90" in s:
        return "F1"
    m = re.search(r"FILIAL\s*(\d+)", s)
    if m:
        n = int(m.group(1))
        if n in (90,99): return "F1"
        if n == 7: return "F8"
        if n == 8: return "F9"
        if n == 9: return "F9"
        return f"F{n}"
    return "F1"

def _filial_key_from_aniversario_filename(path):
    name = os.path.basename(str(path)).upper()
    if "90_99" in name or "FILIAL_90" in name or "90/99" in name:
        return "F1"
    m = re.search(r"FILIAL[_\s-]*(\d+)", name)
    if m:
        n = int(m.group(1))
        if n in (90,99): return "F1"
        if n == 7: return "F8"
        if n == 8: return "F9"
        if n == 9: return "F9"
        return f"F{n}"
    return "F1"

def nome_arquivo_aniversariantes_valido(fname):
    s = str(fname or "").lower().strip()
    if s.startswith("~$") or not (s.endswith(".xls") or s.endswith(".xlsx")):
        return False
    return ("anivers" in s or "aniversario" in s or "aniversarios" in s or "birthday" in s)

def parse_aniversariantes_xls(path):
    try:
        df = pd.read_excel(path, engine="openpyxl")
    except Exception:
        try:
            df = pd.read_excel(path)
        except Exception as e:
            print(f"⚠️ Não consegui ler aniversariantes {path}: {e}")
            return []
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    out=[]
    for _, row in df.iterrows():
        nome = str(row.get("Cliente", "") or "").strip()
        if not nome or nome.lower() in {"nan", "cliente"}:
            continue
        primeiro = nome.split()[0].title() if nome else "Cliente"
        cidade_raw = str(row.get("Cidade", "") or "").strip()
        cidade = cidade_raw
        uf = ""
        m = re.match(r"(.+?)\s*-\s*([A-Z]{2})$", cidade_raw.strip(), flags=re.I)
        if m:
            cidade = m.group(1).strip(); uf = m.group(2).strip().upper()
        if not _cidade_cliente_sem_movimento_permitida(cidade, uf):
            continue
        filial_text = str(row.get("Filial Ult. Venda", row.get("Filial", "")) or "")
        filial = _filial_key_from_aniversario_text(filial_text) or _filial_key_from_aniversario_filename(path)
        contatos=[]
        for c in list(df.columns):
            if str(c).lower().startswith("contato") or "meios" in str(c).lower():
                val = row.get(c, "")
                if val is not None and str(val).lower() != "nan": contatos.append(str(val))
        telefones = _extrair_telefones_mdl(" | ".join(contatos))
        if not telefones:
            continue
        nasc = str(row.get("Data de Nascimento", row.get("Nascimento", "")) or "")
        ult = str(row.get("Data Ult. Venda", row.get("Data Últ. Venda", "")) or "")
        out.append({
            "filial":"F1" if str(filial).upper() in {"F90","F99","90","99"} else str(filial).upper(),
            "cliente":nome,
            "primeiro_nome":primeiro,
            "nascimento":nasc,
            "ultima_venda":ult,
            "cidade":cidade,
            "uf":uf,
            "telefones":telefones,
            "contatos_raw":" | ".join(contatos),
            "origem":os.path.basename(str(path)),
        })
    return out



def _remote_json_colaborador_v95(remote_name, default=None):
    """V9.5: lê JSON já publicado no FTP/URL pública para não sobrescrever listas/configs com vazio em deploy comum."""
    try:
        data = _ftp_json(remote_name, None)
        if data is not None:
            return data
    except Exception:
        pass
    try:
        url = REMOTE_PUBLIC_BASE + '/' + remote_name
        req = urllib.request.Request(url, headers={'Cache-Control': 'no-cache', 'Pragma': 'no-cache'})
        with urllib.request.urlopen(req, timeout=12) as resp:
            raw = resp.read().decode('utf-8', errors='ignore').strip()
        if raw:
            return json.loads(raw)
    except Exception:
        pass
    return default

def _json_has_items_v95(path, list_keys=('clientes','dados','seen')):
    try:
        if not os.path.exists(path):
            return False
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, list):
            return len(data) > 0
        if isinstance(data, dict):
            for k in list_keys:
                v = data.get(k)
                if isinstance(v, list) and len(v) > 0:
                    return True
            # config_meta.json: global/individual contam como conteúdo persistente
            if isinstance(data.get('global'), dict) and data.get('global'):
                return True
            if isinstance(data.get('individual'), dict) and data.get('individual'):
                return True
    except Exception:
        return False
    return False

def _baixou_listas_pesadas_v95():
    return os.getenv('BAIXAR_CLIENTES_SEM_MOVIMENTO','0') == '1' or os.getenv('BAIXAR_ANIVERSARIANTES','0') == '1' or os.getenv('FORCE_DAILY_LISTS_ON_BOOT','0') == '1'

def carregar_aniversariantes_local():
    arquivos=[]
    for fname in os.listdir(pasta):
        if nome_arquivo_aniversariantes_valido(fname):
            arquivos.append(os.path.join(pasta, fname))
    arquivos=sorted(arquivos, key=lambda x: os.path.getmtime(x), reverse=True)
    print(f"🎂 Aniversariantes: {len(arquivos)} XLS encontrado(s) para leitura")
    if not arquivos:
        remote = _remote_json_colaborador_v95('aniversariantes_dia.json', None)
        if isinstance(remote, dict):
            dados = remote.get('dados') or remote.get('clientes') or []
            if isinstance(dados, list) and dados:
                print(f"🎂 Aniversariantes V9.5: sem XLS local; mantendo JSON do FTP com {len(dados)} cliente(s).")
                try:
                    with open(os.path.join(pasta, 'aniversariantes_dia.json'), 'w', encoding='utf-8') as f:
                        json.dump(remote, f, ensure_ascii=False, indent=2)
                except Exception:
                    pass
                return dados
        print("ℹ️ Para testar sem Selenium, coloque aniversarios.xls ou relatorio_aniversariantes*.xls nesta pasta. Para baixar pelo Sólidus, rode com BAIXAR_ANIVERSARIANTES=1. V9.5: não vou sobrescrever o FTP com lista vazia.")
        return []
    todos=[]; seen=set()
    for arq in arquivos:
        for row in parse_aniversariantes_xls(arq):
            key = (row.get('filial',''), _norm_cidade_mdl(row.get('cliente','')), tuple(row.get('telefones') or []))
            if key in seen: continue
            seen.add(key); todos.append(row)
    try:
        with open(os.path.join(pasta, "aniversariantes_dia.json"), "w", encoding="utf-8") as f:
            json.dump({"gerado_em": now_brasilia().isoformat(), "total": len(todos), "dados": todos}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ Não consegui salvar aniversariantes_dia.json: {e}")
    print(f"🎂 Aniversariantes carregados: {len(todos)} cliente(s)")
    return todos

def baixar_aniversariantes_selenium():
    if os.getenv("BAIXAR_ANIVERSARIANTES", "0") != "1":
        print("ℹ️ Download aniversariantes desativado por BAIXAR_ANIVERSARIANTES=0")
        return []
    print("🎂 Iniciando download Aniversariantes do dia...")
    baixados=[]
    try:
        driver.get(URL + "/relatorio_aniversariantes")
        time.sleep(2)
        hoje_dm = now_brasilia().strftime("%d/%m")
        for input_id in ["data_inicial", "data_final"]:
            try:
                el = wait.until(EC.presence_of_element_located((By.ID, input_id)))
                driver.execute_script("arguments[0].removeAttribute('readonly'); arguments[0].value='';", el)
                el.click(); el.send_keys(hoje_dm); driver.execute_script("arguments[0].dispatchEvent(new Event('input',{bubbles:true})); arguments[0].dispatchEvent(new Event('change',{bubbles:true})); arguments[0].blur();", el)
            except Exception:
                pass
        try:
            Select(wait.until(EC.presence_of_element_located((By.ID, "_formato")))).select_by_value("xls")
        except Exception:
            driver.execute_script("var s=document.getElementById('_formato'); if(s){s.value='xls'; s.dispatchEvent(new Event('change',{bubbles:true}));}")
        inicio_download_ts=time.time()
        antes=set(os.listdir(download_dir))
        try:
            btn=wait.until(EC.element_to_be_clickable((By.ID, "gerar")))
            driver.execute_script("arguments[0].click();", btn)
        except Exception:
            driver.execute_script("document.getElementById('gerar')?.click();")
        limite=time.time()+75
        novo=None
        while time.time()<limite:
            time.sleep(1)
            nomes=set(os.listdir(download_dir))
            novos=[n for n in (nomes-antes) if (n.lower().endswith('.xls') or n.lower().endswith('.xlsx')) and not n.startswith('~$') and not n.lower().endswith('.crdownload')]
            recentes=[n for n in nomes if (n.lower().endswith('.xls') or n.lower().endswith('.xlsx')) and not n.startswith('~$') and not n.lower().endswith('.crdownload') and os.path.getmtime(os.path.join(download_dir,n))>=inicio_download_ts-2]
            candidatos=novos or recentes
            if candidatos:
                novo=max([os.path.join(download_dir,n) for n in candidatos], key=os.path.getmtime)
                break
        if novo:
            ext = ".xlsx" if novo.lower().endswith(".xlsx") else ".xls"
            destino=os.path.join(pasta, f"aniversariantes_{now_brasilia().strftime('%Y%m%d')}{ext}")
            shutil.copy2(novo,destino)
            baixados.append(destino)
            print(f"✅ Aniversariantes baixado: {os.path.basename(destino)}")
        else:
            print("⚠️ Aniversariantes: download não encontrado")
    except Exception as e:
        print(f"⚠️ Erro baixando aniversariantes: {e}")
    return baixados

# =========================================================
# 🧡 Selenium - baixar relatório Clientes sem Movimento por filial
# =========================================================
def nome_arquivo_clientes_sem_movimento_valido(fname):
    """
    Aceita o XLS do relatório Clientes sem Movimento.
    O SGI normalmente baixa como relatorio_clientes_sem_movimento_*.xls,
    mas em alguns ambientes pode vir com prefixo UUID ou nome diferente.
    Por isso o download desta rotina usa também detecção por arquivo novo.
    """
    s = str(fname or "").lower().strip()
    if s.startswith("~$"):
        return False
    if not (s.endswith(".xls") or s.endswith(".xlsx")):
        return False
    bloqueados = [
        "relatorio_contas", "contas_pagar_receber", "margem_bruta", "metas_vendas",
        "dashboard", "historico", "fechamentos", "credenciais", "config_meta",
        "cobrancas_log", "quitados_180d"
    ]
    if any(b in s for b in bloqueados):
        return False
    return ("cliente" in s and "mov" in s) or "sem_movimento" in s or "relatorio_clientes" in s


def _qualquer_xls_novo_clientes(fname):
    """Fallback: aceita qualquer XLS novo que não seja dos relatórios de cobrança/vendas/margem."""
    s = str(fname or "").lower().strip()
    if s.startswith("~$") or not (s.endswith(".xls") or s.endswith(".xlsx")):
        return False
    bloqueados = [
        "relatorio_contas", "contas_pagar_receber", "margem_bruta", "metas_vendas",
        "dashboard", "historico", "fechamentos", "credenciais", "config_meta",
        "cobrancas_log", "quitados_180d"
    ]
    return not any(b in s for b in bloqueados)


def _renomear_download_clientes_sem_movimento(caminho, filial_value):
    try:
        filial_value = str(filial_value)
        tag = CLIENTES_SEM_MOV_FILIAL_TAG_BY_VALUE.get(filial_value)
        if not tag:
            tag = f"FILIAL_{int(filial_value):02d}"
        ext = ".xlsx" if str(caminho).lower().endswith(".xlsx") else ".xls"
        destino = os.path.join(pasta, f"{tag}_relatorio_clientes_sem_movimento_{now_brasilia().strftime('%Y%m%d%H%M%S')}{ext}")
        if os.path.abspath(caminho) != os.path.abspath(destino):
            try:
                if os.path.exists(destino):
                    os.remove(destino)
            except Exception:
                pass
            shutil.copy2(caminho, destino)
        return destino
    except Exception as e:
        print(f"⚠️ Não consegui renomear XLS clientes sem movimento: {e}")
        return caminho


def _set_input_clientes_sem_movimento(input_id, valor):
    try:
        el = wait.until(EC.presence_of_element_located((By.ID, input_id)))
        try:
            driver.execute_script("arguments[0].removeAttribute('readonly'); arguments[0].removeAttribute('disabled');", el)
        except Exception:
            pass
        driver.execute_script("arguments[0].value = '';", el)
        if valor is not None and str(valor) != "":
            try:
                el.click()
                el.send_keys(str(valor))
            except Exception:
                driver.execute_script("arguments[0].value = arguments[1];", el, str(valor))
        driver.execute_script("arguments[0].dispatchEvent(new Event('input',{bubbles:true})); arguments[0].dispatchEvent(new Event('change',{bubbles:true})); arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));", el)
        return True
    except Exception:
        return False


def _select_by_id_value(select_id, value):
    try:
        el = wait.until(EC.presence_of_element_located((By.ID, select_id)))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        try:
            Select(el).select_by_value(str(value))
        except Exception:
            driver.execute_script("arguments[0].value = arguments[1];", el, str(value))
        driver.execute_script("""
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input',{bubbles:true}));
            arguments[0].dispatchEvent(new Event('change',{bubbles:true}));
            arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));
            if (document.activeElement) document.activeElement.blur();
        """, el, str(value))
        time.sleep(0.35)
        return True
    except Exception as e:
        print(f"⚠️ Não consegui selecionar {select_id}={value}: {e}")
        return False


def _set_checkbox_switch_by_id(input_id, checked=True):
    try:
        el = driver.find_element(By.ID, input_id)
        atual = bool(el.is_selected())
        if atual != bool(checked):
            driver.execute_script("arguments[0].click();", el)
            time.sleep(0.4)
        return True
    except Exception:
        try:
            return bool(driver.execute_script("""
                const id=arguments[0], checked=!!arguments[1];
                const el=document.getElementById(id);
                if(!el) return false;
                el.checked=checked;
                el.value=checked?'true':'false';
                el.dispatchEvent(new Event('change',{bubbles:true}));
                return true;
            """, input_id, checked))
        except Exception:
            return False


def _preparar_filtros_clientes_sem_movimento(filial_value):
    """
    V2.3 - simples e fiel ao que funciona manualmente no SGI:
    1) seleciona filial_id
    2) garante dias_sem_movimentacao = 45 se o campo existir
    3) seleciona _formato = xls
    4) NÃO mexe em Considerar Movimento, Período, Situação Financeira ou Tipo de Cliente.
    """
    esperar_sumir_overlays(timeout=8)
    _select_by_id_value("filial_id", filial_value)
    time.sleep(0.35)
    _set_input_clientes_sem_movimento("dias_sem_movimentacao", "45")
    _select_by_id_value("_formato", "xls")
    try:
        driver.execute_script("document.body.click();")
    except Exception:
        pass
    esperar_sumir_overlays(timeout=8)


def _clicar_gerar_clientes_sem_movimento():
    """Clica no botão Gerar com JS para não depender de dropdown/overlay aberto."""
    esperar_sumir_overlays(timeout=10)
    try:
        btn = wait.until(EC.presence_of_element_located((By.ID, "gerar")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(0.2)
        driver.execute_script("arguments[0].click();", btn)
        return True
    except Exception as e:
        print(f"⚠️ Clique JS em Gerar falhou, tentando clique seguro: {e}")
        try:
            clicar_seguro_xpath("//button[@id='gerar' or contains(normalize-space(.),'Gerar')]", timeout=15)
            return True
        except Exception as e2:
            print(f"⚠️ Não consegui clicar em Gerar: {e2}")
            return False


def _aguardar_download_clientes_sem_movimento(arquivos_antes, filial_label, timeout_seg=180):
    inicio = time.time()
    while time.time() - inicio < timeout_seg:
        time.sleep(1.5)
        nomes = os.listdir(download_dir)
        baixando = any(str(f).lower().endswith((".crdownload", ".tmp")) for f in nomes)
        if baixando:
            if int(time.time() - inicio) % 15 in (0, 1):
                print(f"   ⏳ {filial_label}: aguardando download...")
            continue
        novos_especificos = set(f for f in nomes if nome_arquivo_clientes_sem_movimento_valido(f)) - arquivos_antes
        novos_fallback = set(f for f in nomes if _qualquer_xls_novo_clientes(f)) - arquivos_antes
        novos = novos_especificos or novos_fallback
        if novos:
            return max([os.path.join(download_dir, f) for f in novos], key=os.path.getctime)
    return None


def baixar_clientes_sem_movimento_selenium():
    """
    Baixa o relatório Clientes sem Movimento filial por filial.
    F90/99 é renomeada e depois unificada na F1 pelo parser.
    Pode desativar com BAIXAR_CLIENTES_SEM_MOVIMENTO=0.
    """
    if os.getenv("BAIXAR_CLIENTES_SEM_MOVIMENTO", "0") == "0":
        print("ℹ️ Download clientes sem movimento desativado por BAIXAR_CLIENTES_SEM_MOVIMENTO=0 (padrão). Scheduler libera 1x ao dia às 07h.")
        return []

    print("\n🧡 Iniciando download Clientes sem Movimento +45 dias por filial...")
    baixados = []
    filiais = [("1", "F1"), ("2", "F2"), ("3", "F3"), ("4", "F4"), ("5", "F5"), ("6", "F6"), ("7", "F8"), ("8", "F9"), ("10", "F90/99→F1")]

    try:
        driver.get(URL + "/relatorio_clientes_sem_movimento")
        wait.until(EC.presence_of_element_located((By.ID, "formulario")))
        wait.until(EC.presence_of_element_located((By.ID, "filial_id")))
        wait.until(EC.presence_of_element_located((By.ID, "_formato")))
        time.sleep(1.0)
    except Exception as e:
        print(f"⚠️ Não consegui abrir relatório Clientes sem Movimento: {e}")
        return baixados

    for filial_value, filial_label in filiais:
        try:
            print(f"📥 Clientes sem movimento: gerando {filial_label}...")
            arquivos_antes = set(f for f in os.listdir(download_dir) if _qualquer_xls_novo_clientes(f) or nome_arquivo_clientes_sem_movimento_valido(f))
            _preparar_filtros_clientes_sem_movimento(filial_value)
            time.sleep(0.8)
            if not _clicar_gerar_clientes_sem_movimento():
                print(f"⚠️ Clientes sem movimento {filial_label}: não clicou em Gerar")
                continue
            caminho = _aguardar_download_clientes_sem_movimento(arquivos_antes, filial_label, timeout_seg=210)
            if not caminho:
                try:
                    vals = driver.execute_script("""
                        return {
                          filial: document.getElementById('filial_id')?.value,
                          formato: document.getElementById('_formato')?.value,
                          dias: document.getElementById('dias_sem_movimentacao')?.value
                        };
                    """)
                    print(f"⚠️ Clientes sem movimento {filial_label}: download não encontrado | tela={vals}")
                except Exception:
                    print(f"⚠️ Clientes sem movimento {filial_label}: download não encontrado")
                continue
            destino = _renomear_download_clientes_sem_movimento(caminho, filial_value)
            baixados.append(destino)
            print(f"✅ Clientes sem movimento {filial_label}: {destino}")
        except Exception as e:
            print(f"⚠️ Erro baixando clientes sem movimento {filial_label}: {e}")
            try:
                driver.get(URL + "/relatorio_clientes_sem_movimento")
                wait.until(EC.presence_of_element_located((By.ID, "formulario")))
                time.sleep(1.0)
            except Exception:
                pass

    print(f"🧡 Clientes sem movimento: {len(baixados)} arquivo(s) baixado(s)")
    return baixados


def _filial_key_from_csm_filename(fn):
    up = os.path.basename(str(fn)).upper()
    if "FILIAL_90_99" in up or "FILIAL_90" in up:
        return "F90_99"
    m = re.search(r"FILIAL[_\s-]*(\d{1,2})", up)
    if not m:
        return None
    n = int(m.group(1))
    if n == 1: return "F1"
    if n == 2: return "F2"
    if n == 3: return "F3"
    if n == 4: return "F4"
    if n == 5: return "F5"
    if n == 6: return "F6"
    if n == 8: return "F8"
    if n == 9: return "F9"
    return f"F{n}"

def carregar_clientes_sem_movimento_local():
    """
    V8.5 - evita duplicar a base de clientes sem movimento.
    Quando existem arquivos FILIAL_* baixados pelo robô, lê somente o XLS mais recente de cada filial.
    Os downloads brutos do navegador (_relatorio_clientes_sem_movimento_*.xls) são ignorados para não dobrar a quantidade.
    """
    json_path = os.path.join(pasta, "clientes_sem_movimento.json")
    base_path = os.path.join(pasta, "clientes_sem_movimento_base.json")
    seen_path = os.path.join(pasta, "clientes_sem_movimento_seen.json")
    global clientes_sem_movimento_meta_py, clientes_sem_movimento_base_py
    clientes_sem_movimento_meta_py = {"modo":"acionaveis_novos_pendentes_retorno_10d", "base_total":0, "acionaveis_total":0, "novos_total":0, "seen_total":0}
    clientes_sem_movimento_base_py = []

    def _csm_key_row(r):
        tels = r.get("telefones") or []
        tel = ""
        if isinstance(tels, list) and tels:
            tel = re.sub(r"\D+", "", str(tels[0]))
        return "|".join([
            str(r.get("filial") or "").strip().upper(),
            normalizar_texto_match(r.get("cliente")),
            tel,
        ])

    def _load_seen_clientes_sem_movimento():
        """V8.7: histórico de clientes já exibidos para o dashboard trabalhar só com NOVOS."""
        seen = set()
        fonte = None
        # 1) Preferir arquivo seen local, se existir.
        for path in [seen_path, json_path]:
            try:
                if os.path.exists(path):
                    with open(path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    rows_old = data.get("clientes", data.get("seen", data)) if isinstance(data, dict) else data
                    if isinstance(rows_old, list):
                        for r0 in rows_old:
                            if isinstance(r0, dict):
                                k0 = r0.get("key") or _csm_key_row(r0)
                                if k0:
                                    seen.add(str(k0))
                            elif r0:
                                seen.add(str(r0))
                        fonte = os.path.basename(path)
                        break
            except Exception as e:
                print(f"⚠️ CSM seen: falha lendo {os.path.basename(path)}: {e}")
        # 2) Em produção, se não existir local, baixa do FTP.
        if not seen:
            for remote_name in ["clientes_sem_movimento_seen.json", "clientes_sem_movimento.json"]:
                try:
                    data = _ftp_json(remote_name, None)
                    rows_old = data.get("clientes", data.get("seen", data)) if isinstance(data, dict) else data
                    if isinstance(rows_old, list):
                        for r0 in rows_old:
                            if isinstance(r0, dict):
                                k0 = r0.get("key") or _csm_key_row(r0)
                                if k0:
                                    seen.add(str(k0))
                            elif r0:
                                seen.add(str(r0))
                        fonte = f"FTP/{remote_name}"
                        break
                except Exception:
                    pass
        if fonte:
            print(f"🧡 Clientes sem movimento V8.7: histórico seen carregado de {fonte} ({len(seen)} chave(s))")
        else:
            print("🧡 Clientes sem movimento V8.7: sem histórico seen; primeira carga pode entrar inteira")
        return seen

    candidatos = []
    for base in list(dict.fromkeys([pasta, download_dir, "/mnt/data"])):
        try:
            for fn in os.listdir(base):
                low = fn.lower()
                if low.startswith("~$"):
                    continue
                if "relatorio_clientes_sem_movimento" in low and (low.endswith(".xls") or low.endswith(".xlsx")):
                    candidatos.append(os.path.join(base, fn))
        except Exception:
            pass
    candidatos = sorted(list(dict.fromkeys(candidatos)))
    por_filial = {}
    for path in candidatos:
        key = _filial_key_from_csm_filename(path)
        if not key:
            continue
        try:
            mt = os.path.getmtime(path)
        except Exception:
            mt = 0
        old = por_filial.get(key)
        if not old or mt > old[0]:
            por_filial[key] = (mt, path)
    if por_filial:
        candidatos = [v[1] for k, v in sorted(por_filial.items())]
        print(f"🧡 Clientes sem movimento: usando {len(candidatos)} XLS FILIAL_* mais recente(s), ignorando downloads brutos duplicados")
    elif candidatos:
        def _date_from_name(path):
            m = re.search(r"(20\d{12}|20\d{6})", os.path.basename(path))
            return m.group(1)[:8] if m else "00000000"
        max_date = max(_date_from_name(x) for x in candidatos)
        candidatos = [x for x in candidatos if _date_from_name(x) == max_date]
        print(f"🧡 Clientes sem movimento: sem FILIAL_*. Usando {len(candidatos)} XLS bruto(s) da data {max_date}")
    rows = []
    if candidatos:
        print(f"🧡 Clientes sem movimento: {len(candidatos)} XLS encontrado(s) para leitura")
        for path in candidatos:
            try:
                parsed = parse_clientes_sem_movimento_xls(path, None)
                rows.extend(parsed)
                print(f"   ✅ {os.path.basename(path)} → {len(parsed)} cliente(s) com telefone")
            except Exception as e:
                print(f"   ⚠️ Erro lendo {os.path.basename(path)}: {e}")
    else:
        print("⚠️ Clientes sem movimento: nenhum XLS encontrado na pasta. Tentando JSON anterior/FTP...")
        for local_path in [json_path, base_path]:
            if os.path.exists(local_path):
                try:
                    with open(local_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    old_rows = data.get("clientes", data) if isinstance(data, dict) else data
                    if isinstance(old_rows, list) and old_rows:
                        print(f"ℹ️ Clientes sem movimento V9.5: usando JSON local {os.path.basename(local_path)} com {len(old_rows)} registro(s)")
                        return old_rows
                except Exception as e:
                    print(f"⚠️ Não consegui ler {os.path.basename(local_path)}: {e}")
        for remote_name in ['clientes_sem_movimento.json', 'clientes_sem_movimento_base.json']:
            remote = _remote_json_colaborador_v95(remote_name, None)
            if isinstance(remote, dict):
                old_rows = remote.get('clientes') or remote.get('dados') or []
                if isinstance(old_rows, list) and old_rows:
                    print(f"🧡 Clientes sem movimento V9.5: sem XLS local; mantendo {remote_name} do FTP com {len(old_rows)} cliente(s).")
                    try:
                        local_dest = json_path if remote_name == 'clientes_sem_movimento.json' else base_path
                        with open(local_dest, 'w', encoding='utf-8') as f:
                            json.dump(remote, f, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                    return old_rows
        print("🧡 Clientes sem movimento V9.5: sem XLS e sem JSON remoto com dados; não vou sobrescrever FTP com vazio.")
        return []
    seen = set(); final = []
    for r in rows:
        tels = r.get("telefones") or []
        k = (r.get("filial"), normalizar_texto_match(r.get("cliente")), tels[0] if tels else "", r.get("ultimo_movimento"))
        if k in seen:
            continue
        seen.add(k)
        final.append(r)
    final.sort(key=lambda x: (str(x.get("filial","")), -int(x.get("dias_sem_movimento") or 0), str(x.get("cliente",""))))

    # V9.2: lista operacional inteligente de clientes sem movimento.
    # Mantém base completa, memória de vistos e usa o histórico de envios:
    # - cliente novo entra;
    # - cliente antigo sem envio continua pendente;
    # - cliente enviado volta após 10 dias se ainda estiver na base do Sólidus;
    # - se comprou e não veio mais no relatório, sai automaticamente.
    seen_antigo = _load_seen_clientes_sem_movimento()
    final_com_key = []
    for r in final:
        try:
            rr = dict(r)
            rr["_csm_key"] = _csm_key_row(rr)
            final_com_key.append(rr)
        except Exception:
            final_com_key.append(r)

    def _parse_dt_csm(v):
        txt = str(v or "").strip()
        if not txt:
            return None
        try:
            if txt.endswith('Z'):
                txt = txt[:-1] + '+00:00'
            return datetime.fromisoformat(txt.replace(' ', 'T')[:19]).date()
        except Exception:
            pass
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"]:
            try:
                return datetime.strptime(txt[:19], fmt).date()
            except Exception:
                continue
        return None

    def _load_reativacao_logs_csm():
        logs = []
        for src in [os.path.join(pasta, 'cobrancas_log.json')]:
            try:
                if os.path.exists(src):
                    with open(src, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    if isinstance(data, list):
                        logs.extend(data)
            except Exception as e:
                print(f"⚠️ CSM logs: falha lendo cobrancas_log local: {e}")
        try:
            data = _ftp_json('cobrancas_log.json', None)
            if isinstance(data, list):
                # FTP é a fonte mais completa; se veio, usa ela no lugar do local.
                logs = data
        except Exception:
            pass
        stats = {}
        for x in logs or []:
            try:
                if str(x.get('titulo') or '').upper() != 'REATIVACAO':
                    continue
                k = str(x.get('cliente_key') or x.get('cobranca_key') or '').strip()
                if not k:
                    parc = str(x.get('parcela') or '')
                    m = re.search(r'CLIENTE_SEM_MOVIMENTO\|([^|]+)', parc)
                    if m:
                        k = m.group(1).strip()
                if not k:
                    continue
                d = _parse_dt_csm(x.get('server_time') or x.get('data') or x.get('created_at') or x.get('criado_em'))
                if not d:
                    continue
                st = stats.setdefault(k, {'qtd':0, 'primeiro':d, 'ultimo':d})
                st['qtd'] += 1
                if d < st['primeiro']:
                    st['primeiro'] = d
                if d > st['ultimo']:
                    st['ultimo'] = d
            except Exception:
                continue
        print(f"🧡 Clientes sem movimento V9.2: histórico de envios REATIVACAO carregado ({len(stats)} cliente(s) com envio)")
        return stats

    clientes_sem_movimento_base_py = final_com_key
    logs_stats = _load_reativacao_logs_csm()
    hoje_csm = now_brasilia().date()
    cooldown_dias = int(os.getenv('REATIVACAO_REENVIO_DIAS', '10') or '10')
    actionable = []
    novos_count = pendentes_count = retorno_count = aguardando_count = 0
    for r in final_com_key:
        rr = dict(r)
        k = str(rr.get('_csm_key') or _csm_key_row(rr))
        st = logs_stats.get(k)
        is_novo = k not in seen_antigo
        rr['_reat_novo'] = bool(is_novo)
        rr['_reat_cooldown_dias'] = cooldown_dias
        if st:
            ultimo = st.get('ultimo')
            primeiro = st.get('primeiro')
            dias = (hoje_csm - ultimo).days if ultimo else 9999
            rr['_reat_qtd_envios'] = int(st.get('qtd') or 0)
            rr['_reat_primeiro_envio'] = primeiro.strftime('%d/%m/%Y') if primeiro else ''
            rr['_reat_ultimo_envio'] = ultimo.strftime('%d/%m/%Y') if ultimo else ''
            rr['_reat_dias_desde_envio'] = int(dias)
            if dias >= cooldown_dias:
                rr['_reat_motivo'] = 'retorno_10d'
                retorno_count += 1
                actionable.append(rr)
            else:
                aguardando_count += 1
        else:
            if is_novo:
                rr['_reat_motivo'] = 'novo'
                novos_count += 1
            else:
                rr['_reat_motivo'] = 'pendente_sem_envio'
                pendentes_count += 1
            actionable.append(rr)

    print(f"🧡 Clientes sem movimento V9.2: {len(actionable)} acionável(is) de {len(final_com_key)} na base | novos={novos_count} pendentes_sem_envio={pendentes_count} retorno_{cooldown_dias}d={retorno_count} aguardando_reenvio={aguardando_count}")
    seen_union = sorted(set(list(seen_antigo) + [str(r.get('_csm_key') or _csm_key_row(r)) for r in final_com_key if (r.get('_csm_key') or _csm_key_row(r))]))
    try:
        with open(seen_path, "w", encoding="utf-8") as f_seen:
            json.dump({"gerado_em": now_brasilia().isoformat(), "versao": DASHBOARD_BUILD_VERSION, "seen": seen_union}, f_seen, ensure_ascii=False, indent=2)
        print(f"💾 Clientes sem movimento seen atualizado: {seen_path} ({len(seen_union)} chave(s))")
    except Exception as e:
        print(f"⚠️ Não consegui salvar clientes_sem_movimento_seen.json: {e}")

    clientes_sem_movimento_meta_py = {
        "modo":"acionaveis_novos_pendentes_retorno_10d",
        "base_total":len(final_com_key),
        "acionaveis_total":len(actionable),
        "novos_total":novos_count,
        "pendentes_sem_envio_total":pendentes_count,
        "retorno_10d_total":retorno_count,
        "aguardando_reenvio_total":aguardando_count,
        "seen_total":len(seen_union),
        "cooldown_dias":cooldown_dias,
        "gerado_em":now_brasilia().isoformat()
    }
    try:
        with open(base_path, "w", encoding="utf-8") as fbase:
            json.dump({"gerado_em": now_brasilia().isoformat(), "versao": DASHBOARD_BUILD_VERSION, "modo": "base_completa", "clientes": final_com_key}, fbase, ensure_ascii=False, indent=2)
        print(f"💾 Clientes sem movimento BASE completa salva: {base_path} ({len(final_com_key)} cliente(s))")
    except Exception as e:
        print(f"⚠️ Não consegui salvar clientes_sem_movimento_base.json: {e}")
    try:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"gerado_em": now_brasilia().isoformat(), "versao": DASHBOARD_BUILD_VERSION, **clientes_sem_movimento_meta_py, "clientes": actionable}, f, ensure_ascii=False, indent=2)
        print(f"💾 Clientes sem movimento JSON atualizado: {json_path} ({len(actionable)} acionável(is); base monitorada {len(final_com_key)}))")
    except Exception as e:
        print(f"⚠️ Não consegui salvar clientes_sem_movimento.json: {e}")
    pf = {}
    for r in actionable:
        pf[r.get("filial") or "?"] = pf.get(r.get("filial") or "?", 0) + 1
    print("🧡 Clientes sem movimento carregados no dashboard (acionáveis): " + (" · ".join(f"{k}: {v}" for k, v in sorted(pf.items())) or "0"))
    return actionable

def relatorio_duplicidades_carteira_py():
    buckets = []
    try:
        for vend_nome, data in (clientes_por_vend_js or {}).items():
            for fx in ["grave", "alerta", "atencao"]:
                for r in (data or {}).get(fx, []) or []:
                    filial = str(r.get("filial") or "")
                    buckets.append((cobranca_row_key_py(r), "vendedor", vend_nome, filial, fx, r))
    except Exception:
        pass
    try:
        for login, data in (clientes_crediarista_js or {}).items():
            for fx in ["grave", "alerta", "atencao"]:
                for r in (data or {}).get(fx, []) or []:
                    buckets.append((cobranca_row_key_py(r), "crediarista", login, str(r.get("filial") or ""), fx, r))
    except Exception:
        pass
    try:
        for fx in ["grave", "alerta", "atencao"]:
            for r in (clientes_terceiro_js or {}).get(fx, []) or []:
                buckets.append((cobranca_row_key_py(r), "terceiro", "Cobrança10", str(r.get("filial") or "FTER"), fx, r))
    except Exception:
        pass
    mp = {}
    for k, tipo, nome, filial, fx, r in buckets:
        if not k or k.count("|") < 2:
            continue
        mp.setdefault(k, []).append({"tipo": tipo, "responsavel": nome, "filial": filial, "faixa": fx, "cliente": r.get("cliente") or r.get("nome"), "titulo": r.get("titulo"), "parcela": r.get("parcela"), "vencimento": r.get("vencimento"), "pendente": r.get("pendente")})
    conflitos = []

    def _is_conflito_real_v81(arr):
        """
        V8.1: o dashboard preserva Vendedor x Crediarista para não zerar a lista dos vendedores.
        Essa duplicidade é operacionalmente permitida e não deve aparecer como alerta vermelho.

        Continua sendo conflito real:
        - mesmo título em mais de um vendedor;
        - mesmo título em mais de um crediarista;
        - qualquer título que ainda duplicar com Cobrança10/terceiro;
        - qualquer tipo desconhecido duplicado.
        """
        tipos = [str(a.get("tipo") or "").lower() for a in (arr or [])]
        vend = {(a.get("responsavel"), a.get("filial")) for a in (arr or []) if str(a.get("tipo") or "").lower() == "vendedor"}
        cred = {(a.get("responsavel"), a.get("filial")) for a in (arr or []) if str(a.get("tipo") or "").lower() == "crediarista"}
        outros = {t for t in tipos if t not in {"vendedor", "crediarista"}}
        if outros:
            return True
        if len(vend) > 1:
            return True
        if len(cred) > 1:
            return True
        # Um vendedor + um crediarista é permitido na V8.1.
        return False

    ignorados_vendedor_crediarista = 0
    for k, arr in mp.items():
        responsaveis = {(a.get("tipo"), a.get("responsavel"), a.get("filial")) for a in arr}
        if len(responsaveis) > 1:
            if _is_conflito_real_v81(arr):
                conflitos.append({"key": k, "qtd": len(arr), "responsaveis": arr})
            else:
                ignorados_vendedor_crediarista += 1
    conflitos.sort(key=lambda x: x.get("qtd",0), reverse=True)
    out = {"gerado_em": now_brasilia().isoformat(), "versao": "V8.1", "regra": "Vendedor x crediarista preservado e ignorado no alerta; conflitos reais continuam sendo exibidos", "total_conflitos": len(conflitos), "ignorados_vendedor_crediarista": ignorados_vendedor_crediarista, "conflitos": conflitos[:300]}
    try:
        with open(os.path.join(pasta, "relatorio_duplicidades_carteira.json"), "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ Não consegui salvar relatório de duplicidades: {e}")
    if conflitos:
        print(f"🚨 Duplicidades reais de carteira encontradas: {len(conflitos)}. Veja relatorio_duplicidades_carteira.json")
    else:
        print("✅ Check anti-duplicidade da carteira: nenhum conflito entre responsáveis.")
    return out

baixar_clientes_sem_movimento_selenium()
baixar_aniversariantes_selenium()
clientes_sem_movimento_meta_py = {"modo":"somente_novos", "base_total":0, "novos_total":0, "seen_total":0}
clientes_sem_movimento_js = carregar_clientes_sem_movimento_local()
aniversariantes_js = carregar_aniversariantes_local()
duplicidades_carteira_js = relatorio_duplicidades_carteira_py()

js_todos    = json.dumps(todos_js,            ensure_ascii=False)
js_filiais  = json.dumps(filiais_js_ordered,  ensure_ascii=False)
# clientes_por_vend_js já construído na serialização acima

# ── Recebimentos detalhados por faixa por vendedor (para relatório 💰) ──
# Mostra títulos pagos após data_corte_parse, por faixa de vencimento
# Só vendedores ativos; inativos/FDEP aparecem com tag no nome
js_recebimentos = json.dumps(recebimentos_det_js, ensure_ascii=False)
js_recebimentos_terceiro = json.dumps(recebimentos_terceiro_js, ensure_ascii=False)
js_clientes      = json.dumps(clientes_js,           ensure_ascii=False)
js_clientes_vend = json.dumps(clientes_por_vend_js,  ensure_ascii=False)
js_clientes_terceiro = json.dumps(clientes_terceiro_js, ensure_ascii=False)
js_clientes_crediarista = json.dumps(clientes_crediarista_js, ensure_ascii=False)
js_recebimentos_crediarista = json.dumps(recebimentos_crediarista_js, ensure_ascii=False)
js_crediaristas_map = json.dumps(CREDIARISTAS_CONFIG, ensure_ascii=False)
js_destaque = json.dumps(destaque_semana or {}, ensure_ascii=False)
js_hist_dash = json.dumps(hist_dash, ensure_ascii=False)
js_quitados_180 = json.dumps((quitados_180_info.get('dados') or {}).get('quitados', []), ensure_ascii=False)
js_hist_recebimentos_mensais = json.dumps(HIST_RECEBIMENTOS_MENSAIS, ensure_ascii=False)
js_clientes_sem_movimento = json.dumps(clientes_sem_movimento_js, ensure_ascii=False)
js_clientes_sem_movimento_meta = json.dumps(clientes_sem_movimento_meta_py, ensure_ascii=False)
js_aniversariantes = json.dumps(aniversariantes_js, ensure_ascii=False)
js_duplicidades_carteira = json.dumps(duplicidades_carteira_js, ensure_ascii=False)

total_dash_p  = round(total_final_p,  2)
total_dash_pg = round(total_final_pg, 2)



# =========================================
# 🔥 HTML COMPLETO v2 — 4 GRÁFICOS META
# =========================================
import re as _re, base64 as _b64

_laranjito = "https://moveisdolar.com.br/colaborador/mascote%20feliz1.png"
_logo      = "https://moveisdolar.com.br/colaborador/Captura%20de%20tela%202026-04-13%20142059.jpg"


template = r"""
<!DOCTYPE html>
<html lang="pt-br">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dashboard - Lojas MDL</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;0,9..40,800;1,9..40,400&family=DM+Mono:wght@400;500&display=swap');
:root{
  --bg-base:#0d0f14;--bg-surface:#13161e;--bg-elevated:#1a1e2a;--bg-hover:#222636;
  --glass:rgba(255,255,255,.04);--glass-border:rgba(255,255,255,.09);--glass-hover:rgba(255,255,255,.07);
  --amber-200:#fcc772;--amber-300:#f9a832;--amber-400:#f58c10;--amber-500:#e06f05;
  --red-400:#f05252;--red-600:#c72a2a;--green-400:#31c48d;--blue-400:#60a5fa;
  --orange-400:#fb923c;--yellow-400:#fbbf24;
  --text-primary:#f0f2f8;--text-secondary:#8b93a9;--text-muted:#4e5669;--text-accent:#f9a832;
  --radius-sm:8px;--radius-md:14px;--radius-lg:20px;--radius-xl:28px;
  --shadow-md:0 8px 28px rgba(0,0,0,.45);--shadow-lg:0 20px 60px rgba(0,0,0,.55);
  --shadow-amber:0 8px 32px rgba(245,140,16,.20);
  --transition:.18s cubic-bezier(.4,0,.2,1);--transition-slow:.32s cubic-bezier(.4,0,.2,1);
  /* legacy aliases */
  --bg:#0d0f14;--bg2:#0d0f14;--card:rgba(255,255,255,.04);--stroke:rgba(255,255,255,.09);
  --text:#f0f2f8;--muted:#8b93a9;--blue:#60a5fa;--blue2:#93c5fd;
  --green:#31c48d;--green2:#6ee7b7;--red:#f05252;--red2:#fca5a5;
  --orange:#fb923c;--orange2:#fdba74;--yellow:#fbbf24;--yellow2:#fde68a;
  --shadow:0 8px 28px rgba(0,0,0,.45);--radius:20px;
  font-family:'DM Sans',system-ui,sans-serif;font-size:14px;line-height:1.5;
  color:var(--text-primary);background:var(--bg-base);
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{min-height:100vh;background:radial-gradient(ellipse 80% 50% at 10% -10%,rgba(245,140,16,.08) 0%,transparent 60%),radial-gradient(ellipse 60% 40% at 90% 110%,rgba(96,165,250,.06) 0%,transparent 60%),var(--bg-base)}
::-webkit-scrollbar{width:6px;height:6px}::-webkit-scrollbar-track{background:transparent}::-webkit-scrollbar-thumb{background:rgba(255,255,255,.12);border-radius:99px}
.app-shell{max-width:1560px;margin:0 auto;padding:20px 24px 100px}
.glass{background:var(--glass);backdrop-filter:blur(20px) saturate(1.4);-webkit-backdrop-filter:blur(20px) saturate(1.4);border:1px solid var(--glass-border);box-shadow:var(--shadow-md)}
.glass:hover{border-color:rgba(255,255,255,.14)}
.header{display:flex;align-items:center;justify-content:space-between;padding:18px 24px;border-radius:var(--radius-xl);margin-bottom:24px;gap:16px;flex-wrap:wrap;background:linear-gradient(135deg,rgba(245,140,16,.12) 0%,rgba(249,168,50,.06) 40%,transparent 70%),var(--glass);border:1px solid rgba(245,140,16,.2);box-shadow:var(--shadow-amber),var(--shadow-md);position:relative;overflow:hidden}
.header::before{content:'';position:absolute;left:-60px;top:-60px;width:220px;height:220px;border-radius:50%;background:radial-gradient(circle,rgba(245,140,16,.14) 0%,transparent 70%);pointer-events:none}
.brand{display:flex;align-items:center;gap:16px;position:relative;z-index:1}
.brand img{width:52px;height:52px;border-radius:var(--radius-md);object-fit:contain;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.12);box-shadow:0 0 0 3px rgba(245,140,16,.18)}
.brand h1{font-size:26px;font-weight:800;letter-spacing:-.03em;background:linear-gradient(135deg,var(--amber-200) 0%,var(--amber-400) 50%,var(--amber-300) 100%);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.sub{font-size:12px;color:#dbe4ff;margin-top:3px}
.header-actions{display:flex;align-items:center;gap:10px;flex-wrap:wrap;position:relative;z-index:1}
.btn{display:inline-flex;align-items:center;gap:7px;padding:10px 16px;border-radius:var(--radius-md);border:1px solid transparent;font-family:inherit;font-size:13px;font-weight:600;cursor:pointer;transition:var(--transition);text-decoration:none;white-space:nowrap;line-height:1;appearance:none}
.btn:hover{transform:translateY(-1px)}.btn:active{transform:translateY(0) scale(.98)}
.btn.primary{background:linear-gradient(135deg,var(--amber-400) 0%,var(--amber-500) 100%);color:#1a0d00;box-shadow:0 4px 16px rgba(245,140,16,.3);border-color:rgba(255,255,255,.15)}
.btn.primary:hover{box-shadow:0 6px 24px rgba(245,140,16,.45)}
.btn.soft{background:rgba(255,255,255,.06);color:var(--text-primary);border-color:var(--glass-border)}
.btn.soft:hover{background:rgba(255,255,255,.09)}
.btn.ghost{background:transparent;color:var(--text-secondary);border-color:transparent}
.btn.ghost:hover{color:var(--text-primary);background:var(--glass)}
.btn.danger{background:rgba(240,82,82,.12);color:var(--red-400);border-color:rgba(240,82,82,.2)}
.btn.danger:hover{background:rgba(240,82,82,.2)}
.btn.wa{background:linear-gradient(135deg,#25D366,#1a9e4a);color:#fff;font-weight:700;box-shadow:0 4px 14px rgba(37,211,102,.25)}
.pulse-alert{animation:hotPulse 1.1s ease-in-out infinite}@keyframes hotPulse{0%,100%{box-shadow:0 0 0 rgba(245,140,16,0)}50%{box-shadow:0 0 22px rgba(245,140,16,.75);transform:translateY(-1px)}}
.badge{display:inline-flex;align-items:center;gap:6px;padding:8px 14px;border-radius:99px;font-size:12px;font-weight:700;letter-spacing:.02em;background:rgba(249,168,50,.1);color:var(--amber-300);border:1px solid rgba(249,168,50,.2)}
.tabs{display:flex;gap:6px;flex-wrap:wrap;padding:6px;background:var(--bg-surface);border-radius:var(--radius-lg);border:1px solid var(--glass-border);margin-bottom:20px}
.tab{padding:10px 16px;border-radius:var(--radius-md);border:none;background:transparent;color:var(--text-secondary);font-family:inherit;font-size:13px;font-weight:600;cursor:pointer;transition:var(--transition)}
.tab:hover{color:var(--text-primary);background:var(--glass)}
.tab.active{background:linear-gradient(135deg,var(--amber-400),var(--amber-500));color:#1a0d00;box-shadow:0 4px 16px rgba(245,140,16,.35)}
.filters{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:20px}
.pill{padding:7px 14px;border-radius:99px;border:1px solid var(--glass-border);background:transparent;color:var(--text-secondary);font-family:inherit;font-size:12px;font-weight:600;cursor:pointer;transition:var(--transition)}
.pill:hover{color:var(--text-primary);border-color:rgba(255,255,255,.2)}
.pill.active{background:rgba(249,168,50,.12);color:var(--amber-300);border-color:rgba(249,168,50,.3)}
.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:24px}
.kpi{padding:20px;border-radius:var(--radius-lg);position:relative;overflow:hidden;transition:var(--transition)}
.kpi:hover{transform:translateY(-2px)}
.kpi::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:var(--accent,var(--amber-400));border-radius:99px 99px 0 0}
.kpi .label{font-size:11px;font-weight:800;letter-spacing:.1em;text-transform:uppercase;color:#dbe4ff;margin-bottom:6px}
.kpi .value{font-size:24px;font-weight:800;letter-spacing:-.03em;color:#ffffff;line-height:1}
.kpi .subline{font-size:12px;color:var(--text-secondary);margin-top:8px}
.section-head{display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:16px;flex-wrap:wrap}
.section-head h2{font-size:20px;font-weight:800;letter-spacing:-.025em;color:var(--text-primary)}
.section-head .hint{font-size:12px;color:var(--text-secondary);margin-top:4px}
.grid-cards{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:14px}
.senhas-table-wrap{overflow:auto;border:1px solid rgba(148,163,184,.20);border-radius:18px;background:rgba(15,23,42,.35)}
.senhas-table{width:100%;border-collapse:separate;border-spacing:0;min-width:980px}
.senhas-table th{position:sticky;top:0;background:#111827;color:#aeb7ca;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.08em;padding:12px;border-bottom:1px solid rgba(148,163,184,.22);z-index:1}
.senhas-table td{padding:10px 12px;border-bottom:1px solid rgba(148,163,184,.12);vertical-align:middle}
.senhas-table tr:hover td{background:rgba(255,255,255,.035)}
.senha-view-row{display:flex;align-items:center;gap:6px;min-width:260px}
.senha-view-row input{min-width:160px;max-width:220px;background:#060a12;border:1px solid rgba(148,163,184,.22);border-radius:10px;color:#e5e7eb;padding:9px 10px;font-family:DM Mono,monospace}
.senha-nova-row{display:flex;align-items:center;gap:8px;min-width:290px}
.senha-nova-row input{width:180px;background:#060a12;border:1px solid rgba(148,163,184,.22);border-radius:10px;color:#e5e7eb;padding:9px 10px}
.btn-xs{padding:8px 10px!important;border-radius:10px!important;font-size:12px!important}
.status-dot{display:inline-flex;gap:6px;align-items:center;white-space:nowrap}.status-dot i{width:8px;height:8px;border-radius:999px;display:inline-block}

.card{padding:20px;border-radius:var(--radius-lg);cursor:pointer;position:relative;overflow:hidden;transition:var(--transition-slow);border:1px solid var(--glass-border);background:var(--glass)}
.card:hover{transform:translateY(-4px) scale(1.01);border-color:rgba(255,255,255,.18);box-shadow:0 20px 60px rgba(0,0,0,.5),0 0 0 1px rgba(249,168,50,.12)}
.card::before{content:'';position:absolute;inset:0;background:linear-gradient(105deg,transparent 40%,rgba(255,255,255,.04) 50%,transparent 60%);background-size:200% 100%;background-position:200% 0;transition:background-position .6s ease;pointer-events:none}
.card:hover::before{background-position:-200% 0}
.card::after{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,var(--card-accent,transparent),rgba(255,255,255,0));border-radius:2px 2px 0 0}
.card .title{font-size:14px;font-weight:700;color:var(--text-primary);line-height:1.3;min-height:auto;margin-bottom:14px}
.card-hit{--card-accent:var(--amber-400);box-shadow:0 0 0 1px rgba(249,168,50,.15),var(--shadow-md);animation:highlightPulse 2.2s ease-in-out infinite}
.card-low-red{--card-accent:var(--red-400);box-shadow:0 0 0 1px rgba(240,82,82,.12),var(--shadow-md)}
.card-low-orange{--card-accent:var(--orange-400);box-shadow:0 0 0 1px rgba(251,146,60,.12),var(--shadow-md)}
.card .numbers{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:10px}
.stat-box{padding:10px 12px;border-radius:var(--radius-sm);background:rgba(0,0,0,.2);border:1px solid rgba(255,255,255,.05)}
.stat-box .mini{font-size:10px;color:var(--text-muted);font-weight:600;text-transform:uppercase;letter-spacing:.08em}
.stat-box .big{font-size:14px;font-weight:800;margin-top:4px;line-height:1;letter-spacing:-.02em;white-space:nowrap;overflow:visible}
.meta-row{display:flex;gap:6px;flex-wrap:wrap;margin-top:10px}
.mini-chip{font-size:11px;font-weight:700;padding:5px 9px;border-radius:var(--radius-sm);background:rgba(255,255,255,.05);border:1px solid var(--glass-border);color:#eaf1ff;display:flex;align-items:center;gap:5px}
.dot{width:6px;height:6px;border-radius:50%;flex-shrink:0;display:inline-block}
.legend-inline{display:flex;gap:12px;flex-wrap:wrap;font-size:12px;color:var(--text-secondary);margin:10px 0 0}
.legend-inline span{display:inline-flex;align-items:center;gap:6px;font-weight:700}
.mascot-status{display:flex;align-items:center;gap:8px;margin-top:10px;padding:9px 11px;border-radius:var(--radius-md);background:rgba(255,255,255,.03);border:1px solid var(--glass-border)}
.mascot-status img{width:32px;height:32px;border-radius:var(--radius-sm);object-fit:cover;animation:mascotPulse 2s ease-in-out infinite}
.mascot-status strong{font-size:12px;color:var(--text-primary);font-weight:700}
.card-sales{margin-top:10px;background:rgba(249,168,50,.04);border:1px solid rgba(249,168,50,.12);border-radius:var(--radius-md);padding:10px}
.card-sales-title{font-size:10px;color:var(--amber-400);font-weight:700;text-transform:uppercase;letter-spacing:.1em;margin-bottom:8px;display:flex;align-items:center;gap:6px}
.card-sales-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:6px}
.sales-mini{padding:7px;border-radius:var(--radius-sm);background:rgba(0,0,0,.25);border:1px solid rgba(255,255,255,.05);text-align:center}
.sales-mini .k{font-size:9px;color:var(--text-muted);font-weight:700;text-transform:uppercase;letter-spacing:.06em;line-height:1.1}
.sales-mini .v{font-size:11px;font-weight:800;color:var(--text-primary);margin-top:4px;line-height:1.1;white-space:nowrap;overflow:visible}
.sales-mini.good{border-color:rgba(49,196,141,.2)}.sales-mini.good .v{color:var(--green-400)}
.sales-mini.warn{border-color:rgba(251,146,60,.2)}.sales-mini.warn .v{color:var(--orange-400)}
.sales-mini.gold{background:linear-gradient(135deg,rgba(249,168,50,.1),rgba(249,168,50,.03));border-color:rgba(249,168,50,.25)}.sales-mini.gold .v{color:var(--amber-300)}
.sales-mini.low{border-color:rgba(240,82,82,.2)}.sales-mini.low .v{color:var(--red-400)}
.sales-mini.realizado{border-color:rgba(49,196,141,.2);animation:realPulse 1.9s ease-in-out infinite}
.sales-mini.meta{border-color:rgba(96,165,250,.15)}
.sales-mini .laranjito-mini{display:block;width:22px;height:22px;border-radius:6px;object-fit:cover;margin:0 auto 3px}
.sales-mini.wrap .v{white-space:normal;overflow-wrap:anywhere;word-break:break-word;line-height:1.05}
.sales-progress{margin-top:6px;height:6px;border-radius:99px;background:rgba(255,255,255,.06);overflow:hidden}
.sales-progress>span{display:block;height:100%;border-radius:99px;background:linear-gradient(90deg,var(--amber-500),var(--amber-300))}
.sales-progress-label{margin-top:4px;font-size:9px;color:var(--text-muted);font-weight:700;text-align:right}
.big-chart-card{padding:22px;border-radius:var(--radius-xl);margin-top:24px}
.groupbars{display:flex;align-items:flex-end;gap:12px;overflow:auto;padding:14px 6px 6px}
.group{min-width:100px}
.group .glabel{text-align:center;font-size:11px;color:var(--text-secondary);font-weight:700;margin-top:8px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.group .bars{height:280px;display:flex;align-items:flex-end;justify-content:center;gap:6px;padding:0 4px;border-left:1px solid rgba(255,255,255,.06)}
.bar{width:20px;border-radius:14px 14px 8px 8px;position:relative;overflow:hidden;box-shadow:inset 0 2px 0 rgba(255,255,255,.35),0 8px 16px rgba(0,0,0,.35)}
.bar::before{content:'';position:absolute;inset:0;background:linear-gradient(180deg,rgba(255,255,255,.45),rgba(255,255,255,.1) 30%,rgba(255,255,255,0) 60%);pointer-events:none}
.axis{display:flex;justify-content:space-between;margin:10px 10px 0;color:var(--text-muted);font-size:11px;font-weight:600}
.detail-screen{margin-top:8px}
.detail-top{display:grid;grid-template-columns:1.1fr .9fr;gap:16px}
.back-row{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:20px;flex-wrap:wrap}
.back-row h2{font-size:24px;font-weight:800;letter-spacing:-.03em}
.note{font-size:12px;color:var(--text-muted)}
.panel{padding:22px;border-radius:var(--radius-xl)}
.panel h3{font-size:16px;font-weight:800;letter-spacing:-.02em;margin-bottom:16px}
.panel h3 .note{font-size:11px;font-weight:500;color:var(--text-muted)}
.mega-progress{display:grid;grid-template-columns:220px 1fr;gap:16px;align-items:center}
.ring-wrap{text-align:center}
.piggy-bank{position:relative;width:190px;min-height:258px;margin:0 auto;display:flex;flex-direction:column;align-items:center;justify-content:flex-start}
.piggy-glow{position:absolute;top:12px;left:50%;transform:translateX(-50%);width:178px;height:178px;border-radius:50%;background:radial-gradient(circle at 50% 35%,rgba(249,168,50,.15),rgba(249,168,50,0) 68%)}
.piggy-shell{position:relative;width:178px;height:178px;border-radius:50%;overflow:hidden;background:conic-gradient(from -90deg,var(--amber-400) calc(var(--pct)*1%),rgba(30,35,50,.9) 0);box-shadow:inset 0 0 0 10px rgba(255,255,255,.08),0 20px 40px rgba(0,0,0,.5)}
.piggy-shell::before{content:'';position:absolute;inset:14px;border-radius:50%;background:linear-gradient(180deg,rgba(255,255,255,.08),rgba(255,255,255,.02));box-shadow:inset 0 2px 0 rgba(255,255,255,.12);z-index:1}
.piggy-slot{position:absolute;left:50%;top:13px;transform:translateX(-50%);width:68px;height:14px;border-radius:99px;background:rgba(0,0,0,.5);border:1px solid rgba(249,168,50,.2);z-index:5}
.piggy-fill{position:absolute;inset:0;z-index:3;pointer-events:none}
.piggy-falling{position:absolute;inset:0;z-index:4;pointer-events:none;overflow:hidden}
.piggy-glass{position:absolute;inset:14px;border-radius:50%;background:radial-gradient(circle at 32% 18%,rgba(255,255,255,.2),rgba(255,255,255,0) 30%);pointer-events:none;z-index:6}
.piggy-label{position:relative;margin-top:14px;text-align:center;z-index:7}
.piggy-label .pct{font-size:44px;font-weight:800;line-height:1;color:var(--text-primary);letter-spacing:-.05em}
.piggy-label .small{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:var(--text-muted);margin-top:6px}
.mdl-coin{position:absolute;display:block;width:30px;height:30px;background-image:url('https://moveisdolar.com.br/colaborador/coin%20png.png');background-size:contain;background-position:center;background-repeat:no-repeat;filter:drop-shadow(0 4px 8px rgba(186,132,14,.3))}
.coin-fill{animation:coinBob 3.2s ease-in-out infinite alternate}
.coin-drop{left:74px;top:16px;opacity:0;animation:coinDropIn 2.8s cubic-bezier(.25,.65,.25,1) infinite}
@keyframes coinDropIn{0%{transform:translate(0,-12px) scale(.55) rotate(0deg);opacity:0}10%{opacity:1}35%{transform:translate(0,4px) scale(.78) rotate(140deg);opacity:1}100%{transform:translate(calc(var(--tx) - 74px),calc(var(--ty) - 16px)) scale(1) rotate(360deg);opacity:0}}
@keyframes coinBob{0%{transform:translateY(0)}100%{transform:translateY(-3px)}}
.metrics-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px}
.metric{padding:14px;border-radius:var(--radius-md);background:rgba(0,0,0,.2);border:1px solid var(--glass-border)}
.metric .k{font-size:11px;color:var(--text-muted);font-weight:600;text-transform:uppercase;letter-spacing:.08em}
.metric .v{font-size:20px;font-weight:800;margin-top:6px;letter-spacing:-.02em}
.meta-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-top:14px}
.meta-card{padding:14px;border-radius:var(--radius-md);background:rgba(0,0,0,.2);border:1px solid var(--glass-border);text-align:center;cursor:pointer;transition:var(--transition)}
.meta-card:hover{border-color:rgba(249,168,50,.25);background:rgba(249,168,50,.04)}
.meta-card .meta-title{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:var(--text-muted)}
.meta-card .meta-main{font-size:30px;font-weight:800;margin-top:8px;letter-spacing:-.04em}
.meta-card .meta-sub{font-size:11px;color:var(--text-secondary);margin-top:6px}
.accordion{border-radius:var(--radius-lg);overflow:hidden;border:1px solid var(--glass-border);background:var(--glass);margin-top:14px}
.accordion+.accordion{margin-top:14px}
.acc-head{padding:16px 20px;display:flex;justify-content:space-between;align-items:center;cursor:pointer;font-weight:700;font-size:14px;color:var(--text-primary);user-select:none;transition:var(--transition)}
.acc-head:hover{background:rgba(255,255,255,.04)}
.acc-head span:last-child{width:20px;height:20px;border-radius:50%;background:rgba(255,255,255,.06);display:flex;align-items:center;justify-content:center;font-size:11px;transition:transform var(--transition)}
.accordion.open .acc-head span:last-child{transform:rotate(180deg)}
.acc-body{display:none;padding:0 16px 16px}
.accordion.open .acc-body{display:block}
.faixa-block{margin-top:12px}
.faixa-title{padding:10px 14px;border-radius:var(--radius-md);font-weight:700;font-size:13px;display:flex;justify-content:space-between;align-items:center;margin-bottom:10px}
.faixa-title.grave{background:rgba(240,82,82,.12);color:var(--red-400);border:1px solid rgba(240,82,82,.2)}
.faixa-title.alerta{background:rgba(251,146,60,.12);color:var(--orange-400);border:1px solid rgba(251,146,60,.2)}
.faixa-title.atencao{background:rgba(251,191,36,.12);color:var(--yellow-400);border:1px solid rgba(251,191,36,.2)}
.tableish{display:grid;gap:8px;margin-top:10px}
.row-item{padding:12px;border-radius:var(--radius-md);background:rgba(255,255,255,.02);border:1px solid var(--glass-border);transition:var(--transition)}
.row-item:hover{background:rgba(255,255,255,.04);border-color:rgba(255,255,255,.12)}
.row-top{display:grid;grid-template-columns:1.3fr repeat(5,.7fr);gap:10px;align-items:start}
.row-top .name{font-weight:800;font-size:13px}
.small{font-size:12px;color:#d8e2ff}.muted{color:#bfc9e6}
.restr-ok{display:inline-flex;align-items:center;gap:4px;padding:3px 8px;border-radius:99px;background:rgba(49,196,141,.1);color:var(--green-400);border:1px solid rgba(49,196,141,.2);font-weight:700;font-size:11px}
.avalista-alert{display:inline-flex;align-items:center;gap:4px;padding:3px 8px;border-radius:99px;background:rgba(251,146,60,.1);color:var(--orange-400);border:1px solid rgba(251,146,60,.2);font-weight:700;font-size:11px;animation:warmPulse 1.8s ease-in-out infinite}
@keyframes warmPulse{0%,100%{opacity:1}50%{opacity:.7}}
.bonus-box{padding:16px;border-radius:var(--radius-xl);background:var(--glass);border:1px solid var(--glass-border)}
.bonus-box h4{margin:0 0 10px;font-size:15px;font-weight:800;color:var(--text-primary)}
.bonus-list{display:grid;gap:8px}
.bonus-item{display:flex;align-items:center;justify-content:space-between;padding:11px 14px;border-radius:var(--radius-md);background:rgba(0,0,0,.18);border:1px solid var(--glass-border);transition:var(--transition)}
.bonus-item .left{display:flex;align-items:center;gap:8px;font-weight:800}
.bonus-item.active{border-color:rgba(249,168,50,.35);background:rgba(249,168,50,.07);animation:glowPulse 2s ease-in-out infinite}
@keyframes glowPulse{0%,100%{box-shadow:0 0 20px rgba(249,168,50,.1)}50%{box-shadow:0 0 32px rgba(249,168,50,.22)}}
.sales-panel{background:linear-gradient(135deg,rgba(249,168,50,.06),rgba(249,168,50,.02));border:1px solid rgba(249,168,50,.15)}
.sales-panel h3{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.sales-note{font-size:11px;color:var(--text-muted);font-weight:700}
.sales-stack{display:flex;flex-direction:column;gap:12px}
.sales-card{background:rgba(0,0,0,.2);border:1px solid rgba(249,168,50,.1);border-radius:var(--radius-md);padding:12px}
.sales-card h4{margin:0 0 8px;font-size:14px;font-weight:700;color:var(--amber-300)}
.sales-sub{font-size:10px;color:var(--amber-400);font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px}
.sales-list{display:flex;flex-direction:column;gap:8px;max-height:240px;overflow:auto;padding-right:4px}
.sales-row{background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.06);border-radius:var(--radius-sm);padding:10px}
.sales-row-title{font-size:12px;font-weight:800;color:var(--text-primary);margin-bottom:8px}
.sales-metrics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:6px}
.sales-empty{padding:14px;border-radius:var(--radius-md);background:rgba(255,255,255,.02);border:1px dashed rgba(249,168,50,.15);color:var(--text-muted);font-weight:700;text-align:center}
.rent-badge{display:inline-flex;align-items:center;gap:8px;padding:7px 12px;border-radius:99px;background:rgba(49,196,141,.08);border:1px solid rgba(49,196,141,.18);color:var(--green-400);font-size:12px;font-weight:700}
.rent-badge strong{font-size:15px;color:#a7f3d0}
.commission-card{background:linear-gradient(135deg,rgba(249,168,50,.06),rgba(249,168,50,.02));border:1px solid rgba(249,168,50,.15);border-radius:var(--radius-xl);padding:20px}
.commission-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px}
.commission-item{background:rgba(0,0,0,.2);border:1px solid var(--glass-border);border-radius:var(--radius-md);padding:12px;transition:var(--transition)}
.commission-item.unlocked{border-color:rgba(249,168,50,.2)}
.commission-item .k{font-size:10px;color:#dbe4ff;font-weight:700;text-transform:uppercase;letter-spacing:.08em}
.commission-item .v{font-size:18px;font-weight:800;color:var(--amber-300);margin-top:6px;letter-spacing:-.02em}
.commission-item.locked{opacity:.4;filter:blur(.5px);position:relative}
.commission-item.locked::after{content:'🔒 Bloqueado';position:absolute;inset:0;display:flex;align-items:center;justify-content:center;background:rgba(13,15,20,.75);font-size:11px;font-weight:700;color:var(--text-muted);border-radius:var(--radius-md)}
.commission-item.total-final{background:linear-gradient(135deg,rgba(249,168,50,.1),rgba(249,168,50,.03));border-color:rgba(249,168,50,.3);box-shadow:0 0 20px rgba(249,168,50,.08)}
.commission-item.total-final .v{color:var(--amber-200);font-size:22px}
.commission-note{margin-top:10px;font-size:11px;color:var(--text-muted);font-weight:600}
.meta-hit-banner{display:flex;align-items:center;gap:12px;padding:12px 14px;border-radius:var(--radius-md);background:rgba(249,168,50,.08);border:1px solid rgba(249,168,50,.2);margin-bottom:14px}
.meta-hit-banner img{width:34px;height:34px;border-radius:var(--radius-sm);object-fit:cover}
.total-locked{animation:lockPulse 1.6s ease-in-out infinite}
@keyframes lockPulse{0%,100%{opacity:.8}50%{opacity:1}}
.delta-pill{display:inline-flex;align-items:center;gap:5px;padding:5px 10px;border-radius:99px;font-size:12px;font-weight:700}
.delta-pill.up{background:rgba(49,196,141,.1);color:var(--green-400);border:1px solid rgba(49,196,141,.2)}
.delta-pill.down{background:rgba(240,82,82,.1);color:var(--red-400);border:1px solid rgba(240,82,82,.2)}
.delta-pill .arrow{font-size:14px;line-height:1}
.campaign-banner{padding:18px 20px;border-radius:var(--radius-lg);background:linear-gradient(135deg,rgba(249,168,50,.08),rgba(249,168,50,.03));border:1px solid rgba(249,168,50,.2);box-shadow:0 8px 32px rgba(249,168,50,.1);margin-bottom:16px;animation:campaignPulse 2.4s ease-in-out infinite}
@keyframes campaignPulse{0%,100%{box-shadow:0 8px 32px rgba(249,168,50,.1)}50%{box-shadow:0 8px 40px rgba(249,168,50,.22)}}
.campaign-card{background:linear-gradient(135deg,rgba(249,168,50,.07),rgba(249,168,50,.02));border:1px solid rgba(249,168,50,.15);border-radius:var(--radius-xl);padding:20px}
.campaign-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px}
.campaign-item{background:rgba(0,0,0,.2);border:1px solid var(--glass-border);border-radius:var(--radius-md);padding:12px}
.campaign-item .k{font-size:10px;color:var(--text-muted);font-weight:700;text-transform:uppercase;letter-spacing:.08em}
.campaign-item .v{font-size:18px;font-weight:800;color:var(--amber-300);margin-top:6px}
.campaign-item.locked{opacity:.4}
.highlight-pulse{animation:highlightPulse 2.2s ease-in-out infinite}
@keyframes highlightPulse{0%,100%{box-shadow:0 0 0 0 rgba(249,168,50,.0)}50%{box-shadow:0 0 0 6px rgba(249,168,50,.08)}}
.msg-banner{padding:18px 20px;border-radius:var(--radius-lg);background:rgba(249,168,50,.05);border:1px solid rgba(249,168,50,.14);margin-bottom:16px}
.login-wrap{min-height:100vh;display:grid;place-items:center;padding:24px;position:relative;z-index:1}
.login-card{width:min(460px,100%);padding:32px;border-radius:var(--radius-xl)}
.login-card .logo-big{width:76px;height:76px;border-radius:20px;object-fit:contain;background:rgba(255,255,255,.05);margin:0 auto 16px;display:block;border:1px solid rgba(255,255,255,.1);box-shadow:0 0 0 4px rgba(249,168,50,.15),var(--shadow-amber)}
.login-card h2{margin:0;text-align:center;font-size:28px;font-weight:800;letter-spacing:-.03em;background:linear-gradient(135deg,var(--amber-200),var(--amber-400));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.login-form{display:grid;gap:12px;margin-top:22px}
.login-form input{padding:14px 16px;border-radius:var(--radius-md);border:1px solid var(--glass-border);background:rgba(0,0,0,.3);color:var(--text-primary);font-family:inherit;font-size:14px;font-weight:500;outline:none;transition:var(--transition)}
.login-form input:focus{border-color:rgba(249,168,50,.4);box-shadow:0 0 0 3px rgba(249,168,50,.08)}
.login-form input::placeholder{color:var(--text-muted)}
#loginMsg{text-align:center;color:var(--orange-400);font-size:13px;font-weight:600}
.modal{position:fixed;inset:0;background:rgba(0,0,0,.65);backdrop-filter:blur(8px);-webkit-backdrop-filter:blur(8px);display:none;place-items:center;z-index:50;padding:20px}
.modal.show{display:grid}
.modal-card{width:min(420px,100%);padding:24px;border-radius:var(--radius-xl)}
.modal-card h3{font-size:18px;font-weight:800;margin-bottom:6px;color:var(--text-primary)}
.modal-list{display:grid;gap:10px;margin-top:16px}
.pulse-notify{animation:pulseNotify 1.1s ease-in-out infinite}@keyframes pulseNotify{0%,100%{box-shadow:0 0 0 rgba(249,168,50,0)}50%{box-shadow:0 0 26px rgba(249,168,50,.65)}}
.toast{position:fixed;right:20px;bottom:20px;z-index:60;display:grid;gap:10px}
.toast-item{min-width:280px;padding:14px 16px;border-radius:var(--radius-lg);background:var(--bg-elevated);border:1px solid var(--glass-border);box-shadow:var(--shadow-lg);display:flex;gap:12px;align-items:center;animation:slideInToast .25s ease}
@keyframes slideInToast{from{transform:translateX(40px);opacity:0}to{transform:translateX(0);opacity:1}}
.toast-item img{width:40px;height:40px;border-radius:var(--radius-sm);object-fit:cover}
.form-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.form-grid.bonus{grid-template-columns:repeat(2,1fr)}
.input-card{padding:14px;border-radius:var(--radius-md);background:rgba(0,0,0,.2);border:1px solid var(--glass-border)}
.input-card label{display:block;font-size:11px;color:var(--text-muted);font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px}
.input-card input,.input-card select,.input-card textarea{width:100%;padding:10px 12px;border-radius:var(--radius-sm);border:1px solid var(--glass-border);background:rgba(0,0,0,.3);color:var(--text-primary);font-family:inherit;font-size:13px;font-weight:500;outline:none;transition:var(--transition)}
.input-card input:focus,.input-card select:focus{border-color:rgba(249,168,50,.35);box-shadow:0 0 0 2px rgba(249,168,50,.07)}
.input-card select option{background:var(--bg-elevated)}
.logs-list{display:grid;gap:8px;margin-top:14px}
.log-row{padding:14px 16px;border-radius:var(--radius-md);background:rgba(255,255,255,.02);border:1px solid var(--glass-border);display:grid;grid-template-columns:1.4fr 1fr .7fr .8fr auto;gap:12px;align-items:center;transition:var(--transition)}
.log-row:hover{background:rgba(255,255,255,.04)}
.search-row{display:grid;grid-template-columns:1fr 180px 180px 220px;gap:10px;align-items:end}
.msg-card{padding:16px;border-radius:var(--radius-md);background:rgba(255,255,255,.03);border:1px solid var(--glass-border)}
.msg-head{display:flex;justify-content:space-between;gap:10px;align-items:center;margin-bottom:8px;flex-wrap:wrap}
.msg-media img,.msg-media video,.msg-media audio{max-width:100%;border-radius:var(--radius-md);margin-top:10px}
.msg-media.compact{width:min(1766px,100%);max-width:100%;height:547px;display:flex;align-items:center;justify-content:center;overflow:hidden;border-radius:var(--radius-md);margin-top:10px;background:rgba(0,0,0,.3)}
.msg-media.compact img,.msg-media.compact video{width:100%;height:100%;object-fit:contain;border-radius:var(--radius-md);margin-top:0}
.msg-media.compact audio{width:min(860px,96%);margin-top:0}
#msgPreviewBody img,#msgPreviewBody video{max-width:100%;max-height:82vh;object-fit:contain;border-radius:var(--radius-md)}
#msgPreviewBody audio{width:min(980px,100%)}
.comm-table{width:100%;border-collapse:separate;border-spacing:0 6px;font-size:12px}
.comm-table th{font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:var(--text-muted);padding:6px 8px}
.comm-table td{background:rgba(0,0,0,.2);border-top:1px solid var(--glass-border);border-bottom:1px solid var(--glass-border);padding:8px}
.comm-table td:first-child{border-left:1px solid var(--glass-border);border-radius:var(--radius-sm) 0 0 var(--radius-sm)}
.comm-table td:last-child{border-right:1px solid var(--glass-border);border-radius:0 var(--radius-sm) var(--radius-sm) 0}
.comm-table input{width:100%;border:none;background:transparent;text-align:center;font-weight:700;color:var(--text-primary);outline:none;font-size:12px}
.comm-table thead th input{background:rgba(0,0,0,.2);border:1px solid var(--glass-border);border-radius:var(--radius-sm);padding:5px 7px;font-size:10px;font-weight:700;color:var(--text-secondary);min-width:100px}
.comm-scroll{overflow-x:auto;padding-bottom:6px}
.comm-wrap{margin-top:18px}
.comm-subtitle{font-size:13px;font-weight:800;color:var(--amber-300);margin:0 0 8px}
.comm-box{background:rgba(0,0,0,.15);border:1px solid var(--glass-border);border-radius:var(--radius-md);padding:14px;margin-top:12px}
.unread-chip{display:inline-flex;align-items:center;gap:5px;padding:3px 9px;border-radius:99px;background:rgba(240,82,82,.1);color:var(--red-400);font-size:10px;font-weight:700;border:1px solid rgba(240,82,82,.18)}
.read-chip{display:inline-flex;align-items:center;gap:5px;padding:3px 9px;border-radius:99px;background:rgba(49,196,141,.1);color:var(--green-400);font-size:10px;font-weight:700;border:1px solid rgba(49,196,141,.18)}
.entity-card .mascot-status{margin-top:10px}
.meta-layout{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.meta-preview-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:14px;margin-top:16px}
.empty{padding:26px;text-align:center;border-radius:var(--radius-lg);border:1px dashed rgba(255,255,255,.1);color:var(--text-muted);font-weight:700;font-size:13px;background:rgba(255,255,255,.01)}
.hidden{display:none!important}
@keyframes mascotPulse{0%,100%{transform:scale(1)}50%{transform:scale(1.06)}}
@keyframes realPulse{0%,100%{transform:scale(1)}50%{transform:scale(1.02)}}
@keyframes liquidGlow{0%{transform:translateY(0)}100%{transform:translateY(12px)}}
@media(max-width:1100px){.detail-top,.meta-layout,.search-row,.mega-progress{grid-template-columns:1fr}.kpis{grid-template-columns:repeat(2,1fr)}.meta-grid{grid-template-columns:repeat(2,1fr)}.sales-metrics{grid-template-columns:repeat(2,minmax(0,1fr))}.card-sales-grid{grid-template-columns:1fr}}
@media(max-width:760px){.app-shell{padding:14px 14px 80px}.brand h1{font-size:20px}.kpis{grid-template-columns:1fr}.grid-cards,.meta-preview-grid{grid-template-columns:1fr}.form-grid,.form-grid.bonus,.meta-grid,.metrics-grid,.commission-grid,.campaign-grid{grid-template-columns:1fr}.row-top,.log-row{grid-template-columns:1fr}.tabs{gap:4px}.tab{padding:9px 12px;font-size:12px}.group{min-width:88px}.group .bars{height:220px}.bar{width:16px}}

/* ===== AJUSTE VISUAL DOS CARDS KPI - CORES POR GRUPO E ÍCONES MAIORES ===== */
.kpi.card-cobranca{
  background:
    radial-gradient(circle at 12% 12%, rgba(96,165,250,.30), transparent 36%),
    linear-gradient(135deg, rgba(30,64,175,.36), rgba(15,23,42,.90) 64%) !important;
}
.kpi.card-financeiro{
  background:
    radial-gradient(circle at 12% 12%, rgba(74,222,128,.28), transparent 36%),
    linear-gradient(135deg, rgba(22,101,52,.36), rgba(15,23,42,.90) 64%) !important;
}
.kpi.card-venda-dia{
  background:
    radial-gradient(circle at 12% 12%, rgba(250,204,21,.34), transparent 36%),
    linear-gradient(135deg, rgba(161,98,7,.38), rgba(15,23,42,.90) 64%) !important;
}
.kpi .label{
  display:flex;
  align-items:center;
  gap:8px;
}
.kpi .kpi-emoji{
  display:inline-flex;
  align-items:center;
  justify-content:center;
  font-size:25px;
  line-height:1;
  min-width:26px;
  filter:drop-shadow(0 0 8px rgba(255,255,255,.20));
}
.kpi .kpi-label-text{
  line-height:1.1;
}


/* ===== LARANJITO NOS CARDS DE ALERTA ===== */
@keyframes laranjitoPulse {
  0%,100% { transform: scale(1); filter: drop-shadow(0 0 7px rgba(255,147,0,.35)); }
  50% { transform: scale(1.13); filter: drop-shadow(0 0 18px rgba(255,147,0,.75)); }
}
.kpi.kpi-laranjito{
  position:relative;
  overflow:hidden;
  padding-right:104px !important;
}
.kpi .laranjito-card{
  position:absolute;
  right:14px;
  top:50%;
  transform:translateY(-50%);
  width:78px;
  height:78px;
  border-radius:22px;
  display:flex;
  align-items:center;
  justify-content:center;
  font-size:54px;
  animation:laranjitoPulse 1.4s ease-in-out infinite;
  background:radial-gradient(circle at 40% 35%, rgba(255,177,59,.22), rgba(255,122,0,.06) 60%, transparent 72%);
  pointer-events:none;
}
.kpi .laranjito-caption{
  position:absolute;
  right:13px;
  bottom:8px;
  font-size:10px;
  color:#fbbf24;
  opacity:.9;
  font-weight:800;
  letter-spacing:.04em;
}


/* ===== IMAGENS REAIS DO LARANJITO E ÍCONE DE OURO ===== */
.kpi .kpi-img-icon{
  width:30px;
  height:30px;
  object-fit:contain;
  display:inline-flex;
  flex:0 0 auto;
  filter:drop-shadow(0 0 8px rgba(255,255,255,.22));
}
.kpi .laranjito-card{
  position:absolute;
  right:12px;
  top:50%;
  transform:translateY(-50%);
  width:96px;
  height:96px;
  object-fit:contain;
  border-radius:0;
  animation:laranjitoPulse 1.4s ease-in-out infinite;
  background:transparent;
  pointer-events:none;
  z-index:2;
}
.kpi .laranjito-card::after{content:none!important;}
.kpi.kpi-laranjito{
  padding-right:120px !important;
}
.kpi .laranjito-caption{display:none!important;}


/* ===== AJUSTE FINAL LARANJITO PROPORCIONAL SEM FUNDO ===== */
.kpi.kpi-laranjito{
  position:relative;
  overflow:hidden;
  padding-right:92px !important;
}
.kpi .laranjito-card{
  position:absolute !important;
  right:12px !important;
  bottom:8px !important;
  top:auto !important;
  transform:none !important;
  width:74px !important;
  height:74px !important;
  max-width:74px !important;
  max-height:74px !important;
  object-fit:contain !important;
  object-position:center !important;
  border-radius:0 !important;
  background:transparent !important;
  box-shadow:none !important;
  border:0 !important;
  outline:0 !important;
  padding:0 !important;
  margin:0 !important;
  animation:laranjitoPulseImg 1.45s ease-in-out infinite !important;
  pointer-events:none !important;
  z-index:2 !important;
}
@keyframes laranjitoPulseImg{
  0%,100% { transform:scale(1); filter:drop-shadow(0 0 6px rgba(255,147,0,.35)); }
  50% { transform:scale(1.08); filter:drop-shadow(0 0 14px rgba(255,147,0,.75)); }
}
.kpi .laranjito-card::after{content:none!important;}
.kpi .laranjito-caption{display:none!important;}



/* ===== V3.9: corrigir botões de acordeão e mascotes sem sobrepor texto ===== */
.acc-head span:last-child,
.accordion.open .acc-head span:last-child{
  transform:none!important;
  writing-mode:horizontal-tb!important;
}
.acc-head .acc-hint,
.acc-head > span:last-child{
  width:auto!important;
  height:auto!important;
  min-width:auto!important;
  max-width:360px!important;
  padding:6px 12px!important;
  border-radius:999px!important;
  background:rgba(255,255,255,.06)!important;
  display:inline-flex!important;
  align-items:center!important;
  justify-content:center!important;
  white-space:nowrap!important;
  overflow:hidden!important;
  text-overflow:ellipsis!important;
  line-height:1.2!important;
  font-size:12px!important;
  text-align:center!important;
  flex:0 0 auto!important;
}
.kpi.kpi-laranjito{
  padding-right:92px!important;
  min-height:94px!important;
}
.kpi .laranjito-card{
  width:62px!important;
  height:62px!important;
  right:14px!important;
  bottom:12px!important;
  top:auto!important;
  transform:none!important;
  object-fit:contain!important;
  border-radius:14px!important;
  background:transparent!important;
  z-index:0!important;
  opacity:.95!important;
}
.kpi.kpi-laranjito .label,
.kpi.kpi-laranjito .value,
.kpi.kpi-laranjito .subline{
  position:relative!important;
  z-index:2!important;
}
.kpi.kpi-laranjito .subline{
  max-width:calc(100% - 12px)!important;
}
.meta-diaria-empty{
  padding:10px 12px;
  border-radius:14px;
  background:rgba(34,197,94,.055);
  border:1px dashed rgba(34,197,94,.22);
  color:#9fb0c8;
  font-weight:700;
}

/* ===== CONFIGURAÇÃO GLOBAL DE MENSAGEM DE COBRANÇA ===== */
.cobranca-config-grid{
  display:grid;
  grid-template-columns: 1.4fr .9fr;
  gap:14px;
}
.cobranca-template{
  min-height:180px;
  resize:vertical;
  line-height:1.45;
}
.placeholder-list{
  display:grid;
  grid-template-columns:repeat(2,minmax(0,1fr));
  gap:6px;
  margin-top:8px;
}
.placeholder-list code{
  font-size:11px;
  background:rgba(255,255,255,.06);
  border:1px solid rgba(255,255,255,.09);
  padding:5px 7px;
  border-radius:9px;
  color:#e5e7eb;
}
.preview-whats{
  white-space:pre-wrap;
  background:rgba(15,23,42,.76);
  border:1px solid rgba(255,255,255,.08);
  border-radius:14px;
  padding:12px;
  color:#dbeafe;
  min-height:130px;
}
@media(max-width:900px){.cobranca-config-grid{grid-template-columns:1fr}}



/* ===== LISTA SEM COBRANÇAS HOJE ===== */
.sem-cobrancas-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:8px;margin-top:12px}

.aviso-ticker{margin-top:10px;overflow:hidden;border:1px solid rgba(245,158,11,.18);border-radius:14px;background:rgba(0,0,0,.16);min-height:44px;display:flex;align-items:center;position:relative}
.aviso-ticker::before,.aviso-ticker::after{content:'';position:absolute;top:0;bottom:0;width:42px;z-index:2;pointer-events:none}.aviso-ticker::before{left:0;background:linear-gradient(90deg,rgba(17,24,39,.98),transparent)}.aviso-ticker::after{right:0;background:linear-gradient(270deg,rgba(17,24,39,.98),transparent)}
.aviso-ticker-track{display:flex;gap:12px;align-items:center;white-space:nowrap;will-change:transform;animation:mdlTicker 520s linear infinite;padding:8px 18px}.aviso-ticker:hover .aviso-ticker-track{animation-play-state:paused}
.aviso-ticker.fast .aviso-ticker-track{animation-duration:130s!important}
.aviso-pill{display:inline-flex;align-items:center;gap:8px;border:1px solid rgba(255,255,255,.08);border-radius:999px;background:rgba(255,255,255,.04);padding:8px 12px;font-weight:900;color:#f4f7ff}.aviso-pill small{font-weight:700;color:#aeb7ca;margin-left:4px}.aviso-pill .red-dot{width:7px;height:7px;border-radius:999px;background:#ef4444;box-shadow:0 0 0 4px rgba(239,68,68,.12)}
@keyframes mdlTicker{0%{transform:translateX(0)}100%{transform:translateX(-50%)}}
.reat-tabs{display:flex;gap:8px;flex-wrap:wrap;margin:10px 0 12px}.reat-tab{border:1px solid rgba(255,255,255,.08);border-radius:999px;background:rgba(255,255,255,.04);color:#dbe4ff;padding:8px 12px;font-weight:900;font-size:12px}.reat-tab.ok{border-color:rgba(34,197,94,.25);color:#86efac}.reat-tab.pending{border-color:rgba(245,158,11,.25);color:#fbbf24}
.sem-cobranca-chip{display:flex;align-items:center;gap:8px;border:1px solid rgba(239,68,68,.18);background:rgba(239,68,68,.06);border-radius:12px;padding:8px 10px;font-size:12px;font-weight:700}
.sem-cobranca-chip small{display:block;color:var(--text-muted);font-weight:600;margin-top:2px}


/* ===== FIX LEGIBILIDADE AVISOS/CAMPANHAS EM FUNDO CLARO ===== */
.campaign-banner{
  color:#1f2937 !important;
}
.campaign-banner h1,
.campaign-banner h2,
.campaign-banner h3,
.campaign-banner strong,
.campaign-banner .msg-head strong{
  color:#111827 !important;
  text-shadow:none !important;
}
.campaign-banner .hint,
.campaign-banner .small,
.campaign-banner .muted,
.campaign-banner .msg-card,
.campaign-banner .msg-card div{
  color:#334155 !important;
  text-shadow:none !important;
}
.campaign-banner .msg-card.campaign-banner,
.campaign-banner .msg-card{
  background:rgba(255,255,255,.58) !important;
  border:1px solid rgba(180,130,40,.22) !important;
  box-shadow:0 8px 24px rgba(120,80,20,.08) !important;
}
.campaign-banner .btn.soft{
  color:#111827 !important;
  background:rgba(255,255,255,.72) !important;
  border-color:rgba(120,80,20,.22) !important;
}
.campaign-banner .unread-chip{
  background:#fff7ed !important;
  border-color:#fb923c !important;
  color:#9a3412 !important;
}
.campaign-banner .mini-chip{
  font-weight:800 !important;
}
.campaign-banner .msg-head .small,
.campaign-banner .msg-head .muted{
  color:#475569 !important;
}
.campaign-banner [style*="white-space:pre-wrap"]{
  color:#1f2937 !important;
  font-weight:700 !important;
}


/* ===== FIX LEGIBILIDADE AVISOS NO MODAL DO SINO ===== */
.modal-card .modal-list .msg-card.campaign-banner,
#bellList .msg-card.campaign-banner{
  background:linear-gradient(180deg,#f7f1df,#f2ead7) !important;
  border:1px solid #e4d8b2 !important;
  color:#1f2937 !important;
}
.modal-card .modal-list .msg-card.campaign-banner *,
#bellList .msg-card.campaign-banner *{
  text-shadow:none !important;
}
.modal-card .modal-list .msg-card.campaign-banner strong,
#bellList .msg-card.campaign-banner strong{
  color:#111827 !important;
}
.modal-card .modal-list .msg-card.campaign-banner .small,
.modal-card .modal-list .msg-card.campaign-banner .muted,
#bellList .msg-card.campaign-banner .small,
#bellList .msg-card.campaign-banner .muted{
  color:#475569 !important;
}
.modal-card .modal-list .msg-card.campaign-banner [style*="white-space:pre-wrap"],
#bellList .msg-card.campaign-banner [style*="white-space:pre-wrap"]{
  color:#1f2937 !important;
  font-weight:700 !important;
}
.modal-card .modal-list .msg-card.campaign-banner .unread-chip,
#bellList .msg-card.campaign-banner .unread-chip{
  background:#fff7ed !important;
  border-color:#fb923c !important;
  color:#9a3412 !important;
}
#bellList .msg-card.campaign-banner .mini-chip{
  background:#fff7ed !important;
  border-color:#fdba74 !important;
  color:#c2410c !important;
}


/* ===== COBRANÇA INTELIGENTE: RECOBRANÇA / CONTADOR ===== */
.cob-history-chip{
  display:inline-flex;
  align-items:center;
  gap:6px;
  margin-left:6px;
  padding:4px 8px;
  border-radius:999px;
  border:1px solid rgba(251,146,60,.55);
  background:rgba(251,146,60,.13);
  color:#fed7aa;
  font-size:11px;
  font-weight:900;
}
.cob-retry-chip{
  display:inline-flex;
  align-items:center;
  gap:6px;
  margin-left:6px;
  padding:4px 8px;
  border-radius:999px;
  border:1px solid rgba(248,113,113,.62);
  background:rgba(248,113,113,.14);
  color:#fecaca;
  font-size:11px;
  font-weight:900;
  animation:pulse 1.35s infinite;
}
.cob-info-box{
  margin-top:7px;
  padding:8px 10px;
  border-radius:12px;
  border:1px solid rgba(251,146,60,.32);
  background:rgba(251,146,60,.08);
  color:#ffd6a3;
  font-size:12px;
  font-weight:800;
}
.cob-info-box.waiting{
  border-color:rgba(96,165,250,.32);
  background:rgba(96,165,250,.08);
  color:#bfdbfe;
}


/* ===== FIX GLOBAL: CAMPANHAS CLARAS EM QUALQUER MODAL/LISTA ===== */
.msg-card.campaign-banner,
.campaign-banner .msg-card,
#bellList .msg-card.campaign-banner,
.modal-card .msg-card.campaign-banner,
.modal-list .msg-card.campaign-banner{
  background:linear-gradient(180deg,#f7f1df,#f2ead7) !important;
  border:1px solid #e4d8b2 !important;
  color:#1f2937 !important;
  text-shadow:none !important;
}
.msg-card.campaign-banner *,
.campaign-banner .msg-card *,
#bellList .msg-card.campaign-banner *,
.modal-card .msg-card.campaign-banner *,
.modal-list .msg-card.campaign-banner *{
  text-shadow:none !important;
}
.msg-card.campaign-banner strong,
.campaign-banner .msg-card strong,
#bellList .msg-card.campaign-banner strong,
.modal-card .msg-card.campaign-banner strong,
.modal-list .msg-card.campaign-banner strong{
  color:#111827 !important;
}
.msg-card.campaign-banner .small,
.msg-card.campaign-banner .muted,
.campaign-banner .msg-card .small,
.campaign-banner .msg-card .muted,
#bellList .msg-card.campaign-banner .small,
#bellList .msg-card.campaign-banner .muted,
.modal-card .msg-card.campaign-banner .small,
.modal-card .msg-card.campaign-banner .muted,
.modal-list .msg-card.campaign-banner .small,
.modal-list .msg-card.campaign-banner .muted{
  color:#475569 !important;
}
.msg-card.campaign-banner [style*="white-space:pre-wrap"],
.campaign-banner .msg-card [style*="white-space:pre-wrap"],
#bellList .msg-card.campaign-banner [style*="white-space:pre-wrap"],
.modal-card .msg-card.campaign-banner [style*="white-space:pre-wrap"],
.modal-list .msg-card.campaign-banner [style*="white-space:pre-wrap"]{
  color:#1f2937 !important;
  font-weight:800 !important;
}
.msg-card.campaign-banner .btn.soft,
.campaign-banner .msg-card .btn.soft{
  color:#111827 !important;
  background:rgba(255,255,255,.78) !important;
  border-color:rgba(120,80,20,.22) !important;
}
.msg-card.campaign-banner .btn.danger,
.campaign-banner .msg-card .btn.danger{
  color:#7f1d1d !important;
  background:#fee2e2 !important;
  border-color:#fecaca !important;
}
.msg-card.campaign-banner .unread-chip,
#bellList .msg-card.campaign-banner .unread-chip,
.modal-card .msg-card.campaign-banner .unread-chip{
  background:#fff7ed !important;
  border-color:#fb923c !important;
  color:#9a3412 !important;
}
.msg-card.campaign-banner .mini-chip,
#bellList .msg-card.campaign-banner .mini-chip{
  background:#fff7ed !important;
  border-color:#fdba74 !important;
  color:#c2410c !important;
}

/* ===== LARANJITO NOTIFICAÇÕES FIXO ===== */
#laranjitoNotify{
  position:fixed;
  top:92px;
  right:22px;
  z-index:9998;
  width:66px;
  height:66px;
  border-radius:999px;
  border:2px solid rgba(255,138,0,.65);
  background:radial-gradient(circle at 35% 25%,rgba(255,255,255,.35),transparent 35%),rgba(17,24,39,.92);
  box-shadow:0 0 0 4px rgba(255,138,0,.12),0 14px 38px rgba(0,0,0,.42),0 0 28px rgba(255,138,0,.25);
  display:none;
  align-items:center;
  justify-content:center;
  cursor:pointer;
  transition:.18s transform ease,.18s filter ease;
}
#laranjitoNotify:hover{transform:scale(1.06);filter:brightness(1.08)}
#laranjitoNotify img{width:54px;height:54px;object-fit:contain;border-radius:999px}
#laranjitoNotifyBadge{
  position:absolute;
  top:-6px;
  right:-5px;
  min-width:22px;
  height:22px;
  border-radius:999px;
  background:#ef4444;
  color:#fff;
  font-size:12px;
  font-weight:950;
  display:flex;
  align-items:center;
  justify-content:center;
  border:2px solid #111827;
}
#laranjitoNotifyPanel{
  position:fixed;
  top:166px;
  right:22px;
  z-index:9999;
  width:min(420px,calc(100vw - 34px));
  max-height:70vh;
  overflow:auto;
  display:none;
  border:1px solid rgba(255,138,0,.35);
  background:rgba(17,20,29,.98);
  border-radius:22px;
  padding:14px;
  box-shadow:0 22px 55px rgba(0,0,0,.55);
}
#laranjitoNotifyPanel.show:not(:empty){display:block}
#laranjitoNotifyPanel:empty{display:none!important;border:none!important;box-shadow:none!important;padding:0!important;height:0!important;min-height:0!important;overflow:hidden!important}
.laranjito-note{
  padding:11px 12px;
  border-radius:15px;
  border:1px solid rgba(255,255,255,.11);
  background:rgba(255,255,255,.05);
  margin-bottom:10px;
  color:#e5e7eb;
  font-weight:800;
}
.laranjito-note .small{color:#aab4c8!important;margin-top:4px}

/* ===== COBRANÇA INTELIGENTE: CONTADOR NO HISTÓRICO ===== */
.log-cob-count{
  display:inline-flex;
  align-items:center;
  gap:6px;
  padding:4px 9px;
  border-radius:999px;
  background:rgba(251,146,60,.12);
  border:1px solid rgba(251,146,60,.45);
  color:#fed7aa;
  font-size:11px;
  font-weight:900;
  margin-top:4px;
}


/* ===== AJUSTE FINAL LARANJITO ALERTAS ===== */
#laranjitoNotify.has-notes{
  display:flex;
  animation:laranjitoPulse 1.25s infinite;
}
@keyframes laranjitoPulse{
  0%{transform:scale(1); box-shadow:0 0 0 0 rgba(255,138,0,.42),0 14px 38px rgba(0,0,0,.42),0 0 28px rgba(255,138,0,.25)}
  60%{transform:scale(1.055); box-shadow:0 0 0 14px rgba(255,138,0,0),0 18px 45px rgba(0,0,0,.48),0 0 38px rgba(255,138,0,.38)}
  100%{transform:scale(1); box-shadow:0 0 0 0 rgba(255,138,0,0),0 14px 38px rgba(0,0,0,.42),0 0 28px rgba(255,138,0,.25)}
}
#laranjitoNotify.no-new{
  animation:none!important;
}
#laranjitoNotify img{
  width:58px!important;
  height:58px!important;
  object-fit:contain!important;
  border-radius:999px!important;
  display:block!important;
}
#laranjitoNotifyPanel .btn.soft{
  color:#fff!important;
}


/* ===== VISUALIZADOR INLINE DA TELA CONGELADA ===== */
#snapshotInlineViewer{
  margin-top:14px;
  border:1px solid rgba(255,138,0,.35);
  background:rgba(9,12,18,.96);
  border-radius:22px;
  padding:14px;
  display:none;
}
#snapshotInlineViewer.show{display:block}
.snapshot-inline-toolbar{
  display:flex;
  justify-content:space-between;
  align-items:center;
  gap:10px;
  margin-bottom:12px;
  padding:10px 12px;
  border-radius:14px;
  background:rgba(255,255,255,.05);
  border:1px solid rgba(255,255,255,.10);
}
.snapshot-inline-body{
  overflow:auto;
  max-height:78vh;
  border-radius:18px;
}


/* ===== HOTFIX FINAL LARANJITO BOLINHA ===== */
#laranjitoNotify{
  overflow:hidden!important;
  font-size:0!important;
  line-height:0!important;
}
#laranjitoNotify img#laranjitoNotifyImg{
  width:58px!important;
  height:58px!important;
  object-fit:contain!important;
  display:block!important;
  border-radius:999px!important;
}
#laranjitoNotify.only-current-user-hidden{display:none!important;}


/* ===== FIX DEFINITIVO BOLINHA LARANJITO ===== */
#laranjitoNotify{
  overflow:hidden!important;
  font-size:0!important;
}
#laranjitoNotify img#laranjitoNotifyImg{
  width:60px!important;
  height:60px!important;
  object-fit:contain!important;
  display:block!important;
  border-radius:999px!important;
  background:transparent!important;
}
#laranjitoNotifyFallback{
  display:none;
  font-size:38px;
  line-height:1;
}
#laranjitoNotify.img-fail #laranjitoNotifyImg{display:none!important}
#laranjitoNotify.img-fail #laranjitoNotifyFallback{display:block!important}
body.master-view #laranjitoNotify, body.diretor-view #laranjitoNotify{display:none!important}


/* ===== V3.1: acordeões e início compacto ===== */
.acc-head .acc-hint{
  width:auto!important;height:auto!important;min-width:180px!important;max-width:520px!important;
  border-radius:999px!important;background:rgba(255,255,255,.06)!important;
  padding:6px 12px!important;font-size:12px!important;line-height:1.2!important;
  color:#dbe4ff!important;white-space:nowrap!important;text-align:right!important;display:inline-flex!important;
  align-items:center!important;justify-content:flex-end!important;transform:none!important;writing-mode:horizontal-tb!important;
}
.accordion.open .acc-head .acc-hint{transform:none!important}
.reat-tabs .reat-tab{cursor:pointer;user-select:none}.reat-tabs .reat-tab.active{box-shadow:0 0 0 2px rgba(249,168,50,.35) inset;background:rgba(249,168,50,.12)}
.inicio-compact .kpis{grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin-bottom:14px}
body.inicio-view .kpi{padding:14px 16px;min-height:92px}body.inicio-view .kpi .value{font-size:21px}body.inicio-view .app-shell{padding-top:14px}
.inicio-murais-grid{display:grid;grid-template-columns:1fr;gap:10px;margin-top:10px}.inicio-murais-grid .glass.panel{margin-bottom:0!important;padding:12px 14px!important}.inicio-murais-grid h2{font-size:16px!important}.inicio-murais-grid .hint{font-size:11px}.inicio-murais-grid .aviso-ticker{min-height:38px}.inicio-murais-grid .aviso-ticker-track{animation-duration:620s!important;padding:7px 14px}.inicio-campaign-compact .campaign-banner{margin-bottom:12px;padding:12px 14px}
.inicio-operacional-compact{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:10px}.inicio-operacional-compact .glass.panel{padding:12px 14px!important;margin-bottom:0!important}.inicio-operacional-compact h2{font-size:16px!important}.inicio-operacional-compact .hint{font-size:11px}.inicio-operacional-compact .aviso-ticker{min-height:38px}.inicio-operacional-compact .aviso-ticker-track{animation-duration:620s!important;padding:7px 14px}.inicio-operacional-compact .aviso-ticker.fast .aviso-ticker-track{animation-duration:150s!important}
/* ===== V3.4: enquadramento menor + murais sem vazar ===== */
html,body{overflow-x:hidden!important}
.app-shell{max-width:1560px!important;margin:0 auto!important;padding:20px 24px 100px!important;overflow-x:hidden!important}
#app,#mainScreen,#inicioSection{max-width:100%!important;overflow-x:hidden!important}
.inicio-compact,.inicio-murais-grid,.inicio-operacional-compact{max-width:100%!important;min-width:0!important;overflow:hidden!important}
.inicio-operacional-compact{display:grid!important;grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:12px!important;align-items:start!important}
.inicio-operacional-compact .glass.panel,.inicio-murais-grid .glass.panel{min-width:0!important;max-width:100%!important;overflow:hidden!important;margin-bottom:0!important}
.inicio-operacional-compact .glass.panel.full{grid-column:1/-1!important}
.aviso-ticker{width:100%!important;max-width:100%!important;min-width:0!important;overflow:hidden!important}
.aviso-ticker-track{animation-duration:1400s!important;will-change:transform}
.aviso-ticker.fast .aviso-ticker-track{animation-duration:260s!important}
.inicio-operacional-compact .aviso-ticker-track{animation-duration:1400s!important}
.inicio-operacional-compact .aviso-ticker.fast .aviso-ticker-track{animation-duration:260s!important}
.aviso-pill{flex:0 0 auto;max-width:420px!important;overflow:hidden!important;text-overflow:ellipsis!important}
.acc-head{gap:12px!important;min-width:0!important;overflow:hidden!important}
.acc-head>span:first-child{min-width:0!important;overflow:hidden!important;text-overflow:ellipsis!important}
.acc-head .acc-hint{position:static!important;margin-left:auto!important;flex:0 0 auto!important;min-width:0!important;max-width:520px!important;white-space:nowrap!important;writing-mode:horizontal-tb!important;transform:none!important;text-align:right!important}
body.inicio-view .kpis{grid-template-columns:repeat(4,minmax(0,1fr))!important;gap:12px!important}
body.inicio-view .kpi{padding:14px 16px!important;min-height:88px!important}
body.inicio-view .kpi .value{font-size:21px!important}
@media(max-width:1180px){.inicio-operacional-compact{grid-template-columns:1fr!important}.app-shell{padding:16px 16px 90px!important}}
@media(max-width:1100px){.inicio-murais-grid,.inicio-operacional-compact{grid-template-columns:1fr}.acc-head .acc-hint{max-width:100%!important;text-align:left!important;white-space:normal!important;justify-content:flex-start!important}}

/* ===== V27D: LISTA COMPACTA MASTER/DIRETOR + PAGINAÇÃO LOGS ===== */
.entity-list{display:grid;gap:8px;margin-bottom:18px}
.entity-row{display:grid;grid-template-columns:minmax(240px,1.5fr) repeat(6,minmax(105px,.7fr)) minmax(140px,.8fr);gap:10px;align-items:center;padding:12px 14px;border-radius:16px;background:rgba(255,255,255,.035);border:1px solid rgba(255,255,255,.09);cursor:pointer;transition:var(--transition)}
.entity-row:hover{background:rgba(255,255,255,.065);border-color:rgba(249,168,50,.25);transform:translateY(-1px)}
.entity-row-head{font-size:12px;color:#dbe4ff;font-weight:900;text-transform:uppercase;letter-spacing:.05em;background:rgba(255,255,255,.05);cursor:default}
.entity-row-head:hover{transform:none;background:rgba(255,255,255,.05)}
.entity-cell .k{font-size:10px;color:var(--text-muted);font-weight:800;text-transform:uppercase;letter-spacing:.08em}.entity-cell .v{font-size:13px;font-weight:900;color:var(--text-primary)}
.entity-cell .v.red{color:var(--red)}.entity-cell .v.green{color:var(--green)}.entity-cell .v.blue{color:var(--blue)}.entity-cell .v.orange{color:var(--orange)}
.log-pager{display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap;margin:12px 0;padding:10px 12px;border-radius:14px;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08)}
@media(max-width:1100px){.entity-row{grid-template-columns:1fr 1fr}.entity-row-head{display:none}}


/* ===== V3.5 ajustes visuais e murais ===== */
.senhas-table-wrap{overflow:hidden!important;width:100%!important;max-width:100%!important}
.senhas-table{min-width:0!important;width:100%!important;table-layout:fixed!important;font-size:12px!important}
.senhas-table th,.senhas-table td{padding:8px 8px!important;font-size:12px!important;line-height:1.25!important;overflow:hidden!important;text-overflow:ellipsis!important;vertical-align:middle!important}
.senhas-table th:nth-child(1),.senhas-table td:nth-child(1){width:12%!important}.senhas-table th:nth-child(2),.senhas-table td:nth-child(2){width:10%!important}.senhas-table th:nth-child(3),.senhas-table td:nth-child(3){width:16%!important}.senhas-table th:nth-child(4),.senhas-table td:nth-child(4){width:5%!important}.senhas-table th:nth-child(5),.senhas-table td:nth-child(5){width:9%!important}.senhas-table th:nth-child(6),.senhas-table td:nth-child(6){width:9%!important}.senhas-table th:nth-child(7),.senhas-table td:nth-child(7){width:17%!important}.senhas-table th:nth-child(8),.senhas-table td:nth-child(8){width:14%!important}.senhas-table th:nth-child(9),.senhas-table td:nth-child(9){width:8%!important}
.senha-view-row,.senha-nova-row{gap:4px!important;min-width:0!important;width:100%!important;flex-wrap:nowrap!important}.senha-view-row input,.senha-nova-row input{min-width:0!important;width:100%!important;max-width:140px!important;padding:7px 8px!important}.senhas-table .btn-xs{padding:7px 8px!important;font-size:11px!important}.senhas-table td strong{white-space:normal!important;word-break:break-word!important}
.admin-accounts-line{display:block!important}.admin-accounts-line .senhas-table-wrap{margin-top:8px!important}
.inicio-operacional-compact{display:block!important;grid-template-columns:1fr!important;gap:0!important}.inicio-operacional-compact .glass.panel{width:100%!important;margin:0 0 10px 0!important}.inicio-operacional-compact .section-head{margin-bottom:6px!important}.inicio-operacional-compact .legend-inline{display:flex!important;flex-wrap:wrap!important;gap:10px!important}.inicio-operacional-compact .legend-inline span{white-space:normal!important}
.aviso-ticker{width:100%!important;max-width:100%!important;overflow:hidden!important}.aviso-ticker-track{animation-duration:1800s!important}.aviso-ticker.fast .aviso-ticker-track{animation-duration:45s!important}.inicio-operacional-compact .aviso-ticker-track{animation-duration:1800s!important}.inicio-operacional-compact .aviso-ticker.fast .aviso-ticker-track{animation-duration:45s!important}.ticker-speed-btn{min-width:104px;justify-content:center}.mural-section-title{margin:14px 0 8px!important}.mural-section-title h2{font-size:18px!important}
@media(max-width:1400px){.senhas-table th,.senhas-table td{font-size:11px!important;padding:7px 6px!important}.senha-view-row input,.senha-nova-row input{max-width:120px!important}.senhas-table .btn-xs{padding:6px 7px!important}}



/* V8.9 HOTFIX MOBILE */
@media (max-width: 760px){
  body{overflow-x:hidden!important}
  .container{padding:12px!important;max-width:100vw!important}
  .hero,.glass.panel,.campaign-banner{border-radius:22px!important;padding:18px!important;margin-left:0!important;margin-right:0!important}
  .hero{display:block!important;text-align:left!important}
  .hero .section-head,.section-head{display:flex!important;flex-direction:column!important;align-items:flex-start!important;gap:10px!important}
  .kpis{display:grid!important;grid-template-columns:1fr 1fr!important;gap:10px!important;overflow:visible!important}
  .kpi-card,.metric,.input-card{min-width:0!important;width:auto!important;padding:14px!important;overflow:hidden!important}
  .kpi-card .k,.metric .k{font-size:10px!important;letter-spacing:.08em!important;white-space:normal!important;word-break:break-word!important}
  .kpi-card .v,.metric .v{font-size:20px!important;white-space:normal!important;line-height:1.12!important;word-break:break-word!important}
  .kpi-card .sub,.metric .sub,.hint,.small{font-size:11px!important;line-height:1.3!important;white-space:normal!important}
  .master-tabs{display:flex!important;gap:8px!important;overflow-x:auto!important;padding:10px!important;scroll-snap-type:x proximity!important}
  .master-tabs .tab{flex:0 0 auto!important;white-space:nowrap!important;scroll-snap-align:start!important}
  .detail-top,.grid-2,.form-grid,.filters-grid{display:grid!important;grid-template-columns:1fr!important;gap:12px!important}
  .log-row,.row-top{grid-template-columns:1fr!important;gap:8px!important}
  .accordion,.acc-body,.logs-list{overflow-x:hidden!important}
  .btn{min-height:42px!important}
}
@media (max-width: 420px){
  .kpis{grid-template-columns:1fr!important}
  .hero h1,.hero h2{font-size:24px!important;line-height:1.15!important}
  .kpi-card .v,.metric .v{font-size:18px!important}
}

</style>
</head>
<body>
<div id="loginScreen" class="login-wrap">
  <div class="glass login-card">
    <img class="logo-big" src="__LOGO__" alt="logo">
    <h2>Dashboard – Lojas MDL</h2>
    <div class="sub" style="text-align:center">Entre com seu usuário de colaborador ou Master</div>
    <div class="note" style="text-align:center;margin-top:8px;color:#f59e0b;font-weight:900">__DASH_VERSION_LABEL__</div>
    <div class="login-form">
      <input id="loginUser" placeholder="Usuário">
      <input id="loginPass" type="password" placeholder="Senha">
      <button id="loginBtn" type="button" class="btn primary" data-login-btn="1" onclick="return fazerLogin()">🔐 Entrar</button>
      <div style="display:flex;gap:10px;justify-content:center;flex-wrap:wrap;margin-top:8px">
        <button class="btn soft" type="button" onclick="openPrimeiroAcesso()">🔑 Primeiro acesso</button>
        <button class="btn soft" type="button" onclick="openRecuperarSenha()">📩 Recuperar senha</button>
      </div>
      <div id="loginMsg" class="note" style="text-align:center;color:#d97706"></div>
    </div>
  </div>
</div>

<div id="app" class="hidden">
  <div class="app-shell">
    <div class="glass header">
      <div class="brand">
        <img src="__LOGO__" alt="logo">
        <div>
          <h1>Dashboard – Lojas MDL</h1>
          <div class="sub">Período Cobrança: __PERIODO__ · Vendedores: painel individual · Gerentes: painel de filial</div>
          <div class="sub" style="color:#f59e0b;font-weight:900">__DASH_VERSION_LABEL__</div>
        </div>
      </div>
      <div class="header-actions">
        <button id="bellBtn" class="btn soft" onclick="openBell()">🔔 Avisos <span id="bellCount" class="badge" style="padding:4px 8px;margin-left:4px;background:#eef5ff">0</span></button>
        <button id="goalNotifBtn" class="btn soft hidden" onclick="openGoalNotifications()">🎉 Notificações <span id="goalNotifCount" class="badge" style="padding:4px 8px;margin-left:4px;background:#fff4d6">0</span></button>
        <div id="userBadge" class="badge">👑 Master</div>
        <button class="btn soft" onclick="logout()">Sair</button>
      </div>
    </div>

    <div id="topMural" class="hidden"></div>

    <div id="kpis" class="kpis"></div>

    <div id="masterTabs" class="tabs hidden">
      <button class="tab active" data-tab="inicio" onclick="setMainTab('inicio')">🏠 Início</button>
      <button class="tab" data-tab="vendedores" onclick="setMainTab('vendedores')">👥 Por Colaborador</button>
      <button class="tab" data-tab="filiais" onclick="setMainTab('filiais')">🏬 Por Filial</button>
      <button class="tab" data-tab="metas" onclick="setMainTab('metas')">🎯 Metas</button>
      <button class="tab" data-tab="servicos" onclick="setMainTab('servicos')">🛠️ Serviços</button>
      <button class="tab" data-tab="cobrancas" onclick="setMainTab('cobrancas')">🧾 Cobranças</button>
      <button class="tab" data-tab="reativacao" onclick="setMainTab('reativacao')">🧡 Clientes sem movimento</button>
      <button class="tab" data-tab="aniversariantes" onclick="setMainTab('aniversariantes')">🎂 Aniversariantes</button>
      <button class="tab" data-tab="avisos" onclick="setMainTab('avisos')">📣 Avisos</button>
      <button class="tab" data-tab="telegram" onclick="setMainTab('telegram')">📲 Telegram</button>
      <button class="tab" data-tab="senhas" onclick="setMainTab('senhas')">🔐 Senhas</button>
      <button class="tab" data-tab="historico" onclick="setMainTab('historico')">🗂️ Histórico</button>
    </div>

    <div id="mainFilters" class="filters hidden"></div>

    <div id="mainScreen">
      <div id="inicioSection" class="hidden"></div>
      <div id="listSection"></div>
      <div id="metaSection" class="hidden"></div>
      <div id="servicesSection" class="hidden"></div>
      <div id="logSection" class="hidden"></div>
      <div id="reativacaoSection" class="hidden"></div>
      <div id="aniversariantesSection" class="hidden"></div>
      <div id="avisosSection" class="hidden"></div>
      <div id="telegramSection" class="hidden"></div>
      <div id="senhasSection" class="hidden"></div>
      <div id="histSection" class="hidden"></div>
    </div>

    <div id="detailScreen" class="detail-screen hidden"></div>
  </div>
</div>

<div id="toast" class="toast"></div>
<div id="phoneModal" class="modal"><div class="glass modal-card"><h3 style="margin:0">Escolher contato</h3><div class="sub">Selecione para abrir o WhatsApp</div><div id="phoneList" class="modal-list"></div><button class="btn soft" style="width:100%;margin-top:10px" onclick="closePhoneModal()">Fechar</button></div></div>
<div id="bellModal" class="modal"><div class="glass modal-card" style="width:min(760px,100%)"><h3 style="margin:0">🔔 Avisos e mensagens</h3><div class="sub">Atualizações enviadas pelo Master para você.</div><div id="bellList" class="modal-list" style="max-height:70vh;overflow:auto"></div><button class="btn soft" style="width:100%;margin-top:10px" onclick="closeBell()">Fechar</button></div></div>
<div id="firstAccessModal" class="modal"><div class="glass modal-card" style="width:min(560px,100%)"><h3 style="margin:0">🔑 Primeiro acesso / troca de senha</h3><div class="sub">Defina sua nova senha para entrar no dashboard.</div><div class="modal-list"><div class="input-card"><label>Usuário</label><input id="faLogin" placeholder="Ex: joaodasilva"></div><div class="input-card"><label>Senha atual</label><input id="faCurrentPass" type="password" placeholder="Senha atual"></div><div class="input-card"><label>Nova senha</label><input id="faNewPass" type="password" placeholder="Nova senha"></div><div class="input-card"><label>Confirmar nova senha</label><input id="faNewPass2" type="password" placeholder="Confirmar nova senha"></div><div id="faMsg" class="note"></div></div><div style="display:flex;gap:10px;margin-top:10px"><button class="btn primary" style="flex:1" onclick="salvarPrimeiroAcesso()">💾 Salvar nova senha</button><button class="btn soft" style="flex:1" onclick="closeFirstAccess()">Fechar</button></div></div></div>
<div id="recoverModal" class="modal"><div class="glass modal-card" style="width:min(560px,100%)"><h3 style="margin:0">📩 Recuperar senha</h3><div class="sub">O pedido será enviado para o Master no dashboard e também pode ser registrado em sac@moveisdolar.com.br.</div><div class="modal-list"><div class="input-card"><label>Usuário</label><input id="recLogin" placeholder="Seu usuário"></div><div class="input-card"><label>Nome / observação</label><input id="recObs" placeholder="Ex: sou da filial F4"></div><div id="recMsg" class="note"></div></div><div style="display:flex;gap:10px;margin-top:10px"><button class="btn primary" style="flex:1" onclick="enviarRecuperacaoSenha()">📨 Enviar solicitação</button><button class="btn soft" style="flex:1" onclick="closeRecover()">Fechar</button></div></div></div>


<div id="laranjitoNotify" onclick="toggleLaranjitoNotifyPanel()" title="Alertas e parabéns">
  <img id="laranjitoNotifyImg" src="" alt="" referrerpolicy="no-referrer">
  <span id="laranjitoNotifyFallback">🍊</span>
  <span id="laranjitoNotifyBadge">0</span>
</div>
<div id="laranjitoNotifyPanel"></div>

<script>
const CREDS=__JS_CREDS__;
const AUTH_BOOT=__JS_AUTH_STATE__;
const TODOS=__JS_TODOS__;
const FILIAIS=__JS_FILIAIS__;
const CLIENTES_FIL=__JS_CLIENTES__;
const CLIENTES_VEND=__JS_CLIENTES_VEND__;
const CLIENTES_TERCEIRO=__JS_CLIENTES_TERCEIRO__;
const CLIENTES_CREDIARISTA=__JS_CLIENTES_CREDIARISTA__||{};
const RECEBIMENTOS=__JS_RECEBIMENTOS__;
const RECEBIMENTOS_TERCEIRO=__JS_RECEBIMENTOS_TERCEIRO__;
const RECEBIMENTOS_CREDIARISTA=__JS_RECEBIMENTOS_CREDIARISTA__||{};
const QUITADOS_180=__JS_QUITADOS_180__||[];
const HIST_RECEBIMENTOS_MENSAIS=__JS_HIST_RECEBIMENTOS_MENSAIS__||{months:{}};
const DASHBOARD_BUILD_VERSION=__JS_DASHBOARD_BUILD_VERSION__;
const DASHBOARD_BUILD_TAG=__JS_DASHBOARD_BUILD_TAG__;

function isAdminLike(){
  const t=String(usuarioAtual?.tipo||'').toLowerCase();
  const l=String(usuarioAtual?.login||'').toLowerCase();
  return ['master','diretor','painel','director'].includes(t) || ['master','diretorcomercial','painel'].includes(l);
}
function renderBackButton(){
  return isAdminLike()?'<button class="btn soft" onclick="backToMain()">↩️ Voltar</button>':'';
}
const CLIENTES_SEM_MOVIMENTO=__JS_CLIENTES_SEM_MOVIMENTO__||[];
const CLIENTES_SEM_MOVIMENTO_BASE=__JS_CLIENTES_SEM_MOVIMENTO_BASE__||[];
window.CLIENTES_SEM_MOVIMENTO_BASE=CLIENTES_SEM_MOVIMENTO_BASE;
const CLIENTES_SEM_MOVIMENTO_META=__JS_CLIENTES_SEM_MOVIMENTO_META__||{modo:'somente_novos',base_total:CLIENTES_SEM_MOVIMENTO.length,novos_total:CLIENTES_SEM_MOVIMENTO.length,seen_total:0};
const ANIVERSARIANTES=__JS_ANIVERSARIANTES__||[];
const DUPLICIDADES_CARTEIRA=__JS_DUPLICIDADES_CARTEIRA__||{total_conflitos:0,conflitos:[]};
let RECEBIMENTOS_CONCILIADOS={};
let CREDIARISTAS_CONFIG=Array.isArray(__JS_CREDIARISTAS_MAP__)?__JS_CREDIARISTAS_MAP__:[];
if(!CREDIARISTAS_CONFIG.length && __JS_CREDIARISTAS_MAP__ && typeof __JS_CREDIARISTAS_MAP__==='object'){
  CREDIARISTAS_CONFIG=Object.entries(__JS_CREDIARISTAS_MAP__).map(([filial,login])=>({filial,login,nome:`Crediarista ${filial}`,pct:100}));
}
function getCrediaristasConfig(){return Array.isArray(CONFIG_META?.crediaristas_config)&&CONFIG_META.crediaristas_config.length?CONFIG_META.crediaristas_config:CREDIARISTAS_CONFIG}
const CREDIARISTAS_MAP=new Proxy({}, {get(t,k){const row=getCrediaristasConfig().find(r=>String(r.filial||'').toUpperCase()===String(k||'').toUpperCase());return row?String(row.login||'').toLowerCase():undefined}, ownKeys(){return getCrediaristasConfig().map(r=>r.filial)}, getOwnPropertyDescriptor(){return {enumerable:true,configurable:true}}});
const COBRANCA10_LOGIN='cobranca10';
const COBRANCA10_NOME='Cobrança10';
let METAS_VENDAS=__JS_METAS_VENDAS__||{metas:{}};
let METAS_VENDAS_DIA=__JS_METAS_VENDAS_DIA__||{metas:{}};
const MARGENS_BRUTAS=__JS_MARGENS_BRUTAS__||{filiais:{},vendedores:{}};
let SALES_EMPRESA=__JS_SALES_EMPRESA__||{};
let RENT_EMPRESA=__JS_RENT_EMPRESA__||{};

function getRentEmpresa(){
  let e = (RENT_EMPRESA && Object.keys(RENT_EMPRESA).length) ? {...RENT_EMPRESA} : {};
  const mb = MARGENS_BRUTAS || {};
  if((!e || !Object.keys(e).length) && mb.empresa) e = {...mb.empresa};

  // Fallback: se por qualquer motivo a empresa não veio no JSON,
  // soma as filiais do relatório de margem bruta.
  if((!Number(e.custo_total||0)) && mb.filiais){
    const vals = Object.values(mb.filiais || {});
    const soma = (campo)=>vals.reduce((acc,x)=>acc + Number(x?.[campo]||0), 0);
    e.custo_total = soma('custo_total');
    e.valor_total = soma('valor_total');
    e.valor_total_liquido = soma('valor_total_liquido');
    e.margem_bruta_valor = soma('margem_bruta_valor');
  }

  if((!Number(e.margem_bruta_pct||0)) && Number(e.valor_total||0)>0){
    e.margem_bruta_pct = Number(e.margem_bruta_valor||0) / Number(e.valor_total||0) * 100;
  }
  if((!Number(e.markup_realizado||0)) && Number(e.custo_total||0)>0){
    e.markup_realizado = Number(e.valor_total||0) / Number(e.custo_total||0);
  }
  return e || {};
}

let SERVICOS_RELATORIO=__JS_SERVICOS_RELATORIO__||{empresa:{},servicos:{},filiais:{},vendedores:{},detalhes:[]};
let CONFIG_META={grave_pct:20,alerta_pct:15,atencao_pct:10,comissao_pagamento_texto:'A comissão reinicia a cada mês e o pagamento é previsto para o dia 25 do mês seguinte.',reativacao_rateio_modo:'igualitario',reativacao_msg_template_filiais:{},aniversario_msg_template_filiais:{},dias_uteis_meta_diaria:25,aniversario_msg_template:`Olá, {primeiro_nome}! Feliz aniversário!

Aqui é da Lojas MDL - Móveis do Lar. Desejamos muita saúde, paz e felicidades neste dia especial.

Preparamos condições especiais para você comemorar com a gente.`,reativacao_msg_template:`Olá, {primeiro_nome}! Tudo bem?

Aqui é da Lojas MDL - Móveis do Lar. Estamos com saudades de você! Faz um tempinho que você não aparece na loja.

Venha conhecer nossas novidades e aproveitar condições especiais que preparamos para nossos clientes.`,peso_grave:60,peso_alerta:30,peso_atencao:10,vendas_min_pct:80,servicos_min_pct:80,gerente_vendas_min_pct:80,gerente_servicos_min_pct:80,vendedor_rentab_min_mercantil_pct:80,gerente_rentab_min_mercantil_pct:80,bonus_50:'',bonus_75:'',bonus_85:'',bonus_100:'',cob_cred_rateio_filial_pct:50,cob_cred_rateio_cred_pct:50,cobranca_global_rateio_pct:20,cobranca_msg_template_terceira:`Olá, {primeiro_nome}. Tudo bem?
Aqui é da Lojas MDL - Móveis do Lar.

Já tentamos contato sobre a parcela vencida em {vencimento}, no valor de {valor}, referente ao título {titulo}/{parcela}.

Para evitar novos encargos e restrições, pedimos que regularize o pagamento o quanto antes.
Caso já tenha pago, por gentileza desconsidere esta mensagem.`,cobranca_msg_template:`Olá, {primeiro_nome} tudo bem?\nAqui é da Lojas MDL - Móveis do Lar.\nPassando para lembrar que tem uma parcelinha vencida na data de {vencimento}, no valor de {valor}.\nCaso o pagamento já tenha sido realizado, por gentileza, desconsidere esta mensagem.\nSe precisar do boleto, chave PIX ou tiver qualquer dúvida, fico à disposição para ajudar.`,...(__CONFIG_META__||{})};
// V8.5: normaliza mensagens padrão se o servidor ainda estiver com versões antigas sem emojis.
const DEFAULT_ANIVERSARIO_MSG_MDL=`Olá, {primeiro_nome}! Feliz aniversário! 🎂🎉

Aqui é da Lojas MDL - Móveis do Lar. Desejamos muita saúde, paz e felicidades neste dia especial. 😍😍

Preparamos condições especiais para você comemorar com a gente.
🕺🎉🤩`;
const DEFAULT_REATIVACAO_MSG_MDL=`Olá, {primeiro_nome}! Tudo bem? 😊

Aqui é da Lojas MDL - Móveis do Lar. Estamos com saudades de você! Faz um tempinho que você não aparece na loja.  🥹

Venha conhecer nossas novidades e aproveitar condições especiais que preparamos para nossos clientes. 👈👈😍😍`;
try{
  const anivRaw=String(CONFIG_META.aniversario_msg_template||'').trim();
  if(!anivRaw || (anivRaw.includes('Feliz aniversário') && !anivRaw.includes('🎂'))){CONFIG_META.aniversario_msg_template=DEFAULT_ANIVERSARIO_MSG_MDL;}
  const reatRaw=String(CONFIG_META.reativacao_msg_template||'').trim();
  if(!reatRaw || (reatRaw.includes('Estamos com saudades') && !reatRaw.includes('🥹'))){CONFIG_META.reativacao_msg_template=DEFAULT_REATIVACAO_MSG_MDL;}
}catch(e){}

let CONFIG_META_IND=__CONFIG_META_IND__||{};
const LOGIN_MASTER=String(__LOGIN_MASTER__);
const SENHA_MASTER=String(__SENHA_MASTER__);
const LOGIN_DIRETOR=String(__LOGIN_DIRETOR__);
const SENHA_DIRETOR=String(__SENHA_DIRETOR__);
const ORDEM=__ORDEM__||[];
const TOTAL_P=Number(__TOTAL_P__||0);
const TOTAL_PG=Number(__TOTAL_PG__||0);
const DESTAQUE_SEMANA=__JS_DESTAQUE__||{};
const LOGO='__LOGO__';
const LARANJITO='__LARANJITO__';
const MDL_COIN_IMG='https://moveisdolar.com.br/colaborador/coin%20png.png';
const API_CFG='config_meta_api.php';
const API_COB='cobrancas_api.php';
const API_MSG='mensagens_api.php';
const API_CRED='credenciais_api.php';
const API_HIST='historico_api.php';
const API_COMIS='historico_comissionamento_api.php';

function sleep(ms){return new Promise(resolve=>setTimeout(resolve,ms));}
function normName(s){
  return String(s||'')
    .normalize('NFD')
    .replace(/[\u0300-\u036f]/g,'')
    .toUpperCase()
    .replace(/[^A-Z0-9]+/g,' ')
    .replace(/\s+/g,' ')
    .trim();
}
function fmtTelBR(t){
  const d=String(t||'').replace(/\D/g,'').replace(/^55/,'');
  return d;
}
function ordenarFiliaisReativacao(filiais){
  const ordem=ORDEM||['F1','F2','F3','F4','F5','F6','F8','F9'];
  return [...filiais].filter(f=>f!=='F7').sort((a,b)=>{
    const ia=ordem.indexOf(a), ib=ordem.indexOf(b);
    return (ia<0?999:ia)-(ib<0?999:ib);
  });
}
function firstNameFromFullName(s){
  const n=String(s||'').replace(/^\s*\d+\s*-\s*/,'').trim();
  return (n.split(/\s+/)[0]||n||'Cliente');
}
async function fetchComTimeout(url, opts={}, ms=3500){
  const ctrl = new AbortController();
  const timer = setTimeout(()=>{try{ctrl.abort()}catch(e){}}, ms);
  try{
    return await fetch(url,{...opts,cache:'no-store',signal:ctrl.signal});
  }finally{
    clearTimeout(timer);
  }
}
async function tentarAtualizarOnlineDepoisLogin(){
  try{
    await Promise.allSettled([
      carregarCredenciaisOnline(),
      carregarConfigOnline(),
      carregarHistoricoOnline(),
      carregarHistoricoComissaoOnline(),
      carregarCobrancasOnline(),
      carregarMsgsOnline()
    ]);
    try{renderKPIs()}catch(e){}
    try{refreshBell()}catch(e){}; try{renderLaranjitoNotify(); showLaranjitoOncePerAccess()}catch(e){}
    try{
      if(usuarioAtual?.tipo==='master'){
        if(!detailScreen.classList.contains('hidden') && currentDetailRef) openEntity(currentDetailRef);
        else renderList();
        if(isUltimoDiaMes23()) setTimeout(()=>salvarSnapshotComissionamentoMensal(true),900);
      }
    }catch(e){}
  }catch(e){console.log('Falha atualização online pós-login',e);}
}

let usuarioAtual=null;
let mainTab='vendedores';
let filtroFilial='TODAS';
let COB_LOGS=[];
let phoneContext=null;
let MSGS=[];
let currentDetailRef=null;
let _logFiltered=[]; let _logPage=1; const LOG_PAGE_SIZE=50;
let AUTH_STATE=AUTH_BOOT||{users:{},director:{},password_reset_requests:[]};
let _pendingFirstAccess=null;
let HIST_DASH=__JS_HIST_DASH__||{dates:{}};
let HIST_COMISSAO={months:{}};
const SESSION_KEY='mdl_dashboard_session_v1';
function getAuthUser(login){const k=String(login||'').trim().toLowerCase(); if(k===LOGIN_DIRETOR.toLowerCase()) return AUTH_STATE?.director||null; return (AUTH_STATE?.users||{})[k]||null}
async function carregarCredenciaisOnline(){try{const r=await fetchComTimeout(API_CRED+'?_='+Date.now(),{},2500); const j=await r.json(); if(j.ok && j.data){AUTH_STATE=j.data;}}catch(e){console.log('Falha ao carregar credenciais online',e);}}
function acessoGeralBloqueado(){return !!(AUTH_STATE && AUTH_STATE.access_blocked);}
function textoBloqueioAcesso(){return String(AUTH_STATE?.access_blocked_reason || 'Sistema em atualização. Aguarde liberação pelo Master.');}

async function carregarHistoricoOnline(){try{const r=await fetchComTimeout(API_HIST+'?_='+Date.now(),{},2500); const j=await r.json(); if(j.ok && j.data){HIST_DASH=j.data;}}catch(e){console.log('Falha ao carregar histórico online',e);}}
async function carregarHistoricoComissaoOnline(){try{const r=await fetchComTimeout(API_COMIS+'?_='+Date.now(),{},3000); const j=await r.json(); if(j.ok && j.data){HIST_COMISSAO=j.data;}}catch(e){console.log('Falha ao carregar histórico de comissionamento',e);}}



let LARANJITO_NOTES=[];
let LARANJITO_UNREAD=0;
function currentSessionNoticeKey(){
  return 'mdl_laranjito_seen_'+(usuarioAtual?.login||usuarioAtual?.nome||usuarioAtual?.tipo||'anon')+'_'+new Date().toISOString().slice(0,10);
}
function shouldShowLaranjitoBubble(){
  // Aparece somente para colaborador logado olhando o próprio painel individual.
  // Master/Diretor nunca veem essa bolinha, mesmo abrindo painel de alguém.
  try{
    const tipo=String(usuarioAtual?.tipo||'').toLowerCase();
    if(tipo==='master' || tipo==='diretor' || tipo==='painel') return false;
    const detail=document.getElementById('detailScreen');
    if(!(detail && !detail.classList.contains('hidden') && currentDetailRef)) return false;
    const login=String(usuarioAtual?.login||'').toLowerCase();
    const refLogin=String(currentDetailRef?.login||'').toLowerCase();
    const uNome=normName(usuarioAtual?.nome||'');
    const rNome=normName(currentDetailRef?.nome||'');
    const uFil=String(usuarioAtual?.filial||'').toUpperCase();
    const rFil=String(currentDetailRef?.filial||'').toUpperCase();
    if(login && refLogin && login!==refLogin) return false;
    if(uNome && rNome && uNome!==rNome && tipo==='vendedor') return false;
    if(uFil && rFil && uFil!==rFil && (tipo==='filial'||tipo==='crediarista'||tipo==='cobranca')) return false;
    return true;
  }catch(e){return false}
}
function addLaranjitoNote(title,msg,kind='info'){
  const key=String(title||'')+'|'+String(msg||'');
  if(LARANJITO_NOTES.some(n=>n.key===key)) return;
  LARANJITO_NOTES.push({key,title,msg,kind,dt:new Date().toLocaleString('pt-BR')});
  LARANJITO_UNREAD++;
  renderLaranjitoNotify();
}
function renderLaranjitoNotify(){
  const btn=document.getElementById('laranjitoNotify');
  const badge=document.getElementById('laranjitoNotifyBadge');
  const panel=document.getElementById('laranjitoNotifyPanel');
  if(!btn||!badge||!panel) return;


  try{
    const img=document.getElementById('laranjitoNotifyImg');
    const btn=document.getElementById('laranjitoNotify');
    if(img && !img.dataset.ready){
      img.dataset.ready='1';
      const urls=[
        '/colaborador/mascote%20feliz1.png',
        'https://moveisdolar.com.br/colaborador/mascote%20feliz1.png',
        (typeof LARANJITO!=='undefined' && LARANJITO)?LARANJITO:''
      ].filter(Boolean);
      let ix=0;
      img.onerror=function(){
        ix++;
        if(ix<urls.length){ this.src=urls[ix]; }
        else{ if(btn) btn.classList.add('img-fail'); }
      };
      img.onload=function(){ if(btn) btn.classList.remove('img-fail'); };
      img.src=urls[0];
    }
  }catch(e){}

  const visible = shouldShowLaranjitoBubble() && LARANJITO_NOTES.length>0;
  btn.style.display = visible ? 'flex' : 'none';
  if(!visible){panel.classList.remove('show'); panel.style.display='none'; panel.innerHTML=''; return;}

  btn.classList.toggle('has-notes', LARANJITO_UNREAD>0);
  btn.classList.toggle('no-new', LARANJITO_UNREAD<=0);
  badge.textContent=String(LARANJITO_UNREAD||0);
  badge.style.display=(LARANJITO_UNREAD>0)?'flex':'none';

  panel.innerHTML=`<div style="display:flex;justify-content:space-between;align-items:center;gap:10px;margin-bottom:10px">
    <strong>🍊 Alertas do Laranjito</strong>
    <button class="btn soft" onclick="clearLaranjitoNotes()">Marcar lidas</button>
  </div>` + LARANJITO_NOTES.map(n=>`<div class="laranjito-note"><div>${esc(n.title||'Aviso')}</div><div class="small">${esc(n.msg||'')}</div><div class="small">${esc(n.dt||'')}</div></div>`).join('');
}
function toggleLaranjitoNotifyPanel(){
  const panel=document.getElementById('laranjitoNotifyPanel');
  if(!panel) return;
  if(!panel.textContent.trim() && !panel.querySelector('*')){panel.classList.remove('show'); panel.style.display='none'; return;}
  panel.style.display='block';
  panel.classList.toggle('show');
}
function clearLaranjitoNotes(){
  // Não apaga as mensagens. Apenas limpa o contador vermelho da bolinha.
  LARANJITO_UNREAD=0;
  try{localStorage.setItem(currentSessionNoticeKey(),'1')}catch(e){}
  renderLaranjitoNotify();
}
function showLaranjitoOncePerAccess(){
  // A bolinha aparece uma vez por acesso/login. Não joga mais toast sozinho toda hora.
  try{
    const k=currentSessionNoticeKey();
    if(localStorage.getItem(k)==='1') return;
    const panel=document.getElementById('laranjitoNotifyPanel');
    if(panel && LARANJITO_NOTES.length && panel.textContent.trim()){
      setTimeout(()=>panel.classList.add('show'),700);
      localStorage.setItem(k,'1');
    }
  }catch(e){}
}

function saveSession(){try{const data={usuarioAtual,exp:Date.now()+30*24*60*60*1000,version:DASHBOARD_BUILD_VERSION}; localStorage.setItem(SESSION_KEY, JSON.stringify(data));}catch(e){}}
function clearSession(){try{localStorage.removeItem(SESSION_KEY);}catch(e){}}
function restoreSession(){try{const raw=localStorage.getItem(SESSION_KEY); if(!raw) return false; const data=JSON.parse(raw); if(!data || !data.usuarioAtual || !data.exp || Date.now()>data.exp){localStorage.removeItem(SESSION_KEY); return false;} usuarioAtual=data.usuarioAtual; return true;}catch(e){return false;}}
function currentUserKey(){return usuarioAtual?.tipo==='master'?'MASTER':(usuarioAtual?.login || `${usuarioAtual?.nome||''}_${usuarioAtual?.filial||''}`)}
function currentUserKeys(){
  try{
    const keys=[]; const add=k=>{k=String(k||'').trim(); if(k && !keys.includes(k)) keys.push(k);};
    add(currentUserKey()); add(usuarioAtual?.login); add(usuarioAtual?.nome); add(`${usuarioAtual?.nome||''}_${usuarioAtual?.filial||''}`); add(`${usuarioAtual?.nome||''} (${usuarioAtual?.filial||''})`);
    add(normName(usuarioAtual?.nome||'')); add(normName(`${usuarioAtual?.nome||''}_${usuarioAtual?.filial||''}`));
    return keys.filter(Boolean);
  }catch(e){return [currentUserKey()].filter(Boolean)}
}

const loginScreen=document.getElementById('loginScreen');
const app=document.getElementById('app');
const userBadge=document.getElementById('userBadge');
const masterTabs=document.getElementById('masterTabs');
const mainFilters=document.getElementById('mainFilters');
const topMural=document.getElementById('topMural');
const inicioSection=document.getElementById('inicioSection');
const listSection=document.getElementById('listSection');
const metaSection=document.getElementById('metaSection');
const servicesSection=document.getElementById('servicesSection');
const logSection=document.getElementById('logSection');
const detailScreen=document.getElementById('detailScreen');
const avisosSection=document.getElementById('avisosSection');
const telegramSection=document.getElementById('telegramSection');
const senhasSection=document.getElementById('senhasSection');
const histSection=document.getElementById('histSection');
const reativacaoSection=document.getElementById('reativacaoSection');
const aniversariantesSection=document.getElementById('aniversariantesSection');

function R(v){return Number(v||0).toLocaleString('pt-BR',{style:'currency',currency:'BRL'})}
function pct(v){return `${Math.round(Number(v||0))}%`}
function esc(s){return String(s??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]))}
function trunc(s,n=18){s=String(s||'');return s.length>n?s.slice(0,n-1)+'…':s}
function filialLabel(f){return String(f||'').replace(/^F/,'Filial F')}
function flattenVendedores(){const out=[];Object.entries(TODOS||{}).forEach(([f,arr])=>(arr||[]).forEach(v=>out.push({type:'vendedor',filial:f,...v})));return out}
function thirdChargeEntity(){const src=CLIENTES_TERCEIRO||{grave:[],alerta:[],atencao:[]}; const rsrc=RECEBIMENTOS_TERCEIRO||{grave:[],alerta:[],atencao:[]}; const pend=[...(src.grave||[]),...(src.alerta||[]),...(src.atencao||[])].reduce((a,b)=>a+Number(b.pendente||0),0); const rec=[...(rsrc.grave||[]),...(rsrc.alerta||[]),...(rsrc.atencao||[])].reduce((a,b)=>a+Number(b.pago||0),0); return {type:'terceiro',login:COBRANCA10_LOGIN,filial:'FTER',nome:COBRANCA10_NOME,is_terceiro:true,is_gerente:false,pendente:pend,pago:rec,total:pend+rec,perc_filial:100,grave_pend:(src.grave||[]).reduce((a,b)=>a+Number(b.pendente||0),0),alerta_pend:(src.alerta||[]).reduce((a,b)=>a+Number(b.pendente||0),0),atencao_pend:(src.atencao||[]).reduce((a,b)=>a+Number(b.pendente||0),0),grave_rec:(rsrc.grave||[]).reduce((a,b)=>a+Number(b.pago||0),0),alerta_rec:(rsrc.alerta||[]).reduce((a,b)=>a+Number(b.pago||0),0),atencao_rec:(rsrc.atencao||[]).reduce((a,b)=>a+Number(b.pago||0),0)}}

function emptyRec(){return {grave:[],alerta:[],atencao:[]}}
function recTotal(src){return ['grave','alerta','atencao'].reduce((acc,fx)=>acc+(src?.[fx]||[]).reduce((a,b)=>a+Number(b.pago||0),0),0)}
function recebimentosSomenteConciliados(ent){
  try{return mergeRecebimentosConciliados(emptyRec(), ent||{})}catch(e){return emptyRec()}
}

function crediaristaEntities(){return getCrediaristasConfig().map((row)=>{
  const filialKey=String(row.filial||'').toUpperCase();
  const loginKey=String(row.login||'').toLowerCase();
  const filData=FILIAIS?.[filialKey]||{};
  const nomeCred=String(row.nome||`Crediarista ${filialKey}`);
  const recOwn=recebimentosSomenteConciliados({type:'crediarista',login:loginKey,filial:filialKey,nome:nomeCred,is_crediarista:true});
  const gRec=(recOwn.grave||[]).reduce((a,b)=>a+Number(b.pago||0),0);
  const aRec=(recOwn.alerta||[]).reduce((a,b)=>a+Number(b.pago||0),0);
  const tRec=(recOwn.atencao||[]).reduce((a,b)=>a+Number(b.pago||0),0);
  const cli=CLIENTES_CREDIARISTA?.[loginKey]||{grave:[],alerta:[],atencao:[]};
  const gp=(cli.grave||[]).reduce((a,b)=>a+Number(b.pendente||0),0);
  const ap=(cli.alerta||[]).reduce((a,b)=>a+Number(b.pendente||0),0);
  const tp=(cli.atencao||[]).reduce((a,b)=>a+Number(b.pendente||0),0);
  const pendCred=gp+ap+tp; const recTotalOwn=gRec+aRec+tRec;
  return {type:'crediarista',login:loginKey,filial:filialKey,nome:nomeCred,pct_base:Number(row.pct||100),is_crediarista:true,is_gerente:false,only_cobranca:true,
    pendente:pendCred,pago:recTotalOwn,total:pendCred+recTotalOwn,perc_filial:100,
    grave_pend:gp,alerta_pend:ap,atencao_pend:tp,grave_rec:gRec,alerta_rec:aRec,atencao_rec:tRec,
    grave_alvo:gp*Number((CONFIG_META||{}).grave_pct||20)/100,alerta_alvo:ap*Number((CONFIG_META||{}).alerta_pct||15)/100,atencao_alvo:tp*Number((CONFIG_META||{}).atencao_pct||10)/100,
    grave_perc:gp>0?(gRec/(gp*Number((CONFIG_META||{}).grave_pct||20)/100)*100):0,
    alerta_perc:ap>0?(aRec/(ap*Number((CONFIG_META||{}).alerta_pct||15)/100)*100):0,
    atencao_perc:tp>0?(tRec/(tp*Number((CONFIG_META||{}).atencao_pct||10)/100)*100):0,
    perc_meta:0,rentabilidade_pct:Number(filData.rentabilidade_pct||0),sem_ativo:Boolean(filData.sem_ativo||false)};
})}
function crediaristaLoginByFilial(filial){const want=String(filial||'').toUpperCase(); for(const [f,l] of Object.entries(CREDIARISTAS_MAP||{})){ if(String(f||'').toUpperCase()===want) return String(l||'').toLowerCase(); } return ''}
function crediaristaEntityByLogin(login){const key=String(login||'').toLowerCase(); return crediaristaEntities().find(x=>String(x.login||'').toLowerCase()===key || String(x.nome||'').toLowerCase()===key || String(x.filial||'').toLowerCase()===key)||null}
function flattenFiliais(){const out=[];ORDEM.forEach(f=>{if(FILIAIS[f]) out.push({type:'filial',filial:f,nome:filialLabel(f),...FILIAIS[f]})});return out}
function metaPureKey(ent){return ent.type==='vendedor'?`${ent.nome}_${ent.filial}`:ent.filial}
function metaCanonKey(ent){return ent.type==='vendedor'?`VEND::${metaPureKey(ent)}`:`FILIAL::${metaPureKey(ent)}`}
function normMetaKey(v){return String(v||'').normalize('NFD').replace(/[̀-ͯ]/g,'').replace(/\s+/g,' ').trim()}
function metaAliasesFromRaw(raw){
  const val=String(raw||'').trim();
  if(!val) return [];
  const parts=val.split('::');
  const hasKind=parts.length>1;
  const kind=hasKind?parts[0]:'';
  const pure=hasKind?parts.slice(1).join('::'):val;
  const alias=new Set([val,pure,String(val).toUpperCase(),String(pure).toUpperCase(),normMetaKey(val),normMetaKey(pure),normMetaKey(val).toUpperCase(),normMetaKey(pure).toUpperCase()]);
  if(kind==='VEND' || (!hasKind && pure.includes('_F'))){ alias.add(`VEND::${pure}`); alias.add(`vend:${pure}`); }
  if(kind==='FILIAL' || (!hasKind && /^F\d+/i.test(pure))){ alias.add(`FILIAL::${pure}`); alias.add(`fil:${pure}`); }
  return Array.from(alias).filter(Boolean);
}
function metaAliasesFromEntity(ent){
  const pure=metaPureKey(ent), canon=metaCanonKey(ent), old=ent.type==='vendedor'?`vend:${pure}`:`fil:${pure}`;
  return Array.from(new Set([...metaAliasesFromRaw(pure),...metaAliasesFromRaw(canon),old]));
}
function mergedMetaConfig(aliases){let cfg={...CONFIG_META}; (aliases||[]).forEach(k=>{ if(CONFIG_META_IND[k]) cfg={...cfg,...CONFIG_META_IND[k]}; }); return cfg}
function entityConfig(ent){return mergedMetaConfig(metaAliasesFromEntity(ent))}
function calcMeta(ent){const cfg=entityConfig(ent);const gPend=Number(ent.grave_pend||0), aPend=Number(ent.alerta_pend||0), tPend=Number(ent.atencao_pend||0);const gRec=Number(ent.grave_rec||0), aRec=Number(ent.alerta_rec||0), tRec=Number(ent.atencao_rec||0);
  const gAlvo=gPend*Number(cfg.grave_pct||0)/100, aAlvo=aPend*Number(cfg.alerta_pct||0)/100, tAlvo=tPend*Number(cfg.atencao_pct||0)/100;
  const gPerc=gAlvo>0?(gRec/gAlvo*100):0, aPerc=aAlvo>0?(aRec/aAlvo*100):0, tPerc=tAlvo>0?(tRec/tAlvo*100):0;
  const sumW=Number(cfg.peso_grave||0)+Number(cfg.peso_alerta||0)+Number(cfg.peso_atencao||0)||1;
  const geral=((Math.min(gPerc,100)*Number(cfg.peso_grave||0))+(Math.min(aPerc,100)*Number(cfg.peso_alerta||0))+(Math.min(tPerc,100)*Number(cfg.peso_atencao||0)))/sumW;
  return {cfg,grave:{pend:gPend,rec:gRec,alvo:gAlvo,perc:gPerc},alerta:{pend:aPend,rec:aRec,alvo:aAlvo,perc:aPerc},atencao:{pend:tPend,rec:tRec,alvo:tAlvo,perc:tPerc},geral:geral};
}
function getBonus(cfg,perc){const p=Number(perc||0);if(p>=100 && cfg.bonus_100) return {thr:100,text:cfg.bonus_100};if(p>=85 && cfg.bonus_85) return {thr:85,text:cfg.bonus_85};if(p>=75 && cfg.bonus_75) return {thr:75,text:cfg.bonus_75};if(p>=50 && cfg.bonus_50) return {thr:50,text:cfg.bonus_50};return null}
function renderDeltaPill(delta,perc){const up=Number(delta||0)>=0;return `<span class="delta-pill ${up?'up':'down'}"><span class="arrow">${up?'⬆️':'⬇️'}</span><span>${pct(Math.abs(Number(perc||0)))}</span></span>`}
function renderPiggyBank(perc){
  const p=Math.max(0,Math.min(100,Number(perc||0)));
  const positions=[[30,110],[60,106],[90,112],[119,108],[45,86],[77,82],[108,86],[60,58],[93,60],[76,132],[106,132],[45,133]];
  const filled=Math.max(1,Math.round((p/100)*positions.length));
  const falling=Math.max(1,Math.min(3,Math.ceil(p/34)));
  const baseCoins=positions.slice(0,filled).map(([x,y],i)=>`<span class="mdl-coin coin-fill" style="left:${x}px;top:${y}px;animation-delay:${i*0.12}s"></span>`).join('');
  const dropTargets=positions.slice(Math.max(0,filled-falling), filled);
  const fallingCoins=dropTargets.map(([x,y],i)=>`<span class="mdl-coin coin-drop" style="--tx:${x}px;--ty:${y}px;animation-delay:${i*0.45}s"></span>`).join('');
  return `<div class="piggy-bank" style="--pct:${p}"><div class="piggy-glow"></div><div class="piggy-shell"><div class="piggy-slot"></div><div class="piggy-falling">${fallingCoins}</div><div class="piggy-fill">${baseCoins}</div><div class="piggy-glass"></div></div><div class="piggy-label"><div class="pct">${pct(p)}</div><div class="small">Meta geral</div></div></div>`;
}
function toast(msg, type='info'){const host=document.getElementById('toast'); const el=document.createElement('div'); el.className='toast-item'; el.innerHTML=`<img src="${LARANJITO}" alt="ok"><div><div style="font-weight:900">${type==='success'?'Tudo certo!':'Aviso'}</div><div class="small">${esc(msg)}</div></div>`; host.appendChild(el); setTimeout(()=>{el.remove();},4200)}
const _toastOriginalLaranjito = toast;
toast = function(msg, type='info'){
  try{
    const s=String(msg||'');
    if(/parab[eé]ns|ganhou|meta|laranjito|comemorando|liberada|ótimo ritmo|otimo ritmo/i.test(s)){
      addLaranjitoNote('Tudo certo!', s, type);
      return;
    }
  }catch(e){}
  return _toastOriginalLaranjito(msg,type);
};

function mascotCongrats(ent){const meta=calcMeta(ent);const b=getBonus(meta.cfg,meta.geral);if(!b) return;toast(`Parabéns, ${ent.type==='vendedor'?ent.nome:filialLabel(ent.filial)}! Meta em ${pct(meta.geral)}. ${b.text||''}`,'success')}
function makeKpi(label,val,accent,sub='',extraClass='',mascote='',iconHtml=''){
  const raw=String(label||'').trim();
  const p=raw.split(' ');
  const emoji=p.length>1?p[0]:'';
  const txt=p.length>1?p.slice(1).join(' '):raw;
  const icon = iconHtml ? iconHtml : (emoji?`<span class="kpi-emoji">${emoji}</span>`:'');
  const masc = mascote ? `<img class="laranjito-card ${esc(mascote)}" src="${laranjitoSrc(mascote)}" alt="Laranjito ${esc(mascote)}" loading="lazy">` : '';
  return `<div class="glass kpi ${extraClass||''} ${mascote?'kpi-laranjito':''}" style="--accent:${accent}">
    <div class="label">${icon}<span class="kpi-label-text">${esc(txt)}</span></div>
    <div class="value">${val}</div>
    ${sub?`<div class="subline">${sub}</div>`:''}
    ${masc}
  </div>`;
}

const LARANJITO_TRANSPARENTE = 'data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAFUAAABaCAYAAADJoxqPAAA08klEQVR4nLW9efBk2VXf+Tn33bdk5m+prWvpfVdvai0tJCEhJCRACIWEQTACHIQdHgO2YWyw/8CDI2xjHDjGEQzjBZtB4EEeHCCjGANCslZQo61ba6/qRb1WdVVXdS2/X/2WfNu998wf976X+asWomeYyYiqzJfv/V6+PO/cs3zP95yUU1dcobzEx59mGfeYjC9nwpYxgJKlfWJAdHjWcdsYUF3sR0CGv0HH1wgw7Ne923/VZ/1m24AiCBCW3te4A2VxXFAZtxXQAEFkPC6ExfEBwb4UYZ4Ww7/Jc76YGSQJJUMRASOk96IgTdo/HCcomCRIuUSospDn8EL4/+eh43/jE5qkZoY3ZVmoy4JUgg6vBRUIqlGgCJKlfQpG9S8X6seyjN+0lk0jo2YZFJO2jYmCMwgiisjiOBHFSFQRSSqYNpNA4/HDN5Jhh16y/f9GRVXQS86nkraDxiuU9Kyg6ThN+0MSuoosXZ0SBDRokkK8KUFZfK58C03dEuF9NucjNkNkr2YakSSwtK2DQAFRDElT03GDgJOiMlyCyGJLBv2RwSSM+vRXfo7ap4t3s/QJKnuO0kGQBoym/Qoh/f2gjSETgkbtNRqFGtL+AN9cqB/ILB+wGReNLJa2JA0dn2WhtaKIkbiMDBgCYiTaVFkIlEHYg4YmW4sutkn7/z+1qYF4IUrUVl3W0Kh5UUNBjaCaNDLZTAnxwkMY/j5ui0JY2g4ar32PUO8Xw2/nlodMttDKQZCGhWaKYJKgZdhOmhuFn26GWdbawa7q3mcYrNceeyrLG7x0W3up11WF6E113KdJ4qP9NItnHTVOF5qaMQo6aFKWpNEBSe8rIQl+FOr7spzftTYJR5ecj4yaKklDDQFjkmDNkubKQpNlaf+g0TLYtGRjB41cRAWaNCCpuC5tL9ms4f09z0v79ZL3dXgeBa0Lbz/YbE0aagwaomKEEFARJHn7UQ4h/l0Ii1uoKggBMQb7lAi/agoetmbJXkYbaoxgJJogY+IHGVm8P/5DMGZhYyUt+UHLR5u6ZEqGr7jQ4EEbNQlZl7R4WQdf2vOol6qoDO/Ii8IlRUcPH22molnSvGQK1CxsaAhRc0MyWaJRyDJobVDsz+cVG8JCO5Mmiomal41LOgooM1FAmYmhSJb+zgwaaxhNw7AtsBTHLhyWsLTMk8ZKslGSluRf1aaOS1xSNLD8XvrooClUSsINISpHSMeE4eYHQTJFPIsoR0A0XXyyyXbDCJnqaAcNSiZ7NTRLyzgTyJJGZulGZIPtHP7WLNnW4WbB6MT2autCQ/do66i1g7wvtZTf+rFsnRcCXAh0Wahj7Jm8UiDa1qCy8Oiq+CBJQ+PxXpWggleQTCFI/H5BsaKaVOSSZTou4fj8J3d/VfI8p6oq6rpGVamqihACbdNQlCXWWuq6BmA2ndK2Lb1zTKfTeFzbkltLXuQ0TYuGwHQ2pet62rZlZTZDgXo+J89ziqIYz1dVFc57+q6jKEuMMezu7jJc03w+B1WqyQTvHG3XUabj6rrGGMNkMqFpGpxzrKys4FxP28bjsiyjqWskHde2LX3fM5tNCUGZz+dUVUWeW3Z35xhjKMsS1/f0zlFVFQDf9Zpb1EahyahVmciSNkZbmYmQ5znGGJxzZFkWjXoIqCo2t6gqPu0DcM6BCLm1eOdQSPsE5xzGRLV1vQMUay3OOwAya1Ggdw4xMYtz3qOqZFlGCB5Uya1FROj7HmMMAnjvUeI+1UDwis0yxBi8j5+b5znOOVTB2nTt3sfrS+eDuM+54bNyVJW+j99RRPDex+OyjBBCvPbkYzAyhDSaYsql5SxgJFBV8a53bUtmDEWR4/qeEAJlWSFA1/cURU6e5zRti4hQlCW9c4TgKYoCRem6nizLsNbSti2qURO7rqfveoqiAKBrW6y1ZJml6zo0BPKiiFrfdZRVhRGhbRryPMfmOX3XoQpFWRJ8oHeOoizJMkPbxGsvy5K2bdO1l/i0Amyek5mMrmsxIhRFQdd19M5RVuViteU5WZbRdR0AeVHgnUvKotjonYVMlMzsfTaGUVPruokXVFW4vgcHNo93r2kajDEURUHbduNy9d7TNA1FOq5t28WycVEry6oiBE9d1xRFgYgsjquqUcvLskTTl4o3JB8/t5pMonapUhQFQZW2bTBZhhVDm25wVVU45+j7fjRdTdOQZRkmz+n7Ph5XRlPj2jZ+rip1XWOtpaoq+r5DiN/DBx8VzdqFpkbvFZZy8b1hz2AWNAEIxkQnMIYUycjHCMCgGkA1Lm9iiCLJc2kIMT5NLl9V0/nkxX8jks6n47ZCPIfEc2hYfNZwfZLOEXwYowtN+aUsXbsZrn0834uP0+G4FPMO1xRCTGqHkGq4vsGMWkmCGsKl5fjTmpiMZMLSnY3qn4tEmycS75hzdF1HUZQANE0TnVJZ0vU9glBWJd55uq4jtxZEaJoWazOqqqJtWxjO5x1tWv7AqKFlVdH3Pc71o/NqmoaiyLnQnNBTZx+nCZsIniKbspYf48jsZsnNKm26dpvM097zKXluUSWaCWspy4K2W6y8Pml5URSoKl3XkRlDXhRp5SmZKNZwCeAxpJtm0L64PTilzBhQ8LrQmBACLGkWyaEo4AdNSNoTtT0hPAGyzACCDwGTRQwteA9IXEbjChk00KfzGXxyDi/UT+vdT/4nntn5MrXbRMRTFJbClEyLVfZNjunta9/HLatvF6traAhkybGFMGi0jCmqMVn6rGH1aPweQGZMRKlER2erSRYgGAN2DNwlauWQ52ey+GeM0jYNNs8py5KmaQghUFXVaOfyoogOqq5RYDab0bUtXddRTSaoBtq2i1qe57RNE8OyyQTXx5BqOp0CUNd1DKnKkqap2TzxgJ790m9z5KbXs3Lbe2UIgXZ2d3jw3N36e4/+Uy70z0bPLgWSKavTipVqBUvGRvss95x/Hyfqr+gbD/w9VrNjMp1Ocb2jazvK5ITbZKPLtGr6vmM6m6JBmdc1VVVibU49nyPLIVXXU6aQyohiY+aRUsMEgAzhsxBShqUUZVzWXddhjCHLMpxzCIzLoe97bG6B6GwEGZeGAEWeE9KyMSn06rouRglFER0FkCfH1nUd7e4Wn/3gr9M8/XHu2HiMaw7fqfnlrxDf93z55J/p+776P3Ohew4xSpFlWBuYTiw1nkIEJ4ZqMiU3ltP11/mzs7/Gmw//fbXdtSJAXli89zGkshaU5NiiV++6GF4VQ9TRtmQ2XnvfdSCCTSEaS0qZvPwibx+11MRcPzOCtSk2c44sCXU0CTbGdzHWs2TG4J0HAWszvPfROSRBeu8jMGNMivVinDp+uSyiZN57qtkqN772Bzh5cYWnHnuCzUc+hHc9T555VP/93f+MZ8+cYHvLU9eBzjs8nrrr6YOncy3zds7WfJvduoNQcKF7ms+e+w22m7OqqtgsxqnB+xjrmvgdQciydO3ep7g3xrPGZIgYnPegMf7WEAghLOLUPWlgigJGQCPtr+d1/JKTGJZ0bUuR5xiT0TRNDI/Kkq5tabuOybRKoUhDURQxY2ma0bE578fQBoS6rinLgqIsadP5qqoCybj59W+X7/zRX+CxJzc48+RX6HfO8Lv3/m88e+oJdhtHawK+DPSZoham1YyimLI6Weey9cOsTdZZqWa0fUvdNBzfup+vbH0AMSaFVCbG010Xv2M1AVW6FFIVZcl8PseISfH0cFwFkpyoteR5Hn3SANuNgMmQVZGcFAOIMoRPg6TTrZDxPwYkI4Y7w02RMVwbHY8OTpERLhqROtUxLBogj8zmvPZdf1NWj72c408/xR9/4lf1kXs+y6qZMMkyCND3Cl5QbzCSkYcKowV9D7mp8KqsTlYRY9itt3nw3Kc4Nf+6sicsTNcq6ZMHWJDla2IREiZHOThiiCvcjpoJI765FwuI+6ezGc456rqmqqpo2FNQPZ1Oo/1rWqrJBID5fE5ZFpTllPl8HoPqScya2qYZs6F5ckrT6ZTd3V1EYu7ddR319kVKd17ZPk1z7gSvvPFyvvCVj1Ns/QH//OXfzebmJvc8cR8PmE3OrgXMYUUOZtTqKVcDO3WH97tsNzsYI1hjEANeDGe2n+PrFz7NldfejmsdQZWyjFnTfD6nKIqIKezOMSZ+x6Zp6OZzJtMJIShN25LndszQSDitNSQM1SyA5QUKxWhTm6ZGxIwxqZdoc9AYGUiK1/quBxkyIB3TOlWla5NTKotoP5PJGLKtooiv+52zmNOf1/L05zH1Cdh6lnzrBa72G+xeH7h2fYdD/R9xxRSuORZ40ynPNx6HZx7uuZAF3KzBru9y2Y1XcvCWV1Cs7UcUOlez0ZzhfPME85VdnrjwZV5z8HndNzkqmQjOpXg6XXvXdRRlTJmHLC8rS/o+OqWYKUaHZbMMRaKmjktz2esnkFgYSs4B5xzW5lhrafqesIRS9SmksllGnYLlcjajbVtc3493tu17cmtjzt+0hJRWOtfTdY7ZtICTd6s+9JvI9jfAX0CkQekp1wJXzVquyCdkmUGMAp5DKFc74bXzQLetbJ50nH7O8/zxFvf8s1zZXMNV7/kbVFfeIWVZUDdzmnauDz7+Mb504k+4cPlpDswux5gFMFNZS9t1uISwDcpRVRXW2mhfjcGmiMU5T5XnyRQE5Ntuu16tiaCzNUqenrNMyQVspuQS+J1PPSU+OZe8KDBi6PqkeXmO8w7vQgRDhDEfzrKMvutimSEhVs57ijyPeX7XkRkhb8+qPvDbmOMfhnAeCVtgHYiL68lqDFFsFhFsM7AzwuhgRT2EgLoAndCfd2yeEeb+MPu/9xfJbn2XFKuHyDIbY23fR+AlgAYlL/JFaJiuvW0TuFLG8GoAdVQV1/cRX8iyEdn6qXfcoktxajTKQ1FsLGkk1R2gLTPAfoQEHKSMCsGYmF2oLnLwwQFAyspYZFQxtwZz9iGVh96HnPoM9OcQ20DmQEPMSEQWpQNMFKSxsUpgcjTEL6Q2j+4hV5hAvhI4dLmD+gL9E/+U4B5VvfN/IqxeLTHWrjBiUDyYiAOM2VuCNkc8IkRHG5ayRpFlfMOMtS8z4KmRXTJURWVE7rPktZu2xYcwBsF915FZGx3WEgTWdR2u7/fYygipmRjoEzXbO49zjvzik6qf/1Xck58i7JxG/S5oB+oX9B8RXA/v/8yMD31lgsdCVuIc3P1Vz3/6ZEnjMyCHbILYCZgS8ilSTWBtRn6ZUO58iOyJf083v0BRFKOzNVk2YqwRDiwinnsplLmURfkQRnvb9RHKzPM88SDG0GW5ZLHMcYr4apHnGBH6LsJjNs9jYBwCRbInMaPKyayNSz4Zc+ccPoTksAK9i8vG7J6h+8g/wZ/4EtTnIdQYIjCBAc3iDe9b5d98eMrvf2Yfjx5fY3c3Q3vo6pLP/HnHB++e8i//S0XdK9gZmq1ANoNsitrZUn29Rs5/lOKp/6gDFpznOSFEn5FZixGh62JmZ62l7x3eufG4ruvIUiK0DGZ7H5VE9mhqUgoz2KfB+yspZcsRY3CuH9PU4AfkPwp1yIayLBsRLGvtokKQZSlj8VBfIHz+P6o/9w3oNkHbaA+H+6vQdsIDz1j+xR+u8rGHVlifWI7as6xWOSIleZZz40FPc3GDp84c4Jd+9xB33+c5fcHgmKJ2H2LXkXwN7AQyg+ic7NQH4fgfKOoi8h8CwfsIsogZqxbWZjHSCSGiZcolWWNMv5czKkGxsQi2VN6V5aLbYrtploDjdFdsbiGB1FmWUSTkfYDvgve03QDfLXJ+mwlu44zqoRtQseAbyALGRpSsD4YHjhd88vGce58qOLUh5NTctP8s73hlDgjqlayqeNOrZzz6zEU+8cwJdueXcX5ryg1XCNcdU157G1x7tEIsSOghKyA0qBGy5ll8c1ZbPSI2s4gwKkJVljjv6Fq/wDySmSjzcoFRDKaw70eQWmKcqmM5OgK3A2dKRzBlzHwGp7OcBS3flCHLWnJOI8CsEDTQ726B38XO1jA3v5n+5EPYx55FxIGA98JHHin4jc+VnNxUNLRcVjjefU3Hj70JDq5ZCD04D9OSYzes87M/HLjxIxv88fGGRy9Meea5KdNJyUNP9vzkD+7j6kMFZKvg5qhzbG5dxeTO78XanODncR+J8zksU2TJbbPIAtPrEeoMYQTs0yJPDBVl8ZTOMlrYpLaT6RTX9zR1TVlV5AkqE2OYTCd0XU/XLmVUu3OKsqCqJtR1jQCu3qSbX9CyLHBBsEXF9K1/n6AX4enfRz3MO3j40Y5y3nPbWsmhw4d5xe03cajK+MZz97C+tkN1zKKTGTS7YAy+Dex72Wv4jqsmXDh3nnrjHFdOL3DNvpw1AdgHUhDcFv3BH6V69d8kkOH6DlsVaH+RTiaU5QQNIeIQRUFRlcznNUaEyXRC07R0fc9kMgD2NbnNqVJGNcjPpqhorMMPqWnMtBYaORS5bJ4TgkcTqCwwwmOZtSPMlxfRzrrkGevt88w3T2lR5jjXU5ZTtG+5ePwBOPkk+zzghYlV/sdvz3jza1+B3vEubnj1G1g7cISsKDn3+AMc/+g/5oaVE2R5DjZnvpvT3fbTvOm9P8hbqlVEDPOtC1x44M+57MJ/Zn1igRwN27QbsHVhg/LQRSaHr0OMldC32KJCQ4t3FiRWW6P3d2Ploe+Tj8iyWGEFrI3HLVeHhcH7J8R72bbq8L+ApNAohEBR5HgfxgBZUqgEEQftUx18CKkiXqrsbp3Vtt6mqeuIxdabPPfQ5zn+1S+Stdvj50lhcIdexuXf91Pc8eZ30mUz/ts9j/DH9zzG6nWvlJVv+wnm8xxCIKhhbq+nPXQX/+fnv8G//dh9PHJmzr7LXybXf89PiN74N2gu9qirkeDIxLPxyP088+k/YOP4gwTXqRhD385VcPi+WYRUKQkoioLc5ktFyypVkf1YqIw2NTpoIdWoSLjpQluXWHxp/2w2wzuXwIZUrk4A83QyxbmepmkiHKaMRIfJdMrO9iZ9F6HDqc0IfUOzs8Gh6+/k8lteh33qSvjCLyNmG5eV1Ptew8rhawkof+s3P8lX2xXeeOuVPPL8PfqTd95B/8wRtH4en83Y4gp+5eNP8MFTGbdduR9XnmR9ZV2vP7omxdVvxJ37BHSnUTWwdjNX/PWf49y28NzjDyDFjH3HborFSu/IbUEwwu7unGXiyDJo1CeSRliqsJZlFUvjafkvNJUxkllo6qi4OmYLUeRLhnjJ7l66Pbzd1rvUOxcpy4pqMsX1HcXsANVsBT8/x/aJx/BtBxm4bEo7vQoV4Xf++xe4+/gW8rH38bcObXAun/Hfv34aUxxEDCgZXzkl/PnzLfUf/zo3Pf4JEM/v3/sQIXiy2QGkOIi2c0KnnPvGJrsXznH0hpdx61v+B9qmodm+oLaYislSZqYB2fOFFt9zqEPt+cqDXGTxN1YY4tSBAxU9oEkgy0DQHcKmSIlp0KCJYKA0dT3Wr9q2AYRJqsX3TU29s6ldM4dZhRgLkmEIbD71NZ7+o//M5Vv3snakARW8ZvSa4VzHF55vMM15ds+c4B//s3/NhVe+ncmbXoZWBjUZiLLZeHbOPQ/NDn/04U8yeUH59u9+O76vFQmosWgfEDrC+VM89+Hf4Ujbcuj2N3HoipsIvgeUoF6Dc6hMZDKZ0fUdbdunOpwmXkJOMaki2G4iYO16lxC2YvRJNgp7QSOMtyHSD3UEbJU8zwFNwX8WcclUThhIFd65WJ5IxlsQrM2Z72zQtbu0dcnO1nlym+F3tnj4Dz9A+cCfcvD2QGZjQ4IYg5tv0jdzDkwM2ZGjVJMJFzY2ma8fZsVdhG4bKQJSFKzlW0yvvIX5dIp1HbvTfVy2XuH7Br+zgbhNjLVIlnPgCsvTn/gij9cdppxy8ObXUa3sEw0BVY8GB8HRj+WUSPtRHWhEMZYdalnLlKWQOAgqQwIlwpBZDQt8jASImEAsFRi6LmZUMU2NMVqeEKchIcjzPNnbWDDr2xrverquwfctBI/4lvDsE1x7REnKG5dOmFPOT9BsX+DHX3mE8tqbse/9O3Q/8jNMr7qGdxzryPrnIReyTLnrsk1uX3OU/+BX0J/4h8y+/c28545j+HYHv/kNsu5kpGXmEyaHCm64LnD881+jfvJedi+cou9qDb5X9SGBQf3oK/I8p+96vPcU5YB5RAdtMkPvUkQ01NdCwCw0lUTNXliRPUjAWGEciA4+aqW1C5A6yyKg0i5lVCGwu7NNU+8QvIuMWe8QzfFtzZrZpaqi2UEUKYWibDlYP8yZU1dy5MqX82s3Oj4wvZwcz9+98gKXX7yb2XQbzBSTFRzbv82/vOkpPrA55fmrZrzzZSWvOiD0GyfInvswZbEB2QQIkMOhq+DoqkN3ztJtn6Nf2Y8tJjFxUY8ETzUr8T7EOlxZgEJTt5jMYCtL1yVZFCU+BLo2gtRD8mNhKUZdcjBmyRgPtB9JBAc/UGxIfM7heJHRhER6TOD86Se13j6H9g19W+HKgl47tp9+nBWzS5al8wQhUzCZ47LZs4Tdj7L5zAu8ZnYtr79xyoTzTDfvZV/4CmaWoVmBoExXDNf5B/iZ6iKuug6bnUee2yVvH2OaPYWZHQSxEV/wHXYmHN2v1OfPkG+eoVw/nOC6jOAdeVZqMnwE1YTiDzSftJqH5GjIGokZ1pCBWoSRh6TLcaoOMUDU1EFDm8RFHUKqZZ5m13WUCQ5rEhNve+M0xy6/gnOnT7CzeZbMz3Gh4/mvfYH97Q69i80HoRNCExmGxbTn8tnTHNTT+G6KdJDTYM0uZhbQLEf6BvpdJMtZXS+Y+Kfw+jziDZIZ7FShFWh3Iw7bnkWCh0I4dlR49JmTFNc/hVk5yFSVvJxg8yqy+eptbLkaHe9QAU70oK5rxorHAMQXRTGGVKOmLnj5CaBm6H9K25J4n6pkl9TJgUTTSdqZqD2RHhQozz3Lls259sY72Dx7krNP3se5Rx+i/uLnOXigJ3ih68BU4FuQTMAHTOWZZB7YjQVJrykeCRAcmADbm2jdgC2x031Y0yfCQgl9HW2kb6G5CH0XwyX1HDgmbHzqNNnjDzEzM9a6jgNX3Mhs337p60010xUG/u2QNY4g/cBVWOIxhOATYK3L3n8ADUY93WNXVRdc0bIsaFJ2VaWQKtJ+coq8oKlrkBhSbT/2OZ1/6JfZyQ5z/NVvZ//1t3DVjXcwf/xx8tBgDDinuF7o6gQitYIpQWqNJRSJrTchCOoCwRu8CziF3nd4CYjsYuwFZusVdv8MsUTky3eQIhRxHog3v5goR2cNT93/GDdefiPVzXchwPZTn9fJyd+jesXf1f6yN0vXdUwmk0gXrWuKsqTIc+qmSSWWEud6+t5RFsXCpg62dK9NXbRCDqajKMvolLpuJOy63i2cUkpJbZ4jAvXFc8y/+Lvsl3OsdS9w9tPP8cinjnGmLSjOPMVVhSMEwfuoRLaEvk5IUEZsxenjjd3dhrM7K8zNPsLkIDLZB/kMs74fiinB9TSbZ+DxR7m8PMXRKxx2xaJZXC0Ej3ofzx0CxsCxy5TTJ7aZP/UEL3SeaXeWy1eeoDp4AR7518i+V1CUB8YUPHJRA533S7SkFklYgUur9Zto6iKfGnP/tN/aCCQ45xKrOhurj5m1hD6yNgZeVfPcw9o/8Rmmybu39Sb3PbjByQ3l5kNQHIqNCq43ZFYJPq5q70BrxdgY871wPufr/W1Mbnodq0evZeXQ5VSr+zBZTl5WZHlJX2/TNXOOP3gvn/zkf+ONW09xw20tZhqX5EDsQAVNDQ/71pRZu8Xzf/rnzKef49brHNPbFVZmsP117AufVr32R2QIr7KyxLUtzvVMUhXA+0BuDVlmRi7VN/H+SzaVvTZ1d3eOtZbZbEZT1wTtRoO9u7NDUZZMJhHmQxXZOI4/f4KQxwhhe0exKOuzipVJh2TgFLpOsVbpG8gLoWuUolA0wMU5/PzvtTy5+RAHDj5LPq245tqrWF2LXNOiKpnvzDl9/CRbF7c4d36btnW89gr4X1cNB66T+A2DxJXvPHgD3lNNhYOryvPnOk5vCTccEZoNWF2Zgzj8qT+jPfwOprMVVGF3Z4eyqqiqMhIssozpZELvHHXdjI0U38Km8iKbam2GSQG+yTJENUKASsq2FuUUQamdo+0gcwHvwDnh2ttv4VXf+SNsffp95N0FxPQEgd5B3inNDhRTcBKpRr/zGeXhUwA955oLHMjh6eOncArbDlZyuCyHSqBzkHXR0d97Aj7xZeWHjyjZFHC6qGaHAN5gMmVl3XLFy2+iPnGC85s77NsQJque3LRIs4M14F38jjbPQRXnFg0Xg/NebqRQoiv4pnHqAjqI70eevk9hUxkbGNLSKBO5q+/7CFKrYq58BebKV7Hz9L2QTzn81ndx8Ht+mur6V8nT9RM6v/9j6PwisfwNXRdL+r4DY6H38NGHlMLAuoVvW4eXr0tMXS280AUOToRj05KiKLmw0/LpEy0fPRPY8fCRrys/8JZAZhkbdumjCYj0ckN16DJe9QO/QNMqZ//rL7O7+zTVeZhphr39FvJqSlOnfoHUBuQTqU41Bf02xxY5XQKpv7VNvSROHWtURZEclI7tMk0dmXMxo4oU87Vr75TqZz/A/Lmv6/TwNdj1oxLEIFnG4e/4YZ578gv4fo7vO4xXghPaHsQK1innPAQPUwu3r8CbrxVuvVU4coOhOihQzRBTIMUUdSXzp8+y/iXH6XsDX9qE8zWoU9SnVKYPSBA0Edm8VJQ3vpbqpm9j/bJr5NDLX6cXPvHv2H32c5iXvZXJnX9HtI3EEWAElGKFtWfovFlUWBdFlLFaNbaL/wVx6kALN5lBnYOgiF3OKGIM16dalikKspX9TG96vZRlgQbFdT0mg5UbXiNrt79Ft77yJygXUQ+9V6QHlymuEI4etXzHHQ6rgXe/Em67wZBnSnCBsAHBzUF2yaptzErF7MY1Xv6ydX7227f58Gd3uXa/Uq74pJ0BcfELBy/4zuCz/UxufRvZ+lERY8iPXCcH3/u/gCwIyME7ioGHGwKSeA6LrDG1q7+oRqXLGVVc7uGbxKmL7r42hU2S6twLPGBAykHHRgpblqlFZkFMC9U+1l//wzTP3kf7fE3QBoPSO9AGVJRsw/OL329QB6FXuvPQk0gecX0lpMwR6h2yzW1MDrftK7jtPVW0nTtz6KLTU5UYXXiDCwV63Ruo7vx+KaYz+r6PSNTQSJGaLIqiXGAeqV3IJQ7AQLhbJmIMmjqiVJfGqYPUB00eZoYM9JYBC4BFtjFUWlWH9p9hnyTAYoELTG9+g6y/4Ucx01Uks3gkBvQOmho2zwQuPufYPRdod8A1im+V4GN2lZnY55WlBmO8oD2w0yEbO+jmLjrXmP72gmuhb4XOFbTVNeRv+Udk60djrV4ipqFJ68xSJTiONVnKqCTRgxLtZ9GStICvzfDHy3jqEKeGUVOj5mmq3wxaafNE+2kj87kocrqEAcQQI9ERiyXaj8Q+ADWW9bf9tKzf9W6y2RqSWie9B5fi1WYObQN9C30DfS+4Tuga6HajU3Ot4ttYtQ4d+EZiMrUl9Lsxoehq6BrofUFbHMO++9eYXftqETF0bUxm8qLA+dTBWC3qa0VRYvNiqeVo6GCMtB9JrB1rbcSVGeNUidDbi+LUBRpTFiUI9AlPHZZ/FGYBqqn6GPtPu7YDIibpEp8zT1DhwJYzmeHQe/65kJc6/+of4TfPYHwsaXifsKIuCtkI2FzIughoOxNtsDGMgx6GDsU00QMfYtoYAjip0EM3M/2uf0hxwxuk69p4fUW+1Gtgx2Utwoingi5xyEKE+UhNIIDNbQy9khKawWbuiVN18Q+NmZ7N7SjIkfYz0HlsKkc7R5ZF7e1dJP9aa/GJcGCTNroExmQmg2LKwXf8vKx/19/G7D+GFCVYGwcSIDgvBA/OC20T6HpoGqVroWuUroG2hnYeAal2N9IBmh1oW6g7y1zW6Q+/kuK7f5HyzneJLUpc79AQsHZoXo7XNDQ1C5HU7PyC9qOq9ClON4n2o6SGXx0aoC/JqKJqRq0djKqmSGAoLVeJOt73PWXqA22aJoEtybADk9Qf1dT1WK5uRlCmHJ1DVU3w+RFWv/Nvy+y6u/TM//Uv6E48jPgGUR8nPoRYO1OBro9a6xykpsGFfU9FOIE0KSInlGsUd/41Vt76D5gcvV565+nqmslkgvd+bPQwJlKWBhp93zv6tqVa6k0thl6xpZ5Y51zsD0s8ByEF/6NNDbECMOKpo+Yu6gDD69E0jOlX/AMZ7w6jYxuclAzLYPkh6bhyyuyOt8k119/Fcx/4Zd345P9B3tfkeXIWwkgzCsRlr31UgACYdH4NA3iuZOuXs++Hf4Xitu8RsQWIQSQsqEtLoRCLy15+MX695eOGV4uaHnuOt6OfX6r7a9qOUFbME/Y0UpQlpijGDrk4CKGLFO5E+6nnkfZTlhXzeaydVwnoHfgBIrHhYmikiIMQLG52FU+c9FRBWZ/C2nrJZLUA7ccqBEYJmNELB0C9YevinL5XXICDByccPHQNk5VVvA9jE0hZluxsb2PzfBysoKpjp/bQSDGpKuZLdf+maWjblsl0igZ/SRW5HaVqDwZlc9QAiXeYheYOlYEBAS+KItrIwc6wqF/ZxFCBCBXqENcmB9Ulmnqe50st50XEa7vYYunqTXYf/DhrpqXzcH5bWH3993HkHe/Fb79At3EKabdR18WGt3JKue8oYXaI+dmznPvtf0Xhdrn+hozDV7xAvvUAXXsHSMwGg/e03kcBAl3XjsySSx1v1/cjrjEwVIqiiK35MNLU+0RtGjX17U3g91ciQLJHU1PdBeIgrD4V+sqiWIRXQ1mh6yiKYu+4j9mMtmnonYtA73IPq7U0bYuqUhSJH9B38W/OPav67Be58ojQ90pjDnD0LT/Evtf9oHjf09Y1RZ6TZYb5fI7NC6rJhLrtyM+d4siXP6XlM5/m2NUZ5SEhbH6Jpn4nplynmFS0bYNznpXZjN45ui6O6jBicG0zsqXbtk2NFLEJpK7nS40UNUYMRZXT9zEhKBPlUgH7Q63n05VwNk+z7BSyFOiHkMYPBZhOp7FGlVTeGBnjzsmkwjlPmzrkhLisrY3ki4GzWk0qvIt1riErq+sam2VMJlOapuXigx9npdhlug59Df2RGzlw53dKm+phk9lq7Dj0gena/uhsmjhXJT98JfvueD39858lMx7EYba/zsQ69bmVsZG4KJnX80gOGQYrqJLncXJGPTjlqqSpm/Qdp/G4rl5gxnUc1DA43iGMMgDvbAcOasr2ReLrIbMQRrLAcgv3kImNr/dUHmWwHeO+IXTbO0Rh0YxgROie+TJVAZMKqkpYu/Uu7PplDMMXhrh5uXIraXiDZJbZNbchkzV8T8weto6jrl7KlDQNThj4N4tr0hRDjo4oLA2AGD5reL3k5DSk72jMIk19e+O5qYteNCjkecHNt93Gt7/hTVx9zXVAbDgYmgeGSqK1FmOy0UgXeT7OQYnw2KKRYpnQVhQRK4gNFwWSGhpsbvEXz0GIgHaeQ3HFy8YGDkiM5tQI1yYTMsCSfd9T7D8KxYy+0Rh/NXPaJpLjyjRXpUkryhiJFdEhf/cO70P0B8mmFslcLbet92MjRWJZ94sS0xinArw3m/KviPHbd3/329jZ2uT8hfPccOPNXHbwQKRNAs5FtGqkurAAqYdJQAIjPGaTUwJGh+VcH3PoLIvdc5qShK5l52IDXVwtWQF2OlskEywoRs65cbqF6/vE15c05Uzo54FQx3KKFYMaE6/JCLnJR4p5NvYkhL2Tilia9oNGpxQCfRLgcJwgI0gdgamU+wPcdmGLu9rAsSuuoO977vnC5/n6ww/z0MMPcuDQodhMli4sG2k/kfw7CtX72NSVMMcocDu2po8ZVcrKIoG2R1HcyUf0qV98nR5uH2H/voygcSnNn/wazDfxfvFFY33IY4tol0fv6zs2n7yfZmeL3bmwezGD3JB/9Huxj/2uutRWPkQfA2VJQ0iIvhnPNyiOcw7vYrEvjKm4HTMviCOkhlRXB5s6PH5qy7P96BN8/OMfwwU/ZlRgxrJ0WcaQqu+68eTD8o+k35RtjX1Uw2iiFy//3rn4um85f+IRshtehz/6GmxRIVmstF685wM0j35KB4Jt17VkSTBd0yWQJ04P2j35hG5+8j/g203Ob0XEC53iV1+H7p7Dih9nsxRlGa+pjS08cSTU0Ee1DKgU2DxfACqJOBKSSYphWZRFnnpV9wh1qvDj2z6VG+Jotquuvo4LGxtjJjTAXWFIDRMUBgPcl6Cw5UxKFscNTQgxG1M0eE7d/yk9d99HmGuPffvPYu/6UYpqwurBkqMHPDqSFVIWkwDxoGHM2uIIuYANHesTuOLyjLo/yM4tP8POte/l4rnn2fraf9V+Z2OE++L5UjDOcE1L1J5wyZwYhs9NDm101AtnHC4VKsArW+UHd6J3v/mWW1lbXefBB+5nMp2OI96GeLVPZelJmn3SNA1lWVKWJfXuPDYgXDJ0QIC2jViBNfD0l/5I+52TXP22H+fmV70ac/9vsT1Zxa+tMt9s6bZr2HqBpp4nAsdkDO2mwzWlDIi+5dwLc86eU3YuOrbyFc6ceprs4fezcuv3s3L9q8jaU1plPXVdEzSM+XvXxW5vk5k0IyUC003T0Hcd08kU7/z4WZnJxmEPZVmlWDuBSJcKFeBdO4Gtm65j340384XPfZrduhm5QmOzGYyUmEXnWzbalbyIx/XOjRN93NJ4Otds05y4X6vnPkd18RG6Rzbg1X+dlRvvonjsv5BNesxh6Lc6uotPk4UOkZjNDPau7zs06Gjfd578GutFzb4ZrK0Ka/1TdA9fYPttv8DqYx8iXHgQXbsaufU9ao+9UdCYeps0oXyg8+R5zKgGvq0I9C46OWvsOI5uqEv5oZEC8Gr+4pnU7376PL9cf4rd3R1sFiunY0bVNoSwNEAxMYnzvEjtPcJ0Fgcodm1LNZ3E9u6miSPcul3qi89rvu8QB9/69zD9HHPyHvr73k82zbBlhjTbiE8xY32O3MSBscN0nswYdudzcmtjLX5es3vyGxGfCEK2BaurSul36O/9DfwV11O87uewR+5AiqkIPd7PaV1sSzJp9IcxJrbwJCQutqYH5glty/Oc+e4uJjNJQ5MPSXX/PSHVpY/VzS1+8vZX8uz33QqujeofPPMEgRljxkxpGEc0ACWKMl9qpBjHwk0mdF3LmWcf1mllKFb2kdmSbLIqrL4b8YWWD/8KtJvQtKgTvM/Y2mkIuzuU+Wqc4dL3dCHEz3UukRuEZ48/Qd7HEnQIEaRe8R2hDDSv+TnswWuQfCqZLWOy0DdxBJ2L2OqQKc2XRtAN3LDpdErfdcy7PsWxYcy8ymrRSBG+lVABbvrcffzekw/yQqH86D/61ZhFsKhlRfxq8WLkCSxlKozvxX8bZ45z/vknkcsOk9scMRnGxMli5qY3s3vfbzLrXiDYy+jWbqa+/h20q1cylaHwK+PnytJnNbtbuGtuRVf30z5xH/3289Tzhsb3uLvewYEAfTMns7lq5P+o6xuMIhQHl77U8vTVSx5DZURiIfHSfZKAqL90zv/3Xgy8/9AiR68mE7o2TpUYQerkeIpv0kgxOK/YSTfn+ONf0tA1bF/cAPVM+oa+nJJlsX9Vb/sxmt1T9IfupF89Rt00FOU609V9IDIOZJQ8j5qS8AUNnvXLb+FiVpBddjl9UyMXn+dis0Ox72q2N86hwdO1u+TlVLMsjY+RRieTw2KyxLdNEGWXBpCVl4LUCSoceLnex8lHeVEgAl5fwi9SvL5WvjKXPVmTmGiMfaL9DF1xQ51nPC45lMGwGyPsXDzLbFKmMfbZANsSYxiFa99A4zq6ziHFPtYvOyx5tQomfv5IsUkOapgQURQlr3zzD8nm2ZN67uQTnH76fnRllelsFamm42fIEqCOyfCuRzXgXYjlHdHU+Bs/a+G87PhZQ41qcMpDaQlewvIfHu/Y1jE4bts2jXHLxt7Usijo+h43zplir2FPM0eqyYybX/V2nrz/z5g/9wwH6m3W9h2ibGcYiTY6K1dZOXiVrB05giJxRhQLhsgAy3kf55v0LvbETiYTpvkqJrtWLrvieu54w/dz/PH79LnHvsjFjfOE4AneEVxHZuPs12K6n/WjN0kIMcYd6ZLDTJjCUieAfTqZ0rQN3rk4Zi/EGa55no9zYAehykv9QZrql36J8sd+TAYu/zBMMUKEIZU8DMH7MWyK4zOGUZwpmE45ete2zLc30dDF4mFesrJ+CJNlkQ2TjhvDlyxDNRBCPJ8QhzPGSUWCT6M9x8/SMI4cca5nvr2B+h7vHUU5ZWX9YBz+PZxPJE4QRmIvf0LlLp36JolJLbL0WWHoRBfe85236EvSVIDu13+d4p3vRPbv3wNADAMIRGI859OF5dbiNE5zjFMe4+sMkCzDZBmz9QOx9JtS1mwY/ZaOM2kbFjUxP4yPE4m0+CzDSobqYvSdaojnEEk3QJitHRyZJN57TLbI15ezppj9pQqr92Q2i+Vu7zE2H48zIkhmCEO11USf4MI3yaj+okd44QW6979f21Q5LZbKvMvDFYu8oEjVARIlKPKS/Ajf9YmNnaecWlXH9u6ROiSxhGGzNHKp6whBl4YzpPHJKb60Q0NDHzW/LGLYM0wnGib3RPC5oh0na6Yxy32PtfmeIWbR8SY2Y1kSgh8bRIbzwTATxtH3DlV56UIFaH/rt+D0aVVN7OQUX10K4C5PyVnc/UtB5WHaT6IOLYPZSzQi1UgOGyYMq6ZZrYOpSZ8VTcMAPjMu5eF8Q34ehrEco7aHgZWUjltM+9FBI7l0RmwYAfZhatEAnDsF85z8hVHZix5a1+iHPrQ3V07DFCANUFzKMIaCYT6gVE0DEgtrA3WorKpFqFQU43BtkHGSr/N+7NxuE6WzSA0dkTsQe0T7vhsH6AzllzzPaZewh+A9bUpSjIlTjGxm09AxR/BhqSiYyjRFTtMOtf7ICRiAc5E47cdaO7Zays9ce53+k7576eo6mbD22c+KrK4uxrAvOa9hmMBg+8zgbCSSwMKgoXucjcEYwaVmh8EuRgewV+OHERuqccRozJ7COCZfNSDEkZ1hdDYLstxwTW6ghppEAh4cb0Lg9jgloqOMk4U1zbLae02DLL7/jXeq+YTJ+HKWfTPxffNHXdN/8IMaZ4su5vxDBKnj1LQ4FQgSzyj46NhS9WAAIMwAUquOUy188JGPpbG5dpmKownoDhpGQByJ0y+MMWOzhxJiXBl0pOwsU5aGUvp4vhBnpJr0ewCu70Eiou+TY7PD+Kc0539UFhZxqvc+/kLFDdfdpK8Njl91L11bz2fCzx+yCOk3VIbx9QzjQYdhjGBMCnVgZJksl8MZt5MZGoJ0kUX6mzDUMT/9S55VGekssRqcMowQtwdiyR5yDYsf9IrPJhJ6ST9Co6lUQ2R4e2T8mQ8NSlCDD0Pwr/BlY/mU8bwt+Jck1INeectu4M8myaCbdEGiGJX4qz/pB7FCiL8eFCTCwANDe8jdF5JdhoqHmavD46Xb/UsfUYCRw6iwxA9dEioxlx8FmxKuPYLUeMz4Sz8q4/shmCTcKGwb0q36E2NfslAB/trcc3ch9Jlg/NJP0y3/TtVYxh6AGFlsSypfG0GSR1/8HtVCjjJI4P+JqjIMP1oIbzgLyKiZqI4JgC69Hz26LH7lZ1lT0w/VDI0uihC84jGExJuwIX3Ul8VwjzG8Pgy38ls/DgT43kb50GTxfXSo7SfBaFzr0ZjrYvkzaKoAnlh/lOHdZQ1eFuxLf+glG3uEurydioujoJX0M3SR97DQ2CXtTYJbvKeRLqWL9y2kFaHwh1hez0u3rT8299yTW86YONTWKEsaGoWRpXUshvRjWMOkdpCwWPYvEt6Spuql+y45bFmIL8q59cUvddn2stemKqTfRU37lbi0WdLMwdamv/NhaNFKv/N31VU36fibUQL/u2+5XV+atgK8YIR/OzU8mpul36FaOKLRQaELu7ksyD3LfHjzJcERL/2xhH3uESzLjoqkxUsOi4VtXba1cb/s0eKFeYD/G3Za/uujbZ8OAAAAAElFTkSuQmCC';
const LARANJITO_IMG = {
  triste: 'https://moveisdolar.com.br/colaborador/mascote%20triste1.png',
  preocupado: 'https://moveisdolar.com.br/colaborador/mascote%20preocupado1.png',
  feliz: 'https://moveisdolar.com.br/colaborador/mascote%20feliz1.png'
};
function laranjitoSrc(status){
  return LARANJITO_IMG[status] || LARANJITO_IMG.triste;
}



function nextUpdateDateSales(now=new Date()){
  const d=new Date(now.getTime());
  d.setSeconds(0,0);
  const m=d.getMinutes();
  const next=Math.floor(m/20)*20+20;
  if(next>=60){d.setHours(d.getHours()+1); d.setMinutes(0);} else {d.setMinutes(next);}
  return d;
}
function nextUpdateDateCobranca(now=new Date()){
  const horas=[7,9,11,13,15,17,19,21];
  const d=new Date(now.getTime());
  d.setSeconds(0,0);
  for(let add=0; add<3; add++){
    const base=new Date(d.getFullYear(), d.getMonth(), d.getDate()+add, 0,0,0,0);
    for(const h of horas){
      const cand=new Date(base.getFullYear(),base.getMonth(),base.getDate(),h,0,0,0);
      if(cand>now) return cand;
    }
  }
  return new Date(now.getTime()+2*60*60*1000);
}
function _fmtClockDelta(target){
  let sec=Math.max(0, Math.floor((target-new Date())/1000));
  const h=Math.floor(sec/3600); sec%=3600;
  const m=Math.floor(sec/60); const s=sec%60;
  return h>0?`${h}h ${String(m).padStart(2,'0')}m ${String(s).padStart(2,'0')}s`:`${String(m).padStart(2,'0')}m ${String(s).padStart(2,'0')}s`;
}
function nextUpdateClockHtml(){
  return `<div class="next-update-clock" style="font-size:12px;color:#fbbf24;font-weight:900">⏳ Próxima atualização: calculando...</div>`;
}
function updateNextUpdateClocks(){
  try{
    const now=new Date();
    const sv=nextUpdateDateSales(now);
    const cb=nextUpdateDateCobranca(now);
    const target=sv<=cb?sv:cb;
    const label=sv<=cb?'vendas':'cobrança';
    document.querySelectorAll('.next-update-clock').forEach(el=>{el.textContent=`⏳ Próxima atualização: ${label} em ${_fmtClockDelta(target)}`;});
  }catch(e){}
}
setInterval(updateNextUpdateClocks,1000);

function renderKPIs(){
  const grave=flattenFiliais().reduce((a,b)=>a+Number(b.grave_pend||0),0);
  const alerta=flattenFiliais().reduce((a,b)=>a+Number(b.alerta_pend||0),0);
  const rentPct=Number(RENT_EMPRESA?.margem_bruta_pct||0);
  const sales=SALES_EMPRESA||{};
  const salesDates=Object.keys(HIST_DASH?.sales_dates||{}).sort();
  const hojeIso=new Date().toISOString().slice(0,10);
  const prevDate=salesDates.filter(d=>String(d)<String(hojeIso)).slice(-1)[0]||'';
  const prevEmpresa=(prevDate?(HIST_DASH?.sales_dates?.[prevDate]?.empresa||{}):{});
  // ✅ Total de Serviços da tela inicial deve bater com o Controle de Meta do SGI.
  // Fonte oficial do TOTAL: sales.servico_realizado_total, vindo da tela /metas/consulta.
  // O relatório de serviços separado continua sendo usado apenas para detalhar por tipo
  // (Ouro, FOB, Cupom Copa etc.), mas não substitui o total oficial.
  const servicoRelatorioTotal = Number(SERVICOS_RELATORIO?.empresa?.real_total || 0);
  const servicoRealizadoOficial = Number(sales.servico_realizado_total || 0);
  const servicoAtingidoOficial = Number(sales.servico_meta_total||0)>0
    ? (servicoRealizadoOficial / Number(sales.servico_meta_total||0) * 100)
    : Number(sales.servico_atingido_total||0);

  const prevServicoReal = Number(prevEmpresa.servico_realizado_total || 0);

  // V6.5: Venda Diária precisa vir do Controle de Meta do Sólidus filtrado HOJE-HOJE,
  // nunca mais da diferença entre snapshots/histórico. A diferença por snapshot podia pegar
  // vários dias acumulados quando o último snapshot era antigo e inflar o card.
  function _mdlBrNumber(v){
    if(typeof v === 'number' && isFinite(v)) return v;
    let raw=String(v??'').trim();
    let s=raw.replace(/R\$/g,'').replace(/%/g,'').replace(/\s+/g,'');
    if(!s) return 0;
    // pt-BR: 1.234.567,89 -> 1234567.89
    if(s.includes(',') && s.includes('.')) s=s.replace(/\./g,'').replace(',', '.');
    else if(s.includes(',')) s=s.replace(',', '.');
    const n=Number(s);
    return isFinite(n)?n:0;
  }
  function _mdlFindField(row, nomesCampos){
    try{
      for(const campo of nomesCampos){
        if(row[campo]!==undefined && row[campo]!==null && String(row[campo]).trim()!==''){
          return {key:campo, raw:row[campo], value:_mdlBrNumber(row[campo])};
        }
        const semEspaco=campo.replace(/\s+/g,'');
        for(const k of Object.keys(row||{})){
          if(String(k).replace(/\s+/g,'')===semEspaco && row[k]!==undefined && row[k]!==null && String(row[k]).trim()!==''){
            return {key:k, raw:row[k], value:_mdlBrNumber(row[k])};
          }
        }
      }
    }catch(e){}
    return null;
  }
  function _mdlDailyMetaTotalValue(chave, nomesCampos){
    try{
      const diaInfo=(METAS_VENDAS_DIA||{});
      const hojeBr=new Date().toLocaleDateString('pt-BR',{timeZone:'America/Sao_Paulo'});
      if(String(diaInfo.data_consulta||'').trim() && String(diaInfo.data_consulta).trim()!==hojeBr){
        console.warn('[MDL venda diaria V8.4] ignorado JSON stale', diaInfo.data_consulta, hojeBr);
        return {ok:false, value:0};
      }
      const linhas=((((diaInfo||{}).metas||{})[chave]||{}).linhas)||[];
      if(!linhas.length) return {ok:false, value:0};
      const total=linhas.find(x=>x && x._is_total) || {};
      const rowsToTry=[total, ...linhas.filter(x=>x && !x._is_total)];
      for(const row of rowsToTry){
        // V8.4: para valores em R$, prioriza campo visual bruto antes de _float.
        // Alguns retornos do Sólidus/HTML podem vir como 5990 quando o visual é 59,90.
        const rawNames=nomesCampos.filter(c=>!String(c).endsWith('_float'));
        const floatNames=nomesCampos.filter(c=>String(c).endsWith('_float'));
        let found=_mdlFindField(row, rawNames) || _mdlFindField(row, floatNames);
        if(found){
          let val=Number(found.value||0);
          // Sanidade: compara com Meta Período × Atingido Período.
          const metaObj=_mdlFindField(row,['Meta (R$) Período','Meta(R$) Período','Meta (R$) Periodo','Meta(R$) Periodo','Meta(R$) Período_float','Meta (R$) Período_float']);
          const atingObj=_mdlFindField(row,['Atingido Período','Atingido Periodo','Atingido Período_float']);
          const calc=(Number(metaObj?.value||0)*Number(atingObj?.value||0))/100;
          if(calc>0 && val>calc*10){
            console.warn('[MDL venda diaria V8.4] corrigindo valor inflado', {chave, bruto:val, calculado:calc, campo:found.key, raw:found.raw});
            val=calc;
          }
          return {ok:true, value:Math.round(val*100)/100};
        }
      }
    }catch(e){ console.warn('Venda diária SGI total falhou', chave, e); }
    return {ok:false, value:0};
  }
  const _diaVenda=_mdlDailyMetaTotalValue('venda_filial_meta',[
    'Realizado (R$) Período_float','Realizado(R$) Período_float','Realizado (R$) Periodo_float','Realizado(R$) Periodo_float',
    'Realizado (R$) Período','Realizado(R$) Período','Realizado (R$) Periodo','Realizado(R$) Periodo'
  ]);
  const _diaServico=_mdlDailyMetaTotalValue('servico_filial_ouro_fob',[
    'Realizado (R$) Período_float','Realizado(R$) Período_float','Realizado (R$) Periodo_float','Realizado(R$) Periodo_float',
    'Realizado (R$) Período','Realizado(R$) Período','Realizado (R$) Periodo','Realizado(R$) Periodo'
  ]);
  // V9.6: o card Venda diária deve priorizar o relatório oficial do worker
  // (venda_diaria_mes_atual.json anexado em metas_vendas_mes_atual.json).
  // O Controle de Meta diário fica apenas como fallback, porque ele pode representar
  // período/projeção diferente do relatório oficial de venda do dia.
  const _diaOficial = (()=>{
    try{
      const vd=(METAS_VENDAS && METAS_VENDAS.venda_diaria) ? METAS_VENDAS.venda_diaria : {};
      const data=String(vd.data||'').trim();
      const hojeBr=new Date().toLocaleDateString('pt-BR',{timeZone:'America/Sao_Paulo'});
      if(data && data!==hojeBr){
        console.warn('[MDL venda diaria V9.6] relatório oficial stale ignorado', data, hojeBr);
        return {ok:false,value:0,fonte:'oficial_stale'};
      }
      const emp=vd.empresa||{};
      let val=_mdlBrNumber(emp.venda_diaria_total);
      if(!val){
        val=_mdlBrNumber(emp.total_vendido)+_mdlBrNumber(emp.valor_servico)+_mdlBrNumber(emp.valor_acrescimo_servico);
      }
      if(val>0) return {ok:true,value:Math.round(val*100)/100,fonte:'relatorio_venda_diaria_oficial'};
    }catch(e){console.warn('[MDL venda diaria V9.6] falha leitura oficial',e)}
    return {ok:false,value:0,fonte:'fallback_meta_diaria'};
  })();
  const vendaDiaria = _diaOficial.ok
    ? Number(_diaOficial.value||0)
    : ((_diaVenda.ok || _diaServico.ok)
      ? Math.max(0, Number(_diaVenda.value||0)) + Math.max(0, Number(_diaServico.value||0))
      : 0);
  try{
    console.log('[MDL venda diaria V9.6]', {
      fonte: _diaOficial.fonte,
      venda_dia_oficial: _diaOficial.value,
      venda_dia_sgi_fallback: _diaVenda.value,
      servico_dia_sgi_fallback: _diaServico.value,
      total_card: vendaDiaria,
      regra: 'prioriza relatório oficial do worker; meta diária só fallback'
    });
  }catch(e){}
  try{
    console.log('[MDL serviços]', {
      total_controle_meta_sgi: Number(sales.servico_realizado_total||0),
      total_relatorio_servicos_tipos: servicoRelatorioTotal,
      usado_no_card_servicos: servicoRealizadoOficial
    });
  }catch(e){}

  const RENT_OK=getRentEmpresa();
  RENT_EMPRESA=RENT_OK;
  const markupBase=(Number(sales.venda_realizado_total||0)+servicoRealizadoOficial);
  const markupCost=Number(RENT_OK?.custo_total||0);
  const markupTotal=markupCost>0?(markupBase/markupCost):0;
  try{
    console.log('[MDL margem]', {
      rent_empresa_boot: RENT_EMPRESA,
      margens_empresa: MARGENS_BRUTAS?.empresa,
      markupBase,
      markupCost,
      markupTotal
    });
  }catch(e){}
  function statusLaranjitoVendaDiaria(v){
    v=Number(v||0);
    if(v < 20000) return 'triste';
    if(v <= 25000) return 'preocupado';
    return 'feliz';
  }
  function statusLaranjitoMarkup(v){
    v=Number(v||0);
    if(!v || v < 2.00) return 'triste';
    if(v <= 2.14) return 'preocupado';
    return 'feliz';
  }
  const isViewer=!!usuarioAtual?.is_viewer;
  const isPrivileged=(usuarioAtual?.tipo==='master' || isViewer);

  // Cards extras por tipo de serviço vindos do relatorio_servicos_mes_atual.json
  function servicoIcone(nome){
    const n=normSalesText(nome||'');
    if(n.includes('OURO')) return '🪙';
    if(n.includes('FOB')) return '🚚';
    if(n.includes('COPA') || n.includes('CUPOM')) return '⚽';
    return '🛠️';
  }
  const topServiceCards = Object.values((SERVICOS_RELATORIO && SERVICOS_RELATORIO.servicos) || {})
    .slice()
    .sort((a,b)=>Number(_srvTotal(b)||0)-Number(_srvTotal(a)||0))
    .slice(0,4)
    .map(s=>makeKpi(`${servicoIcone(s.servico)} ${String(s.servico||'Serviço').slice(0,30)}`, R(_srvTotal(s)||0), 'var(--blue-400)', `${Number(s.quantidade||0).toLocaleString('pt-BR')} item(ns)`));

  const cards=[
    makeKpi('💰 Total pendente',R(TOTAL_P),'var(--red)','', 'card-cobranca'),
    makeKpi('🏦 Total recebido',R(TOTAL_PG),'var(--green)','', 'card-cobranca'),
    makeKpi('🚨 Grave',R(grave),'var(--red)','', 'card-cobranca'),
    makeKpi('🟠 Alerta',R(alerta),'var(--orange)','', 'card-cobranca'),
    makeKpi('📦 Mercantil realizado',R(sales.venda_realizado_total||0),'var(--amber-400)',`Meta ${R(sales.venda_meta_total||0)} · Atingido ${pct(sales.venda_atingido_total||0)}`),
    makeKpi('📈 Mercantil projetado',R(sales.venda_projetado||0),'var(--amber-500)',`Meta período ${R(sales.venda_meta_periodo||0)} · Projetado ${pct((Number(sales.venda_meta_total||0)>0?Number(sales.venda_projetado||0)/Number(sales.venda_meta_total||0)*100:0))}`),
    makeKpi('🛠️ Serviços realizado',R(servicoRealizadoOficial),'var(--blue)',`Meta ${R(sales.servico_meta_total||0)} · Atingido ${pct(servicoAtingidoOficial)} · controle de meta SGI`),
    makeKpi('🧰 Serviços projetado',R(sales.servico_projetado||0),'var(--blue-400)',`Meta período ${R(sales.servico_meta_periodo||0)} · Projetado ${pct((Number(sales.servico_meta_total||0)>0?Number(sales.servico_projetado||0)/Number(sales.servico_meta_total||0)*100:0))}`),
    ...topServiceCards,
    makeKpi('🚚 Caminhão realizado',R(sales.caminhao_realizado_total||0),'var(--yellow)',`Meta ${R(sales.caminhao_meta_total||0)} · Atingido ${pct(sales.caminhao_atingido_total||0)}`),
    makeKpi('🛣️ Caminhão projetado',R(sales.caminhao_projetado||0),'var(--yellow-400)',`Meta período ${R(sales.caminhao_meta_periodo||0)}`),
    (isPrivileged ? makeKpi('💵 Faturamento total',R((Number(sales.venda_realizado_total||0)+servicoRealizadoOficial)),'var(--green-400)','Mercantil + serviços realizado', 'card-financeiro') : ''),
    (isPrivileged ? makeKpi('🕒 Venda diária',R(vendaDiaria),'var(--cyan-400)','Mercantil + serviços do dia', 'card-venda-dia', statusLaranjitoVendaDiaria(vendaDiaria)) : ''),
    makeKpi('📊 Rentabilidade total', rentPct?`${rentPct.toFixed(2).replace('.',',')}%`:'Sem dado','var(--green-400)','Última linha do relatório de margem bruta por filial', 'card-financeiro'),
    makeKpi('🧮 Markup total', markupTotal?String(markupTotal.toFixed(2)).replace('.',','):'0,00','var(--amber-400)', isViewer ? 'Índice mercantil + serviços / custo oculto' : `(Mercantil + serviços) / custo total ${R(markupCost||0)}`, 'card-financeiro', statusLaranjitoMarkup(markupTotal))
  ];
  document.getElementById('kpis').innerHTML=cards.join('') + `<div class="glass" style="grid-column:1/-1;padding:10px 14px;display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;min-height:46px"><div style="font-size:12px;color:#a9b2c7">🕒 Última atualização do dashboard: <strong style="color:#e5e7eb">${esc(latestUpdatedLabel()||'--')}</strong></div>${nextUpdateClockHtml()}</div>`; updateNextUpdateClocks();
}

async function fetchJsonNoCache(url){
  const r=await fetch(url+(url.includes('?')?'&':'?')+'_='+Date.now(),{cache:'no-store'});
  if(!r.ok) throw new Error('HTTP '+r.status);
  return await r.json();
}
const DASHBOARD_UPDATED_AT_LABEL=__DASHBOARD_UPDATED_AT_LABEL__;  // V7.8 gerado no momento da execução

function formatUpdatedLabel(v){
  try{
    const s=String(v||'').trim();
    if(!s) return '';
    if(/^\d{2}\/\d{2}\/\d{4}[, ]+\d{2}:\d{2}:\d{2}$/.test(s)) return s.replace(',', '').replace(/\s+/g,' ').trim();
    const d=new Date(s);
    if(!isNaN(d.getTime())){
      return d.toLocaleString('pt-BR',{timeZone:'America/Sao_Paulo', hour12:false}).replace(',', '');
    }
    return s.replace(',', '').replace(/\s+/g,' ').trim();
  }catch(_e){
    return String(v||'').replace(',', '').trim();
  }
}
function dashboardUpdatedLabel(){
  return formatUpdatedLabel(window.__dashboardUpdatedAtLabel || DASHBOARD_UPDATED_AT_LABEL);
}
function latestUpdatedLabel(){
  const aRaw = window.__dashboardUpdatedAtLabel || DASHBOARD_UPDATED_AT_LABEL || '';
  const bRaw = window.__salesUpdatedAtLabel || '';
  const a = formatUpdatedLabel(aRaw);
  const b = formatUpdatedLabel(bRaw);
  if(!a && !b) return '';
  if(a && !b) return a;
  if(b && !a) return b;
  const da = new Date(aRaw);
  const db = new Date(bRaw);
  if(!isNaN(da.getTime()) && !isNaN(db.getTime())){
    return db > da ? b : a;
  }
  return b || a;
}

function renderUpdateStrip(){
  return `<div class="glass" style="margin:10px 0 14px;padding:10px 14px;display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;min-height:42px;border-color:rgba(148,163,184,.20)"><div style="font-size:12px;color:#a9b2c7">🕒 Última atualização do dashboard: <strong style="color:#e5e7eb">${esc(latestUpdatedLabel()||'--')}</strong></div>${nextUpdateClockHtml()}</div>`;
}


function calcSalesEmpresaFromMetas(payload){
  const metas=((payload||{}).metas)||{};
  function totalOf(key){
    const linhas=((metas[key]||{}).linhas)||[];
    const total=linhas.find(x=>x&&x._is_total) || {};
    return {
      real_total:Number(total['Realizado(R$) Total_float']||0),
      ating_total:Number(total['Atingido Total_float']||0),
      proj:Number(total['Projetado(R$)_float']||0),
      meta_total:Number(total['Meta(R$) Total_float']||0),
      meta_per:Number(total['Meta(R$) Período_float']||0),
    };
  }
  const venda=totalOf('venda_filial_meta');
  const serv=totalOf('servico_filial_ouro_fob');
  const cam=totalOf('venda_filial_subgrupo_20k');
  return {
    venda_realizado_total:venda.real_total,
    venda_atingido_total:venda.ating_total,
    venda_projetado:venda.proj,
    venda_meta_total:venda.meta_total,
    venda_meta_periodo:venda.meta_per,
    servico_realizado_total:serv.real_total,
    servico_atingido_total:serv.ating_total,
    servico_projetado:serv.proj,
    servico_meta_total:serv.meta_total,
    servico_meta_periodo:serv.meta_per,
    caminhao_realizado_total:cam.real_total,
    caminhao_atingido_total:cam.ating_total,
    caminhao_projetado:cam.proj,
    caminhao_meta_total:cam.meta_total,
    caminhao_meta_periodo:cam.meta_per,
  };
}
async function pollSalesLive(){
  try{
    const ver=await fetchJsonNoCache('sales_version.json');
    const stamp=String(ver?.updated_at_label||ver?.updated_at||'');
    if(stamp){ window.__salesUpdatedAtLabel = stamp; }

    // Atualização atômica do pacote de vendas:
    // só busca metas/margens/serviços quando sales_version mudar.
    // Assim o dashboard não mistura parte nova de vendas com serviços/margem antigos.
    if(stamp && window.__lastSalesVersion===stamp && window.__salesBundleLoaded){
      return;
    }

    let metasWrap=null, metasDiaWrap=null, margensWrap=null, servWrap=null;
    try{ metasWrap=await fetchJsonNoCache('metas_vendas_mes_atual.json'); }catch(_e){}
    try{ metasDiaWrap=await fetchJsonNoCache('metas_vendas_dia_atual.json'); }catch(_e){}
    try{ margensWrap=await fetchJsonNoCache('margens_brutas_mes_atual.json'); }catch(_e){}
    try{ servWrap=await fetchJsonNoCache('relatorio_servicos_mes_atual.json'); }catch(_e){}

    if(metasWrap){ METAS_VENDAS=metasWrap||{}; SALES_EMPRESA=calcSalesEmpresaFromMetas(metasWrap||{}); }
    if(metasDiaWrap) METAS_VENDAS_DIA=metasDiaWrap||{};
    if(margensWrap) RENT_EMPRESA=((margensWrap||{}).empresa)||{};
    if(servWrap) SERVICOS_RELATORIO=(servWrap||{});

    if(stamp) window.__lastSalesVersion=stamp;
    window.__salesBundleLoaded=true;

    if(typeof renderKPIs==='function' && document.getElementById('kpis')) renderKPIs();
    if(typeof renderServicosTab==='function' && (!servicesSection.classList.contains('hidden') || usuarioAtual?.is_viewer)) renderServicosTab(!!usuarioAtual?.is_viewer);
    if(!detailScreen.classList.contains('hidden') && currentDetailRef){
      try{ openEntity(currentDetailRef); }catch(_e){}
    }
  }catch(e){console.log('pollSalesLive',e)}
}
async function pollDashboardLiveReload(){
  try{
    const ver=await fetchJsonNoCache('dashboard_version.json');
    const stamp=String(ver?.updated_at||ver?.updated_at_label||'');
    if(stamp){
      window.__dashboardUpdatedAtLabel = stamp;
      if(typeof renderKPIs==='function' && document.getElementById('kpis')) renderKPIs();
    }
    if(!stamp) return;
    if(window.__lastDashboardVersion===undefined){window.__lastDashboardVersion=stamp; return;}
    if(window.__lastDashboardVersion!==stamp){ location.reload(); return; }
  }catch(e){console.log('pollDashboardLiveReload',e)}
}


function _srvTotal(s){return Number((s&&((s.real_total_oficial!=null?s.real_total_oficial:null) ?? (s.total_oficial!=null?s.total_oficial:null) ?? s.real_total))||0);}
function _srvSortedEntries(obj, key='real_total'){
  return Object.values(obj||{}).slice().sort((a,b)=>Number(b?.[key]||0)-Number(a?.[key]||0));
}
function _srvMiniCard(title, value, subtitle=''){
  return `<div class="glass panel" style="padding:14px 16px"><div class="mini" style="margin-bottom:8px">${esc(title)}</div><div class="big" style="font-size:28px;color:#f8fafc">${value}</div>${subtitle?`<div class="hint" style="margin-top:6px">${subtitle}</div>`:''}</div>`;
}
function _srvRankRows(rows, kind){
  if(!rows.length) return `<div class="empty">Nenhum dado encontrado no relatório de serviços.</div>`;
  return rows.map((r,idx)=>{
    const label = kind==='servico' ? (r.servico||'-') : (kind==='filial' ? (r.label||r.filial||'-') : (r.label||r.nome||'-'));
    const qtd = Number(r.quantidade||0);
    const total = Number(_srvTotal(r)||0);
    const details = kind==='vendedor' ? `${esc(r.filial||'')}` : `${qtd.toLocaleString('pt-BR')} item(ns)`;
    return `<div class="log-row"><div><strong>${idx+1}. ${esc(label)}</strong><div class="small muted">${details}</div></div><div><strong>${R(total)}</strong><div class="small muted">Total</div></div></div>`;
  }).join('');
}
function renderServicosTab(isViewer=false){
  const data = SERVICOS_RELATORIO||{};
  const empresa = data.empresa||{};
  const tipos = _srvSortedEntries(data.servicos||{});
  const filiais = _srvSortedEntries(data.filiais||{});
  const vendedores = _srvSortedEntries(data.vendedores||{});

  const totalServ = Number(empresa.real_total||0);
  const totalQtd = Number(empresa.quantidade||0);
  const totalLinhas = Number(empresa.linhas||0);
  const tiposAtivos = tipos.length;

  const topTipo = tipos[0]||{};
  const topFil = filiais[0]||{};
  const topVend = vendedores[0]||{};

  const summary = `
    <div class="kpis" style="margin-bottom:14px">
      ${_srvMiniCard('🧰 Serviços no mês', R(totalServ), `${totalQtd.toLocaleString('pt-BR')} item(ns) · ${totalLinhas.toLocaleString('pt-BR')} linha(s)`)}
      ${_srvMiniCard('🏷️ Tipos de serviço', String(tiposAtivos), topTipo.servico?`Maior: ${esc(topTipo.servico)} · ${R(topTipo.real_total||0)}`:'')}
      ${_srvMiniCard('🏬 Melhor filial', topFil.label?esc(topFil.label):'—', topFil.real_total?`${R(topFil.real_total)} · ${Number(topFil.quantidade||0).toLocaleString('pt-BR')} item(ns)`:'')}
      ${_srvMiniCard('👤 Melhor vendedor', topVend.label?esc(topVend.label):'—', topVend.real_total?`${R(topVend.real_total)} · ${Number(topVend.quantidade||0).toLocaleString('pt-BR')} item(ns)`:'')}
    </div>`;

  const tipoCards = tipos.length ? `
    <div class="glass panel" style="margin-bottom:14px">
      <div class="section-head"><div><h2>🛠️ Serviços por tipo</h2><div class="hint">Totais consolidados por serviço do relatório mensal. Este é o valor oficial usado no card Serviços realizado.</div></div></div>
      <div class="grid-cards">${tipos.map(t=>`
        <div class="glass card" style="padding:14px 16px">
          <div class="title">${esc(t.servico||'-')}</div>
          <div class="numbers" style="grid-template-columns:minmax(0,1fr) minmax(0,1fr)">
            <div class="stat-box"><div class="mini">Quantidade</div><div class="big" style="font-size:18px">${Number(t.quantidade||0).toLocaleString('pt-BR')}</div></div>
            <div class="stat-box"><div class="mini">Total</div><div class="big" style="font-size:18px;color:var(--green)">${R(t.real_total||0)}</div></div>
          </div>
          <div class="legend-inline"><span><i class="dot" style="background:#f59e0b"></i>${Number(t.linhas||0).toLocaleString('pt-BR')} lançamento(s)</span></div>
        </div>`).join('')}</div>
    </div>` : `<div class="empty">Nenhum serviço encontrado no relatório.</div>`;

  const rankings = `
    <div class="grid-2">
      <div class="glass panel">
        <div class="section-head"><div><h2>🏬 Ranking por filial</h2><div class="hint">Top filiais em serviços do mês.</div></div></div>
        <div class="logs-list">${_srvRankRows(filiais.slice(0,20), 'filial')}</div>
      </div>
      <div class="glass panel">
        <div class="section-head"><div><h2>👤 Ranking por vendedor</h2><div class="hint">Top vendedores em serviços do mês.</div></div></div>
        <div class="logs-list">${_srvRankRows(vendedores.slice(0,20), 'vendedor')}</div>
      </div>
    </div>`;

  const host = servicesSection;
  if(!host) return;
  host.innerHTML = summary + tipoCards + rankings;

  if(isViewer){
    host.classList.remove('hidden');
  }
}

function setMainTab(tab){
  const isDiretor=usuarioAtual?.tipo==='master' && usuarioAtual?.roleLabel==='Diretor Comercial';
  if(isDiretor && ['cobrancas','senhas'].includes(tab)){tab='vendedores';}
  mainTab=tab;
  try{document.body.classList.toggle('inicio-view', tab==='inicio')}catch(e){}
  document.querySelectorAll('.tab').forEach(t=>t.classList.toggle('active',t.dataset.tab===tab));
  detailScreen.classList.add('hidden');
  document.getElementById('mainScreen').classList.remove('hidden');
  const hiddenMain=['inicio','metas','servicos','cobrancas','reativacao','aniversariantes','avisos','telegram','senhas','historico'].includes(tab);
  if(inicioSection) inicioSection.classList.toggle('hidden',tab!=='inicio');
  listSection.classList.toggle('hidden',hiddenMain);
  metaSection.classList.toggle('hidden',tab!=='metas');
  servicesSection.classList.toggle('hidden',tab!=='servicos');
  logSection.classList.toggle('hidden',tab!=='cobrancas');
  if(reativacaoSection) reativacaoSection.classList.toggle('hidden',tab!=='reativacao');
  if(aniversariantesSection) aniversariantesSection.classList.toggle('hidden',tab!=='aniversariantes');
  avisosSection.classList.toggle('hidden',tab!=='avisos');
  if(telegramSection) telegramSection.classList.toggle('hidden',tab!=='telegram');
  senhasSection.classList.toggle('hidden',tab!=='senhas');
  histSection.classList.toggle('hidden',tab!=='historico');
  mainFilters.classList.toggle('hidden',!(tab==='vendedores'||tab==='filiais'));
  renderTopMural();
  if(tab==='inicio') renderInicioTab();
  if(tab==='vendedores'||tab==='filiais'){renderFilters();renderList()}
  if(tab==='metas'){
    renderMetasTab();
    if(!window._configMetaOnlineLoaded){
      carregarConfigOnline().then(()=>{ if(mainTab==='metas') renderMetasTab(); }).catch(()=>{});
    }
  }
  if(tab==='servicos') renderServicosTab(false);
  if(tab==='cobrancas') renderLogsTab();
  if(tab==='reativacao') renderReativacaoTab();
  if(tab==='aniversariantes') renderAniversariantesTab();
  if(tab==='avisos') renderAvisosTab();
  if(tab==='telegram') renderTelegramTab();
  if(tab==='senhas') renderSenhasTab();
  if(tab==='historico') renderHistoricoTab();
}
function renderFilters(){if(mainTab!=='vendedores'&&mainTab!=='filiais'){mainFilters.innerHTML='';return} let html=`<button class="pill ${filtroFilial==='TODAS'?'active':''}" onclick="setFiltroFilial('TODAS')">Todas</button>`; ORDEM.forEach(f=>{html+=`<button class="pill ${filtroFilial===f?'active':''}" onclick="setFiltroFilial('${f}')">${f}</button>`}); mainFilters.innerHTML=html;}
function setFiltroFilial(f){
  filtroFilial=f;
  renderFilters();

  const detalheAberto = !detailScreen.classList.contains('hidden');
  if(detalheAberto){
    if(mainTab==='filiais'){
      if(f && f!=='TODAS'){
        openEntity({type:'filial',filial:f,nome:filialLabel(f)});
        renderFilters();
        return;
      }
      detailScreen.classList.add('hidden');
      document.getElementById('mainScreen').classList.remove('hidden');
      renderList();
      return;
    }
    detailScreen.classList.add('hidden');
    document.getElementById('mainScreen').classList.remove('hidden');
  }
  renderList();
}
function currentEntities(){let arr=mainTab==='filiais'?flattenFiliais():flattenVendedores(); if(mainTab==='vendedores' && usuarioAtual?.tipo==='master'){const t=thirdChargeEntity(); const hasThird=Number(t.pendente||0)>0 || Number(t.pago||0)>0 || Number(t.grave_pend||0)>0 || Number(t.alerta_pend||0)>0 || Number(t.atencao_pend||0)>0; if(hasThird) arr=[t,...arr]; const creds=crediaristaEntities().filter(x=>Number(x.pendente||0)>0||Number(x.pago||0)>0); if(creds.length) arr=[...creds,...arr]} return arr.filter(x=>filtroFilial==='TODAS'||x.filial===filtroFilial || x.is_terceiro || x.is_crediarista)}
function renderEntityCard(ent){const m=calcMeta(ent); const isThird=!!(ent?.is_terceiro || ent?.type==='terceiro'); const isCred=!!(ent?.is_crediarista || ent?.type==='crediarista'); if(isThird || isCred){const credLogin=String(ent.login||crediaristaLoginByFilial(ent.filial)||'').toLowerCase(); const credFilial=String(ent.filial||'').toUpperCase(); const credNome=String(ent.nome||`CREDIARISTA${credFilial}`); const label=isThird?'Cobrança terceiro':'Crediarista'; const sub=isThird?'Clique para abrir a carteira terceirizada':'Clique para abrir a carteira do crediarista'; const actionAttr=isThird?`data-action="third-card" role="button" tabindex="0"`:`data-action="cred-card" data-login="${esc(credLogin)}" data-filial="${esc(credFilial)}" data-nome="${esc(credNome)}" role="button" tabindex="0"`; return `<div class="glass card ${m.geral>=50?'card-hit':(m.geral<30?'card-low-red':(m.geral<40?'card-low-orange':''))}" style="box-shadow:0 0 0 2px rgba(239,68,68,.12) inset" ${actionAttr}><div class="title">${esc(credNome)}</div><div class="numbers" style="grid-template-columns:minmax(0,1fr) minmax(0,1fr)"><div class="stat-box" style="min-width:0"><div class="mini">Pendente</div><div class="big" style="color:var(--red);font-size:15px;word-break:break-word">${R(ent.pendente||0)}</div></div><div class="stat-box" style="min-width:0"><div class="mini">Recebido</div><div class="big" style="color:var(--green);font-size:15px;word-break:break-word">${R(ent.pago||0)}</div></div></div><div class="meta-row"><div class="mini-chip">🔴 Grave ${pct(m.grave.perc)}</div><div class="mini-chip">🟠 Alerta ${pct(m.alerta.perc)}</div><div class="mini-chip">🟡 Atenção ${pct(m.atencao.perc)}</div><div class="mini-chip" style="font-size:12px">🔵 Meta geral ${pct(m.geral)}</div></div>${renderMascotStatus(m.geral,label)}<div class="legend-inline"><span><i class="dot" style="background:#2f67f6"></i>${sub}</span></div></div>`} const bonus=getBonus(m.cfg,m.geral);const sales=summarizeSalesCard(ent);const salesPct=sales?.n||0;const salesBorder=salesPct>=100?'box-shadow:0 0 0 2px rgba(242,201,76,.35) inset':salesPct>=80?'box-shadow:0 0 0 2px rgba(34,197,94,.18) inset':salesPct>=50?'box-shadow:0 0 0 2px rgba(249,115,22,.18) inset':'box-shadow:0 0 0 2px rgba(239,68,68,.12) inset';const cls=m.geral>=50?'card-hit':(m.geral<30?'card-low-red':(m.geral<40?'card-low-orange':''));const pulseNote=m.geral>=50?'<div class="legend-inline"><span><i class="dot" style="background:#2f67f6"></i>Meta atingida no mês</span></div>':'';return `<div class="glass card ${cls}" style="${salesBorder}" onclick='openEntity(${JSON.stringify({type:ent.type,filial:ent.filial,nome:ent.nome})})'><div class="title">${esc(ent.nome)} ${ent.type==='vendedor'?`(${ent.filial})`:''}</div><div class="numbers" style="grid-template-columns:minmax(0,1fr) minmax(0,1fr)"><div class="stat-box" style="min-width:0"><div class="mini">Pendente</div><div class="big" style="color:var(--red);font-size:15px;word-break:break-word">${R(ent.pendente||0)}</div></div><div class="stat-box" style="min-width:0"><div class="mini">Recebido</div><div class="big" style="color:var(--green);font-size:15px;word-break:break-word">${R(ent.pago||0)}</div></div></div><div class="meta-row"><div class="mini-chip">🔴 Grave ${pct(m.grave.perc)}</div><div class="mini-chip">🟠 Alerta ${pct(m.alerta.perc)}</div><div class="mini-chip">🟡 Atenção ${pct(m.atencao.perc)}</div><div class="mini-chip" style="font-size:12px">🔵 Meta geral ${pct(m.geral)}</div></div>${renderSalesCardSummary(ent)}${renderDualMascotStatus(ent)}${bonus?`<div class="legend-inline"><span><i class="dot" style="background:#2f67f6"></i>${esc(bonus.text)}</span></div>`:''}${pulseNote}</div>`}
function renderGroupBars(entities){if(!entities.length) return `<div class="empty">Nenhum dado para exibir.</div>`; const max=Math.max(1,...entities.map(e=>Math.max(Number(e.grave_pend||0),Number(e.alerta_pend||0),Number(e.atencao_pend||0),Number(e.pago||0)))); return `<div class="glass big-chart-card"><div class="section-head"><div><h2>📊 Panorama por ${mainTab==='filiais'?'filial':'vendedor'}</h2><div class="hint">Barras por faixa: Grave, Alerta, Atenção e Recebido</div></div><div class="legend-inline"><span><i class="dot" style="background:var(--red)"></i>Grave</span><span><i class="dot" style="background:var(--orange)"></i>Alerta</span><span><i class="dot" style="background:var(--yellow)"></i>Atenção</span><span><i class="dot" style="background:var(--green)"></i>Recebido</span></div></div><div class="groupbars">${entities.map(e=>{const vals=[{c:'var(--red)',v:Number(e.grave_pend||0),t:'Grave'},{c:'var(--orange)',v:Number(e.alerta_pend||0),t:'Alerta'},{c:'var(--yellow)',v:Number(e.atencao_pend||0),t:'Atenção'},{c:'var(--green)',v:Number(e.pago||0),t:'Recebido'}]; return `<div class="group"><div class="bars">${vals.map(v=>`<div title="${v.t}: ${R(v.v)}" class="bar" style="height:${Math.max(12,(v.v/max)*240)}px;background:linear-gradient(180deg,${v.c},${v.c})"></div>`).join('')}<span class="wave one"></span><span class="wave two"></span><span class="bubble b1"></span><span class="bubble b2"></span><span class="bubble b3"></span></div><div class="glabel">${esc(trunc(e.nome,16))}</div></div>`}).join('')}</div><div class="axis"><span>Escala relativa automática</span><span>${entities.length} ${mainTab==='filiais'?'filiais':'colaboradores'}</span></div></div>`}


function toggleTickerSpeed(btn){
  const box=btn?.closest('.glass.panel')?.querySelector('.aviso-ticker');
  if(!box) return;
  const track=box.querySelector('.aviso-ticker-track');
  const fast=!box.classList.contains('fast');
  box.classList.toggle('fast',fast);
  if(track) track.style.animationDuration = fast ? '45s' : '1800s';
  if(btn){btn.textContent=fast?'🐢 Normal':'⚡ Acelerar'; btn.classList.toggle('primary',fast);}
}
function renderAvisoTicker(title,hint,entries,opts={}){
  const arr=(entries||[]).filter(Boolean);
  if(!arr.length) return '';
  const icon=opts.icon||'🔔';
  const color=opts.color||'rgba(245,158,11,.35)';
  const doubled=[...arr,...arr];
  const accel=(usuarioAtual?.tipo==='master' || usuarioAtual?.is_viewer)?`<button class="btn soft btn-xs ticker-speed-btn" onclick="toggleTickerSpeed(this)">⚡ Acelerar</button>`:'';
  return `<div class="glass panel" style="margin-bottom:16px;padding:14px 18px;border-color:${color}">
    <div class="section-head" style="margin:0"><div><h2 style="margin:0;font-size:18px">${icon} ${esc(title)}</h2><div class="hint">${esc(hint||'')}</div></div><div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap"><div class="badge">${arr.length}</div>${accel}</div></div>
    <div class="aviso-ticker"><div class="aviso-ticker-track" style="animation-duration:1800s">${doubled.map(e=>`<span class="aviso-pill"><i class="red-dot"></i>${esc(e.nome||e.label||'')}<small>${esc(e.info||'')}</small></span>`).join('')}</div></div>
  </div>`;
}


function renderMetaDiariaBatidaAlerts(){
  try{
    const hoje=dateOnlyISO(new Date());
    const entry=HIST_DASH?.daily_meta?.dates?.[hoje] || {};
    const arr=[];
    Object.values(entry.vendedores||{}).forEach(v=>{if(Number(v.ponto||0)>0) arr.push({nome:v.nome||v.key, info:`${v.filial||''} · vendeu ${R(v.realizado_dia||0)} / meta ${R(v.meta_diaria||0)}`})});
    Object.values(entry.filiais||{}).forEach(f=>{if(Number(f.ponto||0)>0) arr.push({nome:filialLabel(f.filial||''), info:`${f.filial||''} · vendeu ${R(f.realizado_dia||0)} / meta ${R(f.meta_diaria||0)}`})});
    if(!arr.length){
      return `<div class="glass panel" style="margin-bottom:16px;padding:14px 18px;border-color:rgba(34,197,94,.22)"><div class="section-head" style="margin:0 0 8px"><div><h2 style="margin:0;font-size:18px">🎯 Meta diária BATIDA</h2><div class="hint">Mural ativo. Quando vendedor ou filial bater a meta diária mercantil, aparecerá aqui.</div></div><div class="badge">0</div></div><div class="meta-diaria-empty">Nenhuma meta diária batida até o momento.</div></div>`;
    }
    arr.sort((a,b)=>String(a.nome).localeCompare(String(b.nome),'pt-BR'));
    return renderAvisoTicker('Meta diária BATIDA','Controle de Meta do Sólidus filtrado na data atual: Atingido Período acima de 100%.', arr, {icon:'🎯',color:'rgba(34,197,94,.30)'});
  }catch(e){console.warn('renderMetaDiariaBatidaAlerts',e); return '';}
}

function normCobUserText(v){
  try{
    return String(v||'')
      .normalize('NFD').replace(/[\u0300-\u036f]/g,'')
      .toLowerCase()
      .replace(/^\s*f\d+\s*[-–—:]\s*/,'')
      .replace(/\s+/g,' ')
      .trim();
  }catch(e){return String(v||'').toLowerCase().trim()}
}
function logIsHojeCobrancaReal(x,hoje){
  if(!isLogCobrancaReal(x)) return false;
  return dateOnlyISO(x.server_time||x.criado_em||x.data||x.created_at||'')===hoje;
}
function logMatchesAnyOwner(x,keys,filial=''){
  const filialNorm=String(filial||'').toUpperCase();
  const rawFields=[x.usuario,x.destino_nome,x.login,x.nome,x.responsavel,x.vendedor,x.user_name].filter(Boolean).map(v=>String(v));
  const rawJoined=rawFields.join(' | ');
  const normFields=rawFields.map(normCobUserText);
  const normJoined=normCobUserText(rawJoined);
  const logFilial=String(x.filial||'').toUpperCase();
  const destinoTipo=String(x.destino_tipo||x.tipo||'').toLowerCase();
  const keyset=(keys||[]).map(k=>String(k||'')).filter(Boolean);

  for(const k0 of keyset){
    const kRaw=String(k0||'').trim();
    const kLower=kRaw.toLowerCase();
    const kNorm=normCobUserText(kRaw);
    if(!kNorm) continue;

    // Filial: considera qualquer cobrança registrada naquela filial como ação do painel/gerente.
    if(/^f\d+$/i.test(kRaw) || kLower.startsWith('filial f')){
      const fKey=(kRaw.match(/f\d+/i)||[''])[0].toUpperCase();
      if(fKey && (logFilial===fKey || rawJoined.toUpperCase().includes(fKey+' -') || rawJoined.toUpperCase().includes('FILIAL '+fKey.replace('F','')))) return true;
      continue;
    }

    // Login/nome exato ou nome com prefixo F3 - NOME, etc.
    if(normFields.includes(kNorm)) return true;
    if(kNorm.length>=4 && (normJoined.includes(kNorm) || kNorm.includes(normJoined))) return true;
  }

  // Fallback por filial quando a linha veio marcada como filial/gerente/crediarista.
  if(filialNorm && logFilial===filialNorm && /filial|gerente|crediarista|cobran/i.test(destinoTipo)) return true;
  return false;
}
function countCobrancasHojePorOwner(keys,filial=''){
  const hoje=dateOnlyISO(new Date());
  return (COB_LOGS||[]).filter(x=>logIsHojeCobrancaReal(x,hoje) && logMatchesAnyOwner(x,keys,filial)).length;
}
function renderNoChargeAlerts(){
  const totalPend=(obj)=>((obj?.grave||[]).length+(obj?.alerta||[]).length+(obj?.atencao||[]).length);
  const entries=[];

  // Vendedores / colaboradores: usa nome, login e filial do próprio usuário.
  flattenVendedores().forEach(v=>{
    const pending=totalPend(CLIENTES_VEND[v.nome]||{});
    const done=countCobrancasHojePorOwner([v.nome, v.login], v.filial);
    if(pending>0 && done===0) entries.push({tipo:'Colaborador',nome:v.nome,filial:v.filial,pending,done});
  });

  // Filiais / gerentes: se qualquer cobrança daquela filial foi registrada hoje, não aparece no mural.
  flattenFiliais().forEach(f=>{
    const pending=totalPend(CLIENTES_FIL[f.filial]||{});
    const done=countCobrancasHojePorOwner([f.nome, f.filial, filialLabel(f.filial)], f.filial);
    if(pending>0 && done===0) entries.push({tipo:'Filial',nome:filialLabel(f.filial),filial:f.filial,pending,done});
  });

  // Crediaristas: reconhece login, nome e logs com destino_tipo=crediarista da mesma filial.
  crediaristaEntities().forEach(c=>{
    const key=String(c.login||'').toLowerCase();
    const pending=totalPend(CLIENTES_CREDIARISTA[key]||{});
    const done=countCobrancasHojePorOwner([c.nome, c.login], c.filial);
    if(pending>0 && done===0) entries.push({tipo:'Crediarista',nome:c.nome,filial:c.filial,pending,done});
  });

  // Cobrança terceiro / Cobrança10
  const pendingTer=totalPend(CLIENTES_TERCEIRO||{});
  const doneTer=countCobrancasHojePorOwner([COBRANCA10_NOME, COBRANCA10_LOGIN, 'cobranca10', 'cobranca 10'], 'FTER');
  if(pendingTer>0 && doneTer===0) entries.push({tipo:'Cobrança',nome:COBRANCA10_NOME,filial:'FTER',pending:pendingTer,done:doneTer});

  const uniq=[];
  const seen=new Set();
  entries.forEach(e=>{
    const k=`${e.tipo}|${e.nome}|${e.filial}`;
    if(!seen.has(k)){seen.add(k); uniq.push(e);}
  });
  uniq.sort((a,b)=>String(a.tipo).localeCompare(String(b.tipo),'pt-BR') || Number(b.pending||0)-Number(a.pending||0));
  if(!uniq.length) return '';
  return renderAvisoTicker('Sem cobranças hoje','Lista giratória de usuários/carteiras com cobrança pendente e sem registro de WhatsApp hoje.', uniq.map(e=>({nome:e.nome, info:`${e.tipo}${e.filial?` · ${e.filial}`:''} · ${e.pending} clientes`})), {icon:'⏰',color:'rgba(239,68,68,.30)'});
}

function renderNoReactivationAlerts(){
  try{
    const rows=(CLIENTES_SEM_MOVIMENTO||[]).map((r,i)=>({...r,_idx:i,_owner:reativacaoOwnerInfo(r)}));
    if(!rows.length) return '';
    const hoje=dateOnlyISO(new Date());
    const todayOnly=(x)=>dateOnlyISO(x.server_time||x.criado_em||x.data||x.created_at||'')===hoje;
    const sentLogs=(COB_LOGS||[]).filter(x=>String(x.titulo||'')==='REATIVACAO' && todayOnly(x));
    const limpaNome=x=>normName(String(x||'').replace(/\(F\d+\)/gi,'').replace(/gerente\s*f\d+/gi,'gerente'));
    const byOwner={};
    rows.forEach(r=>{
      const o=r._owner||{}; const key=String(o.key||''); if(!key) return;
      if(!byOwner[key]) byOwner[key]={key,nome:o.label||o.nome||key, filial:o.filial||r.filial, total:0, sent:0, login:String(o.login||'')};
      byOwner[key].total++;
      if(isReativacaoEnviadaHoje(r)) byOwner[key].sent++;
    });
    const usuarioEnviouHoje=(o)=>sentLogs.some(x=>{
      const fLog=String(x.filial||'').toUpperCase();
      const fOwner=String(o.filial||'').toUpperCase();
      const u=limpaNome(x.usuario||'');
      const d=limpaNome(x.destino_nome||'');
      const l=String(x.login||'').toLowerCase();
      const on=limpaNome(o.nome||o.label||'');
      const ol=String(o.login||'').toLowerCase();
      const ok=String(o.key||'');
      const parcela=String(x.parcela||'');
      const mesmaFilial=(!fOwner || !fLog || fOwner===fLog);
      if(!mesmaFilial) return false;
      if(ok && (String(x.owner_key||'')===ok || parcela.includes(ok))) return true;
      if(ol && (l===ol || String(x.usuario||'').toLowerCase()===ol)) return true;
      return !!(on && (u===on || d===on || u.includes(on) || on.includes(u) || d.includes(on) || on.includes(d)));
    });
    const faltantes=Object.values(byOwner)
      .filter(o=>o.total>0 && o.sent===0 && !usuarioEnviouHoje(o) && !usuarioReativacaoLocalHoje(o))
      .sort((a,b)=>String(a.filial).localeCompare(String(b.filial),'pt-BR')||Number(b.total)-Number(a.total));
    if(!faltantes.length) return '';
    return renderAvisoTicker('Clientes sem movimento: ninguém acionou ainda','Usuários com lista de reativação e zero WhatsApp enviado hoje.', faltantes.map(e=>({nome:e.nome, info:`${e.filial||''} · ${e.total} clientes`})), {icon:'🧡',color:'rgba(245,158,11,.30)'});
  }catch(e){console.warn('renderNoReactivationAlerts',e); return '';}
}

function renderTopMural(){
  if(!topMural) return;
  if(mainTab!=='inicio'){topMural.classList.add('hidden'); topMural.innerHTML=''; return;}
  const camp=renderCampaignStrip();
  if(camp){topMural.innerHTML=`<div class="inicio-campaign-compact">${camp}</div>`; topMural.classList.remove('hidden');}
  else {topMural.classList.add('hidden'); topMural.innerHTML='';}
}
function renderInicioTab(){
  if(!inicioSection) return;
  const oper=renderHighlights();
  const murais=oper?`<div class="inicio-operacional-compact">${oper}</div>`:'<div class="glass panel" style="padding:12px 14px;margin-top:10px;border-color:rgba(34,197,94,.22)"><strong>✅ Operação em dia</strong><div class="hint">Sem alertas de cobrança, reativação, vendas ou duplicidade neste momento.</div></div>';
  inicioSection.innerHTML=`<div class="inicio-compact"><div class="glass panel" style="padding:12px 16px;margin-bottom:10px"><div class="section-head" style="margin:0"><div><h2 style="font-size:17px;margin:0">🏠 Início</h2><div class="hint">Avisos do Master/Diretor ficam acima dos cards. Cards principais e murais operacionais ficam reunidos aqui.</div></div></div></div>${murais}</div>`;
}

function renderHighlights(){const cobrarParts=[]; const vendasParts=[]; const filiais=flattenFiliais(); const vendedores=flattenVendedores(); const calcDelta=(e)=>{const delta=Number(e.var_pago_delta||0); const prev=Math.max(Math.abs(Number(e.pago||0)-delta),1); const perc=(Math.abs(delta)/prev)*100; return {delta,perc};}; const bestFil=filiais.filter(e=>Number(e.var_pago_delta||0)>0).sort((a,b)=>Number(b.var_pago_delta||0)-Number(a.var_pago_delta||0))[0]; const bestVend=vendedores.filter(e=>Number(e.var_pago_delta||0)>0).sort((a,b)=>Number(b.var_pago_delta||0)-Number(a.var_pago_delta||0))[0]; if(bestFil){const d=calcDelta(bestFil); cobrarParts.push(`<div class="glass panel highlight-pulse" style="margin-bottom:12px;padding:16px 18px"><div class="section-head" style="margin:0"><div><h2 style="margin:0;font-size:20px">🏆 Destaque da semana · Filial</h2><div class="hint">${esc(filialLabel(bestFil.filial))} recebeu ${R(d.delta)} a mais que a referência anterior</div></div>${renderDeltaPill(d.delta,d.perc)}</div></div>`);} if(bestVend){const d=calcDelta(bestVend); cobrarParts.push(`<div class="glass panel highlight-pulse" style="margin-bottom:16px;padding:16px 18px"><div class="section-head" style="margin:0"><div><h2 style="margin:0;font-size:20px">🥇 Destaque da semana · Vendedor</h2><div class="hint">${esc(bestVend.nome)} recebeu ${R(d.delta)} a mais que a referência anterior</div></div>${renderDeltaPill(d.delta,d.perc)}</div></div>`);}  const achievers=currentEntities().filter(e=>calcMeta(e).geral>=50); if(achievers.length){cobrarParts.push(`<div class="glass panel" style="margin-bottom:16px;padding:14px 18px"><div class="section-head" style="margin:0"><div><h2 style="margin:0;font-size:18px">🔔 Metas atingidas</h2><div class="hint">${achievers.length} colaboradores/filiais com meta alcançada</div></div></div><div class="legend-inline">${achievers.slice(0,10).map(e=>`<span><i class="dot" style="background:#2f67f6"></i>${esc(e.nome)} ${pct(calcMeta(e).geral)}</span>`).join('')}</div></div>`);} cobrarParts.push(renderDuplicidadeCarteiraBanner());
cobrarParts.push(renderNoChargeAlerts());
cobrarParts.push(renderNoReactivationAlerts());
const metaDiariaBatida=renderMetaDiariaBatidaAlerts(); if(metaDiariaBatida) vendasParts.push(metaDiariaBatida);
const topVendaVend=bestLiveSalesEntity(vendedores,'venda_filial_vendedor_meta'); const topServVend=bestLiveSalesEntity(vendedores,'servico_filial_vendedor_ouro_fob'); const topVendaFil=bestLiveSalesEntity(filiais,'venda_filial_meta'); const topServFil=bestLiveSalesEntity(filiais,'servico_filial_ouro_fob');
vendasParts.push(`<div class="glass panel sales-panel" style="margin-bottom:16px;padding:16px 18px"><div class="section-head" style="margin:0 0 10px"><div><h2 style="margin:0;font-size:20px">💲 Vendas do mês</h2><div class="hint">Melhores resultados acumulados de venda e serviço do mês.</div></div></div><div class="legend-inline">${topVendaVend?`<span><i class="dot" style="background:#f97316"></i>Venda vendedor: ${esc(topVendaVend.ent.nome)} ${R(topVendaVend.val||0)}</span>`:''}${topServVend?`<span><i class="dot" style="background:#f59e0b"></i>Serviço vendedor: ${esc(topServVend.ent.nome)} ${R(topServVend.val||0)}</span>`:''}${topVendaFil?`<span><i class="dot" style="background:#fb923c"></i>Venda filial: ${esc(filialLabel(topVendaFil.ent.filial))} ${R(topVendaFil.val||0)}</span>`:''}${topServFil?`<span><i class="dot" style="background:#fdba74"></i>Serviço filial: ${esc(filialLabel(topServFil.ent.filial))} ${R(topServFil.val||0)}</span>`:''}</div></div>`);
return `<div class="section-head" style="margin-bottom:8px"><div><h2>📌 Mural de cobrança</h2><div class="hint">Notificações e destaques do dia da cobrança.</div></div></div>${cobrarParts.join('')}<div class="section-head" style="margin:20px 0 8px"><div><h2>💲 Mural de vendas</h2><div class="hint">Comparativos de venda e serviço do dia/semana.</div></div></div>${vendasParts.join('')}`}

function bindSpecialCards(){document.querySelectorAll('[data-action="cred-card"]').forEach(el=>{el.style.cursor='pointer';el.onclick=(ev)=>{ev.preventDefault();ev.stopPropagation();const node=ev.currentTarget.closest('[data-action="cred-card"]')||ev.currentTarget;const ds=node.dataset||{};openCrediaristaPanel(ds.login||'',ds.filial||'',ds.nome||'');return false};el.onkeydown=(ev)=>{if(ev.key==='Enter' || ev.key===' '){ev.preventDefault();const node=ev.currentTarget.closest('[data-action="cred-card"]')||ev.currentTarget;const ds=node.dataset||{};openCrediaristaPanel(ds.login||'',ds.filial||'',ds.nome||'')}}});document.querySelectorAll('[data-action="third-card"]').forEach(el=>{el.style.cursor='pointer';el.onclick=(ev)=>{ev.preventDefault();ev.stopPropagation();openThirdChargePanel();return false};el.onkeydown=(ev)=>{if(ev.key==='Enter' || ev.key===' '){ev.preventDefault();openThirdChargePanel()}}})}
function openEntityFromRowPayload(payload){try{const ref=JSON.parse(decodeURIComponent(payload)); if(ref?.type==='terceiro') return openThirdChargePanel(); if(ref?.type==='crediarista') return openCrediaristaPanel(ref.login||'',ref.filial||'',ref.nome||''); return openEntity(ref);}catch(e){}}
function renderEntityRow(ent){const m=calcMeta(ent); const isThird=!!(ent?.is_terceiro||ent?.type==='terceiro'); const isCred=!!(ent?.is_crediarista||ent?.type==='crediarista'); const ref=isThird?{type:'terceiro',filial:'FTER',nome:COBRANCA10_NOME}:isCred?{type:'crediarista',login:String(ent.login||crediaristaLoginByFilial(ent.filial)||'').toLowerCase(),filial:ent.filial,nome:ent.nome}:{type:ent.type,filial:ent.filial,nome:ent.nome}; const payload=encodeURIComponent(JSON.stringify(ref)); const sales=summarizeSalesCard(ent)||{}; const role=isThird?'Cobrança terceiro':isCred?'Crediarista':(ent.type==='filial'?'Filial':'Colaborador'); return `<div class="entity-row" onclick="openEntityFromRowPayload('${payload}')"><div class="entity-cell"><div class="v">${esc(ent.nome||filialLabel(ent.filial)||'')}</div><div class="small muted">${esc(role)} ${ent.filial?`· ${esc(ent.filial)}`:''}</div></div><div class="entity-cell"><div class="k">Pendente</div><div class="v red">${R(ent.pendente||0)}</div></div><div class="entity-cell"><div class="k">Recebido</div><div class="v green">${R(ent.pago||0)}</div></div><div class="entity-cell"><div class="k">Grave</div><div class="v red">${pct(m.grave.perc||0)}</div></div><div class="entity-cell"><div class="k">Alerta</div><div class="v orange">${pct(m.alerta.perc||0)}</div></div><div class="entity-cell"><div class="k">Atenção</div><div class="v">${pct(m.atencao.perc||0)}</div></div><div class="entity-cell"><div class="k">Meta geral</div><div class="v blue">${pct(m.geral||0)}</div></div><div class="entity-cell"><div class="k">Vendas/serviços</div><div class="v">${sales.n!=null?pct(sales.n):'—'}</div></div></div>`}
function renderEntityList(entities){return `<div class="entity-list"><div class="entity-row entity-row-head"><div>Nome / tipo</div><div>Pendente</div><div>Recebido</div><div>Grave</div><div>Alerta</div><div>Atenção</div><div>Meta</div><div>Vendas</div></div>${entities.map(renderEntityRow).join('')}</div>`}
function renderList(){const entities=currentEntities();const title=mainTab==='filiais'?'🏬 Filiais':'👥 Colaboradores';const hint=`${entities.length} ${mainTab==='filiais'?'filiais':'colaboradores'} exibidos`; const useRows=usuarioAtual?.tipo==='master'; listSection.innerHTML=`${renderCampaignStrip()}<div class="section-head"><div><h2>${title}</h2><div class="hint">Clique em uma linha para abrir a tela individual completa.</div></div><div class="hint">${hint}</div></div>${useRows?renderEntityList(entities):`<div class="grid-cards">${entities.map(renderEntityCard).join('')}</div>`}`; bindSpecialCards()}
function findEntity(ref){const n=String(ref?.nome||'').toLowerCase(); const f=String(ref?.filial||'').toUpperCase(); const t=String(ref?.type||'').toLowerCase(); if(t==='terceiro' || ref?.is_terceiro || n===String(COBRANCA10_NOME).toLowerCase() || n===String(COBRANCA10_LOGIN).toLowerCase() || f==='FTER'){return thirdChargeEntity()} if(t==='crediarista' || ref?.is_crediarista || n.startsWith('crediarista') || String(ref?.login||'').toLowerCase().startsWith('crediaristaf')){return crediaristaEntityByLogin(ref?.login||ref?.nome||ref?.filial)} if(ref.type==='filial'){return flattenFiliais().find(x=>x.filial===ref.filial)} return flattenVendedores().find(x=>x.filial===ref.filial && x.nome===ref.nome)}
function keysFromLogsForCommission(logs){const out=new Set(); (logs||[]).forEach(l=>{out.add(cobrancaRowKey(l)); const alt=[String(l.cliente||'').trim().toUpperCase(),String(l.titulo||'').trim(),String(l.parcela||'').trim()].join('|'); out.add(alt);}); return out}
function key3Cob(r){return [String(r.cliente||r.nome||'').trim().toUpperCase(),String(r.titulo||'').trim(),String(r.parcela||'').trim()].join('|')}
function renderTerceiroCommission(ent){const isCred=!!(ent?.is_crediarista||ent?.type==='crediarista'); const baseCfg=isCred?entityConfig({type:'vendedor',nome:ent.nome,filial:ent.filial}):entityConfig({type:'vendedor',nome:COBRANCA10_NOME,filial:'FTER'}); const cfg=commissionCfg(baseCfg); const policy=(isCred?(cfg.camp_cob_crediarista||[]):(cfg.camp_cobranca_terceiro||[])); const policyOk=Array.isArray(policy)&&policy.length?policy:(isCred?defaultCampCrediarista():defaultCampTerceiro()); const byFaixa={atencao:{pct:0,cobrado:0,recebido:0,comissao:0},alerta:{pct:0,cobrado:0,recebido:0,comissao:0},grave:{pct:0,cobrado:0,recebido:0,comissao:0}}; policyOk.forEach(r=>{const fx=String(r.faixa||'').toLowerCase(); if(byFaixa[fx]) byFaixa[fx].pct=Number(String(r.pct||0).replace(',','.'))||0}); const mesAtual=dateOnlyISO(new Date()).slice(0,7); const userKeys=isCred?[String(ent.login||'').toLowerCase(),String(ent.nome||'').toLowerCase()]:[COBRANCA10_NOME.toLowerCase(),COBRANCA10_LOGIN]; const cobrados=(COB_LOGS||[]).filter(x=>userKeys.includes(String(x.usuario||'').toLowerCase()) && dateOnlyISO(x.server_time||x.criado_em||x.data||'').slice(0,7)===mesAtual); const keys=keysFromLogsForCommission(cobrados); const srcCli=isCred?(CLIENTES_CREDIARISTA?.[String(ent.login||'').toLowerCase()]||{grave:[],alerta:[],atencao:[]}):(CLIENTES_TERCEIRO||{grave:[],alerta:[],atencao:[]}); const srcRec=getRecebimentos(ent)||{grave:[],alerta:[],atencao:[]}; ['atencao','alerta','grave'].forEach(fx=>{byFaixa[fx].cobrado=(srcCli?.[fx]||[]).filter(r=>keys.has(cobrancaRowKey(r))||keys.has(key3Cob(r))).length; (srcRec?.[fx]||[]).forEach(r=>{const pagMes=dateOnlyISO(r.pagamento||r.data_pagamento||'').slice(0,7); if((keys.has(cobrancaRowKey(r))||keys.has(key3Cob(r))) && pagMes===mesAtual){byFaixa[fx].recebido+=Number(r.pago||0)}}); byFaixa[fx].comissao=byFaixa[fx].recebido*(byFaixa[fx].pct/100)}); const total=Object.values(byFaixa).reduce((a,b)=>a+b.comissao,0); const item=(t,v,s='')=>`<div class="commission-item unlocked ${s}"><div class="k">${t}</div><div class="v">${v}</div></div>`; return `<div class="glass panel commission-card"><h3>💵 ${isCred?'Comissão crediarista':'Comissão cobrança terceiro'} <span class="note">· só títulos cobrados pelo usuário e pagos no mês</span></h3><div class="commission-grid">${item('Atenção %',String(byFaixa.atencao.pct.toFixed(2)).replace('.',',')+'%')}${item('Alerta %',String(byFaixa.alerta.pct.toFixed(2)).replace('.',',')+'%')}${item('Grave %',String(byFaixa.grave.pct.toFixed(2)).replace('.',',')+'%')}${item('Recebido atenção',R(byFaixa.atencao.recebido||0))}${item('Recebido alerta',R(byFaixa.alerta.recebido||0))}${item('Recebido grave',R(byFaixa.grave.recebido||0))}${item('Comissão atenção',R(byFaixa.atencao.comissao||0))}${item('Comissão alerta',R(byFaixa.alerta.comissao||0))}${item('Comissão grave',R(byFaixa.grave.comissao||0))}${item('Total previsto',R(total||0),'total-final')}</div><div class="commission-note">${esc(CONFIG_META?.comissao_pagamento_texto||'A comissão reinicia a cada mês e o pagamento é previsto para o dia 25 do mês seguinte.')}</div></div>`}
function openCrediaristaPanel(login, filial, nome){
  const filialNorm=String(filial||'').toUpperCase();
  const loginNorm=String(login||crediaristaLoginByFilial(filialNorm)||'').toLowerCase();
  const nomeNorm=String(nome||`CREDIARISTA${filialNorm}`);
  // V10.7: abre exatamente a mesma entidade usada na lista inicial.
  // Antes a lista calculava a meta pela carteira do crediarista, mas a tela aberta
  // herdava a meta da filial, causando divergência tipo F3 13% na lista e 1% ao abrir.
  const entLista=(crediaristaEntities()||[]).find(x=>
    String(x.login||'').toLowerCase()===loginNorm ||
    String(x.filial||'').toUpperCase()===filialNorm ||
    String(x.nome||'').toLowerCase()===nomeNorm.toLowerCase()
  );
  let ent=entLista;
  if(!ent){
    const src=(CLIENTES_CREDIARISTA?.[loginNorm])||CLIENTES_FIL?.[filialNorm]||{grave:[],alerta:[],atencao:[]};
    const rsrc=getRecebimentos({type:'crediarista',login:loginNorm,filial:filialNorm,nome:nomeNorm,is_crediarista:true})||{grave:[],alerta:[],atencao:[]};
    const gp=(src.grave||[]).reduce((a,b)=>a+Number(b.pendente||0),0);
    const ap=(src.alerta||[]).reduce((a,b)=>a+Number(b.pendente||0),0);
    const tp=(src.atencao||[]).reduce((a,b)=>a+Number(b.pendente||0),0);
    const gr=(rsrc.grave||[]).reduce((a,b)=>a+Number(b.pago||0),0);
    const ar=(rsrc.alerta||[]).reduce((a,b)=>a+Number(b.pago||0),0);
    const tr=(rsrc.atencao||[]).reduce((a,b)=>a+Number(b.pago||0),0);
    ent={type:'crediarista',login:loginNorm,filial:filialNorm,nome:nomeNorm,is_crediarista:true,only_cobranca:true,
      pendente:gp+ap+tp,pago:gr+ar+tr,total:gp+ap+tp+gr+ar+tr,perc_filial:100,
      grave_pend:gp,alerta_pend:ap,atencao_pend:tp,grave_rec:gr,alerta_rec:ar,atencao_rec:tr};
  }
  currentDetailRef={type:'crediarista',filial:filialNorm,nome:ent.nome||nomeNorm,login:loginNorm};
  mascotCongrats(ent);
  document.getElementById('mainScreen').classList.add('hidden');
  detailScreen.classList.remove('hidden');
  return renderCrediaristaDetail(ent);
}
function openThirdChargePanel(){const ent=thirdChargeEntity(); currentDetailRef={type:'terceiro',filial:'FTER',nome:ent.nome}; mascotCongrats(ent); document.getElementById('mainScreen').classList.add('hidden'); detailScreen.classList.remove('hidden'); renderTerceiroDetail(ent)}

// Clique blindado dos cards especiais: não depende de onclick inline.
document.addEventListener('click', function(ev){
  const cred = ev.target.closest('[data-action="cred-card"]');
  if(cred){
    ev.preventDefault();
    ev.stopPropagation();
    openCrediaristaPanel(cred.dataset.login||'', cred.dataset.filial||'', cred.dataset.nome||'');
    return false;
  }
  const third = ev.target.closest('[data-action="third-card"]');
  if(third){
    ev.preventDefault();
    ev.stopPropagation();
    openThirdChargePanel();
    return false;
  }
}, true);

window.openCrediaristaPanel=openCrediaristaPanel;
window.openThirdChargePanel=openThirdChargePanel;

function renderTerceiroDetail(ent){const src=getClientesEnt(ent); const totalTit=(src.grave?.length||0)+(src.alerta?.length||0)+(src.atencao?.length||0); detailScreen.innerHTML=`${renderUpdateStrip()}<div class="back-row">${renderBackButton()}<div><h2>${esc(ent.nome)}</h2><div class="sub">${ent.is_crediarista?`Painel de cobrança da filial ${ent.filial} · base configurada ${Number(ent.pct_base||100)}% · recebido só por cobrança própria paga`:`Painel de cobrança terceirizada · percentual global sem duplicidade`}</div></div><div class="badge">${ent.is_crediarista?'🧾 Crediarista':'🤝 Cobrança terceiro'}</div></div><div class="detail-top"><div class="glass panel"><h3>🧾 Resumo da carteira</h3><div class="metrics-grid"><div class="metric"><div class="k">Títulos</div><div class="v">${totalTit}</div></div><div class="metric"><div class="k">Pendente</div><div class="v" style="color:var(--red)">${R(ent.pendente||0)}</div></div><div class="metric"><div class="k">Recebido</div><div class="v" style="color:var(--green)">${R(ent.pago||0)}</div></div><div class="metric"><div class="k">Cobrado hoje</div><div class="v">${getCobradosHoje(ent).length}</div></div></div><div class="legend-inline" style="margin-top:12px"><span><i class="dot" style="background:var(--red)"></i>Grave ${src.grave?.length||0}</span><span><i class="dot" style="background:var(--orange)"></i>Alerta ${src.alerta?.length||0}</span><span><i class="dot" style="background:var(--yellow)"></i>Atenção ${src.atencao?.length||0}</span></div></div><div>${renderTerceiroCommission(ent)}</div></div><div class="accordion"><div class="acc-head" onclick="toggleAcc(this)">💰 Recebimentos por faixa <span class="acc-hint">clique para abrir</span></div><div class="acc-body">${renderRecebimentos(ent)}</div></div><div class="accordion"><div class="acc-head" onclick="toggleAcc(this)"><span>🧾 Relatório de cobranças</span><span class="acc-hint">clique para abrir</span></div><div class="acc-body">${renderCobrancasEnt(ent)}</div></div>`}

// ── PAINEL CREDIARISTA ─────────────────────────────────────────────────────
// Mostra: resumo da carteira espelhada + meta de cobrança + recebimentos + cobranças
// NÃO mostra: painel de vendas, comissão mercantil, campanhas de venda
function renderCrediaristaDetail(ent){
  const src=getClientesEnt(ent);
  const totalTit=(src.grave?.length||0)+(src.alerta?.length||0)+(src.atencao?.length||0);
  const meta=calcMeta(ent);
  detailScreen.innerHTML=`
    ${renderInboxBanner()}
    ${renderUpdateStrip()}
    <div class="back-row">
      ${renderBackButton()}
      <div>
        <h2>${esc(ent.nome)}</h2>
        <div class="sub">Painel de cobrança espelhado da filial ${esc(ent.filial)} · mesma base do gerente/filial em tempo real</div>
      </div>
      <div class="badge">🧾 Crediarista</div>
    </div>
    <div class="detail-top">
      <div class="glass panel">
        <h3>🎯 Meta de cobrança <span class="note">· espelhada da filial</span></h3>
        <div class="mega-progress">
          <div class="ring-wrap">${renderPiggyBank(meta.geral)}</div>
          <div>
            <div class="metrics-grid">
              <div class="metric"><div class="k">Títulos na carteira</div><div class="v">${totalTit}</div></div>
              <div class="metric"><div class="k">Pendente total</div><div class="v" style="color:var(--red)">${R(ent.pendente||0)}</div></div>
              <div class="metric"><div class="k">Recebido</div><div class="v" style="color:var(--green)">${R(ent.pago||0)}</div></div>
              <div class="metric"><div class="k">Cobrado hoje</div><div class="v">${getCobradosHoje(ent).length}</div></div>
            </div>
            <div class="legend-inline" style="margin-top:12px">
              <span><i class="dot" style="background:var(--red)"></i>Grave ${src.grave?.length||0} títulos</span>
              <span><i class="dot" style="background:var(--orange)"></i>Alerta ${src.alerta?.length||0} títulos</span>
              <span><i class="dot" style="background:var(--yellow)"></i>Atenção ${src.atencao?.length||0} títulos</span>
            </div>
          </div>
        </div>
        <div class="meta-grid">
          ${renderMetaBox('Grave','var(--red)',meta.grave)}
          ${renderMetaBox('Alerta','var(--orange)',meta.alerta)}
          ${renderMetaBox('Atenção','var(--yellow)',meta.atencao)}
          ${renderMetaBox('Meta geral','var(--blue)',{perc:meta.geral,alvo:meta.grave.alvo+meta.alerta.alvo+meta.atencao.alvo,rec:meta.grave.rec+meta.alerta.rec+meta.atencao.rec})}
        </div>
        <div style="margin-top:14px">${renderMascotStatus(meta.geral,'Cobrança')}</div>
      </div>
      <div>${renderTerceiroCommission(ent)}</div>
    </div>
    <div class="accordion">
      <div class="acc-head" onclick="toggleAcc(this)">💰 Recebimentos por faixa <span class="acc-hint">clique para abrir</span></div>
      <div class="acc-body">${renderRecebimentos(ent)}</div>
    </div>
    <div class="accordion">
      <div class="acc-head" onclick="toggleAcc(this)">🧾 Relatório de cobranças <span class="acc-hint">clique para abrir</span></div>
      <div class="acc-body">${renderCobrancasEnt(ent)}</div>
    </div>
  `;
}

function openEntity(ref){if(ref && (ref.type==='crediarista' || ref.is_crediarista)){return openCrediaristaPanel(ref.login||'', ref.filial||'', ref.nome||'')} const ent=findEntity(ref); if(!ent) return; currentDetailRef={type:ent.type,filial:ent.filial,nome:ent.nome,login:ent.login||''}; mascotCongrats(ent); try{renderLaranjitoNotify(); showLaranjitoOncePerAccess()}catch(e){}; document.getElementById('mainScreen').classList.add('hidden'); detailScreen.classList.remove('hidden'); if(ent.is_terceiro || ent.type==='terceiro'){return renderTerceiroDetail(ent)} if(ent.is_crediarista || ent.type==='crediarista'){return openCrediaristaPanel(ent.login||'', ent.filial||'', ent.nome||'')} const meta=calcMeta(ent); const bonus=getBonus(meta.cfg,meta.geral); const deltaVal=Number(ent.var_pago_delta||0); const prevBase=Math.max(Math.abs(Number(ent.pago||0)-deltaVal),1); const pctFallback=(Math.abs(deltaVal)/prevBase)*100; const compPerc=(ent.var_pago_perc==null || Math.abs(Number(ent.var_pago_perc||0))<0.01)?pctFallback:Math.abs(Number(ent.var_pago_perc||0)); detailScreen.innerHTML=`${usuarioAtual && usuarioAtual.tipo!=='master' ? renderInboxBanner() : ''}${renderUpdateStrip()}<div class="back-row">${renderBackButton()}<div><h2>${ent.type==='filial'?filialLabel(ent.filial):esc(ent.nome)}</h2><div class="sub">${ent.type==='filial'?'Painel individual da filial':'Painel individual do vendedor'} · ${ent.filial}</div></div><div class="badge">${ent.type==='filial'?'🏬 Filial':'👤 Vendedor'}</div></div><div class="detail-top"><div class="glass panel"><h3>🎯 Meta do mês <span class="note">· Não acumulativo</span></h3><div class="mega-progress"><div class="ring-wrap">${renderPiggyBank(meta.geral)}</div><div><div class="metrics-grid"><div class="metric"><div class="k">Pendente</div><div class="v" style="color:var(--red)">${R(ent.pendente||0)}</div></div><div class="metric"><div class="k">Recebido</div><div class="v" style="color:var(--green)">${R(ent.pago||0)}</div></div><div class="metric"><div class="k">% da filial</div><div class="v">${pct(ent.perc_filial||100)}</div></div><div class="metric"><div class="k">Configuração usada</div><div class="v">${Number(meta.cfg.grave_pct||0)}/${Number(meta.cfg.alerta_pct||0)}/${Number(meta.cfg.atencao_pct||0)}</div></div><div class="metric"><div class="k">Comparado a ontem</div><div class="v" style="font-size:16px">${renderDeltaPill(ent.var_pago_delta,compPerc)} <span>${R(Math.abs(Number(ent.var_pago_delta||0)))}</span></div></div></div><div class="legend-inline" style="margin-top:12px"><span><i class="dot" style="background:var(--red)"></i>Grave alvo ${R(meta.grave.alvo)} · recebido ${R(meta.grave.rec)}</span><span><i class="dot" style="background:var(--orange)"></i>Alerta alvo ${R(meta.alerta.alvo)} · recebido ${R(meta.alerta.rec)}</span><span><i class="dot" style="background:var(--yellow)"></i>Atenção alvo ${R(meta.atencao.alvo)} · recebido ${R(meta.atencao.rec)}</span></div></div></div><div class="meta-grid">${renderMetaBox('Grave','var(--red)',meta.grave)}${renderMetaBox('Alerta','var(--orange)',meta.alerta)}${renderMetaBox('Atenção','var(--yellow)',meta.atencao)}${renderMetaBox('Meta geral','var(--blue)',{perc:meta.geral,alvo:meta.grave.alvo+meta.alerta.alvo+meta.atencao.alvo,rec:meta.grave.rec+meta.alerta.rec+meta.atencao.rec})}</div><div style="height:18px"></div><h3>🌊 Gráfico Geral Contas a Receber</h3>${renderSingleBars(ent,meta,true)}<div style="height:16px"></div><div class="glass panel"><h3>🏆 Bônus e premiações <span class="note">· Não acumulativo</span></h3>${renderBonusBox(meta.cfg,meta.geral)}</div></div><div>${renderSalesPanel(ent)}<div style="height:16px"></div>${renderCommissionSummary(ent)}<div style="height:16px"></div>${renderCampaignSummary(ent)}</div></div>${renderReativacaoEnt(ent)}<div class="accordion"><div class="acc-head" onclick="toggleAcc(this)">💰 Recebimentos por faixa <span class="acc-hint">clique para ${'abrir'}</span></div><div class="acc-body">${renderRecebimentos(ent)}</div></div><div class="accordion"><div class="acc-head" onclick="toggleAcc(this)">🧾 Relatório de cobranças <span class="acc-hint">clique para ${'abrir'}</span></div><div class="acc-body">${renderCobrancasEnt(ent)}</div></div>`}
function canVerComissionamento(){return usuarioAtual?.tipo==='master'}
function renderCommissionSummary(ent){if(!canVerComissionamento()) return '';
  const c=calcCommissionSummary(ent);
  const totalLiberado = c.elegivelMercantil && c.elegivelServicos;
  const totalExibido = totalLiberado ? c.totalPrevisto : 0;
  const moneyCell=(title,val,locked=false,extra='')=>`<div class="commission-item ${locked?'locked':''} ${!locked?'unlocked':''} ${extra}"><div class="k">${title}</div><div class="v">${R(val||0)}</div></div>`;
  const pctCell=(title,val,locked=false)=>`<div class="commission-item ${locked?'locked':''} ${!locked?'unlocked':''}"><div class="k">${title}</div><div class="v">${String(Number(val||0).toFixed(2)).replace('.',',')}%</div></div>`;
  const rentNote = c.rentUnlocked
    ? `Rentabilidade atual ${String(Number(c.rentAtual||0).toFixed(2)).replace('.',',')}% · faixa aplicada ${c.rentFaixaTxt}.`
    : `Rentabilidade atual ${String(Number(c.rentAtual||0).toFixed(2)).replace('.',',')}% · bloqueada até bater 50% da meta de cobrança.`;
  return `<div class="glass panel commission-card"><h3>💵 Comissionamento previsto <span class="note">· calculado pela política salva</span></h3>${c.metaAtingida?`<div class="meta-hit-banner"><img src="${LARANJITO}" alt=""><span>Meta liberada! O Laranjito está comemorando sua liberação de comissão/bonus.</span></div>`:''}<div class="commission-grid">${`<div class="commission-item unlocked"><div class="k">Faixa aplicada</div><div class="v" style="font-size:16px">${esc(c.faixaTxt)}</div></div>`}${pctCell('% comissão mercantil',c.comPerc,!c.elegivelMercantil)}${pctCell('% serviços',c.servPct,!c.elegivelServicos)}${pctCell('% caminhão',c.camPct,!c.elegivelServicos)}${moneyCell('Comissão vendas',c.vendasComissao,!c.elegivelMercantil)}${moneyCell('Comissão serviços',c.servicosComissao,!c.elegivelServicos)}${moneyCell('Comissão caminhão',c.caminhaoComissao,!c.elegivelServicos)}${moneyCell('Bônus por meta',c.bonusMeta,!c.bonusLiberado)}${moneyCell('Rentab 48%',c.rent48,!(c.rentUnlocked && c.rentAtual>=48))}${moneyCell('Rentab 52,15%',c.rent52,!(c.rentUnlocked && c.rentAtual>=52.15))}${moneyCell('Rentab 55,50%',c.rent55,!(c.rentUnlocked && c.rentAtual>=55.50))}${moneyCell('Total previsto',totalExibido,!totalLiberado,'total-final '+(!totalLiberado?'total-locked':''))}</div><div class="commission-note">Base mercantil bruta: ${R(c.vendaRealBruto||0)} · Caminhão abatido: ${R(c.camReal||0)} · Mercantil líquido para comissão: ${R(c.vendaReal||0)} · Serviço: ${R(c.servReal||0)}. Mínimo vendas ${pct(c.minVenda)} · mínimo serviços/caminhão ${pct(c.minServico)} · rentab exige cobrança 50% + mercantil ${pct(c.rentMinMercantil)}. ${rentNote}</div></div>`
}

function backToMain(){currentDetailRef=null; try{renderLaranjitoNotify()}catch(e){}; detailScreen.classList.add('hidden');document.getElementById('mainScreen').classList.remove('hidden')}
function renderMetaBox(title,color,obj){return `<div class="meta-card"><div class="meta-title">${title}</div><div class="meta-main" style="color:${color}">${pct(obj.perc||0)}</div><div class="meta-sub">Alvo: ${R(obj.alvo||0)}</div><div class="meta-sub">Recebido: ${R(obj.rec||0)}</div></div>`}
function normSalesText(s){return String(s||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/[^A-Z0-9 ]/gi,' ').replace(/\s+/g,' ').trim().toUpperCase()}
function salesCell(row, keys){for(const k of keys){if(row && row[k]!=null && String(row[k]).trim()!=='') return String(row[k]).trim()} return ''}
function filialAliases(f){const raw=String(f||'').toUpperCase(); if(raw.includes('90')) return ['FILIAL 90/99','FILIAL 90','90/99','DEPOSITO MOVEIS','DEPOSITO']; const n=parseInt((raw.match(/\d+/)||[''])[0]||'0',10); if(!n) return [raw]; const two=String(n).padStart(2,'0'); return [`FILIAL ${two}`,`FILIAL ${n}`,`(${two})`,`(${n})`,`F${n}`]}

function servicosKeyNome(nome){
  return normSalesText(nome).replace(/\s+/g,' ').trim();
}
function servicosEntidade(ent){
  const out=[];
  const data=SERVICOS_RELATORIO||{};
  if(!ent) return out;
  if(ent.type==='filial'){
    const item=(data.filiais||{})[ent.filial]||{};
    Object.entries(item.servicos||{}).forEach(([servico,valor])=>{
      out.push({servico, real_total:Number(valor||0), quantidade:0});
    });
  }else{
    const vend=(data.vendedores||{});
    const key1=`${servicosKeyNome(ent.nome)}_${ent.filial}`;
    let item=vend[key1];
    if(!item){
      item=Object.values(vend).find(v=>{
        return String(v?.filial||'').toUpperCase()===String(ent.filial||'').toUpperCase()
          && servicosKeyNome(v?.nome||'')===servicosKeyNome(ent.nome||'');
      });
    }
    Object.entries((item&&item.servicos)||{}).forEach(([servico,valor])=>{
      out.push({servico, real_total:Number(valor||0), quantidade:0});
    });
  }
  return out.sort((a,b)=>Number(_srvTotal(b)||0)-Number(_srvTotal(a)||0));
}
function renderServicosEntidade(ent){
  const rows=servicosEntidade(ent).filter(x=>Number(x.real_total||0)>0);
  if(!rows.length){
    return `<div class="glass panel"><h3>🛠️ Serviços por tipo</h3><div class="empty">Nenhum serviço localizado para ${ent?.type==='filial'?'esta filial':'este vendedor'} no relatório mensal.</div></div>`;
  }
  const total=rows.reduce((a,b)=>a+Number(b.real_total||0),0);
  return `<div class="glass panel"><div class="section-head" style="margin:0 0 10px"><div><h3 style="margin:0">🛠️ Serviços por tipo</h3><div class="hint">Relatório real de serviços do mês atual · total oficial ${R(total)}</div></div></div><div class="metrics-grid">${rows.slice(0,8).map(r=>`<div class="metric"><div class="k">${esc(String(r.servico||'Serviço').slice(0,34))}</div><div class="v" style="color:var(--blue-400);font-size:18px">${R(r.real_total||0)}</div></div>`).join('')}</div></div>`;
}

function servicosEntidadeTotal(ent){
  return servicosEntidade(ent).reduce((acc,r)=>acc+Number(r.real_total||0),0);
}
function serviceOfficialOverride(ent,key,row){
  if(!String(key||'').includes('servico')) return null;
  const total=servicosEntidadeTotal(ent);
  if(!(total>0)) return null;
  const metaTotalStr=salesCell(row,['Meta (R$) Total','Meta(R$) Total']);
  const metaPeriodoStr=salesCell(row,['Meta (R$) Período','Meta(R$) Período']);
  const metaTotal=salesNum(metaTotalStr);
  const metaPeriodo=salesNum(metaPeriodoStr);
  return {
    total: total,
    realizado: R(total),
    atingidoTotal: metaTotal>0 ? pct(total/metaTotal*100) : '0%',
    atingidoPeriodo: metaPeriodo>0 ? pct(total/metaPeriodo*100) : '0%',
    nota: 'relatório serviços'
  };
}

function rowMatchesFilial(ent,row){const joined=normSalesText([salesCell(row,['Filial','Filial_2','Vendedor','Vendedor_2','Nome','Nome_2']),salesCell(row,['Subgrupo'])].join(' ')); return filialAliases(ent.filial).some(a=>joined.includes(normSalesText(a)))}
function vendedorNomeAlvo(row,key){if(key==='venda_filial_vendedor_meta' || key==='servico_filial_vendedor_ouro_fob') return salesCell(row,['Vendedor_2','Nome_2','Nome','Vendedor']); return salesCell(row,['Vendedor','Vendedor_2','Nome','Nome_2'])}
function fuzzyContainsAllTokens(target, query){const t=normSalesText(target); const q=normSalesText(query); const toks=q.split(' ').filter(x=>x && x.length>1); return toks.length ? toks.every(tok=>t.includes(tok)) : false}
function rowMatchesVendedor(ent,row,key=''){const vendedorCampo=vendedorNomeAlvo(row,key); const filialCampo=salesCell(row,['Filial','Filial_2','Vendedor']); const byNome = fuzzyContainsAllTokens(vendedorCampo, ent.nome) || fuzzyContainsAllTokens(ent.nome, vendedorCampo); const byFilial = !filialCampo || filialAliases(ent.filial).some(a=>normSalesText(filialCampo).includes(normSalesText(a))) || rowMatchesFilial(ent,row); return byNome && byFilial}
function getSalesRows(ent, key){const base=METAS_VENDAS?.metas?.[key]; if(!base || !base.ok) return []; const rows=Array.isArray(base.linhas)?base.linhas:[]; return rows.filter(r=>ent.type==='filial'?rowMatchesFilial(ent,r):rowMatchesVendedor(ent,r,key))}
function salesTitleForRow(ent,row,key=''){return salesCell(row, ent.type==='filial'?['Subgrupo','Filial','Vendedor_2','Vendedor']:['Subgrupo','Vendedor_2','Vendedor','Filial']) || (ent.type==='filial'?filialLabel(ent.filial):ent.nome)}
function salesPercentClass(v){const n=parseFloat(String(v||'').replace('%','').replace(',','.'))||0; if(n>=100) return 'gold'; if(n>=80) return 'good'; if(n>=50) return 'warn'; return 'low'}
function salesNum(v){return parseFloat(String(v||'').replace('%','').replace(/\./g,'').replace(',','.'))||0}
function renderSalesProgress(v){const n=Math.max(0,Math.min(160,salesNum(v))); return `<div class="sales-progress"><span style="width:${Math.min(n,100)}%"></span></div><div class="sales-progress-label">${String(v||'0%')}</div>`}
function renderSalesMini(k,v,kind='',showMascot=false,wrap=false){const cls=[kind, wrap?'wrap':'']; if(kind==='atingido'||kind==='projetado') cls.push(salesPercentClass(v)); return `<div class="sales-mini ${cls.join(' ')}">${showMascot?`<img src="${LARANJITO}" class="laranjito-mini" alt="">`:''}<div class="k">${k}</div><div class="v">${v||'-'}</div>${(kind==='atingido')?renderSalesProgress(v):''}</div>`}
function renderSalesRows(ent, key, label){
  let rows=getSalesRows(ent,key);
  if(!rows.length) return `<div class="sales-card"><h4>${label}</h4><div class="sales-empty">Sem dados desta meta para ${ent.type==='filial'?'esta filial':'este vendedor'}.</div></div>`;

  // Para SERVIÇOS, o realizado oficial vem do relatório real de serviços por tipo.
  // A meta do SGI permanece como alvo, mas os campos "Realizado" e "% atingido" passam a bater
  // exatamente com o painel "Serviços por tipo" da mesma filial/vendedor.
  if(String(key||'').includes('servico') && servicosEntidadeTotal(ent)>0){
    rows=[rows[0]];
  }

  return `<div class="sales-card"><h4>${label}</h4><div class="sales-list">${rows.map(r=>{
    const srv=serviceOfficialOverride(ent,key,r);
    const ating=srv ? srv.atingidoTotal : salesCell(r,['Atingido Total']);
    const atingPeriodo=srv ? srv.atingidoPeriodo : salesCell(r,['Atingido Período']);
    const realizadoTotal=srv ? srv.realizado : esc(salesCell(r,['Realizado (R$) Total','Realizado(R$) Total']));
    const realizadoPeriodo=srv ? srv.realizado : esc(salesCell(r,['Realizado (R$) Período','Realizado(R$) Período']));
    const proj=salesCell(r,['Projetado (R$)','Projetado(R$)']);
    const showMascot=(parseFloat(String(ating).replace('%','').replace(',','.'))||0)>=100;
    const title=srv ? `${salesTitleForRow(ent,r,key)} · relatório serviços` : salesTitleForRow(ent,r,key);
    return `<div class="sales-row"><div class="sales-row-title">${esc(title)}</div><div class="sales-metrics">${renderSalesMini('Meta total',esc(salesCell(r,['Meta (R$) Total','Meta(R$) Total'])),'meta')}${renderSalesMini('Realizado total',realizadoTotal,'realizado')}${renderSalesMini('Atingido total',esc(ating),'atingido',showMascot)}${renderSalesMini('Meta período',esc(salesCell(r,['Meta (R$) Período','Meta(R$) Período'])),'meta')}${renderSalesMini('Realizado período',realizadoPeriodo,'realizado')}${renderSalesMini('Atingido período',esc(atingPeriodo),'atingido')}${renderSalesMini('Projetado',esc(proj),'projetado',showMascot)}${renderSalesMini(ent.type==='filial'?'Filial':'Vendedor',esc(salesCell(r, ent.type==='filial'?['Filial']:['Vendedor_2','Vendedor'])),'realizado',false,true)}</div></div>`;
  }).join('')}</div></div>`;
}
function summarizeSalesCard(ent){const keys=ent.type==='filial'?['venda_filial_meta']:['venda_filial_vendedor_meta']; let best=null; keys.forEach(k=>{getSalesRows(ent,k).forEach(r=>{const ating=salesCell(r,['Atingido Total']); const n=parseFloat(String(ating).replace('%','').replace(',','.'))||0; if(!best || n>best.n){best={row:r,key:k,n};}})}); return best}
function renderSalesCardSummary(ent){const s=summarizeSalesCard(ent); if(!s) return ''; const row=s.row; const showMascot=s.n>=100; return `<div class="card-sales"><div class="card-sales-title">💲 Vendas e metas</div><div class="card-sales-grid">${renderSalesMini('Realizado total',esc(salesCell(row,['Realizado (R$) Total','Realizado(R$) Total'])),'realizado')}${renderSalesMini('Atingido total',esc(salesCell(row,['Atingido Total'])),'atingido',showMascot)}${renderSalesMini('Projetado',esc(salesCell(row,['Projetado (R$)','Projetado(R$)'])),'projetado',showMascot)}</div></div>`}
const MASCOTE_FELIZ='https://moveisdolar.com.br/colaborador/mascote%20feliz1.png';
const MASCOTE_PREOC='https://moveisdolar.com.br/colaborador/mascote%20preocupado1.png';
const MASCOTE_TRISTE='https://moveisdolar.com.br/colaborador/mascote%20triste1.png';
function mascotByPerc(p){const n=Number(p||0); if(n>=80) return {src:MASCOTE_FELIZ,txt:'Laranjito feliz: meta em ótimo ritmo!'}; if(n>=60) return {src:MASCOTE_PREOC,txt:'Laranjito preocupado: atenção para a meta.'}; return {src:MASCOTE_TRISTE,txt:'Laranjito triste: precisa reagir já!'} }
function renderMascotStatus(p,label=''){const m=mascotByPerc(p); return `<div class="mascot-status"><img src="${m.src}" alt=""><div><strong>${label?label+': ':''}${m.txt}</strong></div></div>`}
function renderDualMascotStatus(ent){const metaCob=calcMeta(ent).geral||0; if(ent?.is_terceiro || ent?.type==='terceiro'){return `<div>${renderMascotStatus(metaCob,'Cobrança terceiro')}</div>`} const vendaRow=(getSalesRows(ent, ent.type==='filial'?'venda_filial_meta':'venda_filial_vendedor_meta')[0])||null; const vendaPerc=salesNum(salesCell(vendaRow,['Atingido Total'])); return `<div>${renderMascotStatus(metaCob,'Cobrança')}${renderMascotStatus(vendaPerc,'Vendas/serviços')}</div>`}
function salesCfgHeader(ent){const c=calcMeta(ent).cfg||{}; return ent.type==='filial' ? `mín. vendas ${Number(c.gerente_vendas_min_pct||0)}% · mín. serviços ${Number(c.gerente_servicos_min_pct||0)}% · rentab libera com mercantil ${Number(c.gerente_rentab_min_mercantil_pct||80)}%` : `mín. vendas ${Number(c.vendas_min_pct||0)}% · mín. serviços ${Number(c.servicos_min_pct||0)}% · rentab libera com mercantil ${Number(c.vendedor_rentab_min_mercantil_pct||80)}%`}
function rentabilidadeAtualPct(ent){return Number(ent?.rentabilidade_pct||0)}
function renderRentabilidadeBadge(ent){const r=rentabilidadeAtualPct(ent); const txt=r?`${r.toFixed(2).replace('.',',')}%`:'Sem dado'; return `<div class="rent-badge"><span>📊 Rentabilidade atual</span><strong>${txt}</strong></div>`}
function renderSalesPanel(ent){const blocks = ent.type==='filial' ? [['venda_filial_meta','📈 Venda · Meta Filial'],['servico_filial_ouro_fob','🛠️ Serviço · Ouro / FOB'],['venda_filial_subgrupo_20k','🚚 Venda · Caminhão 20K / Subgrupo']] : [['venda_filial_vendedor_meta','📈 Venda · Meta Vendedor'],['servico_filial_vendedor_ouro_fob','🛠️ Serviço · Ouro / FOB Vendedor'],['venda_vendedor_subgrupo_20k','🚚 Venda · Caminhão 20K / Subgrupo']]; return `<div class="glass panel sales-panel"><div class="section-head" style="margin:0 0 10px;align-items:flex-start"><div><h3 style="margin:0">💲 Vendas e metas <span class="sales-note">· SGI / mês atual</span></h3><div style="margin-top:8px">${renderRentabilidadeBadge(ent)}</div></div><div class="sales-note" style="text-align:right;max-width:280px">${salesCfgHeader(ent)}</div></div>${renderDualMascotStatus(ent)}<div class="sales-stack">${blocks.map(([k,t])=>renderSalesRows(ent,k,t)).join('')}</div><div style="height:14px"></div>${renderServicosEntidade(ent)}</div>`}

function salesMetricFromRow(row,key){if(!row) return 0; if(key==='venda_realizado') return salesNum(salesCell(row,['Realizado (R$) Total','Realizado(R$) Total'])); if(key==='servico_realizado') return salesNum(salesCell(row,['Realizado (R$) Total','Realizado(R$) Total'])); if(key==='venda_atingido') return salesNum(salesCell(row,['Atingido Total'])); if(key==='servico_atingido') return salesNum(salesCell(row,['Atingido Total'])); return 0}
function bestLiveSalesEntity(entities,key){let best=null; entities.forEach(ent=>{const rows=getSalesRows(ent,key); rows.forEach(r=>{const val=salesMetricFromRow(r,key.includes('servico')?'servico_realizado':'venda_realizado'); if(!best || val>best.val) best={ent,row:r,val};})}); return best}
function defaultVendPolicy(){return [
{faixa1:'0',faixa2:'50999.99',comissao:'0.70',bonus90:'100',bonus100:'250',bonus120:'320',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'51000',faixa2:'70999.99',comissao:'0.74',bonus90:'130',bonus100:'270',bonus120:'350',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'71000',faixa2:'90999.99',comissao:'0.76',bonus90:'160',bonus100:'290',bonus120:'380',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'91000',faixa2:'100999.99',comissao:'0.78',bonus90:'190',bonus100:'330',bonus120:'390',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'101000',faixa2:'120999.99',comissao:'0.80',bonus90:'220',bonus100:'360',bonus120:'420',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'121000',faixa2:'140999.99',comissao:'0.82',bonus90:'250',bonus100:'380',bonus120:'430',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'141000',faixa2:'160999.99',comissao:'0.84',bonus90:'270',bonus100:'400',bonus120:'450',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'161000',faixa2:'180999.99',comissao:'0.86',bonus90:'290',bonus100:'410',bonus120:'470',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'181000',faixa2:'200999.99',comissao:'0.88',bonus90:'300',bonus100:'420',bonus120:'500',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'},
{faixa1:'201000',faixa2:'ACIMA',comissao:'0.90',bonus90:'320',bonus100:'430',bonus120:'550',rent48:'200',rent52:'300',rent55:'400',servico_pct:'12',caminhao_pct:'20'}
]}
function defaultGerPolicy(){return [
{faixa1:'0',faixa2:'80999.99',bonusLoja:'300',comissao:'0.50',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'81000',faixa2:'110999.99',bonusLoja:'400',comissao:'0.50',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'111000',faixa2:'140999.99',bonusLoja:'500',comissao:'0.45',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'141000',faixa2:'170999.99',bonusLoja:'600',comissao:'0.45',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'171000',faixa2:'200999.99',bonusLoja:'700',comissao:'0.45',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'201000',faixa2:'230999.99',bonusLoja:'750',comissao:'0.45',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'231000',faixa2:'260999.99',bonusLoja:'800',comissao:'0.45',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'261000',faixa2:'290999.99',bonusLoja:'900',comissao:'0.45',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'291000',faixa2:'320999.99',bonusLoja:'950',comissao:'0.45',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'},
{faixa1:'321000',faixa2:'ACIMA',bonusLoja:'1000',comissao:'0.45',rent48:'100',rent52:'150',rent55:'300',servico_pct:'6',caminhao_pct:'10'}
]}
function commissionCfg(cfg){return {
vendedor_rentab_min_mercantil_pct:Number(cfg?.vendedor_rentab_min_mercantil_pct||80),
gerente_rentab_min_mercantil_pct:Number(cfg?.gerente_rentab_min_mercantil_pct||80),
vendedor_policy:Array.isArray(cfg?.vendedor_policy)&&cfg.vendedor_policy.length?cfg.vendedor_policy:defaultVendPolicy(),
gerente_policy:Array.isArray(cfg?.gerente_policy)&&cfg.gerente_policy.length?cfg.gerente_policy:defaultGerPolicy(),
vendedor_policy_headers:Array.isArray(cfg?.vendedor_policy_headers)&&cfg.vendedor_policy_headers.length?cfg.vendedor_policy_headers:['De','Até','% Comissão','Bônus 90%','Bônus 100%','Bônus 120%','Rentab 48%','Rentab 52,15%','Rentab 55,50%','% Serviços','% Caminhão'],
gerente_policy_headers:Array.isArray(cfg?.gerente_policy_headers)&&cfg.gerente_policy_headers.length?cfg.gerente_policy_headers:['De','Até','Classificação Loja Bônus','% Comissão','Rentab 48%','Rentab 52,15%','Rentab 55,50%','% Serviços','% Caminhão'],
camp_meta_diaria_vend:Array.isArray(cfg?.camp_meta_diaria_vend)&&cfg.camp_meta_diaria_vend.length?cfg.camp_meta_diaria_vend:defaultCampMetaDiariaVend(),
camp_meta_diaria_ger:Array.isArray(cfg?.camp_meta_diaria_ger)&&cfg.camp_meta_diaria_ger.length?cfg.camp_meta_diaria_ger:defaultCampMetaDiariaGer(),
camp_dindin_vend:Array.isArray(cfg?.camp_dindin_vend)&&cfg.camp_dindin_vend.length?cfg.camp_dindin_vend:defaultCampDindinVend(),
camp_dindin_ger:Array.isArray(cfg?.camp_dindin_ger)&&cfg.camp_dindin_ger.length?cfg.camp_dindin_ger:defaultCampDindinGer(),
camp_admin:Array.isArray(cfg?.camp_admin)&&cfg.camp_admin.length?cfg.camp_admin:defaultCampAdmin(),
camp_cobranca_terceiro:Array.isArray(cfg?.camp_cobranca_terceiro)&&cfg.camp_cobranca_terceiro.length?cfg.camp_cobranca_terceiro:defaultCampTerceiro(),
camp_cob_crediarista:Array.isArray(cfg?.camp_cob_crediarista)&&cfg.camp_cob_crediarista.length?cfg.camp_cob_crediarista:defaultCampCrediarista()
}}
function renderPolicyTable(id, rows, cols, headers){return `<div class="comm-scroll"><table class="comm-table"><thead><tr>${cols.map((c,i)=>`<th><input data-comm-head="${id}" data-index="${i}" value="${esc(headers?.[i]??c.label)}"></th>`).join('')}</tr></thead><tbody>${rows.map((r,i)=>`<tr>${cols.map(c=>`<td><input data-comm="${id}" data-row="${i}" data-key="${c.key}" value="${esc(r[c.key]??'')}"></td>`).join('')}</tr>`).join('')}</tbody></table></div>`}

function defaultCampMetaDiariaVend(){return [{dias_uteis:'25',bonus_final:'500'}]}
function defaultCampMetaDiariaGer(){return [{dias_uteis:'25',bonus_final:'1000'}]}
function defaultCampDindinVend(){return [{atingido:'105',extra_pct:'0.50'},{atingido:'110',extra_pct:'1.00'}]}
function defaultCampDindinGer(){return [{atingido:'105',extra_pct:'0.50'},{atingido:'110',extra_pct:'1.00'}]}
function defaultCampAdmin(){return [{atingido:'100',extra_pct:'0.15',colaboradores:'5'},{atingido:'105',extra_pct:'0.20',colaboradores:'5'},{atingido:'110',extra_pct:'0.22',colaboradores:'5'}]}
function defaultCampCrediarista(){return [{faixa:'atencao',pct:'1.00'},{faixa:'alerta',pct:'2.00'},{faixa:'grave',pct:'3.00'}]}
function renderCommissionPanel(cfg){const pc=commissionCfg(cfg);
const vendCols=[{key:'faixa1',label:'De'},{key:'faixa2',label:'Até'},{key:'comissao',label:'% Comissão'},{key:'bonus90',label:'Bônus 90%'},{key:'bonus100',label:'Bônus 100%'},{key:'bonus120',label:'Bônus 120%'},{key:'rent48',label:'Rentab 48%'},{key:'rent52',label:'Rentab 52,15%'},{key:'rent55',label:'Rentab 55,50%'},{key:'servico_pct',label:'% Serviços'},{key:'caminhao_pct',label:'% Caminhão'}];
const gerCols=[{key:'faixa1',label:'De'},{key:'faixa2',label:'Até'},{key:'bonusLoja',label:'Classificação Loja'},{key:'comissao',label:'% Comissão'},{key:'rent48',label:'Rentab 48%'},{key:'rent52',label:'Rentab 52,15%'},{key:'rent55',label:'Rentab 55,50%'},{key:'servico_pct',label:'% Serviços'},{key:'caminhao_pct',label:'% Caminhão'}];
const cmdCols=[{key:'dias_uteis',label:'Dias úteis'},{key:'bonus_final',label:'Bônus final'}];
const cdiCols=[{key:'atingido',label:'Atingiu %'},{key:'extra_pct',label:'Extra % sobre mercantil'}];
const admCols=[{key:'atingido',label:'Atingiu % total geral'},{key:'extra_pct',label:'Extra % sobre mercantil lojas'},{key:'colaboradores',label:'Nº colaboradores'}];
const credCols=[{key:'faixa',label:'Faixa'},{key:'pct',label:'% Comissão'}];
const box=document.getElementById('commissionPanel'); if(!box) return;
box.innerHTML=`<div class="comm-wrap">
<div class="section-head" style="margin-top:14px"><div><h2 style="font-size:18px">💰 Política de comissão</h2><div class="hint">Tabela global/individual para vendedores e gerente/filial.</div></div></div>
<div class="comm-box"><div class="comm-subtitle">🧑‍💼 Política do vendedor</div>${renderPolicyTable('vend',pc.vendedor_policy,vendCols,pc.vendedor_policy_headers)}</div>
<div class="comm-box"><div class="comm-subtitle">🏬 Política do gerente / filial</div>${renderPolicyTable('ger',pc.gerente_policy,gerCols,pc.gerente_policy_headers)}</div>
<div class="comm-box"><div class="comm-subtitle">🎯 CAMPANHA META DIÁRIA · vendedor</div><div class="hint">Meta geral de vendas dividida pelos dias úteis. Cada dia batido pontua 1 ponto no mês.</div>${renderPolicyTable('cmdv',pc.camp_meta_diaria_vend,cmdCols,['Dias úteis','Bônus final'])}</div>
<div class="comm-box"><div class="comm-subtitle">🎯 CAMPANHA META DIÁRIA · gerente / filial</div><div class="hint">A loja/filial que mais bater a meta diária soma pontos e recebe o bônus configurado.</div>${renderPolicyTable('cmdg',pc.camp_meta_diaria_ger,cmdCols,['Dias úteis','Bônus final'])}</div>
<div class="comm-box"><div class="comm-subtitle">💸 CAMPANHA DINDIN NO BOLSO · vendedor</div><div class="hint">Válida só no mercantil. Desbloqueia apenas se recebimentos estiverem acima de 75%.</div>${renderPolicyTable('cdiv',pc.camp_dindin_vend,cdiCols,['Atingiu %','Extra % mercantil'])}</div>
<div class="comm-box"><div class="comm-subtitle">💸 CAMPANHA DINDIN NO BOLSO · gerente / filial</div><div class="hint">Mesmo conceito para filial/gerente, somente sobre mercantil líquido.</div>${renderPolicyTable('cdig',pc.camp_dindin_ger,cdiCols,['Atingiu %','Extra % mercantil'])}</div>
<div class="comm-box"><div class="comm-subtitle">🏢 CAMPANHA DINDIN NO BOLSO · administrativo</div><div class="hint">Tabela apenas para consulta manual. O cálculo do administrativo será feito manualmente depois.</div>${renderPolicyTable('cadm',pc.camp_admin,admCols,['Atingiu % total','Extra % lojas','Nº colaboradores'])}</div>
<div class="comm-box"><div class="comm-subtitle">🤝 COMISSÃO COBRANÇA TERCEIRO · Cobrança10</div><div class="hint">Comissão por faixa somente para clientes realmente cobrados e recebidos no mês.</div>${renderPolicyTable('cter',pc.camp_cobranca_terceiro,credCols,['Faixa','% Comissão'])}</div>
<div class="comm-box"><div class="comm-subtitle">🧾 COMISSÃO CREDIARISTAS · filiais</div><div class="hint">Configuração padrão para os usuários crediaristas por faixa.</div>${renderPolicyTable('ccred',pc.camp_cob_crediarista,credCols,['Faixa','% Comissão'])}</div>
</div>`}
function readCommissionPanel(){const vendRows=[], gerRows=[]; const vendHeaders=[], gerHeaders=[]; const cmdv=[], cmdg=[], cdiv=[], cdig=[], cadm=[], cter=[], ccred=[];
document.querySelectorAll('[data-comm="vend"]').forEach(el=>{const i=Number(el.dataset.row||0); vendRows[i]=vendRows[i]||{}; vendRows[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm="ger"]').forEach(el=>{const i=Number(el.dataset.row||0); gerRows[i]=gerRows[i]||{}; gerRows[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm="cmdv"]').forEach(el=>{const i=Number(el.dataset.row||0); cmdv[i]=cmdv[i]||{}; cmdv[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm="cmdg"]').forEach(el=>{const i=Number(el.dataset.row||0); cmdg[i]=cmdg[i]||{}; cmdg[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm="cdiv"]').forEach(el=>{const i=Number(el.dataset.row||0); cdiv[i]=cdiv[i]||{}; cdiv[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm="cdig"]').forEach(el=>{const i=Number(el.dataset.row||0); cdig[i]=cdig[i]||{}; cdig[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm="cadm"]').forEach(el=>{const i=Number(el.dataset.row||0); cadm[i]=cadm[i]||{}; cadm[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm="cter"]').forEach(el=>{const i=Number(el.dataset.row||0); cter[i]=cter[i]||{}; cter[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm="ccred"]').forEach(el=>{const i=Number(el.dataset.row||0); ccred[i]=ccred[i]||{}; ccred[i][el.dataset.key]=el.value});
document.querySelectorAll('[data-comm-head="vend"]').forEach(el=>{vendHeaders[Number(el.dataset.index||0)] = el.value});
document.querySelectorAll('[data-comm-head="ger"]').forEach(el=>{gerHeaders[Number(el.dataset.index||0)] = el.value});
return {
  vendedor_policy:vendRows.filter(Boolean),
  gerente_policy:gerRows.filter(Boolean),
  vendedor_policy_headers:vendHeaders.filter(v=>v!=null),
  gerente_policy_headers:gerHeaders.filter(v=>v!=null),
  camp_meta_diaria_vend:cmdv.filter(Boolean),
  camp_meta_diaria_ger:cmdg.filter(Boolean),
  camp_dindin_vend:cdiv.filter(Boolean),
  camp_dindin_ger:cdig.filter(Boolean),
  camp_admin:cadm.filter(Boolean),
  camp_cobranca_terceiro:cter.filter(Boolean),
  camp_cob_crediarista:ccred.filter(Boolean)
}}

function moneyNum(s){return parseFloat(String(s||'').replace(/\./g,'').replace(',','.'))||0}
function faixaMatchRealizado(rows, realizado){const n=Number(realizado||0); for(const r of (rows||[])){const de=parseFloat(String(r.faixa1||'0').replace(',','.'))||0; const ateRaw=String(r.faixa2||'').toUpperCase(); if(ateRaw.includes('ACIMA')){ if(n>=de) return r; } else { const ate=parseFloat(String(r.faixa2||'0').replace(',','.'))||0; if(n>=de && n<=ate) return r; }} return (rows||[])[(rows||[]).length-1]||null}
function calcCommissionSummary(ent){
  const cfg=entityConfig(ent);
  const cc=commissionCfg(cfg);
  const vendaRow=(getSalesRows(ent, ent.type==='filial'?'venda_filial_meta':'venda_filial_vendedor_meta')[0])||null;
  const servRow=(getSalesRows(ent, ent.type==='filial'?'servico_filial_ouro_fob':'servico_filial_vendedor_ouro_fob')[0])||null;
  const camRow=(ent.type==='vendedor'?(getSalesRows(ent,'venda_vendedor_subgrupo_20k')[0]||null):(getSalesRows(ent,'venda_filial_subgrupo_20k')[0]||null));
  const vendaRealBruto=moneyNum(salesCell(vendaRow,['Realizado (R$) Total','Realizado(R$) Total']));
  const vendaPerc=salesNum(salesCell(vendaRow,['Atingido Total']));
  const servReal=moneyNum(salesCell(servRow,['Realizado (R$) Total','Realizado(R$) Total']));
  const camReal=moneyNum(salesCell(camRow,['Realizado (R$) Total','Realizado(R$) Total']));
  const vendaReal=Math.max(0,vendaRealBruto-camReal);
  const policy=ent.type==='filial'?cc.gerente_policy:cc.vendedor_policy;
  const faixa=faixaMatchRealizado(policy,vendaReal) || {};
  const comPerc=Number(faixa.comissao||0);
  const servPct=Number(faixa.servico_pct||0);
  const camPct=Number(faixa.caminhao_pct||0);
  const vendasComissao=(vendaReal*comPerc/100);
  let bonusMeta=0;
  let bonusLiberado=false;
  const minVenda = ent.type==='filial' ? Number(cfg.gerente_vendas_min_pct||80) : Number(cfg.vendas_min_pct||80);
  const minServico = ent.type==='filial' ? Number(cfg.gerente_servicos_min_pct||80) : Number(cfg.servicos_min_pct||80);
  const rentMinMercantil = ent.type==='filial' ? Number(cfg.gerente_rentab_min_mercantil_pct||80) : Number(cfg.vendedor_rentab_min_mercantil_pct||80);
  const rentMin50 = 50;
  const geralMeta = calcMeta(ent).geral||0;
  if(ent.type==='filial'){
    if(vendaPerc>=minVenda){ bonusMeta=Number(faixa.bonusLoja||0); bonusLiberado=bonusMeta>0; }
  } else {
    if(vendaPerc>=120){ bonusMeta=Number(faixa.bonus120||0); bonusLiberado=bonusMeta>0; }
    else if(vendaPerc>=100){ bonusMeta=Number(faixa.bonus100||0); bonusLiberado=bonusMeta>0; }
    else if(vendaPerc>=90){ bonusMeta=Number(faixa.bonus90||0); bonusLiberado=bonusMeta>0; }
  }
  const elegivelMercantil=vendaPerc>=minVenda;
  const elegivelServicos=vendaPerc>=minServico;
  const servicosComissao=elegivelServicos?(servReal*servPct/100):0;
  const caminhaoComissao=elegivelServicos?(camReal*camPct/100):0;

  const rentAtual = Number(ent?.rentabilidade_pct||0);
  const rentUnlocked = (geralMeta>=rentMin50) && (vendaPerc>=rentMinMercantil);
  const rentThresholds = [
    {label:'48%', key:'rent48', min:48.0, valor:Number(faixa.rent48||0)},
    {label:'52,15%', key:'rent52', min:52.15, valor:Number(faixa.rent52||0)},
    {label:'55,50%', key:'rent55', min:55.50, valor:Number(faixa.rent55||0)}
  ];
  let rentPremio=0, rentFaixaTxt='Sem dado', rentAppliedKey='';
  if(rentUnlocked){
    rentThresholds.forEach(rt=>{
      if(rentAtual>=rt.min){
        rentPremio=rt.valor||0;
        rentFaixaTxt=rt.label;
        rentAppliedKey=rt.key;
      }
    });
  }
  const rent48=(rentUnlocked && rentAtual>=48.0)?Number(faixa.rent48||0):0;
  const rent52=(rentUnlocked && rentAtual>=52.15)?Number(faixa.rent52||0):0;
  const rent55=(rentUnlocked && rentAtual>=55.50)?Number(faixa.rent55||0):0;
  const metaAtingida=vendaPerc>=minVenda || geralMeta>=rentMin50;

  return {
    vendaRealBruto,vendaReal,vendaPerc,servReal,camReal,comPerc,servPct,camPct,
    vendasComissao,servicosComissao,caminhaoComissao,bonusMeta,bonusLiberado,
    elegivelMercantil,elegivelServicos,rentUnlocked,rent48,rent52,rent55,
    rentAtual,rentPremio,rentFaixaTxt,rentAppliedKey,
    totalPrevisto:(vendasComissao+servicosComissao+caminhaoComissao+bonusMeta+rentPremio),
    faixaTxt:`${faixa.faixa1||'-'} até ${faixa.faixa2||'-'}`,
    metaAtingida,minVenda,minServico,rentMinMercantil,geralMeta
  }
}

function rowFirst(arr){return Array.isArray(arr)&&arr.length?arr[0]:{}}
function campaignMetaDailyData(ent){
  const cfg = commissionCfg(entityConfig(ent));
  const vendaKey = ent.type === 'filial' ? 'venda_filial_meta' : 'venda_filial_vendedor_meta';
  const vendaRow = (getSalesRows(ent, vendaKey)[0]) || null;

  const metaDiaRow = ent.type === 'filial'
    ? rowFirst(cfg.camp_meta_diaria_ger)
    : rowFirst(cfg.camp_meta_diaria_vend);

  const diasUteis = Math.max(1, Number(metaDiaRow.dias_uteis || 25));

  // CORRETO: meta diária usa META TOTAL, não Meta Período.
  const metaTotal = moneyNum(salesCell(vendaRow, [
    'Meta (R$) Total',
    'Meta(R$) Total'
  ]));

  // CORRETO: pontos usam REALIZADO TOTAL acumulado do mês.
  const realizadoTotal = moneyNum(salesCell(vendaRow, [
    'Realizado (R$) Total',
    'Realizado(R$) Total'
  ]));

  const metaDiaria = metaTotal > 0 ? (metaTotal / diasUteis) : 0;

  // REGRA DEFINITIVA: ponto é diário real, salvo no histórico.
  // Não pode pegar realizado acumulado e dividir pela meta diária.
  const mesAtual = new Date().toISOString().slice(0, 7);
  const mdMes = HIST_DASH?.daily_meta?.months?.[mesAtual] || {filiais:{}, vendedores:{}};
  let histObj = null;
  if (ent.type === 'filial') {
    histObj = mdMes?.filiais?.[ent.filial] || null;
  } else {
    const alvoNome = normSalesText(ent.nome);
    const alvoFilial = String(ent.filial || '').toUpperCase();
    const candidatos = Object.entries(mdMes?.vendedores || {});
    for (const [k, v] of candidatos) {
      const nomeOk = normSalesText(v?.nome || k).includes(alvoNome) || alvoNome.includes(normSalesText(v?.nome || k));
      const filialOk = String(v?.filial || '').toUpperCase() === alvoFilial || String(k || '').toUpperCase().includes(alvoFilial);
      if (nomeOk && filialOk) { histObj = v; break; }
    }
  }
  const pontosMetaDiaria = Math.min(diasUteis, Number(histObj?.pontos || 0));
  const hojeKey = new Date().toISOString().slice(0, 10);
  const realizadoDia = Number(histObj?.dias?.[hojeKey]?.realizado_dia || 0);
  const pontoHoje = Number(histObj?.dias?.[hojeKey]?.ponto || 0);

  return {
    cfg,
    vendaRow,
    metaDiaRow,
    diasUteis,
    metaTotal,
    realizadoTotal,
    metaDiaria,
    realizadoDia,
    pontoHoje,
    pontosMetaDiaria
  };
}

function campanhaRankingEntities(ent){
  if (ent.type === 'filial') {
    return Object.entries(FILIAIS || {}).map(([k, v]) => ({
      type: 'filial',
      filial: k,
      nome: v.nome || k,
      ...(v || {})
    }));
  }

  return Object.entries(TODOS || {}).flatMap(([filial, arr]) =>
    (arr || []).map(v => ({
      type: 'vendedor',
      filial,
      nome: v.nome,
      ...(v || {})
    }))
  );
}

function isLiderMetaDiaria(ent){
  const atual = campaignMetaDailyData(ent);

  if (!atual.metaDiaria || atual.metaDiaria <= 0) return false;

  const ranking = campanhaRankingEntities(ent)
    .map(e => {
      const d = campaignMetaDailyData(e);
      return {
        ent: e,
        pontos: Number(d.pontosMetaDiaria || 0),
        realizadoTotal: Number(d.realizadoTotal || 0),
        metaDiaria: Number(d.metaDiaria || 0)
      };
    })
    .filter(x => x.metaDiaria > 0)
    .sort((a, b) => {
      if (b.pontos !== a.pontos) return b.pontos - a.pontos;
      return b.realizadoTotal - a.realizadoTotal;
    });

  if (!ranking.length) return false;

  const lider = ranking[0].ent;

  if (ent.type === 'filial') {
    return String(lider.filial || '').toUpperCase() === String(ent.filial || '').toUpperCase();
  }

  return (
    String(lider.nome || '').toUpperCase() === String(ent.nome || '').toUpperCase() &&
    String(lider.filial || '').toUpperCase() === String(ent.filial || '').toUpperCase()
  );
}

function calcCampaignSummary(ent){
  const base = campaignMetaDailyData(ent);
  const cfg = base.cfg;
  const sale = calcCommissionSummary(ent);
  const meta = calcMeta(ent);

  const dindinRows = ent.type === 'filial'
    ? cfg.camp_dindin_ger
    : cfg.camp_dindin_vend;

  let dindinExtraPct = 0;

  (dindinRows || []).forEach(r => {
    if (sale.vendaPerc >= Number(r.atingido || 0) && meta.geral >= 75) {
      dindinExtraPct = Math.max(dindinExtraPct, Number(r.extra_pct || 0));
    }
  });

  const dindinExtra = sale.vendaReal * (dindinExtraPct / 100);

  return {
    metaDiaria: base.metaDiaria,
    pontosMetaDiaria: base.pontosMetaDiaria,
    diasUteis: base.diasUteis,
    metaTotal: base.metaTotal,
    realizadoTotal: base.realizadoTotal,
    realizadoDia: base.realizadoDia,
    pontoHoje: base.pontoHoje,
    bonusMetaDiaria: String(base.metaDiaRow.bonus_final || '0'),
    dindinExtraPct,
    dindinExtra,
    dindinLiberado: dindinExtraPct > 0,
    liderMetaDiaria: isLiderMetaDiaria(ent)
  };
}

function renderCampaignSummary(ent){
  const c = calcCampaignSummary(ent);

  // REGRA NOVA: se não estiver em 1º lugar, oculta a campanha Meta Diária.
  if (!c.liderMetaDiaria) return '';

  const item = (k, v, locked = false) =>
    `<div class="campaign-item ${locked ? 'locked' : ''}">
      <div class="k">${k}</div>
      <div class="v">${v}</div>
    </div>`;

  return `<div class="glass panel campaign-card">
    <h3>🎯 Campanhas ativas <span class="note">· acompanhamento do mês</span></h3>

    <div class="campaign-grid">
      ${item('Meta diária', R(c.metaDiaria || 0), false)}
      ${item('Pontos no mês', `${c.pontosMetaDiaria || 0} / ${c.diasUteis || 26}`, false)}
      ${item('Bônus meta diária', String(c.bonusMetaDiaria || '0'), false)}
      ${item('Dindin no bolso %', String((c.dindinExtraPct || 0).toFixed(2)).replace('.', ',') + '%', !c.dindinLiberado)}
      ${item('Dindin no bolso valor', R(c.dindinExtra || 0), !c.dindinLiberado)}
    </div>

    <div class="commission-note">
      Campanha Meta Diária aparece somente para o 1º colocado.
      Cálculo: Meta Total ${R(c.metaTotal || 0)} ÷ ${c.diasUteis || 26} dias úteis =
      ${R(c.metaDiaria || 0)} por dia. Pontos: só soma quando a venda DO DIA bate a meta diária.
      Hoje: ${R(c.realizadoDia || 0)} ${c.pontoHoje ? '✅ ponto ganho' : '⏳ sem ponto'} .
    </div>
  </div>`;
}

function renderCommissionSummary(ent){if(!canVerComissionamento()) return '';const c=calcCommissionSummary(ent); const totalLiberado = c.elegivelMercantil && c.elegivelServicos; const totalExibido = totalLiberado ? c.totalPrevisto : 0; const moneyCell=(title,val,locked=false,extra='')=>`<div class="commission-item ${locked?'locked':''} ${!locked?'unlocked':''} ${extra}"><div class="k">${title}</div><div class="v">${R(val||0)}</div></div>`; const pctCell=(title,val,locked=false)=>`<div class="commission-item ${locked?'locked':''} ${!locked?'unlocked':''}"><div class="k">${title}</div><div class="v">${String(Number(val||0).toFixed(2)).replace('.',',')}%</div></div>`; return `<div class="glass panel commission-card"><h3>💵 Comissionamento previsto <span class="note">· calculado pela política salva</span></h3>${c.metaAtingida?`<div class="meta-hit-banner"><img src="${LARANJITO}" alt=""><span>Meta liberada! O Laranjito está comemorando sua liberação de comissão/bonus.</span></div>`:''}<div class="commission-grid">${`<div class="commission-item unlocked"><div class="k">Faixa aplicada</div><div class="v" style="font-size:16px">${esc(c.faixaTxt)}</div></div>`}${pctCell('% comissão mercantil',c.comPerc,!c.elegivelMercantil)}${pctCell('% serviços',c.servPct,!c.elegivelServicos)}${pctCell('% caminhão',c.camPct,!c.elegivelServicos)}${moneyCell('Comissão vendas',c.vendasComissao,!c.elegivelMercantil)}${moneyCell('Comissão serviços',c.servicosComissao,!c.elegivelServicos)}${moneyCell('Comissão caminhão',c.caminhaoComissao,!c.elegivelServicos)}${moneyCell('Bônus por meta',c.bonusMeta,!c.bonusLiberado)}${moneyCell('Rentab 48%',c.rent48,!c.rentUnlocked)}${moneyCell('Rentab 52,15%',c.rent52,!c.rentUnlocked)}${moneyCell('Rentab 55,50%',c.rent55,!c.rentUnlocked)}${moneyCell('Total previsto',totalExibido,!totalLiberado,'total-final '+(!totalLiberado?'total-locked':''))}</div><div class="commission-note">Base mercantil bruta: ${R(c.vendaRealBruto||0)} · Caminhão abatido: ${R(c.camReal||0)} · Mercantil líquido para comissão: ${R(c.vendaReal||0)} · Serviço: ${R(c.servReal||0)}. Mínimo vendas ${pct(c.minVenda)} · mínimo serviços/caminhão ${pct(c.minServico)} · rentab exige cobrança 50% + mercantil ${pct(c.rentMinMercantil)}.</div></div>`}
function backToMain(){currentDetailRef=null; try{renderLaranjitoNotify()}catch(e){}; detailScreen.classList.add('hidden');document.getElementById('mainScreen').classList.remove('hidden')}
function renderMetaBox(title,color,obj){return `<div class="meta-card"><div class="meta-title">${title}</div><div class="meta-main" style="color:${color}">${pct(obj.perc||0)}</div><div class="meta-sub">Alvo: ${R(obj.alvo||0)}</div><div class="meta-sub">Recebido: ${R(obj.rec||0)}</div></div>`}
function renderBonusBox(cfg,geral){const achieved=(geral>=100&&cfg.bonus_100)?100:(geral>=85&&cfg.bonus_85)?85:(geral>=75&&cfg.bonus_75)?75:(geral>=50&&cfg.bonus_50)?50:0; const items=[[50,cfg.bonus_50||'-'],[75,cfg.bonus_75||'-'],[85,cfg.bonus_85||'-'],[100,cfg.bonus_100||'-']];return `<div class="bonus-box"><h4>Faixas configuradas</h4><div class="bonus-list">${items.map(([p,t])=>`<div class="bonus-item ${achieved===p?'active':''}" style="${achieved===p?'box-shadow:0 0 0 2px rgba(59,130,246,.18),0 0 26px rgba(59,130,246,.2);animation:liquid 1.6s ease-in-out infinite alternate':''}"><div class="left"><span>🎯</span><span>${p}%</span></div><div style="display:flex;align-items:center;gap:10px">${achieved===p?`<img src="${LARANJITO}" alt="laranjito" style="width:34px;height:34px;border-radius:10px;object-fit:cover">`:''}<span>${esc(t)}</span></div></div>`).join('')}</div></div>`}
function renderSingleBars(ent,meta,big=false){const vals=[['Grave',meta.grave.pend,'var(--red)'],['Alerta',meta.alerta.pend,'var(--orange)'],['Atenção',meta.atencao.pend,'var(--yellow)'],['Recebido',Number(ent.pago||0),'var(--green)']]; const max=Math.max(1,...vals.map(v=>v[1])); return `<div class="glass big-chart-card" style="margin-top:0"><div class="legend-inline"><span><i class="dot" style="background:var(--red)"></i>Grave</span><span><i class="dot" style="background:var(--orange)"></i>Alerta</span><span><i class="dot" style="background:var(--yellow)"></i>Atenção</span><span><i class="dot" style="background:var(--green)"></i>Recebido</span></div><div class="groupbars" style="justify-content:center"><div class="group" style="min-width:100%"><div class="bars" style="height:${big?320:260}px;justify-content:space-around;border-left:none">${vals.map(v=>`<div style="text-align:center"><div class="bar" style="width:${big?72:54}px;height:${Math.max(18,(v[1]/max)*(big?250:200))}px;background:linear-gradient(180deg,rgba(255,255,255,.18),transparent),linear-gradient(180deg,${v[2]},${v[2]})"><span class="wave one"></span><span class="wave two"></span><span class="bubble b1"></span><span class="bubble b2"></span><span class="bubble b3"></span></div><div class="glabel" style="margin-top:10px">${v[0]}<br><strong>${R(v[1])}</strong></div></div>`).join('')}</div></div></div></div>`}
function recebKeyVend(ent){return `${ent.nome}_${ent.filial}`}

function normConc(s){return String(s||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toUpperCase().replace(/[^A-Z0-9]+/g,' ').trim()}
function parseDataBRjs(s){
  const m=String(s||'').match(/(\d{2})\/(\d{2})\/(\d{4})/);
  if(!m) return null;
  return new Date(Number(m[3]), Number(m[2])-1, Number(m[1]));
}
function similarCliente(a,b){
  const x=normConc(a), y=normConc(b);
  if(!x || !y) return false;
  return x===y || x.includes(y.slice(0,18)) || y.includes(x.slice(0,18));
}
function getQuitadosConciliados(){
  const out={};
  const logs=(COB_LOGS||[]).slice().filter(l=>String(l.acao||'')==='whatsapp' || l.telefone || l.titulo || l.parcela);
  const byTitulo={};
  logs.forEach(l=>{
    const k=`${String(l.titulo||'').trim()}|${String(l.parcela||'').trim()}`;
    if(!byTitulo[k]) byTitulo[k]=[];
    byTitulo[k].push(l);
  });

  (QUITADOS_180||[]).forEach(q=>{
    const k=`${String(q.titulo||'').trim()}|${String(q.parcela||'').trim()}`;
    const cand=(byTitulo[k]||[]).filter(l=>similarCliente(l.cliente,q.cliente));
    if(!cand.length) return;

    const pagto=parseDataBRjs(q.pagamento);
    if(!pagto) return;
    const mesPagto = dateOnlyISO(q.pagamento||'').slice(0,7);
    const mesAtualReceb = (typeof mesAtualComissao==='function' ? mesAtualComissao() : new Date().toISOString().slice(0,7));
    if(mesPagto !== mesAtualReceb) return;

    const antes=cand.filter(l=>{
      const raw=String(l.server_time||l.criado_em||l.data||'');
      const d=raw ? new Date(raw.replace(' ', 'T')) : null;
      if(!d || isNaN(d.getTime())) return true;
      const diffDias=(pagto-d)/(1000*60*60*24);
      return diffDias>=0 && diffDias<=180;
    });
    if(!antes.length) return;

    antes.sort((a,b)=>String(a.server_time||'').localeCompare(String(b.server_time||'')));
    const l=antes[antes.length-1];

    const usuarioKey=normConc(l.usuario||l.destino_nome||'');
    const destinoKey=normConc(l.destino_nome||'');
    const filialKey=String(l.filial||q.filial||'').toUpperCase();
    const faixa=q.faixa||'grave';
    const rec={
      cliente:q.cliente,
      dias:q.dias_atraso_pagamento,
      pago:Number(q.pago||0),
      vencimento:q.vencimento,
      pagamento:q.pagamento,
      parcela:q.parcela,
      titulo:q.titulo,
      vendedor:l.usuario||l.destino_nome||'',
      origem:'conciliado_quitados_180d',
      cobrador:l.usuario||'',
      destino_nome:l.destino_nome||'',
      destino_tipo:l.destino_tipo||'',
      filial:filialKey
    };

    [usuarioKey,destinoKey,normConc(`${l.destino_nome||''}_${filialKey}`),normConc(`${l.usuario||''}_${filialKey}`),normConc(filialKey)].forEach(key=>{
      if(!key) return;
      if(!out[key]) out[key]={grave:[],alerta:[],atencao:[]};
      if(!out[key][faixa]) out[key][faixa]=[];
      // evita duplicar no mesmo key
      const exists=out[key][faixa].some(x=>String(x.titulo)===String(rec.titulo)&&String(x.parcela)===String(rec.parcela)&&similarCliente(x.cliente,rec.cliente));
      if(!exists) out[key][faixa].push(rec);
    });
  });

  return out;
}
function mergeRecebimentosConciliados(base, ent){
  const result={
    grave:[...((base||{}).grave||[])],
    alerta:[...((base||{}).alerta||[])],
    atencao:[...((base||{}).atencao||[])]
  };
  const keys=[
    normConc(ent?.nome),
    normConc(ent?.login),
    normConc(ent?.filial),
    normConc(`${ent?.nome||''}_${ent?.filial||''}`),
    normConc(`${ent?.login||''}_${ent?.filial||''}`)
  ].filter(Boolean);
  keys.forEach(k=>{
    const extra=RECEBIMENTOS_CONCILIADOS[k];
    if(!extra) return;
    ['grave','alerta','atencao'].forEach(fx=>{
      (extra[fx]||[]).forEach(r=>{
        const exists=result[fx].some(x=>String(x.titulo)===String(r.titulo)&&String(x.parcela)===String(r.parcela)&&similarCliente(x.cliente,r.cliente));
        if(!exists) result[fx].push(r);
      });
    });
  });
  ['grave','alerta','atencao'].forEach(fx=>result[fx].sort((a,b)=>Number(b.pago||0)-Number(a.pago||0)));
  return result;
}

function getRecebimentos(ent){
  let base;
  if(ent.type==='terceiro' || ent.is_terceiro) base=RECEBIMENTOS_TERCEIRO||{grave:[],alerta:[],atencao:[]};
  else if(ent.type==='crediarista' || ent.is_crediarista){
    // Crediarista: somente recebimentos conciliados com cobrança feita pelo próprio usuário antes do pagamento.
    base={grave:[],alerta:[],atencao:[]};
  } else if(ent.type==='vendedor') base=RECEBIMENTOS[recebKeyVend(ent)]||{grave:[],alerta:[],atencao:[]};
  else {
    base={grave:[],alerta:[],atencao:[]};
    Object.entries(RECEBIMENTOS||{}).forEach(([k,v])=>{if(k.endsWith('_'+ent.filial)){['grave','alerta','atencao'].forEach(fx=>base[fx].push(...(v[fx]||[])))}}); 
  }
  base=mergeRecebimentosConciliados(base, ent||{});
  ['grave','alerta','atencao'].forEach(fx=>base[fx].sort((a,b)=>Number(b.pago||0)-Number(a.pago||0)));
  return base;
}
function renderRecebimentos(ent){const src=getRecebimentos(ent); let out=''; ['grave','alerta','atencao'].forEach(fx=>{const arr=src[fx]||[]; const label=fx==='grave'?'Grave':fx==='alerta'?'Alerta':'Atenção'; if(!arr.length){out+=`<div class="faixa-block"><div class="faixa-title ${fx}">${label}<span>Sem recebimentos</span></div></div>`; return} out+=`<div class="faixa-block"><div class="faixa-title ${fx}">${label}<span>${arr.length} títulos · ${R(arr.reduce((a,b)=>a+Number(b.pago||0),0))}</span></div><div class="tableish">${arr.slice(0,80).map(r=>`<div class="row-item"><div class="row-top"><div><div class="name">${esc(r.cliente||r.nome||'')}</div><div class="small muted">Título ${esc(r.titulo||'')} · Parcela ${esc(r.parcela||'')}</div></div><div><strong>${esc(r.pagamento||'')}</strong><div class="small muted">Pagamento</div></div><div><strong>${esc(r.vencimento||'')}</strong><div class="small muted">Vencimento</div></div><div><strong>${r.dias||0}d</strong><div class="small muted">Dias</div></div><div><strong>${R(r.pago||0)}</strong><div class="small muted">Recebido</div></div><div><strong>${esc(r.vendedor||'')}</strong><div class="small muted">Origem</div></div></div></div>`).join('')}</div></div>`}); return out}
function cobrancaRowKey(r){return [String(r.cliente||r.nome||'').trim().toUpperCase(),String(r.titulo||'').trim(),String(r.parcela||'').trim(),String(r.vencimento||'').trim()].join('|')}
function dedupeCobrancaBuckets(src){const out={grave:[],alerta:[],atencao:[]}; const seen=new Set(); ['grave','alerta','atencao'].forEach(fx=>{(src?.[fx]||[]).forEach(r=>{const k=cobrancaRowKey(r); if(!seen.has(k)){seen.add(k); out[fx].push(r);}})}); return out}
function vendorAssignedKeysByFilial(filial){const keys=new Set(); const arr=(TODOS?.[filial]||[]); arr.forEach(v=>{const buckets=CLIENTES_VEND?.[v.nome]||{grave:[],alerta:[],atencao:[]}; ['grave','alerta','atencao'].forEach(fx=>(buckets[fx]||[]).forEach(r=>keys.add(cobrancaRowKey(r))))}); return keys}
function getClientesEnt(ent){
  if(ent.type==='terceiro' || ent.is_terceiro) return dedupeCobrancaBuckets(CLIENTES_TERCEIRO||{grave:[],alerta:[],atencao:[]});
  if(ent.type==='crediarista' || ent.is_crediarista){ const filialKey=String(ent.filial||'').toUpperCase(); const loginKey=String(ent.login||crediaristaLoginByFilial(filialKey)||'').toLowerCase(); return dedupeCobrancaBuckets(CLIENTES_CREDIARISTA?.[loginKey]||CLIENTES_FIL?.[filialKey]||{grave:[],alerta:[],atencao:[]}); }
  if(ent.type==='vendedor') return dedupeCobrancaBuckets(CLIENTES_VEND[ent.nome]||{grave:[],alerta:[],atencao:[]});
  const src=CLIENTES_FIL[ent.filial]||{grave:[],alerta:[],atencao:[]};
  const taken=vendorAssignedKeysByFilial(ent.filial);
  const filtered={grave:[],alerta:[],atencao:[]};
  ['grave','alerta','atencao'].forEach(fx=>{(src[fx]||[]).forEach(r=>{const k=cobrancaRowKey(r); if(!taken.has(k)) filtered[fx].push(r);})});
  return dedupeCobrancaBuckets(filtered);
}
function isTodayStr(s){return dateOnlyISO(s)===dateOnlyISO(new Date())}
function isLogCobrancaReal(x){const t=String(x?.titulo||'').toUpperCase(); const p=String(x?.parcela||'').toUpperCase(); const k=String(x?.cobranca_key||'').toUpperCase(); return !(/^REATIVACAO/.test(t) || /^ANIVERSARIO/.test(t) || /^REATIVACAO/.test(p) || /^ANIVERSARIO/.test(p) || /^REATIVACAO/.test(k) || /^ANIVERSARIO/.test(k));}
function getCobradosHoje(ent){
  if(ent.type==='terceiro' || ent.is_terceiro)
    return (COB_LOGS||[]).filter(x=>isLogCobrancaReal(x) && isTodayStr(x.server_time||x.data||'') && (String(x.usuario||'').toLowerCase()===COBRANCA10_LOGIN || String(x.usuario||'').toLowerCase()===COBRANCA10_NOME.toLowerCase()));
  if(ent.type==='crediarista' || ent.is_crediarista){
    const credLogin=String(ent.login||'').toLowerCase();
    const credNome=String(ent.nome||'').toLowerCase();
    const credFilial=String(ent.filial||'').toUpperCase();
    return (COB_LOGS||[]).filter(x=>isLogCobrancaReal(x) && isTodayStr(x.server_time||x.data||'') && (
      String(x.usuario||'').toLowerCase()===credLogin ||
      String(x.usuario||'').toLowerCase()===credNome ||
      (String(x.filial||'').toUpperCase()===credFilial && String(x.destino_tipo||'').toLowerCase()==='crediarista')
    ));
  }
  return (COB_LOGS||[]).filter(x=>isLogCobrancaReal(x) && isTodayStr(x.server_time||x.data||'') && String(x.filial||'')===String(ent.filial||'') && (ent.type==='filial' || String(x.destino_nome||'')===String(ent.nome||'')));
}

const DEFAULT_COBRANCA_TEMPLATE = `Olá, {primeiro_nome} tudo bem?
Aqui é da Lojas MDL - Móveis do Lar.
Passando para lembrar que tem uma parcelinha vencida na data de {vencimento}, no valor de {valor}.
Caso o pagamento já tenha sido realizado, por gentileza, desconsidere esta mensagem.
Se precisar do boleto, chave PIX ou tiver qualquer dúvida, fico à disposição para ajudar.`;

function cobrancaTemplateAtual(){
  return String(CONFIG_META?.cobranca_msg_template || DEFAULT_COBRANCA_TEMPLATE);
}
function primeiroNomeClienteJs(nome){
  const s=String(nome||'').trim();
  return s ? s.split(/\s+/)[0] : 'Cliente';
}
function valorBR(v){
  return Number(v||0).toLocaleString('pt-BR',{style:'currency',currency:'BRL'});
}

function parseCobDate(raw){
  if(!raw) return null;
  const s=String(raw).trim();
  let d=null;
  // aceita ISO, "YYYY-MM-DD HH:mm:ss", ou data local
  d=new Date(s.replace(' ', 'T'));
  if(!d || isNaN(d.getTime())){
    const m=s.match(/(\d{2})\/(\d{2})\/(\d{4})(?:\s+(\d{2}):(\d{2}))?/);
    if(m) d=new Date(Number(m[3]),Number(m[2])-1,Number(m[1]),Number(m[4]||0),Number(m[5]||0));
  }
  return (d && !isNaN(d.getTime()))?d:null;
}
function fmtCobDate(raw){
  const d=parseCobDate(raw);
  if(!d) return String(raw||'');
  return d.toLocaleString('pt-BR',{day:'2-digit',month:'2-digit',year:'numeric',hour:'2-digit',minute:'2-digit'});
}
function daysSinceCob(raw){
  const d=parseCobDate(raw);
  if(!d) return 9999;
  return Math.floor((Date.now()-d.getTime())/(1000*60*60*24));
}
function sameCobTitle(a,b){
  return String(a?.titulo||'').trim()===String(b?.titulo||'').trim()
      && String(a?.parcela||'').trim()===String(b?.parcela||'').trim()
      && (typeof similarCliente==='function' ? similarCliente(a?.cliente||a?.nome||'', b?.cliente||b?.nome||'') : normName(a?.cliente||a?.nome||'')===normName(b?.cliente||b?.nome||''));
}
function entMatchesCobLog(log,ent){
  if(!ent) return true;
  const usuario=String(log.usuario||'').toLowerCase();
  const destinoNome=String(log.destino_nome||'');
  const destinoTipo=String(log.destino_tipo||'').toLowerCase();
  const filial=String(log.filial||'').toUpperCase();
  if(ent.type==='terceiro'||ent.is_terceiro) return destinoTipo==='terceiro' || usuario.includes('terceiro') || String(ent.nome||'')===destinoNome;
  if(ent.type==='crediarista'||ent.is_crediarista){
    const login=String(ent.login||ent.nome||'').toLowerCase();
    return usuario===login || destinoNome===String(ent.nome||'') || (destinoTipo==='crediarista' && filial===String(ent.filial||'').toUpperCase());
  }
  if(ent.type==='filial') return filial===String(ent.filial||'').toUpperCase() && (destinoTipo==='filial' || destinoNome===String(ent.nome||'') || !destinoTipo);
  return destinoNome===String(ent.nome||'') || usuario===String(ent.nome||'').toLowerCase();
}
function cobLogsTitulo(reg,ent=null){
  return (COB_LOGS||[]).filter(x=>String(x.acao||'')==='whatsapp' && sameCobTitle(x,reg) && entMatchesCobLog(x,ent))
    .sort((a,b)=>(parseCobDate(a.server_time||a.data)||0)-(parseCobDate(b.server_time||b.data)||0));
}
function pagamentoAposCobranca(reg,log){
  const dtLog=parseCobDate(log?.server_time||log?.data);
  if(!dtLog) return false;
  return (QUITADOS_180||[]).some(q=>{
    if(!sameCobTitle(q,reg)) return false;
    const dp=parseDataBRjs(q.pagamento);
    return dp && dp.getTime()>=dtLog.getTime();
  });
}
function cobStatusTitulo(reg,ent=null){
  const logs=cobLogsTitulo(reg,ent);
  const last=logs.length?logs[logs.length-1]:null;
  const pago=last?pagamentoAposCobranca(reg,last):false;
  const dias=last?daysSinceCob(last.server_time||last.data):9999;
  const qtd=logs.length;
  return {
    logs,last,pago,dias,qtd,
    ultima:last?(last.server_time||last.data||''):'',
    ultima_fmt:last?fmtCobDate(last.server_time||last.data):'',
    bloqueado: Boolean(last && !pago && dias<3),
    deve_voltar: Boolean(last && !pago && dias>=3),
    proxima_tentativa: qtd+1
  };
}
function cobrancaTemplateTerceiraAtual(){
  return String(CONFIG_META?.cobranca_msg_template_terceira || `Olá, {primeiro_nome}. Tudo bem?
Aqui é da Lojas MDL - Móveis do Lar.

Já tentamos contato sobre a parcela vencida em {vencimento}, no valor de {valor}, referente ao título {titulo}/{parcela}.

Para evitar novos encargos e restrições, pedimos que regularize o pagamento o quanto antes.
Caso já tenha pago, por gentileza desconsidere esta mensagem.`);
}

function montarMensagemCobranca(reg){
  const st=reg?._cob_status||cobStatusTitulo(reg, phoneContext?.entRef||null);
  const tentativa=Number(st?.proxima_tentativa||1);
  let tpl=(tentativa>=3)?cobrancaTemplateTerceiraAtual():cobrancaTemplateAtual();
  const dados={
    primeiro_nome: primeiroNomeClienteJs(reg.cliente||reg.nome||''),
    cliente: String(reg.cliente||reg.nome||''),
    vencimento: String(reg.vencimento||''),
    valor: valorBR(reg.pendente||0),
    titulo: String(reg.titulo||''),
    parcela: String(reg.parcela||''),
    filial: String(reg.filial||''),
    vendedor: String(reg.vendedor||''),
    dias: String(reg.dias||''),
    qtd_cobrancas: String(st?.qtd||0),
    ultima_cobranca: String(st?.ultima_fmt||''),
    tentativa: String(tentativa)
  };
  Object.entries(dados).forEach(([k,v])=>{
    tpl=tpl.replaceAll(`{${k}}`, v);
  });
  return tpl;
}
function exemploMensagemCobranca(){
  return montarMensagemCobranca({
    cliente:'MARIA APARECIDA DA SILVA',
    vencimento:'10/05/2026',
    pendente:199.90,
    titulo:'123456',
    parcela:'03',
    filial:'F1',
    vendedor:'VENDEDOR EXEMPLO',
    dias:25
  });
}
function atualizarPreviewCobranca(){
  const tpl=document.getElementById('cobMsgTemplate')?.value;
  const tpl3=document.getElementById('cobMsgTemplate3')?.value;
  const oldTpl=CONFIG_META.cobranca_msg_template;
  const oldTpl3=CONFIG_META.cobranca_msg_template_terceira;
  CONFIG_META.cobranca_msg_template=tpl || DEFAULT_COBRANCA_TEMPLATE;
  CONFIG_META.cobranca_msg_template_terceira=tpl3 || cobrancaTemplateTerceiraAtual();
  const el=document.getElementById('cobMsgPreview');
  if(el) el.textContent=exemploMensagemCobranca();
  const el3=document.getElementById('cobMsgPreview3');
  if(el3){
    const exemplo={cliente:'MARIA APARECIDA DA SILVA',vencimento:'10/05/2026',pendente:199.90,titulo:'123456',parcela:'03',filial:'F1',vendedor:'VENDEDOR EXEMPLO',dias:25,_cob_status:{qtd:2,proxima_tentativa:3,ultima_fmt:'20/05/2026 10:30'}};
    el3.textContent=montarMensagemCobranca(exemplo);
  }
  CONFIG_META.cobranca_msg_template=oldTpl;
  CONFIG_META.cobranca_msg_template_terceira=oldTpl3;
}
async function salvarMensagemCobrancaGlobal(){
  const msgEl=document.getElementById('cobMsgSaveStatus');
  const tpl=String(document.getElementById('cobMsgTemplate')?.value||'').trim();
  const tpl3=String(document.getElementById('cobMsgTemplate3')?.value||'').trim();
  if(!tpl){
    if(msgEl) msgEl.textContent='⚠️ A mensagem não pode ficar vazia.';
    return;
  }
  CONFIG_META.cobranca_msg_template=tpl;
  if(tpl3) CONFIG_META.cobranca_msg_template_terceira=tpl3;
  try{
    const payload={global:CONFIG_META,individual:CONFIG_META_IND};
    const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
    const j=await resp.json();
    if(j.ok){
      await carregarConfigOnline();
      if(msgEl) msgEl.textContent='✅ Mensagem global de cobrança salva online.';
      toast('Mensagem de cobrança atualizada.','success');
      renderLogsTab();
    }else{
      if(msgEl) msgEl.textContent='⚠️ Não consegui salvar online.';
    }
  }catch(e){
    console.log('Erro ao salvar mensagem de cobrança',e);
    if(msgEl) msgEl.textContent='⚠️ Falha ao salvar online.';
  }
}
function restaurarMensagemCobrancaPadrao(){
  const t=document.getElementById('cobMsgTemplate');
  if(t) t.value=DEFAULT_COBRANCA_TEMPLATE;
  const t3=document.getElementById('cobMsgTemplate3'); if(t3) t3.value=cobrancaTemplateTerceiraAtual();
  atualizarPreviewCobranca();
}

function renderCobrancaConfigPanel(){
  const tpl=esc(cobrancaTemplateAtual());
  const tpl3=esc(cobrancaTemplateTerceiraAtual());
  return `<div class="glass panel" style="margin-bottom:14px">
    <div class="section-head" style="margin:0 0 10px">
      <div>
        <h2 style="margin:0">💬 Mensagens de cobrança</h2>
        <div class="hint">Configuração global usada ao abrir WhatsApp. O sistema usa a mensagem padrão na 1ª e 2ª cobrança. A partir da 3ª, usa a Terceira Mensagem.</div>
      </div>
      <button class="btn primary" onclick="salvarMensagemCobrancaGlobal()">Salvar global</button>
    </div>
    <div class="cobranca-config-grid">
      <div>
        <div class="input-card">
          <label>Mensagem padrão de cobrança</label>
          <textarea id="cobMsgTemplate" class="cobranca-template" oninput="atualizarPreviewCobranca()">${tpl}</textarea>
        </div>
        <div class="input-card" style="margin-top:12px">
          <label>Terceira Mensagem de Cobrança</label>
          <textarea id="cobMsgTemplate3" class="cobranca-template" oninput="atualizarPreviewCobranca()">${tpl3}</textarea>
        </div>
        <div class="placeholder-list">
          <code>{primeiro_nome}</code><code>{cliente}</code>
          <code>{vencimento}</code><code>{valor}</code>
          <code>{titulo}</code><code>{parcela}</code>
          <code>{filial}</code><code>{vendedor}</code>
          <code>{qtd_cobrancas}</code><code>{ultima_cobranca}</code><code>{tentativa}</code>
        </div>
        <div style="display:flex;gap:10px;margin-top:10px;align-items:center;flex-wrap:wrap">
          <button class="btn primary" onclick="salvarMensagemCobrancaGlobal()">Salvar global</button>
          <button class="btn soft" onclick="restaurarMensagemCobrancaPadrao()">Restaurar padrão</button>
          <span id="cobMsgSaveStatus" class="hint"></span>
        </div>
      </div>
      <div>
        <div class="input-card">
          <label>Prévia mensagem padrão</label>
          <div id="cobMsgPreview" class="preview-whats">${esc(exemploMensagemCobranca())}</div>
        </div>
        <div class="input-card" style="margin-top:12px">
          <label>Prévia terceira mensagem</label>
          <div id="cobMsgPreview3" class="preview-whats">${esc((()=>{const old=CONFIG_META.cobranca_msg_template_terceira; const ex={cliente:'MARIA APARECIDA DA SILVA',vencimento:'10/05/2026',pendente:199.90,titulo:'123456',parcela:'03',filial:'F1',vendedor:'VENDEDOR EXEMPLO',dias:25,_cob_status:{qtd:2,proxima_tentativa:3,ultima_fmt:'20/05/2026 10:30'}}; return montarMensagemCobranca(ex);})())}</div>
        </div>
      </div>
    </div>
  </div>`;
}


function normalizarListaTelefones(contatos){let base=[]; if(Array.isArray(contatos)) base=contatos; else if(typeof contatos==='string') base=contatos.split(/[;,/|]+/); const out=[]; const seen=new Set(); base.forEach(item=>{const num=String(item||'').replace(/\D/g,''); if(num.length>=10){const finalNum=num.startsWith('55')?num:'55'+num; if(!seen.has(finalNum)){seen.add(finalNum); out.push(finalNum);}}}); return out}
function matchCob(r,ent=null){return cobLogsTitulo(r,ent).length>0}
let cobExportCounter=0;
function renderCobrancasEnt(ent){
  const src=getClientesEnt(ent);
  const cobradosHoje=getCobradosHoje(ent);
  const allHoje=[...(src.grave||[]),...(src.alerta||[]),...(src.atencao||[])].filter(r=>r.novo);

  const srcAll=[...(src.grave||[]),...(src.alerta||[]),...(src.atencao||[])];

  function decorateRow(r){
    const st=cobStatusTitulo(r,ent);
    return {...r,_cob_status:st};
  }

  function shouldShowInGeral(r){
    const st=r._cob_status||cobStatusTitulo(r,ent);
    if(!st.last) return true;
    if(st.pago) return false;
    return st.deve_voltar; // volta somente depois de 3 dias sem pagamento
  }

  const renderRows=(arr,showFaixa)=>!arr.length?'<div class="empty">Nada nesta aba.</div>':arr.slice(0,150).map(raw=>{
    const r=raw._cob_status?raw:decorateRow(raw);
    const st=r._cob_status||{};
    const cobrado=Boolean(st.last);
    const bloqueado=Boolean(st.bloqueado);
    const retry=Boolean(st.deve_voltar);
    const tentativa=Number(st.proxima_tentativa||1);
    const statusChip = retry
      ? `<span class="cob-retry-chip">🔁 Cobrar novamente · ${tentativa}ª tentativa</span>`
      : (cobrado ? `<span class="cob-history-chip">🕒 Cobrado ${st.qtd}x</span>` : '');
    const info = cobrado
      ? `<div class="cob-info-box ${bloqueado?'waiting':''}">${retry?'⚠️ Cliente já foi cobrado e não pagou em 3 dias.':'✅ Cliente cobrado recentemente.'} Última cobrança: <strong>${esc(st.ultima_fmt||'')}</strong> · Total de cobranças: <strong>${st.qtd||0}</strong>${tentativa>=3?' · A próxima mensagem será a <strong>Terceira Mensagem de Cobrança</strong>.':''}</div>`
      : '';
    const btn = bloqueado
      ? `<button class="btn soft" title="Só volta para cobrança após 3 dias sem pagamento" disabled>⏳ Aguardando 3 dias</button>`
      : `<button class="btn wa" onclick='abrirWhats(${JSON.stringify(r)}, ${JSON.stringify({type:ent.type,filial:ent.filial,nome:ent.nome,login:ent.login||''})})'>💬 ${retry?'Cobrar novamente':'WhatsApp'}</button>`;
    return `<div class="row-item ${retry?'retry-due':''}"><div class="row-top"><div><div class="name">${esc(r.cliente||r.nome||'')} ${r.novo?'<span class="mini-chip" style="margin-left:6px;background:#eef7ff;color:#1e3a8a;border-color:#93c5fd">Novo hoje</span>':''} ${statusChip}</div>${info}<div class="small muted">✍️ Avalista: ${esc((r.avalista && String(r.avalista).toLowerCase()!=='nan')?r.avalista:'Sem Aval')}</div>${(r.avalista && String(r.avalista).toLowerCase()!=='nan')?'<div class="small avalista-alert">⚠️ Atenção, lembre de cobrar o AVALISTA</div>':''}<div class="small muted">🔒 Restrição crédito: ${/sem restr/i.test(String(r.restricao||''))?`<span class="restr-ok">${esc(r.restricao||'Sem Restrição')}</span>`:esc(r.restricao||'Sem informação')}</div><div class="small muted">👤 ${esc(r.vendedor||'')}</div><div class="small muted">☎️ ${esc(Array.isArray(r.telefones)?r.telefones.join(', '):(r.contato||''))}</div></div><div><strong>${esc(r.titulo||'')}</strong><div class="small muted">Título</div></div><div><strong>${r.dias||0}d</strong><div class="small muted">Dias</div></div><div><strong>${esc(r.vencimento||'')}</strong><div class="small muted">Vencimento</div></div><div><strong>${R(r.pendente||0)}</strong><div class="small muted">Pendente</div></div><div>${showFaixa?`<div class="small muted">${esc(r.faixa_label||'')}</div>`:''}${btn}</div></div></div>`;
  }).join('');

  const faixas=['grave','alerta','atencao'];
  const tabs=`<div class="tabs" style="justify-content:flex-start;margin:0 0 12px"><button class="tab active" data-cobtab="geral" onclick="switchCobTab(this,'geral')">Para cobrar</button><button class="tab" data-cobtab="novos" onclick="switchCobTab(this,'novos')">Novos Hoje</button><button class="tab" data-cobtab="cobrados" onclick="switchCobTab(this,'cobrados')">Cobrados Hoje</button><button class="tab" data-cobtab="aguardando" onclick="switchCobTab(this,'aguardando')">Aguardando 3 dias</button></div>`;
  let geral='';
  let aguardando=[];
  faixas.forEach(fx=>{
    const arr=(src[fx]||[]).map(r=>decorateRow({...r,faixa_label:fx}));
    aguardando.push(...arr.filter(r=>r._cob_status?.bloqueado));
    const vis=arr.filter(shouldShowInGeral);
    const label=fx==='grave'?'Grave':fx==='alerta'?'Alerta':'Atenção';
    geral+=`<div class="faixa-block"><div class="faixa-title ${fx}">${label}<span>${vis.length} títulos · ${R(vis.reduce((a,b)=>a+Number(b.pendente||0),0))}</span></div><div class="tableish">${renderRows(vis,false)}</div></div>`;
  });

  const cobradosRows=(cobradosHoje||[]).map(x=>{
    const m=srcAll.find(r=>cobrancaRowKey(r)===cobrancaRowKey(x))||{};
    return decorateRow({cliente:x.cliente,titulo:x.titulo,parcela:x.parcela,vencimento:x.vencimento,pendente:x.pendente,vendedor:x.usuario||m.vendedor||'',dias:m.dias||'',telefones:Array.isArray(m.telefones)&&m.telefones.length?m.telefones:[x.telefone],contato:m.contato||x.telefone,avalista:m.avalista||'',restricao:m.restricao||'',faixa_label:m.faixa||'',novo:false,pagamento:m.pagamento||'',lancamento:m.lancamento||''});
  });
  const exportId='cob_export_'+(++cobExportCounter);
  const exportRows=[];
  faixas.forEach(fx=>{((src[fx]||[]).map(r=>decorateRow({...r,faixa_label:fx}))).forEach(r=>exportRows.push(mdlCobExportRow(r,'Para cobrar')))});
  allHoje.forEach(r=>exportRows.push(mdlCobExportRow(decorateRow({...r,faixa_label:r.faixa||''}),'Novo hoje')));
  cobradosRows.forEach(r=>exportRows.push(mdlCobExportRow(r,'Cobrado hoje')));
  aguardando.forEach(r=>exportRows.push(mdlCobExportRow(r,'Aguardando 3 dias')));
  mdlRegisterExport(exportId, 'Relatorio de cobrancas - '+(ent?.nome||ent?.filial||'usuario'), exportRows);

  return `<div style="display:flex;justify-content:flex-end;margin:0 0 10px">${mdlExportButtons(exportId)}</div>${tabs}<div class="cob-pane" data-cobpane="geral">${geral}</div><div class="cob-pane hidden" data-cobpane="novos">${renderRows(allHoje.map(r=>decorateRow({...r,faixa_label:r.faixa||''})).filter(shouldShowInGeral),true)}</div><div class="cob-pane hidden" data-cobpane="cobrados">${renderRows(cobradosRows,true)}</div><div class="cob-pane hidden" data-cobpane="aguardando">${renderRows(aguardando,true)}</div>`;
}
function switchCobTab(btn,name){const box=btn.closest('.acc-body'); box.querySelectorAll('[data-cobtab]').forEach(b=>b.classList.toggle('active',b===btn)); box.querySelectorAll('[data-cobpane]').forEach(p=>p.classList.toggle('hidden',p.dataset.cobpane!==name));}
function abrirWhats(reg,entRef){const nums=normalizarListaTelefones((reg.telefones&&reg.telefones.length)?reg.telefones:reg.contato); if(!nums.length){toast('Cliente sem telefone válido.'); return} reg._cob_status=cobStatusTitulo(reg,entRef); phoneContext={reg,entRef}; if(nums.length===1){enviarWhats(nums[0]); return} const phoneList=document.getElementById('phoneList'); phoneList.innerHTML=nums.map(n=>`<button class="btn soft" style="width:100%" onclick="enviarWhats('${n}')">${n}</button>`).join(''); document.getElementById('phoneModal').classList.add('show')}
function closePhoneModal(){document.getElementById('phoneModal').classList.remove('show'); phoneContext=null}
function enviarWhats(numero){if(!phoneContext) return; const {reg,entRef}=phoneContext; const msg=montarMensagemCobranca(reg); window.open(`https://wa.me/${numero}?text=${encodeURIComponent(msg)}`,'_blank'); registrarCobrancaOnline(reg,entRef,numero); closePhoneModal()}
async function registrarCobrancaOnline(r,entRef,numero){
  // Para crediarista usa o login como usuario para que getCobradosHoje() funcione corretamente
  const usuarioLog = (entRef.type==='crediarista'||entRef.is_crediarista)
    ? (String(entRef.login||entRef.nome||usuarioAtual?.login||'').toLowerCase() || usuarioAtual?.login || 'master')
    : (usuarioAtual?.nome||usuarioAtual?.login||'master');
  const payload={
    cliente:r.cliente||r.nome||'',titulo:r.titulo||'',parcela:r.parcela||'',
    cliente_key:r.cliente_key||'',cobranca_key:r.cobranca_key||'',owner_key:r.owner_key||'',
    vencimento:r.vencimento||'',pendente:Number(r.pendente||0),telefone:numero,
    usuario:usuarioLog,login:entRef.login||usuarioAtual?.login||'',filial:entRef.filial||'',
    destino_tipo:entRef.type||'',destino_nome:entRef.nome||'',acao:'whatsapp',tentativa:Number(r._cob_status?.proxima_tentativa||1),qtd_cobrancas_antes:Number(r._cob_status?.qtd||0),ultima_cobranca_anterior:String(r._cob_status?.ultima_fmt||'')
  };
  if(payload.titulo==='REATIVACAO') marcarReativacaoUsuarioHoje(payload);
  try{
    const resp=await fetch(API_COB,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
    const j=await resp.json();
    if(j.ok){
      if(payload.titulo==='REATIVACAO') registrarReativacaoLocal(payload);
      await carregarCobrancasOnline();
      toast(payload.titulo==='REATIVACAO'?'Mensagem de reativação registrada.':'Cobrança registrada online com sucesso.','success');
      if(!detailScreen.classList.contains('hidden')){
        if(entRef.type==='crediarista'||entRef.is_crediarista){
          openCrediaristaPanel(entRef.login||'',entRef.filial||'',entRef.nome||'');
        } else {
          openEntity(entRef);
        }
      }
    } else {
      if(payload.titulo==='REATIVACAO'){
        registrarReativacaoLocal(payload); toast('Reativação registrada localmente para teste.','success');
      } else if(window.location.protocol==='file:'){
        registrarCobrancaLocal(payload); toast('Cobrança registrada localmente para teste.','success');
      } else toast('Não consegui gravar a cobrança online.');
    }
  }catch(e){
    if(payload.titulo==='REATIVACAO'){
      registrarReativacaoLocal(payload); toast('Reativação registrada localmente para teste.','success');
    } else if(window.location.protocol==='file:'){
      registrarCobrancaLocal(payload); toast('Cobrança registrada localmente para teste.','success');
    } else toast('Falha ao salvar cobrança online.');
  }
}
function registrarCobrancaLocal(payload){
  try{
    const now=new Date(); const iso=now.toISOString();
    const rec={...payload,id:'LOCAL_COB_'+Date.now(),server_time:iso,criado_em:iso,data:iso,_local:true};
    COB_LOGS=Array.isArray(COB_LOGS)?COB_LOGS:[];
    COB_LOGS.push(rec);
    localStorage.setItem('mdl_cobrancas_log_cache', JSON.stringify(COB_LOGS));
    if(!detailScreen.classList.contains('hidden') && currentDetailRef) openEntity(currentDetailRef);
    if(mainTab==='inicio') renderInicioTab();
  }catch(e){console.warn('registrarCobrancaLocal',e)}
}
function reativacaoOwnerLocalKeyFromPayload(payload){
  const filial=String(payload?.filial||'').toUpperCase();
  const nome=String(payload?.destino_nome||payload?.usuario||payload?.nome||'');
  const login=String(payload?.login||'');
  if(String(payload?.destino_tipo||'')==='filial' || /^GERENTE/i.test(nome)) return `GERENTE_${filial}`;
  return reatUserKeyFromNome(nome||login,filial);
}
function marcarReativacaoUsuarioHoje(payload){
  try{
    const hoje=(new Date()).toISOString().slice(0,10);
    const storageKey='mdl_reativacao_usuarios_'+hoje;
    const arr=JSON.parse(localStorage.getItem(storageKey)||'[]');
    const filial=String(payload?.filial||'').toUpperCase();
    const baseNome=payload?.destino_nome||payload?.usuario||payload?.nome||'';
    const key=reativacaoOwnerLocalKeyFromPayload(payload);
    const aliases=[key, reatUserKeyFromNome(baseNome,filial), String(payload?.login||'').toLowerCase()].filter(Boolean);
    const item={key,nome:baseNome,login:payload?.login||'',filial,aliases,ts:Date.now()};
    if(key && !arr.some(x=>String(x.key)===key || (Array.isArray(x.aliases)&&x.aliases.some(a=>aliases.includes(a))))) arr.push(item);
    localStorage.setItem(storageKey,JSON.stringify(arr));
  }catch(e){console.warn('marcarReativacaoUsuarioHoje',e)}
}
function usuarioReativacaoLocalHoje(o){
  try{
    const hoje=(new Date()).toISOString().slice(0,10);
    const arr=JSON.parse(localStorage.getItem('mdl_reativacao_usuarios_'+hoje)||'[]');
    const key=String(o?.key||'');
    const filial=String(o?.filial||'').toUpperCase();
    const nome=normName(o?.nome||o?.label||'');
    const label=normName(o?.label||'');
    const login=String(o?.login||'').toLowerCase();
    return arr.some(x=>{
      const xk=String(x.key||'');
      const xf=String(x.filial||'').toUpperCase();
      const xn=normName(x.nome||x.label||'');
      const xl=String(x.login||'').toLowerCase();
      const aliases=Array.isArray(x.aliases)?x.aliases.map(a=>String(a||'')):[];
      if(key && (xk===key || aliases.includes(key))) return true;
      if(login && (xl===login || aliases.includes(login))) return true;
      if(filial && xf && filial!==xf) return false;
      return !!(nome && (xn===nome || xn.includes(nome) || nome.includes(xn) || (label && (xn===label || xn.includes(label) || label.includes(xn)))));
    });
  }catch(e){return false}
}
function registrarReativacaoLocal(payload){
  try{
    const now=new Date();
    const iso=now.toISOString();
    const rec={...payload, id:'LOCAL_REAT_'+Date.now(), server_time:iso, criado_em:iso, data:iso, _local:true};
    COB_LOGS=Array.isArray(COB_LOGS)?COB_LOGS:[];
    COB_LOGS.push(rec);
    localStorage.setItem('mdl_cobrancas_log_cache', JSON.stringify(COB_LOGS));
    marcarReativacaoUsuarioHoje(payload);
    if(mainTab==='inicio') renderInicioTab();
    if(mainTab==='reativacao') renderReativacaoTab();
  }catch(e){console.warn('registrarReativacaoLocal',e)}
}
function mergeLocalCobLogs(){
  try{
    const cache=localStorage.getItem('mdl_cobrancas_log_cache');
    if(!cache) return;
    const local=JSON.parse(cache)||[];
    COB_LOGS=Array.isArray(COB_LOGS)?COB_LOGS:[];
    const keyOf=x=>String(x.id||'') || [x.titulo,x.cliente,x.parcela,x.telefone,x.server_time||x.criado_em||x.data].map(v=>String(v||'')).join('|');
    const seen=new Set(COB_LOGS.map(keyOf));
    local.forEach(x=>{const k=keyOf(x); if(k && !seen.has(k)){COB_LOGS.push(x); seen.add(k);}});
  }catch(e){console.warn('mergeLocalCobLogs',e)}
}
async function carregarCobrancasOnline(){
  try{
    const r=await fetchComTimeout(API_COB+'?_='+Date.now(),{},5000);
    const txt=await r.text(); let j={ok:false};
    try{j=JSON.parse(txt);}catch(e){}
    if(j.ok && Array.isArray(j.data)){
      COB_LOGS=j.data;
      mergeLocalCobLogs();
      try{localStorage.setItem('mdl_cobrancas_log_cache', JSON.stringify(COB_LOGS));}catch(e){}
    }else{
      const cache=localStorage.getItem('mdl_cobrancas_log_cache');
      if(cache) COB_LOGS=JSON.parse(cache)||COB_LOGS||[];
      if(txt) console.log('cobrancas_api retorno:', txt);
    }
  }catch(e){
    console.log(e);
    try{
      const cache=localStorage.getItem('mdl_cobrancas_log_cache');
      if(cache) COB_LOGS=JSON.parse(cache)||COB_LOGS||[];
    }catch(_e){}
  }
  mergeLocalCobLogs();
  RECEBIMENTOS_CONCILIADOS=getQuitadosConciliados();
  console.log('🔗 Quitados conciliados:', RECEBIMENTOS_CONCILIADOS);
}

async function removerCobranca(id,cliente='',titulo='',parcela=''){if(!confirm('Remover esta cobrança do histórico?')) return; try{const r=await fetch(API_COB,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({action:'delete',id,cliente,titulo,parcela})}); const txt=await r.text(); let j={ok:false}; try{j=JSON.parse(txt);}catch(e){} if(j.ok){toast('Cobrança removida.','success'); await carregarCobrancasOnline(); renderLogsTab(); renderList(); if(currentDetailRef) openEntity(currentDetailRef);}else{console.log('Falha remover cobrança:', txt); toast('Não consegui remover.')}}catch(e){console.log(e); toast('Falha ao remover cobrança.')}}
function toggleAcc(el){el.parentElement.classList.toggle('open')}
function normalizarConfigMetaPayloadOnline(j){
  try{
    let data=(j && typeof j==='object' && j.data && typeof j.data==='object') ? j.data : j;
    if(!data || typeof data!=='object') return null;
    if(data.global || data.individual){
      return {global:(data.global && typeof data.global==='object')?data.global:{}, individual:(data.individual && typeof data.individual==='object' && !Array.isArray(data.individual))?data.individual:{}};
    }
    const keys=['grave_pct','alerta_pct','atencao_pct','gerente_vendas_min_pct','gerente_servicos_min_pct','comissao_pagamento_texto','vendedor_policy','gerente_policy','camp_meta_diaria_vend','camp_meta_diaria_ger'];
    if(keys.some(k=>Object.prototype.hasOwnProperty.call(data,k))) return {global:data, individual:{}};
  }catch(e){}
  return null;
}
window._configMetaOnlineLoaded=false;
async function carregarConfigOnline(){try{const r=await fetchComTimeout(API_CFG+'?_='+Date.now(),{},8000); const j=await r.json(); const payload=normalizarConfigMetaPayloadOnline(j); if(payload){CONFIG_META={...CONFIG_META,...(payload.global||{})}; CREDIARISTAS_CONFIG=getCrediaristasConfig(); CONFIG_META_IND=payload.individual||{}; window._configMetaOnlineLoaded=true;}}catch(e){console.log('Falha ao carregar config meta',e);}}

function optionTargets(){let opts=''; flattenFiliais().forEach(f=>{opts+=`<option value="FILIAL::${f.filial}">🏬 ${filialLabel(f.filial)}</option>`}); opts+=`<option value="VEND::${COBRANCA10_NOME}_FTER">🤝 ${COBRANCA10_NOME} (Cobranças Terceiro)</option>`; crediaristaEntities().forEach(c=>{opts+=`<option value="VEND::${c.nome}_${c.filial}">🧾 ${c.nome} (${c.filial})</option>`}); flattenVendedores().forEach(v=>{opts+=`<option value="VEND::${v.nome}_${v.filial}">👤 ${v.nome} (${v.filial})</option>`}); return opts}
function fillMetaForm(mode,val){const cfg=mode==='global'?{...CONFIG_META}:mergedMetaConfig(metaAliasesFromRaw(val)); ['grave_pct','alerta_pct','atencao_pct','peso_grave','peso_alerta','peso_atencao','bonus_50','bonus_75','bonus_85','bonus_100','vendas_min_pct','servicos_min_pct','gerente_vendas_min_pct','gerente_servicos_min_pct','vendedor_rentab_min_mercantil_pct','gerente_rentab_min_mercantil_pct','cobranca_global_rateio_pct','comissao_pagamento_texto'].forEach(k=>{const el=document.getElementById('cfg_'+k); if(el) el.value=cfg[k]??''}); renderCommissionPanel(cfg)}
function canonicalMetaLabelFromKey(k){
  const v=String(k||'').trim();
  if(!v) return '';
  if(v.startsWith('VEND::')) return v;
  if(v.startsWith('FILIAL::')) return v;
  if(v.startsWith('vend:')) return 'VEND::'+v.slice(5);
  if(v.startsWith('fil:')) return 'FILIAL::'+v.slice(4);
  if(/^F\d+/i.test(v)) return 'FILIAL::'+v.toUpperCase();
  if(v.includes('_F')) return 'VEND::'+v;
  return '';
}
function renderSavedMetaList(){
  const labels=[]; const seen=new Set();
  Object.keys(CONFIG_META_IND||{}).forEach(k=>{const canon=canonicalMetaLabelFromKey(k); if(canon && !seen.has(canon)){seen.add(canon); labels.push(canon);}});
  labels.sort();
  const lst=document.getElementById('metaSavedList');
  if(lst) lst.innerHTML=labels.length?('Individuais salvos: '+labels.map(k=>`<span class="mini-chip">${esc(k)}</span>`).join(' ')):'Nenhuma configuração individual salva.';
}

function renderCrediaristasConfigPanel(){
  const rows=getCrediaristasConfig();
  return `<div class="section-head" style="margin-top:14px"><div><h2 style="font-size:18px">🧾 Crediaristas configuráveis</h2><div class="hint">Fluxo correto: primeiro crie o login em Senhas. Depois, aqui, vincule esse login à filial/base e ao %. Quando trocar a pessoa, crie novo login, exemplo CREDIARISTAF2_02, e remova o antigo daqui. O histórico antigo permanece salvo. O % é só da carteira que o crediarista enxerga; não altera o rateio 60% filial / 40% vendedores.</div></div></div>
  <div id="credConfigRows">${rows.map(r=>`<div class="glass" style="padding:10px;margin-bottom:8px;border-radius:14px"><div class="form-grid bonus">
    <div class="input-card"><label>Login do usuário</label><input class="cred-login" value="${esc(r.login||'')}" placeholder="crediaristaf2_01"></div>
    <div class="input-card"><label>Nome exibido</label><input class="cred-nome" value="${esc(r.nome||'')}" placeholder="Maria - crediarista F2"></div>
    <div class="input-card"><label>Filial/base</label><select class="cred-filial">${ORDEM.map(f=>`<option value="${f}" ${String(r.filial||'').toUpperCase()===f?'selected':''}>${f}</option>`).join('')}</select></div>
    <div class="input-card"><label>% da base</label><input class="cred-pct" type="number" step="0.01" value="${Number(r.pct||100)}"></div>
  </div><button type="button" class="btn soft" onclick="this.closest('.glass').remove()">🗑️ Remover</button></div>`).join('')}</div>
  <button type="button" class="btn soft" onclick="addCrediaristaConfigRow()">➕ Adicionar crediarista</button>`;
}
function addCrediaristaConfigRow(){
  const box=document.getElementById('credConfigRows'); if(!box) return;
  box.insertAdjacentHTML('beforeend', `<div class="glass" style="padding:10px;margin-bottom:8px;border-radius:14px"><div class="form-grid bonus">
    <div class="input-card"><label>Login do usuário</label><input class="cred-login" placeholder="crediaristaf2_01"></div>
    <div class="input-card"><label>Nome exibido</label><input class="cred-nome" placeholder="Nome da crediarista"></div>
    <div class="input-card"><label>Filial/base</label><select class="cred-filial">${ORDEM.map(f=>`<option value="${f}">${f}</option>`).join('')}</select></div>
    <div class="input-card"><label>% da base</label><input class="cred-pct" type="number" step="0.01" value="100"></div>
  </div><button type="button" class="btn soft" onclick="this.closest('.glass').remove()">🗑️ Remover</button></div>`);
}
function readCrediaristasConfigFromUI(){
  const box=document.getElementById('credConfigRows'); if(!box) return getCrediaristasConfig();
  return Array.from(box.querySelectorAll(':scope > .glass')).map(row=>{
    const login=String(row.querySelector('.cred-login')?.value||'').trim().toLowerCase();
    const filial=String(row.querySelector('.cred-filial')?.value||'').toUpperCase();
    const nome=String(row.querySelector('.cred-nome')?.value||'').trim()||`Crediarista ${filial}`;
    const pct=Math.max(0,Math.min(100,Number(row.querySelector('.cred-pct')?.value||100)));
    return {login,nome,filial,pct};
  }).filter(r=>r.login&&r.filial&&r.pct>0);
}

function renderMetasTab(){const cards=[...flattenVendedores(),...flattenFiliais()]; const currentMode=window._metaMode||'global'; const currentTarget=window._metaSelectedTarget||''; metaSection.innerHTML=`<div class="section-head"><div><h2>🎯 Configuração de metas e bônus</h2><div class="hint">Altere globalmente ou por vendedor/filial. Ao salvar, já fica online.</div></div></div><div class="meta-layout"><div class="glass panel"><div class="tabs" style="justify-content:flex-start;margin-top:0"><button id="btnModeGlobal" class="tab" onclick="setMetaMode('global')">🌐 Padrão global</button><button id="btnModeInd" class="tab" onclick="setMetaMode('individual')">👤 Por vendedor/filial</button></div><div id="metaSelectWrap" class="hidden" style="margin:8px 0 14px"><div class="input-card"><label>Selecionar alvo</label><select id="metaTarget" onchange="loadMetaSelected()"><option value="">Selecione...</option>${optionTargets()}</select></div></div><div class="section-head" style="margin-top:10px"><div><h2 style="font-size:18px">% de meta por faixa</h2></div></div><div class="form-grid"><div class="input-card"><label>Grave</label><input id="cfg_grave_pct" type="number" step="0.01"></div><div class="input-card"><label>Alerta</label><input id="cfg_alerta_pct" type="number" step="0.01"></div><div class="input-card"><label>Atenção</label><input id="cfg_atencao_pct" type="number" step="0.01"></div></div><div class="section-head" style="margin-top:14px"><div><h2 style="font-size:18px">Pesos da meta geral</h2></div></div><div class="form-grid"><div class="input-card"><label>Peso Grave</label><input id="cfg_peso_grave" type="number" step="0.01"></div><div class="input-card"><label>Peso Alerta</label><input id="cfg_peso_alerta" type="number" step="0.01"></div><div class="input-card"><label>Peso Atenção</label><input id="cfg_peso_atencao" type="number" step="0.01"></div></div><div class="section-head" style="margin-top:14px"><div><h2 style="font-size:18px">Bônus / mensagem da faixa <span class="note">· Não acumulativo</span></h2></div></div><div class="form-grid bonus"><div class="input-card"><label>50%</label><input id="cfg_bonus_50" placeholder="Ex: Parabéns, você ganhou R$ 100,00"></div><div class="input-card"><label>75%</label><input id="cfg_bonus_75"></div><div class="input-card"><label>85%</label><input id="cfg_bonus_85"></div><div class="input-card"><label>100%</label><input id="cfg_bonus_100"></div></div><div class="section-head" style="margin-top:14px"><div><h2 style="font-size:18px">💲 Meta mínima Vendas e Serviços</h2><div class="hint">Configuração inicial para comissão de vendedor e gerente/filial.</div></div></div><div class="form-grid bonus"><div class="input-card"><label>Vendedor · mínimo vendas (%)</label><input id="cfg_vendas_min_pct" type="number" step="0.01" placeholder="80"></div><div class="input-card"><label>Vendedor · mínimo serviços (%)</label><input id="cfg_servicos_min_pct" type="number" step="0.01" placeholder="80"></div><div class="input-card"><label>Gerente/Filial · mínimo vendas (%)</label><input id="cfg_gerente_vendas_min_pct" type="number" step="0.01" placeholder="80"></div><div class="input-card"><label>Gerente/Filial · mínimo serviços (%)</label><input id="cfg_gerente_servicos_min_pct" type="number" step="0.01" placeholder="80"></div><div class="input-card"><label>Vendedor · mercantil mínimo para rentab (%)</label><input id="cfg_vendedor_rentab_min_mercantil_pct" type="number" step="0.01" placeholder="80"></div><div class="input-card"><label>Gerente/Filial · mercantil mínimo para rentab (%)</label><input id="cfg_gerente_rentab_min_mercantil_pct" type="number" step="0.01" placeholder="80"></div></div><div class="section-head" style="margin-top:14px"><div><h2 style="font-size:18px">🤝 Rateio cobrança global</h2><div class="hint">Percentual do total único da cobrança geral distribuído para os usuários do tipo cobrança global (ex.: Cobrança10).</div></div></div><div class="form-grid bonus"><div class="input-card"><label>Usuários de cobrança global (%)</label><input id="cfg_cobranca_global_rateio_pct" type="number" step="0.01" placeholder="20"></div></div><div class="section-head" style="margin-top:14px"><div><h2 style="font-size:18px">🧾 Texto da comissão</h2><div class="hint">Frase exibida no card de comissão de crediaristas/cobrança terceira.</div></div></div><div class="form-grid bonus"><div class="input-card" style="grid-column:1/-1"><label>Mensagem abaixo da comissão</label><input id="cfg_comissao_pagamento_texto" placeholder="Ex: Pagamento previsto para o dia 25 do mês seguinte"></div></div>${renderCrediaristasConfigPanel()}<div id="commissionPanel"></div><div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:16px"><button class="btn primary" onclick="salvarMeta()">💾 Salvar configuração</button><button class="btn ghost" onclick="removerMetaIndividual()">🗑️ Remover individual</button></div><div id="metaSaveMsg" class="note" style="margin-top:10px"></div><div id="metaSavedList" class="note" style="margin-top:10px"></div></div></div>`; const sel=document.getElementById('metaTarget'); if(sel && currentTarget) sel.value=currentTarget; setMetaMode(currentMode); renderSavedMetaList();}
function setMetaMode(mode){window._metaMode=mode; const bg=document.getElementById('btnModeGlobal'); const bi=document.getElementById('btnModeInd'); if(bg) bg.classList.toggle('active',mode==='global'); if(bi) bi.classList.toggle('active',mode==='individual'); const wrap=document.getElementById('metaSelectWrap'); if(wrap) wrap.classList.toggle('hidden',mode!=='individual'); if(mode==='global'){fillMetaForm('global')} else {const raw=(document.getElementById('metaTarget')?.value)||window._metaSelectedTarget||''; if(raw){window._metaSelectedTarget=raw; fillMetaForm('individual',raw)} else {fillMetaForm('global')}}}
function loadMetaSelected(){const val=document.getElementById('metaTarget').value; window._metaSelectedTarget=val; if(!val){fillMetaForm('global'); return;} fillMetaForm('individual',val)}
function collectMetaForm(){const out={}; ['grave_pct','alerta_pct','atencao_pct','peso_grave','peso_alerta','peso_atencao','vendas_min_pct','servicos_min_pct','gerente_vendas_min_pct','gerente_servicos_min_pct','vendedor_rentab_min_mercantil_pct','gerente_rentab_min_mercantil_pct','cobranca_global_rateio_pct'].forEach(k=>out[k]=Number(document.getElementById('cfg_'+k).value||0)); ['bonus_50','bonus_75','bonus_85','bonus_100','comissao_pagamento_texto'].forEach(k=>out[k]=document.getElementById('cfg_'+k)?.value||''); return {...out,crediaristas_config:readCrediaristasConfigFromUI(),...readCommissionPanel()}}
async function salvarMeta(){
  const msgEl=document.getElementById('metaSaveMsg');
  const cfg=collectMetaForm();
  if(window._metaMode==='global'){
    CONFIG_META={...CONFIG_META,...cfg};
  } else {
    const raw=(document.getElementById('metaTarget')?.value)||window._metaSelectedTarget||'';
    if(!raw){if(msgEl) msgEl.textContent='Selecione um vendedor ou filial.'; return;}
    window._metaSelectedTarget=raw;
    const aliases=metaAliasesFromRaw(raw);
    const payloadCfg={...cfg,__saved_at:new Date().toISOString(),__target:raw};
    aliases.forEach(k=>CONFIG_META_IND[k]={...payloadCfg});
  }
  try{
    const payload={global:CONFIG_META,individual:CONFIG_META_IND};
    const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
    const j=await resp.json();
    if(msgEl) msgEl.textContent=j.ok?'✅ Salvo online com sucesso.':'⚠️ Não consegui salvar online.';
    if(j.ok){
      await carregarConfigOnline();
      renderMetasTab();
      renderList();
      const restoreMsg=document.getElementById('metaSaveMsg');
      if(restoreMsg) restoreMsg.textContent='✅ Salvo online com sucesso.';
    }
  }catch(e){console.log('Erro ao salvar meta',e); if(msgEl) msgEl.textContent='⚠️ Não consegui salvar online.';}
}
async function removerMetaIndividual(){
  const msgEl=document.getElementById('metaSaveMsg');
  const raw=(document.getElementById('metaTarget')?.value)||window._metaSelectedTarget||'';
  if(window._metaMode!=='individual' || !raw){if(msgEl) msgEl.textContent='Selecione um alvo individual para remover.'; return;}
  metaAliasesFromRaw(raw).forEach(k=>delete CONFIG_META_IND[k]);
  try{
    const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})});
    const j=await resp.json();
    if(msgEl) msgEl.textContent=j.ok?'✅ Configuração individual removida.':'⚠️ Não consegui remover online.';
    if(j.ok){await carregarConfigOnline(); renderMetasTab(); const restoreMsg=document.getElementById('metaSaveMsg'); if(restoreMsg) restoreMsg.textContent='✅ Configuração individual removida.';}
  }catch(e){console.log('Erro ao remover meta individual',e); if(msgEl) msgEl.textContent='⚠️ Não consegui remover online.';}
}


// ===== V7.3 EXPORTAÇÃO INDIVIDUAL PDF/EXCEL =====
window.MDL_EXPORT_LISTS = window.MDL_EXPORT_LISTS || {};
function mdlExportVal(v){return String(v??'').replace(/<[^>]+>/g,'').replace(/&nbsp;/g,' ').trim()}
function mdlExportSafeFile(s){return String(s||'lista').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/[^a-zA-Z0-9_-]+/g,'_').replace(/^_+|_+$/g,'').slice(0,90)||'lista'}
function mdlExportEscapeHtml(s){return String(s??'').replace(/[&<>'"]/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[ch]))}
function mdlExportRowsToHtmlTable(rows,title){
  rows=Array.isArray(rows)?rows:[];
  const headers=[...new Set(rows.flatMap(r=>Object.keys(r||{})))]
  const head=headers.map(h=>`<th>${mdlExportEscapeHtml(h)}</th>`).join('');
  const body=rows.map(r=>`<tr>${headers.map(h=>`<td>${mdlExportEscapeHtml(r?.[h]??'')}</td>`).join('')}</tr>`).join('');
  return `<table border="1" cellspacing="0" cellpadding="5"><thead><tr>${head}</tr></thead><tbody>${body}</tbody></table>`;
}
function mdlDownloadBlob(filename, mime, content){
  const blob=new Blob([content],{type:mime});
  const url=URL.createObjectURL(blob);
  const a=document.createElement('a'); a.href=url; a.download=filename; document.body.appendChild(a); a.click(); a.remove();
  setTimeout(()=>URL.revokeObjectURL(url),800);
}
function mdlExportList(id,fmt){
  try{
    const pack=(window.MDL_EXPORT_LISTS||{})[id];
    if(!pack || !Array.isArray(pack.rows) || !pack.rows.length){toast('Não há dados para exportar.','warn'); return;}
    const title=pack.title||'Lista Dashboard MDL';
    const filename=mdlExportSafeFile(title)+'_'+new Date().toISOString().slice(0,10);
    const table=mdlExportRowsToHtmlTable(pack.rows,title);
    if(fmt==='excel'){
      const html=`<html><head><meta charset="UTF-8"></head><body><h2>${mdlExportEscapeHtml(title)}</h2><p>Gerado em ${new Date().toLocaleString('pt-BR')}</p>${table}</body></html>`;
      mdlDownloadBlob(filename+'.xls','application/vnd.ms-excel;charset=utf-8',html);
      return;
    }
    const w=window.open('','_blank');
    if(!w){toast('Pop-up bloqueado. Libere pop-ups para exportar PDF.','warn'); return;}
    w.document.write(`<html><head><meta charset="UTF-8"><title>${mdlExportEscapeHtml(title)}</title><style>body{font-family:Arial,sans-serif;padding:24px;color:#111}h1{font-size:20px}table{border-collapse:collapse;width:100%;font-size:11px}th{background:#f1f5f9}td,th{border:1px solid #cbd5e1;padding:5px;text-align:left;vertical-align:top}@media print{button{display:none}}</style></head><body><button onclick="window.print()" style="padding:10px 14px;margin-bottom:16px">Imprimir / Salvar PDF</button><h1>${mdlExportEscapeHtml(title)}</h1><p>Gerado em ${new Date().toLocaleString('pt-BR')}</p>${table}<script>setTimeout(()=>window.print(),500)<\/script></body></html>`);
    w.document.close();
  }catch(e){console.error(e); toast('Erro ao exportar lista.','warn')}
}
function mdlExportButtons(id){return `<span class="export-actions" onclick="event.stopPropagation()"><button class="btn soft btn-xs" onclick="mdlExportList('${id}','pdf')">📄 PDF</button><button class="btn soft btn-xs" onclick="mdlExportList('${id}','excel')">📊 Excel</button></span>`}
function mdlRegisterExport(id,title,rows){window.MDL_EXPORT_LISTS=window.MDL_EXPORT_LISTS||{}; window.MDL_EXPORT_LISTS[id]={title,rows:Array.isArray(rows)?rows:[]}; return id;}
function mdlReatExportRow(r,status){return {Cliente:r.cliente||'',Filial:r.filial||'',Cidade:r.cidade||'',Responsavel:r._owner?.label||'',Telefone:(r.telefones||[]).map(fmtTelBR).join(', '),Dias_sem_compra:r.dias_sem_movimento||'',Ultimo_movimento:r.ultimo_movimento||'',Status:status||''}}
function mdlAnivExportRow(r,status){return {Cliente:r.cliente||'',Filial:r.filial||'',Cidade:r.cidade||'',Responsavel:r._owner?.label||'',Telefone:(r.telefones||[]).map(fmtTelBR).join(', '),Nascimento:r.nascimento||'',Status:status||''}}
function mdlCobExportRow(r,status){return {Cliente:r.cliente||r.nome||'',Titulo:r.titulo||'',Parcela:r.parcela||'',Filial:r.filial||'',Vendedor:r.vendedor||'',Faixa:r.faixa_label||r.faixa||'',Vencimento:r.vencimento||'',Dias:r.dias||'',Pendente:R(r.pendente||0),Telefone:Array.isArray(r.telefones)?r.telefones.join(', '):(r.contato||''),Avalista:r.avalista||'',Restricao:r.restricao||'',Status:status||''}}

// ===== V6.8 TELEGRAM CONFIGURÁVEL PELO MASTER =====
function tgBool(v){return v===true || v==='1' || v===1 || String(v||'').toLowerCase()==='true'}
function telegramContacts(){return Array.isArray(CONFIG_META?.telegram_contacts)?CONFIG_META.telegram_contacts:[]}
function tgNovoContato(){
  CONFIG_META.telegram_contacts = telegramContacts();
  CONFIG_META.telegram_contacts.push({id:'tg_'+Date.now(),nome:'',chat_id:'',ativo:true,erros:true,meta_diaria:true,meta_mensal:true,avisos:true,resumo:true});
  renderTelegramTab();
}
function tgRemoverContato(id){
  CONFIG_META.telegram_contacts = telegramContacts().filter(x=>String(x.id)!==String(id));
  renderTelegramTab();
}
function tgReadRows(){
  const rows=[...document.querySelectorAll('.tg-row')];
  return rows.map((row,idx)=>({
    id: row.dataset.id || ('tg_'+Date.now()+'_'+idx),
    nome: row.querySelector('[data-k="nome"]')?.value?.trim() || '',
    chat_id: row.querySelector('[data-k="chat_id"]')?.value?.trim() || '',
    ativo: !!row.querySelector('[data-k="ativo"]')?.checked,
    erros: !!row.querySelector('[data-k="erros"]')?.checked,
    meta_diaria: !!row.querySelector('[data-k="meta_diaria"]')?.checked,
    meta_mensal: !!row.querySelector('[data-k="meta_mensal"]')?.checked,
    avisos: !!row.querySelector('[data-k="avisos"]')?.checked,
    resumo: !!row.querySelector('[data-k="resumo"]')?.checked
  })).filter(x=>x.nome || x.chat_id);
}
function tgRowHtml(c){
  const id=esc(c.id||('tg_'+Math.random().toString(16).slice(2)));
  const ck=(k)=>tgBool(c[k])?'checked':'';
  return `<div class="row-item tg-row" data-id="${id}">
    <div class="row-top" style="grid-template-columns:1.2fr 1.1fr 72px 92px 118px 118px 92px 92px 76px;gap:8px;align-items:center">
      <div class="input-card" style="padding:8px"><label>Nome/grupo</label><input data-k="nome" value="${esc(c.nome||'')}" placeholder="Ex: Grupo Diretoria"></div>
      <div class="input-card" style="padding:8px"><label>Chat ID</label><input data-k="chat_id" value="${esc(c.chat_id||'')}" placeholder="Ex: -100123..."></div>
      <label class="pill" style="justify-content:center"><input data-k="ativo" type="checkbox" ${ck('ativo')}> Ativo</label>
      <label class="pill"><input data-k="erros" type="checkbox" ${ck('erros')}> Erros</label>
      <label class="pill"><input data-k="meta_diaria" type="checkbox" ${ck('meta_diaria')}> Meta diária</label>
      <label class="pill"><input data-k="meta_mensal" type="checkbox" ${ck('meta_mensal')}> Meta mensal</label>
      <label class="pill"><input data-k="avisos" type="checkbox" ${ck('avisos')}> Avisos</label>
      <label class="pill"><input data-k="resumo" type="checkbox" ${ck('resumo')}> Resumo</label>
      <button class="btn soft btn-xs" onclick="tgRemoverContato('${id}')">Remover</button>
    </div>
  </div>`;
}
async function salvarTelegramConfig(){
  CONFIG_META.telegram_contacts=tgReadRows();
  CONFIG_META.telegram_templates={
    meta_diaria: document.getElementById('tgTplMetaDiaria')?.value || '',
    meta_mensal: document.getElementById('tgTplMetaMensal')?.value || ''
  };
  const msg=document.getElementById('tgSaveMsg');
  try{
    const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})});
    const j=await resp.json();
    if(msg) msg.textContent=j.ok?'Configuração Telegram salva.':'Não consegui salvar configuração Telegram.';
    toast(j.ok?'Contatos Telegram salvos.':'Falha ao salvar Telegram.',j.ok?'success':'warn');
  }catch(e){
    if(msg) msg.textContent='Salvo localmente para teste, mas não confirmou no servidor.';
    toast('Falha ao salvar no servidor.','warn');
  }
}
function renderTelegramTab(){
  if(!telegramSection) return;
  const rows=telegramContacts();
  const tpl=CONFIG_META.telegram_templates||{};
  const defDiaria='🎯🚀 PARABÉNS! META DIÁRIA BATIDA\n\n👏 Destaque: {nome}\n📈 Meta atingida: {atingido}\n🛒 Tipo: Venda mercantil\n📅 Data: {data}\n\n🔥 Excelente resultado no Controle de Meta do Sólidus!\n💪 MISSÃO DADA! MISSÃO CUMPRIDA!';
  const defMensal='🏆🚀 PARABÉNS! META MENSAL BATIDA\n\n👏 Destaque: {nome}\n📈 Meta atingida: {atingido}\n🛒 Tipo: Venda mercantil / {tipo}\n🗓️ Competência: {competencia}\n\n🔥 Excelente resultado no Controle de Meta do Sólidus!\n💪 Resultado de time forte!';
  telegramSection.innerHTML=`<div class="section-head"><div><h2>📲 Telegram / Notificações</h2><div class="hint">Configure grupos/contatos e também personalize as mensagens automáticas. O token do bot continua seguro no Railway; aqui entra somente o Chat ID.</div></div></div>
  <div class="glass panel">
    <div class="section-head" style="margin:0 0 12px"><div><h2 style="font-size:18px">Contatos ativos para envio</h2><div class="hint">Para grupo, adicione o bot no grupo e use o Chat ID negativo. Para pessoa individual, ela precisa iniciar conversa com o bot primeiro.</div></div><button class="btn primary" onclick="tgNovoContato()">+ Adicionar contato/grupo</button></div>
    <div class="tableish">${rows.length?rows.map(tgRowHtml).join(''):'<div class="empty">Nenhum contato configurado ainda. Use o botão Adicionar contato/grupo.</div>'}</div>
    <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:14px"><button class="btn primary" onclick="salvarTelegramConfig()">💾 Salvar Telegram</button><button class="btn soft" onclick="toast('Para testar, abra o monitor Railway/local e clique em Teste Telegram.','info')">Como testar?</button></div>
    <div id="tgSaveMsg" class="note" style="margin-top:10px"></div>
  </div>
  <div class="glass panel" style="margin-top:14px"><div class="section-head" style="margin:0 0 12px"><div><h2 style="font-size:18px">✏️ Mensagens automáticas do Telegram</h2><div class="hint">Você pode editar o texto. Variáveis disponíveis: {nome}, {atingido}, {tipo}, {escopo}, {data}, {competencia}. Por segurança, o robô remove valores em R$ das mensagens de meta.</div></div></div><div class="search-row" style="grid-template-columns:1fr 1fr;align-items:stretch"><div class="input-card"><label>Mensagem de META DIÁRIA BATIDA</label><textarea id="tgTplMetaDiaria" rows="9" style="min-height:210px;width:100%;resize:vertical">${esc(tpl.meta_diaria||defDiaria)}</textarea></div><div class="input-card"><label>Mensagem de META MENSAL BATIDA</label><textarea id="tgTplMetaMensal" rows="9" style="min-height:210px;width:100%;resize:vertical">${esc(tpl.meta_mensal||defMensal)}</textarea></div></div><div class="hint" style="margin-top:8px">Não use valores em reais aqui. O sistema mantém apenas percentual de meta atingida.</div></div>
  <div class="glass panel" style="margin-top:14px"><h3>Como pegar o Chat ID</h3><div class="hint">1) Crie/adicone o bot em um grupo. 2) Mande uma mensagem no grupo. 3) No navegador, abra: https://api.telegram.org/botSEU_TOKEN/getUpdates. 4) Procure o campo <b>chat</b> e copie o <b>id</b>. Grupo normalmente começa com -100.</div></div>`;
}

function renderLogsTab(){const cfgPanel=renderCobrancaConfigPanel(); const LOGS_REAIS=(COB_LOGS||[]).filter(isLogCobrancaReal); const filOpts=['<option value="">Todas as filiais</option>',...ORDEM.map(f=>`<option value="${f}">${f}</option>`)].join(''); const vendOpts=['<option value="">Todos os usuários</option>',...Array.from(new Set(LOGS_REAIS.map(x=>x.usuario).filter(Boolean))).sort().map(v=>`<option value="${esc(v)}">${esc(v)}</option>`)].join(''); logSection.innerHTML=cfgPanel+`<div class="section-head"><div><h2>🧾 Histórico de cobranças</h2><div class="hint">Filtre por data, usuário ou filial. Também é possível remover lançamentos indevidos.</div></div></div><div class="glass panel"><div class="search-row"><div class="input-card"><label>Buscar cliente/título</label><input id="logQ" placeholder="Nome, título, parcela"></div><div class="input-card"><label>Data inicial</label><input id="logDe" type="date"></div><div class="input-card"><label>Data final</label><input id="logAte" type="date"></div><div class="input-card"><label>Filial</label><select id="logFil">${filOpts}</select></div></div><div class="search-row" style="margin-top:10px"><div class="input-card"><label>Usuário</label><select id="logVend">${vendOpts}</select></div><div style="display:flex;align-items:end;gap:10px"><button class="btn primary" onclick="applyLogFilter()">Filtrar</button><button class="btn soft" onclick="clearLogFilter()">Limpar</button></div></div><div id="logsList" class="logs-list"></div></div>`; applyLogFilter()}
function parseDateBR(s){if(!s) return null; const v=String(s).trim(); let m=v.match(/^(\d{1,2})\/(\d{1,2})\/(\d{2}|\d{4})$/); if(m){let y=Number(m[3]); if(y<100)y+=2000; const d=new Date(y, Number(m[2])-1, Number(m[1])); return isNaN(d.getTime())?null:d} m=v.match(/^(\d{4})-(\d{2})-(\d{2})(?:[ T](\d{1,2}):(\d{2})(?::(\d{2}))?)?/); if(m){const d=new Date(Number(m[1]),Number(m[2])-1,Number(m[3]),Number(m[4]||0),Number(m[5]||0),Number(m[6]||0)); return isNaN(d.getTime())?null:d} const d=new Date(v); return isNaN(d.getTime())?null:d}
function parseDate(s){return parseDateBR(s)}
function dateOnlyISO(s){const d=parseDateBR(s); if(!d) return ''; const y=d.getFullYear(); const m=String(d.getMonth()+1).padStart(2,'0'); const da=String(d.getDate()).padStart(2,'0'); return `${y}-${m}-${da}`}
function sameDateOnly(a,b){return !!a && !!b && dateOnlyISO(a)===dateOnlyISO(b)}
function renderLogPage(){const host=document.getElementById('logsList'); if(!host) return; const total=_logFiltered.length; const maxPage=Math.max(1,Math.ceil(total/LOG_PAGE_SIZE)); if(_logPage>maxPage)_logPage=maxPage; if(_logPage<1)_logPage=1; const ini=(_logPage-1)*LOG_PAGE_SIZE; const page=_logFiltered.slice(ini,ini+LOG_PAGE_SIZE); const pager=total?`<div class="log-pager"><div><strong>${total}</strong> cobrança(s) encontradas · mostrando ${ini+1}-${Math.min(ini+LOG_PAGE_SIZE,total)} · página ${_logPage}/${maxPage}</div><div style="display:flex;gap:8px"><button class="btn soft" ${_logPage<=1?'disabled':''} onclick="_logPage--;renderLogPage()">⬅️ Anterior</button><button class="btn soft" ${_logPage>=maxPage?'disabled':''} onclick="_logPage++;renderLogPage()">Próxima ➡️</button></div></div>`:''; host.innerHTML=total?pager+page.map((x,idx)=>{const realIdx=ini+idx; return `<div class="log-row"><div><div style="font-weight:900">${esc(x.cliente||'')}</div><div class="small muted">${esc(x.titulo||'')} · Parcela ${esc(x.parcela||'')}</div></div><div><strong>${R(x.pendente||0)}</strong><div class="small muted">${esc(x.filial||'')} · ${esc(x.usuario||'')}</div></div><div><strong>${esc(x.telefone||'')}</strong><div class="small muted">Telefone</div></div><div><strong>${esc((x.server_time||x.data||'').replace('T',' ').slice(0,16))}</strong><div class="small muted">Data</div></div><div><button class="btn danger" onclick="removerCobrancaIdx(${realIdx})">Remover</button></div></div>`}).join('')+pager:'<div class="empty">Nenhuma cobrança encontrada para esse filtro.</div>'}
function applyLogFilter(){const q=(document.getElementById('logQ')?.value||'').toLowerCase(); const deVal=document.getElementById('logDe')?.value||''; const ateVal=document.getElementById('logAte')?.value||''; const fil=document.getElementById('logFil')?.value||''; const vend=document.getElementById('logVend')?.value||''; let arr=[...(COB_LOGS||[]).filter(isLogCobrancaReal)].reverse(); arr=arr.filter(x=>{const txt=`${x.cliente||''} ${x.titulo||''} ${x.parcela||''}`.toLowerCase(); const raw=x.server_time||x.criado_em||x.data||''; if(q && !txt.includes(q)) return false; if(fil && String(x.filial||'')!==fil) return false; if(vend && String(x.usuario||'')!==vend) return false; if(deVal && dateOnlyISO(raw)<deVal) return false; if(ateVal && dateOnlyISO(raw)>ateVal) return false; return true}); _logFiltered=arr; _logPage=1; renderLogPage()}
function clearLogFilter(){['logQ','logDe','logAte'].forEach(id=>{const e=document.getElementById(id); if(e) e.value=''}); ['logFil','logVend'].forEach(id=>{const e=document.getElementById(id); if(e) e.value=''}); applyLogFilter()}
function removerCobrancaIdx(i){const x=_logFiltered?.[i]; if(!x){toast('Cobrança não encontrada.');return;} removerCobranca(x.id||'',x.cliente||'',x.titulo||'',x.parcela||'')}

function msgMatchesUser(m){
  if(!usuarioAtual) return false;
  const targetType=String(m.target_type||'all');
  const targetId=String(m.target_id||'');
  if(usuarioAtual.tipo==='master') return true;
  if(targetType==='all') return true;
  if(targetType==='filial' && String(usuarioAtual.filial||'')===targetId) return true;
  if(targetType==='user'){
    const targetLabel=String(m.target_label||'');
    const keys=[String(usuarioAtual.login||''), String(usuarioAtual.nome||''), `${usuarioAtual.nome||''}_${usuarioAtual.filial||''}`, `${usuarioAtual.nome||''} (${usuarioAtual.filial||''})`];
    if(keys.includes(targetId) || keys.includes(targetLabel)) return true;
    const cleanTarget=normName(targetId.replace(/_F\d+$/i,'').replace(/\([^)]*\)/g,''));
    const cleanLabel=normName(targetLabel.replace(/\([^)]*\)/g,''));
    const userName=normName(usuarioAtual.nome||'');
    const userLogin=String(usuarioAtual.login||'').toLowerCase();
    if(userLogin && String(targetId||'').toLowerCase()===userLogin) return true;
    return !!(userName && (cleanTarget===userName || cleanLabel===userName || cleanTarget.includes(userName) || cleanLabel.includes(userName) || userName.includes(cleanTarget) || userName.includes(cleanLabel)));
  }
  return false;
}
function isCampaign(m){
  if((m.message_kind||'')==='campaign') return true;
  if((m.kind||'')==='campaign') return true;
  return false;
}
function isExpiredCampaign(m){
  if(!isCampaign(m) || !m.expires_at) return false;
  const d=new Date(m.expires_at);
  if(isNaN(d)) return false;
  const end=new Date(d); end.setHours(23,59,59,999);
  return end < new Date();
}
function isReadMsg(m){
  const readBy=Array.isArray(m.read_by)?m.read_by.map(String):[];
  const readNorm=new Set(readBy.concat(readBy.map(x=>normName(x))));
  return currentUserKeys().some(k=>readNorm.has(String(k)) || readNorm.has(normName(k)));
}

function localMsgsKey(){return 'mdl_msgs_local_cache_v40'}
function getLocalMsgs(){try{return JSON.parse(localStorage.getItem(localMsgsKey())||'[]')}catch(e){return []}}
function saveLocalMsgs(arr){try{localStorage.setItem(localMsgsKey(),JSON.stringify(arr||[]))}catch(e){}}
function addMensagemLocal(obj){
  const now=new Date();
  const msg={id:'LOCAL_MSG_'+Date.now(),server_time:now.toISOString(),read_by:[],hidden_on_master:false,_local:true,...obj};
  const arr=getLocalMsgs(); arr.push(msg); saveLocalMsgs(arr); return msg;
}

function allMsgRecipientUsers(m){
  const targetType=String(m?.target_type||'all');
  const targetId=String(m?.target_id||'');
  const users=[];
  const add=(u)=>{if(!u) return; const login=String(u.login||''); const nome=String(u.nome||u.label||login); const filial=String(u.filial||''); const key=login || `${nome}_${filial}`; if(!key || ['master','diretorcomercial'].includes(key.toLowerCase())) return; if(!users.some(x=>String(x.key)===String(key))) users.push({key,nome,filial});};
  Object.values((AUTH_STATE&&AUTH_STATE.users)||{}).forEach(u=>add(u));
  flattenVendedores().forEach(v=>add({login:`${v.nome}_${v.filial}`,nome:v.nome,filial:v.filial}));
  if(targetType==='all') return users;
  if(targetType==='filial') return users.filter(u=>String(u.filial||'')===targetId);
  if(targetType==='user') return users.filter(u=>String(u.key||'')===targetId || String(u.login||'')===targetId || `${u.nome}_${u.filial}`===targetId || normName(u.nome||'')===normName(targetId.replace(/_F\d+$/,'')));
  return users;
}
function naoLidosMsg(m){
  const raw=Array.isArray(m?.read_by)?m.read_by.map(String):[];
  const read=new Set(raw.concat(raw.map(x=>normName(x))));
  const aliasesFor=(u)=>[u.key,u.login,`${u.nome}_${u.filial}`,`${u.nome} (${u.filial})`,u.nome,normName(u.nome||''),normName(`${u.nome}_${u.filial}`)].filter(Boolean).map(String);
  return allMsgRecipientUsers(m).filter(u=>!aliasesFor(u).some(k=>read.has(k)||read.has(normName(k))));
}
function mostrarNaoLidosAviso(id){
  const m=(MSGS||[]).find(x=>String(x.id)===String(id));
  if(!m){alert('Aviso não encontrado.'); return;}
  const arr=naoLidosMsg(m);
  alert(arr.length?('Ainda não leram:\n\n'+arr.map(u=>`${u.nome}${u.filial?` (${u.filial})`:''}`).join('\n')):'Todos os destinatários já leram este aviso/campanha.');
}

async function carregarMsgsOnline(){
  let remote=[];
  try{
    const r=await fetchComTimeout(API_MSG+'?_='+Date.now(),{},3000);
    const txt=await r.text(); let j={ok:false}; try{j=JSON.parse(txt);}catch(e){}
    remote=(j.ok&&Array.isArray(j.data))?j.data:[]; if(!j.ok && txt) console.log('mensagens_api retorno:', txt);
  }catch(e){console.log(e); remote=[]}
  const local=getLocalMsgs();
  const seen=new Set();
  MSGS=[...remote,...local].filter(m=>{const id=String(m.id||''); if(id && seen.has(id)) return false; if(id) seen.add(id); return true;});
  refreshBell();
}

function refreshBell(){const count=(MSGS||[]).filter(m=>msgMatchesUser(m) && !m.hidden_on_master && !isExpiredCampaign(m) && !isReadMsg(m) && !isCampaign(m)).length; const bell=document.getElementById('bellCount'); if(bell) bell.textContent=count;}
function renderMsgCard(m, showRemove=false, showClear=false, compact=false){
  let media = '';
  if(m.media_url){
    if(String(m.media_type||'').startsWith('image')) media = compact ? `<div class="msg-media compact"><img src="${m.media_url}" alt=""></div>` : `<div class="msg-media"><img src="${m.media_url}" alt=""></div>`;
    else if(String(m.media_type||'').startsWith('video')) media = compact ? `<div class="msg-media compact"><video src="${m.media_url}" controls></video></div>` : `<div class="msg-media"><video src="${m.media_url}" controls></video></div>`;
    else if(String(m.media_type||'').startsWith('audio')) media = compact ? `<div class="msg-media compact"><audio src="${m.media_url}" controls></audio></div>` : `<div class="msg-media"><audio src="${m.media_url}" controls></audio></div>`;
    else media = `<div class="msg-media"><a href="${m.media_url}" target="_blank" class="btn soft">Abrir anexo</a></div>`;
  }
  const unreadList = (usuarioAtual?.tipo==='master') ? naoLidosMsg(m) : [];
  const masterRead = (usuarioAtual?.tipo==='master') ? (unreadList.length?`<span class="unread-chip">${unreadList.length} não lido(s)</span>`:`<span class="read-chip">Todos leram</span>`) : null;
  const status = masterRead || (isReadMsg(m) ? '<span class="read-chip">Lido</span>' : '<span class="unread-chip">Não lido</span>');
  const markBtn = (!isReadMsg(m) && usuarioAtual?.tipo!=='master') ? `<button class="btn soft" onclick="marcarMsgLida('${m.id||''}')">✔️ Marcar como lido</button>` : '';
  const unreadBtn = (usuarioAtual?.tipo==='master') ? `<button class="btn soft" onclick="mostrarNaoLidosAviso('${m.id||''}')">👀 Não lido por</button>` : '';
  const removeBtn = showRemove ? `<button class="btn danger" onclick="removerMensagem('${m.id||''}')">Remover</button>` : '';
  const clearBtn = showClear ? `<button class="btn soft" onclick="limparMensagemTela('${m.id||''}')">Limpar</button>` : '';
  const detailBtn = compact && m.media_url ? `<button class="btn soft" onclick="openMsgPreview('${m.id||''}')">🔎 Detalhes</button>` : '';
  const typeTag = isCampaign(m) ? '<span class="mini-chip" style="background:#fff7ed;border-color:#fdba74;color:#c2410c">Campanha</span>' : '<span class="mini-chip">Aviso</span>';
  return `<div class="msg-card ${isCampaign(m)?'campaign-banner':''}"><div class="msg-head"><div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap"><strong>${esc(m.title||'Aviso')}</strong>${typeTag}${status}</div><span class="small muted">${esc((m.server_time||'').replace('T',' ').slice(0,16))}</span></div><div class="small muted">Para: ${esc(m.target_label||m.target_type||'Todos')}${m.expires_at?` · Até ${esc(m.expires_at)}`:''}</div><div style="margin-top:8px;white-space:pre-wrap">${esc(m.body||'')}</div>${media}<div style="display:flex;gap:8px;flex-wrap:wrap;margin-top:10px">${detailBtn}${markBtn}${unreadBtn}${clearBtn}${removeBtn}</div></div>`;
}
function openBell(){
  const arr=(MSGS||[]).filter(m=>msgMatchesUser(m) && !isExpiredCampaign(m)).reverse();
  document.getElementById('bellList').innerHTML=arr.length?arr.map(m=>renderMsgCard(m,false,false,true)).join(''):'<div class="empty">Nenhum aviso para você no momento.</div>';
  document.getElementById('bellModal').classList.add('show')
}
function closeBell(){document.getElementById('bellModal').classList.remove('show')}
function renderCampaignStrip(){
  const arr=(MSGS||[]).filter(m=>msgMatchesUser(m) && isCampaign(m) && !isExpiredCampaign(m)).reverse();
  if(!arr.length) return '';
  return `<div class="campaign-banner highlight-pulse" style="background:linear-gradient(180deg,#f7f1df,#f2ead7);border:1px solid #e4d8b2;color:#1f2937"><div class="section-head" style="margin:0 0 8px"><div><h2 style="margin:0;font-size:22px">📣 Campanha ativa</h2><div class="hint">Mensagem destacada do Master</div></div><button class="btn soft" onclick="openBell()">Ver detalhes</button></div>${arr.map(m=>renderMsgCard(m,false,false,true)).join('')}</div>`;
}
function renderInboxBanner(){
  const arr=(MSGS||[]).filter(m=>msgMatchesUser(m) && !isExpiredCampaign(m));
  const campaigns=arr.filter(m=>isCampaign(m));
  const unread=arr.filter(m=>!isCampaign(m) && !isReadMsg(m)).reverse().slice(0,10);
  const pieces=[];
  if(campaigns.length) pieces.push(renderCampaignStrip());
  if(!unread.length){
    if(!pieces.length) return `<div class="glass panel msg-banner"><h3 style="margin:0 0 6px">🔔 Avisos do Master</h3><div class="small muted">Nenhum aviso novo no momento.</div></div>`;
    return pieces.join('');
  }
  pieces.push(`<div class="glass panel msg-banner highlight-pulse"><div class="section-head" style="margin:0 0 10px"><div><h2 style="font-size:20px;margin:0">🔔 Avisos do Master</h2><div class="hint">Atualizados online em tempo real.</div></div><button class="btn soft" onclick="openBell()">Ver todos</button></div>${unread.map(m=>renderMsgCard(m,false,false,true)).join('')}</div>`);
  return pieces.join('');
}
async function marcarMsgLida(id){
  const fd=new FormData(); fd.append('action','mark_read'); fd.append('id',id); fd.append('user_key',currentUserKey()); fd.append('user_keys', JSON.stringify(currentUserKeys()));
  try{
    const r=await fetch(API_MSG,{method:'POST',body:fd}); const j=await r.json();
    if(j.ok){await carregarMsgsOnline(); if(!detailScreen.classList.contains('hidden')){const titleEl=detailScreen.querySelector('.back-row h2'); const subEl=detailScreen.querySelector('.back-row .sub'); if(titleEl && subEl){}} openBell(); if(usuarioAtual?.tipo!=='master'){const ent=usuarioAtual.is_terceiro?findEntity({type:'terceiro',filial:'FTER',nome:COBRANCA10_NOME}):(usuarioAtual.is_crediarista?findEntity({type:'crediarista',filial:usuarioAtual.filial,login:usuarioAtual.login,nome:usuarioAtual.nome}):(usuarioAtual.is_gerente?findEntity({type:'filial',filial:usuarioAtual.filial}):findEntity({type:'vendedor',filial:usuarioAtual.filial,nome:usuarioAtual.nome}))); if(usuarioAtual?.is_terceiro){openThirdChargePanel()} else if(usuarioAtual?.is_crediarista){openCrediaristaPanel(usuarioAtual.login,usuarioAtual.filial,usuarioAtual.nome)} else if(ent) openEntity({type:ent.type,filial:ent.filial,nome:ent.nome});} }
    else {toast('Não consegui marcar como lido.')}
  }catch(e){
    if(String(id||'').startsWith('LOCAL_MSG_')){
      const arr=getLocalMsgs().map(m=>String(m.id)===String(id)?{...m,read_by:[...(Array.isArray(m.read_by)?m.read_by:[]),currentUserKey()]}:m);
      saveLocalMsgs(arr); await carregarMsgsOnline(); openBell(); refreshBell(); toast('Aviso local marcado como lido.','success');
    } else toast('Não consegui marcar como lido.');
  }
}
function targetOptionsMsg(){
  let o='<option value="all|ALL">Todos</option>';
  ORDEM.forEach(f=>o+=`<option value="filial|${f}">Filial ${f}</option>`);

  const added = new Set();

  const addUserOpt = (value, label)=>{
    const v = String(value||'').trim();
    const l = String(label||'').trim();
    if(!v || !l) return;
    const key = `${v}__${l}`.toLowerCase();
    if(added.has(key)) return;
    added.add(key);
    o += `<option value="user|${esc(v)}">${esc(l)}</option>`;
  };

  // vendedores normais
  flattenVendedores().forEach(v=>{
    addUserOpt(`${v.nome}_${v.filial}`, `${v.nome} (${v.filial})`);
  });

  // usuários especiais das credenciais online
  const users = Object.values((AUTH_STATE && AUTH_STATE.users) || {});
  users.forEach(u=>{
    const login = String((u && u.login) || '').trim();
    const nome = String((u && u.nome) || '').trim();
    const filial = String((u && u.filial) || '').trim();

    if(u && u.is_viewer){
      addUserOpt(login || 'painel', nome || 'Painel');
      return;
    }
    if(login === 'master' || login === 'diretorcomercial') return;

    if((u && u.is_crediarista) || (u && u.is_terceiro) || (u && u.only_cobranca)){
      addUserOpt(login || nome, filial ? `${nome} (${filial})` : nome);
      return;
    }

    if(login && nome){
      addUserOpt(login, filial ? `${nome} (${filial})` : nome);
    }
  });

  if(AUTH_STATE && AUTH_STATE.director){
    addUserOpt('diretorcomercial', (AUTH_STATE.director && AUTH_STATE.director.nome) || 'Diretor Comercial');
  }
  addUserOpt('master', 'Master');

  return o;
}
function _senhaDomKey(key){return String(key||'').replace(/[^a-zA-Z0-9_-]/g,'_')}
function senhaAtualUsuario(u, isDirector=false){
  if(isDirector){return String((AUTH_STATE?.director?.password)||u?.password||u?.initial_password||'');}
  return String(u?.password || u?.senha || u?.initial_password || '');
}
function toggleSenhaAtual(id){
  const el=document.getElementById(id);
  if(!el) return;
  el.type = el.type==='password' ? 'text' : 'password';
}
async function copiarSenhaAtual(id){
  const el=document.getElementById(id);
  if(!el) return;
  try{await navigator.clipboard.writeText(el.value||''); toast('Senha copiada.','success');}
  catch(e){el.focus(); el.select(); document.execCommand('copy'); toast('Senha copiada.','success');}
}
function renderSenhaCard(u, isDirector=false){
  const key=isDirector?'diretorcomercial':u.login;
  const dom=_senhaDomKey(key);
  const senhaAtual=senhaAtualUsuario(u,isDirector);
  const pend=(AUTH_STATE?.password_reset_requests||[]).filter(r=>String(r.login||'').toLowerCase()===String(key).toLowerCase() && String(r.status||'pendente')==='pendente').length;
  const keyJs=JSON.stringify(String(key));
  return `<div class="glass card" style="cursor:default;max-width:520px"><div class="title" style="min-height:auto">${esc(isDirector?'Diretor Comercial':u.nome)} ${!isDirector?`(${u.filial||''})`:''}</div><div class="legend-inline"><span><i class="dot" style="background:${u.must_change_password?'#f59e0b':'#22c55e'}"></i>${u.must_change_password?'Precisa trocar senha':'Senha ativa'}</span>${pend?`<span><i class="dot" style="background:#ef4444"></i>${pend} solicitação(ões)</span>`:''}</div><div class="input-card" style="margin-top:12px"><label>Login</label><strong style="font-family:DM Mono,monospace">${esc(key)}</strong><label style="margin-top:10px">Senha ativa atual</label><div class="senha-view-row"><input id="senha_atual_${dom}" type="password" readonly value="${esc(senhaAtual)}"><button class="btn soft btn-xs" type="button" onclick="toggleSenhaAtual('senha_atual_${dom}')">👁️</button><button class="btn soft btn-xs" type="button" onclick="copiarSenhaAtual('senha_atual_${dom}')">📋</button></div></div><div class="form-grid bonus" style="grid-template-columns:1.1fr .9fr;margin-top:12px"><div class="input-card"><label>Nova senha</label><input id="pwd_${key}" placeholder="Digite a nova senha"></div><div class="input-card"><label>Ações</label><div style="display:flex;gap:8px;flex-wrap:wrap"><button class="btn primary" type="button" onclick='adminSalvarSenha(${keyJs})'>💾 Salvar</button><button class="btn soft" type="button" onclick='adminMarcarTroca(${keyJs})'>🔁 Exigir troca</button></div></div></div>${pend?`<div class="note" style="margin-top:10px">Solicitação pendente de recuperação. <button class="btn soft" style="margin-left:8px" onclick='adminResolverReset(${keyJs})'>Resolver solicitação</button></div>`:''}<div id="pwd_msg_${key}" class="note" style="margin-top:8px"></div></div>`;
}
function renderSenhaRow(u, isDirector=false){
  const key=isDirector?'diretorcomercial':String(u.login||'');
  const dom=_senhaDomKey(key);
  const senhaAtual=senhaAtualUsuario(u,isDirector);
  const pend=(AUTH_STATE?.password_reset_requests||[]).filter(r=>String(r.login||'').toLowerCase()===String(key).toLowerCase() && String(r.status||'pendente')==='pendente').length;
  const keyJs=JSON.stringify(String(key));
  const nome=isDirector?'Diretor Comercial':String(u.nome||key);
  const tipo=isDirector?'Diretor':(u.is_crediarista?'Crediarista':(u.is_terceiro?'Cobrança terceiro':(u.is_gerente?'Gerente':'Vendedor')));
  return `<tr>
    <td><strong>${esc(nome)}</strong>${pend?`<div class="small" style="color:#ef4444">${pend} solicitação(ões)</div>`:''}</td>
    <td><strong style="font-family:DM Mono,monospace">${esc(key)}</strong></td>
    <td><div class="senha-nova-row"><input id="login_new_${dom}" placeholder="Novo login" value="${esc(key)}"><button class="btn soft btn-xs" type="button" onclick='adminAlterarLogin(${keyJs})'>✏️ Alterar</button></div></td>
    <td>${esc(u.filial||'')}</td>
    <td>${esc(tipo)}</td>
    <td><span class="status-dot"><i style="background:${u.must_change_password?'#f59e0b':'#22c55e'}"></i>${u.must_change_password?'Precisa trocar':'Ativa'}</span></td>
    <td><div class="senha-view-row"><input id="senha_atual_${dom}" type="password" readonly value="${esc(senhaAtual)}"><button class="btn soft btn-xs" type="button" onclick="toggleSenhaAtual('senha_atual_${dom}')">👁️</button><button class="btn soft btn-xs" type="button" onclick="copiarSenhaAtual('senha_atual_${dom}')">📋</button></div></td>
    <td><div class="senha-nova-row"><input id="pwd_${key}" placeholder="Nova senha"><button class="btn primary btn-xs" type="button" onclick='adminSalvarSenha(${keyJs})'>💾 Salvar</button></div><div id="pwd_msg_${key}" class="note" style="margin-top:4px"></div></td>
    <td><button class="btn soft btn-xs" type="button" onclick='adminMarcarTroca(${keyJs})'>🔁 Exigir troca</button>${pend?`<button class="btn soft btn-xs" type="button" onclick='adminResolverReset(${keyJs})'>Resolver</button>`:''}</td>
  </tr>`;
}


function _histDates(){return Object.keys(HIST_DASH?.dates||{}).sort().reverse()}
function _histMonths(){return Object.keys(HIST_DASH?.months_closed||{}).sort().reverse()}
function _histCurrentDate(){const el=document.getElementById('histDate'); const dates=_histDates(); return (el?.value||dates[0]||'')}
function _histCurrentMonth(){const el=document.getElementById('histMonth'); const months=_histMonths(); return (el?.value||months[0]||'')}
function _histEntityOptions(dateVal, scope){const d=HIST_DASH?.dates?.[dateVal]||{}; let opts='<option value="">Todos</option>'; if(scope==='vendedores'){Object.entries(d.vendedores||{}).sort((a,b)=>String(a[1].nome||'').localeCompare(String(b[1].nome||''),'pt-BR')).forEach(([k,v])=>{opts+=`<option value="${k}">${esc(v.nome)} (${v.filial})</option>`})} else if(scope==='filiais'){Object.entries(d.filiais||{}).forEach(([k,v])=>{opts+=`<option value="${k}">${esc(v.nome||k)}</option>`})} return opts}
function _histSalesDates(){return Object.keys(HIST_DASH?.sales_dates||{}).sort().reverse()}
function _histSalesMonths(){return Object.keys(HIST_DASH?.sales_months||{}).sort().reverse()}
function _histSalesCurrentDate(){return document.getElementById('histSalesDate')?.value || _histSalesDates()[0] || ''}
function _histSalesCurrentMonth(){return document.getElementById('histSalesMonth')?.value || _histSalesMonths()[0] || ''}
function _histSalesEntityOptions(dateVal, scope){const d=HIST_DASH?.sales_dates?.[dateVal]||{}; let opts='<option value="">Todos</option>'; if(scope==='vendedores'){Object.entries(d.vendedores||{}).sort((a,b)=>String(a[1].nome||'').localeCompare(String(b[1].nome||''),'pt-BR')).forEach(([k,v])=>{opts+=`<option value="${k}">${esc(v.nome)} (${v.filial||''})</option>`})} else if(scope==='filiais'){Object.entries(d.filiais||{}).forEach(([k,v])=>{opts+=`<option value="${k}">${esc(v.nome||k)}</option>`})} return opts}
function _histSalesMonthEntityOptions(monthVal, scope){const d=HIST_DASH?.sales_months?.[monthVal]||{}; let opts='<option value="">Todos</option>'; if(scope==='vendedores'){Object.entries(d.vendedores||{}).sort((a,b)=>String(a[1].nome||'').localeCompare(String(b[1].nome||''),'pt-BR')).forEach(([k,v])=>{opts+=`<option value="${k}">${esc(v.nome)} (${v.filial||''})</option>`})} else if(scope==='filiais'){Object.entries(d.filiais||{}).forEach(([k,v])=>{opts+=`<option value="${k}">${esc(v.nome||k)}</option>`})} return opts}
function updateHistSalesEntityFilter(){const dateVal=_histSalesCurrentDate(); const scope=document.getElementById('histSalesScope')?.value||'empresa'; const wrap=document.getElementById('histSalesEntityWrap'); const sel=document.getElementById('histSalesEntity'); if(!wrap||!sel) return; wrap.classList.toggle('hidden',!(scope==='vendedores'||scope==='filiais')); sel.innerHTML=_histSalesEntityOptions(dateVal, scope);}
function updateHistSalesMonthEntityFilter(){const monthVal=_histSalesCurrentMonth(); const scope=document.getElementById('histSalesMonthScope')?.value||'empresa'; const wrap=document.getElementById('histSalesMonthEntityWrap'); const sel=document.getElementById('histSalesMonthEntity'); if(!wrap||!sel) return; wrap.classList.toggle('hidden',!(scope==='vendedores'||scope==='filiais')); sel.innerHTML=_histSalesMonthEntityOptions(monthVal, scope);}
function renderSalesHistoricoTable(rows,title='💲 Histórico de vendas'){if(!rows.length) return `<div class="empty">Nenhum registro de vendas encontrado para o filtro escolhido.</div>`; return `<div class="glass panel sales-panel"><div class="section-head" style="margin:0 0 10px"><div><h2 style="font-size:18px">${title}</h2></div></div>${rows.map(r=>`<div class="log-row" style="margin-bottom:10px"><div><strong>${esc(r.nome||r.filial||'Empresa')}</strong><div class="small muted">${esc(r.filial||'Resumo')}</div></div><div><strong>${R(r.venda_realizado_total||0)}</strong><div class="small muted">Venda realizada</div></div><div><strong>${pct(r.venda_atingido_total||0)}</strong><div class="small muted">Venda atingida</div></div><div><strong>${R(r.servico_realizado_total||0)}</strong><div class="small muted">Serviço realizado</div></div><div><strong>${pct(r.servico_atingido_total||0)}</strong><div class="small muted">Serviço atingido</div></div></div>`).join('')}</div>`}
function renderHistoricoSalesResults(){const dateVal=_histSalesCurrentDate(); const scope=document.getElementById('histSalesScope')?.value||'empresa'; const entity=document.getElementById('histSalesEntity')?.value||''; const box=document.getElementById('histSalesResults'); const d=HIST_DASH?.sales_dates?.[dateVal]; if(!box) return; if(!d){box.innerHTML='<div class="empty">Nenhum histórico de vendas salvo para esta data.</div>'; return} let top=`<div class="kpis">${makeKpi('Venda realizada',R(d.empresa?.venda_realizado_total||0),'var(--orange)')}${makeKpi('Venda atingida',pct(d.empresa?.venda_atingido_total||0),'var(--orange)')}${makeKpi('Serviço realizado',R(d.empresa?.servico_realizado_total||0),'var(--orange)')}${makeKpi('Serviço atingido',pct(d.empresa?.servico_atingido_total||0),'var(--orange)')}</div>`; if(scope==='empresa'){box.innerHTML=top + renderSalesHistoricoTable([{nome:'Empresa',filial:'Resumo geral',...(d.empresa||{})}],'💲 Histórico diário de vendas da empresa'); return} const source = scope==='filiais' ? Object.entries(d.filiais||{}).map(([k,v])=>({...v,key:k})) : Object.entries(d.vendedores||{}).map(([k,v])=>({...v,key:k})); const rows=entity?source.filter(x=>x.key===entity):source; box.innerHTML=top + renderSalesHistoricoTable(rows,`💲 Histórico diário de vendas ${scope==='filiais'?'de filiais':'de vendedores'}`)}
function renderHistoricoSalesMonthResults(){const monthVal=_histSalesCurrentMonth(); const scope=document.getElementById('histSalesMonthScope')?.value||'empresa'; const entity=document.getElementById('histSalesMonthEntity')?.value||''; const box=document.getElementById('histSalesMonthResults'); const d=HIST_DASH?.sales_months?.[monthVal]; if(!box) return; if(!d){box.innerHTML='<div class="empty">Nenhum fechamento de vendas salvo para este mês.</div>'; return} let top=`<div class="kpis">${makeKpi('Venda realizada',R(d.empresa?.venda_realizado_total||0),'var(--orange)')}${makeKpi('Venda atingida',pct(d.empresa?.venda_atingido_total||0),'var(--orange)')}${makeKpi('Serviço realizado',R(d.empresa?.servico_realizado_total||0),'var(--orange)')}${makeKpi('Serviço atingido',pct(d.empresa?.servico_atingido_total||0),'var(--orange)')}</div>`; if(scope==='empresa'){box.innerHTML=top + renderSalesHistoricoTable([{nome:'Empresa',filial:'Resumo mensal',...(d.empresa||{})}],'🧡 Fechamento mensal de vendas'); return} const source = scope==='filiais' ? Object.entries(d.filiais||{}).map(([k,v])=>({...v,key:k})) : Object.entries(d.vendedores||{}).map(([k,v])=>({...v,key:k})); const rows=entity?source.filter(x=>x.key===entity):source; box.innerHTML=top + renderSalesHistoricoTable(rows,`🧡 Fechamento mensal de vendas ${scope==='filiais'?'de filiais':'de vendedores'}`)}
function _histMonthEntityOptions(monthVal, scope){const d=HIST_DASH?.months_closed?.[monthVal]||{}; let opts='<option value="">Todos</option>'; if(scope==='vendedores'){Object.entries(d.vendedores_finais||{}).sort((a,b)=>String(a[1].nome||'').localeCompare(String(b[1].nome||''),'pt-BR')).forEach(([k,v])=>{opts+=`<option value="${k}">${esc(v.nome)} (${v.filial})</option>`})} else if(scope==='filiais'){Object.entries(d.filiais_finais||{}).forEach(([k,v])=>{opts+=`<option value="${k}">${esc(v.nome||k)}</option>`})} return opts}

function mesAtualComissao(){const d=new Date(); return `${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,'0')}`}
function _histComMeses(){return Object.keys(HIST_COMISSAO?.months||{}).sort().reverse()}
function _comEntKey(ent){return `${ent.type||'ent'}::${ent.filial||''}::${ent.login||ent.nome||''}`}
function _recebResumo(ent){const r=getRecebimentos(ent); const sum=(fx)=>(r[fx]||[]).reduce((a,b)=>a+Number(b.pago||0),0); return {grave:sum('grave'),alerta:sum('alerta'),atencao:sum('atencao'),total:sum('grave')+sum('alerta')+sum('atencao'),qtd:(r.grave||[]).length+(r.alerta||[]).length+(r.atencao||[]).length}}

function snapKV(label,value,color=''){
  return `<div class="snap-kv"><div>${esc(label)}</div><strong style="${color?`color:${color}`:''}">${value}</strong></div>`;
}
function snapMetric(title,value,sub='',color=''){
  return `<div class="snap-card"><div class="snap-title">${esc(title)}</div><div class="snap-value" style="${color?`color:${color}`:''}">${value}</div>${sub?`<div class="snap-sub">${sub}</div>`:''}</div>`;
}
function snapMiniBox(title, perc, alvo, rec, color){
  return `<div class="snap-mini"><div class="snap-title">${title}</div><div class="snap-percent" style="color:${color}">${pct(perc||0)}</div><div class="snap-sub">Alvo: ${R(alvo||0)}</div><div class="snap-sub">Recebido: ${R(rec||0)}</div></div>`;
}
function snapSalesRows(ent, key, label){
  let rows=getSalesRows(ent,key);
  if(String(key||'').includes('servico') && servicosEntidadeTotal(ent)>0 && rows.length) rows=[rows[0]];
  if(!rows.length) return '';
  return `<div class="snap-box"><h3>${label}</h3>${rows.slice(0,1).map(r=>{
    const srv=serviceOfficialOverride(ent,key,r);
    const ating=srv ? srv.atingidoTotal : salesCell(r,['Atingido Total']);
    const atingPeriodo=srv ? srv.atingidoPeriodo : salesCell(r,['Atingido Período']);
    const realizadoTotal=srv ? srv.realizado : esc(salesCell(r,['Realizado (R$) Total','Realizado(R$) Total']));
    const realizadoPeriodo=srv ? srv.realizado : esc(salesCell(r,['Realizado (R$) Período','Realizado(R$) Período']));
    const proj=salesCell(r,['Projetado (R$)','Projetado(R$)']);
    const title=srv ? `${salesTitleForRow(ent,r,key)} · relatório serviços` : salesTitleForRow(ent,r,key);
    return `<div class="snap-sales-title">${esc(title)}</div><div class="snap-grid4">
      ${snapKV('Meta total',esc(salesCell(r,['Meta (R$) Total','Meta(R$) Total'])))}
      ${snapKV('Realizado total',realizadoTotal,'#34d399')}
      ${snapKV('Atingido total',esc(ating),'#f59e0b')}
      ${snapKV('Meta período',esc(salesCell(r,['Meta (R$) Período','Meta(R$) Período'])))}
      ${snapKV('Realizado período',realizadoPeriodo,'#34d399')}
      ${snapKV('Atingido período',esc(atingPeriodo),'#f59e0b')}
      ${snapKV('Projetado',esc(proj),'#fb7185')}
      ${snapKV(ent.type==='filial'?'Filial':'Vendedor',esc(salesCell(r, ent.type==='filial'?['Filial']:['Vendedor_2','Vendedor'])))} 
    </div>`;
  }).join('')}</div>`;
}
function snapServices(ent){
  const rows=servicosEntidade(ent).filter(x=>Number(x.real_total||0)>0);
  const total=rows.reduce((a,b)=>a+Number(b.real_total||0),0);
  return `<div class="snap-box"><h3>🛠️ Serviços por tipo</h3><div class="snap-sub">Relatório real de serviços · total oficial ${R(total)}</div><div class="snap-grid3">${rows.slice(0,6).map(r=>snapKV(String(r.servico||'Serviço').slice(0,36),R(r.real_total||0),'#60a5fa')).join('') || '<div class="snap-sub">Sem serviços localizados.</div>'}</div></div>`;
}
function snapCommission(ent){
  if(!(ent.type==='vendedor'||ent.type==='filial')) return '';
  let c={}; try{c=calcCommissionSummary(ent)||{}}catch(e){return ''}
  const totalLiberado = c.elegivelMercantil && c.elegivelServicos;
  const totalExibido = totalLiberado ? c.totalPrevisto : 0;
  const cell=(t,v,color='')=>snapKV(t, v, color);
  return `<div class="snap-box"><h3>💵 Comissionamento previsto</h3><div class="snap-grid3">
    ${cell('Faixa aplicada',esc(c.faixaTxt||'-'))}
    ${cell('% comissão mercantil',`${String(Number(c.comPerc||0).toFixed(2)).replace('.',',')}%`)}
    ${cell('% serviços',`${String(Number(c.servPct||0).toFixed(2)).replace('.',',')}%`)}
    ${cell('Comissão vendas',R(c.vendasComissao||0),'#34d399')}
    ${cell('Comissão serviços',R(c.servicosComissao||0),'#34d399')}
    ${cell('Comissão caminhão',R(c.caminhaoComissao||0),'#34d399')}
    ${cell('Bônus por meta',R(c.bonusMeta||0),'#fbbf24')}
    ${cell('Rentab 48%',R(c.rent48||0),'#fbbf24')}
    ${cell('Rentab 52,15%',R(c.rent52||0),'#fbbf24')}
    ${cell('Rentab 55,50%',R(c.rent55||0),'#fbbf24')}
    ${cell('Total previsto',R(totalExibido),'#fff')}
  </div><div class="snap-sub">Base mercantil bruta: ${R(c.vendaRealBruto||0)} · Caminhão abatido: ${R(c.camReal||0)} · Serviço: ${R(c.servReal||0)}.</div></div>`;
}
function snapshotEntityHTML(ent){
  try{
    if(ent && (ent.type==='crediarista'||ent.is_crediarista)){
      ent=findEntityBySnapshotRow({key:(typeof _comEntKey==='function'?_comEntKey(ent):`${ent.type||'ent'}::${ent.filial||''}::${ent.login||ent.nome||''}`),nome:ent.nome,filial:ent.filial}) || ent;
    }
    const meta=calcMeta(ent);
    const nome=ent.type==='filial'?filialLabel(ent.filial):esc(ent.nome||'');
    const sub=ent.type==='filial'?'Painel individual da filial':(ent.type==='crediarista'?'Painel de crediarista':'Painel individual do vendedor');
    const bonus=getBonus(meta.cfg,meta.geral);
    const salesBlocks = ent.type==='filial'
      ? [['venda_filial_meta','📈 Venda · Meta Filial'],['servico_filial_ouro_fob','🛠️ Serviço · Ouro / FOB'],['venda_filial_subgrupo_20k','🚚 Venda · Caminhão 20K']]
      : [['venda_filial_vendedor_meta','📈 Venda · Meta Vendedor'],['servico_filial_vendedor_ouro_fob','🛠️ Serviço · Ouro / FOB Vendedor'],['venda_vendedor_subgrupo_20k','🚚 Venda · Caminhão 20K']];
    const vendasHtml=(ent.type==='crediarista'||ent.is_crediarista)?'':salesBlocks.map(([k,l])=>snapSalesRows(ent,k,l)).join('');
    const bonusItems=[[50,meta.cfg.bonus_50||'-'],[75,meta.cfg.bonus_75||'-'],[85,meta.cfg.bonus_85||'-'],[100,meta.cfg.bonus_100||'-']];
    return `<div class="snap-sheet">
      <style>
      .snap-sheet{width:1180px;max-width:100%;margin:0 auto;background:#0d0f14;color:#f3f6ff;font-family:Inter,Arial,sans-serif;padding:22px;border-radius:22px}
      .snap-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:18px;border:1px solid rgba(255,255,255,.12);background:#161922;border-radius:18px;padding:16px 18px}
      .snap-head h1{font-size:28px;margin:0;color:#fff}.snap-head .snap-sub{margin-top:6px}
      .snap-two{display:grid;grid-template-columns:1.04fr .96fr;gap:18px;align-items:start}
      .snap-box,.snap-card,.snap-mini,.snap-kv{border:1px solid rgba(255,255,255,.12);background:#151821;border-radius:16px;padding:14px;box-shadow:0 8px 24px rgba(0,0,0,.22)}
      .snap-box{margin-bottom:16px}.snap-box h2,.snap-box h3{margin:0 0 12px 0;font-size:19px;color:#fff}
      .snap-topmeta{display:grid;grid-template-columns:210px 1fr;gap:18px;align-items:center}.snap-ring{display:flex;align-items:center;justify-content:center}
      .snap-grid2{display:grid;grid-template-columns:1fr 1fr;gap:10px}.snap-grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}.snap-grid4{display:grid;grid-template-columns:repeat(4,1fr);gap:10px}
      .snap-title{font-size:12px;text-transform:uppercase;letter-spacing:.08em;color:#8c96ab;font-weight:800}.snap-value{font-size:26px;font-weight:900;margin-top:6px}.snap-sub{color:#aeb7ca;font-size:13px;line-height:1.35}
      .snap-mini{text-align:center}.snap-percent{font-size:30px;font-weight:950;margin:8px 0}
      .snap-kv strong{display:block;font-size:18px;margin-top:6px;color:#fff}.snap-sales-title{font-weight:900;color:#fff;margin-bottom:10px}
      .snap-chart-bars{height:230px;display:flex;align-items:end;gap:36px;justify-content:center;padding:16px 10px}.snap-barwrap{text-align:center}.snap-bar{width:58px;border-radius:12px 12px 8px 8px;background:linear-gradient(180deg,rgba(255,255,255,.25),transparent),var(--c);box-shadow:0 0 22px color-mix(in srgb,var(--c),transparent 65%)}.snap-barlabel{font-size:12px;color:#c7d0e1;margin-top:8px}
      .snap-bonus{display:grid;gap:8px}.snap-bonus-item{display:flex;justify-content:space-between;gap:12px;border:1px solid rgba(255,255,255,.1);background:#10131a;border-radius:12px;padding:10px 12px;font-size:13px}.snap-bonus-item.active{outline:2px solid rgba(245,158,11,.45)}
      @media print{body{background:#0d0f14}.snap-sheet{width:1180px}}
      </style>
      <div class="snap-head"><div><h1>${nome}</h1><div class="snap-sub">${sub} · ${esc(ent.filial||'')} · ${mesAtualComissao()}</div></div><div class="snap-sub">🕒 ${esc(ULTIMA_ATUALIZACAO_DASHBOARD_BR||'')}</div></div>
      <div class="snap-two">
        <div>
          <div class="snap-box"><h2>🎯 Meta do mês · Não acumulativo</h2><div class="snap-topmeta"><div class="snap-ring">${renderPiggyBank(meta.geral)}</div><div><div class="snap-grid2">
            ${snapMetric('Pendente',R(ent.pendente||0),'','#ff5a6a')}
            ${snapMetric('Recebido',R(ent.pago||0),'','#34d399')}
            ${snapMetric('% da filial',pct(ent.perc_filial||100))}
            ${snapMetric('Configuração usada',`${Number(meta.cfg.grave_pct||0)}/${Number(meta.cfg.alerta_pct||0)}/${Number(meta.cfg.atencao_pct||0)}`)}
          </div><div style="height:12px"></div><div class="snap-sub">🔴 Grave alvo ${R(meta.grave.alvo)} · recebido ${R(meta.grave.rec)}<br>🟠 Alerta alvo ${R(meta.alerta.alvo)} · recebido ${R(meta.alerta.rec)}<br>🟡 Atenção alvo ${R(meta.atencao.alvo)} · recebido ${R(meta.atencao.rec)}</div></div></div><div style="height:14px"></div><div class="snap-grid4">
            ${snapMiniBox('Grave',meta.grave.perc,meta.grave.alvo,meta.grave.rec,'#ff5a6a')}
            ${snapMiniBox('Alerta',meta.alerta.perc,meta.alerta.alvo,meta.alerta.rec,'#fb923c')}
            ${snapMiniBox('Atenção',meta.atencao.perc,meta.atencao.alvo,meta.atencao.rec,'#facc15')}
            ${snapMiniBox('Meta geral',meta.geral,meta.grave.alvo+meta.alerta.alvo+meta.atencao.alvo,meta.grave.rec+meta.alerta.rec+meta.atencao.rec,'#60a5fa')}
          </div></div>
          <div class="snap-box"><h3>🌊 Gráfico Geral Contas a Receber</h3>${(()=>{const vals=[['Grave',meta.grave.pend,'#ff5a6a'],['Alerta',meta.alerta.pend,'#fb923c'],['Atenção',meta.atencao.pend,'#facc15'],['Recebido',Number(ent.pago||0),'#34d399']]; const max=Math.max(1,...vals.map(v=>v[1])); return `<div class="snap-chart-bars">${vals.map(v=>`<div class="snap-barwrap"><div class="snap-bar" style="--c:${v[2]};height:${Math.max(24,(v[1]/max)*190)}px"></div><div class="snap-barlabel">${v[0]}<br><strong>${R(v[1])}</strong></div></div>`).join('')}</div>`})()}</div>
          <div class="snap-box"><h3>🏆 Bônus e premiações · Não acumulativo</h3><div class="snap-bonus">${bonusItems.map(([p,t])=>`<div class="snap-bonus-item ${bonus?.perc===p?'active':''}"><strong>🎯 ${p}%</strong><span>${esc(t)}</span></div>`).join('')}</div></div>
        </div>
        <div>
          ${vendasHtml}
          ${snapServices(ent)}
          ${snapCommission(ent)}
        </div>
      </div>
    </div>`;
  }catch(e){console.log('Falha ao montar snapshot compacto',e); return '';}
}

function isUltimoDiaMes23(){
  const now=new Date(); const tomorrow=new Date(now.getFullYear(),now.getMonth(),now.getDate()+1);
  return tomorrow.getDate()===1 && now.getHours()>=23;
}

function snapshotComissaoEntidade(ent){
  const meta=calcMeta(ent); const rec=_recebResumo(ent); let c={};
  try{ if(ent.type==='vendedor'||ent.type==='filial') c=calcCommissionSummary(ent)||{}; }catch(e){c={erro:String(e)}}
  const totalPrev=Number(c.totalPrevisto||0);
  return {
    key:_comEntKey(ent), tipo:ent.type, nome:ent.nome||filialLabel(ent.filial), filial:ent.filial||'', login:ent.login||'',
    pendente:Number(ent.pendente||0), recebido:Number(ent.pago||rec.total||0), recebido_conciliado:Number(rec.total||0), qtd_recebidos:Number(rec.qtd||0),
    meta_geral:Number(meta.geral||0), grave_rec:Number(meta.grave?.rec||0), alerta_rec:Number(meta.alerta?.rec||0), atencao_rec:Number(meta.atencao?.rec||0),
    grave_alvo:Number(meta.grave?.alvo||0), alerta_alvo:Number(meta.alerta?.alvo||0), atencao_alvo:Number(meta.atencao?.alvo||0),
    venda_real:Number(c.vendaReal||0), servico_real:Number(c.servReal||0), caminhao_real:Number(c.camReal||0),
    faixa:String(c.faixaTxt||''), comissao_vendas:Number(c.vendasComissao||0), comissao_servicos:Number(c.servicosComissao||0), comissao_caminhao:Number(c.caminhaoComissao||0), bonus_meta:Number(c.bonusMeta||0), rent48:Number(c.rent48||0), rent52:Number(c.rent52||0), rent55:Number(c.rent55||0), total_previsto:totalPrev,
    elegivel_mercantil:Boolean(c.elegivelMercantil), elegivel_servicos:Boolean(c.elegivelServicos), html_individual:(()=>{try{return snapshotEntityHTML(ent)||''}catch(e){console.error('snapshot individual erro',e);return ''}})(), observacao:(ent.type==='crediarista'?'Crediarista: somente pagos conciliados após cobrança própria.':(ent.type==='terceiro'?'Cobrança terceiro.':''))
  };
}
function buildSnapshotComissionamentoMensal(month=mesAtualComissao()){
  const ents=[...flattenVendedores(),...flattenFiliais(),...crediaristaEntities(),thirdChargeEntity()].filter(Boolean);
  const entidades=ents.map(snapshotComissaoEntidade);
  const total=entidades.reduce((a,b)=>a+Number(b.total_previsto||0),0);
  return {month, gerado_em:new Date().toISOString(), versao_snapshot:'snapshot_visual_compacto_v2', atualizado_em_br:new Date().toLocaleString('pt-BR'), total_previsto:total, entidades};
}
async function salvarSnapshotComissionamentoMensal(auto=false){
  if(usuarioAtual?.tipo!=='master') return;
  const month=document.getElementById('histComMonthSave')?.value||mesAtualComissao();
  const payload=buildSnapshotComissionamentoMensal(month);
  const fd=new FormData(); fd.append('month',month); fd.append('payload',JSON.stringify(payload));
  try{
    const r=await fetch(API_COMIS,{method:'POST',body:fd}); const j=await r.json();
    if(j.ok){HIST_COMISSAO=j.data||{months:{}}; if(!auto) toast('✅ Histórico de comissionamento salvo.'); renderHistoricoComissaoResults();}
    else if(!auto) toast('⚠️ Não consegui salvar histórico de comissionamento.');
  }catch(e){console.log('Erro salvar comissionamento',e); if(!auto) toast('⚠️ Erro ao salvar histórico de comissionamento.');}
}
function renderComissionamentoHistoricoTable(rows){
  if(!rows.length) return '<div class="empty">Nenhum comissionamento salvo para este mês.</div>';
  return `<div class="glass panel"><div class="tableish">${rows.map(r=>`<div class="row-item"><div class="row-top"><div><div class="name">${esc(r.nome||'')}</div><div class="small muted">${esc(r.tipo||'')} · ${esc(r.filial||'')}</div></div><div><strong>${pct(r.meta_geral||0)}</strong><div class="small muted">Meta cobrança</div></div><div><strong>${R(r.recebido||0)}</strong><div class="small muted">Recebido</div></div><div><strong>${R(r.comissao_vendas||0)}</strong><div class="small muted">Vendas</div></div><div><strong>${R(r.comissao_servicos||0)}</strong><div class="small muted">Serviços</div></div><div><strong>${R(r.bonus_meta||0)}</strong><div class="small muted">Bônus</div></div><div><strong>${R(r.total_previsto||0)}</strong><div class="small muted">Total previsto</div></div></div>${r.observacao?`<div class="small muted" style="margin-top:6px">${esc(r.observacao)}</div>`:''}</div>`).join('')}</div></div>`;
}

function findEntityBySnapshotRow(row){
  if(!row) return null;
  const key=String(row.key||'');
  const nome=String(row.nome||'');
  const filial=String(row.filial||'');
  let buckets=[];
  try{buckets=[...buckets,...flattenVendedores()]}catch(e){}
  try{buckets=[...buckets,...flattenFiliais()]}catch(e){}
  try{buckets=[...buckets,...crediaristaEntities()]}catch(e){}
  try{buckets=[...buckets,thirdChargeEntity()]}catch(e){}
  const getKey=(e)=>{try{return (typeof _comEntKey==='function')?_comEntKey(e):`${e.type||'ent'}::${e.filial||''}::${e.login||e.nome||''}`}catch(_){return ''}};
  return buckets.find(e=>String(getKey(e))===key)
      || buckets.find(e=>normName(e.nome)===normName(nome) && String(e.filial||'')===filial)
      || buckets.find(e=>normName(e.nome)===normName(nome))
      || null;
}
function snapshotHtmlFallbackFromRow(row,month){
  try{
    const R2=(v)=>typeof R==='function'?R(Number(v||0)):('R$ '+Number(v||0).toFixed(2).replace('.',','));
    const pct2=(v)=>typeof pct==='function'?pct(Number(v||0)):(Number(v||0).toFixed(0)+'%');
    return `<div class="snap-sheet"><style>
      .snap-sheet{width:1120px;max-width:100%;margin:0 auto;background:#0d0f14;color:#f3f6ff;font-family:Inter,Arial,sans-serif;padding:22px;border-radius:22px}
      .snap-head{border:1px solid rgba(255,255,255,.12);background:#161922;border-radius:18px;padding:16px 18px;margin-bottom:16px}.snap-head h1{margin:0;font-size:28px;color:#fff}.snap-sub{color:#aeb7ca;font-size:13px;margin-top:6px}
      .snap-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}.snap-card{border:1px solid rgba(255,255,255,.12);background:#151821;border-radius:16px;padding:16px}.snap-title{font-size:12px;text-transform:uppercase;color:#8c96ab;font-weight:900;letter-spacing:.08em}.snap-value{font-size:25px;font-weight:950;margin-top:8px}.green{color:#34d399}.orange{color:#f59e0b}.blue{color:#60a5fa}.red{color:#fb7185}
      .snap-note{margin-top:14px;border:1px solid rgba(245,158,11,.35);background:rgba(245,158,11,.08);border-radius:16px;padding:14px;color:#fde68a;font-weight:800}
    </style><div class="snap-head"><h1>${esc(row.nome||'')}</h1><div class="snap-sub">${esc(row.tipo||'')} · ${esc(row.filial||'')} · fechamento ${esc(month||'')}</div></div><div class="snap-grid">
      <div class="snap-card"><div class="snap-title">Meta cobrança</div><div class="snap-value blue">${pct2(row.meta_geral||0)}</div></div>
      <div class="snap-card"><div class="snap-title">Recebido</div><div class="snap-value green">${R2(row.recebido||0)}</div></div>
      <div class="snap-card"><div class="snap-title">Vendas</div><div class="snap-value orange">${R2(row.comissao_vendas||0)}</div></div>
      <div class="snap-card"><div class="snap-title">Serviços</div><div class="snap-value orange">${R2(row.comissao_servicos||0)}</div></div>
      <div class="snap-card"><div class="snap-title">Bônus meta</div><div class="snap-value green">${R2(row.bonus_meta||0)}</div></div>
      <div class="snap-card"><div class="snap-title">Comissão caminhão</div><div class="snap-value orange">${R2(row.comissao_caminhao||0)}</div></div>
      <div class="snap-card"><div class="snap-title">Faixa</div><div class="snap-value blue">${esc(row.faixa||'-')}</div></div>
      <div class="snap-card"><div class="snap-title">Total previsto</div><div class="snap-value green">${R2(row.total_previsto||0)}</div></div>
    </div><div class="snap-note">Resumo reconstruído automaticamente porque a tela visual completa não estava salva neste fechamento. Clique em salvar fechamento novamente para gerar o modelo visual completo.</div></div>`;
  }catch(e){return `<div style="padding:20px;background:#111;color:#fff">Resumo indisponível: ${esc(String(e))}</div>`}
}

function getSnapshotHtmlFromSelection(){
  const month=document.getElementById('histComMonth')?.value || _histComMeses()[0] || mesAtualComissao();
  const key=document.getElementById('histComEntityView')?.value||'';
  const snap=HIST_COMISSAO?.months?.[month];
  const row=(snap?.entidades||[]).find(x=>String(x.key)===String(key));
  if(!row) return {ok:false,msg:'Registro não encontrado no histórico.'};

  let html=String(row.html_individual||'');
  let fonte='tela congelada salva no fechamento';
  if(!html){
    const ent=findEntityBySnapshotRow(row);
    if(ent){
      try{html=String(snapshotEntityHTML(ent)||'');}catch(e){console.error('snapshotEntityHTML falhou',e); html='';}
      fonte='gerada agora porque este fechamento antigo não tinha a tela salva';
    }
  }
  if(!html){
    html=snapshotHtmlFallbackFromRow(row,month);
    fonte='resumo reconstruído do histórico salvo';
  }
  return {ok:true,html,row,month,fonte};
}
function abrirTelaComissionamentoCongeladaPorSelect(){
  try{
    const data=getSnapshotHtmlFromSelection();
    if(!data.ok){toast(data.msg); return;}

    let viewer=document.getElementById('snapshotInlineViewer');
    if(!viewer){
      const box=document.getElementById('histComResults') || document.querySelector('.hist-results') || document.body;
      viewer=document.createElement('div');
      viewer.id='snapshotInlineViewer';
      box.prepend(viewer);
    }

    viewer.innerHTML=`<div class="snapshot-inline-toolbar">
      <div><strong>📌 Tela congelada</strong><div class="hint">${esc(data.row.nome||'')} · ${esc(data.row.filial||'')} · ${esc(data.month)} · ${esc(data.fonte)}</div></div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn primary" onclick="imprimirSnapshotInline()">🖨️ Imprimir / Salvar PDF</button>
        <button class="btn soft" onclick="abrirSnapshotNovaAba()">Abrir em nova aba</button>
        <button class="btn soft" onclick="fecharSnapshotInline()">Fechar</button>
      </div>
    </div><div class="snapshot-inline-body" id="snapshotInlineBody">${data.html}</div>`;
    viewer.classList.add('show');
    viewer.scrollIntoView({behavior:'smooth',block:'start'});
  }catch(e){
    console.error('Erro ao abrir tela congelada inline:', e);
    toast('Erro ao abrir tela congelada: '+(e && e.message ? e.message : 'erro desconhecido'));
  }
}
function fecharSnapshotInline(){
  const v=document.getElementById('snapshotInlineViewer');
  if(v){v.classList.remove('show');v.innerHTML='';}
}
function imprimirSnapshotInline(){
  try{
    const data=getSnapshotHtmlFromSelection();
    if(!data.ok){toast(data.msg);return;}
    const w=window.open('','_blank');
    if(!w){toast('Pop-up bloqueado pelo navegador.');return;}
    w.document.open();
    w.document.write(`<!doctype html><html><head><meta charset="utf-8"><title>Comissionamento</title><style>body{margin:0;background:#080a0f;color:#f4f6fb;font-family:Inter,Arial,sans-serif;padding:12px}@media print{body{padding:0}}</style></head><body>${data.html}<script>setTimeout(()=>window.print(),700)<\/script></body></html>`);
    w.document.close();
  }catch(e){console.error(e);toast('Erro ao imprimir tela congelada.');}
}
function abrirSnapshotNovaAba(){
  try{
    const data=getSnapshotHtmlFromSelection();
    if(!data.ok){toast(data.msg);return;}
    const w=window.open('','_blank');
    if(!w){toast('Pop-up bloqueado pelo navegador.');return;}
    w.document.open();
    w.document.write(`<!doctype html><html><head><meta charset="utf-8"><title>Comissionamento</title><style>body{margin:0;background:#080a0f;color:#f4f6fb;font-family:Inter,Arial,sans-serif;padding:12px}</style></head><body>${data.html}</body></html>`);
    w.document.close();
  }catch(e){console.error(e);toast('Erro ao abrir nova aba.');}
}

function renderHistoricoComissaoResults(){
  const months=_histComMeses(); const current=document.getElementById('histComMonth')?.value || months[0] || mesAtualComissao();
  const box=document.getElementById('histComResults'); if(!box) return;
  const snap=HIST_COMISSAO?.months?.[current];
  const saveInput=document.getElementById('histComMonthSave'); if(saveInput && !saveInput.value) saveInput.value=mesAtualComissao();
  if(!snap){box.innerHTML=`<div class="empty">Nenhum histórico salvo para ${esc(current)}. Clique em “Salvar fechamento do mês atual” para gravar o resultado visível agora.</div>`; return;}
  const rows=[...(snap.entidades||[])].sort((a,b)=>String(a.tipo).localeCompare(String(b.tipo),'pt-BR')||String(a.nome).localeCompare(String(b.nome),'pt-BR'));
  box.innerHTML=`<div class="kpis">${makeKpi('Mês',esc(snap.month||current),'var(--blue)')}${makeKpi('Total previsto',R(snap.total_previsto||0),'var(--green)')}${makeKpi('Entidades',String(rows.length),'var(--orange)')}${makeKpi('Salvo em',esc((snap.atualizado_em_br||snap.gerado_em||'').replace('T',' ').slice(0,19)),'var(--blue)')}</div>`+`<div class="glass panel"><div class="form-grid"><div class="input-card"><label>Ver tela congelada individual</label><div class="hint">Agora a tela congelada salva um resumo visual compacto, sem a lista enorme de recebimentos/cobranças. Para atualizar este mês com o novo modelo, clique em Salvar fechamento do mês atual novamente.</div><select id="histComEntityView">${rows.map(r=>`<option value="${esc(r.key||'')}">${esc(r.nome||'')} · ${esc(r.filial||'')}</option>`).join('')}</select></div><div style="display:flex;align-items:end"><button class="btn primary" onclick="abrirTelaComissionamentoCongeladaPorSelect()">Abrir tela congelada</button></div></div></div>`+renderComissionamentoHistoricoTable(rows);
}

function setHistMode(mode){window._histMode=mode; document.getElementById('histDailyPane')?.classList.toggle('hidden',mode!=='daily'); document.getElementById('histMonthPane')?.classList.toggle('hidden',mode!=='monthly'); document.getElementById('histSalesPane')?.classList.toggle('hidden',mode!=='sales'); document.getElementById('histThirdPane')?.classList.toggle('hidden',mode!=='third'); document.getElementById('histComPane')?.classList.toggle('hidden',mode!=='comissao'); document.getElementById('histTabDaily')?.classList.toggle('active',mode==='daily'); document.getElementById('histTabMonthly')?.classList.toggle('active',mode==='monthly'); document.getElementById('histTabSales')?.classList.toggle('active',mode==='sales'); document.getElementById('histTabThird')?.classList.toggle('active',mode==='third'); document.getElementById('histTabCom')?.classList.toggle('active',mode==='comissao'); if(mode==='daily'){updateHistEntityFilter(); renderHistoricoResults();} else if(mode==='monthly'){updateHistMonthEntityFilter(); renderHistoricoMonthResults();} else if(mode==='third'){renderHistoricoTerceiro();} else if(mode==='comissao'){renderHistoricoComissaoResults();} else {updateHistSalesEntityFilter(); updateHistSalesMonthEntityFilter(); renderHistoricoSalesResults(); renderHistoricoSalesMonthResults();}}
function updateHistEntityFilter(){const dateVal=_histCurrentDate(); const scope=document.getElementById('histScope')?.value||'empresa'; const wrap=document.getElementById('histEntityWrap'); const sel=document.getElementById('histEntity'); if(!wrap||!sel) return; wrap.classList.toggle('hidden',!(scope==='vendedores'||scope==='filiais')); sel.innerHTML=_histEntityOptions(dateVal, scope);}
function updateHistMonthEntityFilter(){const monthVal=_histCurrentMonth(); const scope=document.getElementById('histMonthScope')?.value||'empresa'; const wrap=document.getElementById('histMonthEntityWrap'); const sel=document.getElementById('histMonthEntity'); if(!wrap||!sel) return; wrap.classList.toggle('hidden',!(scope==='vendedores'||scope==='filiais')); sel.innerHTML=_histMonthEntityOptions(monthVal, scope);}
function renderHistoricoTable(rows, scope, title='📋 Histórico'){if(!rows.length) return `<div class="empty">Nenhum registro encontrado para o filtro escolhido.</div>`; return `<div class="glass panel"><div class="section-head" style="margin:0 0 10px"><div><h2 style="font-size:18px">${title}</h2></div></div>${rows.map(r=>`<div class="log-row" style="margin-bottom:10px"><div><strong>${esc(r.nome||r.filial||'Empresa')}</strong><div class="small muted">${esc(r.filial||'Resumo')}</div></div><div><strong>${R(r.pendente||0)}</strong><div class="small muted">Pendente</div></div><div><strong>${R(r.recebido||0)}</strong><div class="small muted">Recebido</div></div><div><strong>${pct(r.perc_meta||0)}</strong><div class="small muted">Meta</div></div><div><strong>${R(r.grave_alvo||0)} / ${R(r.alerta_alvo||0)} / ${R(r.atencao_alvo||0)}</strong><div class="small muted">Alvos G/A/At</div></div></div>`).join('')}</div>`}
function renderHistoricoResults(){const dateVal=_histCurrentDate(); const scope=document.getElementById('histScope')?.value||'empresa'; const entity=document.getElementById('histEntity')?.value||''; const box=document.getElementById('histResults'); const d=HIST_DASH?.dates?.[dateVal]; if(!box){return} if(!d){box.innerHTML='<div class="empty">Nenhum histórico salvo para esta data.</div>'; return} let top=`<div class="kpis">${makeKpi('Pendente do dia',R(d.empresa?.pendente||0),'var(--red)')}${makeKpi('Recebido do dia',R(d.empresa?.recebido||0),'var(--green)')}${makeKpi('Grave do dia',R(d.empresa?.grave||0),'var(--red)')}${makeKpi('Alerta do dia',R(d.empresa?.alerta||0),'var(--orange)')}</div><div class="glass panel" style="margin-bottom:14px"><div class="section-head" style="margin:0"><div><h2 style="font-size:18px">⚙️ Meta usada no dia</h2><div class="hint">Global: G ${Number(d.empresa?.config_global?.grave_pct||0)}% · A ${Number(d.empresa?.config_global?.alerta_pct||0)}% · At ${Number(d.empresa?.config_global?.atencao_pct||0)}% · Pesos ${Number(d.empresa?.config_global?.peso_grave||0)}/${Number(d.empresa?.config_global?.peso_alerta||0)}/${Number(d.empresa?.config_global?.peso_atencao||0)}</div></div><div class="small muted">${esc(dateVal)}</div></div></div>`; if(scope==='empresa'){box.innerHTML=top + renderHistoricoTable([{nome:'Empresa',filial:'Resumo geral',pendente:d.empresa?.pendente||0,recebido:d.empresa?.recebido||0,perc_meta:0,grave_alvo:0,alerta_alvo:0,atencao_alvo:0}], 'empresa','📋 Histórico diário da empresa'); return} const source = scope==='filiais' ? Object.entries(d.filiais||{}).map(([k,v])=>({...v,key:k})) : Object.entries(d.vendedores||{}).map(([k,v])=>({...v,key:k})); const rows=entity?source.filter(x=>x.key===entity):source; box.innerHTML=top + renderHistoricoTable(rows, scope, `📋 Histórico diário ${scope==='filiais'?'de filiais':'de vendedores'}`);}
function renderHistoricoMonthResults(){const monthVal=_histCurrentMonth(); const scope=document.getElementById('histMonthScope')?.value||'empresa'; const entity=document.getElementById('histMonthEntity')?.value||''; const box=document.getElementById('histMonthResults'); const d=HIST_DASH?.months_closed?.[monthVal]; if(!box){return} if(!d){box.innerHTML='<div class="empty">Nenhum fechamento mensal salvo para este mês.</div>'; return} const cfg=d.config_global_fechamento||{}; let top=`<div class="kpis">${makeKpi('Mês fechado',esc(monthVal),'var(--blue)')}${makeKpi('Último dia',esc(d.ultimo_dia_historico||'-'),'var(--blue)')}${makeKpi('Snapshot final',esc(d.snapshot_final_data||'-'),'var(--blue)')}${makeKpi('Meta mês',esc(d.meta_file||'-'),'var(--blue)')}</div><div class="glass panel" style="margin-bottom:14px"><div class="section-head" style="margin:0"><div><h2 style="font-size:18px">📦 Fechamento mensal travado</h2><div class="hint">Global no fechamento: G ${Number(cfg.grave_pct||0)}% · A ${Number(cfg.alerta_pct||0)}% · At ${Number(cfg.atencao_pct||0)}% · Pesos ${Number(cfg.peso_grave||0)}/${Number(cfg.peso_alerta||0)}/${Number(cfg.peso_atencao||0)}</div></div><div class="small muted">Fechado em ${esc((d.fechado_em||'').replace('T',' ').slice(0,16))}</div></div></div>`;
 if(scope==='empresa'){const e=d.empresa_final||{}; box.innerHTML=top+renderHistoricoTable([{nome:'Empresa',filial:'Resultado final do mês',pendente:e.pendente||0,recebido:e.recebido||0,perc_meta:e.perc_meta||0,grave_alvo:e.grave_alvo||0,alerta_alvo:e.alerta_alvo||0,atencao_alvo:e.atencao_alvo||0}], 'empresa','📋 Resultado final mensal da empresa'); return}
 const source=scope==='filiais'?Object.entries(d.filiais_finais||{}).map(([k,v])=>({...v,key:k})) : Object.entries(d.vendedores_finais||{}).map(([k,v])=>({...v,key:k}));
 const rows=entity?source.filter(x=>x.key===entity):source; box.innerHTML=top + renderHistoricoTable(rows, scope, `📋 Resultado final mensal ${scope==='filiais'?'de filiais':'de vendedores'}`);}
function renderHistoricoTerceiro(){const box=document.getElementById('histThirdResults'); if(!box) return; const logs=(COB_LOGS||[]).filter(x=>String(x.usuario||'').toLowerCase()===COBRANCA10_LOGIN || String(x.usuario||'').toLowerCase()===COBRANCA10_NOME.toLowerCase()).slice().reverse(); const ent=thirdChargeEntity(); const top=`<div class="kpis">${makeKpi('Títulos na carteira',String((CLIENTES_TERCEIRO?.grave?.length||0)+(CLIENTES_TERCEIRO?.alerta?.length||0)+(CLIENTES_TERCEIRO?.atencao?.length||0)),'var(--blue)')}${makeKpi('Pendente',R(ent.pendente||0),'var(--red)')}${makeKpi('Recebido',R(ent.pago||0),'var(--green)')}${makeKpi('Cobranças lançadas',String(logs.length),'var(--orange)')}</div>`; const comm=renderTerceiroCommission(ent); const rows=logs.length?logs.map(x=>`<div class="log-row"><div><strong>${esc(x.cliente||'')}</strong><div class="small muted">${esc(x.titulo||'')} · Parcela ${esc(x.parcela||'')}</div></div><div><strong>${R(x.pendente||0)}</strong><div class="small muted">${esc(x.filial||'')}</div></div><div><strong>${esc((x.server_time||'').replace('T',' ').slice(0,16))}</strong><div class="small muted">Data</div></div><div><strong>${esc(x.telefone||'')}</strong><div class="small muted">Telefone</div></div></div>`).join(''):'<div class="empty">Nenhuma cobrança do Cobrança10 encontrada.</div>'; box.innerHTML=top+comm+`<div class="glass panel"><div class="section-head"><div><h2 style="font-size:18px">🤝 Cobranças Terceiro</h2><div class="hint">Histórico apenas do usuário Cobrança10.</div></div></div><div class="logs-list">${rows}</div></div>`}
function setHistMode(mode){window._histMode=mode; document.getElementById('histDailyPane')?.classList.toggle('hidden',mode!=='daily'); document.getElementById('histMonthPane')?.classList.toggle('hidden',mode!=='monthly'); document.getElementById('histSalesPane')?.classList.toggle('hidden',mode!=='sales'); document.getElementById('histThirdPane')?.classList.toggle('hidden',mode!=='third'); document.getElementById('histTabDaily')?.classList.toggle('active',mode==='daily'); document.getElementById('histTabMonthly')?.classList.toggle('active',mode==='monthly'); document.getElementById('histTabSales')?.classList.toggle('active',mode==='sales'); document.getElementById('histTabThird')?.classList.toggle('active',mode==='third'); if(mode==='daily'){updateHistEntityFilter(); renderHistoricoResults();} else if(mode==='monthly'){updateHistMonthEntityFilter(); renderHistoricoMonthResults();} else if(mode==='sales'){updateHistSalesEntityFilter(); updateHistSalesMonthEntityFilter(); renderHistoricoSalesResults(); renderHistoricoSalesMonthResults();} else {renderHistoricoTerceiro();}}
function renderHistoricoTab(){const dates=_histDates(); const months=_histMonths(); const salesDates=_histSalesDates(); const salesMonths=_histSalesMonths(); histSection.innerHTML=`<div class="section-head"><div><h2>🗂️ Histórico</h2><div class="hint">Consulte o histórico diário, vendas e os fechamentos mensais travados do Master/Diretor.</div></div></div><div class="tabs" style="justify-content:flex-start;margin:0 0 14px"><button id="histTabDaily" class="tab active" onclick="setHistMode('daily')">📅 Diário</button><button id="histTabMonthly" class="tab" onclick="setHistMode('monthly')">📦 Fechamento mensal</button><button id="histTabSales" class="tab" onclick="setHistMode('sales')">🧡 Vendas</button><button id="histTabThird" class="tab" onclick="setHistMode('third')">🤝 Cobranças Terceiro</button></div><div id="histDailyPane"><div class="glass panel" style="margin-bottom:14px"><div class="search-row"><div class="input-card"><label>Data</label><select id="histDate" onchange="updateHistEntityFilter();renderHistoricoResults()">${dates.map(d=>`<option value="${d}">${d}</option>`).join('')}</select></div><div class="input-card"><label>Escopo</label><select id="histScope" onchange="updateHistEntityFilter();renderHistoricoResults()"><option value="empresa">Empresa</option><option value="filiais">Filiais</option><option value="vendedores">Vendedores</option></select></div><div id="histEntityWrap" class="input-card hidden"><label>Filtro</label><select id="histEntity" onchange="renderHistoricoResults()"><option value="">Todos</option></select></div></div></div><div id="histResults"></div></div><div id="histMonthPane" class="hidden"><div class="glass panel" style="margin-bottom:14px"><div class="search-row"><div class="input-card"><label>Mês fechado</label><select id="histMonth" onchange="updateHistMonthEntityFilter();renderHistoricoMonthResults()">${months.map(m=>`<option value="${m}">${m}</option>`).join('')}</select></div><div class="input-card"><label>Escopo</label><select id="histMonthScope" onchange="updateHistMonthEntityFilter();renderHistoricoMonthResults()"><option value="empresa">Empresa</option><option value="filiais">Filiais</option><option value="vendedores">Vendedores</option></select></div><div id="histMonthEntityWrap" class="input-card hidden"><label>Filtro</label><select id="histMonthEntity" onchange="renderHistoricoMonthResults()"><option value="">Todos</option></select></div></div></div><div id="histMonthResults"></div></div><div id="histSalesPane" class="hidden"><div class="glass panel" style="margin-bottom:14px"><div class="search-row"><div class="input-card"><label>Data de vendas</label><select id="histSalesDate" onchange="updateHistSalesEntityFilter();renderHistoricoSalesResults()">${salesDates.map(d=>`<option value="${d}">${d}</option>`).join('')}</select></div><div class="input-card"><label>Escopo</label><select id="histSalesScope" onchange="updateHistSalesEntityFilter();renderHistoricoSalesResults()"><option value="empresa">Empresa</option><option value="filiais">Filiais</option><option value="vendedores">Vendedores</option></select></div><div id="histSalesEntityWrap" class="input-card hidden"><label>Filtro</label><select id="histSalesEntity" onchange="renderHistoricoSalesResults()"><option value="">Todos</option></select></div></div></div><div id="histSalesResults"></div><div class="glass panel" style="margin:14px 0"><div class="search-row"><div class="input-card"><label>Mês de vendas</label><select id="histSalesMonth" onchange="updateHistSalesMonthEntityFilter();renderHistoricoSalesMonthResults()">${salesMonths.map(m=>`<option value="${m}">${m}</option>`).join('')}</select></div><div class="input-card"><label>Escopo</label><select id="histSalesMonthScope" onchange="updateHistSalesMonthEntityFilter();renderHistoricoSalesMonthResults()"><option value="empresa">Empresa</option><option value="filiais">Filiais</option><option value="vendedores">Vendedores</option></select></div><div id="histSalesMonthEntityWrap" class="input-card hidden"><label>Filtro</label><select id="histSalesMonthEntity" onchange="renderHistoricoSalesMonthResults()"><option value="">Todos</option></select></div></div></div><div id="histSalesMonthResults"></div></div><div id="histThirdPane" class="hidden"><div id="histThirdResults"></div></div>`; if(dates.length){document.getElementById('histDate').value=dates[0]} if(months.length){document.getElementById('histMonth').value=months[0]} if(salesDates.length){document.getElementById('histSalesDate').value=salesDates[0]} if(salesMonths.length){document.getElementById('histSalesMonth').value=salesMonths[0]} setHistMode(window._histMode||'daily');}

// ===== OVERRIDE HISTÓRICO COMISSIONAMENTO VISÍVEL =====

function setHistMode(mode){
  window._histMode=mode;
  document.getElementById('histDailyPane')?.classList.toggle('hidden',mode!=='daily');
  document.getElementById('histMonthPane')?.classList.toggle('hidden',mode!=='monthly');
  document.getElementById('histSalesPane')?.classList.toggle('hidden',mode!=='sales');
  document.getElementById('histThirdPane')?.classList.toggle('hidden',mode!=='third');
  document.getElementById('histComPane')?.classList.toggle('hidden',mode!=='comissao');

  document.getElementById('histTabDaily')?.classList.toggle('active',mode==='daily');
  document.getElementById('histTabMonthly')?.classList.toggle('active',mode==='monthly');
  document.getElementById('histTabSales')?.classList.toggle('active',mode==='sales');
  document.getElementById('histTabThird')?.classList.toggle('active',mode==='third');
  document.getElementById('histTabCom')?.classList.toggle('active',mode==='comissao');

  if(mode==='daily'){
    updateHistEntityFilter(); renderHistoricoResults();
  } else if(mode==='monthly'){
    updateHistMonthEntityFilter(); renderHistoricoMonthResults();
  } else if(mode==='third'){
    renderHistoricoTerceiro();
  } else if(mode==='comissao'){
    try{carregarHistoricoComissaoOnline().then(()=>renderHistoricoComissaoResults())}catch(e){renderHistoricoComissaoResults()}
  } else {
    updateHistSalesEntityFilter(); updateHistSalesMonthEntityFilter(); renderHistoricoSalesResults(); renderHistoricoSalesMonthResults();
  }
}


function renderHistoricoTab(){
  const dates=_histDates();
  const months=_histMonths();
  const salesDates=_histSalesDates();
  const salesMonths=_histSalesMonths();
  const comMonths=_histComMeses();
  const currentComMonth=mesAtualComissao();

  histSection.innerHTML=`
    <div class="section-head">
      <div>
        <h2>🗂️ Histórico</h2>
        <div class="hint">Consulte o histórico diário, vendas, fechamentos mensais e o histórico de comissionamento para folha.</div>
      </div>
    </div>

    <div class="tabs" style="justify-content:flex-start;margin:0 0 14px">
      <button id="histTabDaily" class="tab active" onclick="setHistMode('daily')">📅 Diário</button>
      <button id="histTabMonthly" class="tab" onclick="setHistMode('monthly')">📦 Fechamento mensal</button>
      <button id="histTabSales" class="tab" onclick="setHistMode('sales')">🧡 Vendas</button>
      <button id="histTabThird" class="tab" onclick="setHistMode('third')">🤝 Cobranças Terceiro</button>
      <button id="histTabCom" class="tab" onclick="setHistMode('comissao')">💰 Comissionamento</button>
    </div>

    <div id="histDailyPane">
      <div class="glass panel" style="margin-bottom:14px">
        <div class="search-row">
          <div class="input-card"><label>Data</label><select id="histDate" onchange="updateHistEntityFilter();renderHistoricoResults()">${dates.map(d=>`<option value="${d}">${d}</option>`).join('')}</select></div>
          <div class="input-card"><label>Escopo</label><select id="histScope" onchange="updateHistEntityFilter();renderHistoricoResults()"><option value="empresa">Empresa</option><option value="filiais">Filiais</option><option value="vendedores">Vendedores</option></select></div>
          <div id="histEntityWrap" class="input-card hidden"><label>Filtro</label><select id="histEntity" onchange="renderHistoricoResults()"><option value="">Todos</option></select></div>
        </div>
      </div>
      <div id="histResults"></div>
    </div>

    <div id="histMonthPane" class="hidden">
      <div class="glass panel" style="margin-bottom:14px">
        <div class="search-row">
          <div class="input-card"><label>Mês fechado</label><select id="histMonth" onchange="updateHistMonthEntityFilter();renderHistoricoMonthResults()">${months.map(m=>`<option value="${m}">${m}</option>`).join('')}</select></div>
          <div class="input-card"><label>Escopo</label><select id="histMonthScope" onchange="updateHistMonthEntityFilter();renderHistoricoMonthResults()"><option value="empresa">Empresa</option><option value="filiais">Filiais</option><option value="vendedores">Vendedores</option></select></div>
          <div id="histMonthEntityWrap" class="input-card hidden"><label>Filtro</label><select id="histMonthEntity" onchange="renderHistoricoMonthResults()"><option value="">Todos</option></select></div>
        </div>
      </div>
      <div id="histMonthResults"></div>
    </div>

    <div id="histSalesPane" class="hidden">
      <div class="glass panel" style="margin-bottom:14px">
        <div class="search-row">
          <div class="input-card"><label>Data de vendas</label><select id="histSalesDate" onchange="updateHistSalesEntityFilter();renderHistoricoSalesResults()">${salesDates.map(d=>`<option value="${d}">${d}</option>`).join('')}</select></div>
          <div class="input-card"><label>Escopo</label><select id="histSalesScope" onchange="updateHistSalesEntityFilter();renderHistoricoSalesResults()"><option value="empresa">Empresa</option><option value="filiais">Filiais</option><option value="vendedores">Vendedores</option></select></div>
          <div id="histSalesEntityWrap" class="input-card hidden"><label>Filtro</label><select id="histSalesEntity" onchange="renderHistoricoSalesResults()"><option value="">Todos</option></select></div>
        </div>
      </div>
      <div id="histSalesResults"></div>
      <div class="glass panel" style="margin:14px 0">
        <div class="search-row">
          <div class="input-card"><label>Mês de vendas</label><select id="histSalesMonth" onchange="updateHistSalesMonthEntityFilter();renderHistoricoSalesMonthResults()">${salesMonths.map(m=>`<option value="${m}">${m}</option>`).join('')}</select></div>
          <div class="input-card"><label>Escopo</label><select id="histSalesMonthScope" onchange="updateHistSalesMonthEntityFilter();renderHistoricoSalesMonthResults()"><option value="empresa">Empresa</option><option value="filiais">Filiais</option><option value="vendedores">Vendedores</option></select></div>
          <div id="histSalesMonthEntityWrap" class="input-card hidden"><label>Filtro</label><select id="histSalesMonthEntity" onchange="renderHistoricoSalesMonthResults()"><option value="">Todos</option></select></div>
        </div>
      </div>
      <div id="histSalesMonthResults"></div>
    </div>

    <div id="histThirdPane" class="hidden"><div id="histThirdResults"></div></div>

    <div id="histComPane" class="hidden">
      <div class="glass panel" style="margin-bottom:14px">
        <div class="section-head" style="margin:0 0 10px">
          <div>
            <h2 style="font-size:18px">💰 Histórico mensal de comissionamento</h2>
            <div class="hint">Use para fechar o mês e consultar depois os valores para a folha de pagamento.</div>
          </div>
        </div>
        <div class="search-row">
          <div class="input-card">
            <label>Mês para consultar</label>
            <select id="histComMonth" onchange="renderHistoricoComissaoResults()">
              ${(comMonths.length?comMonths:[currentComMonth]).map(m=>`<option value="${m}">${m}</option>`).join('')}
            </select>
          </div>
          <div class="input-card">
            <label>Mês para salvar/atualizar</label>
            <input id="histComMonthSave" value="${currentComMonth}" placeholder="AAAA-MM">
          </div>
          <div class="input-card" style="display:flex;align-items:end">
            <button class="btn" onclick="salvarSnapshotComissionamentoMensal(false)">💾 Salvar fechamento do mês atual</button>
          </div>
        </div>
      </div>
      <div id="histComResults"></div>
    </div>
  `;

  if(dates.length){document.getElementById('histDate').value=dates[0]}
  if(months.length){document.getElementById('histMonth').value=months[0]}
  if(salesDates.length){document.getElementById('histSalesDate').value=salesDates[0]}
  if(salesMonths.length){document.getElementById('histSalesMonth').value=salesMonths[0]}
  if(comMonths.length){document.getElementById('histComMonth').value=comMonths[0]}
  setHistMode(window._histMode||'daily');
}

function colabBool(u,k,def=true){return u && Object.prototype.hasOwnProperty.call(u,k) ? !!u[k] : def}
function colabStatusBadge(u){const st=String(u?.status_operacional||'ativo').toLowerCase(); return st==='inativo'?'<span class="mini-chip" style="background:#450a0a;color:#fecaca;border:1px solid #991b1b">Inativo</span>':'<span class="mini-chip" style="background:#052e16;color:#bbf7d0;border:1px solid #166534">Ativo</span>'}
function renderColaboradorStatusPanel(users){
  const normalUsers=(users||[]).filter(u=>u && !u.is_viewer);
  const options=normalUsers.map(u=>`<option value="${esc(u.login||'')}">${esc(u.nome||u.login||'')} ${u.filial?`- ${esc(u.filial)}`:''}</option>`).join('');
  return `<div class="glass panel" style="margin-bottom:14px;border-color:rgba(34,197,94,.28)">
    <div class="section-head" style="margin:0 0 10px"><div><h2 style="font-size:18px">👥 Status operacional dos colaboradores</h2><div class="hint">Use quando alguém sair, entrar ou trocar de função. Inativo não acessa e, na próxima geração, sai do rateio de cobrança. A data de saída é só histórico/controle interno. As flags controlam em quais murais/listas ele participa.</div></div></div>
    <div class="senhas-table-wrap"><table class="senhas-table"><thead><tr><th>Colaborador</th><th>Filial</th><th>Tipo</th><th>Status</th><th>Cobrança</th><th>Sem movimento</th><th>Aniversário</th><th>Murais</th><th>Saída</th><th>Substituto</th><th>Obs</th><th>Ações</th></tr></thead><tbody>${normalUsers.map(u=>{
      const login=String(u.login||'').toLowerCase(); const dom=_senhaDomKey(login);
      return `<tr>
        <td><strong>${esc(u.nome||login)}</strong><div class="small muted">${esc(login)}</div></td>
        <td>${esc(u.filial||'-')}</td>
        <td>${u.is_crediarista?'Crediarista':(u.is_terceiro?'Terceiro':(u.is_gerente?'Gerente':'Vendedor'))}</td>
        <td><select id="colab_status_${dom}" style="min-width:110px"><option value="ativo" ${String(u.status_operacional||'ativo')!=='inativo'?'selected':''}>Ativo</option><option value="inativo" ${String(u.status_operacional||'ativo')==='inativo'?'selected':''}>Inativo</option></select><div style="margin-top:6px">${colabStatusBadge(u)}</div></td>
        <td style="text-align:center"><input type="checkbox" id="colab_cob_${dom}" ${colabBool(u,'participa_cobrancas')?'checked':''}></td>
        <td style="text-align:center"><input type="checkbox" id="colab_mov_${dom}" ${colabBool(u,'participa_sem_movimento')?'checked':''}></td>
        <td style="text-align:center"><input type="checkbox" id="colab_ani_${dom}" ${colabBool(u,'participa_aniversariantes')?'checked':''}></td>
        <td style="text-align:center"><input type="checkbox" id="colab_mur_${dom}" ${colabBool(u,'participa_murais')?'checked':''}></td>
        <td><input id="colab_saida_${dom}" type="date" value="${esc(u.data_saida||'')}" style="min-width:130px"></td>
        <td><select id="colab_sub_${dom}" style="min-width:180px"><option value="">Sem substituto</option>${options}</select></td>
        <td><input id="colab_obs_${dom}" value="${esc(u.obs||'')}" placeholder="Ex: saiu, férias, troca filial" style="min-width:210px"></td>
        <td><button class="btn primary" onclick="adminSalvarStatusColaborador('${login}')">💾 Salvar</button><div id="colab_msg_${dom}" class="small muted" style="margin-top:6px"></div></td>
      </tr>`
    }).join('')}</tbody></table></div>
  </div>`;
}
async function adminSalvarStatusColaborador(login){
  const dom=_senhaDomKey(login); const msg=document.getElementById(`colab_msg_${dom}`);
  const fd=new FormData();
  fd.append('action','admin_update_user_status'); fd.append('login',login);
  fd.append('status',document.getElementById(`colab_status_${dom}`)?.value||'ativo');
  fd.append('participa_cobrancas',document.getElementById(`colab_cob_${dom}`)?.checked?'1':'0');
  fd.append('participa_sem_movimento',document.getElementById(`colab_mov_${dom}`)?.checked?'1':'0');
  fd.append('participa_aniversariantes',document.getElementById(`colab_ani_${dom}`)?.checked?'1':'0');
  fd.append('participa_murais',document.getElementById(`colab_mur_${dom}`)?.checked?'1':'0');
  fd.append('data_saida',document.getElementById(`colab_saida_${dom}`)?.value||'');
  fd.append('substituto',document.getElementById(`colab_sub_${dom}`)?.value||'');
  fd.append('obs',document.getElementById(`colab_obs_${dom}`)?.value||'');
  try{
    const r=await fetch(API_CRED,{method:'POST',body:fd}); const j=await r.json();
    if(j.ok){ if(msg) msg.textContent='✅ Salvo online. Rode o dashboard novamente para recalcular rateio/listas.'; await carregarCredenciaisOnline(); renderSenhasTab(); toast('Status salvo. Próxima execução recalcula rateio/listas.','success'); }
    else{ throw new Error(j.error||'erro'); }
  }catch(e){
    // fallback local para testar a tela abrindo o HTML pelo arquivo, sem FTP/API.
    const u=(AUTH_STATE?.users||{})[login];
    if(u){u.status_operacional=fd.get('status'); u.access_disabled=(u.status_operacional==='inativo'); ['participa_cobrancas','participa_sem_movimento','participa_aniversariantes','participa_murais'].forEach(k=>u[k]=fd.get(k)==='1'); u.data_saida=fd.get('data_saida'); u.substituto=fd.get('substituto'); u.obs=fd.get('obs');}
    localStorage.setItem('mdl_colab_status_teste_'+login, JSON.stringify(u||{}));
    if(msg) msg.textContent='🧪 Salvo só no navegador para teste local. Para recalcular rateio, precisa salvar no JSON/API e rodar o robô.';
    renderSenhasTab();
  }
}

function renderSenhasTab(){
  const users=Object.values(AUTH_STATE?.users||{}).sort((a,b)=>String(a.nome||'').localeCompare(String(b.nome||''),'pt-BR'));
  const reqs=[...(AUTH_STATE?.password_reset_requests||[])].reverse();
  const resets=reqs.filter(r=>String(r.status||'pendente')==='pendente');
  const resolved=reqs.filter(r=>String(r.status||'pendente')==='resolvido');
  senhasSection.innerHTML=`<div class="section-head"><div><h2>🔐 Gerenciar senhas</h2><div class="hint">Master e Diretor Comercial podem redefinir senhas online, exigir troca no primeiro acesso e visualizar solicitações.</div></div></div>
  <div class="glass panel" style="margin-bottom:14px;border-color:${acessoGeralBloqueado()?'rgba(239,68,68,.65)':'rgba(34,197,94,.35)'}">
    <div class="section-head" style="margin:0;gap:14px;align-items:center">
      <div>
        <h2 style="font-size:18px;margin:0">${acessoGeralBloqueado()?'🔒 Acessos bloqueados':'✅ Acessos liberados'}</h2>
        <div class="hint">${acessoGeralBloqueado()?esc(textoBloqueioAcesso()):'Todos os usuários podem acessar normalmente.'}</div>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn ${acessoGeralBloqueado()?'soft':'primary'}" onclick="adminSetAccessBlock(true)">🔒 Bloquear acessos</button>
        <button class="btn ${acessoGeralBloqueado()?'primary':'soft'}" onclick="adminSetAccessBlock(false)">🔓 Liberar acessos</button>
      </div>
    </div>
  </div>
  <div class="glass panel" style="margin-bottom:14px">
    <div class="section-head" style="margin:0 0 8px">
      <div><h2 style="font-size:18px">📩 Solicitações de recuperação</h2><div class="hint">Pedidos enviados pelo botão de recuperar senha.</div></div>
      <div style="display:flex;gap:8px;flex-wrap:wrap">
        <button class="btn soft" onclick="adminLimparHistoricoReset('all')">🧹 Limpar histórico</button>
      </div>
    </div>
    ${resets.length?resets.map(r=>`<div class="log-row" style="margin-bottom:8px">
      <div><div style="font-weight:900">${esc(r.login||'')}</div><div class="small muted">${esc(r.obs||'Sem observação')}</div></div>
      <div><strong>${esc((r.created_at||'').replace('T',' ').slice(0,16))}</strong><div class="small muted">Data</div></div>
      <div><span class="mini-chip" style="background:#fff7ed;color:#b45309;border:1px solid #fdba74">Pendente</span></div>
      <div><button class="btn primary" onclick="adminResolverReset('${'${'}String(r.login||'').replace(/'/g,"\\'")${'}'}')">Resolver</button></div>
    </div>`).join(''):'<div class="empty">Nenhuma solicitação pendente.</div>'}
    ${resolved.length?`<div style="margin-top:14px"><h4 style="margin:0 0 8px">Histórico resolvido</h4>${resolved.map(r=>`<div class="log-row" style="margin-bottom:8px;opacity:.9">
      <div><div style="font-weight:900">${esc(r.login||'')}</div><div class="small muted">${esc(r.obs||'Sem observação')}</div></div>
      <div><strong>${esc(((r.resolved_at||r.created_at||'').replace('T',' ').slice(0,16)))}</strong><div class="small muted">Resolvido em</div></div>
      <div><span class="mini-chip" style="background:#ecfdf5;color:#166534;border:1px solid #86efac">Resolvido</span></div>
      <div></div>
    </div>`).join('')}</div>`:''}
  </div>
  <div class="glass panel admin-accounts-line" style="margin-bottom:14px"><div class="section-head" style="margin:0 0 8px"><div><h2 style="font-size:18px">👑 Contas administrativas</h2><div class="hint">Contas administrativas em linha para caber melhor na tela.</div></div></div><div class="senhas-table-wrap"><table class="senhas-table"><thead><tr><th>Usuário</th><th>Login atual</th><th>Alterar login</th><th>Filial</th><th>Tipo</th><th>Status</th><th>Senha ativa atual</th><th>Nova senha</th><th>Ações</th></tr></thead><tbody>${renderSenhaRow(AUTH_STATE?.director||{login:'diretorcomercial',nome:'Diretor Comercial',must_change_password:true}, true)}</tbody></table></div></div>
  <div class="glass panel" style="margin-bottom:14px"><div class="section-head" style="margin:0 0 8px"><div><h2 style="font-size:18px">➕ Criar usuário de acesso</h2><div class="hint">Aqui você cria apenas o login/senha de acesso. Depois vá em Metas > Crediaristas configuráveis para vincular esse usuário à filial/base e ao percentual.</div></div></div><div class="form-grid bonus"><div class="input-card"><label>Login</label><input id="newUserLogin" placeholder="ex: crediaristaf08"></div><div class="input-card"><label>Nome</label><input id="newUserNome" placeholder="ex: CREDIARISTAF08"></div><div class="input-card"><label>Filial</label><input id="newUserFilial" placeholder="ex: F8"></div><div class="input-card"><label>Senha inicial</label><input id="newUserSenha" placeholder="mín. 4 caracteres"></div></div><div class="form-grid bonus" style="margin-top:10px"><div class="input-card"><label>Tipo</label><select id="newUserTipo"><option value="crediarista">Crediarista</option><option value="cobranca">Cobrança</option></select></div></div><div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:14px"><button class="btn primary" onclick="adminCriarUsuarioCobranca()">💾 Criar usuário</button></div><div id="newUserMsg" class="note" style="margin-top:10px"></div></div>
  ${renderColaboradorStatusPanel(users)}
  <div class="section-head"><div><h2>👥 Usuários do dashboard</h2><div class="hint">Visualização em linha para conferir login, senha ativa, status e alterar rapidamente.</div></div></div>
  <div class="senhas-table-wrap"><table class="senhas-table"><thead><tr><th>Usuário</th><th>Login atual</th><th>Alterar login</th><th>Filial</th><th>Tipo</th><th>Status</th><th>Senha ativa atual</th><th>Nova senha</th><th>Ações</th></tr></thead><tbody>${users.map(u=>renderSenhaRow(u,false)).join('')}</tbody></table></div>`;
}
async function adminSetAccessBlock(flag){
  try{
    const fd=new FormData();
    fd.append('action','admin_set_access_block');
    fd.append('blocked', flag ? '1' : '0');
    fd.append('reason', flag ? 'Sistema em atualização. Aguarde liberação pelo Master.' : '');
    const r=await fetch(API_CRED,{method:'POST',body:fd});
    const j=await r.json();
    if(j.ok){
      await carregarCredenciaisOnline();
      toast(flag?'Acessos bloqueados.':'Acessos liberados.','success');
      renderSenhasTab();
    }else{
      toast('Não consegui alterar o bloqueio de acessos.');
    }
  }catch(e){
    toast('Não consegui alterar o bloqueio de acessos.');
  }
}

async function adminAlterarLogin(login){
  const dom=_senhaDomKey(login);
  const box=document.getElementById(`pwd_msg_${login}`);
  const novo=(document.getElementById(`login_new_${dom}`)?.value||'').trim().toLowerCase();
  if(!novo || novo.length<3){if(box) box.textContent='Digite um login com pelo menos 3 caracteres.'; return;}
  if(!confirm(`Alterar login de ${login} para ${novo}?`)) return;
  try{
    const fd=new FormData(); fd.append('action','admin_change_login'); fd.append('old_login',login); fd.append('new_login',novo);
    const r=await fetch(API_CRED,{method:'POST',body:fd}); const j=await r.json();
    if(box) box.textContent=j.ok?'✅ Login atualizado online.':'⚠️ Não consegui alterar o login: '+(j.error||'');
    if(j.ok){await carregarCredenciaisOnline(); renderSenhasTab();}
  }catch(e){if(box) box.textContent='⚠️ Não consegui alterar o login.';}
}

async function adminSalvarSenha(login){const box=document.getElementById(`pwd_msg_${login}`); const senha=(document.getElementById(`pwd_${login}`)?.value||'').trim(); if(!senha || senha.length<4){if(box) box.textContent='Digite uma senha com pelo menos 4 caracteres.'; return;} try{const fd=new FormData(); fd.append('action','admin_set_password'); fd.append('login',login); fd.append('new_password',senha); fd.append('must_change_password','0'); const r=await fetch(API_CRED,{method:'POST',body:fd}); const j=await r.json(); if(box) box.textContent=j.ok?'✅ Senha atualizada online.':'⚠️ Não consegui atualizar a senha.'; if(j.ok){await carregarCredenciaisOnline(); renderSenhasTab();}}catch(e){if(box) box.textContent='⚠️ Não consegui atualizar a senha.';}}
async function adminMarcarTroca(login){const box=document.getElementById(`pwd_msg_${login}`); try{const fd=new FormData(); fd.append('action','admin_force_change'); fd.append('login',login); const r=await fetch(API_CRED,{method:'POST',body:fd}); const j=await r.json(); if(box) box.textContent=j.ok?'✅ Usuário marcado para trocar a senha no próximo acesso.':'⚠️ Não consegui marcar troca de senha.'; if(j.ok){await carregarCredenciaisOnline(); renderSenhasTab();}}catch(e){if(box) box.textContent='⚠️ Não consegui marcar troca de senha.';}}
async function adminResolverReset(login){try{const fd=new FormData(); fd.append('action','resolve_reset'); fd.append('login',login); const r=await fetch(API_CRED,{method:'POST',body:fd}); const j=await r.json(); if(j.ok){toast('Solicitação resolvida. Usuário terá que criar nova senha no próximo acesso.','success'); await carregarCredenciaisOnline(); renderSenhasTab();}else{toast('Não consegui resolver a solicitação.')}}catch(e){toast('Não consegui resolver a solicitação.')}}
async function adminLimparHistoricoReset(mode='resolved'){try{const fd=new FormData(); fd.append('action','clear_reset_history'); fd.append('mode',mode); const r=await fetch(API_CRED,{method:'POST',body:fd}); const j=await r.json(); if(j.ok){toast(mode==='all'?'Solicitações limpas da tela.':'Histórico de solicitações limpo.','success'); await carregarCredenciaisOnline(); renderSenhasTab();}else{toast('Não consegui limpar o histórico.')}}catch(e){toast('Não consegui limpar o histórico.')}}

async function adminCriarUsuarioCobranca(){
  const msg=document.getElementById('newUserMsg');
  const login=(document.getElementById('newUserLogin')?.value||'').trim().toLowerCase();
  const nome=(document.getElementById('newUserNome')?.value||'').trim();
  const filial=(document.getElementById('newUserFilial')?.value||'').trim().toUpperCase();
  const senha=(document.getElementById('newUserSenha')?.value||'').trim();
  const tipo=(document.getElementById('newUserTipo')?.value||'crediarista').trim();
  if(!login||!nome||!filial||!senha){ if(msg) msg.textContent='Preencha login, nome, filial e senha.'; return; }
  try{
    const fd=new FormData();
    fd.append('action','admin_create_user');
    fd.append('login',login); fd.append('nome',nome); fd.append('filial',filial); fd.append('password',senha); fd.append('tipo',tipo);
    const r=await fetch(API_CRED,{method:'POST',body:fd});
    const j=await r.json();
    if(msg) msg.textContent=j.ok?'✅ Usuário criado online.':'⚠️ Não consegui criar o usuário.';
    if(j.ok){ await carregarCredenciaisOnline(); renderSenhasTab(); }
  }catch(e){ if(msg) msg.textContent='⚠️ Não consegui criar o usuário.'; }
}


function reativacaoTemplateAtual(filial=""){filial=String(filial||"").toUpperCase(); const porFilial=CONFIG_META?.reativacao_msg_template_filiais||{}; if(filial && String(porFilial[filial]||"").trim()) return String(porFilial[filial]); return String(CONFIG_META?.reativacao_msg_template||DEFAULT_REATIVACAO_MSG_MDL)}
function primeiroNomeClienteJs(nome){return String(nome||'Cliente').trim().split(/\s+/)[0]||'Cliente'}
function montarMensagemReativacao(c){let tpl=reativacaoTemplateAtual(c?.filial||""); const dados={primeiro_nome:c.primeiro_nome||primeiroNomeClienteJs(c.cliente||''),nome:c.cliente||'',filial:c.filial||'',dias:String(c.dias_sem_movimento||''),ultimo_movimento:c.ultimo_movimento||''}; Object.entries(dados).forEach(([k,v])=>{tpl=tpl.replaceAll(`{${k}}`,v)}); return tpl;}
function reatUserKeyFromNome(nome,filial){return normName(nome)+'_'+String(filial||'').toUpperCase()}
function reativacaoDestinatariosFilial(filial){
  filial=String(filial||'').toUpperCase();
  const vends=flattenVendedores().filter(v=>String(v.filial||'').toUpperCase()===filial).map(v=>({tipo:'vendedor',nome:v.nome,filial,key:reatUserKeyFromNome(v.nome,filial),label:`${v.nome} (${filial})`}));
  const temGerente=!!CREDS && Object.values(CREDS).some(u=>u && u.is_gerente && String(u.filial||'').toUpperCase()===filial);
  const gerente=temGerente?[{tipo:'gerente',nome:`GERENTE ${filial}`,filial,key:`GERENTE_${filial}`,label:`Gerente ${filial}`}]:[];
  const arr=[...gerente,...vends].filter((x,i,a)=>a.findIndex(y=>y.key===x.key)===i);
  return arr.length?arr:[{tipo:'filial',nome:`FILIAL ${filial}`,filial,key:`GERENTE_${filial}`,label:`Filial ${filial}`}];
}
function hashStr(s){
  s=String(s||'');
  let h=0;
  for(let i=0;i<s.length;i++){
    h=((h<<5)-h)+s.charCodeAt(i);
    h|=0;
  }
  return h;
}
function reativacaoClienteKey(c){
  const filial=String(c?.filial||'').toUpperCase();
  const codigo=String(c?.codigo||'').replace(/\D/g,'');
  const nome=normName(c?.cliente||'');
  const ultimo=String(c?.ultimo_movimento||'');
  const cidade=normName(c?.cidade||'');
  return `REAT|${filial}|${codigo||nome}|${ultimo}|${cidade}`;
}
function reativacaoOwnerInfo(c){
  const filial=String(c.filial||'F1').toUpperCase(); const arr=reativacaoDestinatariosFilial(filial);
  const k=hashStr(String(c.codigo||'')+'|'+String(c.cliente||'')+'|'+filial);
  return arr[Math.abs(k)%arr.length];
}
function reativacaoOwnerKey(c){return reativacaoOwnerInfo(c).key}
function reativacaoCurrentKey(){
  if(!usuarioAtual || usuarioAtual.tipo==='master' || usuarioAtual.is_viewer) return '';
  const filial=String(usuarioAtual.filial||'').toUpperCase();
  if(usuarioAtual.is_gerente) return `GERENTE_${filial}`;
  return reatUserKeyFromNome(usuarioAtual.nome||usuarioAtual.login||'',filial);
}
function isReativacaoEnviadaHoje(c){
  const hoje=new Date().toISOString().slice(0,10);
  const rowKey=reativacaoClienteKey(c);
  const cod=String(c.codigo||'').replace(/\D/g,'');
  const nome=normName(c.cliente||'');
  const filialRow=String(c.filial||'').toUpperCase();
  return (COB_LOGS||[]).some(x=>{
    if(String(x.titulo||'').toUpperCase()!=='REATIVACAO') return false;
    if(String(x.filial||'').toUpperCase()!==filialRow) return false;
    if(String(x.server_time||x.data||x.created_at||x.criado_em||'').slice(0,10)!==hoje) return false;
    const logKey=String(x.cliente_key||x.cobranca_key||'');
    const parc=String(x.parcela||'');
    if(rowKey && (logKey===rowKey || parc.includes(rowKey))) return true;
    // fallback seguro para registros antigos: precisa bater cliente + código/telefone, nunca apenas owner_key.
    const nomeLog=normName(x.cliente||'');
    if(nome && nomeLog && nomeLog===nome){
      if(cod && parc.includes(cod)) return true;
      const tels=(c.telefones||[]).map(t=>String(t||'').replace(/\D/g,''));
      const telLog=String(x.telefone||'').replace(/\D/g,'');
      if(telLog && tels.includes(telLog)) return true;
      if(!cod && !telLog) return true;
    }
    return false;
  });
}
function reativacaoRowsPermitidas(){
  let rows=(CLIENTES_SEM_MOVIMENTO||[]).map((r,i)=>({...r,_idx:i,_owner:reativacaoOwnerInfo(r)}));
  const ck=reativacaoCurrentKey();
  if(ck) rows=rows.filter(r=>String(r._owner?.key||'')===ck);
  return rows;
}
function abrirWhatsReativacao(idx,tel){
  const c=(CLIENTES_SEM_MOVIMENTO||[])[idx]; if(!c) return;
  const num=String(tel||((c.telefones||[])[0]||'')).replace(/\D/g,'');
  const msg=montarMensagemReativacao(c);
  const owner=reativacaoOwnerInfo(c)||{};
  const clienteKey=reativacaoClienteKey(c);
  const entRef={type:owner.tipo||'reativacao',filial:c.filial,nome:owner.nome||owner.label||usuarioAtual?.nome||'',login:owner.login||''};
  registrarCobrancaOnline({cliente:c.cliente,titulo:'REATIVACAO',parcela:`CLIENTE_SEM_MOVIMENTO|${clienteKey}|${owner.key||''}`,cliente_key:clienteKey,cobranca_key:clienteKey,owner_key:owner.key||'',vencimento:c.ultimo_movimento||'',pendente:0,filial:c.filial,telefones:[num]}, entRef, num);
  window.open(`https://wa.me/${num}?text=${encodeURIComponent(msg)}`,'_blank');
  setTimeout(()=>carregarCobrancasOnline().then(()=>{ if(detailScreen && !detailScreen.classList.contains('hidden') && currentDetailRef){ openEntity(currentDetailRef); } else { renderReativacaoTab(); renderInicioTab(); } }),800)
}
let reatBuscaState='';
let reatBuscaTimer=null;
let reatFilialState='';
let reatPageState=1;
const REAT_PAGE_SIZE=20;
function reatBuscaChanged(v){reatBuscaState=String(v||''); reatPageState=1; clearTimeout(reatBuscaTimer); reatBuscaTimer=setTimeout(()=>renderReativacaoTab(),380)}
function reatFilialChanged(v){reatFilialState=String(v||''); reatPageState=1; renderReativacaoTab()}
function reatSetPage(p){reatPageState=Math.max(1,Number(p)||1); renderReativacaoTab();}
async function salvarMensagemReativacaoGlobal(){const el=document.getElementById('reatMsgTemplate'); CONFIG_META.reativacao_msg_template=el?el.value:reativacaoTemplateAtual(); try{const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})}); const j=await resp.json(); toast(j.ok?'Mensagem global de reativação salva.':'Não consegui salvar mensagem.',j.ok?'success':'warn')}catch(e){toast('Falha ao salvar mensagem.','warn')}}
async function salvarMensagemReativacaoFilial(){const f=String(document.getElementById('reatMsgFilial')?.value||'').toUpperCase(); const el=document.getElementById('reatMsgTemplateFilial'); if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn'); return} CONFIG_META.reativacao_msg_template_filiais=CONFIG_META.reativacao_msg_template_filiais||{}; CONFIG_META.reativacao_msg_template_filiais[f]=el?el.value:''; try{const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})}); const j=await resp.json(); toast(j.ok?`Mensagem da ${f} salva.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn')}catch(e){toast('Falha ao salvar mensagem por filial.','warn')}}
function trocarMensagemReativacaoFilial(f){const el=document.getElementById('reatMsgTemplateFilial'); if(el) el.value=reativacaoTemplateAtual(f)}
function reatInfoExtra(r){
  const qtd=Number(r?._reat_qtd_envios||0);
  const motivo=String(r?._reat_motivo||'');
  if(qtd>0){
    const ult=r._reat_ultimo_envio||''; const pri=r._reat_primeiro_envio||''; const dias=Number(r._reat_dias_desde_envio||0);
    if(motivo==='retorno_10d') return ` · ${qtd} envio(s) · último ${esc(ult)} · voltou após ${dias} dias`;
    return ` · ${qtd} envio(s) · último ${esc(ult)} · aguardar ${Number(r._reat_cooldown_dias||10)} dias`;
  }
  if(motivo==='pendente_sem_envio') return ' · pendente sem envio anterior';
  if(motivo==='novo') return ' · novo na base';
  return '';
}
function renderReativacaoTab(){
  const box=reativacaoSection; if(!box) return;
  try{
    const filial=String(reatFilialState||'');
    const q=String(reatBuscaState||'').toLowerCase();
    let rows=reativacaoRowsPermitidas();
    if(filial) rows=rows.filter(r=>String(r.filial||'')===filial);
    if(q) rows=rows.filter(r=>`${r.cliente||''} ${r.cidade||''} ${r.bairro||''} ${r._owner?.label||''}`.toLowerCase().includes(q));
    const filiais=mdlV91FiliaisFromRows((CLIENTES_SEM_MOVIMENTO_BASE&&CLIENTES_SEM_MOVIMENTO_BASE.length)?CLIENTES_SEM_MOVIMENTO_BASE:(CLIENTES_SEM_MOVIMENTO||[]));
    const enviadosHoje=rows.filter(isReativacaoEnviadaHoje).length;
    const totalPermitido=reativacaoRowsPermitidas().length;
    const porFilial=filiais.map(f=>`${f}: ${(CLIENTES_SEM_MOVIMENTO||[]).filter(r=>r.filial===f).length}`).join(' · ');
    const semBase=!(CLIENTES_SEM_MOVIMENTO||[]).length;
    const tituloLista=(usuarioAtual?.tipo==='master'||usuarioAtual?.is_viewer)?'Lista geral / filtro':'Minha lista de reativação';
    box.innerHTML=`<div class="section-head"><div><h2>🧡 Clientes sem movimento +45 dias <span class="note" style="color:#f59e0b">${esc(DASHBOARD_BUILD_VERSION)}</span></h2><div class="hint">Base do Sólidus para reativação por WhatsApp. ${esc(porFilial||'Nenhum XLS carregado ainda.')}</div></div></div>
    ${semBase?'<div class="glass panel" style="border-color:rgba(245,158,11,.35);margin-bottom:14px"><strong>⚠️ Sem base de clientes carregada</strong><div class="hint">Coloque ou baixe os XLS de Clientes sem Movimento na pasta do dashboard e rode o script novamente.</div></div>':''}
    <div class="kpis" style="margin-bottom:14px">
      ${makeKpi('Para acionar',String(CLIENTES_SEM_MOVIMENTO.length),'var(--amber-400)',`Novos ${Number(CLIENTES_SEM_MOVIMENTO_META.novos_total||0).toLocaleString('pt-BR')} · retorno 10d ${Number(CLIENTES_SEM_MOVIMENTO_META.retorno_10d_total||0).toLocaleString('pt-BR')} · base ${Number(CLIENTES_SEM_MOVIMENTO_META.base_total||0).toLocaleString('pt-BR')}`)}
      ${makeKpi(tituloLista,String(rows.length),'var(--blue)',usuarioAtual?.tipo==='master'?'Filtro atual':'Distribuída sem duplicar')}
      ${makeKpi('Enviados hoje',String(enviadosHoje),'var(--green)','Da lista exibida')}
      ${makeKpi('Minha base total',String(totalPermitido),'var(--orange)','Rateio automático por filial')}
    </div>
    <div class="glass panel" style="margin-bottom:14px">
      <div class="search-row" style="grid-template-columns:1.5fr 220px minmax(420px,1fr);align-items:stretch">
        <div class="input-card"><label>Buscar</label><input id="reatBusca" value="${esc(reatBuscaState)}" oninput="reatBuscaChanged(this.value)" placeholder="Cliente, bairro, cidade, responsável"></div>
        <div class="input-card"><label>Filial</label><select id="reatFilial" onchange="reatFilialChanged(this.value)"><option value="">Todas</option>${filiais.map(f=>`<option value="${esc(f)}" ${filial===f?'selected':''}>${esc(f)}</option>`).join('')}</select></div>
        <div class="input-card" style="min-width:420px"><label>Mensagem padrão global</label><textarea id="reatMsgTemplate" rows="6" style="min-height:135px;width:100%;resize:vertical" oninput="CONFIG_META.reativacao_msg_template=this.value">${esc(reativacaoTemplateAtual())}</textarea><div class="hint">Variáveis: {primeiro_nome}, {nome}, {filial}, {dias}, {ultimo_movimento}</div><button class="btn primary" style="margin-top:8px" onclick="salvarMensagemReativacaoGlobal()">Salvar mensagem global</button></div>
      </div>
      <div class="search-row" style="grid-template-columns:220px minmax(520px,1fr);align-items:stretch;margin-top:12px">
        <div class="input-card"><label>Mensagem por filial</label><select id="reatMsgFilial" onchange="trocarMensagemReativacaoFilial(this.value)">${filiais.map(f=>`<option value="${esc(f)}">${esc(f)}</option>`).join('')}</select></div>
        <div class="input-card"><label>Texto específico da filial selecionada</label><textarea id="reatMsgTemplateFilial" rows="5" style="min-height:120px;width:100%;resize:vertical">${esc(reativacaoTemplateAtual(filiais[0]||''))}</textarea><div class="hint">Se vazio, usa a mensagem global.</div><button class="btn primary" style="margin-top:8px" onclick="salvarMensagemReativacaoFilial()">Salvar mensagem desta filial</button></div>
      </div>
    </div>
    <div class="faixa-title atencao" style="margin-bottom:10px"><span>📋 ${esc(tituloLista)}</span><span>${rows.length} cliente(s) · ${enviadosHoje} enviado(s) hoje</span></div>
    <div class="logs-list">${(()=>{const total=rows.length; const maxPage=Math.max(1,Math.ceil(total/REAT_PAGE_SIZE)); if(reatPageState>maxPage) reatPageState=maxPage; if(reatPageState<1) reatPageState=1; const ini=(reatPageState-1)*REAT_PAGE_SIZE; const pageRows=rows.slice(ini,ini+REAT_PAGE_SIZE); const pager=total>REAT_PAGE_SIZE?`<div class="log-pager"><div><strong>${total}</strong> cliente(s) · mostrando ${ini+1}-${Math.min(ini+REAT_PAGE_SIZE,total)} · página ${reatPageState}/${maxPage}</div><div style="display:flex;gap:8px"><button class="btn soft" ${reatPageState<=1?'disabled':''} onclick="reatSetPage(${reatPageState-1})">⬅️ Anterior</button><button class="btn soft" ${reatPageState>=maxPage?'disabled':''} onclick="reatSetPage(${reatPageState+1})">Próxima ➡️</button></div></div>`:''; return pager + pageRows.map(r=>{const tels=(r.telefones||[]); const enviado=isReativacaoEnviadaHoje(r); return `<div class="log-row" style="grid-template-columns:1.45fr .85fr .9fr auto"><div><strong>${esc(r.cliente||'')}</strong><div class="small muted">${esc(r.filial||'')} · ${esc(r.cidade||'')} · ${Number(r.dias_sem_movimento||0)} dias sem comprar · último ${esc(r.ultimo_movimento||'')}${reatInfoExtra(r)}</div></div><div><strong>${esc(r._owner?.label||'')}</strong><div class="small muted">Responsável pelo envio</div></div><div><strong>${enviado?'✅ Enviado hoje':esc(tels.length+' WhatsApp(s)')}</strong><div class="small muted">${esc(tels.map(fmtTelBR).join(', '))}</div></div><div style="display:flex;gap:8px;flex-wrap:wrap">${tels.map(t=>`<button class="btn wa" ${enviado?'disabled style="opacity:.45"':''} onclick="abrirWhatsReativacao(${r._idx},'${esc(t)}')">Whats ${esc(fmtTelBR(t))}</button>`).join('')}</div></div>`}).join('') + pager})() || '<div class="empty">Nenhum cliente encontrado.</div>'}</div>`;
  }catch(e){console.error('Erro renderReativacaoTab',e); box.innerHTML=`<div class="glass panel" style="border-color:rgba(239,68,68,.35)"><strong>⚠️ Erro na aba Clientes sem movimento</strong><div class="hint">${esc(e.message||e)}</div></div>`;}
}

function reativacaoRowsParaEnt(ent){
  if(!ent) return [];
  const filial=String(ent.filial||'').toUpperCase();
  let rows=(CLIENTES_SEM_MOVIMENTO||[]).map((r,i)=>({...r,_idx:i,_owner:reativacaoOwnerInfo(r)}));
  if(ent.type==='filial') return rows.filter(r=>String(r.filial||'').toUpperCase()===filial);
  const key=reatUserKeyFromNome(ent.nome||ent.login||'',filial);
  return rows.filter(r=>String(r._owner?.key||'')===key);
}
function showReatPanel(id,mode){
  const pend=document.getElementById(id+'_pend'); const sent=document.getElementById(id+'_sent');
  const bp=document.getElementById(id+'_btn_pend'); const bs=document.getElementById(id+'_btn_sent');
  if(!pend||!sent) return;
  pend.style.display=mode==='sent'?'none':'block';
  sent.style.display=mode==='sent'?'block':'none';
  if(bp) bp.classList.toggle('active',mode!=='sent');
  if(bs) bs.classList.toggle('active',mode==='sent');
}
let reatPanelCounter=0;
function renderReativacaoEnt(ent){
  const rows=reativacaoRowsParaEnt(ent);
  const enviadosRows=rows.filter(isReativacaoEnviadaHoje);
  const pendentesRows=rows.filter(r=>!isReativacaoEnviadaHoje(r));
  const tipo=ent?.type==='filial'?'filial':'vendedor';
  const panelId='reat_panel_'+(++reatPanelCounter);
  const exportId='export_'+panelId;
  mdlRegisterExport(exportId, 'Clientes sem movimento - '+(ent?.nome||ent?.filial||'usuario'), [
    ...pendentesRows.map(r=>mdlReatExportRow(r,'Pendente')),
    ...enviadosRows.map(r=>mdlReatExportRow(r,'Enviado hoje'))
  ]);
  const rowHtml=(r,enviado=false)=>{const tels=(r.telefones||[]); return `<div class="log-row" style="grid-template-columns:1.5fr .7fr .8fr auto"><div><strong>${esc(r.cliente||'')}</strong><div class="small muted">${esc(r.filial||'')} · ${esc(r.cidade||'')} · ${Number(r.dias_sem_movimento||0)} dias sem comprar · último ${esc(r.ultimo_movimento||'')}</div></div><div><strong>${esc(r._owner?.label||'')}</strong><div class="small muted">Responsável</div></div><div><strong>${enviado?'✅ Enviado hoje':esc(tels.length+' WhatsApp(s)')}</strong><div class="small muted">${esc(tels.map(fmtTelBR).join(', '))}</div></div><div style="display:flex;gap:8px;flex-wrap:wrap">${tels.map(t=>`<button class="btn wa" ${enviado?'disabled style="opacity:.45"':''} onclick="abrirWhatsReativacao(${r._idx},'${esc(t)}')">Whats ${esc(fmtTelBR(t))}</button>`).join('')}</div></div>`};
  return `<div class="accordion"><div class="acc-head" onclick="toggleAcc(this)"><span>🧡 Clientes sem movimento para reativação</span><span style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">${mdlExportButtons(exportId)}<span class="acc-hint">${rows.length} cliente(s) · ${enviadosRows.length} enviado(s) hoje · clique para abrir</span></span></div><div class="acc-body">
    <div class="faixa-title atencao" style="margin-bottom:10px"><span>Lista individual de ${esc(tipo)}</span><span>sem duplicar com outros usuários</span></div>
    <div class="reat-tabs"><span id="${panelId}_btn_pend" class="reat-tab pending active" onclick="showReatPanel('${panelId}','pend')">📋 Pendentes: ${pendentesRows.length}</span><span id="${panelId}_btn_sent" class="reat-tab ok" onclick="showReatPanel('${panelId}','sent')">✅ Enviados hoje: ${enviadosRows.length}</span></div>
    <div id="${panelId}_pend" class="faixa-block"><div class="faixa-title alerta"><span>📋 Para enviar hoje</span><span>${pendentesRows.length} cliente(s)</span></div><div class="logs-list">${pendentesRows.slice(0,20).map(r=>rowHtml(r,false)).join('') || '<div class="empty">Nenhum cliente pendente para este usuário/filial.</div>'}</div></div>
    <div id="${panelId}_sent" class="faixa-block" style="display:none"><div class="faixa-title atencao"><span>✅ Enviados hoje</span><span>${enviadosRows.length} cliente(s)</span></div><div class="logs-list">${enviadosRows.slice(0,20).map(r=>rowHtml(r,true)).join('') || '<div class="empty">Nenhuma mensagem de reativação enviada hoje.</div>'}</div></div>
  </div></div>`;
}

function renderDuplicidadeCarteiraBanner(){const d=DUPLICIDADES_CARTEIRA||{}; const n=Number(d.total_conflitos||0); if(!n) return '<div class="glass panel" style="margin-bottom:14px;border-color:rgba(34,197,94,.22)"><strong>✅ Check anti-duplicidade</strong><div class="hint">Nenhum cliente/título duplicado entre responsáveis na carteira gerada.</div></div>'; const exemplos=(d.conflitos||[]).slice(0,5).map(c=>`<div class="sem-cobranca-chip"><i class="dot" style="background:#ef4444"></i><div>${esc(c.responsaveis?.[0]?.cliente||c.key||'Título')}<small>${(c.responsaveis||[]).map(r=>esc(`${r.tipo}:${r.responsavel}`)).join(' · ')}</small></div></div>`).join(''); return `<div class="glass panel" style="margin-bottom:14px;border-color:rgba(239,68,68,.35)"><div class="section-head" style="margin:0 0 8px"><div><h2 style="font-size:18px;margin:0">🚨 Duplicidade na carteira</h2><div class="hint">${n} conflito(s). Revise antes de disparar cobranças para evitar dois usuários cobrando o mesmo título.</div></div></div><div class="sem-cobrancas-grid">${exemplos}</div></div>`;}

function renderAvisosTab(){
  avisosSection.innerHTML=`<div class="section-head"><div><h2>📣 Central de avisos</h2><div class="hint">Envie mensagens, imagens, vídeos e áudios para um usuário, uma filial ou todos.</div></div></div>
  ${renderCampaignStrip()}
  <div class="meta-layout">
    <div class="glass panel">
      <div class="form-grid bonus" style="grid-template-columns:1fr 1fr">
        <div class="input-card"><label>Destino</label><select id="msgTarget">${targetOptionsMsg()}</select></div>
        <div class="input-card"><label>Título</label><input id="msgTitle" placeholder="Ex: Campanha do dia"></div>
      </div>
      <div class="form-grid bonus" style="grid-template-columns:1fr 1fr;margin-top:12px">
        <div class="input-card"><label>Tipo</label><select id="msgKind"><option value="notice">Aviso</option><option value="campaign">Campanha</option></select></div>
        <div class="input-card"><label>Duração da campanha (até)</label><input id="msgExpires" type="date"></div>
      </div>
      <div class="input-card" style="margin-top:12px"><label>Mensagem</label><input id="msgBody" placeholder="Digite a mensagem"></div>
      <div class="input-card" style="margin-top:12px"><label>Anexo (imagem, vídeo ou áudio)</label><input id="msgFile" type="file" accept="image/*,video/*,audio/*"></div>
      <div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:14px"><button class="btn primary" onclick="enviarMensagemOnline()">📨 Enviar agora</button></div>
      <div id="msgSendInfo" class="note" style="margin-top:10px"></div>
    </div>
    <div class="glass panel">
      <h3>🗂️ Histórico de avisos</h3>
      <div id="msgHistory"></div>
    </div>
  </div>`;
  renderMsgHistory();
}
function renderMsgHistory(){
  const host=document.getElementById('msgHistory'); if(!host) return;
  const q=(document.getElementById('msgDate')?.value||'');
  const arr=[...(MSGS||[])].reverse();
  const ativos=arr.filter(m=>!m.hidden_on_master && (!q || String(m.server_time||'').startsWith(q)));
  const hist=arr.filter(m=>m.hidden_on_master && (!q || String(m.server_time||'').startsWith(q)));
  host.innerHTML=`<div style="display:flex;gap:10px;align-items:end;flex-wrap:wrap;margin-bottom:10px"><div class="input-card" style="min-width:220px"><label>Filtrar por data</label><input id="msgDate" type="date" value="${q}"></div><button class="btn soft" onclick="renderMsgHistory()">Filtrar</button></div>${ativos.length?'<h4 style="margin:6px 0">Na tela</h4>'+ativos.map(m=>renderMsgCard(m,true,true)).join(''):'<div class="empty">Nenhum aviso ativo na tela.</div>'}${hist.length?'<h4 style="margin:16px 0 6px">Histórico</h4>'+hist.map(m=>renderMsgCard(m,true,false)).join(''):''}`;
}
async function enviarMensagemOnline(){
  const target=document.getElementById('msgTarget').value;
  const [target_type,target_id]=target.split('|');
  const title=document.getElementById('msgTitle').value||'Aviso';
  const body=document.getElementById('msgBody').value||'';
  const kind=document.getElementById('msgKind').value||'notice';
  const expires=document.getElementById('msgExpires').value||'';
  const file=document.getElementById('msgFile').files[0];
  const target_label=document.getElementById('msgTarget').selectedOptions[0]?.textContent || target;
  const fd=new FormData();
  fd.append('target_type',target_type); fd.append('target_id',target_id); fd.append('target_label', target_label);
  fd.append('title',title); fd.append('body',body); fd.append('message_kind', kind); if(expires) fd.append('expires_at', expires);
  if(file) fd.append('media', file);
  try{
    const r=await fetch(API_MSG,{method:'POST',body:fd});
    const j=await r.json();
    if(!j.ok) throw new Error('api_not_ok');
    document.getElementById('msgSendInfo').textContent='✅ Aviso enviado online com sucesso.';
    document.getElementById('msgTitle').value=''; document.getElementById('msgBody').value=''; document.getElementById('msgFile').value=''; document.getElementById('msgExpires').value='';
    await carregarMsgsOnline(); renderMsgHistory(); renderList(); renderTopMural();
  }catch(e){
    addMensagemLocal({target_type,target_id,target_label,title,body,message_kind:kind,expires_at:expires});
    document.getElementById('msgSendInfo').textContent='✅ Aviso salvo localmente para teste. Em produção vai pelo FTP/API.';
    document.getElementById('msgTitle').value=''; document.getElementById('msgBody').value=''; document.getElementById('msgFile').value=''; document.getElementById('msgExpires').value='';
    await carregarMsgsOnline(); renderMsgHistory(); renderList(); renderTopMural();
  }
}
async function removerMensagem(id){
  if(!confirm('Remover esta mensagem?')) return;
  if(String(id||'').startsWith('LOCAL_MSG_')){saveLocalMsgs(getLocalMsgs().filter(m=>String(m.id)!==String(id))); await carregarMsgsOnline(); renderMsgHistory(); renderList(); renderTopMural(); toast('Mensagem local removida.','success'); return;}
  const fd=new FormData(); fd.append('action','delete'); fd.append('id',id);
  try{const r=await fetch(API_MSG,{method:'POST',body:fd}); const j=await r.json(); if(j.ok){toast('Mensagem removida.','success'); await carregarMsgsOnline(); renderMsgHistory(); renderList();}else{toast('Não consegui remover a mensagem.')}}catch(e){toast('Não consegui remover a mensagem.')}}
async function limparMensagemTela(id){
  if(String(id||'').startsWith('LOCAL_MSG_')){const arr=getLocalMsgs().map(m=>String(m.id)===String(id)?{...m,hidden_on_master:true}:m); saveLocalMsgs(arr); await carregarMsgsOnline(); renderMsgHistory(); renderList(); renderTopMural(); toast('Mensagem local enviada ao histórico.','success'); return;}
  const fd=new FormData(); fd.append('action','archive_master'); fd.append('id',id);
  try{const r=await fetch(API_MSG,{method:'POST',body:fd}); const j=await r.json(); if(j.ok){toast('Mensagem removida da tela e enviada ao histórico.','success'); await carregarMsgsOnline(); renderMsgHistory(); renderList();}else{toast('Não consegui limpar a mensagem.')}}catch(e){toast('Não consegui limpar a mensagem.')}}
async function openPrimeiroAcesso(prefillLogin=''){document.getElementById('faLogin').value=prefillLogin||document.getElementById('loginUser').value||''; document.getElementById('faCurrentPass').value=document.getElementById('loginPass').value||''; document.getElementById('faNewPass').value=''; document.getElementById('faNewPass2').value=''; document.getElementById('faMsg').textContent=''; document.getElementById('firstAccessModal').classList.add('show')}
function closeFirstAccess(){document.getElementById('firstAccessModal').classList.remove('show')}
function openRecuperarSenha(){document.getElementById('recLogin').value=document.getElementById('loginUser').value||''; document.getElementById('recObs').value=''; document.getElementById('recMsg').textContent=''; document.getElementById('recoverModal').classList.add('show')}
function closeRecover(){document.getElementById('recoverModal').classList.remove('show')}
async function enviarRecuperacaoSenha(){const login=(document.getElementById('recLogin').value||'').trim().toLowerCase(); const obs=(document.getElementById('recObs').value||'').trim(); const box=document.getElementById('recMsg'); if(!login){box.textContent='Informe o usuário.'; return;} try{const fd=new FormData(); fd.append('action','request_reset'); fd.append('login',login); fd.append('obs',obs); const r=await fetch(API_CRED,{method:'POST',body:fd}); const j=await r.json(); box.textContent=j.ok?'✅ Solicitação enviada ao Master.':'⚠️ Não consegui enviar a solicitação.';}catch(e){box.textContent='⚠️ Não consegui enviar a solicitação.';}}
async function salvarPrimeiroAcesso(){const login=(document.getElementById('faLogin').value||'').trim().toLowerCase(); const atual=(document.getElementById('faCurrentPass').value||'').trim(); const nova=(document.getElementById('faNewPass').value||'').trim(); const nova2=(document.getElementById('faNewPass2').value||'').trim(); const box=document.getElementById('faMsg'); box.textContent=''; if(!login||!atual||!nova||!nova2){box.textContent='Preencha todos os campos.'; return;} if(nova.length<4){box.textContent='A nova senha deve ter pelo menos 4 caracteres.'; return;} if(nova!==nova2){box.textContent='A confirmação da senha não confere.'; return;} const auth=getAuthUser(login); if(!auth || String(auth.password)!==atual){box.textContent='Usuário ou senha atual inválidos.'; return;} try{const fd=new FormData(); fd.append('action','change_password'); fd.append('login',login); fd.append('current_password',atual); fd.append('new_password',nova); const r=await fetch(API_CRED,{method:'POST',body:fd}); const j=await r.json(); box.textContent=j.ok?'✅ Senha alterada com sucesso. Entre com a nova senha.':'⚠️ Não consegui alterar a senha.'; if(j.ok){await carregarCredenciaisOnline(); document.getElementById('loginPass').value=''; setTimeout(()=>{closeFirstAccess();},700);}}catch(e){box.textContent='⚠️ Não consegui alterar a senha.';}}
async function fazerLogin(){
  const u=(document.getElementById('loginUser').value||'').trim().toLowerCase();
  const s=(document.getElementById('loginPass').value||'').trim();
  const msg=document.getElementById('loginMsg');
  msg.textContent='';
  if(!u || !s){msg.textContent='Informe usuário e senha.'; return;}

  // Master abre imediatamente, sem depender de API PHP online.
  if(u===LOGIN_MASTER.toLowerCase() && s===SENHA_MASTER){
    usuarioAtual={tipo:'master',nome:'Master',roleLabel:'Master'};
    saveSession();
    return abrirApp();
  }

  // Demais usuários tentam atualizar credenciais, mas com timeout curto.
  try{await carregarCredenciaisOnline();}catch(e){console.log('Login seguindo com credenciais embutidas',e);}

  // V23: bloqueio geral para manutenção/atualização.
  // Master e Diretor Comercial continuam com acesso para poder desbloquear.
  if(acessoGeralBloqueado() && u!==LOGIN_DIRETOR.toLowerCase()){
    msg.innerHTML='🔒 Acesso temporariamente bloqueado pelo Master.<br><small>'+esc(textoBloqueioAcesso())+'</small>';
    return;
  }

  if(u===LOGIN_DIRETOR.toLowerCase()){
    const authDir=getAuthUser(u);
    const senhaDir=authDir?.password || SENHA_DIRETOR;
    if(String(senhaDir)===s){
      if(authDir?.must_change_password){
        msg.textContent='Primeiro acesso do Diretor Comercial: defina uma nova senha.';
        return openPrimeiroAcesso(u);
      }
      usuarioAtual={tipo:'master',nome:'Diretor Comercial',roleLabel:'Diretor Comercial'};
      saveSession();
      return abrirApp();
    }
  }

  const auth=getAuthUser(u);
  if(CREDS[u] && auth && String(auth.password)===s){
    if(auth.access_disabled || auth.status_operacional==='inativo' || CREDS[u]?.access_disabled || CREDS[u]?.status_operacional==='inativo'){
      msg.innerHTML='🔒 Usuário inativo/bloqueado pelo Master.<br><small>Peça liberação ao responsável.</small>';
      return;
    }
    if(auth.must_change_password){
      msg.textContent='Primeiro acesso: defina sua nova senha.';
      return openPrimeiroAcesso(u);
    }
    usuarioAtual={tipo:'user',login:u,...CREDS[u]};
    saveSession();
    return abrirApp();
  }

  msg.textContent='Login ou senha inválidos.';
}
async function abrirApp(){
  try{document.body.classList.toggle('master-view', String(usuarioAtual?.tipo||'').toLowerCase()==='master'); document.body.classList.toggle('diretor-view', String(usuarioAtual?.tipo||'').toLowerCase()==='diretor');}catch(e){}
 loginScreen.classList.add('hidden'); app.classList.remove('hidden'); if(usuarioAtual.tipo==='master'){document.getElementById('kpis').classList.remove('hidden'); renderKPIs(); const isDiretor=usuarioAtual?.roleLabel==='Diretor Comercial'; userBadge.textContent=isDiretor?'👑 Diretor Comercial':'👑 Master'; masterTabs.classList.remove('hidden'); document.querySelectorAll('#masterTabs .tab').forEach(btn=>{const t=btn.dataset.tab; btn.classList.toggle('hidden', isDiretor && ['cobrancas','senhas'].includes(t));}); setMainTab('inicio')} else if(usuarioAtual.is_viewer){document.getElementById('kpis').classList.remove('hidden'); renderKPIs(); userBadge.textContent='📺 Painel'; masterTabs.classList.add('hidden'); mainFilters.classList.add('hidden'); listSection.classList.add('hidden'); metaSection.classList.add('hidden'); logSection.classList.add('hidden'); avisosSection.classList.add('hidden'); senhasSection.classList.add('hidden'); histSection.classList.add('hidden'); document.getElementById('mainScreen').classList.remove('hidden'); detailScreen.classList.add('hidden'); mainTab='inicio'; renderTopMural(); renderInicioTab();} else {document.getElementById('kpis').classList.add('hidden'); userBadge.textContent=usuarioAtual.is_terceiro?`🤝 ${usuarioAtual.nome}`:(usuarioAtual.is_crediarista?`🧾 ${usuarioAtual.nome}`:(usuarioAtual.is_gerente?`🏬 ${usuarioAtual.filial}`:`👤 ${usuarioAtual.nome}`)); masterTabs.classList.add('hidden'); mainFilters.classList.add('hidden'); const ent=usuarioAtual.is_terceiro?findEntity({type:'terceiro',filial:'FTER',nome:COBRANCA10_NOME}):(usuarioAtual.is_crediarista?findEntity({type:'crediarista',filial:usuarioAtual.filial,login:usuarioAtual.login,nome:usuarioAtual.nome}):(usuarioAtual.is_gerente?findEntity({type:'filial',filial:usuarioAtual.filial}):findEntity({type:'vendedor',filial:usuarioAtual.filial,nome:usuarioAtual.nome}))); document.getElementById('mainScreen').classList.add('hidden'); detailScreen.classList.remove('hidden'); if(usuarioAtual.is_terceiro){openThirdChargePanel()} else if(usuarioAtual.is_crediarista){openCrediaristaPanel(usuarioAtual.login,usuarioAtual.filial,usuarioAtual.nome)} else if(ent) openEntity({type:ent.type,filial:ent.filial,nome:ent.nome,login:ent.login}) }
  setTimeout(()=>{tentarAtualizarOnlineDepoisLogin();}, 80);
}
function logout(){clearSession(); location.reload()}
window.addEventListener('load',async ()=>{
  const u=document.getElementById('loginUser');
  const p=document.getElementById('loginPass');
  if(u) u.focus();
  if(u) u.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault(); if(p) p.focus();}});
  if(p) p.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault(); fazerLogin();}});
  await carregarCredenciaisOnline();
  try{
    const verDash = await fetchJsonNoCache('dashboard_version.json');
    const stampDash = String((verDash && (verDash.updated_at_label||verDash.updated_at)) || '');
    if(stampDash) window.__dashboardUpdatedAtLabel = stampDash;
  }catch(_e){}
  try{
    const verSales = await fetchJsonNoCache('sales_version.json');
    const stampSales = String((verSales && (verSales.updated_at_label||verSales.updated_at)) || '');
    if(stampSales) window.__salesUpdatedAtLabel = stampSales;
  }catch(_e){}
  if(restoreSession()){abrirApp();}
  setInterval(pollSalesLive,60000);
  setInterval(pollDashboardLiveReload,60000);
  setTimeout(pollSalesLive,3000);
  setTimeout(pollDashboardLiveReload,5000);
})

// ===== V4.1 HOTFIX: notificações individuais, histórico comissionamento atual e aniversariantes visível =====
function _entAtualUsuario(){
  try{
    if(currentDetailRef){ const e=findEntity(currentDetailRef); if(e) return e; }
    const login=String(usuarioAtual?.login||'').toLowerCase();
    const nome=normName(usuarioAtual?.nome||'');
    return flattenVendedores().find(e=>String(e.login||'').toLowerCase()===login || normName(e.nome||'')===nome) || null;
  }catch(e){return null;}
}
function _goalNotifsFor(ent){
  const arr=[]; if(!ent) return arr;
  try{ const meta=calcMeta(ent); if(Number(meta.geral||0)>=100) arr.push({k:'meta_cobranca',t:'Meta geral de recebimentos atingida',d:`Você chegou em ${pct(meta.geral||0)} da meta de cobrança.`}); }
  catch(e){}
  try{ const c=calcCommissionSummary(ent); if(c?.bonusLiberado) arr.push({k:'bonus_meta',t:'Bônus liberado',d:`Bônus previsto: ${R(c.bonusMeta||0)}.`}); if(c?.elegivelMercantil) arr.push({k:'meta_vendas',t:'Meta mercantil liberada',d:`Comissão de vendas liberada: ${R(c.vendasComissao||0)}.`}); if(c?.elegivelServicos) arr.push({k:'meta_servicos',t:'Meta de serviços liberada',d:`Comissão de serviços liberada: ${R(c.servicosComissao||0)}.`}); }
  catch(e){}
  try{ const b=calcularMetaDiariaBatida().find(x=>normName(x.nome||'')===normName(ent.nome||'') || (x.filial && x.filial===ent.filial && ent.type==='filial')); if(b) arr.push({k:'meta_diaria',t:'Meta diária BATIDA',d:`${b.nome||filialLabel(b.filial)} bateu a meta diária: ${R(b.realizado||0)} / ${R(b.meta_diaria||0)}.`}); }catch(e){}
  return arr;
}
function updateGoalNotifications(){
  const btn=document.getElementById('goalNotifBtn'), count=document.getElementById('goalNotifCount'); if(!btn||!count) return;
  const ent=_entAtualUsuario(); const arr=_goalNotifsFor(ent);
  if(!ent || isAdminLike() || !arr.length){btn.classList.add('hidden'); count.textContent='0'; return;}
  btn.classList.remove('hidden');
  const readKey='mdl_goal_notifs_read_'+(usuarioAtual?.login||usuarioAtual?.nome||'user')+'_'+mesAtualComissao();
  let read=[]; try{read=JSON.parse(localStorage.getItem(readKey)||'[]')}catch(e){}
  const unread=arr.filter(n=>!read.includes(n.k)); count.textContent=String(unread.length||arr.length);
  btn.classList.toggle('pulse-alert', unread.length>0);
}
function openGoalNotifications(){
  const ent=_entAtualUsuario(); const arr=_goalNotifsFor(ent); if(!arr.length){toast('Nenhuma notificação de meta no momento.'); return;}
  const readKey='mdl_goal_notifs_read_'+(usuarioAtual?.login||usuarioAtual?.nome||'user')+'_'+mesAtualComissao();
  try{localStorage.setItem(readKey, JSON.stringify(arr.map(n=>n.k)))}catch(e){}
  updateGoalNotifications();
  const html=`<div class="glass panel" style="max-width:760px;margin:40px auto"><div class="section-head"><div><h2>🎉 Notificações de metas</h2><div class="hint">Metas e bônus atingidos no mês.</div></div><button class="btn soft" onclick="this.closest('.goal-modal')?.remove()">Fechar</button></div><div class="tableish">${arr.map(n=>`<div class="row-item"><div class="row-top" style="grid-template-columns:72px 1fr"><img src="${LARANJITO}" style="width:58px;height:58px;object-fit:contain;border-radius:14px"><div><div class="name">${esc(n.t)}</div><div class="small muted">${esc(n.d)}</div></div></div></div>`).join('')}</div></div>`;
  const div=document.createElement('div'); div.className='goal-modal'; div.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.62);z-index:99999;padding:20px;overflow:auto'; div.innerHTML=html; document.body.appendChild(div);
}

// Histórico comissionamento: sempre permite gerar fechamento atual e abrir tela visual por usuário
function _allComissaoEntitiesNow(){
  let ents=[]; try{ents=ents.concat(flattenVendedores())}catch(e){} try{ents=ents.concat(flattenFiliais())}catch(e){} try{ents=ents.concat(crediaristaEntities())}catch(e){} try{ents.push(thirdChargeEntity())}catch(e){}
  return ents.filter(Boolean);
}
function _comKeyNow(e){try{return (typeof _comEntKey==='function')?_comEntKey(e):`${e.type||'ent'}::${e.filial||''}::${e.login||e.nome||''}`}catch(_){return `${e.type||'ent'}::${e.filial||''}::${e.login||e.nome||''}`}}
function abrirTelaComissionamentoAtual(){
  const key=document.getElementById('histComCurrentEntity')?.value||''; const ent=_allComissaoEntitiesNow().find(e=>_comKeyNow(e)===key) || _allComissaoEntitiesNow()[0];
  if(!ent){toast('Nenhum usuário encontrado.'); return;}
  const html=snapshotEntityHTML(ent)||'<div>Não foi possível montar tela congelada.</div>';
  const w=window.open('','_blank'); if(!w){toast('Pop-up bloqueado.');return}
  w.document.write(`<!doctype html><html><head><meta charset="utf-8"><title>Fechamento ${esc(ent.nome||ent.filial||'')}</title><style>body{margin:0;background:#080a0f;color:#f4f6fb;font-family:Inter,Arial,sans-serif;padding:12px}@media print{body{padding:0}}</style></head><body><div style="display:flex;gap:8px;margin:0 0 12px"><button onclick="window.print()">Salvar PDF / Imprimir</button></div>${html}</body></html>`); w.document.close();
}
function exportarComissaoAtualExcel(){
  const rows=_allComissaoEntitiesNow().map(e=>snapshotComissaoEntidade(e));
  const header=['Tipo','Nome','Filial','Login','Pendente','Recebido','Meta cobrança %','Venda mercantil','Serviços','Caminhão','Comissão vendas','Comissão serviços','Comissão caminhão','Bônus meta','Rentab 48','Rentab 52,15','Rentab 55,50','Total previsto'];
  const lines=[header].concat(rows.map(r=>[r.tipo,r.nome,r.filial,r.login,r.pendente,r.recebido,r.meta_geral,r.venda_real,r.servico_real,r.caminhao_real,r.comissao_vendas,r.comissao_servicos,r.comissao_caminhao,r.bonus_meta,r.rent48,r.rent52,r.rent55,r.total_previsto]));
  const csv=lines.map(row=>row.map(v=>'"'+String(v??'').replace(/"/g,'""')+'"').join(';')).join('\n');
  const blob=new Blob(['\ufeff'+csv],{type:'text/csv;charset=utf-8'}); const a=document.createElement('a'); a.href=URL.createObjectURL(blob); a.download='comissionamento_atual_'+mesAtualComissao()+'.csv'; a.click();
}
const _oldRenderHistoricoComissaoResults = typeof renderHistoricoComissaoResults==='function' ? renderHistoricoComissaoResults : null;
renderHistoricoComissaoResults=function(){
  const months=_histComMeses(); const current=document.getElementById('histComMonth')?.value || months[0] || mesAtualComissao();
  const box=document.getElementById('histComResults'); if(!box) return;
  const ents=_allComissaoEntitiesNow();
  const currentPanel=`<div class="glass panel" style="margin-bottom:14px"><div class="section-head"><div><h2 style="font-size:18px">🧊 Tela individual congelada atual</h2><div class="hint">Escolha qualquer usuário/filial/crediarista/terceiro e abra a tela visual para salvar em PDF. Também exporta o resultado geral para Excel/CSV.</div></div></div><div class="form-grid"><div class="input-card"><label>Usuário / filial</label><select id="histComCurrentEntity">${ents.map(e=>`<option value="${esc(_comKeyNow(e))}">${esc(e.nome||filialLabel(e.filial)||'Entidade')} · ${esc(e.filial||'')}</option>`).join('')}</select></div><div style="display:flex;align-items:end;gap:8px;flex-wrap:wrap"><button class="btn primary" onclick="abrirTelaComissionamentoAtual()">Abrir tela congelada</button><button class="btn soft" onclick="exportarComissaoAtualExcel()">Exportar Excel</button></div></div></div>`;
  const snap=HIST_COMISSAO?.months?.[current];
  if(!snap){box.innerHTML=currentPanel+`<div class="empty">Nenhum histórico salvo para ${esc(current)}. Use “Salvar fechamento do mês atual” para gravar no histórico.</div>`; return;}
  const rows=[...(snap.entidades||[])].sort((a,b)=>String(a.tipo).localeCompare(String(b.tipo),'pt-BR')||String(a.nome).localeCompare(String(b.nome),'pt-BR'));
  box.innerHTML=currentPanel+`<div class="kpis">${makeKpi('Mês',esc(snap.month||current),'var(--blue)')}${makeKpi('Total previsto',R(snap.total_previsto||0),'var(--green)')}${makeKpi('Entidades',String(rows.length),'var(--orange)')}${makeKpi('Salvo em',esc((snap.atualizado_em_br||snap.gerado_em||'').replace('T',' ').slice(0,19)),'var(--blue)')}</div>`+`<div class="glass panel"><div class="form-grid"><div class="input-card"><label>Ver tela congelada salva</label><select id="histComEntityView">${rows.map(r=>`<option value="${esc(r.key||'')}">${esc(r.nome||'')} · ${esc(r.filial||'')}</option>`).join('')}</select></div><div style="display:flex;align-items:end"><button class="btn primary" onclick="abrirTelaComissionamentoCongeladaPorSelect()">Abrir salva</button></div></div></div>`+renderComissionamentoHistoricoTable(rows);
};

// Botões extras na cobrança terceiro
const _oldRenderHistoricoTerceiro = typeof renderHistoricoTerceiro==='function' ? renderHistoricoTerceiro : null;
renderHistoricoTerceiro=function(){
  if(_oldRenderHistoricoTerceiro) _oldRenderHistoricoTerceiro();
  const box=document.getElementById('histThirdResults'); if(!box) return;
  box.insertAdjacentHTML('afterbegin', `<div class="glass panel" style="margin-bottom:14px"><button class="btn primary" onclick="window.print()">Salvar tela/PDF</button> <button class="btn soft" onclick="exportarComissaoAtualExcel()">Exportar comissões Excel</button></div>`);
};

// Reprocessa notificações após abrir painel individual
const _oldOpenEntity = typeof openEntity==='function' ? openEntity : null;
if(_oldOpenEntity){ openEntity=function(ref){ const ret=_oldOpenEntity(ref); setTimeout(updateGoalNotifications,600); return ret; } }


// ===== V4.3: hotfix real - aniversariantes, notificações sempre visíveis e botão voltar só admin =====
try{
  const st=document.createElement('style');
  st.textContent=`body.individual-view .back-row .btn{display:none!important} body.individual-view .back-row{grid-template-columns:1fr!important}`;
  document.head.appendChild(st);
}catch(e){}
function _setIndividualViewFlag(){try{document.body.classList.toggle('individual-view', !!usuarioAtual && !isAdminLike() && !usuarioAtual.is_viewer)}catch(e){}}
try{setInterval(_setIndividualViewFlag,1000)}catch(e){}

function aniversarioTemplateAtual(filial){
  filial=String(filial||'').toUpperCase();
  const map=CONFIG_META.aniversario_msg_template_filiais||{};
  return (filial && map[filial]) ? map[filial] : (CONFIG_META.aniversario_msg_template||`Olá, {primeiro_nome}! Feliz aniversário! 🎂🎉\n\nAqui é da Lojas MDL – Móveis do Lar. Desejamos muita saúde, paz e felicidades neste dia especial.\n\nPreparamos condições especiais para você comemorar com a gente. 🧡`);
}
function aniversarioClienteKey(c){
  const filial=String(c?.filial||'').toUpperCase();
  const nome=normName(c?.cliente||'');
  const nasc=String(c?.nascimento||'');
  const cidade=normName(c?.cidade||'');
  return `ANIV|${filial}|${nome}|${nasc}|${cidade}`;
}
function aniversarioOwnerInfo(c){
  const filial=String(c?.filial||'F1').toUpperCase();
  const arr=reativacaoDestinatariosFilial(filial);
  const k=hashStr(String(c?.cliente||'')+'|'+String(c?.nascimento||'')+'|'+filial);
  return arr[Math.abs(k)%arr.length];
}
function aniversarioCurrentKey(){return reativacaoCurrentKey();}
function isAniversarioEnviadoHoje(c){
  const hoje=new Date().toISOString().slice(0,10);
  const rowKey=aniversarioClienteKey(c);
  const nome=normName(c?.cliente||'');
  const filial=String(c?.filial||'').toUpperCase();
  return (COB_LOGS||[]).some(x=>{
    if(String(x.titulo||'').toUpperCase()!=='ANIVERSARIO') return false;
    if(String(x.filial||'').toUpperCase()!==filial) return false;
    if(String(x.server_time||x.data||x.created_at||x.criado_em||'').slice(0,10)!==hoje) return false;
    const lk=String(x.cliente_key||x.cobranca_key||'');
    if(lk && lk===rowKey) return true;
    return nome && normName(x.cliente||'')===nome;
  });
}
function aniversarioRowsPermitidas(){
  let rows=(ANIVERSARIANTES||[]).map((r,i)=>({...r,_idx:i,_owner:aniversarioOwnerInfo(r)}));
  const ck=aniversarioCurrentKey();
  if(ck) rows=rows.filter(r=>String(r._owner?.key||'')===ck);
  return rows;
}
function montarMensagemAniversario(c){
  let msg=aniversarioTemplateAtual(c?.filial||'');
  const vars={primeiro_nome:c?.primeiro_nome||firstNameFromFullName(c?.cliente||''),nome:c?.cliente||'',filial:c?.filial||'',cidade:c?.cidade||'',nascimento:c?.nascimento||''};
  Object.entries(vars).forEach(([k,v])=>{msg=msg.replaceAll('{'+k+'}',String(v||''));});
  return msg;
}
function abrirWhatsAniversario(idx,tel){
  const c=(ANIVERSARIANTES||[])[idx]; if(!c) return;
  const num=String(tel||((c.telefones||[])[0]||'')).replace(/\D/g,'');
  const owner=aniversarioOwnerInfo(c)||{};
  const key=aniversarioClienteKey(c);
  const entRef={type:owner.tipo||'aniversario',filial:c.filial,nome:owner.nome||owner.label||usuarioAtual?.nome||'',login:owner.login||''};
  registrarCobrancaOnline({cliente:c.cliente,titulo:'ANIVERSARIO',parcela:`ANIVERSARIO|${key}|${owner.key||''}`,cliente_key:key,cobranca_key:key,owner_key:owner.key||'',vencimento:c.nascimento||'',pendente:0,filial:c.filial,telefones:[num]}, entRef, num);
  window.open(`https://wa.me/${num}?text=${encodeURIComponent(montarMensagemAniversario(c))}`,'_blank');
  setTimeout(()=>carregarCobrancasOnline().then(()=>{ if(detailScreen && !detailScreen.classList.contains('hidden') && currentDetailRef){ openEntity(currentDetailRef); } else { renderAniversariantesTab(); renderInicioTab(); } }),800);
}
let anivBuscaState=''; let anivFilialState=''; let anivBuscaTimer=null;
function anivBuscaChanged(v){anivBuscaState=String(v||''); clearTimeout(anivBuscaTimer); anivBuscaTimer=setTimeout(()=>renderAniversariantesTab(),350)}
function anivFilialChanged(v){anivFilialState=String(v||''); renderAniversariantesTab();}
async function salvarMensagemAniversarioGlobal(){const el=document.getElementById('anivMsgTemplate'); CONFIG_META.aniversario_msg_template=el?el.value:aniversarioTemplateAtual(); try{const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})}); const j=await resp.json(); toast(j.ok?'Mensagem global de aniversário salva.':'Não consegui salvar mensagem.',j.ok?'success':'warn')}catch(e){toast('Mensagem de aniversário salva localmente para teste.','success')}}
async function salvarMensagemAniversarioFilial(){const f=String(document.getElementById('anivMsgFilial')?.value||'').toUpperCase(); const el=document.getElementById('anivMsgTemplateFilial'); CONFIG_META.aniversario_msg_template_filiais=CONFIG_META.aniversario_msg_template_filiais||{}; CONFIG_META.aniversario_msg_template_filiais[f]=el?el.value:''; try{const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})}); const j=await resp.json(); toast(j.ok?`Mensagem da ${f} salva.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn')}catch(e){toast('Mensagem da filial salva localmente para teste.','success')}}
function trocarMensagemAniversarioFilial(f){const el=document.getElementById('anivMsgTemplateFilial'); if(el) el.value=aniversarioTemplateAtual(f)}
function renderAniversariantesTab(){
  const box=aniversariantesSection; if(!box) return;
  try{
    const filial=String(anivFilialState||''); const q=String(anivBuscaState||'').toLowerCase();
    let rows=aniversarioRowsPermitidas();
    if(filial) rows=rows.filter(r=>String(r.filial||'')===filial);
    if(q) rows=rows.filter(r=>`${r.cliente||''} ${r.cidade||''} ${r._owner?.label||''}`.toLowerCase().includes(q));
    const filiais=mdlV91FiliaisFromRows(ANIVERSARIANTES||[]);
    const enviados=rows.filter(isAniversarioEnviadoHoje).length;
    const totalPermitido=aniversarioRowsPermitidas().length;
    const titulo=(usuarioAtual?.tipo==='master'||usuarioAtual?.is_viewer)?'Lista geral / filtro':'Minha lista de aniversariantes';
    const semBase=!(ANIVERSARIANTES||[]).length;
    box.innerHTML=`<div class="section-head"><div><h2>🎂 Aniversariantes do dia</h2><div class="hint">Base do Sólidus para mensagem de aniversário. ${filiais.length?esc(filiais.map(f=>`${f}: ${(ANIVERSARIANTES||[]).filter(r=>r.filial===f).length}`).join(' · ')):'Nenhum XLS carregado ainda.'}</div></div></div>
    ${semBase?'<div class="glass panel" style="border-color:rgba(245,158,11,.35);margin-bottom:14px"><strong>⚠️ Sem base de aniversariantes carregada</strong><div class="hint">Coloque aniversarios.xls/relatorio_aniversariantes*.xls nesta pasta ou rode com BAIXAR_ANIVERSARIANTES=1.</div></div>':''}
    <div class="kpis" style="margin-bottom:14px">${makeKpi('Aniversariantes na base',String((ANIVERSARIANTES||[]).length),'var(--amber-400)','WhatsApp válido e cidade permitida')}${makeKpi(titulo,String(rows.length),'var(--blue)',usuarioAtual?.tipo==='master'?'Filtro atual':'Distribuída sem duplicar')}${makeKpi('Enviados hoje',String(enviados),'var(--green)','Da lista exibida')}${makeKpi('Minha base total',String(totalPermitido),'var(--orange)','Rateio automático por filial')}</div>
    <div class="glass panel" style="margin-bottom:14px"><div class="search-row" style="grid-template-columns:1.5fr 220px minmax(420px,1fr);align-items:stretch"><div class="input-card"><label>Buscar</label><input id="anivBusca" value="${esc(anivBuscaState)}" oninput="anivBuscaChanged(this.value)" placeholder="Cliente, cidade, responsável"></div><div class="input-card"><label>Filial</label><select id="anivFilial" onchange="anivFilialChanged(this.value)"><option value="">Todas</option>${filiais.map(f=>`<option value="${esc(f)}" ${filial===f?'selected':''}>${esc(f)}</option>`).join('')}</select></div><div class="input-card" style="min-width:420px"><label>Mensagem padrão global</label><textarea id="anivMsgTemplate" rows="6" style="min-height:135px;width:100%;resize:vertical">${esc(aniversarioTemplateAtual())}</textarea><div class="hint">Variáveis: {primeiro_nome}, {nome}, {filial}, {cidade}, {nascimento}</div><button class="btn primary" style="margin-top:8px" onclick="salvarMensagemAniversarioGlobal()">Salvar mensagem global</button></div></div><div class="search-row" style="grid-template-columns:220px minmax(520px,1fr);align-items:stretch;margin-top:12px"><div class="input-card"><label>Mensagem por filial</label><select id="anivMsgFilial" onchange="trocarMensagemAniversarioFilial(this.value)">${filiais.map(f=>`<option value="${esc(f)}">${esc(f)}</option>`).join('')}</select></div><div class="input-card"><label>Texto específico da filial selecionada</label><textarea id="anivMsgTemplateFilial" rows="5" style="min-height:120px;width:100%;resize:vertical">${esc(aniversarioTemplateAtual(filiais[0]||''))}</textarea><div class="hint">Se vazio, usa a mensagem global.</div><button class="btn primary" style="margin-top:8px" onclick="salvarMensagemAniversarioFilial()">Salvar mensagem desta filial</button></div></div></div>
    <div class="faixa-title atencao" style="margin-bottom:10px"><span>🎂 ${esc(titulo)}</span><span>${rows.length} cliente(s) · ${enviados} enviado(s) hoje</span></div><div class="logs-list">${(()=>{const total=rows.length; const maxPage=Math.max(1,Math.ceil(total/REAT_PAGE_SIZE)); if(reatPageState>maxPage) reatPageState=maxPage; if(reatPageState<1) reatPageState=1; const ini=(reatPageState-1)*REAT_PAGE_SIZE; const pageRows=rows.slice(ini,ini+REAT_PAGE_SIZE); const pager=total>REAT_PAGE_SIZE?`<div class="log-pager"><div><strong>${total}</strong> cliente(s) · mostrando ${ini+1}-${Math.min(ini+REAT_PAGE_SIZE,total)} · página ${reatPageState}/${maxPage}</div><div style="display:flex;gap:8px"><button class="btn soft" ${reatPageState<=1?'disabled':''} onclick="reatSetPage(${reatPageState-1})">⬅️ Anterior</button><button class="btn soft" ${reatPageState>=maxPage?'disabled':''} onclick="reatSetPage(${reatPageState+1})">Próxima ➡️</button></div></div>`:''; const body=pageRows.map(r=>{const tels=(r.telefones||[]); const enviado=isAniversarioEnviadoHoje(r); return `<div class="log-row" style="grid-template-columns:1.45fr .85fr .9fr auto"><div><strong>${esc(r.cliente||'')}</strong><div class="small muted">${esc(r.filial||'')} · ${esc(r.cidade||'')} · nasc. ${esc(r.nascimento||'')}</div></div><div><strong>${esc(r._owner?.label||'')}</strong><div class="small muted">Responsável pelo envio</div></div><div><strong>${enviado?'✅ Enviado hoje':esc(tels.length+' WhatsApp(s)')}</strong><div class="small muted">${esc(tels.map(fmtTelBR).join(', '))}</div></div><div style="display:flex;gap:8px;flex-wrap:wrap">${tels.map(t=>`<button class="btn wa" ${enviado?'disabled style="opacity:.45"':''} onclick="abrirWhatsAniversario(${r._idx},'${esc(t)}')">Whats ${esc(fmtTelBR(t))}</button>`).join('')}</div></div>`}).join(''); return pager + (body || '<div class="empty">Nenhum aniversariante encontrado.</div>') + pager;})()}</div>`;
  }catch(e){console.error('Erro renderAniversariantesTab',e); box.innerHTML=`<div class="glass panel" style="border-color:rgba(239,68,68,.35)"><strong>⚠️ Erro na aba Aniversariantes</strong><div class="hint">${esc(e.message||e)}</div></div>`;}
}
function aniversariantesRowsParaEnt(ent){
  if(!ent) return [];
  const filial=String(ent.filial||'').toUpperCase();
  let rows=(ANIVERSARIANTES||[]).map((r,i)=>({...r,_idx:i,_owner:aniversarioOwnerInfo(r)}));
  if(ent.type==='filial') return rows.filter(r=>String(r.filial||'').toUpperCase()===filial);
  const key=reatUserKeyFromNome(ent.nome||ent.login||'',filial);
  return rows.filter(r=>String(r._owner?.key||'')===key);
}
let anivPanelCounter=0;
function showAnivPanel(id,mode){const p=document.getElementById(id+'_pend'), s=document.getElementById(id+'_sent'), bp=document.getElementById(id+'_btn_pend'), bs=document.getElementById(id+'_btn_sent'); if(!p||!s) return; p.style.display=mode==='sent'?'none':'block'; s.style.display=mode==='sent'?'block':'none'; if(bp)bp.classList.toggle('active',mode!=='sent'); if(bs)bs.classList.toggle('active',mode==='sent');}
function renderAniversariantesEnt(ent){
  const rows=aniversariantesRowsParaEnt(ent);
  const enviadosRows=rows.filter(isAniversarioEnviadoHoje); const pendentesRows=rows.filter(r=>!isAniversarioEnviadoHoje(r)); const panelId='aniv_panel_'+(++anivPanelCounter);
  const exportId='export_'+panelId;
  mdlRegisterExport(exportId, 'Aniversariantes do dia - '+(ent?.nome||ent?.filial||'usuario'), [
    ...pendentesRows.map(r=>mdlAnivExportRow(r,'Pendente')),
    ...enviadosRows.map(r=>mdlAnivExportRow(r,'Enviado hoje'))
  ]);
  const rowHtml=(r,enviado=false)=>{const tels=(r.telefones||[]); return `<div class="log-row" style="grid-template-columns:1.5fr .7fr .8fr auto"><div><strong>${esc(r.cliente||'')}</strong><div class="small muted">${esc(r.filial||'')} · ${esc(r.cidade||'')} · nasc. ${esc(r.nascimento||'')}</div></div><div><strong>${esc(r._owner?.label||'')}</strong><div class="small muted">Responsável</div></div><div><strong>${enviado?'✅ Enviado hoje':esc(tels.length+' WhatsApp(s)')}</strong><div class="small muted">${esc(tels.map(fmtTelBR).join(', '))}</div></div><div style="display:flex;gap:8px;flex-wrap:wrap">${tels.map(t=>`<button class="btn wa" ${enviado?'disabled style="opacity:.45"':''} onclick="abrirWhatsAniversario(${r._idx},'${esc(t)}')">Whats ${esc(fmtTelBR(t))}</button>`).join('')}</div></div>`};
  return `<div class="accordion" data-anniv-panel="1"><div class="acc-head" onclick="toggleAcc(this)"><span>🎂 Aniversariantes do dia</span><span style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">${mdlExportButtons(exportId)}<span class="acc-hint">${rows.length} cliente(s) · ${enviadosRows.length} enviado(s) hoje · clique para abrir</span></span></div><div class="acc-body"><div class="faixa-title atencao" style="margin-bottom:10px"><span>Lista individual de aniversário</span><span>sem duplicar com outros usuários</span></div><div class="reat-tabs"><span id="${panelId}_btn_pend" class="reat-tab pending active" onclick="showAnivPanel('${panelId}','pend')">🎂 Pendentes: ${pendentesRows.length}</span><span id="${panelId}_btn_sent" class="reat-tab ok" onclick="showAnivPanel('${panelId}','sent')">✅ Enviados hoje: ${enviadosRows.length}</span></div><div id="${panelId}_pend" class="faixa-block"><div class="faixa-title alerta"><span>🎂 Para enviar hoje</span><span>${pendentesRows.length} cliente(s)</span></div><div class="logs-list">${pendentesRows.slice(0,20).map(r=>rowHtml(r,false)).join('') || '<div class="empty">Nenhum aniversariante pendente para este usuário/filial.</div>'}</div></div><div id="${panelId}_sent" class="faixa-block" style="display:none"><div class="faixa-title atencao"><span>✅ Enviados hoje</span><span>${enviadosRows.length} cliente(s)</span></div><div class="logs-list">${enviadosRows.slice(0,20).map(r=>rowHtml(r,true)).join('') || '<div class="empty">Nenhuma mensagem de aniversário enviada hoje.</div>'}</div></div></div></div>`;
}

// Mural de aniversariantes do dia no início, mesmo quando vazio
function renderMuralAniversariantesDia(){
  const rows=(ANIVERSARIANTES||[]).map((r,i)=>({...r,_idx:i,_owner:aniversarioOwnerInfo(r)}));
  const items=rows.slice(0,80).map(r=>`${r.cliente||''} · ${r.filial||''} · resp. ${r._owner?.label||''}`);
  return `<div class="glass panel full" style="border-color:rgba(249,168,50,.25)"><div class="section-head" style="margin-bottom:6px"><div><h2 style="font-size:18px">🎂 Aniversariantes do dia</h2><div class="hint">Clientes de aniversário para contato por WhatsApp.</div></div><span class="badge">${rows.length}</span></div>${items.length?renderTicker(items,'anivHoje'): '<div class="empty" style="padding:12px">Nenhum aniversariante carregado para hoje.</div>'}</div>`;
}
const _oldRenderInicioTabV42=typeof renderInicioTab==='function'?renderInicioTab:null;
if(_oldRenderInicioTabV42){ renderInicioTab=function(){ _oldRenderInicioTabV42(); try{ const host=document.querySelector('#inicioSection .inicio-operacional-compact')||document.getElementById('inicioSection'); if(host && !host.querySelector('[data-mural-aniversariantes]')){ const wrap=document.createElement('div'); wrap.setAttribute('data-mural-aniversariantes','1'); wrap.innerHTML=renderMuralAniversariantesDia(); host.appendChild(wrap); } }catch(e){console.log(e)} } }

// Injeta aba aniversariantes no detalhe individual depois do painel original renderizar
const _oldOpenEntityV42=typeof openEntity==='function'?openEntity:null;
if(_oldOpenEntityV42){ openEntity=function(ref){ const ret=_oldOpenEntityV42(ref); setTimeout(()=>{try{_setIndividualViewFlag(); const ent=findEntity(ref)||_entAtualUsuario(); const h=renderAniversariantesEnt(ent); if(detailScreen && !detailScreen.querySelector('[data-anniv-panel]')){ const firstAcc=detailScreen.querySelector('.accordion'); if(firstAcc) firstAcc.insertAdjacentHTML('beforebegin',h); else detailScreen.insertAdjacentHTML('beforeend',h); }}catch(e){console.log('aniv detail',e)}},120); return ret; } }

// Botão de notificações aparece para usuário individual mesmo quando zerado
updateGoalNotifications=function(){
  const btn=document.getElementById('goalNotifBtn'), count=document.getElementById('goalNotifCount'); if(!btn||!count) return;
  const ent=_entAtualUsuario(); const arr=_goalNotifsFor(ent);
  if(!usuarioAtual || isAdminLike() || usuarioAtual.is_viewer){btn.classList.add('hidden'); count.textContent='0'; return;}
  btn.classList.remove('hidden');
  const readKey='mdl_goal_notifs_read_'+(usuarioAtual?.login||usuarioAtual?.nome||'user')+'_'+mesAtualComissao();
  let read=[]; try{read=JSON.parse(localStorage.getItem(readKey)||'[]')}catch(e){}
  const unread=arr.filter(n=>!read.includes(n.k)); count.textContent=String(unread.length); btn.classList.toggle('pulse-alert', unread.length>0);
};
openGoalNotifications=function(){
  const ent=_entAtualUsuario(); const arr=_goalNotifsFor(ent);
  const readKey='mdl_goal_notifs_read_'+(usuarioAtual?.login||usuarioAtual?.nome||'user')+'_'+mesAtualComissao();
  try{localStorage.setItem(readKey, JSON.stringify(arr.map(n=>n.k)))}catch(e){}
  updateGoalNotifications();
  const html=`<div class="glass panel" style="max-width:760px;margin:40px auto"><div class="section-head"><div><h2>🎉 Notificações de metas</h2><div class="hint">Metas, bônus e avisos de desempenho do mês.</div></div><button class="btn soft" onclick="this.closest('.goal-modal')?.remove()">Fechar</button></div><div class="tableish">${arr.length?arr.map(n=>`<div class="row-item"><div class="row-top" style="grid-template-columns:72px 1fr"><img src="${LARANJITO}" style="width:58px;height:58px;object-fit:contain;border-radius:14px"><div><div class="name">${esc(n.t)}</div><div class="small muted">${esc(n.d)}</div></div></div></div>`).join(''):'<div class="empty">Nenhuma meta ou bônus atingido ainda.</div>'}</div></div>`;
  const div=document.createElement('div'); div.className='goal-modal'; div.style.cssText='position:fixed;inset:0;background:rgba(0,0,0,.62);z-index:99999;padding:20px;overflow:auto'; div.innerHTML=html; document.body.appendChild(div);
};

// Histórico comissionamento: força render avançado na troca para aba e melhora os botões
const _oldSetHistModeV42=typeof setHistMode==='function'?setHistMode:null;
if(_oldSetHistModeV42){ setHistMode=function(mode){ _oldSetHistModeV42(mode); if(mode==='comissao'){setTimeout(()=>{try{renderHistoricoComissaoResults()}catch(e){console.log(e)}},150)} } }



// ===== V4.3: estabiliza botão de notificações individuais e reforça aniversariantes Selenium =====
try{
  const st=document.createElement('style');
  st.textContent=`body.individual-view #goalNotifBtn.hidden{display:inline-flex!important} #goalNotifBtn{min-width:142px;justify-content:center}`;
  document.head.appendChild(st);
}catch(e){}
try{
  setInterval(()=>{try{updateGoalNotifications()}catch(e){}}, 7000);
  setTimeout(()=>{try{updateGoalNotifications()}catch(e){}}, 1200);
}catch(e){}


// ===== V4.5: WhatsApp sem emoji quebrado + mural de aniversariantes pendentes =====
function safeWhatsTextMDL(s){
  return String(s||'')
    .replace(/[\uD800-\uDBFF][\uDC00-\uDFFF]/g,'')
    .replace(/[\u2600-\u27BF]/g,'')
    .replace(/\uFFFD/g,'')
    .replace(/[“”]/g,'"')
    .replace(/[‘’]/g,"'")
    .replace(/[–—]/g,'-')
    .replace(/[ \t]+\n/g,'\n')
    .replace(/\n{3,}/g,'\n\n')
    .trim();
}
function aniversarioTemplateAtual(filial){
  filial=String(filial||'').toUpperCase();
  const map=CONFIG_META.aniversario_msg_template_filiais||{};
  const def=`Olá, {primeiro_nome}! Feliz aniversário!\n\nAqui é da Lojas MDL - Móveis do Lar. Desejamos muita saúde, paz e felicidades neste dia especial.\n\nPreparamos condições especiais para você comemorar com a gente.`;
  return safeWhatsTextMDL((filial && map[filial]) ? map[filial] : (CONFIG_META.aniversario_msg_template||def));
}
function reativacaoTemplateAtual(filial=""){
  filial=String(filial||"").toUpperCase();
  const porFilial=CONFIG_META?.reativacao_msg_template_filiais||{};
  const def=`Olá, {primeiro_nome}! Tudo bem?\n\nAqui é da Lojas MDL - Móveis do Lar. Estamos com saudades de você! Faz um tempinho que você não aparece na loja.\n\nVenha conhecer nossas novidades e aproveitar condições especiais que preparamos para nossos clientes.`;
  if(filial && String(porFilial[filial]||"").trim()) return safeWhatsTextMDL(String(porFilial[filial]));
  return safeWhatsTextMDL(String(CONFIG_META?.reativacao_msg_template||def));
}
function montarMensagemAniversario(c){
  let msg=aniversarioTemplateAtual(c?.filial||'');
  const vars={primeiro_nome:c?.primeiro_nome||firstNameFromFullName(c?.cliente||''),nome:c?.cliente||'',filial:c?.filial||'',cidade:c?.cidade||'',nascimento:c?.nascimento||''};
  Object.entries(vars).forEach(([k,v])=>{msg=msg.replaceAll('{'+k+'}',String(v||''));});
  return safeWhatsTextMDL(msg);
}
function montarMensagemReativacao(c){
  let tpl=reativacaoTemplateAtual(c?.filial||"");
  const dados={primeiro_nome:c.primeiro_nome||primeiroNomeClienteJs(c.cliente||''),nome:c.cliente||'',filial:c.filial||'',dias:String(c.dias_sem_movimento||''),ultimo_movimento:c.ultimo_movimento||''};
  Object.entries(dados).forEach(([k,v])=>{tpl=tpl.replaceAll(`{${k}}`,v)});
  return safeWhatsTextMDL(tpl);
}
function abrirWhatsReativacao(idx,tel){
  const c=(CLIENTES_SEM_MOVIMENTO||[])[idx]; if(!c) return;
  const num=String(tel||((c.telefones||[])[0]||'')).replace(/\D/g,'');
  const msg=montarMensagemReativacao(c);
  const owner=reativacaoOwnerInfo(c)||{};
  const clienteKey=reativacaoClienteKey(c);
  const entRef={type:owner.tipo||'reativacao',filial:c.filial,nome:owner.nome||owner.label||usuarioAtual?.nome||'',login:owner.login||''};
  registrarCobrancaOnline({cliente:c.cliente,titulo:'REATIVACAO',parcela:`CLIENTE_SEM_MOVIMENTO|${clienteKey}|${owner.key||''}`,cliente_key:clienteKey,cobranca_key:clienteKey,owner_key:owner.key||'',vencimento:c.ultimo_movimento||'',pendente:0,filial:c.filial,telefones:[num]}, entRef, num);
  window.open(`https://wa.me/${num}?text=${encodeURIComponent(msg)}`,'_blank');
  setTimeout(()=>carregarCobrancasOnline().then(()=>{ if(detailScreen && !detailScreen.classList.contains('hidden') && currentDetailRef){ openEntity(currentDetailRef); } else { renderReativacaoTab(); renderInicioTab(); } }),800);
}
function abrirWhatsAniversario(idx,tel){
  const c=(ANIVERSARIANTES||[])[idx]; if(!c) return;
  const num=String(tel||((c.telefones||[])[0]||'')).replace(/\D/g,'');
  const owner=aniversarioOwnerInfo(c)||{};
  const key=aniversarioClienteKey(c);
  const entRef={type:owner.tipo||'aniversario',filial:c.filial,nome:owner.nome||owner.label||usuarioAtual?.nome||'',login:owner.login||''};
  registrarCobrancaOnline({cliente:c.cliente,titulo:'ANIVERSARIO',parcela:`ANIVERSARIO|${key}|${owner.key||''}`,cliente_key:key,cobranca_key:key,owner_key:owner.key||'',vencimento:c.nascimento||'',pendente:0,filial:c.filial,telefones:[num]}, entRef, num);
  window.open(`https://wa.me/${num}?text=${encodeURIComponent(montarMensagemAniversario(c))}`,'_blank');
  setTimeout(()=>carregarCobrancasOnline().then(()=>{ if(detailScreen && !detailScreen.classList.contains('hidden') && currentDetailRef){ openEntity(currentDetailRef); } else { renderAniversariantesTab(); renderInicioTab(); } }),800);
}
function renderMuralAniversariantesDia(){
  const rows=(ANIVERSARIANTES||[]).map((r,i)=>({...r,_idx:i,_owner:aniversarioOwnerInfo(r)}));
  const pendentes=rows.filter(r=>!isAniversarioEnviadoHoje(r));
  const items=pendentes.slice(0,100).map(r=>`${r.cliente||''} · ${r.filial||''} · responsável ${r._owner?.label||''}`);
  return `<div class="glass panel full" style="border-color:rgba(249,168,50,.25)"><div class="section-head" style="margin-bottom:6px"><div><h2 style="font-size:18px">🎂 Aniversariantes do dia</h2><div class="hint">Clientes de aniversário pendentes para saudação por WhatsApp. Ao enviar, sai deste mural.</div></div><span class="badge">${pendentes.length}/${rows.length}</span></div>${items.length?renderTicker(items,'anivHoje'): '<div class="empty" style="padding:12px">Nenhum aniversariante pendente para hoje.</div>'}</div>`;
}


// ===== V4.6 HOTFIX: backups leves, aniversariantes no início, percentuais destacados e histórico/exportação =====
(function(){
  try{
    document.documentElement.style.setProperty('--kpi-badge-bg','rgba(249,168,50,.13)');
    const css = document.createElement('style');
    css.textContent = `
      .kpi .subline{padding-right:88px;line-height:1.35}
      .kpi .kpi-pct-badge{position:absolute;right:16px;bottom:15px;min-width:64px;text-align:center;padding:6px 9px;border-radius:12px;background:var(--kpi-badge-bg);border:1px solid rgba(249,168,50,.32);color:#fbbf24;font-size:20px;font-weight:900;letter-spacing:-.03em;box-shadow:0 8px 25px rgba(249,168,50,.08)}
      .kpi.card-venda-dia .subline,.kpi.card-financeiro .subline{padding-right:88px}
      .inicio-compact .glass.panel.full{margin-bottom:14px;overflow:hidden;max-width:100%}
      .mural-aniversariantes-dia{border-color:rgba(249,168,50,.35)!important}
      .hist-freeze-screen{background:#0d0f14;color:#eef2ff;font-family:Inter,Arial,sans-serif;padding:22px;min-height:100vh}
      .hist-freeze-header{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;margin-bottom:18px;border-bottom:1px solid rgba(255,255,255,.12);padding-bottom:14px}
      .hist-freeze-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:12px 0 18px}
      .hist-freeze-card{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.12);border-radius:14px;padding:12px}
      .hist-freeze-card .k{font-size:11px;text-transform:uppercase;color:#94a3b8;font-weight:800;letter-spacing:.08em}.hist-freeze-card .v{font-size:22px;font-weight:900;margin-top:6px}.hist-freeze-table{width:100%;border-collapse:collapse;margin-top:12px}.hist-freeze-table th,.hist-freeze-table td{border-bottom:1px solid rgba(255,255,255,.1);padding:9px;text-align:left}.hist-freeze-table th{color:#f59e0b;font-size:12px;text-transform:uppercase}
      @media print{button{display:none!important}.hist-freeze-screen{background:#fff;color:#111}.hist-freeze-card{border:1px solid #ddd;background:#fafafa}.hist-freeze-table th,.hist-freeze-table td{border-bottom:1px solid #ddd}}
    `;
    document.head.appendChild(css);
  }catch(e){}
})();

function _v46PctFromSub(sub){
  const s=String(sub||'');
  let m=s.match(/(?:Atingido|Projetado)\s+(\d+(?:[,.]\d+)?)%/i);
  if(!m) m=s.match(/(\d+(?:[,.]\d+)?)%/);
  return m?m[1].replace('.',',')+'%':'';
}
const _oldMakeKpiV46 = typeof makeKpi==='function' ? makeKpi : null;
if(_oldMakeKpiV46){
  makeKpi=function(label,val,accent,sub='',extraClass='',mascote='',iconHtml=''){
    const html=_oldMakeKpiV46(label,val,accent,sub,extraClass,mascote,iconHtml);
    const raw=String(label||'');
    if(/Mercantil realizado|Mercantil projetado|Serviços realizado|Serviços projetado|Caminhão realizado|Caminhão projetado/i.test(raw)){
      const p=_v46PctFromSub(sub);
      if(p && !html.includes('kpi-pct-badge')) return html.replace('</div>', `<div class="kpi-pct-badge">${esc(p)}</div></div>`);
    }
    return html;
  }
}

function renderMuralAniversariantesDia(){
  try{
    const rows=(ANIVERSARIANTES||[]).map((r,i)=>({...r,_idx:i,_owner:aniversarioOwnerInfo(r)}));
    const pendentes=rows.filter(r=>!isAniversarioEnviadoHoje(r));
    const items=pendentes.slice(0,120).map(r=>({nome:r.cliente||'',info:`${r.filial||''} · resp. ${r._owner?.label||''}`}));
    return `<div class="glass panel full mural-aniversariantes-dia" data-mural-aniversariantes="1"><div class="section-head" style="margin-bottom:6px"><div><h2 style="font-size:18px">🎂 Aniversariantes do dia</h2><div class="hint">Clientes de aniversário pendentes para saudação por WhatsApp. Ao enviar, sai deste mural.</div></div><span class="badge">${pendentes.length}/${rows.length}</span></div>${items.length?renderAvisoTicker('Aniversariantes do dia','Cliente e responsável pela saudação.',items,{icon:'🎂',color:'rgba(249,168,50,.25)'}):'<div class="empty" style="padding:12px">Nenhum aniversariante pendente para hoje.</div>'}</div>`;
  }catch(e){return `<div class="glass panel full mural-aniversariantes-dia"><strong>🎂 Aniversariantes do dia</strong><div class="hint">Erro ao montar mural: ${esc(e.message||e)}</div></div>`}
}

const _oldRenderHighlightsV46 = typeof renderHighlights==='function' ? renderHighlights : null;
if(_oldRenderHighlightsV46){
  renderHighlights=function(){
    let html=_oldRenderHighlightsV46()||'';
    if(!html.includes('data-mural-aniversariantes')){
      html += `<div class="section-head" style="margin:20px 0 8px"><div><h2>🎂 Mural de aniversariantes</h2><div class="hint">Saudações de aniversário do dia.</div></div></div>` + renderMuralAniversariantesDia();
    }
    return html;
  }
}

function _v46Num(v){return Number(String(v??0).replace(/\./g,'').replace(',','.'))||0}
function _v46CsvMoney(v){return 'R$ '+_v46Num(v).toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2})}
function _v46CsvPct(v){return (_v46Num(v)).toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2})+'%'}
function _v46CurrentCommissionRows(){
  try{return currentEntities().map(e=>{const m=calcMeta(e); const s=summarizeSalesCard(e)||{}; const c=commissionSummaryForEntity?commissionSummaryForEntity(e):{}; return {tipo:e.type||'',nome:e.nome||filialLabel(e.filial)||'',filial:e.filial||'',login:e.login||'',pendente:e.pendente||0,recebido:e.pago||0,meta_cobranca:m.geral||0,venda_mercantil:s.venda_realizado_total||s.venda||s.valor||0,servicos:s.servico_realizado_total||s.servicos||0,caminhao:s.caminhao_realizado_total||0,comissao_vendas:c.comissao_vendas||0,comissao_cobranca:c.comissao_cobranca||0,bonus:c.bonus_meta||0,rentab48:c.rentab48||0,rentab52:c.rentab52||0,rentab55:c.rentab55||0,total_previsto:c.total_previsto||0}})}catch(e){console.warn(e);return []}
}
function exportarComissaoAtualExcel(){
  const rows=_v46CurrentCommissionRows();
  const head=['Tipo','Nome','Filial','Login','Pendente','Recebido','Meta cobrança %','Venda mercantil','Serviços','Caminhão','Comissão vendas','Comissão cobrança','Bônus meta','Rentab 48','Rentab 52','Rentab 55','Total previsto'];
  const lines=[head.join(';')].concat(rows.map(r=>[r.tipo,r.nome,r.filial,r.login,_v46CsvMoney(r.pendente),_v46CsvMoney(r.recebido),_v46CsvPct(r.meta_cobranca),_v46CsvMoney(r.venda_mercantil),_v46CsvMoney(r.servicos),_v46CsvMoney(r.caminhao),_v46CsvMoney(r.comissao_vendas),_v46CsvMoney(r.comissao_cobranca),_v46CsvMoney(r.bonus),_v46CsvMoney(r.rentab48),_v46CsvMoney(r.rentab52),_v46CsvMoney(r.rentab55),_v46CsvMoney(r.total_previsto)].map(x=>'"'+String(x??'').replace(/"/g,'""')+'"').join(';')));
  const blob=new Blob(['\ufeff'+lines.join('\n')],{type:'text/csv;charset=utf-8'}); const a=document.createElement('a'); a.href=URL.createObjectURL(blob); a.download='comissionamento_atual_'+mesAtualComissao()+'.csv'; a.click(); setTimeout(()=>URL.revokeObjectURL(a.href),1000);
}
function abrirTelaComissionamentoAtual(){
  try{
    const key=document.getElementById('histComCurrentEntity')?.value||'';
    const ent=currentEntities().find(e=>_comKeyNow(e)===key) || currentEntities()[0];
    const m=calcMeta(ent); const s=summarizeSalesCard(ent)||{}; const c=commissionSummaryForEntity?commissionSummaryForEntity(ent):{};
    const rows=[['Pendente',_v46CsvMoney(ent.pendente)],['Recebido',_v46CsvMoney(ent.pago)],['Meta cobrança',_v46CsvPct(m.geral)],['Venda mercantil',_v46CsvMoney(s.venda_realizado_total||s.venda||0)],['Serviços',_v46CsvMoney(s.servico_realizado_total||s.servicos||0)],['Caminhão',_v46CsvMoney(s.caminhao_realizado_total||0)],['Comissão vendas',_v46CsvMoney(c.comissao_vendas)],['Comissão cobrança',_v46CsvMoney(c.comissao_cobranca)],['Bônus/meta',_v46CsvMoney(c.bonus_meta)],['Total previsto',_v46CsvMoney(c.total_previsto)]];
    const html=`<html><head><title>Comissionamento ${esc(ent.nome||'')}</title></head><body><div class="hist-freeze-screen"><button onclick="window.print()">Salvar PDF / Imprimir</button><div class="hist-freeze-header"><div><h1>${esc(ent.nome||filialLabel(ent.filial)||'Entidade')}</h1><div>${esc(ent.type||'')} · ${esc(ent.filial||'')} · mês ${esc(mesAtualComissao())}</div></div><div><strong>Dashboard MDL ${esc(typeof DASHBOARD_BUILD_VERSION !== 'undefined'?DASHBOARD_BUILD_VERSION:'')}</strong><br>${new Date().toLocaleString('pt-BR')}</div></div><div class="hist-freeze-grid">${rows.map(r=>`<div class="hist-freeze-card"><div class="k">${esc(r[0])}</div><div class="v">${esc(r[1])}</div></div>`).join('')}</div><h2>Resumo para folha</h2><table class="hist-freeze-table"><tbody>${rows.map(r=>`<tr><th>${esc(r[0])}</th><td>${esc(r[1])}</td></tr>`).join('')}</tbody></table></div></body></html>`;
    const w=window.open('about:blank','_blank'); w.document.write(html); w.document.close();
  }catch(e){console.error(e); toast('Não foi possível montar tela congelada: '+(e.message||e));}
}

// Mensagem de aniversário do Diretor Comercial para saudação opcional.
function mensagemDiretorAniversario(c){
  const nome=(c.primeiro_nome||c.cliente||'').split(' ')[0]||'cliente';
  return `Olá, ${nome}! Tudo bem?\n\nSou o Diretor Comercial das Lojas MDL - Móveis do Lar. Estou passando para desejar um feliz aniversário, muita saúde, paz e felicidades.\n\nAgradecemos por fazer parte da nossa história. Será sempre um prazer atender você em nossas lojas.`;
}
function abrirWhatsAniversarioDiretor(idx,tel){
  const c=(ANIVERSARIANTES||[])[idx]; if(!c) return;
  const num=String(tel||((c.telefones||[])[0]||'')).replace(/\D/g,'');
  const owner={key:'diretor_comercial',nome:'Diretor Comercial',label:'Diretor Comercial',filial:c.filial||'',login:'diretorcomercial'};
  const key=aniversarioClienteKey(c)+'|DIRETOR';
  registrarCobrancaOnline({cliente:c.cliente,titulo:'ANIVERSARIO_DIRETOR',parcela:`ANIVERSARIO_DIRETOR|${key}`,cliente_key:key,cobranca_key:key,owner_key:owner.key,vencimento:c.nascimento||'',pendente:0,filial:c.filial,telefones:[num]}, {type:'diretor',filial:c.filial,nome:'Diretor Comercial',login:'diretorcomercial'}, num);
  window.open(`https://wa.me/${num}?text=${encodeURIComponent(mensagemDiretorAniversario(c))}`,'_blank');
}
const _oldRenderAniversariantesTabV46 = typeof renderAniversariantesTab==='function' ? renderAniversariantesTab : null;
if(_oldRenderAniversariantesTabV46){
  renderAniversariantesTab=function(){
    _oldRenderAniversariantesTabV46();
    try{
      if(!isPrivilegedUser()) return;
      const box=document.getElementById('aniversariantesSection'); if(!box || box.querySelector('[data-diretor-aniv]')) return;
      const rows=(ANIVERSARIANTES||[]).map((r,i)=>({...r,_idx:i,_owner:aniversarioOwnerInfo(r)}));
      const html=`<div class="glass panel" data-diretor-aniv="1" style="margin-top:14px;border-color:rgba(249,168,50,.25)"><div class="section-head"><div><h2 style="font-size:18px">👔 Saudação do Diretor</h2><div class="hint">Lista geral para o Diretor Comercial também poder enviar uma saudação especial de aniversário.</div></div></div><div class="logs-list">${rows.slice(0,100).map(r=>`<div class="log-row" style="grid-template-columns:1.4fr .7fr 1fr auto"><div><strong>${esc(r.cliente||'')}</strong><div class="small muted">${esc(r.filial||'')} · ${esc(r.cidade||'')} · nasc. ${esc(r.nascimento||'')}</div></div><div><strong>${esc(r._owner?.label||'')}</strong><div class="small muted">Responsável original</div></div><div>${esc((r.telefones||[]).map(fmtTelBR).join(', '))}</div><div>${(r.telefones||[]).map(t=>`<button class="btn wa" onclick="abrirWhatsAniversarioDiretor(${r._idx},'${esc(t)}')">Diretor Whats ${esc(fmtTelBR(t))}</button>`).join('')}</div></div>`).join('') || '<div class="empty">Nenhum aniversariante carregado.</div>'}</div></div>`;
      box.insertAdjacentHTML('beforeend',html);
    }catch(e){console.warn(e)}
  }
}



// ===== V4.7 HOTFIX: murais sem corte + comissionamento/freeze + diretor aniversariantes =====
try{
  const st=document.createElement('style');
  st.textContent=`
  /* Murais rotativos: nome em cima, detalhes abaixo, sem cortar responsável */
  .aviso-ticker{overflow:hidden!important;width:100%!important;max-width:100%!important}
  .aviso-ticker-track{display:flex!important;align-items:stretch!important;gap:10px!important}
  .aviso-pill{
    display:inline-grid!important;grid-template-columns:10px minmax(160px,1fr)!important;grid-template-rows:auto auto!important;
    align-items:center!important;column-gap:8px!important;row-gap:2px!important;
    max-width:360px!important;min-width:230px!important;white-space:normal!important;line-height:1.15!important;
    overflow:hidden!important;text-overflow:clip!important;padding:8px 12px!important;border-radius:16px!important;
  }
  .aviso-pill .red-dot{grid-row:1 / span 2!important;grid-column:1!important;align-self:center!important;flex:0 0 auto!important}
  .aviso-pill small{display:block!important;grid-column:2!important;grid-row:2!important;margin-left:0!important;font-size:10.5px!important;line-height:1.2!important;color:#aeb7ca!important;white-space:normal!important;overflow:hidden!important;text-overflow:ellipsis!important}
  .aviso-pill .ticker-main{grid-column:2!important;grid-row:1!important;font-size:12.5px!important;font-weight:950!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;color:#f4f7ff!important}
  @media(max-width:900px){.aviso-pill{min-width:210px!important;max-width:300px!important}.aviso-pill .ticker-main{font-size:11.5px!important}.aviso-pill small{font-size:10px!important}}
  /* Diretor não edita mensagem global/filial de aniversário; só mensagem própria */
  body.diretor-view #anivMsgTemplate, body.diretor-view #anivMsgTemplateFilial, body.diretor-view #anivMsgFilial{ }
  body.diretor-view .aniv-admin-only{display:none!important}
  .director-msg-card{border-color:rgba(249,168,50,.35)!important;background:rgba(249,168,50,.06)!important}
  `;
  document.head.appendChild(st);
}catch(e){}

// Render ticker substituído para aceitar quebra de linha: cliente/nome em cima; responsável/detalhes embaixo.
function renderAvisoTicker(title,hint,entries,opts={}){
  const icon=opts.icon||'•';
  const color=opts.color||'rgba(249,168,50,.25)';
  const arr=(entries||[]).map(e=>typeof e==='string'?{nome:e,info:''}:e).filter(e=>(e.nome||e.label||'').trim());
  const doubled=arr.concat(arr).concat(arr);
  const safeId='ticker_'+Math.random().toString(36).slice(2);
  if(!arr.length) return `<div class="empty">Nenhuma informação no momento.</div>`;
  return `<div class="glass panel full aviso-rotativo" style="border-color:${color}"><div class="section-head" style="margin-bottom:6px"><div><h2 style="font-size:18px">${esc(icon)} ${esc(title||'Mural')}</h2><div class="hint">${esc(hint||'')}</div></div><div style="display:flex;gap:8px;align-items:center"><span class="badge">${arr.length}</span><button class="btn soft ticker-speed-btn" onclick="toggleTickerSpeed('${safeId}',this)">⚡ Acelerar</button></div></div><div id="${safeId}" class="aviso-ticker"><div class="aviso-ticker-track" style="animation-duration:1800s">${doubled.map(e=>`<span class="aviso-pill"><i class="red-dot"></i><span class="ticker-main">${esc(e.nome||e.label||'')}</span><small>${esc(e.info||'')}</small></span>`).join('')}</div></div></div>`;
}

// Fallback seguro para resumo de comissionamento. Evita erro commissionSummaryForEntity is not defined.
function _v47CommissionSummarySafe(ent){
  try{ if(typeof commissionSummaryForEntity==='function') return commissionSummaryForEntity(ent)||{}; }catch(e){}
  try{ if(ent && ent._commission) return ent._commission; }catch(e){}
  return {comissao_vendas:0,comissao_cobranca:0,bonus_meta:0,rentab48:0,rentab52:0,rentab55:0,total_previsto:0};
}
function _v47MoneyNum(v){
  if(typeof v==='number' && isFinite(v)) return v;
  let s=String(v??'').trim();
  if(!s) return 0;
  s=s.replace(/R\$/g,'').replace(/\s/g,'');
  if(s.includes(',') && s.includes('.')) s=s.replace(/\./g,'').replace(',','.');
  else if(s.includes(',')) s=s.replace(',','.');
  const n=Number(s); return isFinite(n)?n:0;
}
function _v47FmtMoney(v){return _v47MoneyNum(v).toLocaleString('pt-BR',{style:'currency',currency:'BRL'});}
function _v47FmtPct(v){return _v47MoneyNum(v).toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2})+'%';}
function _v46CsvMoney(v){return _v47FmtMoney(v)}
function _v46CsvPct(v){return _v47FmtPct(v)}
function _v47SalesSafe(ent){try{return summarizeSalesCard(ent)||{}}catch(e){return {}}}
function _v47MetaSafe(ent){try{return calcMeta(ent)||{}}catch(e){return {}}}
function _v46CurrentCommissionRows(){
  try{return currentEntities().map(e=>{const m=_v47MetaSafe(e); const s=_v47SalesSafe(e); const c=_v47CommissionSummarySafe(e); return {
    tipo:e.type||'',nome:e.nome||filialLabel(e.filial)||'',filial:e.filial||'',login:e.login||'',
    pendente:_v47MoneyNum(e.pendente),recebido:_v47MoneyNum(e.pago),meta_cobranca:_v47MoneyNum(m.geral),
    venda_mercantil:_v47MoneyNum(s.venda_realizado_total??s.venda??s.valor??0),servicos:_v47MoneyNum(s.servico_realizado_total??s.servicos??0),caminhao:_v47MoneyNum(s.caminhao_realizado_total??0),
    comissao_vendas:_v47MoneyNum(c.comissao_vendas),comissao_cobranca:_v47MoneyNum(c.comissao_cobranca),bonus:_v47MoneyNum(c.bonus_meta),rentab48:_v47MoneyNum(c.rentab48),rentab52:_v47MoneyNum(c.rentab52),rentab55:_v47MoneyNum(c.rentab55),total_previsto:_v47MoneyNum(c.total_previsto)
  }})}catch(e){console.warn('export comissao rows',e);return []}
}
function exportarComissaoAtualExcel(){
  const rows=_v46CurrentCommissionRows();
  const head=['Tipo','Nome','Filial','Login','Pendente','Recebido','Meta cobrança %','Venda mercantil','Serviços','Caminhão','Comissão vendas','Comissão cobrança','Bônus meta','Rentab 48','Rentab 52','Rentab 55','Total previsto'];
  const lines=[head.join(';')].concat(rows.map(r=>[r.tipo,r.nome,r.filial,r.login,_v47FmtMoney(r.pendente),_v47FmtMoney(r.recebido),_v47FmtPct(r.meta_cobranca),_v47FmtMoney(r.venda_mercantil),_v47FmtMoney(r.servicos),_v47FmtMoney(r.caminhao),_v47FmtMoney(r.comissao_vendas),_v47FmtMoney(r.comissao_cobranca),_v47FmtMoney(r.bonus),_v47FmtMoney(r.rentab48),_v47FmtMoney(r.rentab52),_v47FmtMoney(r.rentab55),_v47FmtMoney(r.total_previsto)].map(x=>'"'+String(x??'').replace(/"/g,'""')+'"').join(';')));
  const blob=new Blob(['\ufeff'+lines.join('\n')],{type:'text/csv;charset=utf-8'}); const a=document.createElement('a'); a.href=URL.createObjectURL(blob); a.download='comissionamento_atual_'+(typeof mesAtualComissao==='function'?mesAtualComissao():'mes')+'.csv'; a.click(); setTimeout(()=>URL.revokeObjectURL(a.href),1000);
  toast(`Excel/CSV de comissionamento gerado com ${rows.length} linha(s).`,'success');
}
function _v47ResumoRows(ent){
  const m=_v47MetaSafe(ent); const s=_v47SalesSafe(ent); const c=_v47CommissionSummarySafe(ent);
  return [
    ['Pendente cobrança',_v47FmtMoney(ent?.pendente)],['Recebido cobrança',_v47FmtMoney(ent?.pago)],['Meta cobrança',_v47FmtPct(m.geral)],
    ['Venda mercantil',_v47FmtMoney(s.venda_realizado_total??s.venda??0)],['Serviços',_v47FmtMoney(s.servico_realizado_total??s.servicos??0)],['Caminhão',_v47FmtMoney(s.caminhao_realizado_total??0)],
    ['Comissão vendas',_v47FmtMoney(c.comissao_vendas)],['Comissão cobrança',_v47FmtMoney(c.comissao_cobranca)],['Bônus/meta',_v47FmtMoney(c.bonus_meta)],['Rentab 48',_v47FmtMoney(c.rentab48)],['Rentab 52',_v47FmtMoney(c.rentab52)],['Rentab 55',_v47FmtMoney(c.rentab55)],['Total previsto',_v47FmtMoney(c.total_previsto)]
  ];
}
function abrirTelaComissionamentoAtual(){
  try{
    const key=document.getElementById('histComCurrentEntity')?.value||'';
    const ents=currentEntities();
    const ent=ents.find(e=>(typeof _comKeyNow==='function'?_comKeyNow(e):(e.nome+'_'+e.filial))===key) || ents[0];
    if(!ent) throw new Error('Nenhum usuário/filial encontrado para congelar.');
    const rows=_v47ResumoRows(ent);
    const title=ent.nome||filialLabel(ent.filial)||'Entidade';
    const html=`<!doctype html><html><head><meta charset="utf-8"><title>Comissionamento ${esc(title)}</title><style>body{font-family:Arial,Helvetica,sans-serif;background:#080b10;color:#f4f7ff;padding:24px}.hist-freeze-screen{max-width:1100px;margin:auto}.hist-freeze-header{display:flex;justify-content:space-between;gap:16px;border:1px solid #334155;border-radius:18px;padding:18px;margin-bottom:16px;background:#111827}.hist-freeze-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}.hist-freeze-card{background:#111827;border:1px solid #334155;border-radius:16px;padding:14px}.k{font-size:11px;text-transform:uppercase;color:#9ca3af;font-weight:800}.v{font-size:20px;font-weight:900;margin-top:6px}.hist-freeze-table{width:100%;border-collapse:collapse;margin-top:18px}.hist-freeze-table th,.hist-freeze-table td{border:1px solid #334155;padding:10px;text-align:left}button{padding:10px 14px;border-radius:10px;border:0;font-weight:800}@media print{button{display:none}body{background:white;color:black}.hist-freeze-header,.hist-freeze-card{background:white;color:black;border-color:#ddd}}</style></head><body><div class="hist-freeze-screen"><button onclick="window.print()">Salvar PDF / Imprimir</button><div class="hist-freeze-header"><div><h1>${esc(title)}</h1><div>${esc(ent.type||'')} · ${esc(ent.filial||'')} · mês ${esc(typeof mesAtualComissao==='function'?mesAtualComissao():'')}</div></div><div><strong>Dashboard MDL ${esc(typeof DASHBOARD_BUILD_VERSION !== 'undefined'?DASHBOARD_BUILD_VERSION:'')}</strong><br>${new Date().toLocaleString('pt-BR')}</div></div><div class="hist-freeze-grid">${rows.map(r=>`<div class="hist-freeze-card"><div class="k">${esc(r[0])}</div><div class="v">${esc(r[1])}</div></div>`).join('')}</div><h2>Resumo para folha</h2><table class="hist-freeze-table"><tbody>${rows.map(r=>`<tr><th>${esc(r[0])}</th><td>${esc(r[1])}</td></tr>`).join('')}</tbody></table></div></body></html>`;
    const w=window.open('about:blank','_blank'); if(!w) throw new Error('Pop-up bloqueado pelo navegador.'); w.document.open(); w.document.write(html); w.document.close();
  }catch(e){console.error(e); toast('Não foi possível montar tela congelada: '+(e.message||e),'warn');}
}

// Diretor: usa somente mensagem própria de aniversário.
function diretorAnivTemplateAtual(){return CONFIG_META.aniversario_msg_template_diretor||`Olá, {primeiro_nome}! Tudo bem?\n\nSou o Diretor Comercial das Lojas MDL - Móveis do Lar. Estou passando para desejar um feliz aniversário, muita saúde, paz e felicidades.\n\nAgradecemos por fazer parte da nossa história. Será sempre um prazer atender você em nossas lojas.`}
async function salvarMensagemAniversarioDiretor(){CONFIG_META.aniversario_msg_template_diretor=document.getElementById('anivMsgDiretorTemplate')?.value||diretorAnivTemplateAtual(); try{const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})}); const j=await resp.json(); toast(j.ok?'Mensagem do Diretor salva.':'Mensagem do Diretor salva localmente.',j.ok?'success':'warn')}catch(e){toast('Mensagem do Diretor salva localmente para teste.','success')}}
function mensagemDiretorAniversario(c){
  const vars={primeiro_nome:(c.primeiro_nome||c.cliente||'cliente').split(' ')[0],nome:c.cliente||'',filial:c.filial||'',cidade:c.cidade||'',nascimento:c.nascimento||''};
  let msg=diretorAnivTemplateAtual(); Object.keys(vars).forEach(k=>{msg=msg.replaceAll('{'+k+'}', vars[k])}); return sanitizeWhatsText(msg);
}
function _v47ApplyDiretorAnivLayout(){
  try{
    if(!document.body.classList.contains('diretor-view')) return;
    const box=aniversariantesSection; if(!box || box.querySelector('#anivMsgDiretorTemplate')) return;
    box.querySelectorAll('.input-card').forEach(card=>{ if(card.querySelector('#anivMsgTemplate')||card.querySelector('#anivMsgTemplateFilial')||card.querySelector('#anivMsgFilial')) card.classList.add('aniv-admin-only'); });
    const host=box.querySelector('.glass.panel');
    if(host){host.insertAdjacentHTML('beforeend',`<div class="input-card director-msg-card" style="margin-top:12px"><label>Mensagem própria do Diretor</label><textarea id="anivMsgDiretorTemplate" rows="5" style="min-height:120px;width:100%;resize:vertical">${esc(diretorAnivTemplateAtual())}</textarea><div class="hint">Variáveis: {primeiro_nome}, {nome}, {filial}, {cidade}, {nascimento}</div><button class="btn primary" style="margin-top:8px" onclick="salvarMensagemAniversarioDiretor()">Salvar minha mensagem</button></div>`)}
  }catch(e){console.log('diretor aniv layout',e)}
}
const _oldRenderAnivV47=typeof renderAniversariantesTab==='function'?renderAniversariantesTab:null;
if(_oldRenderAnivV47){renderAniversariantesTab=function(){const r=_oldRenderAnivV47(); setTimeout(_v47ApplyDiretorAnivLayout,60); return r;}}



// ===== V4.8 HOTFIX: acelerar, comissionamento completo e diretor aniversariantes enxuto =====
(function(){
  try{
    const st=document.createElement('style');
    st.textContent=`
      .aviso-ticker{overflow:hidden!important;max-width:100%!important;position:relative}
      .aviso-ticker-track{will-change:transform;display:flex;gap:10px;align-items:stretch;width:max-content;animation-name:tickerMove;animation-timing-function:linear;animation-iteration-count:infinite;animation-duration:1800s}
      .aviso-ticker.fast .aviso-ticker-track{animation-duration:28s!important}
      .aviso-pill{min-width:230px;max-width:360px;white-space:normal!important;line-height:1.22;align-items:flex-start!important;padding:10px 13px!important;display:inline-flex!important;flex-direction:column!important;gap:3px!important}
      .aviso-pill .ticker-main{font-size:13px;font-weight:950;color:#f4f7ff;display:block;max-width:100%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
      .aviso-pill small{font-size:11px!important;color:#9ca3af!important;display:block;max-width:100%;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
      .aniv-diretor-only-card{border-color:rgba(249,168,50,.35)!important}
      body.diretor-view .aniv-admin-only{display:none!important}
    `;
    document.head.appendChild(st);
  }catch(e){}
})();

// Aceita os dois formatos já usados no código: toggleTickerSpeed(this) e toggleTickerSpeed('id', this)
function toggleTickerSpeed(arg1,arg2){
  try{
    let box=null, btn=null;
    if(typeof arg1==='string'){
      box=document.getElementById(arg1);
      btn=arg2||null;
    }else{
      btn=arg1||null;
      box=btn?.closest('.glass.panel,.aviso-rotativo')?.querySelector('.aviso-ticker');
    }
    if(!box) return;
    const fast=!box.classList.contains('fast');
    box.classList.toggle('fast',fast);
    const track=box.querySelector('.aviso-ticker-track');
    if(track) track.style.animationDuration=fast?'28s':'1800s';
    if(btn){btn.textContent=fast?'🐢 Normal':'⚡ Acelerar'; btn.classList.toggle('primary',fast);}
  }catch(e){console.warn('toggleTickerSpeed',e)}
}

function _v48AllComissaoEntities(){
  const out=[]; const seen=new Set();
  function add(e){if(!e) return; const k=_v48EntKey(e); if(seen.has(k)) return; seen.add(k); out.push(e);}
  try{flattenVendedores().forEach(add)}catch(e){}
  try{flattenFiliais().forEach(add)}catch(e){}
  try{crediaristaEntities().forEach(add)}catch(e){}
  try{add(thirdChargeEntity())}catch(e){}
  return out;
}
function _v48EntKey(e){return `${e?.type||'ent'}::${e?.filial||''}::${String(e?.login||e?.nome||'').toLowerCase()}`}
function _v48SelectedEnt(){
  const key=document.getElementById('histComCurrentEntity')?.value||'';
  const ents=_v48AllComissaoEntities();
  return ents.find(e=>_v48EntKey(e)===key || (typeof _comKeyNow==='function' && _comKeyNow(e)===key)) || ents[0] || null;
}
function _v48MoneyTextToNum(v){
  if(typeof v==='number') return v;
  let s=String(v||'').replace(/R\$/g,'').replace(/%/g,'').trim();
  if(!s) return 0;
  s=s.replace(/\./g,'').replace(',', '.').replace(/[^0-9.-]/g,'');
  const n=parseFloat(s); return Number.isFinite(n)?n:0;
}
function _v48RowVal(row, keys){try{return salesCell(row,keys)||''}catch(e){return ''}}
function _v48SalesRow(ent,key){try{return (getSalesRows(ent,key)||[])[0]||null}catch(e){return null}}
function _v48SalesBlock(ent,key){
  const row=_v48SalesRow(ent,key);
  if(!row) return {meta_total:0,real_total:0,ating_total:0,meta_periodo:0,real_periodo:0,ating_periodo:0,projetado:0};
  return {
    meta_total:_v48MoneyTextToNum(_v48RowVal(row,['Meta (R$) Total','Meta(R$) Total'])),
    real_total:_v48MoneyTextToNum(_v48RowVal(row,['Realizado (R$) Total','Realizado(R$) Total'])),
    ating_total:_v48MoneyTextToNum(_v48RowVal(row,['Atingido Total'])),
    meta_periodo:_v48MoneyTextToNum(_v48RowVal(row,['Meta (R$) Período','Meta(R$) Período'])),
    real_periodo:_v48MoneyTextToNum(_v48RowVal(row,['Realizado (R$) Período','Realizado(R$) Período'])),
    ating_periodo:_v48MoneyTextToNum(_v48RowVal(row,['Atingido Período'])),
    projetado:_v48MoneyTextToNum(_v48RowVal(row,['Projetado (R$)','Projetado(R$)']))
  };
}
function _v48ServiceTotal(ent){try{return Number(servicosEntidadeTotal(ent)||0)}catch(e){return 0}}
function _v48FmtMoney(v){try{return R(Number(v||0))}catch(e){return 'R$ 0,00'}}
function _v48FmtPct(v){return (Number(v||0)).toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2})+'%'}
function _v48Commission(ent){try{return calcCommissionSummary(ent)||{}}catch(e){return {}}}

// MDL_V103_CREDIARISTA_FREEZE_FIX: calcula comissão de crediarista/terceiro para tela congelada/exportação.
function _v103CobCommissionSummary(ent){
  try{
    const isCred=!!(ent?.is_crediarista||ent?.type==='crediarista');
    const baseCfg=isCred?entityConfig({type:'vendedor',nome:ent.nome,filial:ent.filial}):entityConfig({type:'vendedor',nome:COBRANCA10_NOME,filial:'FTER'});
    const cfg=commissionCfg(baseCfg);
    const policy=(isCred?(cfg.camp_cob_crediarista||[]):(cfg.camp_cobranca_terceiro||[]));
    const policyOk=Array.isArray(policy)&&policy.length?policy:(isCred?defaultCampCrediarista():defaultCampTerceiro());
    const byFaixa={atencao:{pct:0,cobrado:0,recebido:0,comissao:0},alerta:{pct:0,cobrado:0,recebido:0,comissao:0},grave:{pct:0,cobrado:0,recebido:0,comissao:0}};
    policyOk.forEach(r=>{const fx=String(r.faixa||'').toLowerCase(); if(byFaixa[fx]) byFaixa[fx].pct=Number(String(r.pct||0).replace(',','.'))||0});
    const mesAtual=dateOnlyISO(new Date()).slice(0,7);
    const userKeys=isCred?[String(ent.login||'').toLowerCase(),String(ent.nome||'').toLowerCase()]:[String(COBRANCA10_NOME||'').toLowerCase(),String(COBRANCA10_LOGIN||'').toLowerCase()];
    const cobrados=(COB_LOGS||[]).filter(x=>userKeys.includes(String(x.usuario||'').toLowerCase()) && dateOnlyISO(x.server_time||x.criado_em||x.created_at||x.data||x.server_date||'').slice(0,7)===mesAtual);
    const keys=(typeof keysFromLogsForCommission==='function')?keysFromLogsForCommission(cobrados):new Set();
    const srcCli=isCred?(CLIENTES_CREDIARISTA?.[String(ent.login||'').toLowerCase()]||{grave:[],alerta:[],atencao:[]}):(CLIENTES_TERCEIRO||{grave:[],alerta:[],atencao:[]});
    const srcRec=getRecebimentos(ent)||{grave:[],alerta:[],atencao:[]};
    ['atencao','alerta','grave'].forEach(fx=>{
      byFaixa[fx].cobrado=(srcCli?.[fx]||[]).filter(r=>keys.has(cobrancaRowKey(r))||keys.has(key3Cob(r))).length;
      (srcRec?.[fx]||[]).forEach(r=>{
        const pagMes=dateOnlyISO(r.pagamento||r.data_pagamento||r.pagto||'').slice(0,7);
        if((keys.has(cobrancaRowKey(r))||keys.has(key3Cob(r))) && pagMes===mesAtual){byFaixa[fx].recebido+=Number(r.pago||0)}
      });
      byFaixa[fx].comissao=byFaixa[fx].recebido*(byFaixa[fx].pct/100);
    });
    const total=Object.values(byFaixa).reduce((a,b)=>a+Number(b.comissao||0),0);
    return {isCred, byFaixa, total, totalPrevisto:total,
      recebidoAtencao:byFaixa.atencao.recebido, recebidoAlerta:byFaixa.alerta.recebido, recebidoGrave:byFaixa.grave.recebido,
      comissaoAtencao:byFaixa.atencao.comissao, comissaoAlerta:byFaixa.alerta.comissao, comissaoGrave:byFaixa.grave.comissao,
      pctAtencao:byFaixa.atencao.pct, pctAlerta:byFaixa.alerta.pct, pctGrave:byFaixa.grave.pct};
  }catch(e){console.warn('MDL_V103_CREDIARISTA_FREEZE_FIX falhou', e); return {byFaixa:{},total:0,totalPrevisto:0};}
}

function _v48CommissionRow(ent){
  const meta=calcMeta(ent);
  const isCobOnly=!!(ent?.type==='crediarista'||ent?.is_crediarista||ent?.type==='terceiro'||ent?.is_terceiro);
  if(isCobOnly){
    const cs=_v103CobCommissionSummary(ent);
    const rec=_recebResumo(ent);
    return {
      tipo:ent.type||'', nome:ent.nome||filialLabel(ent.filial)||'', filial:ent.filial||'', login:ent.login||'',
      pendente:Number(ent.pendente||0), recebido:Number(ent.pago||rec.total||0), meta_cobranca:Number(meta.geral||0),
      grave_alvo:Number(meta.grave?.alvo||0), grave_rec:Number(meta.grave?.rec||0), alerta_alvo:Number(meta.alerta?.alvo||0), alerta_rec:Number(meta.alerta?.rec||0), atencao_alvo:Number(meta.atencao?.alvo||0), atencao_rec:Number(meta.atencao?.rec||0),
      venda_meta_total:0, venda_real_total:0, venda_ating_total:0, venda_meta_periodo:0, venda_real_periodo:0, venda_ating_periodo:0, venda_projetado:0,
      servico_meta_total:0, servico_real_total:0, servico_ating_total:0, servico_meta_periodo:0, servico_real_periodo:0, servico_ating_periodo:0, servico_projetado:0,
      caminhao_meta_total:0, caminhao_real_total:0, caminhao_ating_total:0, caminhao_projetado:0,
      faixa:'Cobrança', comissao_mercantil:Number(cs.totalPrevisto||0), comissao_servicos:0, comissao_caminhao:0, bonus_meta:0, rentab48:0, rentab52:0, rentab55:0, total_previsto:Number(cs.totalPrevisto||0),
      recebido_atencao:Number(cs.recebidoAtencao||0), recebido_alerta:Number(cs.recebidoAlerta||0), recebido_grave:Number(cs.recebidoGrave||0),
      comissao_atencao:Number(cs.comissaoAtencao||0), comissao_alerta:Number(cs.comissaoAlerta||0), comissao_grave:Number(cs.comissaoGrave||0),
      pct_atencao:Number(cs.pctAtencao||0), pct_alerta:Number(cs.pctAlerta||0), pct_grave:Number(cs.pctGrave||0), cob_only:true
    };
  }
  const venda=_v48SalesBlock(ent, ent.type==='filial'?'venda_filial_meta':'venda_filial_vendedor_meta');
  const serv=_v48SalesBlock(ent, ent.type==='filial'?'servico_filial_ouro_fob':'servico_filial_vendedor_ouro_fob');
  const cam=_v48SalesBlock(ent, ent.type==='filial'?'venda_filial_subgrupo_20k':'venda_vendedor_subgrupo_20k');
  const c=_v48Commission(ent);
  const servReal=_v48ServiceTotal(ent)||serv.real_total||Number(c.servReal||0);
  const vendaReal=venda.real_total||Number(c.vendaRealBruto||0)||Number(c.vendaReal||0);
  const camReal=cam.real_total||Number(c.camReal||0);
  return {
    tipo:ent.type||'', nome:ent.nome||filialLabel(ent.filial)||'', filial:ent.filial||'', login:ent.login||'',
    pendente:Number(ent.pendente||0), recebido:Number(ent.pago||0), meta_cobranca:Number(meta.geral||0),
    grave_alvo:Number(meta.grave?.alvo||0), grave_rec:Number(meta.grave?.rec||0), alerta_alvo:Number(meta.alerta?.alvo||0), alerta_rec:Number(meta.alerta?.rec||0), atencao_alvo:Number(meta.atencao?.alvo||0), atencao_rec:Number(meta.atencao?.rec||0),
    venda_meta_total:venda.meta_total, venda_real_total:vendaReal, venda_ating_total:venda.ating_total, venda_meta_periodo:venda.meta_periodo, venda_real_periodo:venda.real_periodo, venda_ating_periodo:venda.ating_periodo, venda_projetado:venda.projetado,
    servico_meta_total:serv.meta_total, servico_real_total:servReal, servico_ating_total:serv.ating_total, servico_meta_periodo:serv.meta_periodo, servico_real_periodo:serv.real_periodo, servico_ating_periodo:serv.ating_periodo, servico_projetado:serv.projetado,
    caminhao_meta_total:cam.meta_total, caminhao_real_total:camReal, caminhao_ating_total:cam.ating_total, caminhao_projetado:cam.projetado,
    faixa:String(c.faixaTxt||''), comissao_mercantil:Number(c.vendasComissao||0), comissao_servicos:Number(c.servicosComissao||0), comissao_caminhao:Number(c.caminhaoComissao||0), bonus_meta:Number(c.bonusMeta||0), rentab48:Number(c.rent48||0), rentab52:Number(c.rent52||0), rentab55:Number(c.rent55||0), total_previsto:Number(c.totalPrevisto||0)
  };
}
function _v48CSVCell(v){return '"'+String(v??'').replace(/"/g,'""')+'"'}
function exportarComissaoAtualExcel(){
  const rows=_v48AllComissaoEntities().map(_v48CommissionRow);
  const head=['Tipo','Nome','Filial','Login','Pendente','Recebido','Meta cobrança %','Grave alvo','Grave recebido','Alerta alvo','Alerta recebido','Atenção alvo','Atenção recebido','Venda meta total','Venda realizado total','Venda atingido total %','Venda meta período','Venda realizado período','Venda atingido período %','Venda projetado','Serviço meta total','Serviço realizado total','Serviço atingido total %','Serviço meta período','Serviço realizado período','Serviço atingido período %','Serviço projetado','Caminhão meta total','Caminhão realizado total','Caminhão atingido %','Caminhão projetado','Faixa comissão','Comissão mercantil','Comissão serviços','Comissão caminhão','Bônus meta','Rentab 48','Rentab 52,15','Rentab 55,50','Total previsto'];
  const moneyFields=new Set(['Pendente','Recebido','Grave alvo','Grave recebido','Alerta alvo','Alerta recebido','Atenção alvo','Atenção recebido','Venda meta total','Venda realizado total','Venda meta período','Venda realizado período','Venda projetado','Serviço meta total','Serviço realizado total','Serviço meta período','Serviço realizado período','Serviço projetado','Caminhão meta total','Caminhão realizado total','Caminhão projetado','Comissão mercantil','Comissão serviços','Comissão caminhão','Bônus meta','Rentab 48','Rentab 52,15','Rentab 55,50','Total previsto']);
  const pctFields=new Set(['Meta cobrança %','Venda atingido total %','Venda atingido período %','Serviço atingido total %','Serviço atingido período %','Caminhão atingido %']);
  const vals=r=>[r.tipo,r.nome,r.filial,r.login,r.pendente,r.recebido,r.meta_cobranca,r.grave_alvo,r.grave_rec,r.alerta_alvo,r.alerta_rec,r.atencao_alvo,r.atencao_rec,r.venda_meta_total,r.venda_real_total,r.venda_ating_total,r.venda_meta_periodo,r.venda_real_periodo,r.venda_ating_periodo,r.venda_projetado,r.servico_meta_total,r.servico_real_total,r.servico_ating_total,r.servico_meta_periodo,r.servico_real_periodo,r.servico_ating_periodo,r.servico_projetado,r.caminhao_meta_total,r.caminhao_real_total,r.caminhao_ating_total,r.caminhao_projetado,r.faixa,r.comissao_mercantil,r.comissao_servicos,r.comissao_caminhao,r.bonus_meta,r.rentab48,r.rentab52,r.rentab55,r.total_previsto];
  const lines=[head.join(';')].concat(rows.map(r=>vals(r).map((v,i)=>moneyFields.has(head[i])?_v48CSVCell(_v48FmtMoney(v)):(pctFields.has(head[i])?_v48CSVCell(_v48FmtPct(v)):_v48CSVCell(v))).join(';')));
  const blob=new Blob(['\ufeff'+lines.join('\n')],{type:'text/csv;charset=utf-8'}); const a=document.createElement('a'); a.href=URL.createObjectURL(blob); a.download='comissionamento_completo_'+(typeof mesAtualComissao==='function'?mesAtualComissao():'mes')+'.csv'; a.click(); setTimeout(()=>URL.revokeObjectURL(a.href),1200);
  toast(`Excel/CSV completo gerado com ${rows.length} linha(s).`,'success');
}
function _v48FreezeCard(t,v){return `<div class="hist-freeze-card"><div class="k">${esc(t)}</div><div class="v">${esc(v)}</div></div>`}
function abrirTelaComissionamentoAtual(){
  try{
    const ent=_v48SelectedEnt(); if(!ent) throw new Error('Nenhum usuário/filial encontrado.');
    const r=_v48CommissionRow(ent); const title=r.nome||'Entidade';
    const cards=r.cob_only
      ? [['Pendente cobrança',_v48FmtMoney(r.pendente)],['Recebido cobrança',_v48FmtMoney(r.recebido)],['Meta cobrança',_v48FmtPct(r.meta_cobranca)],['Atenção %',_v48FmtPct(r.pct_atencao)],['Alerta %',_v48FmtPct(r.pct_alerta)],['Grave %',_v48FmtPct(r.pct_grave)],['Recebido atenção',_v48FmtMoney(r.recebido_atencao)],['Recebido alerta',_v48FmtMoney(r.recebido_alerta)],['Recebido grave',_v48FmtMoney(r.recebido_grave)],['Comissão atenção',_v48FmtMoney(r.comissao_atencao)],['Comissão alerta',_v48FmtMoney(r.comissao_alerta)],['Comissão grave',_v48FmtMoney(r.comissao_grave)],['Total previsto',_v48FmtMoney(r.total_previsto)]]
      : [['Pendente cobrança',_v48FmtMoney(r.pendente)],['Recebido cobrança',_v48FmtMoney(r.recebido)],['Meta cobrança',_v48FmtPct(r.meta_cobranca)],['Venda mercantil',_v48FmtMoney(r.venda_real_total)],['Venda atingido',_v48FmtPct(r.venda_ating_total)],['Venda projetado',_v48FmtMoney(r.venda_projetado)],['Serviços',_v48FmtMoney(r.servico_real_total)],['Serviço atingido',_v48FmtPct(r.servico_ating_total)],['Serviço projetado',_v48FmtMoney(r.servico_projetado)],['Caminhão',_v48FmtMoney(r.caminhao_real_total)],['Comissão mercantil',_v48FmtMoney(r.comissao_mercantil)],['Comissão serviços',_v48FmtMoney(r.comissao_servicos)],['Comissão caminhão',_v48FmtMoney(r.comissao_caminhao)],['Bônus/meta',_v48FmtMoney(r.bonus_meta)],['Rentab 48%',_v48FmtMoney(r.rentab48)],['Rentab 52,15%',_v48FmtMoney(r.rentab52)],['Rentab 55,50%',_v48FmtMoney(r.rentab55)],['Total previsto',_v48FmtMoney(r.total_previsto)]];
    const html=`<!doctype html><html><head><meta charset="utf-8"><title>Comissionamento ${esc(title)}</title><style>body{font-family:Arial,Helvetica,sans-serif;background:#080b10;color:#f4f7ff;padding:24px}.hist-freeze-screen{max-width:1180px;margin:auto}.hist-freeze-header{display:flex;justify-content:space-between;gap:16px;border:1px solid #334155;border-radius:18px;padding:18px;margin-bottom:16px;background:#111827}.hist-freeze-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}.hist-freeze-card{background:#111827;border:1px solid #334155;border-radius:16px;padding:14px}.k{font-size:11px;text-transform:uppercase;color:#9ca3af;font-weight:800}.v{font-size:20px;font-weight:900;margin-top:6px}.hist-freeze-table{width:100%;border-collapse:collapse;margin-top:18px}.hist-freeze-table th,.hist-freeze-table td{border:1px solid #334155;padding:10px;text-align:left}button{padding:10px 14px;border-radius:10px;border:0;font-weight:800}@media print{button{display:none}body{background:white;color:black}.hist-freeze-header,.hist-freeze-card{background:white;color:black;border-color:#ddd}}</style></head><body><div class="hist-freeze-screen"><button onclick="window.print()">Salvar PDF / Imprimir</button><div class="hist-freeze-header"><div><h1>${esc(title)}</h1><div>${esc(r.tipo)} · ${esc(r.filial)} · mês ${esc(typeof mesAtualComissao==='function'?mesAtualComissao():'')}</div></div><div><strong>Dashboard MDL ${esc(typeof DASHBOARD_BUILD_VERSION !== 'undefined'?DASHBOARD_BUILD_VERSION:'')}</strong><br>${new Date().toLocaleString('pt-BR')}</div></div><div class="hist-freeze-grid">${cards.map(x=>_v48FreezeCard(x[0],x[1])).join('')}</div><h2>Resumo para folha</h2><table class="hist-freeze-table"><tbody>${cards.map(x=>`<tr><th>${esc(x[0])}</th><td>${esc(x[1])}</td></tr>`).join('')}</tbody></table></div></body></html>`;
    const w=window.open('about:blank','_blank'); if(!w) throw new Error('Pop-up bloqueado pelo navegador.'); w.document.open(); w.document.write(html); w.document.close();
  }catch(e){console.error(e); toast('Não foi possível montar tela congelada: '+(e.message||e),'warn');}
}
function renderHistoricoComissaoResults(){
  const box=document.getElementById('histComResults'); if(!box) return;
  const ents=_v48AllComissaoEntities();
  const currentPanel=`<div class="glass panel" style="margin-bottom:14px"><div class="section-head"><div><h2 style="font-size:18px">🧊 Tela individual congelada atual</h2><div class="hint">Escolha usuário/filial/crediarista/terceiro. A tela e o Excel usam vendas, serviços, cobrança, bônus e comissões atuais.</div></div></div><div class="form-grid"><div class="input-card"><label>Usuário / filial</label><select id="histComCurrentEntity">${ents.map(e=>`<option value="${esc(_v48EntKey(e))}">${esc(e.nome||filialLabel(e.filial)||'Entidade')} · ${esc(e.filial||'')}</option>`).join('')}</select></div><div style="display:flex;align-items:end;gap:8px;flex-wrap:wrap"><button class="btn primary" onclick="abrirTelaComissionamentoAtual()">Abrir tela congelada</button><button class="btn soft" onclick="exportarComissaoAtualExcel()">Exportar Excel completo</button></div></div></div>`;
  const months=(typeof _histComMeses==='function'?_histComMeses():[]); const current=document.getElementById('histComMonth')?.value || months[0] || (typeof mesAtualComissao==='function'?mesAtualComissao():'');
  const snap=HIST_COMISSAO?.months?.[current];
  if(!snap){box.innerHTML=currentPanel+`<div class="empty">Nenhum histórico salvo para ${esc(current)}. Você já pode abrir a tela congelada atual ou exportar o Excel completo acima.</div>`; return;}
  const rows=[...(snap.entidades||[])].sort((a,b)=>String(a.tipo).localeCompare(String(b.tipo),'pt-BR')||String(a.nome).localeCompare(String(b.nome),'pt-BR'));
  box.innerHTML=currentPanel+`<div class="kpis">${makeKpi('Mês',esc(snap.month||current),'var(--blue)')}${makeKpi('Total previsto',R(snap.total_previsto||0),'var(--green)')}${makeKpi('Entidades',String(rows.length),'var(--orange)')}${makeKpi('Salvo em',esc((snap.atualizado_em_br||snap.gerado_em||'').replace('T',' ').slice(0,19)),'var(--blue)')}</div>`+(typeof renderComissionamentoHistoricoTable==='function'?renderComissionamentoHistoricoTable(rows):'');
}
function _v48IsDiretorView(){return !!(usuarioAtual && usuarioAtual.tipo==='master' && usuarioAtual.roleLabel==='Diretor Comercial')}
function _v48ApplyAnivPermissions(){
  try{
    const box=aniversariantesSection; if(!box) return;
    if(_v48IsDiretorView()){
      // Diretor: só lista + mensagem própria. Remove config global/filial do master.
      box.querySelectorAll('textarea#anivMsgTemplate, textarea#anivMsgTemplateFilial, select#anivMsgFilial').forEach(el=>{const card=el.closest('.input-card')||el.closest('.glass.panel'); if(card) card.style.display='none';});
      box.querySelectorAll('button').forEach(b=>{const tx=(b.textContent||'').toLowerCase(); if(tx.includes('salvar mensagem global')||tx.includes('salvar mensagem desta filial')){const card=b.closest('.input-card')||b; card.style.display='none';}});
      if(!box.querySelector('#anivMsgDiretorTemplate')){
        const panel=box.querySelector('.glass.panel');
        if(panel) panel.insertAdjacentHTML('beforeend',`<div class="input-card aniv-diretor-only-card" style="margin-top:12px"><label>Minha mensagem de aniversário como Diretor</label><textarea id="anivMsgDiretorTemplate" rows="5" style="min-height:120px;width:100%;resize:vertical">${esc(diretorAnivTemplateAtual())}</textarea><div class="hint">Esta mensagem é só para o envio do Diretor. O Master configura mensagens globais/filiais.</div><button class="btn primary" style="margin-top:8px" onclick="salvarMensagemAniversarioDiretor()">Salvar minha mensagem</button></div>`);
      }
    }else if(usuarioAtual?.tipo==='master'){
      // Master: mantém global/filial e também configura mensagem individual do Diretor.
      if(!box.querySelector('#anivMsgDiretorTemplateMaster')){
        const panel=box.querySelector('.glass.panel');
        if(panel) panel.insertAdjacentHTML('beforeend',`<div class="input-card aniv-diretor-only-card" style="margin-top:12px"><label>Mensagem individual do Diretor Comercial</label><textarea id="anivMsgDiretorTemplateMaster" rows="5" style="min-height:120px;width:100%;resize:vertical">${esc(diretorAnivTemplateAtual())}</textarea><div class="hint">Mensagem usada quando o Diretor enviar saudação de aniversário.</div><button class="btn primary" style="margin-top:8px" onclick="CONFIG_META.aniversario_msg_template_diretor=document.getElementById('anivMsgDiretorTemplateMaster').value; salvarMensagemAniversarioDiretor()">Salvar mensagem do Diretor</button></div>`);
      }
    }
  }catch(e){console.warn('aniv permissions',e)}
}
const _v48RenderAnivBase = typeof renderAniversariantesTab==='function' ? renderAniversariantesTab : null;
if(_v48RenderAnivBase){ renderAniversariantesTab=function(){ const r=_v48RenderAnivBase(); setTimeout(_v48ApplyAnivPermissions,80); return r; } }
const _v48SetHistModeBase = typeof setHistMode==='function' ? setHistMode : null;
if(_v48SetHistModeBase){ setHistMode=function(mode){ const r=_v48SetHistModeBase(mode); if(mode==='comissao') setTimeout(renderHistoricoComissaoResults,80); return r; } }
try{setTimeout(()=>{if((typeof mainTab!=='undefined') && mainTab==='historico' && (window._histMode==='comissao')) renderHistoricoComissaoResults();},1200)}catch(e){}


// ===== V4.9 HOTFIX: diretor aniversario somente leitura + mensagem correta + ticker funcional =====
(function(){
  try{
    const css=document.createElement('style');
    css.textContent=`
      .aviso-ticker{overflow:hidden!important;max-width:100%!important;position:relative!important}
      .aviso-ticker-track{display:flex!important;gap:10px!important;align-items:stretch!important;width:max-content!important;white-space:nowrap!important;will-change:transform!important;animation-name:mdlTicker!important;animation-timing-function:linear!important;animation-iteration-count:infinite!important;animation-duration:900s!important;animation-play-state:running!important}
      .aviso-ticker.fast .aviso-ticker-track{animation-duration:26s!important}
      .aviso-ticker .aviso-pill{display:inline-flex!important;flex-direction:column!important;align-items:flex-start!important;justify-content:center!important;min-width:210px!important;max-width:300px!important;white-space:normal!important;line-height:1.15!important;overflow:hidden!important}
      .aviso-ticker .aviso-pill .ticker-main{display:block!important;max-width:100%!important;overflow:hidden!important;text-overflow:ellipsis!important;white-space:nowrap!important;font-size:13px!important;font-weight:900!important}
      .aviso-ticker .aviso-pill small{display:block!important;max-width:100%!important;overflow:hidden!important;text-overflow:ellipsis!important;white-space:nowrap!important;font-size:11px!important;opacity:.78!important;margin-top:3px!important}
      .ticker-speed-btn{min-width:106px!important;justify-content:center!important}
      .aniv-diretor-readonly-card textarea{opacity:.92!important;background:rgba(255,255,255,.035)!important}
    `;
    document.head.appendChild(css);
  }catch(e){}
})();

function toggleTickerSpeed(a,b){
  try{
    let box=null, btn=null;
    if(typeof a==='string'){
      box=document.getElementById(a);
      btn=b || null;
    }else{
      btn=a || null;
      box=btn?.closest('.glass.panel,.aviso-rotativo')?.querySelector('.aviso-ticker') || btn?.closest('.aviso-ticker');
    }
    if(!box) box=document.querySelector('.aviso-ticker');
    if(!box) return false;
    const fast=!box.classList.contains('fast');
    box.classList.toggle('fast',fast);
    const track=box.querySelector('.aviso-ticker-track');
    if(track){
      track.style.animationName='mdlTicker';
      track.style.animationTimingFunction='linear';
      track.style.animationIterationCount='infinite';
      track.style.animationPlayState='running';
      track.style.animationDuration=fast?'26s':'900s';
    }
    if(btn){
      btn.textContent=fast?'🐢 Normal':'⚡ Acelerar';
      btn.classList.toggle('primary',fast);
    }
    return false;
  }catch(e){console.warn('toggleTickerSpeed V4.9',e); return false;}
}

function diretorAnivTemplateAtual(){
  return safeWhatsTextMDL(CONFIG_META.aniversario_msg_template_diretor || `Olá, {primeiro_nome}! Tudo bem?\n\nSou o Diretor Comercial das Lojas MDL - Móveis do Lar. Estou passando para desejar um feliz aniversário, muita saúde, paz e felicidades.\n\nAgradecemos por fazer parte da nossa história. Será sempre um prazer atender você em nossas lojas.`);
}
function mensagemDiretorAniversario(c){
  let msg=diretorAnivTemplateAtual();
  const vars={primeiro_nome:c?.primeiro_nome||firstNameFromFullName(c?.cliente||''),nome:c?.cliente||'',filial:c?.filial||'',cidade:c?.cidade||'',nascimento:c?.nascimento||''};
  Object.entries(vars).forEach(([k,v])=>{msg=msg.replaceAll('{'+k+'}',String(v||''));});
  return safeWhatsTextMDL(msg);
}

const _v49AbrirWhatsAniversarioBase = typeof abrirWhatsAniversario==='function' ? abrirWhatsAniversario : null;
abrirWhatsAniversario=function(idx,tel){
  try{
    if(typeof _v48IsDiretorView==='function' && _v48IsDiretorView()){
      return abrirWhatsAniversarioDiretor(idx,tel);
    }
  }catch(e){}
  return _v49AbrirWhatsAniversarioBase ? _v49AbrirWhatsAniversarioBase(idx,tel) : null;
}

function _v49ApplyDiretorAnivReadonly(){
  try{
    const box=aniversariantesSection; if(!box) return;
    if(typeof _v48IsDiretorView==='function' && _v48IsDiretorView()){
      // Remove qualquer configuração editável; Diretor só visualiza a mensagem configurada pelo Master.
      box.querySelectorAll('textarea#anivMsgTemplate, textarea#anivMsgTemplateFilial, select#anivMsgFilial, #anivMsgDiretorTemplate').forEach(el=>{
        const card=el.closest('.input-card')||el.closest('.glass.panel'); if(card) card.remove();
      });
      box.querySelectorAll('button').forEach(b=>{
        const tx=(b.textContent||'').toLowerCase();
        if(tx.includes('salvar mensagem')){ const card=b.closest('.input-card')||b; card.remove(); }
      });
      const panel=box.querySelector('.glass.panel');
      if(panel && !box.querySelector('[data-diretor-msg-preview]')){
        panel.insertAdjacentHTML('beforeend',`<div class="input-card aniv-diretor-readonly-card" data-diretor-msg-preview="1" style="margin-top:12px;border-color:rgba(249,168,50,.32)"><label>Mensagem do Diretor configurada pelo Master</label><textarea rows="5" readonly style="min-height:120px;width:100%;resize:vertical">${esc(diretorAnivTemplateAtual())}</textarea><div class="hint">Somente o Master altera esta mensagem. Ao clicar no WhatsApp como Diretor, esta será a mensagem enviada.</div></div>`);
      }
    }else if(usuarioAtual?.tipo==='master'){
      // Master configura a mensagem individual do Diretor; mantém global/filial.
      const panel=box.querySelector('.glass.panel');
      if(panel && !box.querySelector('#anivMsgDiretorTemplateMaster')){
        panel.insertAdjacentHTML('beforeend',`<div class="input-card aniv-diretor-only-card" style="margin-top:12px;border-color:rgba(249,168,50,.32)"><label>Mensagem individual do Diretor Comercial</label><textarea id="anivMsgDiretorTemplateMaster" rows="5" style="min-height:120px;width:100%;resize:vertical">${esc(diretorAnivTemplateAtual())}</textarea><div class="hint">Mensagem usada quando o Diretor enviar saudação de aniversário. Variáveis: {primeiro_nome}, {nome}, {filial}, {cidade}, {nascimento}</div><button class="btn primary" style="margin-top:8px" onclick="CONFIG_META.aniversario_msg_template_diretor=document.getElementById('anivMsgDiretorTemplateMaster').value; salvarMensagemAniversarioDiretor()">Salvar mensagem do Diretor</button></div>`);
      }
    }
  }catch(e){console.warn('diretor readonly V4.9',e)}
}

const _v49RenderAniversariantesBase = typeof renderAniversariantesTab==='function' ? renderAniversariantesTab : null;
if(_v49RenderAniversariantesBase){
  renderAniversariantesTab=function(){
    const r=_v49RenderAniversariantesBase();
    setTimeout(_v49ApplyDiretorAnivReadonly,80);
    setTimeout(()=>{try{document.querySelectorAll('.aviso-ticker-track').forEach(t=>{t.style.animationName='mdlTicker'; t.style.animationPlayState='running';});}catch(e){}},120);
    return r;
  }
}
try{setTimeout(_v49ApplyDiretorAnivReadonly,1200)}catch(e){}
try{setInterval(()=>{try{document.querySelectorAll('.aviso-ticker-track').forEach(t=>{t.style.animationName='mdlTicker'; if(!t.style.animationDuration)t.style.animationDuration='900s';});}catch(e){}},5000)}catch(e){}


// ===== V5.1: MURAIS ORGANIZADOS + CARD RESUMO NO TOPO =====
(function(){
  try{
    const st=document.createElement('style');
    st.textContent=`
      .mdl-hero-mural-card{grid-column:span 2!important;min-height:112px!important;padding:14px 16px!important;border:1px solid rgba(249,168,50,.36)!important;background:radial-gradient(circle at 12% 12%,rgba(249,168,50,.18),transparent 34%),linear-gradient(135deg,rgba(17,24,39,.96),rgba(10,13,20,.94))!important;box-shadow:0 18px 42px rgba(0,0,0,.28), inset 0 0 0 1px rgba(255,255,255,.04)!important;position:relative;overflow:hidden!important;display:flex;flex-direction:column;justify-content:center}
      .mdl-hero-mural-card:before{content:"";position:absolute;right:-26px;top:-30px;width:130px;height:130px;border-radius:999px;background:rgba(249,168,50,.12);filter:blur(2px)}
      .mdl-hero-title{font-size:12px;font-weight:950;text-transform:uppercase;letter-spacing:.08em;color:#fbbf24;margin-bottom:7px;display:flex;align-items:center;gap:7px}
      .mdl-hero-main{font-size:22px;font-weight:950;color:#fff;line-height:1.08;margin-bottom:6px;position:relative;z-index:1}
      .mdl-hero-detail{font-size:12px;color:#aab4cb;font-weight:750;line-height:1.35;position:relative;z-index:1;max-width:86%}
      .mdl-hero-count{position:absolute;right:18px;bottom:14px;border-radius:999px;padding:8px 12px;font-weight:950;color:#fff;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.13)}
      .mdl-hero-dot{display:inline-block;width:9px;height:9px;border-radius:99px;background:#f59e0b;box-shadow:0 0 18px rgba(249,168,50,.65)}
      .mdl-mural-zone{display:grid;gap:14px;margin-top:12px}.mdl-mural-group{border-radius:24px;padding:15px 16px;border:1px solid rgba(255,255,255,.09);background:linear-gradient(135deg,rgba(16,22,34,.72),rgba(10,12,18,.66));box-shadow:0 18px 42px rgba(0,0,0,.2)}
      .mdl-mural-group.cobranca{border-color:rgba(239,68,68,.24);background:linear-gradient(135deg,rgba(127,29,29,.18),rgba(15,23,42,.66))}.mdl-mural-group.reativacao{border-color:rgba(245,158,11,.28);background:linear-gradient(135deg,rgba(120,53,15,.18),rgba(15,23,42,.66))}.mdl-mural-group.vendas{border-color:rgba(34,197,94,.28);background:linear-gradient(135deg,rgba(6,78,59,.18),rgba(15,23,42,.66))}.mdl-mural-group.aniversarios{border-color:rgba(236,72,153,.24);background:linear-gradient(135deg,rgba(131,24,67,.16),rgba(15,23,42,.66))}
      .mdl-mural-group-head{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;margin-bottom:10px}.mdl-mural-group-head h2{font-size:18px!important;margin:0!important}.mdl-mural-group-head .hint{font-size:12px}.mdl-mural-group-grid{display:grid;grid-template-columns:1fr;gap:10px}.mdl-mural-group .glass.panel{margin-bottom:0!important;background:rgba(255,255,255,.035)!important}.mdl-mural-group.compact .glass.panel{padding:12px 14px!important}
      .aviso-rotativo,.mdl-mural-group .glass.panel{overflow:hidden!important}.aviso-ticker{overflow:hidden!important;width:100%!important;max-width:100%!important;position:relative}.aviso-ticker-track{display:flex!important;gap:10px!important;align-items:stretch!important;width:max-content!important;animation-name:mdlTicker!important;animation-timing-function:linear!important;animation-iteration-count:infinite!important;animation-duration:900s!important;will-change:transform!important}.aviso-ticker.fast .aviso-ticker-track{animation-duration:22s!important}.aviso-pill{min-width:260px!important;max-width:330px!important;white-space:normal!important;display:inline-flex!important;flex-direction:column!important;align-items:flex-start!important;justify-content:center!important;gap:4px!important;padding:10px 13px!important;line-height:1.22!important;overflow:visible!important}.aviso-pill .ticker-main{display:block!important;white-space:normal!important;overflow:visible!important;text-overflow:clip!important;font-size:12.5px!important;line-height:1.22!important}.aviso-pill small{display:block!important;white-space:normal!important;overflow:visible!important;text-overflow:clip!important;font-size:10.8px!important;line-height:1.18!important;color:#aab4cb!important}.ticker-speed-btn{min-width:105px!important;justify-content:center!important}.ticker-speed-btn.primary{background:linear-gradient(135deg,#f59e0b,#f97316)!important;color:#111827!important}
      .mdl-sales-line{display:flex;gap:10px;flex-wrap:wrap}.mdl-sales-line span{border:1px solid rgba(255,255,255,.09);border-radius:999px;background:rgba(255,255,255,.045);padding:8px 11px;font-weight:850;font-size:12px;color:#dbe4ff}.mdl-empty-mural{padding:12px 14px;border-radius:16px;border:1px dashed rgba(255,255,255,.12);background:rgba(255,255,255,.035);color:#aab4cb;font-weight:800}
      @media(max-width:1180px){.mdl-hero-mural-card{grid-column:1/-1!important}.mdl-hero-detail{max-width:100%}}
    `;
    document.head.appendChild(st);
  }catch(e){console.warn('v5.1 css',e)}
})();

function mdlV51CountFromHtml(html){try{const m=String(html||'').match(/class=["']badge["'][^>]*>(\d+)/i);return m?Number(m[1]||0):0}catch(e){return 0}}
function mdlV51TextFromHtml(html){try{const tmp=document.createElement('div'); tmp.innerHTML=String(html||''); const pill=tmp.querySelector('.aviso-pill'); return (pill?.innerText||tmp.innerText||'').replace(/\s+/g,' ').trim().slice(0,130)}catch(e){return ''}}
function mdlV51AnivPendentes(){try{return (ANIVERSARIANTES||[]).filter(c=>!(typeof isAniversarioEnviadoHoje==='function'&&isAniversarioEnviadoHoje(c)))}catch(e){return []}}
function mdlV51MetaDiariaCount(){try{return mdlV51CountFromHtml(renderMetaDiariaBatidaAlerts())}catch(e){return 0}}
function mdlV51HeroSlides(){
  const noCobHtml=(()=>{try{return renderNoChargeAlerts()||''}catch(e){return ''}})();
  const noReatHtml=(()=>{try{return renderNoReactivationAlerts()||''}catch(e){return ''}})();
  const metaHtml=(()=>{try{return renderMetaDiariaBatidaAlerts()||''}catch(e){return ''}})();
  const anivs=mdlV51AnivPendentes();
  const slides=[];
  slides.push({icon:'⏰',title:'Sem cobranças hoje',count:mdlV51CountFromHtml(noCobHtml),main:'Carteiras pendentes de ação',detail:mdlV51TextFromHtml(noCobHtml)||'Nenhum alerta crítico de cobrança agora.',dot:'#ef4444'});
  slides.push({icon:'🧡',title:'Clientes sem movimento',count:mdlV51CountFromHtml(noReatHtml),main:'Usuários sem reativação hoje',detail:mdlV51TextFromHtml(noReatHtml)||'Todos os responsáveis acionaram ou não há lista pendente.',dot:'#f59e0b'});
  slides.push({icon:'🎂',title:'Aniversariantes do dia',count:anivs.length,main:anivs.length?'Saudações pendentes':'Aniversariantes em dia',detail:anivs.slice(0,2).map(c=>`${c.cliente||''} · resp. ${(typeof aniversarioOwnerInfo==='function'?aniversarioOwnerInfo(c).label:'')||''}`).join('  |  ')||'Nenhum aniversariante pendente no momento.',dot:'#ec4899'});
  slides.push({icon:'🎯',title:'Meta diária BATIDA',count:mdlV51MetaDiariaCount(),main:mdlV51MetaDiariaCount()?'Meta diária batida hoje':'Aguardando metas diárias',detail:mdlV51TextFromHtml(metaHtml)||'Quando vendedor ou filial bater, aparecerá aqui automaticamente.',dot:'#22c55e'});
  return slides;
}
function mdlV51HeroHtml(){const slides=mdlV51HeroSlides();return `<div id="mdlHeroMural" class="glass kpi mdl-hero-mural-card" data-idx="0">${mdlV51HeroSlideHtml(slides[0]||{})}</div>`}
function mdlV51HeroSlideHtml(s){return `<div class="mdl-hero-title"><span class="mdl-hero-dot" style="background:${esc(s.dot||'#f59e0b')}"></span>${esc(s.icon||'🔔')} ${esc(s.title||'Mural operacional')}</div><div class="mdl-hero-main">${esc(s.main||'Operação em andamento')}</div><div class="mdl-hero-detail">${esc(s.detail||'Resumo automático dos murais do dia.')}</div><div class="mdl-hero-count">${esc(String(s.count??0))}</div>`}
function mdlV51StartHeroMural(){try{const el=document.getElementById('mdlHeroMural'); if(!el) return; const slides=mdlV51HeroSlides(); if(window._mdlHeroTimer) clearInterval(window._mdlHeroTimer); window._mdlHeroTimer=setInterval(()=>{try{const fresh=mdlV51HeroSlides(); const idx=(Number(el.dataset.idx||0)+1)%Math.max(fresh.length,1); el.dataset.idx=String(idx); el.innerHTML=mdlV51HeroSlideHtml(fresh[idx]||{});}catch(e){}},5200);}catch(e){}}
function mdlV51InstallHeroMural(){try{const kpis=document.getElementById('kpis'); if(!kpis || document.getElementById('mdlHeroMural')) return; const allow=(usuarioAtual?.tipo==='master'||usuarioAtual?.is_viewer||usuarioAtual?.roleLabel==='Diretor Comercial'); if(!allow) return; const update=[...kpis.children].find(x=>String(x.textContent||'').includes('Última atualização do dashboard')); const holder=document.createElement('div'); holder.innerHTML=mdlV51HeroHtml(); const card=holder.firstElementChild; if(update) kpis.insertBefore(card,update); else kpis.appendChild(card); mdlV51StartHeroMural();}catch(e){console.warn('install hero mural',e)}}
if(typeof renderKPIs==='function' && !window._renderKPIsV51Wrapped){window._renderKPIsV51Wrapped=true; const _oldRenderKPIsV51=renderKPIs; renderKPIs=function(){const r=_oldRenderKPIsV51.apply(this,arguments); setTimeout(mdlV51InstallHeroMural,80); return r;}}

// Botão acelerar blindado: funciona por onclick antigo e por clique delegado.
function toggleTickerSpeed(arg1,arg2){
  try{
    let box=null,btn=null;
    if(typeof arg1==='string'){box=document.getElementById(arg1);btn=arg2||null}else{btn=arg1||null;box=btn?.closest('.glass.panel,.aviso-rotativo,.mdl-mural-group')?.querySelector('.aviso-ticker')}
    if(!box && btn) box=btn.parentElement?.parentElement?.parentElement?.querySelector?.('.aviso-ticker');
    if(!box) return;
    const fast=!box.classList.contains('fast'); box.classList.toggle('fast',fast);
    const track=box.querySelector('.aviso-ticker-track'); if(track){track.style.animationName='mdlTicker'; track.style.animationDuration=fast?'22s':'900s'; track.style.animationPlayState='running';}
    if(btn){btn.textContent=fast?'🐢 Normal':'⚡ Acelerar'; btn.classList.toggle('primary',fast)}
  }catch(e){console.warn('toggle ticker v51',e)}
}
document.addEventListener('click',function(ev){const btn=ev.target.closest?.('.ticker-speed-btn'); if(btn){ev.preventDefault();ev.stopPropagation();toggleTickerSpeed(btn);}},true);

function renderAvisoTicker(title,hint,entries,opts={}){
  const arr=(entries||[]).filter(Boolean); if(!arr.length) return '';
  const icon=opts.icon||'🔔'; const color=opts.color||'rgba(245,158,11,.35)'; const safeId='ticker_'+Math.random().toString(36).slice(2);
  const doubled=arr; const accel=(usuarioAtual?.tipo==='master'||usuarioAtual?.is_viewer||usuarioAtual?.roleLabel==='Diretor Comercial')?`<button class="btn soft btn-xs ticker-speed-btn" onclick="toggleTickerSpeed('${safeId}',this)">⚡ Acelerar</button>`:'';
  return `<div class="glass panel aviso-rotativo" style="margin-bottom:12px;padding:13px 15px;border-color:${color}"><div class="section-head" style="margin:0 0 8px"><div><h2 style="margin:0;font-size:17px">${esc(icon)} ${esc(title)}</h2><div class="hint">${esc(hint||'')}</div></div><div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap"><div class="badge">${arr.length}</div>${accel}</div></div><div id="${safeId}" class="aviso-ticker"><div class="aviso-ticker-track" style="animation-duration:900s">${doubled.map(e=>`<span class="aviso-pill"><i class="red-dot"></i><span class="ticker-main">${esc(e.nome||e.label||'')}</span><small>${esc(e.info||'')}</small></span>`).join('')}</div></div></div>`;
}

function mdlV51DestaquesCobranca(){
  try{
    const cobrarParts=[]; const filiais=flattenFiliais(); const vendedores=flattenVendedores().filter(e=>!e.access_disabled && e.participa_murais!==false);
    const calcDelta=(e)=>{const delta=Number(e.var_pago_delta||0); const prev=Math.max(Math.abs(Number(e.pago||0)-delta),1); return {delta,perc:(Math.abs(delta)/prev)*100};};
    const bestFil=filiais.filter(e=>Number(e.var_pago_delta||0)>0).sort((a,b)=>Number(b.var_pago_delta||0)-Number(a.var_pago_delta||0))[0];
    const bestVend=vendedores.filter(e=>Number(e.var_pago_delta||0)>0).sort((a,b)=>Number(b.var_pago_delta||0)-Number(a.var_pago_delta||0))[0];
    if(bestFil){const d=calcDelta(bestFil); cobrarParts.push(`<div class="glass panel highlight-pulse" style="padding:13px 15px"><div class="section-head" style="margin:0"><div><h2 style="margin:0;font-size:17px">🏆 Destaque · Filial</h2><div class="hint">${esc(filialLabel(bestFil.filial))} recebeu ${R(d.delta)} a mais que a referência anterior</div></div>${renderDeltaPill(d.delta,d.perc)}</div></div>`)}
    else {cobrarParts.push(`<div class="glass panel" style="padding:13px 15px"><h2 style="margin:0;font-size:17px">🏆 Destaque · Filial</h2><div class="hint">Nenhuma filial com crescimento positivo de recebimento na referência atual.</div></div>`)}
    if(bestVend){const d=calcDelta(bestVend); cobrarParts.push(`<div class="glass panel highlight-pulse" style="padding:13px 15px"><div class="section-head" style="margin:0"><div><h2 style="margin:0;font-size:17px">🥇 Destaque · Vendedor</h2><div class="hint">${esc(bestVend.nome)} recebeu ${R(d.delta)} a mais que a referência anterior</div></div>${renderDeltaPill(d.delta,d.perc)}</div></div>`)}
    else {cobrarParts.push(`<div class="glass panel" style="padding:13px 15px"><h2 style="margin:0;font-size:17px">🥇 Destaque · Vendedor</h2><div class="hint">Nenhum vendedor com crescimento positivo de recebimento na referência atual.</div></div>`)}
    try{cobrarParts.push(renderDuplicidadeCarteiraBanner())}catch(e){}
    try{const no=renderNoChargeAlerts(); if(no)cobrarParts.push(no)}catch(e){}
    return cobrarParts.join('') || `<div class="mdl-empty-mural">Sem alertas de cobrança no momento.</div>`;
  }catch(e){return `<div class="mdl-empty-mural">Não foi possível montar mural de cobrança.</div>`}
}
function mdlV51VendasResumo(){
  try{
    const vendedores=flattenVendedores().filter(e=>!e.access_disabled && e.participa_murais!==false), filiais=flattenFiliais();
    const topVendaVend=bestLiveSalesEntity(vendedores,'venda_filial_vendedor_meta'); const topServVend=bestLiveSalesEntity(vendedores,'servico_filial_vendedor_ouro_fob'); const topVendaFil=bestLiveSalesEntity(filiais,'venda_filial_meta'); const topServFil=bestLiveSalesEntity(filiais,'servico_filial_ouro_fob');
    const meta=renderMetaDiariaBatidaAlerts();
    const linha=`<div class="glass panel sales-panel" style="padding:14px 15px"><div class="section-head" style="margin:0 0 10px"><div><h2 style="margin:0;font-size:17px">💲 Vendas do mês</h2><div class="hint">Melhores resultados acumulados de venda e serviço do mês.</div></div></div><div class="mdl-sales-line">${topVendaVend?`<span>Venda vendedor: ${esc(topVendaVend.ent.nome)} ${R(topVendaVend.val||0)}</span>`:''}${topServVend?`<span>Serviço vendedor: ${esc(topServVend.ent.nome)} ${R(topServVend.val||0)}</span>`:''}${topVendaFil?`<span>Venda filial: ${esc(filialLabel(topVendaFil.ent.filial))} ${R(topVendaFil.val||0)}</span>`:''}${topServFil?`<span>Serviço filial: ${esc(filialLabel(topServFil.ent.filial))} ${R(topServFil.val||0)}</span>`:''}</div></div>`;
    return (meta||'')+linha;
  }catch(e){return `<div class="mdl-empty-mural">Sem resumo de vendas no momento.</div>`}
}
function mdlV51Group(cls,icon,title,hint,body){return `<div class="mdl-mural-group ${cls} compact"><div class="mdl-mural-group-head"><div><h2>${icon} ${title}</h2><div class="hint">${hint}</div></div></div><div class="mdl-mural-group-grid">${body||'<div class="mdl-empty-mural">Sem dados no momento.</div>'}</div></div>`}
function renderInicioTab(){
  if(!inicioSection) return;
  const cobra=mdlV51DestaquesCobranca();
  const reat=(()=>{try{return renderNoReactivationAlerts()||'<div class="mdl-empty-mural">Todos os responsáveis de reativação já acionaram ou não há pendência.</div>'}catch(e){return '<div class="mdl-empty-mural">Não foi possível montar reativação.</div>'}})();
  const venda=mdlV51VendasResumo();
  const aniv=(()=>{try{return renderMuralAniversariantesDia()||'<div class="mdl-empty-mural">Sem aniversariantes pendentes.</div>'}catch(e){return '<div class="mdl-empty-mural">Não foi possível montar aniversariantes.</div>'}})();
  inicioSection.innerHTML=`<div class="inicio-compact"><div class="glass panel" style="padding:12px 16px;margin-bottom:10px"><div class="section-head" style="margin:0"><div><h2 style="font-size:17px;margin:0">🏠 Início</h2><div class="hint">Resumo visual da operação. O card acima dos botões alterna os alertas principais automaticamente.</div></div></div></div><div class="mdl-mural-zone">${mdlV51Group('cobranca','📌','Mural de cobrança','Destaques, duplicidade e usuários sem cobrança.',cobra)}${mdlV51Group('reativacao','🧡','Mural de reativação','Clientes sem movimento e responsáveis sem ação.',reat)}${mdlV51Group('vendas','💲','Mural de vendas','Meta diária e melhores resultados do mês.',venda)}${mdlV51Group('aniversarios','🎂','Mural de aniversariantes','Clientes aniversariantes e responsáveis pelo envio.',aniv)}</div></div>`;
}
try{setInterval(()=>{document.querySelectorAll('.aviso-ticker-track').forEach(t=>{t.style.animationName='mdlTicker'; if(!t.style.animationDuration)t.style.animationDuration='900s';});},4000)}catch(e){}





// ===== V9.0: mobile início + salvamento de emojis nas mensagens =====
(function(){
  try{
    const st=document.createElement('style');
    st.id='mdl-v90-mobile-css';
    st.textContent=`
@media(max-width:760px){
  html,body{max-width:100%!important;overflow-x:hidden!important;}
  body{font-size:14px!important;}
  .app-shell{width:100%!important;max-width:100vw!important;padding:8px 8px 86px!important;overflow-x:hidden!important;box-sizing:border-box!important;}
  .glass.header{display:flex!important;flex-direction:column!important;align-items:stretch!important;gap:12px!important;padding:16px 14px!important;border-radius:24px!important;margin:6px 0 12px!important;}
  .brand{display:grid!important;grid-template-columns:62px minmax(0,1fr)!important;gap:12px!important;align-items:center!important;width:100%!important;}
  .brand img{width:58px!important;height:58px!important;min-width:58px!important;}
  .brand h1{font-size:23px!important;line-height:1.05!important;margin:0 0 4px!important;word-break:normal!important;}
  .brand .sub{font-size:12.5px!important;line-height:1.35!important;white-space:normal!important;word-break:normal!important;}
  .header-actions{display:flex!important;flex-wrap:wrap!important;gap:8px!important;justify-content:flex-start!important;width:100%!important;}
  .header-actions .btn,.header-actions .badge{min-height:42px!important;border-radius:16px!important;font-size:13px!important;padding:10px 12px!important;}
  #userBadge{max-width:100%!important;overflow:hidden!important;text-overflow:ellipsis!important;white-space:nowrap!important;}
  #topMural,.campaign-feature,.glass.panel{max-width:100%!important;box-sizing:border-box!important;}
  body.inicio-view .kpis,.kpis,#kpis{display:grid!important;grid-template-columns:1fr!important;gap:10px!important;width:100%!important;max-width:100%!important;margin-bottom:12px!important;overflow:visible!important;}
  body.inicio-view .kpi,.kpi{width:100%!important;min-width:0!important;max-width:100%!important;min-height:auto!important;padding:15px 16px!important;border-radius:20px!important;overflow:hidden!important;box-sizing:border-box!important;}
  .kpi .label{font-size:11px!important;letter-spacing:.12em!important;line-height:1.15!important;display:flex!important;align-items:center!important;gap:7px!important;white-space:normal!important;}
  .kpi .value{font-size:25px!important;line-height:1.1!important;white-space:normal!important;word-break:break-word!important;letter-spacing:-.03em!important;}
  .kpi .subline{font-size:12.5px!important;line-height:1.35!important;padding-right:0!important;white-space:normal!important;word-break:normal!important;}
  .kpi .kpi-pct-badge{position:static!important;display:inline-flex!important;margin-top:10px!important;min-width:76px!important;justify-content:center!important;font-size:18px!important;}
  .kpi .laranjito-card{position:absolute!important;right:12px!important;bottom:12px!important;width:58px!important;height:58px!important;opacity:.92!important;}
  .kpi.kpi-laranjito .subline{padding-right:70px!important;}
  .tabs,#masterTabs{display:flex!important;flex-wrap:nowrap!important;overflow-x:auto!important;overflow-y:hidden!important;gap:8px!important;padding:8px!important;margin-bottom:12px!important;-webkit-overflow-scrolling:touch!important;scroll-snap-type:x proximity!important;}
  .tabs .tab,#masterTabs .tab{flex:0 0 auto!important;white-space:nowrap!important;padding:10px 12px!important;font-size:13px!important;border-radius:16px!important;scroll-snap-align:start!important;}
  .inicio-compact,.mdl-mural-zone{width:100%!important;max-width:100%!important;overflow:hidden!important;}
  .mdl-mural-group{padding:12px!important;border-radius:22px!important;margin-bottom:12px!important;overflow:hidden!important;}
  .mdl-mural-group-head h2{font-size:18px!important;line-height:1.15!important;}
  .mdl-mural-group-grid{display:block!important;width:100%!important;overflow:hidden!important;}
  .aviso-rotativo,.aviso-ticker{max-width:100%!important;overflow:hidden!important;}
  .aviso-pill{min-width:220px!important;max-width:250px!important;padding:9px 10px!important;}
  .mdl-sales-line{display:flex!important;overflow-x:auto!important;flex-wrap:nowrap!important;padding-bottom:6px!important;-webkit-overflow-scrolling:touch!important;}
  .mdl-sales-line span{flex:0 0 auto!important;max-width:280px!important;white-space:normal!important;}
  .section-head{display:flex!important;flex-direction:column!important;align-items:flex-start!important;gap:8px!important;}
  .search-row,.form-grid,.campaign-grid,.meta-layout,.detail-top,.row-top,.log-row{grid-template-columns:1fr!important;display:grid!important;}
  .input-card{min-width:0!important;width:100%!important;box-sizing:border-box!important;}
  .input-card textarea{min-height:120px!important;font-size:14px!important;}
  .logs-list,.acc-body,.accordion{max-width:100%!important;overflow-x:hidden!important;}
}
@media(max-width:390px){.brand h1{font-size:21px!important}.kpi .value{font-size:22px!important}.header-actions .btn,.header-actions .badge{font-size:12px!important;padding:9px 10px!important}.kpi .laranjito-card{width:50px!important;height:50px!important}}
    `;
    document.head.appendChild(st);
  }catch(e){console.warn('V9.0 mobile css',e)}
})();

const DEFAULT_REATIVACAO_MSG_V90=`Olá, {primeiro_nome}! Tudo bem? 😊

Aqui é da Lojas MDL - Móveis do Lar. Estamos com saudades de você! Faz um tempinho que você não aparece na loja. 🥹🧡

Venha conhecer nossas novidades e aproveitar condições especiais que preparamos para nossos clientes. 👈😍

Acesse nosso site e confira nossas ofertas:
www.moveisdolar.com.br`;
const DEFAULT_ANIVERSARIO_MSG_V90=`Olá, {primeiro_nome}! Feliz aniversário! 🎂🎉🥳

Aqui é da Lojas MDL - Móveis do Lar. Desejamos muita saúde, paz e felicidades neste dia especial. 😍🙏

Preparamos condições especiais para você comemorar com a gente. Acesse nosso site e confira nossas ofertas:
www.moveisdolar.com.br

🕺🎁🎉🤩`;
function mdlV90IsOldReatMsg(s){s=String(s||'');return s.includes('Estamos com saudades de você') && (!s.includes('www.moveisdolar.com.br') || !/[😊🥹😍👈🧡]/u.test(s))}
function mdlV90IsOldAnivMsg(s){s=String(s||'');return s.includes('Feliz aniversário') && (!s.includes('www.moveisdolar.com.br') || !/[🎂🎉😍🕺🤩🥳🎁]/u.test(s))}
function mdlV90NormalizeMessageDefaults(){
  try{
    CONFIG_META=CONFIG_META||{};
    if(!String(CONFIG_META.reativacao_msg_template||'').trim() || mdlV90IsOldReatMsg(CONFIG_META.reativacao_msg_template)) CONFIG_META.reativacao_msg_template=DEFAULT_REATIVACAO_MSG_V90;
    if(!String(CONFIG_META.aniversario_msg_template||'').trim() || mdlV90IsOldAnivMsg(CONFIG_META.aniversario_msg_template)) CONFIG_META.aniversario_msg_template=DEFAULT_ANIVERSARIO_MSG_V90;
    const maps=['reativacao_msg_template_filiais','aniversario_msg_template_filiais'];
    maps.forEach(k=>{const isAniv=k.includes('aniversario'); const m=CONFIG_META[k]||{}; Object.keys(m).forEach(f=>{if(isAniv && mdlV90IsOldAnivMsg(m[f])) m[f]=DEFAULT_ANIVERSARIO_MSG_V90; if(!isAniv && mdlV90IsOldReatMsg(m[f])) m[f]=DEFAULT_REATIVACAO_MSG_V90;}); CONFIG_META[k]=m;});
  }catch(e){console.warn('normalize messages v90',e)}
}
async function mdlV90SaveConfigMeta(){
  mdlV90NormalizeMessageDefaults();
  const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json; charset=utf-8','Accept':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})});
  return await resp.json();
}
try{mdlV90NormalizeMessageDefaults();}catch(e){}
window.salvarMensagemReativacaoGlobal=async function(){
  const el=document.getElementById('reatMsgTemplate');
  CONFIG_META.reativacao_msg_template=el?el.value:reativacaoTemplateAtual();
  try{const j=await mdlV90SaveConfigMeta(); toast(j.ok?'Mensagem global de reativação salva com emojis.':'Não consegui salvar mensagem.',j.ok?'success':'warn'); if(el) el.value=CONFIG_META.reativacao_msg_template;}catch(e){toast('Falha ao salvar mensagem.','warn')}
};
window.salvarMensagemReativacaoFilial=async function(){
  const f=String(document.getElementById('reatMsgFilial')?.value||'').toUpperCase(); const el=document.getElementById('reatMsgTemplateFilial');
  if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return}
  CONFIG_META.reativacao_msg_template_filiais=CONFIG_META.reativacao_msg_template_filiais||{}; CONFIG_META.reativacao_msg_template_filiais[f]=el?el.value:'';
  try{const j=await mdlV90SaveConfigMeta(); toast(j.ok?`Mensagem da ${f} salva com emojis.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn'); if(el) el.value=CONFIG_META.reativacao_msg_template_filiais[f]||'';}catch(e){toast('Falha ao salvar mensagem por filial.','warn')}
};
window.salvarMensagemAniversarioGlobal=async function(){
  const el=document.getElementById('anivMsgTemplate');
  CONFIG_META.aniversario_msg_template=el?el.value:aniversarioTemplateAtual();
  try{const j=await mdlV90SaveConfigMeta(); toast(j.ok?'Mensagem global de aniversário salva com emojis.':'Não consegui salvar mensagem.',j.ok?'success':'warn'); if(el) el.value=CONFIG_META.aniversario_msg_template;}catch(e){toast('Falha ao salvar mensagem de aniversário.','warn')}
};
window.salvarMensagemAniversarioFilial=async function(){
  const f=String(document.getElementById('anivMsgFilial')?.value||'').toUpperCase(); const el=document.getElementById('anivMsgTemplateFilial');
  if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return}
  CONFIG_META.aniversario_msg_template_filiais=CONFIG_META.aniversario_msg_template_filiais||{}; CONFIG_META.aniversario_msg_template_filiais[f]=el?el.value:'';
  try{const j=await mdlV90SaveConfigMeta(); toast(j.ok?`Mensagem de aniversário da ${f} salva com emojis.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn'); if(el) el.value=CONFIG_META.aniversario_msg_template_filiais[f]||'';}catch(e){toast('Falha ao salvar mensagem por filial.','warn')}
};
// Ao renderizar as abas, se vierem mensagens antigas do servidor, atualiza visualmente para os padrões com emoji.
(function(){
  const wrap=(name)=>{try{const fn=window[name]||eval(name); if(typeof fn==='function' && !fn._v90){const nw=function(){mdlV90NormalizeMessageDefaults(); return fn.apply(this,arguments)}; nw._v90=true; window[name]=nw; try{eval(name+'=window["'+name+'"]')}catch(e){} }}catch(e){}};
  ['renderReativacaoTab','renderAniversariantesTab'].forEach(wrap);
})();



// ===== V9.1: mobile reforçado + mensagens WhatsApp com emoji/site + filiais fallback + salvar UTF-8 direto =====
(function mdlV91Patch(){
  try{
    const st=document.createElement('style');
    st.id='mdl-v91-mobile-css';
    st.textContent=`
@media(max-width:760px){
  html,body{max-width:100%!important;overflow-x:hidden!important;}
  body{font-size:14px!important;}
  .app,.wrap,.container,.main,.content,#app{max-width:100%!important;width:100%!important;overflow-x:hidden!important;padding-left:8px!important;padding-right:8px!important;box-sizing:border-box!important;}
  .hero,.topbar,.brand-card,.header,.dash-header,.glass,.panel{max-width:100%!important;width:100%!important;box-sizing:border-box!important;border-radius:18px!important;}
  .brand,.brand-row,.header-inner{display:flex!important;flex-direction:column!important;align-items:flex-start!important;gap:10px!important;}
  .brand h1,.hero h1{font-size:24px!important;line-height:1.1!important;white-space:normal!important;}
  .header-actions,.top-actions,.actions{width:100%!important;display:flex!important;flex-wrap:wrap!important;justify-content:flex-start!important;gap:8px!important;}
  .tabs,.nav-tabs,.menu-tabs{display:flex!important;overflow-x:auto!important;white-space:nowrap!important;gap:8px!important;padding-bottom:8px!important;}
  .tabs .tab,.nav-tabs .tab,.menu-tabs .tab{flex:0 0 auto!important;}
  .kpis,.cards-grid,.grid-cards,.dashboard-grid{display:grid!important;grid-template-columns:1fr!important;gap:12px!important;width:100%!important;}
  .kpi,.card,.metric-card{width:100%!important;min-width:0!important;max-width:100%!important;box-sizing:border-box!important;min-height:auto!important;padding:16px!important;}
  .kpi .value,.metric-card .value{font-size:28px!important;line-height:1.15!important;white-space:normal!important;overflow-wrap:anywhere!important;}
  .kpi .label,.metric-card .label{font-size:12px!important;letter-spacing:.08em!important;white-space:normal!important;}
  .kpi .hint,.metric-card .hint,.small,.muted{font-size:12px!important;white-space:normal!important;overflow-wrap:anywhere!important;}
  .laranjito-card,.mascot,.kpi img{max-width:58px!important;max-height:58px!important;right:10px!important;bottom:10px!important;}
  .search-row,.form-grid,.meta-layout,.split,.row-grid{display:grid!important;grid-template-columns:1fr!important;gap:12px!important;width:100%!important;}
  .input-card{min-width:0!important;width:100%!important;max-width:100%!important;box-sizing:border-box!important;}
  .input-card textarea,.input-card input,.input-card select{width:100%!important;max-width:100%!important;box-sizing:border-box!important;font-size:14px!important;}
  .log-row,.list-row{display:grid!important;grid-template-columns:1fr!important;gap:8px!important;}
  .section-head{display:flex!important;flex-direction:column!important;align-items:flex-start!important;gap:8px!important;}
  .faixa-title{display:flex!important;flex-direction:column!important;align-items:flex-start!important;gap:6px!important;}
  .modal,.modal-content,.dialog{max-width:96vw!important;width:96vw!important;}
}
@media(min-width:761px) and (max-width:1180px){.kpis,.cards-grid,.grid-cards,.dashboard-grid{grid-template-columns:repeat(2,minmax(0,1fr))!important;}.search-row{grid-template-columns:1fr 220px!important;}}
    `;
    document.head.appendChild(st);
  }catch(e){console.warn('mdl v91 css',e)}
})();

const MDL_FILIAIS_PADRAO_V91=['F1','F2','F3','F4','F5','F6','F8','F9'];
const DEFAULT_REATIVACAO_MSG_V91=`Olá, {primeiro_nome}! Tudo bem? 😊

Aqui é da Lojas MDL - Móveis do Lar. Estamos com saudades de você! Faz um tempinho que você não aparece na loja. 🥹🧡

Venha conhecer nossas novidades e aproveitar condições especiais que preparamos para nossos clientes. 👈😍

Acesse nosso site e confira nossas ofertas:
www.moveisdolar.com.br`;
const DEFAULT_ANIVERSARIO_MSG_V91=`Olá, {primeiro_nome}! Feliz aniversário! 🎂🎉🥳

Aqui é da Lojas MDL - Móveis do Lar. Desejamos muita saúde, paz e felicidades neste dia especial. 😍🙏

Preparamos condições especiais para você comemorar com a gente. Acesse nosso site e confira nossas ofertas:
www.moveisdolar.com.br

🕺🎁🎉🤩`;
function mdlV91TemEmoji(s){try{return /[\u{1F300}-\u{1FAFF}\u{2600}-\u{27BF}]/u.test(String(s||''));}catch(e){return false}}
function mdlV91EnsureDefaults(){
  try{
    CONFIG_META=CONFIG_META||{};
    const rg=String(CONFIG_META.reativacao_msg_template||'').trim();
    const ag=String(CONFIG_META.aniversario_msg_template||'').trim();
    if(!rg || (rg.includes('Estamos com saudades') && (!rg.includes('www.moveisdolar.com.br') || !mdlV91TemEmoji(rg)))) CONFIG_META.reativacao_msg_template=DEFAULT_REATIVACAO_MSG_V91;
    if(!ag || (ag.includes('Feliz aniversário') && (!ag.includes('www.moveisdolar.com.br') || !mdlV91TemEmoji(ag)))) CONFIG_META.aniversario_msg_template=DEFAULT_ANIVERSARIO_MSG_V91;
  }catch(e){console.warn('mdl v91 defaults',e)}
}
function mdlV91FiliaisFromRows(rows){
  const s=new Set();
  try{(rows||[]).forEach(r=>{const f=String(r?.filial||'').toUpperCase(); if(/^F\d+$/.test(f)) s.add(f);});}catch(e){}
  try{Object.keys(CONFIG_META?.reativacao_msg_template_filiais||{}).forEach(f=>/^F\d+$/.test(String(f).toUpperCase())&&s.add(String(f).toUpperCase()));}catch(e){}
  try{Object.keys(CONFIG_META?.aniversario_msg_template_filiais||{}).forEach(f=>/^F\d+$/.test(String(f).toUpperCase())&&s.add(String(f).toUpperCase()));}catch(e){}
  MDL_FILIAIS_PADRAO_V91.forEach(f=>s.add(f));
  return ordenarFiliaisReativacao([...s]);
}
async function mdlV91SaveConfigDireto(){
  const resp=await fetch(API_CFG,{method:'POST',headers:{'Content-Type':'application/json; charset=utf-8','Accept':'application/json'},body:JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND})});
  return await resp.json();
}
function mdlV91SyncMessageAreas(){
  const r=document.getElementById('reatMsgTemplate'); if(r && (!String(r.value||'').trim() || !String(r.value||'').includes('www.moveisdolar.com.br') || !mdlV91TemEmoji(r.value))) r.value=CONFIG_META.reativacao_msg_template||DEFAULT_REATIVACAO_MSG_V91;
  const a=document.getElementById('anivMsgTemplate'); if(a && (!String(a.value||'').trim() || !String(a.value||'').includes('www.moveisdolar.com.br') || !mdlV91TemEmoji(a.value))) a.value=CONFIG_META.aniversario_msg_template||DEFAULT_ANIVERSARIO_MSG_V91;
}
window.salvarMensagemReativacaoGlobal=async function(){
  const el=document.getElementById('reatMsgTemplate');
  CONFIG_META.reativacao_msg_template=el?el.value:DEFAULT_REATIVACAO_MSG_V91;
  try{const j=await mdlV91SaveConfigDireto(); toast(j.ok?'Mensagem global de reativação salva com emojis.':'Não consegui salvar mensagem.',j.ok?'success':'warn');}catch(e){toast('Falha ao salvar mensagem.','warn')}
};
window.salvarMensagemReativacaoFilial=async function(){
  const f=String(document.getElementById('reatMsgFilial')?.value||'').toUpperCase(); const el=document.getElementById('reatMsgTemplateFilial');
  if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return}
  CONFIG_META.reativacao_msg_template_filiais=CONFIG_META.reativacao_msg_template_filiais||{}; CONFIG_META.reativacao_msg_template_filiais[f]=el?el.value:'';
  try{const j=await mdlV91SaveConfigDireto(); toast(j.ok?`Mensagem da ${f} salva com emojis.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn');}catch(e){toast('Falha ao salvar mensagem por filial.','warn')}
};
window.salvarMensagemAniversarioGlobal=async function(){
  const el=document.getElementById('anivMsgTemplate');
  CONFIG_META.aniversario_msg_template=el?el.value:DEFAULT_ANIVERSARIO_MSG_V91;
  try{const j=await mdlV91SaveConfigDireto(); toast(j.ok?'Mensagem global de aniversário salva com emojis.':'Não consegui salvar mensagem.',j.ok?'success':'warn');}catch(e){toast('Falha ao salvar mensagem de aniversário.','warn')}
};
window.salvarMensagemAniversarioFilial=async function(){
  const f=String(document.getElementById('anivMsgFilial')?.value||'').toUpperCase(); const el=document.getElementById('anivMsgTemplateFilial');
  if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return}
  CONFIG_META.aniversario_msg_template_filiais=CONFIG_META.aniversario_msg_template_filiais||{}; CONFIG_META.aniversario_msg_template_filiais[f]=el?el.value:'';
  try{const j=await mdlV91SaveConfigDireto(); toast(j.ok?`Mensagem de aniversário da ${f} salva com emojis.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn');}catch(e){toast('Falha ao salvar mensagem por filial.','warn')}
};
// Corrige selects de filiais quando a lista "novos do dia" vier vazia.
(function(){
  mdlV91EnsureDefaults();
  const wrap=(name)=>{try{const fn=window[name]||eval(name); if(typeof fn==='function' && !fn._v91){const nw=function(){mdlV91EnsureDefaults(); const ret=fn.apply(this,arguments); setTimeout(()=>{try{
      const rf=document.getElementById('reatMsgFilial'); if(rf && rf.options.length===0){rf.innerHTML=mdlV91FiliaisFromRows(CLIENTES_SEM_MOVIMENTO_BASE||CLIENTES_SEM_MOVIMENTO||[]).map(f=>`<option value="${f}">${f}</option>`).join(''); trocarMensagemReativacaoFilial(rf.value||'F1');}
      const af=document.getElementById('anivMsgFilial'); if(af && af.options.length===0){af.innerHTML=mdlV91FiliaisFromRows(window.ANIVERSARIANTES||[]).map(f=>`<option value="${f}">${f}</option>`).join(''); trocarMensagemAniversarioFilial(af.value||'F1');}
      mdlV91SyncMessageAreas();
    }catch(e){}},60); return ret;}; nw._v91=true; window[name]=nw; try{eval(name+'=window["'+name+'"]')}catch(e){} }}catch(e){}};
  ['renderReativacaoTab','renderAniversariantesTab','renderInicioTab'].forEach(wrap);
})();



// ===== V9.2 HOTFIX: emojis persistentes + reativação com retorno 10 dias =====
(function(){
  try{
    const st=document.createElement('style'); st.id='mdl-v92-mobile-css';
    st.textContent=`@media(max-width:760px){.mobile .kpis,.kpis{grid-template-columns:1fr!important}.mobile .kpi,.kpi{min-width:0!important;max-width:100%!important}.search-row{grid-template-columns:1fr!important}.input-card textarea{min-height:130px!important}.topbar,.hero,.header-actions{max-width:100vw!important}.nav-tabs,.tabs{overflow-x:auto!important;white-space:nowrap!important}}`;
    document.head.appendChild(st);
  }catch(e){}
})();
function mdlV92MaybeForceDefaultMessages(){
  try{
    // Só troca textos antigos do sistema sem emoji/site; não sobrescreve texto manual que o usuário digitou com emoji.
    const r=String(CONFIG_META.reativacao_msg_template||'');
    const a=String(CONFIG_META.aniversario_msg_template||'');
    if(!r.trim() || (!mdlV91TemEmoji(r) && r.includes('Estamos com saudades'))) CONFIG_META.reativacao_msg_template=DEFAULT_REATIVACAO_MSG_V91;
    if(!a.trim() || (!mdlV91TemEmoji(a) && a.includes('Feliz aniversário'))) CONFIG_META.aniversario_msg_template=DEFAULT_ANIVERSARIO_MSG_V91;
  }catch(e){}
}
async function mdlV92SaveConfigComEmoji(){
  const payload=JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND});
  const resp=await fetch(API_CFG+'?utf8=1&_=' + Date.now(),{method:'POST',cache:'no-store',headers:{'Content-Type':'application/json;charset=UTF-8','Accept':'application/json, text/plain, */*'},body:payload});
  const txt=await resp.text();
  try{return JSON.parse(txt)}catch(e){return {ok:resp.ok,raw:txt}}
}
window.salvarMensagemReativacaoGlobal=async function(){
  const el=document.getElementById('reatMsgTemplate');
  CONFIG_META.reativacao_msg_template=el?el.value:DEFAULT_REATIVACAO_MSG_V91;
  try{const j=await mdlV92SaveConfigComEmoji(); toast(j.ok?'Mensagem global de reativação salva com emojis.':'Não consegui salvar mensagem.',j.ok?'success':'warn'); if(el) el.value=CONFIG_META.reativacao_msg_template;}catch(e){toast('Falha ao salvar mensagem.','warn')}
};
window.salvarMensagemReativacaoFilial=async function(){
  const f=String(document.getElementById('reatMsgFilial')?.value||'').toUpperCase(); const el=document.getElementById('reatMsgTemplateFilial');
  if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return}
  CONFIG_META.reativacao_msg_template_filiais=CONFIG_META.reativacao_msg_template_filiais||{}; CONFIG_META.reativacao_msg_template_filiais[f]=el?el.value:'';
  try{const j=await mdlV92SaveConfigComEmoji(); toast(j.ok?`Mensagem da ${f} salva com emojis.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn');}catch(e){toast('Falha ao salvar mensagem por filial.','warn')}
};
window.salvarMensagemAniversarioGlobal=async function(){
  const el=document.getElementById('anivMsgTemplate');
  CONFIG_META.aniversario_msg_template=el?el.value:DEFAULT_ANIVERSARIO_MSG_V91;
  try{const j=await mdlV92SaveConfigComEmoji(); toast(j.ok?'Mensagem global de aniversário salva com emojis.':'Não consegui salvar mensagem.',j.ok?'success':'warn'); if(el) el.value=CONFIG_META.aniversario_msg_template;}catch(e){toast('Falha ao salvar mensagem de aniversário.','warn')}
};
window.salvarMensagemAniversarioFilial=async function(){
  const f=String(document.getElementById('anivMsgFilial')?.value||'').toUpperCase(); const el=document.getElementById('anivMsgTemplateFilial');
  if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return}
  CONFIG_META.aniversario_msg_template_filiais=CONFIG_META.aniversario_msg_template_filiais||{}; CONFIG_META.aniversario_msg_template_filiais[f]=el?el.value:'';
  try{const j=await mdlV92SaveConfigComEmoji(); toast(j.ok?`Mensagem de aniversário da ${f} salva com emojis.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn');}catch(e){toast('Falha ao salvar mensagem por filial.','warn')}
};
(function(){
  mdlV92MaybeForceDefaultMessages();
  setTimeout(()=>{try{mdlV91SyncMessageAreas();}catch(e){}},120);
})();


// ===== V9.3 HOTFIX: salvar mensagem por filial + emojis via base64 UTF-8 =====
function mdlV93B64Utf8(str){
  const bytes=new TextEncoder().encode(String(str||''));
  let bin='';
  bytes.forEach(b=>bin+=String.fromCharCode(b));
  return btoa(bin);
}
function mdlV93FilialKey(v){
  const m=String(v||'').toUpperCase().match(/F\d+/);
  return m?m[0]:String(v||'').toUpperCase().trim();
}
function mdlV93GetMap(kind){
  CONFIG_META=CONFIG_META||{};
  const key=kind==='aniv'?'aniversario_msg_template_filiais':'reativacao_msg_template_filiais';
  if(!CONFIG_META[key] || typeof CONFIG_META[key]!=='object' || Array.isArray(CONFIG_META[key])) CONFIG_META[key]={};
  return CONFIG_META[key];
}
async function mdlV93SaveConfigB64(){
  const payload=JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND});
  const fd=new FormData();
  fd.append('payload_b64', mdlV93B64Utf8(payload));
  fd.append('v','9.3');
  const resp=await fetch(API_CFG+'?v=9.3&_='+Date.now(),{method:'POST',cache:'no-store',body:fd});
  const txt=await resp.text();
  try{return JSON.parse(txt)}catch(e){return {ok:resp.ok,raw:txt}}
}
function mdlV93LocalBackup(kind, filial, text){
  try{localStorage.setItem('mdl_msg_'+kind+'_'+mdlV93FilialKey(filial), String(text||''));}catch(e){}
}
function mdlV93LocalRead(kind, filial){
  try{return localStorage.getItem('mdl_msg_'+kind+'_'+mdlV93FilialKey(filial))||''}catch(e){return ''}
}
const _v93ReatTplBase = typeof reativacaoTemplateAtual==='function' ? reativacaoTemplateAtual : null;
reativacaoTemplateAtual=function(filial=''){
  const f=mdlV93FilialKey(filial);
  const map=mdlV93GetMap('reat');
  if(f && String(map[f]||'').trim()) return String(map[f]);
  const lb=mdlV93LocalRead('reat',f);
  if(f && lb.trim()) return lb;
  return _v93ReatTplBase ? _v93ReatTplBase(f) : (CONFIG_META.reativacao_msg_template||DEFAULT_REATIVACAO_MSG_V91||'');
};
const _v93AnivTplBase = typeof aniversarioTemplateAtual==='function' ? aniversarioTemplateAtual : null;
aniversarioTemplateAtual=function(filial=''){
  const f=mdlV93FilialKey(filial);
  const map=mdlV93GetMap('aniv');
  if(f && String(map[f]||'').trim()) return String(map[f]);
  const lb=mdlV93LocalRead('aniv',f);
  if(f && lb.trim()) return lb;
  return _v93AnivTplBase ? _v93AnivTplBase(f) : (CONFIG_META.aniversario_msg_template||DEFAULT_ANIVERSARIO_MSG_V91||'');
};
window.salvarMensagemReativacaoGlobal=async function(){
  const el=document.getElementById('reatMsgTemplate');
  const val=el?el.value:(CONFIG_META.reativacao_msg_template||DEFAULT_REATIVACAO_MSG_V91||'');
  CONFIG_META.reativacao_msg_template=val;
  try{
    const j=await mdlV93SaveConfigB64();
    if(j.ok){await carregarConfigOnline(); CONFIG_META.reativacao_msg_template=val; if(el) el.value=val;}
    toast(j.ok?'Mensagem global de reativação salva no servidor.':'Não consegui salvar mensagem global.',j.ok?'success':'warn');
  }catch(e){console.warn(e); toast('Falha ao salvar mensagem global.','warn')}
};
window.salvarMensagemReativacaoFilial=async function(){
  const f=mdlV93FilialKey(document.getElementById('reatMsgFilial')?.value||'');
  const el=document.getElementById('reatMsgTemplateFilial');
  if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return}
  const val=el?el.value:'';
  const map=mdlV93GetMap('reat'); map[f]=val;
  mdlV93LocalBackup('reat',f,val);
  try{
    const j=await mdlV93SaveConfigB64();
    if(j.ok){await carregarConfigOnline(); mdlV93GetMap('reat')[f]=val; if(el) el.value=val;}
    toast(j.ok?`Mensagem da ${f} salva no servidor.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn');
  }catch(e){console.warn(e); toast('Falha ao salvar mensagem por filial.','warn')}
};
window.salvarMensagemAniversarioGlobal=async function(){
  const el=document.getElementById('anivMsgTemplate');
  const val=el?el.value:(CONFIG_META.aniversario_msg_template||DEFAULT_ANIVERSARIO_MSG_V91||'');
  CONFIG_META.aniversario_msg_template=val;
  try{
    const j=await mdlV93SaveConfigB64();
    if(j.ok){await carregarConfigOnline(); CONFIG_META.aniversario_msg_template=val; if(el) el.value=val;}
    toast(j.ok?'Mensagem global de aniversário salva no servidor.':'Não consegui salvar mensagem global.',j.ok?'success':'warn');
  }catch(e){console.warn(e); toast('Falha ao salvar mensagem de aniversário.','warn')}
};
window.salvarMensagemAniversarioFilial=async function(){
  const f=mdlV93FilialKey(document.getElementById('anivMsgFilial')?.value||'');
  const el=document.getElementById('anivMsgTemplateFilial');
  if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return}
  const val=el?el.value:'';
  const map=mdlV93GetMap('aniv'); map[f]=val;
  mdlV93LocalBackup('aniv',f,val);
  try{
    const j=await mdlV93SaveConfigB64();
    if(j.ok){await carregarConfigOnline(); mdlV93GetMap('aniv')[f]=val; if(el) el.value=val;}
    toast(j.ok?`Mensagem de aniversário da ${f} salva no servidor.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn');
  }catch(e){console.warn(e); toast('Falha ao salvar mensagem por filial.','warn')}
};
function trocarMensagemReativacaoFilial(f){const el=document.getElementById('reatMsgTemplateFilial'); if(el) el.value=reativacaoTemplateAtual(mdlV93FilialKey(f));}
function trocarMensagemAniversarioFilial(f){const el=document.getElementById('anivMsgTemplateFilial'); if(el) el.value=aniversarioTemplateAtual(mdlV93FilialKey(f));}

// ===== V5.5: META DIARIA PRECALCULADA PELO PYTHON + FALLBACK ROBUSTO =====
(function(){
  try{
    const st=document.createElement('style');
    st.textContent=`
      .mdl-hero-mural-card{grid-column:span 2!important;min-height:128px!important;padding:16px 18px!important;border-color:rgba(245,158,11,.42)!important;}
      .mdl-hero-mural-card.v52-cobranca{border-color:rgba(239,68,68,.45)!important;background:linear-gradient(135deg,rgba(127,29,29,.28),rgba(15,23,42,.94))!important}
      .mdl-hero-mural-card.v52-reativacao{border-color:rgba(245,158,11,.48)!important;background:linear-gradient(135deg,rgba(120,53,15,.30),rgba(15,23,42,.94))!important}
      .mdl-hero-mural-card.v52-aniversario{border-color:rgba(236,72,153,.46)!important;background:linear-gradient(135deg,rgba(131,24,67,.26),rgba(15,23,42,.94))!important}
      .mdl-hero-mural-card.v52-meta{border-color:rgba(34,197,94,.48)!important;background:linear-gradient(135deg,rgba(6,78,59,.28),rgba(15,23,42,.94))!important}
      .mdl-hero-main{font-size:20px!important;line-height:1.12!important;max-width:78%;}
      .mdl-hero-detail{font-size:12.5px!important;max-width:82%!important;display:block!important;}
      .mdl-hero-mini-list{display:flex;gap:8px;margin-top:8px;max-width:88%;overflow:hidden;position:relative;z-index:2}
      .mdl-hero-mini{border:1px solid rgba(255,255,255,.10);background:rgba(255,255,255,.045);border-radius:999px;padding:6px 9px;font-size:10.5px;font-weight:850;color:#cbd5e1;white-space:nowrap;max-width:210px;overflow:hidden;text-overflow:ellipsis}
      .mdl-mural-group.cobranca .glass.panel{border-color:rgba(239,68,68,.32)!important}.mdl-mural-group.reativacao .glass.panel{border-color:rgba(245,158,11,.34)!important}.mdl-mural-group.vendas .glass.panel{border-color:rgba(34,197,94,.32)!important}.mdl-mural-group.aniversarios .glass.panel{border-color:rgba(236,72,153,.32)!important}
      .aviso-ticker{overflow-x:auto!important;overflow-y:hidden!important;scrollbar-width:thin;cursor:grab!important;user-select:none;scroll-behavior:auto!important;}
      .aviso-ticker.dragging{cursor:grabbing!important}.aviso-ticker.dragging .aviso-ticker-track,.aviso-ticker:hover .aviso-ticker-track{animation-play-state:paused!important}
      .aviso-ticker::-webkit-scrollbar{height:8px}.aviso-ticker::-webkit-scrollbar-thumb{background:rgba(255,255,255,.18);border-radius:999px}.aviso-ticker::-webkit-scrollbar-track{background:rgba(255,255,255,.04);border-radius:999px}
      .aviso-pill{min-width:230px!important;max-width:290px!important;padding:9px 12px!important;}
      .aviso-pill .ticker-main{font-size:12px!important;font-weight:950!important;}
      .aviso-pill small{font-size:10.5px!important;opacity:.92!important;}
      .mdl-mural-zone{gap:18px!important}.mdl-mural-group{padding:16px 18px!important}.mdl-mural-group-head h2{letter-spacing:-.01em}.mdl-mural-group.cobranca{box-shadow:0 0 0 1px rgba(239,68,68,.10),0 20px 48px rgba(127,29,29,.13)}.mdl-mural-group.reativacao{box-shadow:0 0 0 1px rgba(245,158,11,.10),0 20px 48px rgba(120,53,15,.13)}.mdl-mural-group.vendas{box-shadow:0 0 0 1px rgba(34,197,94,.10),0 20px 48px rgba(6,78,59,.13)}.mdl-mural-group.aniversarios{box-shadow:0 0 0 1px rgba(236,72,153,.10),0 20px 48px rgba(131,24,67,.13)}
      @media(max-width:1180px){.mdl-hero-main,.mdl-hero-detail,.mdl-hero-mini-list{max-width:100%!important}.mdl-hero-mural-card{grid-column:1/-1!important}}
    `;
    document.head.appendChild(st);
  }catch(e){console.warn('v5.2 css',e)}
})();

function mdlV52EntriesFromTickerHtml(html){
  try{
    const tmp=document.createElement('div'); tmp.innerHTML=String(html||'');
    return [...tmp.querySelectorAll('.aviso-pill')].map(p=>({nome:(p.querySelector('.ticker-main')?.textContent||p.textContent||'').trim(), info:(p.querySelector('small')?.textContent||'').trim()})).filter(x=>x.nome);
  }catch(e){return []}
}
function mdlV52NumFromPctText(v){ return salesNum ? salesNum(v) : (parseFloat(String(v||'').replace('%','').replace(',','.'))||0); }
function mdlV52MetaDiariaEntries(){
  const arr=[];
  const wrap=(METAS_VENDAS_DIA&&METAS_VENDAS_DIA.metas)?METAS_VENDAS_DIA:{metas:{}};
  function rowAtingPeriodo(row){
    let v=salesCell(row,['Atingido Período','Atingido Periodo','Atingido(R$) Período','Atingido(R$) Periodo']);
    if(v==='' || v==null){
      for(const k of Object.keys(row||{})){
        const ku=String(k).toUpperCase();
        if(ku.includes('ATINGIDO') && (ku.includes('PER')||ku.includes('PERÍODO')) && ku.endsWith('_FLOAT')) return Number(row[k]||0);
      }
    }
    return mdlV52NumFromPctText(v);
  }
  function rowRealPeriodo(row){return salesCell(row,['Realizado (R$) Período','Realizado(R$) Período','Realizado (R$) Periodo','Realizado(R$) Periodo'])||'';}
  function rowMetaPeriodo(row){return salesCell(row,['Meta (R$) Período','Meta(R$) Período','Meta (R$) Periodo','Meta(R$) Periodo'])||'';}
  try{
    const mf=wrap.metas?.venda_filial_meta?.linhas||[];
    mf.forEach(row=>{
      if(row?._is_total) return;
      const ating=rowAtingPeriodo(row);
      if(ating>=100){
        const fil=_filialFromAny ? _filialFromAny(salesCell(row,['Filial','Nome','Vendedor'])||'') : '';
        arr.push({nome:filialLabel(fil)||salesCell(row,['Filial'])||'Filial', info:`${fil||''} · ${ating.toFixed(2).replace('.',',')}% no dia · ${rowRealPeriodo(row)||'R$ 0,00'} / ${rowMetaPeriodo(row)||'R$ 0,00'}`, kind:'filial', ating});
      }
    });
    const mv=wrap.metas?.venda_filial_vendedor_meta?.linhas||[];
    mv.forEach(row=>{
      if(row?._is_total) return;
      const ating=rowAtingPeriodo(row);
      if(ating>=100){
        const nome=salesCell(row,['Vendedor_2','Nome_2','Vendedor','Nome'])||'Vendedor';
        const fil=_filialFromAny ? _filialFromAny(salesCell(row,['Filial','Vendedor'])||'') : '';
        arr.push({nome:nome, info:`${fil||''} · ${ating.toFixed(2).replace('.',',')}% no dia · ${rowRealPeriodo(row)||'R$ 0,00'} / ${rowMetaPeriodo(row)||'R$ 0,00'}`, kind:'vendedor', ating});
      }
    });
  }catch(e){console.warn('mdlV52MetaDiariaEntries SGI dia',e)}
  arr.sort((a,b)=>Number(b.ating||0)-Number(a.ating||0)||String(a.nome).localeCompare(String(b.nome),'pt-BR'));
  return arr;
}
function renderMetaDiariaBatidaAlerts(){
  try{
    const arr=mdlV52MetaDiariaEntries();
    if(!arr.length){
      return `<div class="glass panel" style="margin-bottom:16px;padding:14px 18px;border-color:rgba(34,197,94,.22)"><div class="section-head" style="margin:0 0 8px"><div><h2 style="margin:0;font-size:18px">🎯 Meta diária BATIDA</h2><div class="hint">Mural ativo. Usa o Controle de Meta do Sólidus filtrado no dia atual: quem ficar com Atingido Período acima de 100% aparece aqui.</div></div><div class="badge">0</div></div><div class="meta-diaria-empty">Nenhuma meta diária batida até o momento.</div></div>`;
    }
    return renderAvisoTicker('Meta diária BATIDA','Controle de Meta Sólidus do dia atual: Atingido Período acima de 100%.', arr, {icon:'🎯',color:'rgba(34,197,94,.34)'});
  }catch(e){console.warn('renderMetaDiariaBatidaAlerts v52',e); return '';}
}
function mdlV51MetaDiariaCount(){try{return mdlV52MetaDiariaEntries().length}catch(e){return 0}}

function mdlV52SlideItemsFromHtml(title,icon,html,emptyMain,emptyDetail,dot,kind){
  const items=mdlV52EntriesFromTickerHtml(html);
  if(!items.length) return [{icon,title,count:0,main:emptyMain,detail:emptyDetail,dot,kind,mini:[]}];
  return items.slice(0,12).map((it,idx)=>({icon,title,count:items.length,main:it.nome,detail:it.info,dot,kind,mini:items.slice(idx+1,idx+4).map(x=>x.nome)}));
}
function mdlV51HeroSlides(){
  let slides=[];
  const noCobHtml=(()=>{try{return renderNoChargeAlerts()||''}catch(e){return ''}})();
  const noReatHtml=(()=>{try{return renderNoReactivationAlerts()||''}catch(e){return ''}})();
  const metaHtml=(()=>{try{return renderMetaDiariaBatidaAlerts()||''}catch(e){return ''}})();
  const anivHtml=(()=>{try{return renderMuralAniversariantesDia()||''}catch(e){return ''}})();
  slides=slides.concat(mdlV52SlideItemsFromHtml('Sem cobranças hoje','⏰',noCobHtml,'Cobranças em dia','Nenhum alerta crítico de cobrança agora.','#ef4444','cobranca'));
  slides=slides.concat(mdlV52SlideItemsFromHtml('Clientes sem movimento','🧡',noReatHtml,'Reativação em dia','Todos os responsáveis acionaram ou não há lista pendente.','#f59e0b','reativacao'));
  slides=slides.concat(mdlV52SlideItemsFromHtml('Aniversariantes do dia','🎂',anivHtml,'Aniversariantes em dia','Nenhuma saudação pendente no momento.','#ec4899','aniversario'));
  slides=slides.concat(mdlV52SlideItemsFromHtml('Meta diária BATIDA','🎯',metaHtml,'Aguardando metas diárias','Quem passar de 100% no período do Sólidus aparece aqui.','#22c55e','meta'));
  return slides.length?slides:[{icon:'📌',title:'Mural operacional',count:0,main:'Operação em andamento',detail:'Resumo automático dos murais do dia.',dot:'#f59e0b',kind:'cobranca',mini:[]}];
}
function mdlV51HeroSlideHtml(s){
  const mini=(s.mini||[]).slice(0,3).map(x=>`<span class="mdl-hero-mini">${esc(x)}</span>`).join('');
  return `<div class="mdl-hero-title"><span class="mdl-hero-dot" style="background:${esc(s.dot||'#f59e0b')}"></span>${esc(s.icon||'🔔')} ${esc(s.title||'Mural operacional')}</div><div class="mdl-hero-main">${esc(s.main||'Operação em andamento')}</div><div class="mdl-hero-detail">${esc(s.detail||'Resumo automático dos murais do dia.')}</div>${mini?`<div class="mdl-hero-mini-list">${mini}</div>`:''}`;
}
function mdlV51HeroHtml(){const slides=mdlV51HeroSlides(); const k=slides[0]?.kind||'cobranca'; return `<div id="mdlHeroMural" class="glass kpi mdl-hero-mural-card v52-${esc(k)}" data-idx="0">${mdlV51HeroSlideHtml(slides[0]||{})}</div>`}
function mdlV51StartHeroMural(){
  try{
    const el=document.getElementById('mdlHeroMural'); if(!el) return;
    if(window._mdlHeroTimer) clearInterval(window._mdlHeroTimer);
    window._mdlHeroTimer=setInterval(()=>{try{const fresh=mdlV51HeroSlides(); const idx=(Number(el.dataset.idx||0)+1)%Math.max(fresh.length,1); const s=fresh[idx]||{}; el.dataset.idx=String(idx); el.className='glass kpi mdl-hero-mural-card v52-'+String(s.kind||'cobranca'); el.innerHTML=mdlV51HeroSlideHtml(s);}catch(e){}},3600);
  }catch(e){}
}

function renderAvisoTicker(title,hint,entries,opts={}){
  const arr=(entries||[]).filter(Boolean); if(!arr.length) return '';
  const icon=opts.icon||'🔔'; const color=opts.color||'rgba(245,158,11,.35)'; const safeId='ticker_'+Math.random().toString(36).slice(2);
  const doubled=arr; const accel=(usuarioAtual?.tipo==='master'||usuarioAtual?.is_viewer||usuarioAtual?.roleLabel==='Diretor Comercial')?`<button class="btn soft btn-xs ticker-speed-btn" onclick="toggleTickerSpeed('${safeId}',this)">⚡ Acelerar</button>`:'';
  return `<div class="glass panel aviso-rotativo" style="margin-bottom:12px;padding:13px 15px;border-color:${color}"><div class="section-head" style="margin:0 0 8px"><div><h2 style="margin:0;font-size:17px">${esc(icon)} ${esc(title)}</h2><div class="hint">${esc(hint||'')}</div></div><div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap"><div class="badge">${arr.length}</div>${accel}</div></div><div id="${safeId}" class="aviso-ticker"><div class="aviso-ticker-track" style="animation-duration:900s">${doubled.map(e=>`<span class="aviso-pill"><i class="red-dot"></i><span class="ticker-main">${esc(e.nome||e.label||'')}</span><small>${esc(e.info||'')}</small></span>`).join('')}</div></div></div>`;
}
function toggleTickerSpeed(arg1,arg2){
  try{
    let box=null,btn=null;
    if(typeof arg1==='string'){box=document.getElementById(arg1);btn=arg2||null}else{btn=arg1||null;box=btn?.closest('.glass.panel,.aviso-rotativo,.mdl-mural-group')?.querySelector('.aviso-ticker')}
    if(!box && btn) box=btn.closest?.('.glass.panel,.aviso-rotativo,.mdl-mural-group')?.querySelector?.('.aviso-ticker');
    if(!box) return;
    const fast=!box.classList.contains('fast'); box.classList.toggle('fast',fast);
    const track=box.querySelector('.aviso-ticker-track'); if(track){track.style.animationName='mdlTicker'; track.style.animationDuration=fast?'18s':'900s'; track.style.animationPlayState='running';}
    if(btn){btn.textContent=fast?'🐢 Normal':'⚡ Acelerar'; btn.classList.toggle('primary',fast)}
  }catch(e){console.warn('toggle ticker v52',e)}
}
(function(){
  let active=null,startX=0,startScroll=0;
  document.addEventListener('mousedown',function(ev){const box=ev.target.closest?.('.aviso-ticker'); if(!box) return; active=box; startX=ev.clientX; startScroll=box.scrollLeft; box.classList.add('dragging'); const tr=box.querySelector('.aviso-ticker-track'); if(tr)tr.style.animationPlayState='paused';},true);
  document.addEventListener('mousemove',function(ev){if(!active)return; active.scrollLeft=startScroll-(ev.clientX-startX);},true);
  document.addEventListener('mouseup',function(){if(!active)return; const tr=active.querySelector('.aviso-ticker-track'); if(tr && !active.matches(':hover'))tr.style.animationPlayState='running'; active.classList.remove('dragging'); active=null;},true);
  document.addEventListener('mouseleave',function(){if(!active)return; active.classList.remove('dragging'); active=null;},true);
})();
try{setInterval(()=>{document.querySelectorAll('.aviso-ticker-track').forEach(t=>{t.style.animationName='mdlTicker'; if(!t.style.animationDuration)t.style.animationDuration='900s'; if(!t.closest('.aviso-ticker')?.classList.contains('dragging')) t.style.animationPlayState='running';});},5000)}catch(e){}



// ===== V5.5 HOTFIX: usa lista pré-calculada pelo Python para Meta diária BATIDA =====
window.META_DIARIA_BATIDA_PRECALC = __JS_META_DIARIA_BATIDA_PRECALC__ || [];
const _v54_oldMetaDiariaEntries = (typeof mdlV52MetaDiariaEntries === 'function') ? mdlV52MetaDiariaEntries : null;
function _v54FormatPct(v){
  const n=Number(v||0);
  return Number.isFinite(n) ? n.toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2})+'%' : '0,00%';
}
function mdlV52MetaDiariaEntries(){
  try{
    const pre = Array.isArray(window.META_DIARIA_BATIDA_PRECALC) ? window.META_DIARIA_BATIDA_PRECALC : [];
    if(pre.length){
      return pre.map(x=>({
        nome: String(x.nome||x.label||''),
        info: String(x.info||`${x.filial||''} · ${_v54FormatPct(x.ating)} no dia`),
        kind: x.kind||x.tipo||'',
        filial: x.filial||'',
        ating: Number(x.ating||0)
      })).filter(x=>x.nome).sort((a,b)=>Number(b.ating||0)-Number(a.ating||0)||String(a.nome).localeCompare(String(b.nome),'pt-BR'));
    }
  }catch(e){console.warn('META_DIARIA_BATIDA_PRECALC falhou',e)}
  // V10.1: não usa fallback antigo da meta diária, porque ele podia ler Projetado/Total como Realizado Período.
  // Se o pré-cálculo Python vier vazio, a regra correta é mostrar zero meta diária batida.
  return [];
}
function mdlV51MetaDiariaCount(){try{return mdlV52MetaDiariaEntries().length}catch(e){return 0}}
function renderMetaDiariaBatidaAlerts(){
  try{
    const arr=mdlV52MetaDiariaEntries();
    if(!arr.length){
      return `<div class="glass panel" style="margin-bottom:16px;padding:14px 18px;border-color:rgba(34,197,94,.22)"><div class="section-head" style="margin:0 0 8px"><div><h2 style="margin:0;font-size:18px">🎯 Meta diária BATIDA</h2><div class="hint">Mural ativo. Usa o Controle de Meta do Sólidus filtrado no dia atual: quem ficar com Atingido Período acima de 100% aparece aqui.</div></div><div class="badge">0</div></div><div class="meta-diaria-empty">Nenhuma meta diária batida até o momento.</div></div>`;
    }
    return renderAvisoTicker('Meta diária BATIDA','Controle de Meta Sólidus do dia atual: Atingido Período acima de 100%.', arr, {icon:'🎯',color:'rgba(34,197,94,.34)'});
  }catch(e){console.warn('renderMetaDiariaBatidaAlerts v54',e); return ''}
}

// ===== V5.5 HOTFIX: notificações individuais usam a lista pré-calculada da meta diária =====
const _v55_oldGoalNotifsFor = (typeof _goalNotifsFor === 'function') ? _goalNotifsFor : null;
function _v55NormFilial(v){
  const s=String(v||'').toUpperCase();
  const m=s.match(/F(\d{1,2})/); if(m) return 'F'+String(Number(m[1]));
  const m2=s.match(/FILIAL\s*0?(\d{1,2})/); if(m2) return 'F'+String(Number(m2[1]));
  return s;
}
function _v55MetaDiariaMatchesEntity(ent){
  try{
    const pre = Array.isArray(window.META_DIARIA_BATIDA_PRECALC) ? window.META_DIARIA_BATIDA_PRECALC : [];
    if(!ent || !pre.length) return [];
    const entType=String(ent.type||'').toLowerCase();
    const entFil=_v55NormFilial(ent.filial||'');
    const entName=normName(ent.nome||ent.login||'');
    return pre.filter(x=>{
      const kind=String(x.kind||x.tipo||'').toLowerCase();
      const xf=_v55NormFilial(x.filial||'');
      const xn=normName(x.nome||x.label||'');
      if(entType==='filial') return kind==='filial' && xf && xf===entFil;
      if(entType==='vendedor') return kind==='vendedor' && ((xn && xn===entName) || (xn && entName && (xn.includes(entName)||entName.includes(xn))));
      return false;
    });
  }catch(e){return []}
}
function _goalNotifsFor(ent){
  let arr=[];
  try{ arr = _v55_oldGoalNotifsFor ? (_v55_oldGoalNotifsFor(ent)||[]) : []; }catch(e){ arr=[]; }
  try{
    const hits=_v55MetaDiariaMatchesEntity(ent);
    hits.forEach((b,idx)=>{
      const key='meta_diaria_solidus_'+(b.kind||b.tipo||'')+'_'+(b.filial||'')+'_'+normName(b.nome||'')+'_'+idx;
      if(!arr.some(n=>String(n.k||'')===key || String(n.k||'')==='meta_diaria')){
        arr.push({k:key,t:'Meta diária BATIDA',d:`${b.nome||filialLabel(b.filial||'')} bateu a meta diária no Sólidus: ${_v54FormatPct(b.ating)} no período do dia. Realizado: ${b.realizado_periodo||''} / Meta: ${b.meta_periodo||''}.`});
      }
    });
  }catch(e){}
  return arr;
}



// ===== V6.4 HOTFIX: notificação vazia, aniversariantes sem duplicar, filtro filial e hora Brasília =====
(function(){
  try{
    const st=document.createElement('style');
    st.textContent=`
      #laranjitoNotifyPanel:empty,#laranjitoNotifyPanel.show:empty{display:none!important;border:0!important;box-shadow:none!important;padding:0!important;height:0!important;min-height:0!important;max-height:0!important;overflow:hidden!important;opacity:0!important;pointer-events:none!important}
      .ticker-row,.aviso-ticker,.ticker-track{max-width:100%!important;overflow:hidden!important}
      .ticker-item{min-width:210px!important;max-width:340px!important;white-space:normal!important;line-height:1.15!important;align-items:flex-start!important;flex-direction:column!important;gap:3px!important}
      .ticker-item .small,.ticker-item small{display:block!important;white-space:normal!important;overflow:visible!important;text-overflow:clip!important;font-size:11px!important;line-height:1.15!important;opacity:.82}
    `;
    document.head.appendChild(st);
  }catch(e){}
  function _hideEmptyNotifyPanel(){
    try{
      const p=document.getElementById('laranjitoNotifyPanel');
      if(!p) return;
      const hasText=(p.textContent||'').trim().length>0;
      const hasChild=!!p.querySelector('*');
      if(!hasText && !hasChild){
        p.classList.remove('show');
        p.style.display='none';
        p.innerHTML='';
      }
    }catch(e){}
  }
  window.hideEmptyNotifyPanelMDL=_hideEmptyNotifyPanel;
  setInterval(_hideEmptyNotifyPanel,900);
  setTimeout(_hideEmptyNotifyPanel,200);
  try{
    const _oldUpdateGoal=typeof updateGoalNotifications==='function'?updateGoalNotifications:null;
    if(_oldUpdateGoal){
      updateGoalNotifications=function(){
        const r=_oldUpdateGoal.apply(this,arguments);
        setTimeout(_hideEmptyNotifyPanel,60);
        return r;
      }
    }
  }catch(e){}

  // Filial/Vendedor: aniversário e reativação não podem duplicar no painel da filial e no vendedor.
  // A filial/gerente fica somente com o que caiu no bucket GERENTE_Fx; vendedor fica somente com seu key.
  function _ownerKeyForEntityV64(ent){
    try{
      if(!ent) return '';
      const f=String(ent.filial||'').toUpperCase();
      if(!f) return '';
      if(ent.type==='filial' || ent.is_filial || String(ent.nome||'').toUpperCase().startsWith('FILIAL ')) return `GERENTE_${f}`;
      if(ent.is_gerente || /GERENTE/i.test(String(ent.tipo||''))) return `GERENTE_${f}`;
      return reatUserKeyFromNome(ent.nome||ent.login||'', f);
    }catch(e){return ''}
  }
  window.ownerKeyForEntityV64=_ownerKeyForEntityV64;

  if(typeof aniversariantesRowsParaEnt==='function'){
    aniversariantesRowsParaEnt=function(ent){
      const key=_ownerKeyForEntityV64(ent);
      const rows=(ANIVERSARIANTES||[]).map((r,i)=>({...r,_idx:i,_owner:aniversarioOwnerInfo(r)}));
      if(!key) return rows;
      return rows.filter(r=>String(r._owner?.key||'')===key);
    }
  }
  if(typeof reativacaoRowsParaEnt==='function'){
    reativacaoRowsParaEnt=function(ent){
      const key=_ownerKeyForEntityV64(ent);
      const rows=(CLIENTES_SEM_MOVIMENTO||[]).map((r,i)=>({...r,_idx:i,_owner:reativacaoOwnerInfo(r)}));
      if(!key) return rows;
      return rows.filter(r=>String(r._owner?.key||'')===key);
    }
  }

  // Filtro por filial: se selecionar F5, não trazer crediarista de outra filial nem cobrança terceiro.
  if(typeof currentEntities==='function'){
    const _oldCurrentEntitiesV64=currentEntities;
    currentEntities=function(){
      let arr=[];
      try{arr=_oldCurrentEntitiesV64.apply(this,arguments)||[]}catch(e){arr=[]}
      if(String(filtroFilial||'TODAS')==='TODAS') return arr;
      const f=String(filtroFilial||'').toUpperCase();
      return arr.filter(x=>{
        if(!x) return false;
        if(x.is_terceiro || x.type==='terceiro') return false;
        return String(x.filial||'').toUpperCase()===f;
      });
    }
  }

  // Botão voltar só para master/diretor/painel; vendedor/filial/crediarista individual não precisa.
  try{
    const st2=document.createElement('style');
    st2.textContent=`body.individual-view .back-row, body.individual-view .btn[onclick*="goBack"]{display:none!important}`;
    document.head.appendChild(st2);
  }catch(e){}

  // Re-renderiza detalhes antigos após o hotfix, evitando painel duplicado aberto.
  setTimeout(()=>{try{_hideEmptyNotifyPanel(); if(detailScreen && !detailScreen.classList.contains('hidden') && currentDetailRef){ openEntity(currentDetailRef); }}catch(e){}},800);
})();


// ===== V7.4: botão LISTA nos murais rotativos da tela inicial =====
(function(){
  try{
    window.MDL_MURAL_LISTAS = window.MDL_MURAL_LISTAS || {};
    const css=document.createElement('style');
    css.textContent=`
      .mdl-mural-actions{display:flex;gap:8px;align-items:center;justify-content:flex-end;flex-wrap:wrap}
      .mdl-mural-list-btn{min-width:84px!important;justify-content:center!important;background:rgba(255,255,255,.055)!important;border:1px solid rgba(255,255,255,.13)!important;color:#dbeafe!important}
      .mdl-mural-list-btn:hover{background:rgba(59,130,246,.16)!important;border-color:rgba(96,165,250,.28)!important;color:#fff!important}
      .mdl-mural-modal-backdrop{position:fixed;inset:0;background:rgba(2,6,23,.72);backdrop-filter:blur(5px);z-index:99998;display:flex;align-items:center;justify-content:center;padding:18px}
      .mdl-mural-modal{width:min(920px,96vw);max-height:86vh;overflow:hidden;border-radius:24px;background:linear-gradient(135deg,rgba(17,24,39,.98),rgba(10,12,18,.98));border:1px solid rgba(255,255,255,.12);box-shadow:0 30px 90px rgba(0,0,0,.55);display:flex;flex-direction:column}
      .mdl-mural-modal-head{display:flex;align-items:flex-start;justify-content:space-between;gap:14px;padding:18px 20px;border-bottom:1px solid rgba(255,255,255,.08)}
      .mdl-mural-modal-head h2{margin:0;font-size:20px;color:#fff}.mdl-mural-modal-head .hint{font-size:12px;color:#9ca3af;margin-top:4px}
      .mdl-mural-modal-tools{display:flex;gap:8px;align-items:center;padding:12px 18px;border-bottom:1px solid rgba(255,255,255,.07)}
      .mdl-mural-modal-tools input{flex:1;min-width:180px;background:rgba(255,255,255,.045);border:1px solid rgba(255,255,255,.10);color:#fff;border-radius:14px;padding:10px 12px;font-weight:750;outline:none}
      .mdl-mural-modal-list{overflow:auto;padding:12px 18px 18px;display:grid;gap:8px}
      .mdl-mural-modal-row{display:grid;grid-template-columns:40px minmax(180px,1.1fr) minmax(120px,.9fr);gap:12px;align-items:center;padding:11px 12px;border-radius:16px;background:rgba(255,255,255,.045);border:1px solid rgba(255,255,255,.075)}
      .mdl-mural-modal-row:hover{background:rgba(255,255,255,.065)}
      .mdl-mural-modal-row .idx{width:28px;height:28px;border-radius:999px;display:flex;align-items:center;justify-content:center;background:rgba(249,168,50,.14);border:1px solid rgba(249,168,50,.25);color:#fbbf24;font-weight:950;font-size:12px}
      .mdl-mural-modal-row .nome{font-weight:950;color:#f8fafc;line-height:1.25}.mdl-mural-modal-row .info{font-size:12px;color:#aab4cb;line-height:1.25;font-weight:750}
      @media(max-width:720px){.mdl-mural-modal-row{grid-template-columns:34px 1fr}.mdl-mural-modal-row .info{grid-column:2}.mdl-mural-actions{justify-content:flex-start}.mdl-mural-modal{max-height:92vh}}
    `;
    document.head.appendChild(css);
  }catch(e){console.warn('V7.4 css mural lista',e)}
})();

function mdlV74CleanMuralEntries(entries){
  try{
    const seen=new Set();
    return (entries||[]).map(e=>typeof e==='string'?{nome:e,info:''}:e).filter(e=>(e.nome||e.label||'').trim()).map(e=>({
      nome:String(e.nome||e.label||'').trim(),
      info:String(e.info||e.detalhe||e.filial||'').trim()
    })).filter(e=>{
      const k=(e.nome+'|'+e.info).toLowerCase();
      if(seen.has(k)) return false; seen.add(k); return true;
    });
  }catch(e){return []}
}
function mdlV74OpenMuralLista(id){
  try{
    const data=(window.MDL_MURAL_LISTAS||{})[id];
    if(!data){toast('Lista não encontrada neste mural.','warn'); return;}
    const arr=mdlV74CleanMuralEntries(data.entries||[]);
    const old=document.getElementById('mdlMuralListaModal'); if(old) old.remove();
    const html=`<div id="mdlMuralListaModal" class="mdl-mural-modal-backdrop" onclick="if(event.target.id==='mdlMuralListaModal')mdlV74CloseMuralLista()">
      <div class="mdl-mural-modal">
        <div class="mdl-mural-modal-head">
          <div><h2>${esc(data.icon||'📋')} ${esc(data.title||'Lista do mural')}</h2><div class="hint">${arr.length} item(ns) encontrados. Use a busca para localizar rápido.</div></div>
          <button class="btn soft" onclick="mdlV74CloseMuralLista()">Fechar</button>
        </div>
        <div class="mdl-mural-modal-tools"><input id="mdlMuralListaBusca" placeholder="Buscar nome, filial ou detalhe..." oninput="mdlV74FiltrarMuralLista(this.value)"><button class="btn soft" onclick="mdlV74CopiarMuralLista('${esc(id)}')">Copiar lista</button></div>
        <div id="mdlMuralListaRows" class="mdl-mural-modal-list">${arr.map((e,i)=>`<div class="mdl-mural-modal-row" data-search="${esc((e.nome+' '+e.info).toLowerCase())}"><div class="idx">${i+1}</div><div class="nome">${esc(e.nome)}</div><div class="info">${esc(e.info||'')}</div></div>`).join('') || '<div class="empty">Nenhum item para listar.</div>'}</div>
      </div>
    </div>`;
    document.body.insertAdjacentHTML('beforeend',html);
    setTimeout(()=>document.getElementById('mdlMuralListaBusca')?.focus(),80);
  }catch(e){console.warn('mdlV74OpenMuralLista',e); toast('Erro ao abrir lista do mural.','warn')}
}
function mdlV74CloseMuralLista(){try{document.getElementById('mdlMuralListaModal')?.remove()}catch(e){}}
function mdlV74FiltrarMuralLista(q){
  try{q=String(q||'').trim().toLowerCase(); document.querySelectorAll('#mdlMuralListaRows .mdl-mural-modal-row').forEach(r=>{r.style.display=(!q || String(r.dataset.search||'').includes(q))?'grid':'none'});}catch(e){}
}
async function mdlV74CopiarMuralLista(id){
  try{
    const data=(window.MDL_MURAL_LISTAS||{})[id]||{};
    const arr=mdlV74CleanMuralEntries(data.entries||[]);
    const txt=[`${data.title||'Lista do mural'} (${arr.length})`,...arr.map((e,i)=>`${i+1}. ${e.nome}${e.info?' - '+e.info:''}`)].join('\n');
    await navigator.clipboard.writeText(txt);
    toast('Lista copiada.','success');
  }catch(e){toast('Não consegui copiar a lista.','warn')}
}


// ===== V8.4: remove contadores numéricos dos murais/cards rotativos =====
(function(){
  try{
    const st=document.createElement('style');
    st.textContent=`.mdl-hero-count,.mdl-mural-actions .badge,.aviso-rotativo .section-head .badge{display:none!important}`;
    document.head.appendChild(st);
  }catch(e){}
})();

// Override final do ticker para incluir botão Lista ao lado do Acelerar.
function renderAvisoTicker(title,hint,entries,opts={}){
  const icon=opts.icon||'•';
  const color=opts.color||'rgba(249,168,50,.25)';
  const arr=mdlV74CleanMuralEntries(entries||[]);
  if(!arr.length) return `<div class="empty">Nenhuma informação no momento.</div>`;
  const doubled=arr.concat(arr).concat(arr);
  const safeId='ticker_'+Math.random().toString(36).slice(2);
  window.MDL_MURAL_LISTAS=window.MDL_MURAL_LISTAS||{};
  window.MDL_MURAL_LISTAS[safeId]={title:title||'Mural',hint:hint||'',icon,entries:arr};
  return `<div class="glass panel full aviso-rotativo" style="border-color:${color}"><div class="section-head" style="margin-bottom:6px"><div><h2 style="font-size:18px">${esc(icon)} ${esc(title||'Mural')}</h2><div class="hint">${esc(hint||'')}</div></div><div class="mdl-mural-actions"><button class="btn soft mdl-mural-list-btn" onclick="mdlV74OpenMuralLista('${safeId}')">📋 Lista</button><button class="btn soft ticker-speed-btn" onclick="toggleTickerSpeed('${safeId}',this)">⚡ Acelerar</button></div></div><div id="${safeId}" class="aviso-ticker"><div class="aviso-ticker-track" style="animation-duration:900s">${doubled.map(e=>`<span class="aviso-pill"><i class="red-dot"></i><span class="ticker-main">${esc(e.nome||'')}</span><small>${esc(e.info||'')}</small></span>`).join('')}</div></div></div>`;
}


// ===== V9.4 HOTFIX: login robusto após atualização do pacote de vendas =====
(function(){
  function _loginMsg(txt){try{const m=document.getElementById('loginMsg'); if(m) m.textContent=txt||'';}catch(e){}}
  function _val(id){return String((document.getElementById(id)||{}).value||'').trim();}
  async function _loginStable(ev){
    try{ if(ev && ev.preventDefault) ev.preventDefault(); }catch(_e){}
    const u=_val('loginUser').toLowerCase();
    const s=_val('loginPass');
    _loginMsg('');
    if(!u || !s){_loginMsg('Informe usuário e senha.'); return false;}
    try{
      if(typeof LOGIN_MASTER!=='undefined' && typeof SENHA_MASTER!=='undefined' && u===String(LOGIN_MASTER).toLowerCase() && s===String(SENHA_MASTER)){
        usuarioAtual={tipo:'master',nome:'Master',roleLabel:'Master'};
        try{saveSession();}catch(_e){}
        try{await abrirApp();}catch(e){console.error('Erro abrirApp master',e); _loginMsg('Erro ao abrir dashboard. Atualize com Ctrl+F5 ou chame suporte.');}
        return false;
      }
      try{ if(typeof carregarCredenciaisOnline==='function') await Promise.race([carregarCredenciaisOnline(), new Promise(resolve=>setTimeout(resolve,1800))]); }catch(e){console.warn('Credenciais online ignoradas no login estável',e);}
      if(typeof acessoGeralBloqueado==='function' && acessoGeralBloqueado() && u!==String(LOGIN_DIRETOR||'').toLowerCase()){
        _loginMsg('Acesso temporariamente bloqueado pelo Master.'); return false;
      }
      if(typeof LOGIN_DIRETOR!=='undefined' && u===String(LOGIN_DIRETOR).toLowerCase()){
        const authDir=(typeof getAuthUser==='function')?getAuthUser(u):null;
        const senhaDir=(authDir&&authDir.password) || (typeof SENHA_DIRETOR!=='undefined'?SENHA_DIRETOR:'');
        if(String(senhaDir)===s){ usuarioAtual={tipo:'master',nome:'Diretor Comercial',roleLabel:'Diretor Comercial'}; try{saveSession();}catch(_e){}; try{await abrirApp();}catch(e){console.error(e); _loginMsg('Erro ao abrir dashboard.');} return false; }
      }
      const auth=(typeof getAuthUser==='function')?getAuthUser(u):null;
      if(typeof CREDS!=='undefined' && CREDS[u] && auth && String(auth.password)===s){
        if(auth.access_disabled || auth.status_operacional==='inativo' || CREDS[u]?.access_disabled || CREDS[u]?.status_operacional==='inativo'){
          _loginMsg('Usuário inativo/bloqueado pelo Master.'); return false;
        }
        if(auth.must_change_password && typeof openPrimeiroAcesso==='function'){
          _loginMsg('Primeiro acesso: defina sua nova senha.'); openPrimeiroAcesso(u); return false;
        }
        usuarioAtual={tipo:'user',login:u,...CREDS[u]}; try{saveSession();}catch(_e){}; try{await abrirApp();}catch(e){console.error(e); _loginMsg('Erro ao abrir dashboard.');} return false;
      }
      _loginMsg('Login ou senha inválidos.');
      return false;
    }catch(e){ console.error('Erro geral login V9.4',e); _loginMsg('Erro no login. Atualize com Ctrl+F5 e tente novamente.'); return false; }
  }
  window.fazerLogin=_loginStable;
  window.mdlLoginV94=_loginStable;
  document.addEventListener('DOMContentLoaded',()=>{
    const b=document.getElementById('loginBtn') || document.querySelector('[data-login-btn]') || document.querySelector('#loginScreen .btn.primary');
    if(b){ b.onclick=_loginStable; b.addEventListener('click',_loginStable,true); }
    const p=document.getElementById('loginPass'); if(p){ p.addEventListener('keydown',e=>{if(e.key==='Enter'){_loginStable(e);}},true); }
  });
  window.addEventListener('error',e=>{try{ if(!document.getElementById('loginScreen')?.classList.contains('hidden')){console.error('Erro JS capturado V9.4',e.error||e.message);}}catch(_e){} });
})();



// ===== V9.6 HOTFIX: venda diária live + murais consistentes + clientes sem movimento lazy-load =====
(function(){
  try{
    window.MDL_V96_HOTFIX = true;

    window._mdlCsmLoadedV96 = Array.isArray(CLIENTES_SEM_MOVIMENTO) && CLIENTES_SEM_MOVIMENTO.length>0;
    window._mdlCsmLoadingV96 = false;
    async function mdlV96LoadClientesSemMovimento(force=false){
      if(window._mdlCsmLoadedV96 && !force) return true;
      if(window._mdlCsmLoadingV96) return false;
      window._mdlCsmLoadingV96 = true;
      try{
        const payload = await fetchJsonNoCache('clientes_sem_movimento.json');
        const rows = Array.isArray(payload) ? payload : (Array.isArray(payload?.clientes) ? payload.clientes : []);
        if(Array.isArray(CLIENTES_SEM_MOVIMENTO)){
          CLIENTES_SEM_MOVIMENTO.splice(0, CLIENTES_SEM_MOVIMENTO.length, ...rows);
        }
        if(payload && typeof payload === 'object' && !Array.isArray(payload) && typeof CLIENTES_SEM_MOVIMENTO_META === 'object'){
          Object.assign(CLIENTES_SEM_MOVIMENTO_META, payload);
          delete CLIENTES_SEM_MOVIMENTO_META.clientes;
        }
        window._mdlCsmLoadedV96 = true;
        window._mdlCsmLoadingV96 = false;
        console.log('[MDL V9.6] clientes sem movimento carregados sob demanda:', rows.length);
        return true;
      }catch(e){
        console.warn('[MDL V9.6] falha ao carregar clientes_sem_movimento.json', e);
        window._mdlCsmLoadingV96 = false;
        return false;
      }
    }
    window.mdlV96LoadClientesSemMovimento = mdlV96LoadClientesSemMovimento;

    if(typeof renderReativacaoTab === 'function' && !window._renderReativacaoV96Wrapped){
      window._renderReativacaoV96Wrapped = true;
      const _oldRenderReat = renderReativacaoTab;
      renderReativacaoTab = function(){
        try{
          const semDados = !(Array.isArray(CLIENTES_SEM_MOVIMENTO) && CLIENTES_SEM_MOVIMENTO.length);
          if(semDados && !window._mdlCsmLoadedV96){
            if(reativacaoSection){
              reativacaoSection.innerHTML = `<div class="section-head"><div><h2>🧡 Clientes sem movimento +45 dias <span class="note" style="color:#f59e0b">${esc(DASHBOARD_BUILD_VERSION)}</span></h2><div class="hint">Carregando lista do FTP somente agora para deixar o dashboard mais leve.</div></div></div><div class="glass panel" style="padding:18px"><strong>⏳ Carregando clientes sem movimento...</strong><div class="hint">Aguarde alguns segundos. A lista grande não fica mais embutida na tela inicial.</div></div>`;
            }
            mdlV96LoadClientesSemMovimento().then(()=>{try{_oldRenderReat.apply(this,arguments)}catch(e){console.warn(e)}});
            return;
          }
        }catch(e){console.warn('[MDL V9.6] wrapper reativação',e)}
        return _oldRenderReat.apply(this,arguments);
      };
    }

    if(typeof renderNoReactivationAlerts === 'function' && !window._renderNoReactivationV96Wrapped){
      const _oldNoReat = renderNoReactivationAlerts;
      renderNoReactivationAlerts = function(){
        try{
          if(!(Array.isArray(CLIENTES_SEM_MOVIMENTO) && CLIENTES_SEM_MOVIMENTO.length) && !window._mdlCsmLoadedV96){
            const total = Number(CLIENTES_SEM_MOVIMENTO_META?.acionaveis_total || CLIENTES_SEM_MOVIMENTO_META?.base_total || 0);
            if(total>0){
              return `<div class="glass panel aviso-rotativo" style="margin-bottom:12px;padding:13px 15px;border-color:rgba(245,158,11,.30)"><div class="section-head" style="margin:0"><div><h2 style="margin:0;font-size:17px">🧡 Clientes sem movimento</h2><div class="hint">${total.toLocaleString('pt-BR')} cliente(s) monitorado(s). Abra a aba Clientes sem movimento para carregar a lista completa.</div></div><button class="btn soft btn-xs" onclick="setMainTab('reativacao')">📋 Abrir lista</button></div></div>`;
            }
          }
        }catch(e){}
        return _oldNoReat.apply(this,arguments);
      };
    }

    function mdlV96EntriesFromTickerHtml(html){
      try{
        const tmp=document.createElement('div'); tmp.innerHTML=String(html||'');
        return Array.from(tmp.querySelectorAll('.aviso-pill')).map(p=>({
          nome:(p.querySelector('.ticker-main')?.textContent||p.childNodes?.[1]?.textContent||p.textContent||'').trim(),
          info:(p.querySelector('small')?.textContent||'').trim()
        })).filter(x=>x.nome);
      }catch(e){return []}
    }
    window.mdlV96EntriesFromTickerHtml = mdlV96EntriesFromTickerHtml;

    if(!window._mdlHeroV96Wrapped){
      window._mdlHeroV96Wrapped = true;
      window.mdlV51HeroSlides = function(){
        const groups = [
          {title:'Sem cobranças hoje', icon:'⏰', kind:'cobranca', dot:'#ef4444', html:(()=>{try{return renderNoChargeAlerts()||''}catch(e){return ''}})(), emptyMain:'Cobranças em dia', emptyDetail:'Nenhum alerta crítico de cobrança agora.'},
          {title:'Clientes sem movimento', icon:'🧡', kind:'reativacao', dot:'#f59e0b', html:(()=>{try{return renderNoReactivationAlerts()||''}catch(e){return ''}})(), emptyMain:'Reativação', emptyDetail:'Abra a aba Clientes sem movimento para ver a lista.'},
          {title:'Aniversariantes do dia', icon:'🎂', kind:'aniversario', dot:'#ec4899', html:(()=>{try{return renderMuralAniversariantesDia()||''}catch(e){return ''}})(), emptyMain:'Aniversariantes', emptyDetail:'Nenhum aniversariante pendente no momento.'},
          {title:'Meta diária BATIDA', icon:'🎯', kind:'meta', dot:'#22c55e', html:(()=>{try{return renderMetaDiariaBatidaAlerts()||''}catch(e){return ''}})(), emptyMain:'Aguardando metas diárias', emptyDetail:'Quem passar de 100% no período aparece aqui.'}
        ];
        let slides=[];
        groups.forEach(g=>{
          const items=mdlV96EntriesFromTickerHtml(g.html);
          if(items.length){
            slides = slides.concat(items.slice(0,12).map((it,idx)=>({icon:g.icon,title:g.title,count:items.length,main:it.nome,detail:it.info,dot:g.dot,kind:g.kind,mini:items.slice(idx+1,idx+4).map(x=>x.nome)})));
          }else{
            slides.push({icon:g.icon,title:g.title,count:0,main:g.emptyMain,detail:g.emptyDetail,dot:g.dot,kind:g.kind,mini:[]});
          }
        });
        return slides;
      };
      window.mdlV51HeroSlideHtml = function(s){
        const mini=(s.mini||[]).slice(0,3).map(x=>`<span class="mdl-hero-mini">${esc(x)}</span>`).join('');
        return `<div class="mdl-hero-title"><span class="mdl-hero-dot" style="background:${esc(s.dot||'#f59e0b')}"></span>${esc(s.icon||'🔔')} ${esc(s.title||'Mural operacional')}</div><div class="mdl-hero-main">${esc(s.main||'Operação em andamento')}</div><div class="mdl-hero-detail">${esc(s.detail||'Resumo automático dos murais do dia.')}</div>${mini?`<div class="mdl-hero-mini-list">${mini}</div>`:''}`;
      };
    }

    if(typeof pollSalesLive === 'function' && !window._pollSalesLiveV96Wrapped){
      const _oldPollSalesLive = pollSalesLive;
      pollSalesLive = async function(){
        const r = await _oldPollSalesLive.apply(this,arguments);
        try{
          if(mainTab==='inicio'){
            renderInicioTab();
            const hero=document.getElementById('mdlHeroMural');
            if(hero){ hero.dataset.idx='0'; const s=mdlV51HeroSlides()[0]||{}; hero.className='glass kpi mdl-hero-mural-card v52-'+String(s.kind||'cobranca'); hero.innerHTML=mdlV51HeroSlideHtml(s); }
          }
        }catch(e){console.warn('[MDL V9.6] refresh mural pós vendas',e)}
        return r;
      };
    }

    try{
      const st=document.createElement('style');
      st.id='mdl-v96-lazy-light-css';
      st.textContent=`
        .mdl-hero-mini-list{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px;position:relative;z-index:1}
        .mdl-hero-mini{border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.055);border-radius:999px;padding:5px 8px;font-size:11px;color:#dbe4ff;font-weight:800;max-width:170px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
        @media(max-width:760px){.mdl-hero-mini-list{display:none!important}.mdl-hero-detail{max-width:100%!important}.mdl-hero-main{font-size:17px!important}}
      `;
      document.head.appendChild(st);
    }catch(e){}

    console.log('[MDL V9.6] hotfix ativo: venda diária oficial + CSM lazy-load + murais sincronizados');
  }catch(e){console.warn('MDL V9.6 hotfix falhou',e)}
})();

// ===== V9.7: config preservada, diretor com abas configuráveis, emojis WhatsApp e mobile reforçado =====
(function(){
  try{
    window.MDL_V97_HOTFIX = true;
    const DEFAULT_REATIVACAO_MSG_V97 = `Olá, {primeiro_nome}! Tudo bem? 😊

Aqui é da Lojas MDL - Móveis do Lar. Estamos com saudades de você! Faz um tempinho que você não aparece na loja. 🥹

Venha conhecer nossas novidades e aproveitar condições especiais que preparamos para nossos clientes. 👉👉😍😍`;
    const DEFAULT_ANIVERSARIO_MSG_V97 = `Olá, {primeiro_nome}! Feliz aniversário! 🎂🎉

Aqui é da Lojas MDL - Móveis do Lar. Desejamos muita saúde, paz e felicidades neste dia especial. 😍😍

Preparamos condições especiais para você comemorar com a gente.
🕺🎉🤩`;
    function mdlV97HasEmoji(s){ return /[\u{1F300}-\u{1FAFF}\u{2600}-\u{27BF}]/u.test(String(s||'')); }
    function mdlV97EnsureMsgDefaults(){
      try{
        CONFIG_META = CONFIG_META || {};
        if(!String(CONFIG_META.reativacao_msg_template||'').trim() || !mdlV97HasEmoji(CONFIG_META.reativacao_msg_template)){ CONFIG_META.reativacao_msg_template = DEFAULT_REATIVACAO_MSG_V97; }
        if(!String(CONFIG_META.aniversario_msg_template||'').trim() || !mdlV97HasEmoji(CONFIG_META.aniversario_msg_template)){ CONFIG_META.aniversario_msg_template = DEFAULT_ANIVERSARIO_MSG_V97; }
        const r=document.getElementById('reatMsgTemplate'); if(r && (!String(r.value||'').trim() || !mdlV97HasEmoji(r.value))) r.value=CONFIG_META.reativacao_msg_template;
        const a=document.getElementById('anivMsgTemplate'); if(a && (!String(a.value||'').trim() || !mdlV97HasEmoji(a.value))) a.value=CONFIG_META.aniversario_msg_template;
      }catch(e){console.warn('[MDL V9.7] ensure msg defaults', e)}
    }
    window.mdlV97EnsureMsgDefaults = mdlV97EnsureMsgDefaults;
    function mdlV97FilialKey(v){ return String(v||'').toUpperCase().replace(/[^A-Z0-9]/g,'').replace(/^FILIAL/,'F'); }
    const _oldReatTplV97 = typeof reativacaoTemplateAtual==='function' ? reativacaoTemplateAtual : null;
    reativacaoTemplateAtual = function(filial=''){
      try{ const f=mdlV97FilialKey(filial), map=CONFIG_META?.reativacao_msg_template_filiais||{}; if(f && String(map[f]||'').trim()) return String(map[f]); const base=_oldReatTplV97?_oldReatTplV97(filial):(CONFIG_META?.reativacao_msg_template||''); return (String(base||'').trim() && mdlV97HasEmoji(base)) ? base : DEFAULT_REATIVACAO_MSG_V97; }catch(e){return DEFAULT_REATIVACAO_MSG_V97}
    };
    const _oldAnivTplV97 = typeof aniversarioTemplateAtual==='function' ? aniversarioTemplateAtual : null;
    aniversarioTemplateAtual = function(filial=''){
      try{ const f=mdlV97FilialKey(filial), map=CONFIG_META?.aniversario_msg_template_filiais||{}; if(f && String(map[f]||'').trim()) return String(map[f]); const base=_oldAnivTplV97?_oldAnivTplV97(filial):(CONFIG_META?.aniversario_msg_template||''); return (String(base||'').trim() && mdlV97HasEmoji(base)) ? base : DEFAULT_ANIVERSARIO_MSG_V97; }catch(e){return DEFAULT_ANIVERSARIO_MSG_V97}
    };
    async function mdlV97SaveConfig(){
      if(typeof mdlV93SaveConfigB64 === 'function') return await mdlV93SaveConfigB64();
      const fd=new FormData(); fd.append('payload_json', JSON.stringify({global:CONFIG_META,individual:CONFIG_META_IND||{}}));
      const r=await fetch(API_CFG+'?_='+Date.now(),{method:'POST',body:fd,cache:'no-store'}); return await r.json();
    }
    window.salvarMensagemReativacaoGlobal = async function(){ const el=document.getElementById('reatMsgTemplate'); const val=String(el?.value||DEFAULT_REATIVACAO_MSG_V97); CONFIG_META.reativacao_msg_template=val; try{ const j=await mdlV97SaveConfig(); if(j.ok){ if(el) el.value=val; toast('Mensagem global de reativação salva com emojis.','success'); } else toast('Não consegui salvar mensagem global.','warn'); }catch(e){console.warn(e); toast('Falha ao salvar mensagem global.','warn')} };
    window.salvarMensagemReativacaoFilial = async function(){ const f=mdlV97FilialKey(document.getElementById('reatMsgFilial')?.value||''), el=document.getElementById('reatMsgTemplateFilial'); if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return} CONFIG_META.reativacao_msg_template_filiais=CONFIG_META.reativacao_msg_template_filiais||{}; CONFIG_META.reativacao_msg_template_filiais[f]=String(el?.value||''); try{ const j=await mdlV97SaveConfig(); toast(j.ok?`Mensagem da ${f} salva com emojis.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn'); }catch(e){console.warn(e); toast('Falha ao salvar mensagem por filial.','warn')} };
    window.salvarMensagemAniversarioGlobal = async function(){ const el=document.getElementById('anivMsgTemplate'); const val=String(el?.value||DEFAULT_ANIVERSARIO_MSG_V97); CONFIG_META.aniversario_msg_template=val; try{ const j=await mdlV97SaveConfig(); if(j.ok){ if(el) el.value=val; toast('Mensagem global de aniversário salva com emojis.','success'); } else toast('Não consegui salvar mensagem global.','warn'); }catch(e){console.warn(e); toast('Falha ao salvar mensagem de aniversário.','warn')} };
    window.salvarMensagemAniversarioFilial = async function(){ const f=mdlV97FilialKey(document.getElementById('anivMsgFilial')?.value||''), el=document.getElementById('anivMsgTemplateFilial'); if(!f){toast('Selecione uma filial para salvar mensagem individual.','warn');return} CONFIG_META.aniversario_msg_template_filiais=CONFIG_META.aniversario_msg_template_filiais||{}; CONFIG_META.aniversario_msg_template_filiais[f]=String(el?.value||''); try{ const j=await mdlV97SaveConfig(); toast(j.ok?`Mensagem de aniversário da ${f} salva com emojis.`:'Não consegui salvar mensagem por filial.',j.ok?'success':'warn'); }catch(e){console.warn(e); toast('Falha ao salvar mensagem por filial.','warn')} };
    window.saveSession = function(){ try{ const data={usuarioAtual,exp:Date.now()+45*24*60*60*1000,version:DASHBOARD_BUILD_VERSION}; localStorage.setItem(SESSION_KEY, JSON.stringify(data)); sessionStorage.setItem(SESSION_KEY, JSON.stringify(data)); }catch(e){} };
    window.restoreSession = function(){ try{ const raw=localStorage.getItem(SESSION_KEY)||sessionStorage.getItem(SESSION_KEY); if(!raw) return false; const data=JSON.parse(raw); if(!data||!data.usuarioAtual||!data.exp||Date.now()>Number(data.exp)){localStorage.removeItem(SESSION_KEY);sessionStorage.removeItem(SESSION_KEY);return false;} usuarioAtual=data.usuarioAtual; return true; }catch(e){return false} };
    const DEFAULT_DIRECTOR_TABS_V97=['inicio','vendedores','filiais','servicos','cobrancas','avisos','historico'];
    const ALL_TABS_V97=[['inicio','Início'],['vendedores','Por Colaborador'],['filiais','Por Filial'],['metas','Metas'],['servicos','Serviços'],['cobrancas','Cobranças'],['reativacao','Clientes sem movimento'],['aniversariantes','Aniversariantes'],['avisos','Avisos'],['telegram','Telegram'],['senhas','Senhas'],['historico','Histórico']];
    function isDiretorV97(){ return String(usuarioAtual?.roleLabel||'').toLowerCase().includes('diretor') || String(usuarioAtual?.login||'').toLowerCase()==='diretorcomercial' || String(usuarioAtual?.tipo||'').toLowerCase()==='diretor'; }
    function diretorTabsV97(){ const arr=Array.isArray(CONFIG_META?.director_visible_tabs)?CONFIG_META.director_visible_tabs:[]; return (arr.length?arr:DEFAULT_DIRECTOR_TABS_V97).map(String); }
    function applyDirectorTabsV97(){ try{ if(!isDiretorV97()) return; const allowed=new Set(diretorTabsV97()); document.querySelectorAll('#masterTabs .tab').forEach(btn=>{ const t=String(btn.dataset.tab||''); btn.classList.toggle('hidden', !allowed.has(t)); }); if(!allowed.has(String(window.mainTab||'inicio')) && typeof setMainTab==='function') setMainTab('inicio'); }catch(e){console.warn('[MDL V9.7] apply tabs',e)} }
    window.applyDirectorTabsV97=applyDirectorTabsV97;
    const _oldAbrirV97=typeof abrirApp==='function'?abrirApp:null; if(_oldAbrirV97){ abrirApp=async function(){ const r=await _oldAbrirV97.apply(this,arguments); setTimeout(()=>{applyDirectorTabsV97();mdlV97EnsureMsgDefaults();},80); return r; }; }
    const _oldSetMainTabV97=typeof setMainTab==='function'?setMainTab:null; if(_oldSetMainTabV97){ setMainTab=function(tab){ if(isDiretorV97()&&!diretorTabsV97().includes(String(tab||''))) tab='inicio'; const r=_oldSetMainTabV97.apply(this,[tab]); setTimeout(()=>{applyDirectorTabsV97();mdlV97EnsureMsgDefaults();},30); return r; }; }
    function renderDirectorTabsConfigV97(){ try{ if(String(usuarioAtual?.tipo||'').toLowerCase()!=='master') return; const host=document.getElementById('senhasSection'); if(!host||document.getElementById('directorTabsConfigV97')) return; const allowed=new Set(diretorTabsV97()); const html=`<div id="directorTabsConfigV97" class="glass panel" style="margin:14px 0;border-color:rgba(96,165,250,.35)"><div class="section-head" style="margin-bottom:8px"><div><h2 style="font-size:18px;margin:0">👑 Visualização do Diretor Comercial</h2><div class="hint">Escolha quais abas o Diretor Comercial pode visualizar. Padrão leve: sem Metas, Clientes sem movimento e Aniversariantes.</div></div></div><div style="display:flex;gap:8px;flex-wrap:wrap">${ALL_TABS_V97.map(([k,n])=>`<label class="pill" style="cursor:pointer"><input type="checkbox" class="dir-tab-v97" value="${esc(k)}" ${allowed.has(k)?'checked':''}> ${esc(n)}</label>`).join('')}</div><button class="btn primary" style="margin-top:12px" onclick="salvarDirectorTabsV97()">💾 Salvar visualização do Diretor</button></div>`; const firstPanel=host.querySelector('.glass.panel'); if(firstPanel) firstPanel.insertAdjacentHTML('beforebegin',html); else host.insertAdjacentHTML('afterbegin',html); }catch(e){console.warn('[MDL V9.7] render director config',e)} }
    window.salvarDirectorTabsV97=async function(){ try{ const vals=[...document.querySelectorAll('.dir-tab-v97:checked')].map(x=>x.value); CONFIG_META.director_visible_tabs=vals.length?vals:DEFAULT_DIRECTOR_TABS_V97; const j=await mdlV97SaveConfig(); toast(j.ok?'Visualização do Diretor salva.':'Não consegui salvar visualização do Diretor.',j.ok?'success':'warn'); }catch(e){console.warn(e); toast('Falha ao salvar visualização do Diretor.','warn')} };
    const _oldRenderSenhasV97=typeof renderSenhasTab==='function'?renderSenhasTab:null; if(_oldRenderSenhasV97){ renderSenhasTab=function(){ const r=_oldRenderSenhasV97.apply(this,arguments); setTimeout(renderDirectorTabsConfigV97,30); return r; }; }
    try{ const st=document.createElement('style'); st.id='mdl-v97-mobile-home-css'; st.textContent=`@media(max-width:760px){html,body{max-width:100vw!important;overflow-x:hidden!important}.app-shell,.container{width:100%!important;max-width:100vw!important;padding:10px!important;box-sizing:border-box!important}.hero,.header-card{display:block!important;padding:16px!important;border-radius:22px!important}.brand{display:flex!important;align-items:center!important;gap:12px!important;flex-wrap:wrap!important}.brand h1{font-size:22px!important;line-height:1.05!important}.brand p,.hero .sub,.hint{font-size:12px!important;line-height:1.35!important}.header-actions{display:flex!important;flex-wrap:wrap!important;gap:8px!important;margin-top:12px!important;width:100%!important}.header-actions .btn,.header-actions .badge{font-size:13px!important;padding:10px 12px!important;min-height:42px!important}.kpis{display:grid!important;grid-template-columns:1fr!important;gap:10px!important;width:100%!important;overflow:visible!important}.kpi{min-width:0!important;width:100%!important;max-width:100%!important;min-height:0!important;padding:14px 16px!important;border-radius:18px!important;box-sizing:border-box!important}.kpi .label{font-size:10px!important;white-space:normal!important;letter-spacing:.08em!important}.kpi .value{font-size:22px!important;line-height:1.12!important;white-space:normal!important;word-break:break-word!important}.kpi .subline{font-size:12px!important;line-height:1.3!important;white-space:normal!important}.kpi img,.kpi .mascote{max-width:56px!important;max-height:56px!important;right:10px!important;bottom:10px!important}.tabs{display:flex!important;overflow-x:auto!important;flex-wrap:nowrap!important;gap:8px!important;padding:8px!important;margin:10px 0!important}.tabs .tab{flex:0 0 auto!important;white-space:nowrap!important;padding:10px 12px!important;font-size:12px!important}.inicio-murais-grid,.inicio-operacional-compact,.grid-cards,.form-grid,.meta-grid{grid-template-columns:1fr!important}.glass.panel,.campaign-banner{padding:14px!important;border-radius:18px!important;max-width:100%!important;overflow:hidden!important}.section-head{display:flex!important;flex-direction:column!important;align-items:flex-start!important;gap:8px!important}.mdl-hero-mural-card{grid-column:1/-1!important;min-height:0!important}}`; document.head.appendChild(st); }catch(e){}
    setTimeout(()=>{mdlV97EnsureMsgDefaults();applyDirectorTabsV97();},300);
    console.log('[MDL V9.7] hotfix ativo: config preservada, diretor abas configuráveis, emojis WhatsApp e mobile reforçado');
  }catch(e){console.warn('[MDL V9.7] hotfix falhou', e)}
})();



// ===== V9.8 HOTFIX: clientes sem movimento visíveis no acesso individual =====
(function(){
  try{
    window.MDL_V99_RESUMO_TELEGRAM_LOGS_FIX = true;

    function mdlV98FilialNorm(v){
      v=String(v||'').trim().toUpperCase();
      if(!v) return '';
      if(/^\d+$/.test(v)) return 'F'+Number(v);
      if(/^F0\d+$/.test(v)) return 'F'+Number(v.replace(/^F0*/,''));
      return v;
    }
    function mdlV98Norm(v){
      try{return normName(v||'')}catch(e){return String(v||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/[^a-z0-9]+/gi,' ').trim().toLowerCase()}
    }
    function mdlV98Key(nome,filial){return mdlV98Norm(nome)+'_'+mdlV98FilialNorm(filial)}
    function mdlV98FindVendByLoginNome(login,nome,filial){
      const fil=mdlV98FilialNorm(filial);
      const log=String(login||'').toLowerCase().trim();
      const nn=mdlV98Norm(nome);
      let arr=[];
      try{arr=flattenVendedores()||[]}catch(e){arr=[]}
      return arr.find(v=>{
        const vf=mdlV98FilialNorm(v.filial);
        if(fil && vf!==fil) return false;
        const vlogin=String(v.login||'').toLowerCase().trim();
        const vn=mdlV98Norm(v.nome);
        return (log && vlogin && vlogin===log) || (nn && vn===nn) || (nn && (vn.includes(nn)||nn.includes(vn)));
      }) || null;
    }
    function mdlV98CandidateKeysFrom(ent){
      const filial=mdlV98FilialNorm(ent?.filial || usuarioAtual?.filial || '');
      const names=[];
      function add(v){v=String(v||'').trim(); if(v && !names.includes(v)) names.push(v)}
      add(ent?.nome); add(ent?.login); add(usuarioAtual?.nome); add(usuarioAtual?.login);
      const found=mdlV98FindVendByLoginNome(ent?.login||usuarioAtual?.login, ent?.nome||usuarioAtual?.nome, filial);
      add(found?.nome); add(found?.login);
      const keys=[];
      names.forEach(n=>{const k=mdlV98Key(n,filial); if(k!=='_'+filial && !keys.includes(k)) keys.push(k)});
      return keys;
    }
    function mdlV98RowsWithOwners(){
      let base=[];
      try{base=(CLIENTES_SEM_MOVIMENTO||[])}catch(e){base=[]}
      return base.map((r,i)=>({...r,_idx:(r._idx!=null?r._idx:i),_owner:reativacaoOwnerInfo(r)}));
    }

    // Corrige a chave do usuário logado: usa o vendedor real da carteira quando existir.
    window.reativacaoCurrentKey = function(){
      if(!usuarioAtual || usuarioAtual.tipo==='master' || usuarioAtual.is_viewer) return '';
      const filial=mdlV98FilialNorm(usuarioAtual.filial||'');
      if(usuarioAtual.is_gerente) return `GERENTE_${filial}`;
      const found=mdlV98FindVendByLoginNome(usuarioAtual.login, usuarioAtual.nome, filial);
      return mdlV98Key(found?.nome || usuarioAtual.nome || usuarioAtual.login || '', filial);
    };

    // Corrige a aba individual do usuário, aceitando nome/login/entidade encontrada.
    window.reativacaoRowsPermitidas = function(){
      let rows=mdlV98RowsWithOwners();
      if(!usuarioAtual || usuarioAtual.tipo==='master' || usuarioAtual.is_viewer) return rows;
      const filial=mdlV98FilialNorm(usuarioAtual.filial||'');
      if(usuarioAtual.is_gerente) return rows.filter(r=>mdlV98FilialNorm(r.filial)===filial && String(r._owner?.key||'')===`GERENTE_${filial}`);
      const keys=mdlV98CandidateKeysFrom(usuarioAtual);
      let filtered=rows.filter(r=>keys.includes(String(r._owner?.key||'')));
      // Fallback seguro: se o login/nome salvo vier abreviado, compara label/nome do responsável.
      if(!filtered.length){
        const nomes=keys.map(k=>mdlV98Norm(k.split('_')[0])).filter(Boolean);
        filtered=rows.filter(r=>mdlV98FilialNorm(r.filial)===filial && nomes.some(n=>{
          const o=mdlV98Norm(r._owner?.nome||r._owner?.label||'');
          return o && n && (o===n || o.includes(n) || n.includes(o));
        }));
      }
      return filtered;
    };

    // Corrige o painel individual aberto pelo Master e também pelo próprio usuário.
    window.reativacaoRowsParaEnt = function(ent){
      if(!ent) return [];
      const filial=mdlV98FilialNorm(ent.filial||'');
      let rows=mdlV98RowsWithOwners();
      if(ent.type==='filial') return rows.filter(r=>mdlV98FilialNorm(r.filial)===filial);
      const keys=mdlV98CandidateKeysFrom(ent);
      let filtered=rows.filter(r=>keys.includes(String(r._owner?.key||'')));
      if(!filtered.length){
        const nomes=keys.map(k=>mdlV98Norm(k.split('_')[0])).filter(Boolean);
        filtered=rows.filter(r=>mdlV98FilialNorm(r.filial)===filial && nomes.some(n=>{
          const o=mdlV98Norm(r._owner?.nome||r._owner?.label||'');
          return o && n && (o===n || o.includes(n) || n.includes(o));
        }));
      }
      return filtered;
    };

    // Se a lista grande ainda não foi carregada no login individual, carrega e redesenha a tela do usuário.
    if(typeof renderReativacaoEnt === 'function' && !window._renderReatEntV98Wrapped){
      window._renderReatEntV98Wrapped = true;
      const _oldRenderReatEnt = renderReativacaoEnt;
      window.renderReativacaoEnt = renderReativacaoEnt = function(ent){
        try{
          const semDados = !(Array.isArray(CLIENTES_SEM_MOVIMENTO) && CLIENTES_SEM_MOVIMENTO.length);
          if(semDados && !window._mdlCsmLoadedV96 && typeof mdlV96LoadClientesSemMovimento==='function'){
            mdlV96LoadClientesSemMovimento().then(()=>{
              try{
                if(detailScreen && !detailScreen.classList.contains('hidden') && currentDetailRef){openEntity(currentDetailRef)}
                else if(typeof renderReativacaoTab==='function'){renderReativacaoTab()}
              }catch(e){console.warn('[MDL V9.8] redraw reativacao individual',e)}
            });
            return `<div class="accordion open"><div class="acc-head"><span>🧡 Clientes sem movimento para reativação</span><span class="acc-hint">carregando lista...</span></div><div class="acc-body"><div class="glass panel"><strong>⏳ Carregando clientes sem movimento</strong><div class="hint">A lista fica no FTP para deixar o dashboard leve. Aguarde alguns segundos.</div></div></div></div>`;
          }
        }catch(e){console.warn('[MDL V9.8] wrapper renderReativacaoEnt',e)}
        return _oldRenderReatEnt.apply(this,arguments);
      };
    }
  }catch(e){console.warn('[MDL V9.8] hotfix reativacao individual falhou',e)}
})();



// ===== V10.0 HOTFIX: reativação abre no acesso individual sem travar =====
(function(){
  try{
    window.MDL_V100_REAT_INDIVIDUAL_LAZY_BASE = true;

    function mdlV100Norm(v){
      try{return normName(v||'')}catch(e){return String(v||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/[^a-z0-9]+/gi,' ').trim().toLowerCase()}
    }
    function mdlV100Filial(v){
      v=String(v||'').trim().toUpperCase();
      if(/^0?\d+$/.test(v)) return 'F'+Number(v);
      if(/^F0\d+$/.test(v)) return 'F'+Number(v.replace(/^F0*/,''));
      return v;
    }
    function mdlV100ExtractRows(payload){
      if(Array.isArray(payload)) return payload;
      if(payload && typeof payload==='object'){
        for(const k of ['clientes','data','rows','items','lista']){
          if(Array.isArray(payload[k])) return payload[k];
        }
      }
      return [];
    }
    async function mdlV100FetchJson(name){
      if(typeof fetchJsonNoCache==='function') return await fetchJsonNoCache(name);
      const sep=String(name).includes('?')?'&':'?';
      const r=await fetch(name+sep+'_='+Date.now(),{cache:'no-store'});
      if(!r.ok) throw new Error(name+' HTTP '+r.status);
      return await r.json();
    }
    function mdlV100SpliceArray(target, rows){
      if(!Array.isArray(target)) return;
      target.splice(0, target.length, ...(Array.isArray(rows)?rows:[]));
    }
    window._mdlV100CsmLoadingPromise = null;
    async function mdlV100LoadClientesSemMovimento(force=false){
      const hasAction = Array.isArray(CLIENTES_SEM_MOVIMENTO) && CLIENTES_SEM_MOVIMENTO.length>0;
      const hasBase = Array.isArray(CLIENTES_SEM_MOVIMENTO_BASE) && CLIENTES_SEM_MOVIMENTO_BASE.length>0;
      if(!force && (hasAction || hasBase)) return true;
      if(window._mdlV100CsmLoadingPromise) return window._mdlV100CsmLoadingPromise;
      window._mdlCsmLoadingV96 = true;
      window._mdlV100CsmLoadingPromise = (async()=>{
        let actionRows=[];
        let baseRows=[];
        try{ actionRows = mdlV100ExtractRows(await mdlV100FetchJson('clientes_sem_movimento.json')); }catch(e){ console.warn('[MDL V10.0] falha clientes_sem_movimento.json',e); }
        try{ baseRows = mdlV100ExtractRows(await mdlV100FetchJson('clientes_sem_movimento_base.json')); }catch(e){ console.warn('[MDL V10.0] falha clientes_sem_movimento_base.json',e); }
        // Para o usuário individual, o principal é não ficar travado. Se a lista acionável estiver vazia,
        // usa a base completa como fallback visual, mantendo o histórico de enviados no cobrancas_log.json.
        if(Array.isArray(CLIENTES_SEM_MOVIMENTO)){
          mdlV100SpliceArray(CLIENTES_SEM_MOVIMENTO, actionRows.length ? actionRows : baseRows);
        }
        try{ mdlV100SpliceArray(CLIENTES_SEM_MOVIMENTO_BASE, baseRows.length ? baseRows : actionRows); }catch(e){}
        try{ window.CLIENTES_SEM_MOVIMENTO_BASE = CLIENTES_SEM_MOVIMENTO_BASE; }catch(e){}
        if((actionRows.length || baseRows.length) && typeof CLIENTES_SEM_MOVIMENTO_META==='object'){
          CLIENTES_SEM_MOVIMENTO_META.base_total = Number(CLIENTES_SEM_MOVIMENTO_META.base_total || baseRows.length || actionRows.length || 0);
          CLIENTES_SEM_MOVIMENTO_META.acionaveis_total = Number(CLIENTES_SEM_MOVIMENTO_META.acionaveis_total || actionRows.length || baseRows.length || 0);
          CLIENTES_SEM_MOVIMENTO_META._loaded_frontend_v100 = true;
        }
        window._mdlCsmLoadedV96 = true;
        window._mdlCsmLoadingV96 = false;
        window._mdlV100CsmLoadingPromise = null;
        console.log('[MDL V10.0] clientes sem movimento carregados no acesso individual:', {acionaveis:actionRows.length, base:baseRows.length});
        return true;
      })().catch(e=>{
        console.warn('[MDL V10.0] falha geral ao carregar reativação',e);
        window._mdlCsmLoadingV96 = false;
        window._mdlV100CsmLoadingPromise = null;
        return false;
      });
      return window._mdlV100CsmLoadingPromise;
    }
    window.mdlV100LoadClientesSemMovimento = mdlV100LoadClientesSemMovimento;

    function mdlV100FindVend(login,nome,filial){
      const fil=mdlV100Filial(filial);
      const log=String(login||'').toLowerCase().trim();
      const nn=mdlV100Norm(nome);
      let arr=[]; try{arr=flattenVendedores()||[]}catch(e){}
      return arr.find(v=>{
        const vf=mdlV100Filial(v.filial);
        if(fil && vf!==fil) return false;
        const vl=String(v.login||'').toLowerCase().trim();
        const vn=mdlV100Norm(v.nome);
        return (log && vl && vl===log) || (nn && vn && (vn===nn || vn.includes(nn) || nn.includes(vn)));
      }) || null;
    }
    function mdlV100CandidateKeys(ent){
      const filial=mdlV100Filial(ent?.filial || usuarioAtual?.filial || '');
      const names=[];
      const add=v=>{v=String(v||'').trim(); if(v && !names.includes(v)) names.push(v)};
      add(ent?.nome); add(ent?.login); add(usuarioAtual?.nome); add(usuarioAtual?.login);
      const found=mdlV100FindVend(ent?.login||usuarioAtual?.login, ent?.nome||usuarioAtual?.nome, filial);
      add(found?.nome); add(found?.login);
      return names.map(n=>mdlV100Norm(n)+'_'+filial).filter(k=>k && k!=='_'+filial);
    }
    function mdlV100RowsFromSource(src){
      const arr=Array.isArray(src)?src:[];
      return arr.map((r,i)=>{
        let idx=Number.isInteger(r._idx)?r._idx:-1;
        if(idx<0 && Array.isArray(CLIENTES_SEM_MOVIMENTO)){
          idx=CLIENTES_SEM_MOVIMENTO.indexOf(r);
          if(idx<0){ CLIENTES_SEM_MOVIMENTO.push(r); idx=CLIENTES_SEM_MOVIMENTO.length-1; }
        }
        return {...r,_idx:idx>=0?idx:i,_owner:reativacaoOwnerInfo(r)};
      });
    }
    function mdlV100FilterRowsForEnt(ent, src){
      if(!ent) return [];
      const filial=mdlV100Filial(ent.filial||usuarioAtual?.filial||'');
      let rows=mdlV100RowsFromSource(src);
      if(ent.type==='filial' || ent.is_gerente){
        return rows.filter(r=>mdlV100Filial(r.filial)===filial && (ent.type==='filial' || String(r._owner?.key||'')===`GERENTE_${filial}` || String(r._owner?.tipo||'')==='filial'));
      }
      const keys=mdlV100CandidateKeys(ent);
      let out=rows.filter(r=>keys.includes(String(r._owner?.key||'')));
      if(!out.length){
        const nomes=keys.map(k=>mdlV100Norm(String(k).split('_')[0])).filter(Boolean);
        out=rows.filter(r=>mdlV100Filial(r.filial)===filial && nomes.some(n=>{
          const o=mdlV100Norm(r._owner?.nome||r._owner?.label||r.responsavel||'');
          return o && n && (o===n || o.includes(n) || n.includes(o));
        }));
      }
      return out;
    }

    // Recria a regra de visibilidade usando lista acionável e, se necessário, a base completa do FTP.
    window.reativacaoRowsParaEnt = function(ent){
      let out=mdlV100FilterRowsForEnt(ent, CLIENTES_SEM_MOVIMENTO||[]);
      if(!out.length && Array.isArray(CLIENTES_SEM_MOVIMENTO_BASE) && CLIENTES_SEM_MOVIMENTO_BASE.length){
        out=mdlV100FilterRowsForEnt(ent, CLIENTES_SEM_MOVIMENTO_BASE);
      }
      return out;
    };
    window.reativacaoRowsPermitidas = function(){
      if(!usuarioAtual || usuarioAtual.tipo==='master' || usuarioAtual.is_viewer){
        const src=(Array.isArray(CLIENTES_SEM_MOVIMENTO)&&CLIENTES_SEM_MOVIMENTO.length)?CLIENTES_SEM_MOVIMENTO:(CLIENTES_SEM_MOVIMENTO_BASE||[]);
        return mdlV100RowsFromSource(src);
      }
      const ent={type:usuarioAtual.is_gerente?'filial':'vendedor', filial:usuarioAtual.filial, nome:usuarioAtual.nome, login:usuarioAtual.login, is_gerente:usuarioAtual.is_gerente};
      return window.reativacaoRowsParaEnt(ent);
    };

    if(typeof renderReativacaoEnt==='function' && !window._renderReatEntV100Wrapped){
      window._renderReatEntV100Wrapped=true;
      const oldRender=renderReativacaoEnt;
      window.renderReativacaoEnt = renderReativacaoEnt = function(ent){
        try{
          const hasAction=Array.isArray(CLIENTES_SEM_MOVIMENTO) && CLIENTES_SEM_MOVIMENTO.length>0;
          const hasBase=Array.isArray(CLIENTES_SEM_MOVIMENTO_BASE) && CLIENTES_SEM_MOVIMENTO_BASE.length>0;
          if(!hasAction && !hasBase){
            mdlV100LoadClientesSemMovimento(false).then(()=>{
              try{
                if(detailScreen && !detailScreen.classList.contains('hidden') && currentDetailRef){openEntity(currentDetailRef)}
                else if(typeof renderReativacaoTab==='function'){renderReativacaoTab()}
              }catch(e){console.warn('[MDL V10.0] redraw individual',e)}
            });
            return `<div class="accordion open"><div class="acc-head"><span>🧡 Clientes sem movimento para reativação</span><span class="acc-hint">carregando lista...</span></div><div class="acc-body"><div class="glass panel"><strong>⏳ Carregando clientes sem movimento</strong><div class="hint">A lista fica no FTP para deixar o dashboard leve. Aguarde alguns segundos. Se demorar, atualize a página uma vez.</div></div></div></div>`;
          }
        }catch(e){console.warn('[MDL V10.0] wrapper renderReativacaoEnt',e)}
        return oldRender.apply(this,arguments);
      };
    }

    // Se o usuário comum entrar direto no painel, dispara o pré-carregamento logo após abrir a tela.
    if(typeof abrirApp==='function' && !window._abrirAppV100Wrapped){
      window._abrirAppV100Wrapped=true;
      const oldAbrirApp=abrirApp;
      window.abrirApp = abrirApp = async function(){
        const r=await oldAbrirApp.apply(this,arguments);
        try{
          if(usuarioAtual && usuarioAtual.tipo!=='master' && !usuarioAtual.is_viewer){
            setTimeout(()=>mdlV100LoadClientesSemMovimento(false).then(()=>{
              try{ if(detailScreen && !detailScreen.classList.contains('hidden') && currentDetailRef) openEntity(currentDetailRef); }catch(e){}
            }), 350);
          }
        }catch(e){}
        return r;
      };
    }

    console.log('[MDL V10.0] hotfix ativo: reativação individual carrega do FTP com fallback para base completa');
  }catch(e){console.warn('[MDL V10.0] hotfix reativação individual falhou',e)}
})();


// ===== V10.2 HOTFIX: meta diária apenas por Realizado Período validado no navegador =====
(function(){
  try{
    const V102_TAG='MDL_V102_META_DIARIA_JS_STRICT';
    function brMoney(v){
      if(typeof v==='number') return Number.isFinite(v)?v:0;
      let s=String(v??'').trim();
      if(!s || /^(nan|null|none)$/i.test(s)) return 0;
      s=s.replace(/R\$/ig,'').replace(/%/g,'').replace(/\s/g,'');
      if(s.includes(',')) s=s.replace(/\./g,'').replace(',','.');
      const n=Number(s); return Number.isFinite(n)?n:0;
    }
    function brPct(v){return brMoney(v)}
    function cell(row, keys){
      row=row||{};
      for(const k of keys){ if(Object.prototype.hasOwnProperty.call(row,k) && row[k]!==null && row[k]!==undefined && String(row[k]).trim()!=='') return row[k]; }
      const norm=k=>String(k||'').normalize('NFD').replace(/[\u0300-\u036f]/g,'').toLowerCase().replace(/[^a-z0-9]+/g,'');
      const wanted=keys.map(norm);
      for(const [k,v] of Object.entries(row)){
        const nk=norm(k);
        if(wanted.includes(nk) && v!==null && v!==undefined && String(v).trim()!=='') return v;
      }
      return '';
    }
    function todayBR(){const d=new Date(); return String(d.getDate()).padStart(2,'0')+'/'+String(d.getMonth()+1).padStart(2,'0')+'/'+d.getFullYear();}
    function filFromText(v){
      const s=String(v||'').toUpperCase();
      let m=s.match(/F(?:ILIAL)?\s*0?(\d{1,2})/); if(m) return 'F'+String(Number(m[1]));
      m=s.match(/\bF\s*0?(\d{1,2})\b/); if(m) return 'F'+String(Number(m[1]));
      return '';
    }
    function isGerMeta(nome){return /\(\s*GER\s*F?\d*\s*\)|\bGERF?\d*\b/i.test(String(nome||''));}
    function validateMetaRow(row){
      const metaTxt=cell(row,['Meta (R$) Período','Meta (R$) Periodo','Meta(R$) Período','Meta(R$) Periodo']);
      const realTxt=cell(row,['Realizado (R$) Período','Realizado (R$) Periodo','Realizado(R$) Período','Realizado(R$) Periodo']);
      const totalTxt=cell(row,['Realizado (R$) Total','Realizado(R$) Total']);
      const atingTxt=cell(row,['Atingido Período','Atingido Periodo','Atingido Período_float','Atingido Periodo_float']);
      const meta=brMoney(metaTxt), real=brMoney(realTxt), total=brMoney(totalTxt);
      if(!(meta>0) || !(real>0)) return null;
      if(total>0 && real>total+0.01) return null;
      const calc=Number(((real/meta)*100).toFixed(2));
      if(!(calc>=100)) return null;
      if(calc>500) return null; // coluna errada/projetado/total lido como realizado período
      const atingSgi=brPct(atingTxt);
      if(atingSgi>0 && Math.abs(atingSgi-calc)>Math.max(2,calc*0.05)) return null;
      return {ating:calc, real, meta, realTxt:String(realTxt||real), metaTxt:String(metaTxt||meta)};
    }
    function getWrap(){return (typeof METAS_VENDAS_DIA==='object' && METAS_VENDAS_DIA) ? METAS_VENDAS_DIA : {metas:{}};}
    window.mdlV102MetaDiariaEntries=function(){
      const wrap=getWrap();
      if(String(wrap.data_consulta||'').trim() && String(wrap.data_consulta||'').trim()!==todayBR()) return [];
      const metas=wrap.metas||{};
      const arr=[]; const seen=new Set();
      function add(item){ const k=[item.kind,item.filial||'',item.nome||''].join('|').toUpperCase(); if(seen.has(k)) return; seen.add(k); arr.push(item); }
      try{
        const mf=(metas.venda_filial_meta&&Array.isArray(metas.venda_filial_meta.linhas))?metas.venda_filial_meta.linhas:[];
        mf.forEach(row=>{
          if(!row || row._is_total) return;
          const v=validateMetaRow(row); if(!v) return;
          const raw=String(cell(row,['Filial','Nome'])||'Filial').trim();
          const filial=filFromText(raw);
          add({nome:raw, filial, kind:'filial', tipo:'filial', ating:v.ating, info:`${filial||raw} · ${v.ating.toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2})}% no dia · ${v.realTxt} / ${v.metaTxt}`, realizado_periodo:v.realTxt, meta_periodo:v.metaTxt});
        });
        const mv=(metas.venda_filial_vendedor_meta&&Array.isArray(metas.venda_filial_vendedor_meta.linhas))?metas.venda_filial_vendedor_meta.linhas:[];
        mv.forEach(row=>{
          if(!row || row._is_total) return;
          const nome=String(cell(row,['Vendedor_2','Nome_2','Nome','Vendedor'])||'Vendedor').trim();
          if(isGerMeta(nome)) return;
          const v=validateMetaRow(row); if(!v) return;
          const filial=filFromText(cell(row,['Vendedor','Filial'])||nome);
          add({nome, filial, kind:'vendedor', tipo:'vendedor', ating:v.ating, info:`${filial||''} · ${v.ating.toLocaleString('pt-BR',{minimumFractionDigits:2,maximumFractionDigits:2})}% no dia · ${v.realTxt} / ${v.metaTxt}`, realizado_periodo:v.realTxt, meta_periodo:v.metaTxt});
        });
      }catch(e){console.warn('[V10.2] meta diária strict falhou',e)}
      arr.sort((a,b)=>Number(b.ating||0)-Number(a.ating||0)||String(a.nome).localeCompare(String(b.nome),'pt-BR'));
      return arr;
    };
    window.mdlV52MetaDiariaEntries = function(){ return window.mdlV102MetaDiariaEntries(); };
    window.mdlV51MetaDiariaCount = function(){ try{return window.mdlV102MetaDiariaEntries().length}catch(e){return 0} };
    window.renderMetaDiariaBatidaAlerts = function(){
      const arr=window.mdlV102MetaDiariaEntries();
      if(!arr.length){
        return `<div class="glass panel" style="margin-bottom:16px;padding:14px 18px;border-color:rgba(34,197,94,.22)"><div class="section-head" style="margin:0 0 8px"><div><h2 style="margin:0;font-size:18px">🎯 Meta diária BATIDA</h2><div class="hint">Mural ativo. Usa somente Realizado Período / Meta Período do Sólidus no dia atual.</div></div><div class="badge">0</div></div><div class="meta-diaria-empty">Nenhuma meta diária batida até o momento.</div></div>`;
      }
      return renderAvisoTicker('Meta diária BATIDA','Controle de Meta Sólidus do dia atual: somente Realizado Período acima de 100%.', arr, {icon:'🎯',color:'rgba(34,197,94,.34)'});
    };
    setTimeout(()=>{try{ if(typeof renderTopMural==='function') renderTopMural(); if(mainTab==='inicio' && typeof renderInicioTab==='function') renderInicioTab(); }catch(e){}},400);
    console.log('[V10.2] hotfix ativo: meta diária strict por Realizado Período', V102_TAG);
  }catch(e){console.warn('[V10.2] falha meta diária strict',e)}
})();

</script>
</body>
</html>
"""

html = template
repls = {
    '__JS_CREDS__': js_creds,
    '__JS_AUTH_STATE__': js_auth_state,
    '__JS_TODOS__': js_todos,
    '__JS_FILIAIS__': js_filiais,
    '__JS_CLIENTES__': js_clientes,
    '__JS_CLIENTES_VEND__': js_clientes_vend,
    '__JS_CLIENTES_TERCEIRO__': js_clientes_terceiro,
    '__JS_CLIENTES_CREDIARISTA__': js_clientes_crediarista,
    '__JS_RECEBIMENTOS__': js_recebimentos,
    '__JS_RECEBIMENTOS_TERCEIRO__': js_recebimentos_terceiro,
    '__JS_RECEBIMENTOS_CREDIARISTA__': js_recebimentos_crediarista,
    '__JS_CREDIARISTAS_MAP__': js_crediaristas_map,
    '__JS_METAS_VENDAS__': js_metas_vendas,
    '__JS_METAS_VENDAS_DIA__': js_metas_vendas_dia,
    '__JS_META_DIARIA_BATIDA_PRECALC__': js_meta_diaria_batida_precalc,
    '__JS_MARGENS_BRUTAS__': js_margens_brutas,
    '__JS_SALES_EMPRESA__': js_sales_empresa,
    '__JS_RENT_EMPRESA__': js_rent_empresa,
    '__JS_SERVICOS_RELATORIO__': js_servicos_relatorio,
    '__CONFIG_META__': json.dumps(CONFIG_META, ensure_ascii=False),
    '__CONFIG_META_IND__': json.dumps(CONFIG_META_IND, ensure_ascii=False),
    '__JS_DESTAQUE__': js_destaque,
    '__JS_HIST_DASH__': js_hist_dash,
    '__JS_QUITADOS_180__': js_quitados_180,
    '__JS_HIST_RECEBIMENTOS_MENSAIS__': js_hist_recebimentos_mensais,
    '__JS_CLIENTES_SEM_MOVIMENTO__': '[]',  # V9.6: lazy-load do JSON no navegador para não pesar o HTML
    '__JS_CLIENTES_SEM_MOVIMENTO_BASE__': '[]',  # V9.6: base completa fica no FTP, não embutida no HTML
    '__JS_CLIENTES_SEM_MOVIMENTO_META__': js_clientes_sem_movimento_meta,
    '__JS_ANIVERSARIANTES__': js_aniversariantes,
    '__JS_DUPLICIDADES_CARTEIRA__': js_duplicidades_carteira,
    '__LOGIN_MASTER__': json.dumps(LOGIN_MASTER, ensure_ascii=False),
    '__SENHA_MASTER__': json.dumps(SENHA_MASTER, ensure_ascii=False),
    '__LOGIN_DIRETOR__': json.dumps(LOGIN_DIRETOR, ensure_ascii=False),
    '__SENHA_DIRETOR__': json.dumps(SENHA_DIRETOR, ensure_ascii=False),
    '__LOGIN_PAINEL__': json.dumps(LOGIN_PAINEL, ensure_ascii=False),
    '__SENHA_PAINEL__': json.dumps(SENHA_PAINEL, ensure_ascii=False),
    '__ORDEM__': json.dumps(ORDEM_FILIAIS, ensure_ascii=False),
    '__TOTAL_P__': json.dumps(total_dash_p, ensure_ascii=False),
    '__TOTAL_PG__': json.dumps(total_dash_pg, ensure_ascii=False),
    '__LARANJITO__': _laranjito,
    '__LOGO__': _logo,
    '__PERIODO__': f"{data_inicio} a {data_fim}",
    '__DASH_VERSION_LABEL__': DASHBOARD_BUILD_VERSION,
    '__JS_DASHBOARD_BUILD_VERSION__': json.dumps(DASHBOARD_BUILD_VERSION, ensure_ascii=False),
    '__JS_DASHBOARD_BUILD_TAG__': json.dumps(DASHBOARD_BUILD_TAG, ensure_ascii=False),
    '__DASHBOARD_UPDATED_AT_LABEL__': json.dumps(now_brasilia().strftime('%d/%m/%Y %H:%M:%S'), ensure_ascii=False),
}
for k,v in repls.items():
    html = html.replace(k, v)

html_path = os.path.join(pasta, 'dashboard_vendedores.html')
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'🌐 HTML salvo: {html_path}')

# =========================================
# 📡 APIs PHP PARA LOG ONLINE / CONFIG ONLINE
# =========================================

HISTORICO_COMISSIONAMENTO_API_PHP = r"""<?php
header('Content-Type: application/json; charset=utf-8');
header('Access-Control-Allow-Origin: *');
header('Access-Control-Allow-Methods: GET, POST, OPTIONS');
header('Access-Control-Allow-Headers: Content-Type');
if ($_SERVER['REQUEST_METHOD'] === 'OPTIONS') { http_response_code(204); exit; }
$file = __DIR__ . '/historico_comissionamento.json';
if (!file_exists($file)) file_put_contents($file, json_encode(['months'=>new stdClass()], JSON_UNESCAPED_UNICODE|JSON_PRETTY_PRINT));
$data = json_decode(file_get_contents($file), true);
if (!is_array($data)) $data = ['months'=>[]];
if (!isset($data['months']) || !is_array($data['months'])) $data['months'] = [];
if ($_SERVER['REQUEST_METHOD'] === 'GET') { echo json_encode(['ok'=>true,'data'=>$data], JSON_UNESCAPED_UNICODE); exit; }
$month = $_POST['month'] ?? date('Y-m');
$payloadRaw = $_POST['payload'] ?? '';
$payload = json_decode($payloadRaw, true);
if (!is_array($payload)) { echo json_encode(['ok'=>false,'error'=>'payload_invalido'], JSON_UNESCAPED_UNICODE); exit; }
$payload['month'] = $month;
$payload['salvo_em_php'] = date('c');
$data['months'][$month] = $payload;
file_put_contents($file, json_encode($data, JSON_UNESCAPED_UNICODE|JSON_PRETTY_PRINT));
echo json_encode(['ok'=>true,'data'=>$data], JSON_UNESCAPED_UNICODE); exit;
?>"""

COBRANCAS_API_PHP = r"""<?php
header('Content-Type: application/json; charset=utf-8');
header('Access-Control-Allow-Origin: *');
header('Access-Control-Allow-Methods: GET, POST, OPTIONS');
header('Access-Control-Allow-Headers: Content-Type');
if ($_SERVER['REQUEST_METHOD'] === 'OPTIONS') { http_response_code(204); exit; }
date_default_timezone_set('America/Sao_Paulo');

$file = __DIR__ . '/cobrancas_log.json';
$backupDir = __DIR__ . '/backups_cobrancas';
if (!file_exists($backupDir)) @mkdir($backupDir, 0777, true);

function read_json_safe($path, $default = []){
  if (!file_exists($path)) return $default;
  $raw = @file_get_contents($path);
  if ($raw === false || trim($raw) === '') return $default;
  $j = json_decode($raw, true);
  return is_array($j) ? $j : $default;
}
function write_json_atomic($path, $data){
  $tmp = $path . '.tmp_' . uniqid('', true);
  $json = json_encode($data, JSON_PRETTY_PRINT|JSON_UNESCAPED_UNICODE);
  if (@file_put_contents($tmp, $json, LOCK_EX) === false) return false;
  return @rename($tmp, $path);
}
function cleanup_backups($backupDir){
  // V4.6: evita estourar FTP. Mantém somente:
  // - latest do dia
  // - latest dos últimos 10 dias
  // - append NDJSON dos últimos 10 dias
  $keepDays = 10;
  $limit = time() - ($keepDays * 86400);
  foreach (glob($backupDir . '/cobrancas_log_*.json') ?: [] as $f) {
    $base = basename($f);
    if (preg_match('/_latest\.json$/', $base)) {
      if (@filemtime($f) < $limit) @unlink($f);
    } else {
      // remove backups antigos por horário/segundo das versões anteriores
      @unlink($f);
    }
  }
  foreach (glob($backupDir . '/cobrancas_log_append_*.ndjson') ?: [] as $f) {
    if (@filemtime($f) < $limit) @unlink($f);
  }
}
function make_backup($file, $backupDir, $tag='before_write'){
  if (!file_exists($file)) return;
  $day = date('Y-m-d');
  // V4.6: não cria 1 JSON a cada clique. Atualiza somente o latest do dia.
  @copy($file, $backupDir . "/cobrancas_log_{$day}_latest.json");
  cleanup_backups($backupDir);
}
function restore_latest_backup($file, $backupDir){
  if (file_exists($file) && filesize($file) > 2) return false;
  $files = glob($backupDir . '/cobrancas_log_*_latest.json');
  if (!$files) $files = glob($backupDir . '/cobrancas_log_*.json');
  if (!$files) return false;
  usort($files, function($a,$b){ return filemtime($b) <=> filemtime($a); });
  if (isset($files[0]) && file_exists($files[0])) { @copy($files[0], $file); return true; }
  return false;
}
function norm_key($s){
  $s = strtoupper(trim((string)$s));
  $s = preg_replace('/\s+/', ' ', $s);
  return $s;
}
function cobranca_key($p){
  if (!empty($p['cobranca_key'])) return (string)$p['cobranca_key'];
  return norm_key($p['cliente'] ?? '') . '|' . trim((string)($p['titulo'] ?? '')) . '|' . trim((string)($p['parcela'] ?? '')) . '|' . trim((string)($p['vencimento'] ?? ''));
}
function cliente_key($p){
  if (!empty($p['cliente_key'])) return (string)$p['cliente_key'];
  return substr(norm_key($p['cpf'] ?? ($p['documento'] ?? ($p['cliente'] ?? ''))), 0, 90);
}

if (!file_exists($file)) {
  restore_latest_backup($file, $backupDir);
  if (!file_exists($file)) write_json_atomic($file, []);
}
$data = read_json_safe($file, []);
if (!is_array($data)) $data = [];

if ($_SERVER['REQUEST_METHOD'] === 'GET') {
  echo json_encode(['ok'=>true,'data'=>$data,'count'=>count($data)], JSON_UNESCAPED_UNICODE); exit;
}

if ($_SERVER['REQUEST_METHOD'] === 'POST') {
  $raw = file_get_contents('php://input');
  if (!mb_check_encoding($raw, 'UTF-8')) { $raw = mb_convert_encoding($raw, 'UTF-8', 'auto'); }
  $payload = json_decode($raw, true);
  if (!is_array($payload)) { echo json_encode(['ok'=>false,'error'=>'payload_invalido','json_error'=>json_last_error_msg()]); exit; }

  if (($payload['action'] ?? '') === 'delete') {
    make_backup($file, $backupDir, 'before_delete');
    $id = $payload['id'] ?? '';
    $cliente = $payload['cliente'] ?? '';
    $titulo = $payload['titulo'] ?? '';
    $parcela = $payload['parcela'] ?? '';
    $novo = [];
    foreach($data as $item){
      $sameId = ($id !== '' && (string)($item['id'] ?? '') === (string)$id);
      $sameComposite = ($cliente !== '' && (string)($item['cliente'] ?? '') === (string)$cliente && (string)($item['titulo'] ?? '') === (string)$titulo && (string)($item['parcela'] ?? '') === (string)$parcela);
      if(!$sameId && !$sameComposite) $novo[] = $item;
    }
    write_json_atomic($file, $novo);
    echo json_encode(['ok'=>true,'count'=>count($novo)], JSON_UNESCAPED_UNICODE); exit;
  }

  if (empty($payload['id'])) $payload['id'] = uniqid('cob_', true);
  $payload['server_time'] = date('c');
  $payload['server_date'] = date('Y-m-d');
  $payload['cliente_key'] = cliente_key($payload);
  $payload['cobranca_key'] = cobranca_key($payload);

  // Antiduplo clique: mesmo usuário + mesmo título + telefone em até 2 minutos não duplica.
  $now = time();
  foreach(array_reverse($data) as $item){
    if (($item['cobranca_key'] ?? '') === $payload['cobranca_key']
      && strtolower((string)($item['usuario'] ?? '')) === strtolower((string)($payload['usuario'] ?? ''))
      && (string)($item['telefone'] ?? '') === (string)($payload['telefone'] ?? '')) {
        $ts = strtotime((string)($item['server_time'] ?? ''));
        if ($ts && abs($now - $ts) <= 120) {
          echo json_encode(['ok'=>true,'id'=>$item['id'] ?? $payload['id'],'duplicado'=>true], JSON_UNESCAPED_UNICODE); exit;
        }
    }
  }

  make_backup($file, $backupDir, 'before_write');
  $data[] = $payload;
  $ok = write_json_atomic($file, $data);
  @file_put_contents($backupDir . '/cobrancas_log_append_' . date('Y-m-d') . '.ndjson', json_encode($payload, JSON_UNESCAPED_UNICODE) . "\n", FILE_APPEND|LOCK_EX);
  @copy($file, $backupDir . '/cobrancas_log_' . date('Y-m-d') . '_latest.json');
  echo json_encode(['ok'=>$ok,'id'=>$payload['id'],'count'=>count($data)], JSON_UNESCAPED_UNICODE); exit;
}
echo json_encode(['ok'=>false,'error'=>'metodo_nao_suportado'], JSON_UNESCAPED_UNICODE);
?>"""

CONFIG_META_API_PHP = r"""<?php
header('Content-Type: application/json; charset=utf-8');
header('Access-Control-Allow-Origin: *');
header('Access-Control-Allow-Methods: GET, POST, OPTIONS');
header('Access-Control-Allow-Headers: Content-Type');
if ($_SERVER['REQUEST_METHOD'] === 'OPTIONS') { http_response_code(204); exit; }
$file = __DIR__ . '/config_meta.json';
if (!file_exists($file)) file_put_contents($file, '{"global":{},"individual":{}}');
if ($_SERVER['REQUEST_METHOD'] === 'GET') {
  $data = json_decode(@file_get_contents($file), true);
  if (!is_array($data)) $data = ['global'=>[], 'individual'=>[]];
  if (!isset($data['global']) || !is_array($data['global'])) $data['global'] = [];
  if (!isset($data['individual']) || !is_array($data['individual'])) $data['individual'] = [];
  echo json_encode(['ok'=>true,'data'=>$data], JSON_UNESCAPED_UNICODE); exit;
}
if ($_SERVER['REQUEST_METHOD'] === 'POST') {
  // V9.3: aceita JSON normal e também payload_b64 via FormData.
  // Isso evita perda de emoji em alguns navegadores/servidores quando o texto vem do teclado do Windows.
  $raw = '';
  if (isset($_POST['payload_b64']) && $_POST['payload_b64'] !== '') {
    $raw = base64_decode((string)$_POST['payload_b64']);
  } else if (isset($_POST['payload_json']) && $_POST['payload_json'] !== '') {
    $raw = (string)$_POST['payload_json'];
  } else {
    $raw = file_get_contents('php://input');
  }
  if (!is_string($raw)) $raw = '';
  if (function_exists('mb_check_encoding') && !mb_check_encoding($raw, 'UTF-8')) {
    $raw = mb_convert_encoding($raw, 'UTF-8', 'auto');
  }
  $payload = json_decode($raw, true);
  if (!is_array($payload)) { echo json_encode(['ok'=>false,'error'=>'payload_invalido','json_error'=>json_last_error_msg()], JSON_UNESCAPED_UNICODE); exit; }
  $global = isset($payload['global']) && is_array($payload['global']) ? $payload['global'] : [];
  $individual = isset($payload['individual']) && is_array($payload['individual']) ? $payload['individual'] : [];
  $save = ['global'=>$global, 'individual'=>$individual, 'updated_at'=>date('c')];
  $json = json_encode($save, JSON_PRETTY_PRINT|JSON_UNESCAPED_UNICODE|JSON_UNESCAPED_SLASHES);
  $ok = @file_put_contents($file, $json, LOCK_EX) !== false;
  echo json_encode(['ok'=>$ok,'saved_keys'=>count($individual),'global_keys'=>count($global)], JSON_UNESCAPED_UNICODE); exit;
}
echo json_encode(['ok'=>false,'error'=>'metodo_nao_suportado'], JSON_UNESCAPED_UNICODE);
?>"""

MESSAGES_API_PHP = r"""<?php
header('Content-Type: application/json; charset=utf-8');
header('Access-Control-Allow-Origin: *');
header('Access-Control-Allow-Methods: GET, POST, OPTIONS');
header('Access-Control-Allow-Headers: Content-Type');
if ($_SERVER['REQUEST_METHOD'] === 'OPTIONS') { http_response_code(204); exit; }
$file = __DIR__ . '/mensagens_log.json';
$uploadDir = __DIR__ . '/uploads_mural';
if (!file_exists($uploadDir)) @mkdir($uploadDir, 0777, true);
if (!file_exists($file)) file_put_contents($file, '[]');
$data = json_decode(@file_get_contents($file), true);
if (!is_array($data)) $data = [];
if ($_SERVER['REQUEST_METHOD'] === 'GET') {
  echo json_encode(['ok'=>true,'data'=>$data], JSON_UNESCAPED_UNICODE); exit;
}
if ($_SERVER['REQUEST_METHOD'] === 'POST') {
  if ((($_POST['action'] ?? '') === 'delete')) {
    $id = $_POST['id'] ?? '';
    if (!$id) { echo json_encode(['ok'=>false,'error'=>'id_obrigatorio']); exit; }
    $novo = [];
    foreach($data as $item){ if((string)($item['id'] ?? '') !== (string)$id) $novo[] = $item; }
    file_put_contents($file, json_encode($novo, JSON_PRETTY_PRINT|JSON_UNESCAPED_UNICODE));
    echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
  }
  if ((($_POST['action'] ?? '') === 'archive_master')) {
    $id = $_POST['id'] ?? '';
    foreach($data as &$item){ if((string)($item['id'] ?? '') === (string)$id) { $item['hidden_on_master'] = true; } }
    file_put_contents($file, json_encode($data, JSON_PRETTY_PRINT|JSON_UNESCAPED_UNICODE));
    echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
  }
  if ((($_POST['action'] ?? '') === 'mark_read')) {
    $id = $_POST['id'] ?? '';
    $userKey = $_POST['user_key'] ?? '';
    $userKeysRaw = $_POST['user_keys'] ?? '';
    $keysToSave = [];
    if ($userKey) $keysToSave[] = (string)$userKey;
    if ($userKeysRaw) {
      $tmpKeys = json_decode($userKeysRaw, true);
      if (is_array($tmpKeys)) foreach($tmpKeys as $k) if ($k) $keysToSave[] = (string)$k;
    }
    $keysToSave = array_values(array_unique($keysToSave));
    if (!$id || !count($keysToSave)) { echo json_encode(['ok'=>false,'error'=>'parametros_obrigatorios']); exit; }
    foreach($data as &$item){
      if ((string)($item['id'] ?? '') === (string)$id) {
        if (!isset($item['read_by']) || !is_array($item['read_by'])) $item['read_by'] = [];
        foreach($keysToSave as $kSave){ if (!in_array($kSave, $item['read_by'])) $item['read_by'][] = $kSave; }
      }
    }
    file_put_contents($file, json_encode($data, JSON_PRETTY_PRINT|JSON_UNESCAPED_UNICODE));
    echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
  }
  $item = [
    'id' => uniqid('msg_', true),
    'target_type' => $_POST['target_type'] ?? 'all',
    'target_id' => $_POST['target_id'] ?? 'ALL',
    'target_label' => $_POST['target_label'] ?? ($_POST['target_id'] ?? 'ALL'),
    'title' => $_POST['title'] ?? 'Aviso',
    'body' => $_POST['body'] ?? '',
    'message_kind' => $_POST['message_kind'] ?? 'notice',
    'expires_at' => $_POST['expires_at'] ?? '',
    'read_by' => [],
    'server_time' => date('c')
  ];
  if (!empty($_FILES['media']['name'])) {
    $name = preg_replace('/[^A-Za-z0-9._-]/', '_', basename($_FILES['media']['name']));
    $dest = $uploadDir . '/' . time() . '_' . $name;
    if (move_uploaded_file($_FILES['media']['tmp_name'], $dest)) {
      $scheme = (!empty($_SERVER['HTTPS']) && $_SERVER['HTTPS'] !== 'off') ? 'https' : 'http';
      $base = $scheme . '://' . $_SERVER['HTTP_HOST'] . rtrim(dirname($_SERVER['SCRIPT_NAME']), '/\\');
      $item['media_url'] = $base . '/uploads_mural/' . basename($dest);
      $item['media_type'] = $_FILES['media']['type'] ?? '';
    }
  }
  $data[] = $item;
  file_put_contents($file, json_encode($data, JSON_PRETTY_PRINT|JSON_UNESCAPED_UNICODE));
  echo json_encode(['ok'=>true,'data'=>$item], JSON_UNESCAPED_UNICODE); exit;
}
echo json_encode(['ok'=>false,'error'=>'metodo_nao_suportado'], JSON_UNESCAPED_UNICODE);
?>"""



CREDENCIAIS_API_PHP = r"""<?php
header('Content-Type: application/json; charset=utf-8');
header('Access-Control-Allow-Origin: *');
header('Access-Control-Allow-Methods: GET, POST, OPTIONS');
header('Access-Control-Allow-Headers: Content-Type');
if ($_SERVER['REQUEST_METHOD'] === 'OPTIONS') { http_response_code(204); exit; }

$file = __DIR__ . '/credenciais_dashboard.json';
$resetLog = __DIR__ . '/password_reset_requests.json';
$masterEmail = 'sac@moveisdolar.com.br';

if (!file_exists($file)) file_put_contents($file, json_encode(['users'=>new stdClass(),'director'=>new stdClass(),'password_reset_requests'=>[]], JSON_UNESCAPED_UNICODE));
if (!file_exists($resetLog)) file_put_contents($resetLog, '[]');

$data = json_decode(@file_get_contents($file), true);
if (!is_array($data)) $data = ['users'=>[], 'director'=>[], 'password_reset_requests'=>[]];
if (!isset($data['users']) || !is_array($data['users'])) $data['users'] = [];
if (!isset($data['director']) || !is_array($data['director'])) $data['director'] = [];
if (!isset($data['password_reset_requests']) || !is_array($data['password_reset_requests'])) $data['password_reset_requests'] = [];

function save_all($file, $data){ file_put_contents($file, json_encode($data, JSON_PRETTY_PRINT|JSON_UNESCAPED_UNICODE)); }
function resolve_login_ref(&$data, $login){
  $login = strtolower(trim((string)$login));
  if ($login === 'diretorcomercial') return ['type'=>'director', 'key'=>'director'];
  foreach(($data['users'] ?? []) as $k=>$u){ if(strtolower((string)$k)===$login || strtolower((string)($u['login'] ?? ''))===$login) return ['type'=>'user', 'key'=>$k]; }
  return null;
}
function mark_reset_resolved(&$data, $login){ foreach(($data['password_reset_requests'] ?? []) as &$req){ if (strtolower((string)($req['login'] ?? ''))===strtolower((string)$login) && (($req['status'] ?? 'pendente')==='pendente')) $req['status']='resolvido'; } }
function norm_colab_key($s){ $s = strtoupper(trim((string)$s)); $s = iconv('UTF-8','ASCII//TRANSLIT//IGNORE',$s); $s = preg_replace('/[^A-Z0-9]+/', ' ', $s); return trim(preg_replace('/\s+/', ' ', $s)); }
function colab_status_key($nome, $filial, $isGerente=false){ return norm_colab_key($nome).'|'.strtoupper(trim((string)$filial)).'|'.($isGerente ? 'GERENTE' : 'VENDEDOR'); }
function ensure_colab_status(&$data){ if(!isset($data['colaborador_status']) || !is_array($data['colaborador_status'])) $data['colaborador_status'] = []; }

if ($_SERVER['REQUEST_METHOD'] === 'GET') { ensure_colab_status($data); echo json_encode(['ok'=>true,'data'=>$data], JSON_UNESCAPED_UNICODE); exit; }
$action = $_POST['action'] ?? '';
if ($action === 'admin_set_access_block') {
  $blocked = (($_POST['blocked'] ?? '0') === '1');
  $reason = trim((string)($_POST['reason'] ?? ''));
  $data['access_blocked'] = $blocked;
  $data['access_blocked_reason'] = $blocked ? ($reason ?: 'Sistema em atualização. Aguarde liberação pelo Master.') : '';
  if ($blocked) { $data['access_blocked_at'] = date('c'); }
  else { $data['access_unblocked_at'] = date('c'); }
  save_all($file, $data);
  echo json_encode(['ok'=>true,'data'=>$data], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'admin_change_login') {
  $old = strtolower(trim($_POST['old_login'] ?? ''));
  $new = strtolower(trim($_POST['new_login'] ?? ''));
  if (!$old || !$new || strlen($new) < 3) { echo json_encode(['ok'=>false,'error'=>'dados_invalidos']); exit; }
  if (!preg_match('/^[a-z0-9._-]+$/', $new)) { echo json_encode(['ok'=>false,'error'=>'login_invalido']); exit; }
  if ($new === 'master' || $new === 'diretorcomercial') { echo json_encode(['ok'=>false,'error'=>'login_reservado']); exit; }
  if (resolve_login_ref($data, $new)) { echo json_encode(['ok'=>false,'error'=>'login_ja_existe']); exit; }
  $ref = resolve_login_ref($data, $old);
  if (!$ref || $ref['type'] !== 'user') { echo json_encode(['ok'=>false,'error'=>'login_nao_encontrado']); exit; }
  $oldKey = $ref['key'];
  $user = $data['users'][$oldKey];
  unset($data['users'][$oldKey]);
  $user['login'] = $new;
  $user['login_original'] = $old;
  $user['login_changed_at'] = date('c');
  $data['users'][$new] = $user;
  foreach(($data['password_reset_requests'] ?? []) as &$req){
    if (strtolower((string)($req['login'] ?? '')) === $old) $req['login'] = $new;
  }
  save_all($file, $data);
  echo json_encode(['ok'=>true,'old_login'=>$old,'new_login'=>$new,'data'=>$data], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'change_password') {
  $login = strtolower(trim($_POST['login'] ?? '')); $current = strval($_POST['current_password'] ?? ''); $new = strval($_POST['new_password'] ?? '');
  if (!$login || !$current || !$new) { echo json_encode(['ok'=>false,'error'=>'parametros_obrigatorios']); exit; }
  if (strlen($new) < 4) { echo json_encode(['ok'=>false,'error'=>'senha_curta']); exit; }
  $ref = resolve_login_ref($data, $login); if (!$ref) { echo json_encode(['ok'=>false,'error'=>'login_nao_encontrado']); exit; }
  if ($ref['type'] === 'director') { if (($data['director']['password'] ?? '') !== $current) { echo json_encode(['ok'=>false,'error'=>'senha_atual_invalida']); exit; } $data['director']['password']=$new; $data['director']['must_change_password']=false; }
  else { if (($data['users'][$ref['key']]['password'] ?? '') !== $current) { echo json_encode(['ok'=>false,'error'=>'senha_atual_invalida']); exit; } $data['users'][$ref['key']]['password']=$new; $data['users'][$ref['key']]['must_change_password']=false; }
  mark_reset_resolved($data, $login); save_all($file, $data); echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'admin_update_user_status') {
  ensure_colab_status($data);
  $login = strtolower(trim($_POST['login'] ?? ''));
  $ref = resolve_login_ref($data, $login);
  if (!$ref || $ref['type'] !== 'user') { echo json_encode(['ok'=>false,'error'=>'login_nao_encontrado']); exit; }
  $key = $ref['key'];
  $u =& $data['users'][$key];
  $status = strtolower(trim($_POST['status'] ?? 'ativo'));
  $status = in_array($status, ['inativo','desligado','bloqueado']) ? 'inativo' : 'ativo';
  $flags = ['participa_cobrancas','participa_sem_movimento','participa_aniversariantes','participa_murais'];
  $u['status_operacional'] = $status;
  $u['access_disabled'] = ($status !== 'ativo');
  foreach($flags as $fl){ $u[$fl] = (($_POST[$fl] ?? '1') === '1'); }
  $u['data_entrada'] = trim((string)($_POST['data_entrada'] ?? ($u['data_entrada'] ?? '')));
  $u['data_saida'] = trim((string)($_POST['data_saida'] ?? ''));
  $u['substituto'] = trim((string)($_POST['substituto'] ?? ''));
  $u['obs'] = trim((string)($_POST['obs'] ?? ''));
  $sk = colab_status_key($u['nome'] ?? $login, $u['filial'] ?? '', !empty($u['is_gerente']));
  $data['colaborador_status'][$sk] = [
    'login'=>$key, 'nome'=>($u['nome'] ?? $login), 'filial'=>($u['filial'] ?? ''),
    'tipo'=>!empty($u['is_gerente'])?'Gerente':(!empty($u['is_crediarista'])?'Crediarista':(!empty($u['is_terceiro'])?'Cobrança terceiro':(!empty($u['is_viewer'])?'Painel':'Vendedor'))),
    'status'=>$status,
    'participa_cobrancas'=>$u['participa_cobrancas'],
    'participa_sem_movimento'=>$u['participa_sem_movimento'],
    'participa_aniversariantes'=>$u['participa_aniversariantes'],
    'participa_murais'=>$u['participa_murais'],
    'data_entrada'=>$u['data_entrada'], 'data_saida'=>$u['data_saida'], 'substituto'=>$u['substituto'], 'obs'=>$u['obs'],
    'updated_at'=>date('c')
  ];
  save_all($file, $data);
  echo json_encode(['ok'=>true,'data'=>$data], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'admin_set_password') {
  $login = strtolower(trim($_POST['login'] ?? '')); $new = strval($_POST['new_password'] ?? ''); $must = ($_POST['must_change_password'] ?? '0') === '1';
  if (!$login || !$new || strlen($new) < 4) { echo json_encode(['ok'=>false,'error'=>'dados_invalidos']); exit; }
  $ref = resolve_login_ref($data, $login); if (!$ref) { echo json_encode(['ok'=>false,'error'=>'login_nao_encontrado']); exit; }
  if ($ref['type'] === 'director') { $data['director']['password']=$new; $data['director']['must_change_password']=$must; }
  else { $data['users'][$ref['key']]['password']=$new; $data['users'][$ref['key']]['must_change_password']=$must; }
  mark_reset_resolved($data, $login); save_all($file, $data); echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'admin_force_change') {
  $login = strtolower(trim($_POST['login'] ?? '')); if (!$login) { echo json_encode(['ok'=>false,'error'=>'login_obrigatorio']); exit; }
  $ref = resolve_login_ref($data, $login); if (!$ref) { echo json_encode(['ok'=>false,'error'=>'login_nao_encontrado']); exit; }
  if ($ref['type'] === 'director') $data['director']['must_change_password']=true; else $data['users'][$ref['key']]['must_change_password']=true;
  mark_reset_resolved($data, $login); save_all($file, $data); echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'admin_create_user') {
  $login = strtolower(trim($_POST['login'] ?? '')); $nome = trim($_POST['nome'] ?? ''); $filial = strtoupper(trim($_POST['filial'] ?? '')); $senha = strval($_POST['password'] ?? ''); $tipo = strtolower(trim($_POST['tipo'] ?? 'crediarista'));
  if (!$login || !$nome || !$filial || !$senha || strlen($senha) < 4) { echo json_encode(['ok'=>false,'error'=>'dados_invalidos']); exit; }
  if (resolve_login_ref($data, $login)) { echo json_encode(['ok'=>false,'error'=>'login_ja_existe']); exit; }
  $data['users'][$login] = ['login'=>$login,'nome'=>$nome,'filial'=>$filial,'password'=>$senha,'must_change_password'=>true,'is_gerente'=>false,'is_terceiro'=>($tipo==='cobranca'),'is_crediarista'=>($tipo!=='cobranca'),'only_cobranca'=>true];
  save_all($file, $data); echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'resolve_reset') {
  $login = strtolower(trim($_POST['login'] ?? '')); if (!$login) { echo json_encode(['ok'=>false,'error'=>'login_obrigatorio']); exit; }
  $ref = resolve_login_ref($data, $login); if (!$ref) { echo json_encode(['ok'=>false,'error'=>'login_nao_encontrado']); exit; }
  if ($ref['type'] === 'director') $data['director']['must_change_password']=true; else $data['users'][$ref['key']]['must_change_password']=true;
  foreach(($data['password_reset_requests'] ?? []) as &$req){
    if (strtolower((string)($req['login'] ?? ''))===strtolower((string)$login) && (($req['status'] ?? 'pendente')==='pendente')) {
      $req['status']='resolvido';
      $req['resolved_at']=date('c');
    }
  }
  save_all($file, $data); echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'clear_reset_history') {
  $mode = strtolower(trim($_POST['mode'] ?? 'resolved'));
  if ($mode === 'all') $data['password_reset_requests'] = [];
  else $data['password_reset_requests'] = array_values(array_filter(($data['password_reset_requests'] ?? []), function($r){ return (($r['status'] ?? 'pendente') !== 'resolvido'); }));
  save_all($file, $data); echo json_encode(['ok'=>true], JSON_UNESCAPED_UNICODE); exit;
}
if ($action === 'request_reset') {
  $login = strtolower(trim($_POST['login'] ?? '')); $obs = trim($_POST['obs'] ?? ''); if (!$login) { echo json_encode(['ok'=>false,'error'=>'login_obrigatorio']); exit; }
  $req = ['id'=>uniqid('reset_', true), 'login'=>$login, 'obs'=>$obs, 'created_at'=>date('c'), 'status'=>'pendente'];
  $data['password_reset_requests'][] = $req; save_all($file, $data);
  $log = json_decode(@file_get_contents($resetLog), true); if (!is_array($log)) $log = []; $log[] = $req; file_put_contents($resetLog, json_encode($log, JSON_PRETTY_PRINT|JSON_UNESCAPED_UNICODE));
  $assunto = 'Dashboard MDL - solicitacao de recuperacao de senha'; $mensagem = "Usuario: {$login}\nObservacao: {$obs}\nData: {$req['created_at']}\n"; $headers = "From: no-reply@moveisdolar.com.br\r\n"; $mail_ok = @mail($masterEmail, $assunto, $mensagem, $headers);
  echo json_encode(['ok'=>true,'email_sent'=>$mail_ok], JSON_UNESCAPED_UNICODE); exit;
}
echo json_encode(['ok'=>false,'error'=>'acao_nao_suportada'], JSON_UNESCAPED_UNICODE);
?>"""

HISTORICO_API_PHP = r"""<?php
header('Content-Type: application/json; charset=utf-8');
header('Access-Control-Allow-Origin: *');
header('Access-Control-Allow-Methods: GET, OPTIONS');
header('Access-Control-Allow-Headers: Content-Type');
if ($_SERVER['REQUEST_METHOD'] === 'OPTIONS') { http_response_code(204); exit; }
$file = __DIR__ . '/historico_dashboard.json';
if (!file_exists($file)) file_put_contents($file, '{"dates":{},"months_closed":{}}');
$data = json_decode(@file_get_contents($file), true);
if (!is_array($data)) $data = ['dates'=>[], 'months_closed'=>[]];
echo json_encode(['ok'=>true,'data'=>$data], JSON_UNESCAPED_UNICODE); exit;
?>"""

# =========================================
# 📤 UPLOAD AUTOMÁTICO VIA FTP
# =========================================
import ftplib
from io import BytesIO

FTP_HOST  = 'moveisdolar.com.br'
FTP_USER  = 'moveisdolar3'
FTP_PASS  = 'Deg27ll02mdl2301#'
FTP_DIR   = '/public_html/colaborador'

MODO_TESTE_LOCAL = os.getenv('MODO_TESTE_LOCAL', '0') == '1'
if MODO_TESTE_LOCAL:
    print('🧪 MODO_TESTE_LOCAL=1 ativo: NÃO envia arquivos ao FTP de produção.')

if FTP_USER and FTP_PASS and not MODO_TESTE_LOCAL:
    try:
        print('\n📤 Enviando arquivos para o servidor FTP...')
        ftp = ftplib.FTP()
        ftp.connect(FTP_HOST, 21, timeout=30)
        ftp.login(FTP_USER, FTP_PASS)
        ftp.encoding = 'utf-8'
        try:
            ftp.cwd(FTP_DIR)
        except Exception:
            try:
                ftp.mkd(FTP_DIR)
            except Exception:
                pass
            ftp.cwd(FTP_DIR)
        with open(html_path, 'rb') as f_html:
            ftp.storbinary('STOR dashboard_vendedores.html', f_html)
        ftp.storbinary('STOR cobrancas_api.php', BytesIO(COBRANCAS_API_PHP.encode('utf-8')))
        ftp.storbinary('STOR config_meta_api.php', BytesIO(CONFIG_META_API_PHP.encode('utf-8')))
        ftp.storbinary('STOR mensagens_api.php', BytesIO(MESSAGES_API_PHP.encode('utf-8')))
        ftp.storbinary('STOR credenciais_api.php', BytesIO(CREDENCIAIS_API_PHP.encode('utf-8')))
        ftp.storbinary('STOR historico_api.php', BytesIO(HISTORICO_API_PHP.encode('utf-8')))
        ftp.storbinary('STOR historico_comissionamento_api.php', BytesIO(HISTORICO_COMISSIONAMENTO_API_PHP.encode('utf-8')))
        try:
            # V9.5: config_meta.json salvo pelo dashboard/API é fonte de verdade no FTP.
            # Se o main não conseguiu sincronizar do servidor, NÃO sobe default/local vazio por cima.
            if _config_meta_loaded_from_remote or os.getenv('FORCE_UPLOAD_CONFIG_META','0') == '1':
                with open(_config_meta_path, 'rb') as f_cfg:
                    ftp.storbinary('STOR config_meta.json', f_cfg)
                    print('✅ FTP: config_meta.json remoto preservado/enviado')
            else:
                print('🛡️ V9.7: config_meta.json do FTP não foi sobrescrito pelo main; configurações salvas no dashboard permanecem.')
        except Exception as e_cfg_upload:
            print(f'⚠️ V9.7: não enviei config_meta.json para evitar reset indevido: {e_cfg_upload}')
        try:
            with open(cred_state_path, 'rb') as f_credstate:
                ftp.storbinary('STOR credenciais_dashboard.json', f_credstate)
        except Exception:
            ftp.storbinary('STOR credenciais_dashboard.json', BytesIO(b'{"users":{},"director":{},"password_reset_requests":[]}'))
        try:
            with open(_hist_dash_path, 'rb') as f_hist:
                ftp.storbinary('STOR historico_dashboard.json', f_hist)
        except Exception:
            ftp.storbinary('STOR historico_dashboard.json', BytesIO(b'{"dates":{},"months_closed":{}}'))
        try:
            with open(COMISSAO_HIST_PATH, 'rb') as f_ch:
                ftp.storbinary('STOR historico_comissao_cobranca.json', f_ch)
        except Exception:
            ftp.storbinary('STOR historico_comissao_cobranca.json', BytesIO(b'{"months":{}}'))
        try:
            with open(_fechamento_path, 'rb') as f_fech:
                ftp.storbinary('STOR fechamentos_mensais.json', f_fech)
        except Exception:
            ftp.storbinary('STOR fechamentos_mensais.json', BytesIO(b'{"months":{}}'))
        try:
            if os.path.exists(RECEBIMENTOS_MENSAL_PATH):
                with open(RECEBIMENTOS_MENSAL_PATH, 'rb') as f_hrm:
                    ftp.storbinary('STOR historico_recebimentos_mensais.json', f_hrm)
                print('✅ FTP: historico_recebimentos_mensais.json enviado')
        except Exception as e_hrm_ftp:
            print(f'⚠️ Erro enviando historico_recebimentos_mensais.json: {e_hrm_ftp}')
        try:
            _csm_path = os.path.join(pasta, 'clientes_sem_movimento.json')
            _ani_path = os.path.join(pasta, 'aniversariantes_dia.json')
            _csm_seen_path = os.path.join(pasta, 'clientes_sem_movimento_seen.json')
            _csm_base_path = os.path.join(pasta, 'clientes_sem_movimento_base.json')
            _listas_forcadas = _baixou_listas_pesadas_v95()

            # V9.5: em deploy comum, não sobrescreve listas remotas com JSON vazio/local incompleto.
            if os.path.exists(_csm_path) and (_listas_forcadas or _json_has_items_v95(_csm_path, ('clientes',))):
                with open(_csm_path, 'rb') as f_csm:
                    ftp.storbinary('STOR clientes_sem_movimento.json', f_csm)
                    print('✅ FTP: clientes_sem_movimento.json enviado')
            else:
                print('🛡️ V9.5: pulando clientes_sem_movimento.json para não zerar lista do FTP')

            if os.path.exists(_ani_path) and (_listas_forcadas or _json_has_items_v95(_ani_path, ('dados','clientes'))):
                with open(_ani_path, 'rb') as f_ani:
                    ftp.storbinary('STOR aniversariantes_dia.json', f_ani)
                    print('✅ FTP: aniversariantes_dia.json enviado')
            else:
                print('🛡️ V9.5: pulando aniversariantes_dia.json para não zerar lista do FTP')

            if os.path.exists(_csm_seen_path) and (_listas_forcadas or _json_has_items_v95(_csm_seen_path, ('seen',))):
                with open(_csm_seen_path, 'rb') as f_csm_seen:
                    ftp.storbinary('STOR clientes_sem_movimento_seen.json', f_csm_seen)
                    print('✅ FTP: clientes_sem_movimento_seen.json enviado')
            else:
                print('🛡️ V9.5: pulando clientes_sem_movimento_seen.json vazio/local')

            if os.path.exists(_csm_base_path) and (_listas_forcadas or _json_has_items_v95(_csm_base_path, ('clientes',))):
                with open(_csm_base_path, 'rb') as f_csm_base:
                    ftp.storbinary('STOR clientes_sem_movimento_base.json', f_csm_base)
                    print('✅ FTP: clientes_sem_movimento_base.json enviado')
            else:
                print('🛡️ V9.5: pulando clientes_sem_movimento_base.json para não zerar base do FTP')
            _dup_path = os.path.join(pasta, 'relatorio_duplicidades_carteira.json')
            if os.path.exists(_dup_path):
                with open(_dup_path, 'rb') as f_dup:
                    ftp.storbinary('STOR relatorio_duplicidades_carteira.json', f_dup)
        except Exception as e_extra_ftp:
            print(f'⚠️ Erro ao enviar extras dashboard 2.0 ao FTP: {e_extra_ftp}')

        try:
            if quitados_180_info.get('json_path') and os.path.exists(quitados_180_info['json_path']):
                with open(quitados_180_info['json_path'], 'rb') as f_q_json:
                    ftp.storbinary('STOR quitados_180d_contas_receber.json', f_q_json)
            if quitados_180_info.get('xlsx_path') and os.path.exists(quitados_180_info['xlsx_path']):
                with open(quitados_180_info['xlsx_path'], 'rb') as f_q_xlsx:
                    ftp.storbinary('STOR quitados_180d_contas_receber.xlsx', f_q_xlsx)
        except Exception as e_q_ftp:
            print(f'⚠️ Erro ao enviar quitados 180d ao FTP: {e_q_ftp}')

        # Pacote de vendas/margem/rentabilidade/serviços/diária fica exclusivo do dashboard_sales_worker_headless.py.
        # Isso evita o navegador misturar arquivos de horários diferentes.
        print('ℹ️ MAIN: não publica metas_vendas/margens/sales_version; pacote de vendas é exclusivo do sales worker.')
        try:
            ftp.size('cobrancas_log.json')
        except Exception:
            ftp.storbinary('STOR cobrancas_log.json', BytesIO(b'[]'))
        try:
            ftp.size('mensagens_log.json')
        except Exception:
            ftp.storbinary('STOR mensagens_log.json', BytesIO(b'[]'))
        try:
            _dashboard_ver = json.dumps({'updated_at': now_brasilia().isoformat(), 'updated_at_label': now_brasilia().strftime('%d/%m/%Y %H:%M:%S'), 'timezone': 'America/Sao_Paulo', 'scope': 'dashboard_full'}, ensure_ascii=False).encode('utf-8')
            ftp.storbinary('STOR dashboard_version.json', BytesIO(_dashboard_ver))
            # sales_version.json é exclusivo do dashboard_sales_worker_headless.py para sincronizar vendas/margens/serviços juntos.
        except Exception as e_ver_ftp:
            print(f'⚠️ Erro enviando arquivos de versão: {e_ver_ftp}')
        ftp.quit()
        print('✅ Upload concluído → https://moveisdolar.com.br/colaborador/')
    except Exception as e:
        print(f'⚠️ Erro no upload FTP: {e}')
else:
    print('\nℹ️ FTP não configurado. Envie manualmente:')
    print(f'   {html_path} → {FTP_DIR}/dashboard_vendedores.html')

print('\n🔥 FINALIZADO!')
print("🔎 Fechamento automático em ambiente Railway")
driver.quit()


# MDL_V99_RESUMO_TELEGRAM_LOGS_FIX: dashboard mantido com hotfix de resumo Telegram via telegram_monitor_mdl_v23.

# MDL_V101_META_DIARIA_VALIDACAO_PERIODO: usa somente Realizado Período / Meta Período e bloqueia Projetado/Total.
# MDL_V102_META_DIARIA_JS_STRICT: renderização do mural ignora precalc antigo e valida no navegador.

# MDL_V103_CREDIARISTA_FREEZE_FIX
