# VERSAO: COBRANCA10_WORKER_V17_META_REAL_TOTAL_DIARIA_MARGEM_FLUXO
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pandas as pd
import time
import re
import os
import webbrowser
import random
import json
import unicodedata
import urllib.request
import urllib.error
import tempfile
from zoneinfo import ZoneInfo
from bs4 import BeautifulSoup

LOGIN = "administrativo01.moveisdolar"
SENHA = "mdladm01"
URL   = "https://smart.sgisistemas.com.br"
APP_TZ = ZoneInfo(os.getenv("APP_TZ", "America/Sao_Paulo"))

def now_brasilia():
    return datetime.now(APP_TZ)

# ===== DATAS
hoje        = datetime.now()
data_inicio = (hoje - timedelta(days=90)).strftime("%d/%m/%Y")
data_fim    = (hoje - timedelta(days=15)).strftime("%d/%m/%Y")

# ===== CHROME
options = Options()
IS_RAILWAY = bool(os.getenv("RAILWAY_ENVIRONMENT") or os.getenv("RUN_ON_RAILWAY") == "1")
pasta = os.path.dirname(os.path.abspath(__file__))
download_dir = pasta if not IS_RAILWAY else tempfile.gettempdir()

if IS_RAILWAY:
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1600,2200")
    options.add_argument("--disable-blink-features=AutomationControlled")
else:
    options.add_argument("--start-maximized")

prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(options=options)
wait   = WebDriverWait(driver, 40)


# ===== SELENIUM HELPERS
def esperar_sumir_overlays(timeout=10):
    fim = time.time() + timeout
    overlays = [
        ".blockUI", ".modal-backdrop", ".loading", ".spinner",
        ".ui-widget-overlay", ".select2-drop-mask"
    ]
    while time.time() < fim:
        try:
            ativos = driver.execute_script("""
                const sels = arguments[0];
                return sels.some(sel =>
                    Array.from(document.querySelectorAll(sel)).some(el => {
                        const s = window.getComputedStyle(el);
                        return s.display !== 'none' && s.visibility !== 'hidden' && el.offsetParent !== null;
                    })
                );
            """, overlays)
            if not ativos:
                return True
        except Exception:
            return True
        time.sleep(0.2)
    return False


def localizar_xpath_clickavel(xpath, timeout=20):
    elem = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center', inline:'center'});", elem
    )
    esperar_sumir_overlays()
    try:
        elem = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, xpath))
        )
    except Exception:
        elem = driver.find_element(By.XPATH, xpath)
    return elem


def clicar_seguro_xpath(xpath, timeout=20):
    ultimo_erro = None
    for _ in range(3):
        try:
            elem = localizar_xpath_clickavel(xpath, timeout=timeout)
            try:
                elem.click()
            except Exception:
                driver.execute_script("arguments[0].click();", elem)
            return elem
        except Exception as e:
            ultimo_erro = e
            time.sleep(0.5)
    raise ultimo_erro


def abrir_multiselect_por_label(label_texto, timeout=20):
    xpath_botao = (
        f"//label[contains(normalize-space(.),'{label_texto}')]/following::button[1]"
    )
    botao = clicar_seguro_xpath(xpath_botao, timeout=timeout)
    time.sleep(1)
    return botao


def marcar_multiselect_por_label(label_texto, valores, timeout=20, limpar_antes=True):
    abrir_multiselect_por_label(label_texto, timeout=timeout)
    xpath_ul = f"//label[contains(normalize-space(.),'{label_texto}')]/following::ul[1]"
    container = wait.until(EC.presence_of_element_located((By.XPATH, xpath_ul)))
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center', inline:'nearest'});", container
    )
    time.sleep(0.5)
    checkboxes = container.find_elements(By.XPATH, ".//input[@type='checkbox']")
    valores = {str(v) for v in valores}

    if limpar_antes:
        for c in checkboxes:
            if c.is_selected():
                driver.execute_script("arguments[0].click();", c)
                time.sleep(0.05)

    for c in checkboxes:
        if c.get_attribute("value") in valores and not c.is_selected():
            driver.execute_script("arguments[0].click();", c)
            time.sleep(0.05)

    driver.execute_script("document.body.click();")
    time.sleep(0.5)
    return checkboxes


def normalizar_texto_match(texto):
    s = str(texto or "").strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def set_input_value_safe(by, locator, value):
    el = wait.until(EC.presence_of_element_located((by, locator)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    try:
        el.click()
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.DELETE)
        el.send_keys(value)
    except Exception:
        driver.execute_script("arguments[0].value = arguments[1];", el, value)
        driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", el)
    return el


def clicar_primeiro_filtro_metas():
    xpaths = [
        "//button[contains(.,'Filtrar')]",
        "//a[contains(.,'Filtrar')]",
        "//button[@type='submit' and contains(.,'Filtrar')]",
    ]
    ultimo = None
    for xp in xpaths:
        try:
            return clicar_seguro_xpath(xp, timeout=10)
        except Exception as e:
            ultimo = e
    if ultimo:
        raise ultimo



def extrair_linhas_tabela_metas():
    """
    Lê a tabela de metas de forma robusta usando os atributos data-title/data-value.
    O SGI muda os IDs das metas todo mês (ex.: 208, 209, 210, 211), então não usamos ID fixo.
    """
    wait.until(EC.presence_of_all_elements_located((By.XPATH, "//table//tbody/tr")))
    linhas = []

    # Primeira tentativa: via JavaScript, pegando data-title/data-value da própria tabela.
    try:
        rows = driver.execute_script("""
            const out = [];
            document.querySelectorAll('table tbody tr').forEach(tr => {
                const obj = {};
                tr.querySelectorAll('td').forEach((td, idx) => {
                    const title = (td.getAttribute('data-title') || '').trim();
                    const val = (td.getAttribute('data-value') || td.innerText || '').trim();
                    if (title) obj[title] = val;
                    obj['col_' + idx] = val;
                    const a = td.querySelector('a[href]');
                    if (a && !obj['_href_id']) obj['_href_id'] = a.getAttribute('href') || '';
                    if (a && (a.getAttribute('href')||'').includes('/metas/consulta/')) obj['_href_resultado'] = a.getAttribute('href') || '';
                });
                if (Object.keys(obj).length) out.push(obj);
            });
            return out;
        """) or []

        for obj in rows:
            id_meta = str(obj.get('ID') or obj.get('Id') or obj.get('id') or obj.get('col_0') or '').strip()
            id_meta = re.sub(r"\D", "", id_meta)
            if not id_meta:
                continue

            tipo = str(obj.get('Tipo') or obj.get('tipo') or obj.get('col_1') or '').strip()
            escopo = str(obj.get('Escopo') or obj.get('escopo') or obj.get('col_2') or '').strip()
            descricao = str(obj.get('Descrição') or obj.get('Descricao') or obj.get('descricao') or obj.get('col_3') or '').strip()
            data_inicial = str(obj.get('Data Inicial') or obj.get('data inicial') or obj.get('col_4') or '').strip()
            data_final = str(obj.get('Data Final') or obj.get('data final') or obj.get('col_5') or '').strip()

            href_resultado = str(obj.get('_href_resultado') or '').strip()
            if href_resultado.startswith('/'):
                href_resultado = URL + href_resultado
            if not href_resultado:
                href_resultado = f"{URL}/metas/consulta/{id_meta}"

            linhas.append({
                "id": id_meta,
                "tipo": tipo,
                "escopo": escopo,
                "descricao": descricao,
                "data_inicial": data_inicial,
                "data_final": data_final,
                "href_resultado": href_resultado,
            })

        if linhas:
            return linhas
    except Exception as e:
        print(f"⚠️ Leitura JS da tabela de metas falhou, tentando Selenium comum: {e}")

    # Fallback: leitura Selenium tradicional.
    for tr in driver.find_elements(By.XPATH, "//table//tbody/tr"):
        tds = tr.find_elements(By.TAG_NAME, "td")
        if len(tds) < 4:
            continue
        vals = []
        for td in tds:
            vals.append((td.get_attribute('data-value') or td.text or '').strip())
        id_meta = re.sub(r"\D", "", vals[0] if vals else "")
        if not id_meta:
            continue
        item = {
            "id": id_meta,
            "tipo": vals[1] if len(vals) > 1 else "",
            "escopo": vals[2] if len(vals) > 2 else "",
            "descricao": vals[3] if len(vals) > 3 else "",
            "data_inicial": vals[4] if len(vals) > 4 else "",
            "data_final": vals[5] if len(vals) > 5 else "",
            "href_resultado": f"{URL}/metas/consulta/{id_meta}",
        }
        linhas.append(item)
    return linhas


def linha_meta_match(item, spec):
    tipo = normalizar_texto_match(item.get("tipo"))
    escopo = normalizar_texto_match(item.get("escopo"))
    desc = normalizar_texto_match(item.get("descricao"))

    tipo_spec = normalizar_texto_match(spec.get("tipo", ""))
    escopo_spec = normalizar_texto_match(spec.get("escopo", ""))

    if tipo_spec and tipo_spec != tipo:
        return False
    if escopo_spec and escopo_spec != escopo:
        return False

    partes = [normalizar_texto_match(p) for p in spec.get("descricao_contem", [])]
    return all(parte in desc for parte in partes if parte)


def escolher_melhor_linha_meta(linhas, spec):
    """
    Escolhe a meta do mês pela linha exibida na tela.
    Não usa IDs fixos, porque o SGI cria novos IDs todo mês.
    """
    candidatos = [x for x in linhas if linha_meta_match(x, spec)]
    if not candidatos:
        return None

    def id_int(x):
        return int(re.sub(r"\D", "", str(x.get("id") or "0")) or "0")

    candidatos.sort(key=id_int, reverse=True)
    escolhido = dict(candidatos[0])
    id_meta = str(escolhido.get("id", "")).strip()
    if id_meta and not escolhido.get("href_resultado"):
        escolhido["href_resultado"] = f"{URL}/metas/consulta/{id_meta}"
    return escolhido

def limpar_df_tabela(df):
    df = df.copy()
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    return df


def encontrar_df_resultado_metas():
    # Lê direto da tabela HTML visível via Selenium, sem depender de lxml/pandas.read_html
    possiveis = [
        "//table[@id='tabela_metas']",
        "//table[contains(@class,'tabela-metas')]",
        "(//table[contains(@class,'table') and .//tbody/tr])[1]",
    ]
    tabela = None
    ultimo_erro = None

    for xp in possiveis:
        try:
            tabela = wait.until(EC.presence_of_element_located((By.XPATH, xp)))
            if tabela and tabela.find_elements(By.XPATH, ".//tbody/tr"):
                break
        except Exception as e:
            ultimo_erro = e
            tabela = None

    if tabela is None:
        raise Exception(f"Tabela de resultados das metas não encontrada na tela: {ultimo_erro}")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tabela)
    time.sleep(0.6)

    headers = []
    for th in tabela.find_elements(By.XPATH, ".//thead//th"):
        txt = th.text.strip()
        if txt:
            headers.append(re.sub(r"\s+", " ", txt))

    # Alguns layouts têm a 1ª coluna sem texto no header. Ajusta pelo data-title do tbody.
    linhas = []
    trs = tabela.find_elements(By.XPATH, ".//tbody/tr[td]")
    for tr in trs:
        tds = tr.find_elements(By.XPATH, "./td")
        vals = []
        for td in tds:
            txt = td.text.strip()
            if not txt:
                txt = td.get_attribute("data-value") or ""
            vals.append(re.sub(r"\s+", " ", txt).strip())
        if any(v for v in vals):
            linhas.append(vals)

        # Le tfoot - o SGI coloca a linha de Total aqui
    trs_tfoot = tabela.find_elements(By.XPATH, ".//tfoot/tr[td]")
    for tr in trs_tfoot:
        tds = tr.find_elements(By.XPATH, "./td")
        vals = []
        for td in tds:
            txt = td.text.strip()
            if not txt:
                txt = td.get_attribute("data-value") or ""
            vals.append(re.sub(r"\\s+", " ", txt).strip())
        if any(v for v in vals):
            linhas.append(vals)

    if not linhas:
        raise Exception("Tabela de metas encontrada, mas sem linhas no tbody")

    max_cols = max(len(x) for x in linhas)
    while len(headers) < max_cols:
        idx = len(headers)
        if idx == 0:
            # tenta descobrir pelo data-title da primeira célula
            try:
                data_title = trs[0].find_elements(By.XPATH, "./td")[0].get_attribute("data-title") or ""
                headers.append(data_title.strip() or "Nome")
            except Exception:
                headers.append("Nome")
        else:
            headers.append(f"Coluna_{idx+1}")

    headers = headers[:max_cols]
    linhas = [row + [""] * (max_cols - len(row)) for row in linhas]
    df = pd.DataFrame(linhas, columns=headers)

    # Normaliza nomes esperados das colunas
    ren = {}
    for i, c in enumerate(df.columns):
        uc = normalizar_texto_match(c)
        nome = c
        if uc in ("FILIAL", "VENDEDOR", "SUBGRUPO", "FILIAL/VENDEDOR", "VENDEDOR/SUBGRUPO"):
            nome = c.title()
        elif "META(R$) TOTAL" in uc:
            nome = "Meta(R$) Total"
        elif "REALIZADO(R$) TOTAL" in uc:
            nome = "Realizado(R$) Total"
        elif "ATINGIDO TOTAL" in uc:
            nome = "Atingido Total"
        elif "META(R$) PERIODO" in uc:
            nome = "Meta(R$) Período"
        elif "REALIZADO(R$) PERIODO" in uc:
            nome = "Realizado(R$) Período"
        elif "ATINGIDO PERIODO" in uc:
            nome = "Atingido Período"
        elif "PROJETADO(R$)" in uc:
            nome = "Projetado(R$)"
        ren[c] = nome
    df = df.rename(columns=ren)

    # Garante nomes únicos: ex. Vendedor, Vendedor_2
    cols = []
    usados = {}
    for c in df.columns:
        if c not in usados:
            usados[c] = 1
            cols.append(c)
        else:
            usados[c] += 1
            cols.append(f"{c}_{usados[c]}")
    df.columns = cols
    return limpar_df_tabela(df)


def valor_float_brasil(v):
    try:
        s = str(v).strip()
        if s in ("", "nan", "None"):
            return None
        # Se já é um número puro (ex: 0.1995 vindo da seção numérica do xlsx),
        # não faz replace de "." que quebraria o valor
        s_clean = s.replace("%", "").strip()
        # Detecta formato brasileiro (tem vírgula como decimal): "13.031,84"
        if "," in s_clean and s_clean.index(",") > s_clean.index(".") if "." in s_clean else False:
            # Formato BR: remove pontos de milhar, troca vírgula por ponto
            s_clean = s_clean.replace(".", "").replace(",", ".")
        elif "," in s_clean and "." not in s_clean:
            # Só vírgula: "13031,84" -> "13031.84"
            s_clean = s_clean.replace(",", ".")
        # else: formato puro com ponto: "13031.84" ou "0.1995" — usa direto
        return float(s_clean)
    except Exception:
        return None


def nome_arquivo_contas_valido(fname):
    s = str(fname or "").lower()
    if s.startswith("~$"):
        return False
    if not (s.endswith(".xls") or s.endswith(".xlsx")):
        return False
    # Exclui arquivos auxiliares gerados pelo próprio script
    bloqueados = [
        "metas_vendas_mes_atual",
        "dashboard",
        "historico",
        "fechamentos",
        "credenciais",
        "config_meta",
        "cobrancas_log",
    ]
    if any(b in s for b in bloqueados):
        return False
    # Dá preferência a nomes típicos do relatório de contas a receber
    preferidos = [
        "relatorio_contas",
        "contas_pagar_receber",
        "contas_receber",
        "administrativo",
    ]
    return any(p in s for p in preferidos)


def _is_total_row(reg):
    """Detecta se uma linha é a linha de Total (tfoot ou primeira coluna contém "Total")."""
    # Verifica colunas conhecidas de identificação
    for k in ["Filial", "Vendedor", "Subgrupo", "Nome", "col_0"]:
        v = str(reg.get(k, "") or "").strip()
        if re.search(r"\btotal\b", v, re.IGNORECASE):
            return True
    # Verifica a primeira coluna de dados (não meta)
    for k, v in reg.items():
        if k.startswith("_"):
            continue
        sv = str(v or "").strip()
        if re.search(r"\btotal\b", sv, re.IGNORECASE):
            return True
        break  # só verifica a primeira coluna de dados
    return False


def df_meta_para_registros(df, spec, origem):
    saida = []
    for _, row in df.iterrows():
        reg = {}
        for c in df.columns:
            v = row[c]
            if isinstance(v, pd.Series):
                v = " | ".join("" if pd.isna(x) else str(x).strip() for x in v.tolist())
            reg[str(c)] = "" if pd.isna(v) else str(v).strip()
        reg["_meta_chave"] = spec["chave"]
        reg["_meta_label"] = spec["label"]
        reg["_meta_tipo"] = origem.get("tipo")
        reg["_meta_escopo"] = origem.get("escopo")
        reg["_meta_descricao"] = origem.get("descricao")
        reg["_meta_id"] = origem.get("id")
        reg["_meta_href"] = origem.get("href_resultado")
        # Marca linha de total (tfoot)
        reg["_is_total"] = _is_total_row(reg)
        for k in list(reg.keys()):
            if "META(R$)" in k.upper() or "REALIZADO(R$)" in k.upper() or "PROJETADO(R$)" in k.upper():
                reg[k + "_float"] = valor_float_brasil(reg[k])
            if "ATINGIDO" in k.upper():
                reg[k + "_float"] = valor_float_brasil(reg[k])
        saida.append(reg)
    return saida




def _extrair_totais_do_xlsx(xlsx_path):
    """
    Extrai os totais das abas de metas a partir das células finais do XLSX.
    Regra pedida pelo usuário:
      - venda_filial_meta           -> linha 11
      - servico_filial_ouro_fob     -> linha 10
      - venda_filial_subgrupo_20k   -> linha 10

    Mapeamento usado (ordem real do SGI):
      B = Meta Total
      C = Realizado Total  ← valor oficial do card
      D = Atingido Total (%)
      E = Meta Período
      F = Realizado Período
      H = Projetado

    Se a célula fixa não existir, faz fallback para localizar a última linha "Total"
    ou somar as linhas de filial.
    """
    from openpyxl import load_workbook

    def _br_float(v):
        try:
            s = str(v).strip()
            if s in ("", "None", "nan"):
                return 0.0
            s = s.replace("R$", "").replace("%", "").strip()
            if "," in s:
                s = s.replace(".", "").replace(",", ".")
            return float(s)
        except Exception:
            return 0.0

    def _sheet_by_prefix(wb, nome_base):
        for s in wb.sheetnames:
            if s == nome_base or s.startswith(nome_base[:20]):
                return wb[s]
        return None

    def _fallback_from_sheet(ws):
        # 1) tenta achar a última linha "Total"
        max_row = ws.max_row
        for r in range(max_row, 1, -1):
            a = str(ws.cell(r, 1).value or "").strip().upper()
            if a == "TOTAL":
                return {
                    "meta_total":  round(_br_float(ws.cell(r, 2).value), 2),
                    "real_total":  round(_br_float(ws.cell(r, 3).value), 2),
                    "ating_total": round(_br_float(ws.cell(r, 4).value), 2),
                    "meta_per":    round(_br_float(ws.cell(r, 5).value), 2),
                    "real_per":    round(_br_float(ws.cell(r, 6).value), 2),
                    "proj":        round(_br_float(ws.cell(r, 8).value), 2),
                }

        # 2) fallback final: soma linhas FILIAL
        meta_total = real_total = meta_per = proj = 0.0
        for r in range(2, max_row + 1):
            a = str(ws.cell(r, 1).value or "").strip().upper()
            if a.startswith("FILIAL"):
                meta_total += _br_float(ws.cell(r, 2).value)
                real_total += _br_float(ws.cell(r, 3).value)
                meta_per   += _br_float(ws.cell(r, 5).value)
                proj       += _br_float(ws.cell(r, 8).value)
        ating = round((real_total / meta_total * 100.0), 2) if meta_total > 0 else 0.0
        return {
            "meta_total": round(meta_total, 2),
            "real_total": round(real_total, 2),
            "ating_total": ating,
            "meta_per": round(meta_per, 2),
            "real_per": 0.0,
            "proj": round(proj, 2),
        }

    _MAP = {
        "venda_filial_meta":       {"sheet": "venda_filial_meta", "row": 11},
        "servico_filial_ouro_fob": {"sheet": "servico_filial_ouro_fob", "row": 10},
        "venda_filial_subgrupo_20k": {"sheet": "venda_filial_subgrupo_20k", "row": 10},
    }
    out = {}

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"⚠️ Não consegui abrir xlsx de metas: {e}")
        return out

    for chave, cfg in _MAP.items():
        ws = _sheet_by_prefix(wb, cfg["sheet"])
        if ws is None:
            print(f"⚠️ Sheet não encontrada no xlsx: {cfg['sheet']}")
            out[chave] = {"meta_total": 0.0, "real_total": 0.0, "ating_total": 0.0, "meta_per": 0.0, "real_per": 0.0, "proj": 0.0}
            continue

        r = int(cfg["row"])
        a = str(ws.cell(r, 1).value or "").strip().upper()
        # Usa a linha fixa se houver conteúdo e parecer ser a linha final/total
        if a in ("TOTAL",) or ws.cell(r, 2).value not in (None, ""):
            meta_total = round(_br_float(ws.cell(r, 2).value), 2)
            real_total = round(_br_float(ws.cell(r, 3).value), 2)
            ating_total = round(_br_float(ws.cell(r, 4).value), 2)
            meta_per = round(_br_float(ws.cell(r, 5).value), 2)
            real_per = round(_br_float(ws.cell(r, 6).value), 2)
            proj = round(_br_float(ws.cell(r, 8).value), 2)
            item = {
                "meta_total": meta_total,
                "real_total": real_total,
                "ating_total": ating_total,
                "meta_per": meta_per,
                "real_per": real_per,
                "proj": proj,
            }
        else:
            item = _fallback_from_sheet(ws)

        out[chave] = item
        print(f"✅ V17 Total xlsx [{chave}]: meta={item['meta_total']:.2f}  real_total_col_C={item['real_total']:.2f}  real_periodo_col_F={item['real_per']:.2f}  ating={item['ating_total']:.2f}%  proj={item['proj']:.2f}")

    return out



def coletar_metas_vendas_mes_atual():
    primeiro_dia = hoje.replace(day=1).strftime("%d/%m/%Y")
    prox_mes = (hoje.replace(day=28) + timedelta(days=4)).replace(day=1)
    ultimo_dia = (prox_mes - timedelta(days=1)).strftime("%d/%m/%Y")

    print("\n📊 Iniciando coleta de metas/vendas do mês atual...")
    driver.get(URL + "/metas")
    wait.until(EC.presence_of_element_located((By.ID, "data_inicial_maior_igual")))
    set_input_value_safe(By.ID, "data_inicial_maior_igual", primeiro_dia)
    set_input_value_safe(By.ID, "data_final_menor_igual", ultimo_dia)
    clicar_primeiro_filtro_metas()
    time.sleep(2)

    linhas = extrair_linhas_tabela_metas()
    print(f"📋 Metas listadas na tela: {len(linhas)}")

    # IDs das metas mudam todo mês. Buscar sempre pelo Tipo + Escopo + Descrição da tela filtrada.
    specs = [
        {"chave": "venda_filial_meta", "label": "Venda / Filial / Meta Filial", "tipo": "Venda", "escopo": "Filial", "descricao_contem": ["META FILIAL"]},
        {"chave": "venda_filial_vendedor_meta", "label": "Venda / Filial-Vendedor / Meta Vendedor", "tipo": "Venda", "escopo": "Filial/Vendedor", "descricao_contem": ["META VENDEDOR"]},
        {"chave": "venda_filial_subgrupo_20k", "label": "Venda / Filial-Subgrupo / Caminhão 20K", "tipo": "Venda", "escopo": "Filial/Subgrupo", "descricao_contem": ["CAMINHAO", "20K"]},
        {"chave": "servico_filial_ouro_fob", "label": "Serviço / Filial / Ouro-FOB", "tipo": "Serviço", "escopo": "Filial", "descricao_contem": ["OURO", "FOB"]},
        {"chave": "servico_filial_vendedor_ouro_fob", "label": "Serviço / Filial-Vendedor / Ouro-FOB", "tipo": "Serviço", "escopo": "Filial/Vendedor", "descricao_contem": ["OURO", "FOB"]},
        {"chave": "venda_vendedor_subgrupo_20k", "label": "Venda / Vendedor-Subgrupo / Caminhão 20K", "tipo": "Venda", "escopo": "Vendedor/Subgrupo", "descricao_contem": ["CAMINHAO", "20K"]},
    ]

    resultados = {
        "coletado_em": now_brasilia().strftime("%Y-%m-%d %H:%M:%S"),
        "periodo_meta": {"data_inicial": primeiro_dia, "data_final": ultimo_dia},
        "lista_index": linhas,
        "metas": {}
    }

    xlsx_path = os.path.join(pasta, "metas_vendas_mes_atual.xlsx")
    json_path = os.path.join(pasta, "metas_vendas_mes_atual.json")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        pd.DataFrame(linhas).to_excel(writer, sheet_name="index_metas", index=False)

        for spec in specs:
            origem = escolher_melhor_linha_meta(linhas, spec)
            if not origem:
                print(f"⚠️ Meta não encontrada: {spec['label']}")
                resultados["metas"][spec["chave"]] = {"ok": False, "erro": "meta_nao_encontrada", "spec": spec}
                continue

            try:
                print(f"➡️ Abrindo resultados: {spec['label']} -> {origem['href_resultado']}")
                driver.get(origem["href_resultado"])
                time.sleep(2)
                df_result = encontrar_df_resultado_metas()
                df_result = limpar_df_tabela(df_result)
                print(f"   ↳ Colunas lidas: {list(df_result.columns)}")
                print(f"   ↳ Primeiras linhas: {len(df_result)}")
                resultados["metas"][spec["chave"]] = {
                    "ok": True,
                    "spec": spec,
                    "origem": origem,
                    "colunas": list(df_result.columns),
                    "linhas": df_meta_para_registros(df_result, spec, origem),
                }
                sheet = spec["chave"][:31]
                df_result.to_excel(writer, sheet_name=sheet, index=False)
                print(f"✅ {spec['label']}: {len(df_result)} linha(s)")
            except Exception as e:
                print(f"⚠️ Erro em {spec['label']}: {e}")
                resultados["metas"][spec["chave"]] = {
                    "ok": False,
                    "spec": spec,
                    "origem": origem,
                    "erro": str(e),
                }

    # ── Pós-coleta: relê o xlsx para extrair totais numéricos limpos ──────────────
    # O xlsx já está fechado aqui. Relemos e injetamos a linha Total em cada spec
    # que a precisa, garantindo que _sales_emp receba os valores corretos.
    _totais_xlsx = _extrair_totais_do_xlsx(xlsx_path)
    for _chave_t, _tot in _totais_xlsx.items():
        _meta_obj_t = resultados["metas"].get(_chave_t)
        if not isinstance(_meta_obj_t, dict) or not _meta_obj_t.get("ok"):
            continue
        _linhas_t = _meta_obj_t.get("linhas") or []
        # Remove qualquer linha Total anterior (evita duplicata)
        _linhas_t = [_r for _r in _linhas_t if not _r.get("_is_total")]
        # Constrói linha Total canônica com todos os campos _float prontos
        _total_reg = {
            "Filial": "Total",
            "Meta(R$) Total":           str(_tot["meta_total"]),
            "Realizado(R$) Total":       str(_tot["real_total"]),
            "Atingido Total":            f"{_tot['ating_total']:.2f}%",
            "Meta(R$) Período":          str(_tot["meta_per"]),
            "Realizado(R$) Período":     str(_tot["real_per"]),
            "Projetado(R$)":             str(_tot["proj"]),
            "Meta(R$) Total_float":      _tot["meta_total"],
            "Realizado(R$) Total_float": _tot["real_total"],
            "Atingido Total_float":      _tot["ating_total"],
            "Meta(R$) Período_float":    _tot["meta_per"],
            "Realizado(R$) Período_float": _tot["real_per"],
            "Projetado(R$)_float":       _tot["proj"],
            "_is_total": True,
            "_meta_chave": _chave_t,
        }
        _linhas_t.append(_total_reg)
        resultados["metas"][_chave_t]["linhas"] = _linhas_t
    # ── Fim do pós-coleta ──────────────────────────────────────────────────────────

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)

    print(f"💾 Metas JSON: {json_path}")
    print(f"💾 Metas XLSX: {xlsx_path}")
    return {"json_path": json_path, "xlsx_path": xlsx_path, "dados": resultados}


# =========================================
# 📈 RELATÓRIO DE MARGEM BRUTA / RENTABILIDADE
# Coleta 2 relatórios do SGI:
#   1) Lucratividade por FILIAL
#   2) Lucratividade por VENDEDOR
# Alimenta o dashboard com Margem Bruta (%) do mês atual.
# =========================================

def _select_by_text_contains(select_el, texto_alvo):
    texto_alvo_n = normalizar_texto_match(texto_alvo)
    sel = Select(select_el)
    for opt in sel.options:
        if texto_alvo_n in normalizar_texto_match(opt.text):
            sel.select_by_visible_text(opt.text)
            try:
                driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", select_el)
            except Exception:
                pass
            return True
    return False


def _label_contains_xpath(txt):
    return f"//label[contains(normalize-space(.),'{txt}')] | //strong[contains(normalize-space(.),'{txt}')] | //div[contains(normalize-space(.),'{txt}') and not(*)]"


def _find_relatorio_block_by_label(label_texto):
    ultimo = None
    xp = _label_contains_xpath(label_texto)
    candidatos = [
        f"({xp})[1]/ancestor::div[contains(@class,'row') or contains(@class,'col') or contains(@class,'form-group')][1]",
        f"({xp})[1]/parent::*",
        f"({xp})[1]/ancestor::div[1]",
    ]
    for cp in candidatos:
        try:
            bloco = wait.until(EC.presence_of_element_located((By.XPATH, cp)))
            if bloco:
                return bloco
        except Exception as e:
            ultimo = e
    raise ultimo or Exception(f"Bloco do label {label_texto} não encontrado")


def _set_relatorio_periodo_emissao(data_inicial, data_final):
    bloco = _find_relatorio_block_by_label('Período de Emissão')
    try:
        sels = [s for s in bloco.find_elements(By.TAG_NAME, 'select') if s.is_displayed()]
        if sels:
            try:
                Select(sels[0]).select_by_value('intervalo')
            except Exception:
                _select_by_text_contains(sels[0], 'Intervalo')
            driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", sels[0])
    except Exception:
        pass

    inputs = [i for i in bloco.find_elements(By.XPATH, ".//input[not(@type='hidden')]") if i.is_displayed()]
    if len(inputs) < 2:
        # fallback geral: pega os 2 primeiros inputs visíveis de data/texto do formulário
        form = wait.until(EC.presence_of_element_located((By.XPATH, "//form[contains(@action,'relatorio_margens_brutas')]")))
        inputs = [i for i in form.find_elements(By.XPATH, ".//input[not(@type='hidden')]") if i.is_displayed()]
    if len(inputs) < 2:
        raise Exception('Não encontrei os 2 campos de data do relatório de margens')

    for el, valor in [(inputs[0], data_inicial), (inputs[1], data_final)]:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
        try:
            el.click()
            el.send_keys(Keys.CONTROL, 'a')
            el.send_keys(Keys.DELETE)
            el.send_keys(valor)
        except Exception:
            driver.execute_script("arguments[0].value = arguments[1];", el, valor)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", el)
        driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", el)


def _select_relatorio_por_label(label_texto, texto_opcao=None, value=None):
    bloco = _find_relatorio_block_by_label(label_texto)
    selects = [s for s in bloco.find_elements(By.TAG_NAME, 'select') if s.is_displayed()]
    if not selects:
        # fallback: primeiro select visível após o label no DOM
        xp = f"({_label_contains_xpath(label_texto)})[1]/following::select[1]"
        selects = [wait.until(EC.presence_of_element_located((By.XPATH, xp)))]
    el = selects[0]
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    sel = Select(el)
    if value is not None:
        try:
            sel.select_by_value(str(value))
        except Exception:
            pass
    if texto_opcao is not None:
        if not _select_by_text_contains(el, texto_opcao):
            try:
                sel.select_by_visible_text(texto_opcao)
            except Exception:
                # último fallback: tentativa por option text normalizado
                achou = False
                alvo = normalizar_texto_match(texto_opcao)
                for opt in sel.options:
                    if alvo == normalizar_texto_match(opt.text):
                        sel.select_by_visible_text(opt.text)
                        achou = True
                        break
                if not achou:
                    raise Exception(f'Opção {texto_opcao} não encontrada em {label_texto}')
    driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", el)
    return el



def _tratar_valor_margem(v):
    """Converte valores BR e valores numéricos puros sem quebrar markup 2.13 -> 213."""
    try:
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s in ("", "nan", "None"):
            return 0.0
        s = s.replace("R$", "").replace("%", "").strip()
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


def _limpar_nome_margem(nome):
    s = str(nome or "").strip()
    s = re.sub(r"^Vendedor:\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"^\d+\s*-\s*", "", s)
    s = re.sub(r"^C\.\d+\s*-\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*\(GER\s*F\d+\)\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*\(F\d+\)\s*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _norm_key_margem(nome):
    s = _limpar_nome_margem(nome).upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _filial_key_from_text(texto):
    s = str(texto or "").upper()
    m = re.search(r"FILIAL\s*0*(\d{1,3})", s)
    if not m:
        return ""
    n = int(m.group(1))
    return "FDEP" if n in (90, 99) else f"F{n}"


def _select_all_multiselect_by_label(label_texto):
    abrir_multiselect_por_label(label_texto, timeout=20)
    xpath_ul = f"//label[contains(normalize-space(.),'{label_texto}')]/following::ul[1]"
    container = wait.until(EC.presence_of_element_located((By.XPATH, xpath_ul)))
    driver.execute_script("arguments[0].scrollIntoView({block:'center', inline:'nearest'});", container)
    time.sleep(0.5)
    checkboxes = container.find_elements(By.XPATH, ".//input[@type='checkbox']")
    for c in checkboxes:
        try:
            if not c.is_selected():
                driver.execute_script("arguments[0].click();", c)
                time.sleep(0.05)
        except Exception:
            pass
    driver.execute_script("document.body.click();")
    time.sleep(0.5)
    return len(checkboxes)


def _nome_arquivo_margem_valido(fname):
    s = str(fname or '').lower()
    if s.startswith('~$'):
        return False
    if not (s.endswith('.xls') or s.endswith('.xlsx')):
        return False
    return ('margem' in s) or ('margens' in s)


def _ler_xls_ou_xlsx(caminho):
    erros = []
    for engine in ('openpyxl', None):
        try:
            if engine:
                return pd.read_excel(caminho, header=None, engine=engine)
            return pd.read_excel(caminho, header=None)
        except Exception as e:
            erros.append(f'read_excel[{engine}]: {e}')
    try:
        tabelas = pd.read_html(caminho, decimal=',', thousands='.')
        if tabelas:
            return tabelas[0]
    except Exception as e:
        erros.append(f'read_html: {e}')
    raise Exception('Não consegui ler XLS/XLSX de margem: ' + ' | '.join(erros))


def _parse_relatorio_margem_xls(caminho, tipo):
    df_m = _ler_xls_ou_xlsx(caminho).fillna('')
    filiais = {}
    vendedores = {}
    empresa = {}

    for _, row in df_m.iterrows():
        vals = [str(x).strip() for x in row.tolist()]
        vals += [''] * max(0, 8 - len(vals))

        if tipo == 'filial':
            filial_txt = vals[0]
            norm0 = normalizar_texto_match(filial_txt)
            if not filial_txt or norm0 == 'FILIAL':
                continue
            if norm0 == 'TOTAL':
                empresa = {
                    'label': 'TOTAL',
                    'margem_bruta_pct': round(_tratar_valor_margem(vals[5]), 2),
                    'valor_total_liquido': _tratar_valor_margem(vals[1]),
                    'valor_total': _tratar_valor_margem(vals[2]),
                    'custo_total': _tratar_valor_margem(vals[3]),
                    'margem_bruta_valor': _tratar_valor_margem(vals[4]),
                    'markup_realizado': _tratar_valor_margem(vals[6]) if len(vals) > 6 else 0.0,
                    'pmr_vendas_prazo': _tratar_valor_margem(vals[7]) if len(vals) > 7 else 0.0,
                }
                continue
            if 'FILIAL' not in norm0:
                continue
            fk = _filial_key_from_text(filial_txt)
            if not fk:
                continue
            margem_pct = _tratar_valor_margem(vals[5])
            filiais[fk] = {
                'filial': fk,
                'label': filial_txt,
                'margem_bruta_pct': round(margem_pct, 2),
                'valor_total_liquido': _tratar_valor_margem(vals[1]),
                'valor_total': _tratar_valor_margem(vals[2]),
                'custo_total': _tratar_valor_margem(vals[3]),
                'margem_bruta_valor': _tratar_valor_margem(vals[4]),
                'markup_realizado': _tratar_valor_margem(vals[6]) if len(vals) > 6 else 0.0,
            }

        else:
            filial_txt = vals[0]
            vendedor_txt = vals[1]
            if not vendedor_txt or normalizar_texto_match(vendedor_txt) in ('VENDEDOR', 'TOTAL'):
                continue
            fk = _filial_key_from_text(filial_txt)
            if not fk:
                continue
            nome = _limpar_nome_margem(vendedor_txt)
            if not nome:
                continue
            margem_pct = _tratar_valor_margem(vals[6])
            key = f'{_norm_key_margem(nome)}_{fk}'
            vendedores[key] = {
                'nome': nome,
                'filial': fk,
                'key': key,
                'label_filial': filial_txt,
                'margem_bruta_pct': round(margem_pct, 2),
                'valor_total_liquido': _tratar_valor_margem(vals[2]),
                'valor_total': _tratar_valor_margem(vals[3]),
                'custo_total': _tratar_valor_margem(vals[4]),
                'margem_bruta_valor': _tratar_valor_margem(vals[5]),
                'markup_realizado': _tratar_valor_margem(vals[7]) if len(vals) > 7 else 0.0,
            }

    return {'filiais': filiais, 'vendedores': vendedores, 'empresa': empresa}


def _gerar_relatorio_margem(tipo):
    primeiro_dia = hoje.replace(day=1).strftime("%d/%m/%Y")
    prox_mes = (hoje.replace(day=28) + timedelta(days=4)).replace(day=1)
    ultimo_dia = (prox_mes - timedelta(days=1)).strftime("%d/%m/%Y")

    driver.get(URL + "/relatorio_margens_brutas")
    wait.until(EC.presence_of_element_located((By.ID, "gerar")))
    time.sleep(1.5)
    esperar_sumir_overlays(10)

    _set_relatorio_periodo_emissao(primeiro_dia, ultimo_dia)

    try:
        _select_relatorio_por_label("Tipo de Custo", texto_opcao="Custo Médio Gerencial")
    except Exception as e:
        print(f"⚠️ Não consegui selecionar Tipo de Custo automaticamente: {e}")

    try:
        _select_relatorio_por_label("Lucratividade por", texto_opcao=("Filial" if tipo == "filial" else "Vendedor"))
    except Exception as e:
        print(f"⚠️ Não consegui selecionar Lucratividade por {tipo}: {e}")

    try:
        total_ops = _select_all_multiselect_by_label("Operações de Devolução")
        print(f"✅ Operações de Devolução: {total_ops} opções marcadas")
    except Exception as e:
        print(f"⚠️ Não consegui marcar Operações de Devolução: {e}")

    try:
        Select(wait.until(EC.presence_of_element_located((By.ID, "_formato")))).select_by_value("xls")
        print("✅ Formato XLS para Margem Bruta")
    except Exception as e:
        print(f"⚠️ Não consegui selecionar formato XLS de Margem Bruta: {e}")

    arquivos_antes = set(f for f in os.listdir(download_dir) if _nome_arquivo_margem_valido(f))

    gerar_btn = wait.until(EC.presence_of_element_located((By.ID, 'gerar')))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", gerar_btn)
    esperar_sumir_overlays(10)
    try:
        wait.until(EC.element_to_be_clickable((By.ID, 'gerar'))).click()
    except Exception:
        driver.execute_script("arguments[0].click();", gerar_btn)
    print(f"📈 Gerando Margem Bruta por {tipo.upper()}...")

    caminho = None
    for _ in range(60):
        time.sleep(2)
        baixando = any(f.endswith('.crdownload') or f.endswith('.tmp') for f in os.listdir(download_dir))
        if baixando:
            continue
        novos = set(f for f in os.listdir(download_dir) if _nome_arquivo_margem_valido(f)) - arquivos_antes
        if novos:
            caminho = max([os.path.join(download_dir, f) for f in novos], key=os.path.getctime)
            break

    if not caminho:
        todos = [os.path.join(download_dir, f) for f in os.listdir(download_dir) if _nome_arquivo_margem_valido(f)]
        if not todos:
            raise Exception('Nenhum arquivo XLS/XLSX de margem foi baixado')
        caminho = max(todos, key=os.path.getctime)
        print(f"⚠️ Usando último arquivo de margem encontrado: {caminho}")
    else:
        print(f"✅ Download margem OK: {caminho}")

    return _parse_relatorio_margem_xls(caminho, tipo)


def coletar_margens_brutas_mes_atual():
    print("\n📈 Iniciando coleta de Margem Bruta/Rentabilidade do mês atual...")
    json_path = os.path.join(pasta, "margens_brutas_mes_atual.json")
    xlsx_path = os.path.join(pasta, "margens_brutas_mes_atual.xlsx")
    resultados = {
        "coletado_em": now_brasilia().strftime("%Y-%m-%d %H:%M:%S"),
        "periodo": {
            "data_inicial": hoje.replace(day=1).strftime("%d/%m/%Y"),
            "data_final": ((hoje.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)).strftime("%d/%m/%Y"),
        },
        "filiais": {},
        "vendedores": {},
        "empresa": {},
        "ok": False,
    }

    try:
        res_fil = _gerar_relatorio_margem("filial")
        resultados["filiais"] = res_fil.get("filiais", {})
        resultados["empresa"] = res_fil.get("empresa", {})
        print(f"✅ Margem por FILIAL: {len(resultados['filiais'])} registro(s)")
    except Exception as e:
        resultados["erro_filial"] = str(e)
        print(f"⚠️ Erro ao coletar margem por FILIAL: {e}")

    try:
        res_vend = _gerar_relatorio_margem("vendedor")
        resultados["vendedores"] = res_vend.get("vendedores", {})
        print(f"✅ Margem por VENDEDOR: {len(resultados['vendedores'])} registro(s)")
    except Exception as e:
        resultados["erro_vendedor"] = str(e)
        print(f"⚠️ Erro ao coletar margem por VENDEDOR: {e}")

    resultados["ok"] = bool(resultados.get("filiais") or resultados.get("vendedores"))

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)

    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            pd.DataFrame(list(resultados.get("filiais", {}).values())).to_excel(writer, sheet_name="filiais", index=False)
            pd.DataFrame([resultados.get("empresa", {})]).to_excel(writer, sheet_name="empresa", index=False)
            pd.DataFrame(list(resultados.get("vendedores", {}).values())).to_excel(writer, sheet_name="vendedores", index=False)
    except Exception as e:
        print(f"⚠️ Não consegui salvar XLSX de margens: {e}")
        xlsx_path = None

    print(f"💾 Margens JSON: {json_path}")
    if xlsx_path:
        print(f"💾 Margens XLSX: {xlsx_path}")
    return {"json_path": json_path, "xlsx_path": xlsx_path, "dados": resultados}





def _br_money(v):
    try:
        s = str(v or '').strip()
        if s in ('', 'None', 'nan'):
            return 0.0
        s = s.replace('R$', '').replace('%', '').strip()
        s = s.replace('.', '').replace(',', '.')
        return float(s)
    except Exception:
        return 0.0


def _first_day_last_day_current_month():
    primeiro_dia = hoje.replace(day=1)
    prox_mes = (hoje.replace(day=28) + timedelta(days=4)).replace(day=1)
    ultimo_dia = prox_mes - timedelta(days=1)
    return primeiro_dia.strftime('%d/%m/%Y'), ultimo_dia.strftime('%d/%m/%Y')


def _wait_container_relatorio_ready(timeout=60):
    fim = time.time() + timeout
    while time.time() < fim:
        try:
            html = driver.execute_script("""
                const el = document.getElementById('container_relatorio');
                return el ? (el.innerHTML || '') : '';
            """) or ''
            if 'label-agrupamento' in html and ('DADOS DO SERVIÇO/RECUPERAÇÃO' in html or 'TOTAL' in html):
                return html
        except Exception:
            pass
        time.sleep(1)
    raise Exception('container_relatorio não carregou com os dados do relatório de serviços')


def _parse_relatorio_servicos_html(html):
    soup = BeautifulSoup(html, 'html.parser')
    container = soup.find(id='container_relatorio') or soup

    detalhes = []
    servicos = {}
    filiais = {}
    vendedores = {}
    empresa_total = 0.0
    empresa_qtd = 0.0
    servico_atual = ''
    totais_oficiais_servico = {}

    def _txt(el):
        return ' '.join(el.get_text(' ', strip=True).split())

    def _is_money_text(txt):
        txt = str(txt or '').strip()
        if not txt or txt.upper() in {'X', '-', '*'}:
            return False
        if re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', txt):
            return False
        return bool(re.fullmatch(r'-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2}', txt))

    def _money_candidates(vals):
        out = []
        for raw in vals:
            txt = str(raw or '').strip()
            if _is_money_text(txt):
                out.append(_br_money(txt))
        return out

    def _first_number(vals):
        for raw in vals:
            txt = str(raw or '').strip()
            if re.fullmatch(r'\d+(?:,\d+)?', txt):
                return _br_money(txt)
        return 0.0

    def _is_total_row(vals):
        if not vals:
            return False
        first = unicodedata.normalize('NFD', str(vals[0] or '')).encode('ascii', 'ignore').decode('ascii').upper().strip()
        return first == 'TOTAL'

    def _is_total_cancelado(vals):
        raw = ' '.join(str(v or '') for v in vals).upper()
        raw = unicodedata.normalize('NFD', raw).encode('ascii', 'ignore').decode('ascii')
        return 'CANCELAD' in raw

    def _aplicar_total_oficial(servico, vals):
        if not servico or _is_total_cancelado(vals):
            return

        # TOTAL correto do tipo = último valor monetário da linha TOTAL do bloco SERVIÇO.
        # Exemplo do SGI: TOTAL | 327 | ... | Vl. Unitário 209,79 | Vl. Acréscimo 812,52 | Vl. Total 4.079,25
        nums = _money_candidates(vals)
        if not nums:
            return

        qtd = _first_number(vals[1:3]) if len(vals) > 1 else 0.0
        vl_total = nums[-1]
        vl_acrescimo = nums[-2] if len(nums) >= 2 else 0.0
        vl_unit_total = nums[-3] if len(nums) >= 3 else 0.0

        totais_oficiais_servico[servico] = {
            'servico': servico,
            'quantidade': round(float(qtd or 0), 2),
            'vl_unitario_total': round(float(vl_unit_total or 0), 2),
            'vl_acrescimo_total': round(float(vl_acrescimo or 0), 2),
            'real_total': round(float(vl_total or 0), 2),
            'linhas': 0,
            'fonte_total': 'linha_TOTAL_SGI_VL_TOTAL_com_acrescimo',
        }
        print(f"✅ Total oficial serviço [{servico}]: qtd={qtd} unit_total={vl_unit_total} acresc_total={vl_acrescimo} vl_total={vl_total}")

    def _servico_area_vals(vals):
        # A área SERVIÇO fica antes do bloco DADOS DA VENDA.
        # Mantém até a coluna Número dos Bilhetes; evita pegar preço de produto da venda.
        return vals[:13] if len(vals) >= 13 else vals

    def _pegar_total_linha(vals):
        servico_vals = _servico_area_vals(vals)
        money = _money_candidates(servico_vals[6:])
        # Último monetário da área serviço = Vl. Total já com acréscimo.
        return round(float(money[-1]), 2) if money else 0.0

    def _pegar_vl_unit_acrescimo(vals):
        servico_vals = _servico_area_vals(vals)
        money = _money_candidates(servico_vals[6:])
        # Normalmente: Vl.Unitário, Vl.Acréscimo, Vl.Total
        vl_unit = money[-3] if len(money) >= 3 else (money[-2] if len(money) >= 2 else 0.0)
        vl_acresc = money[-2] if len(money) >= 2 else 0.0
        return round(float(vl_unit), 2), round(float(vl_acresc), 2)

    for tr in container.find_all('tr'):
        th = tr.find('th', class_=lambda c: c and 'label-agrupamento' in c)
        if th:
            servico_atual = _txt(th)
            servicos.setdefault(servico_atual, {
                'servico': servico_atual,
                'quantidade': 0.0,
                'real_total': 0.0,
                'linhas': 0,
            })
            continue

        if tr.find_all('th'):
            continue

        tds = tr.find_all('td')
        if len(tds) < 2:
            continue

        vals = [_txt(td) for td in tds]
        if not vals or not servico_atual:
            continue

        if _is_total_row(vals):
            _aplicar_total_oficial(servico_atual, vals)
            continue

        if len(vals) < 6 or _is_total_cancelado(vals):
            continue

        filial_txt = vals[0] if len(vals) > 0 else ''
        filial_key = _filial_key_from_text(filial_txt)
        num_lanc = vals[1] if len(vals) > 1 else ''
        data_lanc = vals[2] if len(vals) > 2 else ''
        cliente = vals[3] if len(vals) > 3 else ''
        periodo_validade = vals[4] if len(vals) > 4 else ''
        vendedor_txt = vals[5] if len(vals) > 5 else ''
        quantidade = _br_money(vals[6]) if len(vals) > 6 else 0.0

        vl_total = _pegar_total_linha(vals)
        vl_unit, vl_acrescimo = _pegar_vl_unit_acrescimo(vals)
        numeros_bilhetes = vals[12] if len(vals) > 12 else ''

        vendedor_nome = _limpar_nome_margem(vendedor_txt)
        vendedor_key = f'{vendedor_nome}_{filial_key}' if vendedor_nome and filial_key else vendedor_nome

        row = {
            'servico': servico_atual,
            'filial': filial_key,
            'filial_label': filial_txt,
            'num_lanc': num_lanc,
            'data_lanc': data_lanc,
            'cliente': cliente,
            'periodo_validade': periodo_validade,
            'vendedor': vendedor_nome,
            'vendedor_label': vendedor_txt,
            'quantidade': quantidade,
            'vl_unitario': round(vl_unit, 2),
            'vl_acrescimo': round(vl_acrescimo, 2),
            'vl_total': round(vl_total, 2),
            'numeros_bilhetes': numeros_bilhetes,
            'vendedor_key': vendedor_key,
        }
        detalhes.append(row)

        empresa_total += vl_total
        empresa_qtd += quantidade

        srv = servicos.setdefault(servico_atual, {
            'servico': servico_atual,
            'quantidade': 0.0,
            'real_total': 0.0,
            'linhas': 0,
        })
        srv['quantidade'] += quantidade
        srv['real_total'] += vl_total
        srv['linhas'] += 1

        if filial_key:
            fil = filiais.setdefault(filial_key, {
                'filial': filial_key,
                'label': filial_txt,
                'quantidade': 0.0,
                'real_total': 0.0,
                'linhas': 0,
                'servicos': {},
            })
            fil['quantidade'] += quantidade
            fil['real_total'] += vl_total
            fil['linhas'] += 1
            fil['servicos'][servico_atual] = round(float(fil['servicos'].get(servico_atual, 0.0)) + vl_total, 2)

        if vendedor_key:
            vend = vendedores.setdefault(vendedor_key, {
                'key': vendedor_key,
                'nome': vendedor_nome,
                'filial': filial_key,
                'label': vendedor_txt,
                'quantidade': 0.0,
                'real_total': 0.0,
                'linhas': 0,
                'servicos': {},
            })
            vend['quantidade'] += quantidade
            vend['real_total'] += vl_total
            vend['linhas'] += 1
            vend['servicos'][servico_atual] = round(float(vend['servicos'].get(servico_atual, 0.0)) + vl_total, 2)

    # Cards por tipo = linha TOTAL do SGI.
    # Filial/vendedor = soma da coluna Vl. Total de cada linha.
    if totais_oficiais_servico:
        for servico, total_item in totais_oficiais_servico.items():
            antigo = servicos.get(servico, {})
            servicos[servico] = {
                **antigo,
                **total_item,
                'linhas': int(antigo.get('linhas', 0) or 0),
                'real_total_linhas': round(float(antigo.get('real_total', 0.0) or 0.0), 2),
                'quantidade_linhas': round(float(antigo.get('quantidade', 0.0) or 0.0), 2),
            }
        empresa_total = sum(float(v.get('real_total', 0.0) or 0.0) for v in servicos.values())
        empresa_qtd = sum(float(v.get('quantidade', 0.0) or 0.0) for v in servicos.values())

    for bucket in (servicos, filiais, vendedores):
        for _, item in bucket.items():
            if 'quantidade' in item:
                item['quantidade'] = round(float(item['quantidade']), 2)
            if 'real_total' in item:
                item['real_total'] = round(float(item['real_total']), 2)

    return {
        'empresa': {
            'quantidade': round(float(empresa_qtd), 2),
            'real_total': round(float(empresa_total), 2),
            'linhas': len(detalhes),
        },
        'servicos': servicos,
        'filiais': filiais,
        'vendedores': vendedores,
        'detalhes': detalhes,
    }


def coletar_relatorio_servicos_mes_atual():
    primeiro_dia, ultimo_dia = _first_day_last_day_current_month()
    print('\n🧰 Iniciando coleta do relatório de serviços do mês atual...')

    driver.get(URL + '/relatorio_servicos')
    wait.until(EC.presence_of_element_located((By.ID, 'data_emissao_inicial')))
    set_input_value_safe(By.ID, 'data_emissao_inicial', primeiro_dia)
    set_input_value_safe(By.ID, 'data_emissao_final', ultimo_dia)

    try:
        Select(wait.until(EC.presence_of_element_located((By.ID, '_formato')))).select_by_value('html')
        print('✅ Formato HTML para relatório de serviços')
    except Exception as e:
        print(f'⚠️ Não consegui selecionar formato HTML do relatório de serviços: {e}')

    try:
        clicar_seguro_xpath("//button[@id='gerar'] | //input[@id='gerar'] | //button[contains(.,'Gerar')] | //input[@value='Gerar']", timeout=20)
    except Exception:
        try:
            driver.find_element(By.ID, 'gerar').click()
        except Exception as e:
            raise Exception(f'Não consegui clicar em Gerar do relatório de serviços: {e}')

    html_rel = _wait_container_relatorio_ready(timeout=90)
    parsed = _parse_relatorio_servicos_html(driver.page_source)
    parsed['coletado_em'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    parsed['periodo'] = {'data_inicial': primeiro_dia, 'data_final': ultimo_dia}

    json_path = os.path.join(pasta, 'relatorio_servicos_mes_atual.json')
    xlsx_path = os.path.join(pasta, 'relatorio_servicos_mes_atual.xlsx')

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(parsed, f, ensure_ascii=False, indent=2)

    df_det = pd.DataFrame(parsed['detalhes'])
    df_srv = pd.DataFrame(list(parsed['servicos'].values()))
    df_fil = pd.DataFrame(list(parsed['filiais'].values()))
    df_vend = pd.DataFrame(list(parsed['vendedores'].values()))

    serv_fil_rows = []
    for fkey, fitem in parsed['filiais'].items():
        for srv, total in (fitem.get('servicos') or {}).items():
            serv_fil_rows.append({'filial': fkey, 'servico': srv, 'real_total': round(float(total), 2)})
    serv_vend_rows = []
    for vkey, vitem in parsed['vendedores'].items():
        for srv, total in (vitem.get('servicos') or {}).items():
            serv_vend_rows.append({'vendedor_key': vkey, 'nome': vitem.get('nome',''), 'filial': vitem.get('filial',''), 'servico': srv, 'real_total': round(float(total), 2)})

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        pd.DataFrame([parsed['empresa']]).to_excel(writer, sheet_name='empresa', index=False)
        df_srv.to_excel(writer, sheet_name='resumo_servico', index=False)
        df_fil.to_excel(writer, sheet_name='resumo_filial', index=False)
        df_vend.to_excel(writer, sheet_name='resumo_vendedor', index=False)
        pd.DataFrame(serv_fil_rows).to_excel(writer, sheet_name='servico_filial', index=False)
        pd.DataFrame(serv_vend_rows).to_excel(writer, sheet_name='servico_vendedor', index=False)
        df_det.to_excel(writer, sheet_name='detalhado', index=False)

    print(f"✅ Relatório de serviços coletado: {len(parsed['detalhes'])} linha(s)")
    print(f"💾 Serviços JSON: {json_path}")
    print(f"💾 Serviços XLSX: {xlsx_path}")
    return {'json_path': json_path, 'xlsx_path': xlsx_path, 'dados': parsed}


def _num_from_row(row, *keys):
    for k in keys:
        if k in row and row.get(k) not in (None, ''):
            return float(row.get(k) or 0)
    return 0.0


def _set_money_fields(row, valor, use_vendor=False, filial_label=''):
    meta_total = _num_from_row(row, 'Meta(R$) Total_float', 'Meta (R$) Total_float')
    meta_per = _num_from_row(row, 'Meta(R$) Período_float', 'Meta (R$) Período_float')
    proj = _num_from_row(row, 'Projetado(R$)_float', 'Projetado (R$)_float')
    ating_total = round((valor / meta_total * 100.0), 2) if meta_total > 0 else 0.0
    ating_per = round((valor / meta_per * 100.0), 2) if meta_per > 0 else 0.0

    row['Realizado(R$) Total_float'] = round(valor, 2)
    row['Realizado(R$) Total'] = f"{valor:.2f}"
    row['Realizado(R$) Período_float'] = round(valor, 2)
    row['Realizado(R$) Período'] = f"{valor:.2f}"
    row['Atingido Total_float'] = ating_total
    row['Atingido Total'] = f"{ating_total:.2f}%"
    row['Atingido Período_float'] = ating_per
    row['Atingido Período'] = f"{ating_per:.2f}%"
    if 'Projetado(R$)_float' not in row:
        row['Projetado(R$)_float'] = proj
        row['Projetado(R$)'] = f"{proj:.2f}"
    if 'Meta(R$) Total_float' not in row:
        row['Meta(R$) Total_float'] = meta_total
        row['Meta(R$) Total'] = f"{meta_total:.2f}"
    if 'Meta(R$) Período_float' not in row:
        row['Meta(R$) Período_float'] = meta_per
        row['Meta(R$) Período'] = f"{meta_per:.2f}"
    if filial_label and 'Filial' not in row:
        row['Filial'] = filial_label
    if use_vendor:
        row.setdefault('Vendedor', filial_label)
        row.setdefault('Vendedor_2', '')
    return row


def aplicar_relatorio_servicos_em_metas(metas_info, servicos_info):
    try:
        dados_metas = metas_info.get('dados', {}) if isinstance(metas_info, dict) else {}
        metas = dados_metas.get('metas', {}) if isinstance(dados_metas, dict) else {}
        dados_srv = servicos_info.get('dados', {}) if isinstance(servicos_info, dict) else {}
        filiais_srv = dados_srv.get('filiais', {}) or {}
        vendedores_srv = dados_srv.get('vendedores', {}) or {}
        empresa_srv = dados_srv.get('empresa', {}) or {}

        # Filiais
        chave_fil = 'servico_filial_ouro_fob'
        obj_fil = metas.get(chave_fil)
        if isinstance(obj_fil, dict) and obj_fil.get('ok'):
            linhas = obj_fil.get('linhas') or []
            novas = []
            vistos = set()
            total_model = next((dict(r) for r in linhas if r.get('_is_total')), {'_is_total': True, '_meta_chave': chave_fil})
            for row in linhas:
                if row.get('_is_total'):
                    continue
                nr = dict(row)
                fk = _filial_key_from_text(nr.get('Filial', ''))
                vistos.add(fk)
                valor = round(float((filiais_srv.get(fk, {}) or {}).get('real_total', 0.0) or 0.0), 2)
                _set_money_fields(nr, valor, filial_label=nr.get('Filial', ''))
                novas.append(nr)
            for fk, item in filiais_srv.items():
                if fk in vistos:
                    continue
                nr = {
                    'Filial': f'FILIAL {fk.replace("F", "").zfill(2)}',
                    '_meta_chave': chave_fil,
                }
                _set_money_fields(nr, round(float(item.get('real_total', 0.0) or 0.0), 2), filial_label=nr['Filial'])
                novas.append(nr)
            total_val = round(float(empresa_srv.get('real_total', 0.0) or 0.0), 2)
            total_row = dict(total_model)
            total_row['_is_total'] = True
            total_row['_meta_chave'] = chave_fil
            total_row['Filial'] = 'Total'
            _set_money_fields(total_row, total_val, filial_label='Total')
            obj_fil['linhas'] = novas + [total_row]

        # Vendedores
        chave_vend = 'servico_filial_vendedor_ouro_fob'
        obj_v = metas.get(chave_vend)
        if isinstance(obj_v, dict) and obj_v.get('ok'):
            linhas = obj_v.get('linhas') or []
            novas = []
            vistos = set()
            total_model = next((dict(r) for r in linhas if r.get('_is_total')), {'_is_total': True, '_meta_chave': chave_vend})
            for row in linhas:
                if row.get('_is_total'):
                    continue
                nr = dict(row)
                nome = str(nr.get('Vendedor_2') or nr.get('Vendedor') or '').strip()
                filial_txt = str(nr.get('Vendedor') or nr.get('Filial') or '').strip()
                fk = _filial_key_from_text(filial_txt)
                key = f'{_limpar_nome_margem(nome)}_{fk}' if nome and fk else ''
                if key:
                    vistos.add(key)
                valor = round(float((vendedores_srv.get(key, {}) or {}).get('real_total', 0.0) or 0.0), 2)
                _set_money_fields(nr, valor, use_vendor=True, filial_label=filial_txt)
                novas.append(nr)
            for key, item in vendedores_srv.items():
                if key in vistos:
                    continue
                fk = item.get('filial', '')
                filial_label = f'FILIAL {str(fk).replace("F", "").zfill(2)}' if fk else ''
                nome = item.get('nome', '')
                nr = {
                    'Vendedor': filial_label,
                    'Vendedor_2': nome,
                    '_meta_chave': chave_vend,
                }
                _set_money_fields(nr, round(float(item.get('real_total', 0.0) or 0.0), 2), use_vendor=True, filial_label=filial_label)
                novas.append(nr)
            total_val = round(float(empresa_srv.get('real_total', 0.0) or 0.0), 2)
            total_row = dict(total_model)
            total_row['_is_total'] = True
            total_row['_meta_chave'] = chave_vend
            total_row['Vendedor'] = 'Total'
            total_row['Vendedor_2'] = 'Total'
            _set_money_fields(total_row, total_val, use_vendor=True, filial_label='Total')
            obj_v['linhas'] = novas + [total_row]

        dados_metas['servicos_relatorio'] = dados_srv

        # regrava JSON consolidado de metas para o dashboard ler os serviços reais do relatório
        if metas_info.get('json_path'):
            with open(metas_info['json_path'], 'w', encoding='utf-8') as f:
                json.dump(dados_metas, f, ensure_ascii=False, indent=2)
        print('✅ Relatório de serviços aplicado sobre os blocos de serviços em metas_vendas_mes_atual.json')
    except Exception as e:
        print(f'⚠️ Erro ao aplicar relatório de serviços sobre metas: {e}')
    return metas_info



# =========================================
# 🕒 VENDA DIÁRIA OFICIAL — /relatorio_analises_totais_vendas
# Soma Total Vendido + Valor Serviço + Valor Acréscimo Serviço.
# Mantém valores negativos quando houver devolução/estorno.
# =========================================
def _nome_arquivo_venda_diaria_valido(fname):
    s = str(fname or '').lower().strip()
    if s.startswith('~$'):
        return False
    if not (s.endswith('.xls') or s.endswith('.xlsx')):
        return False
    return ('analise_total_venda' in s) or ('analises_totais_vendas' in s) or ('total_venda' in s)

def _set_all_visible_date_inputs(valor):
    """Preenche datas no relatório de análise total de vendas.
    O SGI muda IDs/nomes, então usamos seletores amplos e também forçamos via JS.
    """
    selectors = [
        "//input[not(@type='hidden') and (contains(translate(@id,'DATA','data'),'data') or contains(translate(@name,'DATA','data'),'data') or contains(translate(@class,'DATE','date'),'date'))]",
        "//input[not(@type='hidden') and (@type='text' or @type='date')]",
    ]
    seen = set(); inputs = []
    for xp in selectors:
        try:
            for el in driver.find_elements(By.XPATH, xp):
                try:
                    key = el.id
                    if key not in seen:
                        seen.add(key); inputs.append(el)
                except Exception:
                    pass
        except Exception:
            pass
    count = 0
    for el in inputs:
        try:
            if not el.is_displayed():
                continue
            driver.execute_script("arguments[0].removeAttribute('readonly'); arguments[0].removeAttribute('disabled');", el)
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            try:
                el.click(); el.send_keys(Keys.CONTROL, "a"); el.send_keys(Keys.DELETE); el.send_keys(valor)
            except Exception:
                driver.execute_script("arguments[0].value = arguments[1];", el, valor)
            driver.execute_script("arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", el)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change',{bubbles:true}));", el)
            driver.execute_script("arguments[0].dispatchEvent(new Event('blur',{bubbles:true}));", el)
            count += 1
        except Exception:
            pass
    print(f"✅ V17 Venda diária: campos de data preenchidos: {count} campo(s) = {valor}")
    return count

def _parse_venda_diaria_xls(caminho):
    dfv = _ler_xls_ou_xlsx(caminho).fillna('')
    if dfv.empty:
        raise Exception('Relatório de venda diária veio vazio')
    headers = [normalizar_texto_match(x) for x in dfv.iloc[0].tolist()]
    data = dfv.iloc[1:].copy()

    def col_idx(*names):
        targets = [normalizar_texto_match(n) for n in names]
        for i, h in enumerate(headers):
            for t in targets:
                if t and t in h:
                    return i
        return None

    idx_filial = col_idx('Filial')
    idx_total = col_idx('Total Vendido')
    idx_serv = col_idx('Valor Serviço', 'Valor Servico')
    idx_acresc = col_idx('Valor Acréscimo Serviço', 'Valor Acrescimo Servico')
    idx_qtd = col_idx('Qtde Vendida')

    if idx_total is None:
        raise Exception(f'Coluna Total Vendido não encontrada no relatório de venda diária. Headers={headers}')

    def val(row, idx):
        if idx is None:
            return 0.0
        return _tratar_valor_margem(row.iloc[idx])

    total_row = None
    if idx_filial is not None:
        for _, row in data.iterrows():
            if normalizar_texto_match(row.iloc[idx_filial]) == 'TOTAL':
                total_row = row
                break

    rows_filiais = []
    if total_row is None:
        total_vendido = valor_servico = valor_acrescimo = qtd = 0.0
        for _, row in data.iterrows():
            label = str(row.iloc[idx_filial] if idx_filial is not None else '').strip()
            if not label or normalizar_texto_match(label) in ('FILIAL', 'TOTAL'):
                continue
            tv = val(row, idx_total)
            vs = val(row, idx_serv)
            va = val(row, idx_acresc)
            qv = val(row, idx_qtd)
            total_vendido += tv
            valor_servico += vs
            valor_acrescimo += va
            qtd += qv
            rows_filiais.append({'filial': label, 'total_vendido': tv, 'valor_servico': vs, 'valor_acrescimo_servico': va, 'venda_diaria_total': round(tv+vs+va, 2), 'qtde_vendida': qv})
    else:
        total_vendido = val(total_row, idx_total)
        valor_servico = val(total_row, idx_serv)
        valor_acrescimo = val(total_row, idx_acresc)
        qtd = val(total_row, idx_qtd)

    venda_diaria_total = round(total_vendido + valor_servico + valor_acrescimo, 2)
    return {
        'empresa': {
            'total_vendido': round(total_vendido, 2),
            'valor_servico': round(valor_servico, 2),
            'valor_acrescimo_servico': round(valor_acrescimo, 2),
            'venda_diaria_total': venda_diaria_total,
            'qtde_vendida': qtd,
        },
        'filiais': rows_filiais,
    }

def coletar_venda_diaria_oficial():
    data_hoje = hoje.strftime("%d/%m/%Y")
    print("\\n🕒 Iniciando coleta de venda diária oficial...")
    driver.get(URL + "/relatorio_analises_totais_vendas")
    wait.until(EC.presence_of_element_located((By.ID, "gerar")))
    time.sleep(1.5)
    esperar_sumir_overlays(10)
    _set_all_visible_date_inputs(data_hoje)

    try:
        Select(wait.until(EC.presence_of_element_located((By.ID, "_formato")))).select_by_value("xls")
        print("✅ Formato XLS para venda diária")
    except Exception as e:
        print(f"⚠️ Não consegui selecionar XLS na venda diária: {e}")

    arquivos_antes = set(f for f in os.listdir(download_dir) if _nome_arquivo_venda_diaria_valido(f))
    gerar_btn = wait.until(EC.presence_of_element_located((By.ID, "gerar")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", gerar_btn)
    esperar_sumir_overlays(10)
    try:
        wait.until(EC.element_to_be_clickable((By.ID, "gerar"))).click()
    except Exception:
        driver.execute_script("arguments[0].click();", gerar_btn)
    print("📥 Gerando XLS venda diária... aguardando download...")

    caminho = None
    for _ in range(60):
        time.sleep(2)
        baixando = any(f.endswith(".crdownload") or f.endswith(".tmp") for f in os.listdir(download_dir))
        if baixando:
            continue
        novos = set(f for f in os.listdir(download_dir) if _nome_arquivo_venda_diaria_valido(f)) - arquivos_antes
        if novos:
            caminho = max([os.path.join(download_dir, f) for f in novos], key=os.path.getctime)
            break

    if not caminho:
        todos = [os.path.join(download_dir, f) for f in os.listdir(download_dir) if _nome_arquivo_venda_diaria_valido(f)]
        if not todos:
            raise Exception("Nenhum arquivo XLS/XLSX de venda diária foi baixado")
        caminho = max(todos, key=os.path.getctime)
        print(f"⚠️ Usando último arquivo de venda diária encontrado: {caminho}")
    else:
        print(f"✅ Download venda diária OK: {caminho}")

    dados = _parse_venda_diaria_xls(caminho)
    payload = {'coletado_em': now_brasilia().strftime("%Y-%m-%d %H:%M:%S"), 'data': data_hoje, **dados}
    json_path = os.path.join(pasta, "venda_diaria_mes_atual.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"✅ Venda diária oficial: total_vendido={dados['empresa']['total_vendido']:.2f} servico={dados['empresa']['valor_servico']:.2f} acrescimo={dados['empresa']['valor_acrescimo_servico']:.2f} total={dados['empresa']['venda_diaria_total']:.2f}")
    return {"json_path": json_path, "xlsx_path": caminho, "dados": payload}

def aplicar_venda_diaria_em_metas(metas_info, venda_diaria_info):
    try:
        if not metas_info.get("json_path") or not os.path.exists(metas_info["json_path"]):
            return metas_info
        with open(metas_info["json_path"], "r", encoding="utf-8") as f:
            dados_metas = json.load(f)
        dados_metas["venda_diaria"] = venda_diaria_info.get("dados", {})
        with open(metas_info["json_path"], "w", encoding="utf-8") as f:
            json.dump(dados_metas, f, ensure_ascii=False, indent=2)
        print("✅ Venda diária oficial anexada ao metas_vendas_mes_atual.json")
    except Exception as e:
        print(f"⚠️ Erro ao anexar venda diária ao metas: {e}")
    return metas_info


# ===== LOGIN ROBUSTO — UMA ÚNICA VEZ
def fazer_login_sgi():
    driver.get(URL)
    time.sleep(5)
    print("TITLE:", driver.title)
    print("URL ATUAL:", driver.current_url)

    campo_usuario = None
    seletores_usuario = [
        (By.NAME, "usuario"),
        (By.ID, "usuario"),
        (By.CSS_SELECTOR, "input[name='usuario']"),
        (By.CSS_SELECTOR, "input[id='usuario']"),
        (By.XPATH, "//input[contains(@name,'usuario') or contains(@id,'usuario')]"),
        (By.XPATH, "//input[@type='text' and not(@disabled)]"),
    ]

    for by, sel in seletores_usuario:
        try:
            cand = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((by, sel))
            )
            if cand:
                campo_usuario = cand
                print(f"✅ Campo usuário encontrado em: {by} = {sel}")
                break
        except Exception:
            pass

    if not campo_usuario:
        print("❌ Campo usuário não encontrado")
        print("URL:", driver.current_url)
        print("TITLE:", driver.title)
        try:
            driver.save_screenshot(os.path.join(download_dir, "debug_login_worker.png"))
        except Exception:
            pass
        raise Exception("Campo usuário não encontrado na tela de login")

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", campo_usuario)
        esperar_sumir_overlays(8)
        campo_usuario.click()
        campo_usuario.send_keys(Keys.CONTROL, "a")
        campo_usuario.send_keys(Keys.DELETE)
        campo_usuario.send_keys(LOGIN)
    except Exception:
        # Railway às vezes acha o input, mas ele fica não-interagível. Força via JS.
        driver.execute_script("""
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
        """, campo_usuario, LOGIN)

    senha_field = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='password']")))
    try:
        senha_field.click()
        senha_field.send_keys(Keys.CONTROL, "a")
        senha_field.send_keys(Keys.DELETE)
        senha_field.send_keys(SENHA)
        senha_field.send_keys(Keys.ENTER)
    except Exception:
        driver.execute_script("""
            arguments[0].value = arguments[1];
            arguments[0].dispatchEvent(new Event('input', {bubbles:true}));
            arguments[0].dispatchEvent(new Event('change', {bubbles:true}));
        """, senha_field, SENHA)
        try:
            btn = driver.find_element(By.XPATH, "//button[@type='submit'] | //input[@type='submit']")
            driver.execute_script("arguments[0].click();", btn)
        except Exception:
            senha_field.send_keys(Keys.ENTER)

    try:
        btn_filial = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "botao_prosseguir_informa_local_trabalho"))
        )
        driver.execute_script("arguments[0].click();", btn_filial)
        print("✅ Filial OK")
    except Exception:
        print("⚠️ Tela de filial não apareceu")

    time.sleep(5)

fazer_login_sgi()
print('✅ COBRANCA10_WORKER_V17_META_REAL_TOTAL_DIARIA_MARGEM_FLUXO')

metas_vendas_info = {"json_path": None, "xlsx_path": None, "dados": {}}
try:
    metas_vendas_info = coletar_metas_vendas_mes_atual()
except Exception as e:
    print(f"⚠️ Erro ao coletar metas/vendas do SGI: {e}")

margens_brutas_info = {"json_path": None, "xlsx_path": None, "dados": {}}
try:
    margens_brutas_info = coletar_margens_brutas_mes_atual()
except Exception as e:
    print(f"⚠️ Erro ao coletar Margem Bruta/Rentabilidade do SGI: {e}")

# V16: Serviços são SOMENTE do Controle de Metas SGI.
# Não coletar /relatorio_servicos e não aplicar relatório individual nos blocos de meta.
relatorio_servicos_info = {"json_path": None, "xlsx_path": None, "dados": {"empresa": {}, "servicos": {}, "filiais": {}, "vendedores": {}, "detalhes": []}}
print("✅ V17_SERVICOS_SOMENTE_CONTROLE_META: relatório individual de serviços desativado; valores de serviço vêm da coluna Realizado Total da meta SGI")

venda_diaria_info = {"json_path": None, "xlsx_path": None, "dados": {}}
try:
    venda_diaria_info = coletar_venda_diaria_oficial()
    metas_vendas_info = aplicar_venda_diaria_em_metas(metas_vendas_info, venda_diaria_info)
except Exception as e:
    print(f"⚠️ Erro ao coletar venda diária oficial do SGI: {e}")

FTP_HOST  = "moveisdolar.com.br"
FTP_USER  = "moveisdolar3"
FTP_PASS  = "Deg27ll02mdl2301#"
FTP_DIR   = "/public_html/colaborador"

try:
    import ftplib
    from io import BytesIO
    ftp = ftplib.FTP()
    ftp.connect(FTP_HOST, 21, timeout=30)
    ftp.login(FTP_USER, FTP_PASS)
    ftp.encoding = 'utf-8'
    ftp.cwd(FTP_DIR)

    if metas_vendas_info.get('json_path') and os.path.exists(metas_vendas_info['json_path']):
        with open(metas_vendas_info['json_path'], 'rb') as f_json:
            ftp.storbinary('STOR metas_vendas_mes_atual.json', f_json)
    if metas_vendas_info.get('xlsx_path') and os.path.exists(metas_vendas_info['xlsx_path']):
        with open(metas_vendas_info['xlsx_path'], 'rb') as f_xlsx:
            ftp.storbinary('STOR metas_vendas_mes_atual.xlsx', f_xlsx)

    if margens_brutas_info.get('json_path') and os.path.exists(margens_brutas_info['json_path']):
        with open(margens_brutas_info['json_path'], 'rb') as f_mjson:
            ftp.storbinary('STOR margens_brutas_mes_atual.json', f_mjson)
    if margens_brutas_info.get('xlsx_path') and os.path.exists(margens_brutas_info['xlsx_path']):
        with open(margens_brutas_info['xlsx_path'], 'rb') as f_mxlsx:
            ftp.storbinary('STOR margens_brutas_mes_atual.xlsx', f_mxlsx)

    # V16: limpa arquivo remoto antigo do relatório individual de serviços para não contaminar a tela.
    ftp.storbinary('STOR relatorio_servicos_mes_atual.json', BytesIO(b'{"empresa":{},"servicos":{},"filiais":{},"vendedores":{},"detalhes":[]}'))
    if venda_diaria_info.get('json_path') and os.path.exists(venda_diaria_info['json_path']):
        with open(venda_diaria_info['json_path'], 'rb') as f_djson:
            ftp.storbinary('STOR venda_diaria_mes_atual.json', f_djson)

    ver = json.dumps({'updated_at': now_brasilia().isoformat(), 'updated_at_label': now_brasilia().strftime('%d/%m/%Y %H:%M:%S'), 'timezone': 'America/Sao_Paulo', 'scope': 'sales_only'}, ensure_ascii=False).encode('utf-8')
    ftp.storbinary('STOR sales_version.json', BytesIO(ver))
    ftp.quit()
    print('✅ Upload vendas/margens/serviços concluído')
except Exception as e:
    print(f'⚠️ Erro no upload FTP vendas/margens: {e}')

driver.quit()
print('🔥 WORKER DE VENDAS FINALIZADO')
